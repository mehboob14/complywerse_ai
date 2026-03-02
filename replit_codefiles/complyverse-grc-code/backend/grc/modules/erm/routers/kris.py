from typing import List, Optional
from datetime import datetime, timedelta
from io import BytesIO
import re
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ....models import (
    Risk, RiskKRI, RiskKRIMeasurement, GRCUser, get_db
)
from ....schemas import (
    RiskKRICreate, RiskKRIUpdate, RiskKRIResponse,
    RiskKRIMeasurementCreate, RiskKRIMeasurementResponse,
    MessageResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/kris", tags=["ERM - Key Risk Indicators"])


def calculate_kri_status(value: float, kri: RiskKRI) -> str:
    if kri.green_threshold is None or kri.amber_threshold is None:
        return "unknown"
    
    if kri.threshold_direction == "lower_is_better":
        if value <= kri.green_threshold:
            return "green"
        elif value <= kri.amber_threshold:
            return "amber"
        else:
            return "red"
    else:
        if value >= kri.green_threshold:
            return "green"
        elif value >= kri.amber_threshold:
            return "amber"
        else:
            return "red"


@router.get("", response_model=List[RiskKRIResponse])
def list_kris(
    risk_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    is_active: Optional[bool] = True,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RiskKRI).outerjoin(Risk).filter(
        or_(
            Risk.tenant_id.in_(user_tenants),
            RiskKRI.tenant_id.in_(user_tenants)
        )
    )
    
    if risk_id:
        query = query.filter(RiskKRI.risk_id == risk_id)
    if is_active is not None:
        query = query.filter(RiskKRI.is_active == is_active)
    
    kris = query.offset(skip).limit(limit).all()
    
    result = []
    for kri in kris:
        kri_data = RiskKRIResponse.model_validate(kri)
        if kri.current_value is not None:
            kri_data.current_status = calculate_kri_status(kri.current_value, kri)
        result.append(kri_data)
    
    if status_filter:
        result = [k for k in result if k.current_status == status_filter]
    
    return result


@router.post("", response_model=RiskKRIResponse, status_code=status.HTTP_201_CREATED)
def create_kri(
    kri: RiskKRICreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    tenant_id = get_user_primary_tenant(current_user, db)
    
    if kri.risk_id:
        risk = db.query(Risk).filter(
            Risk.id == kri.risk_id,
            Risk.tenant_id.in_(user_tenants)
        ).first()
        if not risk:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Risk not found"
            )
    
    db_kri = RiskKRI(
        tenant_id=tenant_id,
        risk_id=kri.risk_id,
        name=kri.name,
        description=kri.description,
        metric_type=kri.metric_type,
        unit=kri.unit,
        green_threshold=kri.green_threshold,
        amber_threshold=kri.amber_threshold,
        threshold_direction=kri.threshold_direction,
        frequency=kri.frequency,
        data_source=kri.data_source,
        owner_id=kri.owner_id
    )
    db.add(db_kri)
    db.commit()
    db.refresh(db_kri)
    return db_kri


@router.get("/alerts", response_model=List[RiskKRIResponse])
def get_kri_alerts(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    kris = db.query(RiskKRI).outerjoin(Risk).filter(
        or_(
            Risk.tenant_id.in_(user_tenants),
            RiskKRI.tenant_id.in_(user_tenants)
        ),
        RiskKRI.is_active == True,
        RiskKRI.current_value.isnot(None)
    ).all()
    
    alerts = []
    for kri in kris:
        kri_status = calculate_kri_status(kri.current_value, kri)
        if kri_status in ["red", "amber"]:
            kri_data = RiskKRIResponse.model_validate(kri)
            kri_data.current_status = kri_status
            alerts.append(kri_data)
    
    return sorted(alerts, key=lambda x: 0 if x.current_status == "red" else 1)


@router.get("/{kri_id}", response_model=RiskKRIResponse)
def get_kri(
    kri_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).options(
        joinedload(RiskKRI.measurements)
    ).outerjoin(Risk).filter(
        RiskKRI.id == kri_id,
        or_(
            Risk.tenant_id.in_(user_tenants),
            RiskKRI.tenant_id.in_(user_tenants)
        )
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    kri_data = RiskKRIResponse.model_validate(kri)
    if kri.current_value is not None:
        kri_data.current_status = calculate_kri_status(kri.current_value, kri)
    return kri_data


@router.put("/{kri_id}", response_model=RiskKRIResponse)
def update_kri(
    kri_id: int,
    kri_update: RiskKRIUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).outerjoin(Risk).filter(
        RiskKRI.id == kri_id,
        or_(Risk.tenant_id.in_(user_tenants), RiskKRI.tenant_id.in_(user_tenants))
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    update_data = kri_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(kri, key, value)
    
    db.commit()
    db.refresh(kri)
    
    kri_data = RiskKRIResponse.model_validate(kri)
    if kri.current_value is not None:
        kri_data.current_status = calculate_kri_status(kri.current_value, kri)
    return kri_data


@router.delete("/{kri_id}", response_model=MessageResponse)
def delete_kri(
    kri_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).outerjoin(Risk).filter(
        RiskKRI.id == kri_id,
        or_(Risk.tenant_id.in_(user_tenants), RiskKRI.tenant_id.in_(user_tenants))
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    db.delete(kri)
    db.commit()
    return {"message": "KRI deleted successfully", "id": kri_id}


@router.post("/{kri_id}/measure", response_model=RiskKRIMeasurementResponse)
def record_kri_measurement(
    kri_id: int,
    measurement: RiskKRIMeasurementCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).outerjoin(Risk).filter(
        RiskKRI.id == kri_id,
        or_(Risk.tenant_id.in_(user_tenants), RiskKRI.tenant_id.in_(user_tenants))
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    status_val = calculate_kri_status(measurement.value, kri)
    
    db_measurement = RiskKRIMeasurement(
        kri_id=kri_id,
        value=measurement.value,
        status=status_val,
        measured_by=current_user.id,
        notes=measurement.notes
    )
    db.add(db_measurement)
    
    kri.current_value = measurement.value
    kri.last_measured_at = datetime.utcnow()
    
    db.commit()
    db.refresh(db_measurement)
    return db_measurement


@router.get("/{kri_id}/trend", response_model=List[RiskKRIMeasurementResponse])
def get_kri_trend(
    kri_id: int,
    days: int = 365,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).outerjoin(Risk).filter(
        RiskKRI.id == kri_id,
        or_(Risk.tenant_id.in_(user_tenants), RiskKRI.tenant_id.in_(user_tenants))
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    start_date = datetime.utcnow() - timedelta(days=days)
    measurements = db.query(RiskKRIMeasurement).filter(
        RiskKRIMeasurement.kri_id == kri_id,
        RiskKRIMeasurement.measured_at >= start_date
    ).order_by(RiskKRIMeasurement.measured_at.asc()).all()
    
    return measurements


HEADER_KEYWORDS = {
    'name': ['kri name', 'kri_name', 'indicator name', 'indicator', 'name', 'key risk indicator'],
    'kri_id': ['kri id', 'kri_id', 'id', 'indicator id', 'ref', 'reference'],
    'category': ['risk category', 'category', 'risk_category', 'risk area', 'domain'],
    'description': ['description', 'desc', 'details', 'definition'],
    'measurement': ['measurement', 'metric', 'measure', 'kpi', 'metric_type', 'measurement type'],
    'unit': ['unit', 'unit of measure', 'uom', 'units'],
    'frequency': ['frequency', 'monitoring frequency', 'reporting frequency', 'review frequency', 'cadence'],
    'owner': ['owner', 'kri owner', 'responsible', 'assigned to', 'assignee'],
    'current_value': ['current value', 'current_value', 'value', 'current', 'latest value', 'actual'],
    'previous_value': ['previous value', 'previous_value', 'prior value', 'last value'],
    'trend': ['trend', 'direction', 'movement'],
    'green_threshold': ['green threshold', 'green_threshold', 'green', 'target', 'green limit'],
    'amber_threshold': ['amber threshold', 'amber_threshold', 'amber', 'warning', 'amber limit', 'yellow threshold', 'yellow'],
    'red_threshold': ['red threshold', 'red_threshold', 'red', 'critical', 'red limit', 'breach'],
    'status': ['status', 'current status', 'rag status', 'rag', 'indicator status'],
    'data_source': ['data source', 'data_source', 'source', 'system'],
    'threshold_direction': ['threshold direction', 'direction', 'threshold_direction', 'polarity'],
}


def _match_header(header_text: str) -> Optional[str]:
    if not header_text:
        return None
    text = str(header_text).strip().lower()
    for field, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            if text == kw or text.replace('_', ' ') == kw:
                return field
    for field, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            if kw in text or text in kw:
                return field
    return None


def _parse_threshold(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    nums = re.findall(r'[-+]?\d*\.?\d+', text)
    if nums:
        return float(nums[0])
    return None


def _parse_frequency(value) -> str:
    if not value:
        return "monthly"
    text = str(value).strip().lower()
    freq_map = {
        'daily': 'daily', 'day': 'daily',
        'weekly': 'weekly', 'week': 'weekly',
        'monthly': 'monthly', 'month': 'monthly',
        'quarterly': 'quarterly', 'quarter': 'quarterly',
        'annually': 'annually', 'annual': 'annually', 'yearly': 'annually', 'year': 'annually',
    }
    for key, val in freq_map.items():
        if key in text:
            return val
    return "monthly"


def _detect_threshold_direction(green_val, red_val) -> str:
    if green_val is not None and red_val is not None:
        if green_val < red_val:
            return "lower_is_better"
        else:
            return "higher_is_better"
    return "lower_is_better"


@router.post("/upload", response_model=None)
def upload_kris(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No tenant found for user")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    ws = None
    preferred_names = ['kris', 'kri', 'key risk indicators', 'indicators', 'sheet1']
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in preferred_names:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_row_idx = None
    column_map = {}
    for idx, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c else '' for c in row]
        matched = {}
        for col_idx, cell in enumerate(cells):
            field = _match_header(cell)
            if field and field not in matched:
                matched[field] = col_idx
        if 'name' in matched and len(matched) >= 3:
            header_row_idx = idx
            column_map = matched
            break

    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not detect header row. Ensure columns include at least 'KRI Name' and 2 other recognized columns.")

    data_rows = rows[header_row_idx + 1:]
    created = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
        try:
            def get_val(field):
                col = column_map.get(field)
                if col is not None and col < len(row):
                    v = row[col]
                    return v if v is not None else None
                return None

            name = get_val('name')
            if not name or str(name).strip() == '':
                skipped += 1
                continue
            name = str(name).strip()

            description_parts = []
            if get_val('description'):
                description_parts.append(str(get_val('description')).strip())
            if get_val('category'):
                description_parts.append(f"Category: {str(get_val('category')).strip()}")
            if get_val('kri_id'):
                description_parts.append(f"KRI ID: {str(get_val('kri_id')).strip()}")

            description = ' | '.join(description_parts) if description_parts else None

            measurement = get_val('measurement')
            unit = get_val('unit')
            metric_type = "numeric"
            if measurement:
                m_text = str(measurement).strip().lower()
                if any(k in m_text for k in ['percent', '%', 'ratio', 'rate']):
                    metric_type = "percentage"
                    if not unit:
                        unit = "%"
                elif any(k in m_text for k in ['count', 'number', 'total', '#']):
                    metric_type = "count"
                elif any(k in m_text for k in ['bool', 'yes/no', 'true/false']):
                    metric_type = "boolean"
                if not unit:
                    unit = str(measurement).strip()

            green = _parse_threshold(get_val('green_threshold'))
            amber = _parse_threshold(get_val('amber_threshold'))
            red = _parse_threshold(get_val('red_threshold'))

            if green is None and red is not None and amber is not None:
                green = amber
                amber = red
            elif green is not None and amber is None and red is not None:
                amber = (green + red) / 2

            threshold_dir = get_val('threshold_direction')
            if threshold_dir:
                td_text = str(threshold_dir).strip().lower()
                if any(k in td_text for k in ['higher', 'more', 'above', 'increase']):
                    threshold_direction = "higher_is_better"
                else:
                    threshold_direction = "lower_is_better"
            else:
                threshold_direction = _detect_threshold_direction(green, red)

            current_value = _parse_threshold(get_val('current_value'))
            frequency = _parse_frequency(get_val('frequency'))

            data_source_parts = []
            if get_val('data_source'):
                data_source_parts.append(str(get_val('data_source')).strip())
            if get_val('kri_id'):
                data_source_parts.append(f"Ref: {str(get_val('kri_id')).strip()}")
            data_source = ' | '.join(data_source_parts) if data_source_parts else None

            db_kri = RiskKRI(
                tenant_id=tenant_id,
                risk_id=None,
                name=name,
                description=description,
                metric_type=metric_type,
                unit=unit,
                current_value=current_value,
                green_threshold=green,
                amber_threshold=amber,
                threshold_direction=threshold_direction,
                frequency=frequency,
                data_source=data_source,
                is_active=True,
            )
            db.add(db_kri)
            created += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            skipped += 1

    if created > 0:
        db.commit()

    return {
        "message": f"Successfully imported {created} KRIs",
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
