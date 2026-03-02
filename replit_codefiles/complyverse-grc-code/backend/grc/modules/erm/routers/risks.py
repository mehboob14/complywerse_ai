from typing import List, Optional
from datetime import datetime
from io import BytesIO
import os
import json
import logging
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from pydantic import BaseModel
import openpyxl
from openai import OpenAI

logger = logging.getLogger(__name__)

from ....models import (
    Risk, RiskControlLink, RiskAssetLink, RiskEvidenceLink,
    RiskFrameworkControlLink, RiskGovernanceLink, RiskAuditFindingLink,
    NormalizedControl, FrameworkControl, ITAsset, Evidence,
    GovernanceObjective, Issue, GRCUser, Tenant, get_db,
    ParsedFrameworkControl, UploadedFramework
)
from ....schemas import (
    RiskCreate, RiskUpdate, RiskResponse,
    RiskAssessment, RiskTreatment,
    RiskControlLinkCreate, RiskAssetLinkCreate, RiskEvidenceLinkCreate,
    RiskFrameworkControlLinkCreate, RiskGovernanceLinkCreate,
    RiskDetailResponse, RiskHeatmapData, MessageResponse,
    RiskAuditFindingLinkCreate, RiskAuditFindingLinkResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/risks", tags=["ERM - Risk Register"])


def calculate_risk_score(likelihood: int, impact: int) -> float:
    return likelihood * impact


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


@router.get("", response_model=List[RiskResponse])
def list_risks(
    tenant_id: Optional[int] = None,
    category: Optional[str] = None,
    status_filter: Optional[str] = None,
    min_score: Optional[float] = None,
    max_score: Optional[float] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    if category:
        query = query.filter(Risk.category == category)
    if status_filter:
        query = query.filter(Risk.status == status_filter)
    if min_score is not None:
        query = query.filter(Risk.inherent_score >= min_score)
    if max_score is not None:
        query = query.filter(Risk.inherent_score <= max_score)
    
    risks = query.order_by(Risk.created_at.desc()).offset(skip).limit(limit).all()
    return risks


@router.post("", response_model=RiskResponse, status_code=status.HTTP_201_CREATED)
def create_risk(
    risk: RiskCreate,
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tenant not found"
        )
    
    db_risk = Risk(
        tenant_id=tenant_id,
        title=risk.title,
        description=risk.description,
        category=risk.category,
        risk_category=getattr(risk, 'risk_category', risk.category),
        owner_id=risk.owner_id
    )
    db.add(db_risk)
    db.commit()
    db.refresh(db_risk)
    return db_risk


@router.get("/dashboard")
def get_risk_dashboard(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {
            "total_risks": 0,
            "by_category": {},
            "by_status": {},
            "by_score_range": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "avg_inherent_score": 0,
            "avg_residual_score": 0,
            "open_risks": 0,
            "risks_needing_review": 0
        }
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    by_category = {}
    by_status = {}
    by_score_range = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    total_inherent_score = 0
    total_residual_score = 0
    risks_with_score = 0
    
    for risk in risks:
        cat = risk.risk_category or risk.category or "operational"
        by_category[cat] = by_category.get(cat, 0) + 1
        
        status_val = risk.status or "open"
        by_status[status_val] = by_status.get(status_val, 0) + 1
        
        score = risk.residual_score or risk.inherent_score or 0
        if score >= 20:
            by_score_range["critical"] += 1
        elif score >= 12:
            by_score_range["high"] += 1
        elif score >= 6:
            by_score_range["medium"] += 1
        else:
            by_score_range["low"] += 1
        
        if risk.inherent_score:
            total_inherent_score += risk.inherent_score
            risks_with_score += 1
        if risk.residual_score:
            total_residual_score += risk.residual_score
    
    return {
        "total_risks": len(risks),
        "by_category": by_category,
        "by_status": by_status,
        "by_score_range": by_score_range,
        "avg_inherent_score": round(total_inherent_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "avg_residual_score": round(total_residual_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "open_risks": by_status.get("open", 0),
        "risks_needing_review": sum(1 for r in risks if r.review_date and r.review_date < datetime.utcnow())
    }


@router.get("/heatmap")
def get_risk_heatmap(
    risk_type: Optional[str] = None,
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    heatmap = {}
    risk_type_prefix = risk_type if risk_type in ["inherent", "residual"] else "inherent"
    
    for risk in risks:
        likelihood = getattr(risk, f"{risk_type_prefix}_likelihood") or 0
        impact = getattr(risk, f"{risk_type_prefix}_impact") or 0
        
        if likelihood > 0 and impact > 0:
            key = f"{likelihood}-{impact}"
            if key not in heatmap:
                heatmap[key] = {"likelihood": likelihood, "impact": impact, "count": 0, "risks": []}
            heatmap[key]["count"] += 1
            heatmap[key]["risks"].append({
                "id": risk.id,
                "title": risk.title,
                "score": getattr(risk, f"{risk_type_prefix}_score")
            })
    
    return list(heatmap.values())


@router.get("/{risk_id}/detail")
def get_risk_detail(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links).joinedload(RiskControlLink.normalized_control),
        joinedload(Risk.asset_links).joinedload(RiskAssetLink.asset),
        joinedload(Risk.evidence_links).joinedload(RiskEvidenceLink.evidence),
        joinedload(Risk.framework_control_links).joinedload(RiskFrameworkControlLink.framework_control),
        joinedload(Risk.governance_links).joinedload(RiskGovernanceLink.governance_objective),
        joinedload(Risk.owner)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    linked_controls = []
    for link in risk.control_links:
        if link.normalized_control:
            linked_controls.append({
                "id": link.id,
                "control_id": link.normalized_control.id,
                "code": link.normalized_control.code,
                "name": link.normalized_control.name
            })
    
    linked_framework_controls = []
    for link in risk.framework_control_links:
        if link.framework_control:
            linked_framework_controls.append({
                "id": link.id,
                "framework_control_id": link.framework_control.id,
                "code": link.framework_control.code,
                "name": link.framework_control.name,
                "mitigation_effectiveness": link.mitigation_effectiveness,
                "notes": link.notes
            })
    
    linked_assets = []
    for link in risk.asset_links:
        if link.asset:
            linked_assets.append({
                "id": link.id,
                "asset_id": link.asset.id,
                "name": link.asset.name,
                "asset_type": link.asset.asset_type
            })
    
    linked_evidence = []
    for link in risk.evidence_links:
        if link.evidence:
            linked_evidence.append({
                "id": link.id,
                "evidence_id": link.evidence.id,
                "name": link.evidence.name,
                "status": link.evidence.status
            })
    
    linked_governance = []
    for link in risk.governance_links:
        if link.governance_objective:
            linked_governance.append({
                "id": link.id,
                "governance_objective_id": link.governance_objective.id,
                "name": link.governance_objective.name,
                "impact_level": link.impact_level
            })
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "owner_id": risk.owner_id,
        "owner_name": risk.owner.display_name if risk.owner else None,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "due_date": risk.due_date.isoformat() if risk.due_date else None,
        "review_date": risk.review_date.isoformat() if risk.review_date else None,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": linked_controls,
        "linked_framework_controls": linked_framework_controls,
        "linked_assets": linked_assets,
        "linked_evidence": linked_evidence,
        "linked_governance": linked_governance
    }


@router.get("/{risk_id}", response_model=dict)
def get_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links),
        joinedload(Risk.asset_links),
        joinedload(Risk.evidence_links)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "owner_id": risk.owner_id,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": [link.normalized_control_id for link in risk.control_links],
        "linked_assets": [link.asset_id for link in risk.asset_links],
        "linked_evidence": [link.evidence_id for link in risk.evidence_links]
    }


@router.put("/{risk_id}", response_model=RiskResponse)
def update_risk(
    risk_id: int,
    risk_update: RiskUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    update_data = risk_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(risk, field, value)
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.delete("/{risk_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    db.delete(risk)
    db.commit()
    return None


@router.post("/{risk_id}/assess", response_model=RiskResponse)
def assess_risk(
    risk_id: int,
    assessment: RiskAssessment,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.inherent_likelihood = assessment.inherent_likelihood
    risk.inherent_impact = assessment.inherent_impact
    risk.inherent_score = calculate_risk_score(
        assessment.inherent_likelihood,
        assessment.inherent_impact
    )
    
    if assessment.residual_likelihood and assessment.residual_impact:
        risk.residual_likelihood = assessment.residual_likelihood
        risk.residual_impact = assessment.residual_impact
        risk.residual_score = calculate_risk_score(
            assessment.residual_likelihood,
            assessment.residual_impact
        )
    
    if assessment.risk_appetite:
        risk.risk_appetite = assessment.risk_appetite
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/treatment", response_model=RiskResponse)
def add_treatment_plan(
    risk_id: int,
    treatment: RiskTreatment,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.treatment_plan = treatment.treatment_plan
    risk.status = "mitigating"
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_control(
    risk_id: int,
    link: RiskControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    control = db.query(NormalizedControl).filter(
        NormalizedControl.id == link.normalized_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control not found"
        )
    
    existing = db.query(RiskControlLink).filter(
        RiskControlLink.risk_id == risk_id,
        RiskControlLink.normalized_control_id == link.normalized_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskControlLink(
        risk_id=risk_id,
        normalized_control_id=link.normalized_control_id
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Control linked successfully")


@router.post("/{risk_id}/link-framework-control", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_framework_control(
    risk_id: int,
    link: RiskFrameworkControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    framework_control = db.query(FrameworkControl).filter(
        FrameworkControl.id == link.framework_control_id
    ).first()
    if not framework_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control not found"
        )
    
    existing = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.risk_id == risk_id,
        RiskFrameworkControlLink.framework_control_id == link.framework_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskFrameworkControlLink(
        risk_id=risk_id,
        framework_control_id=link.framework_control_id,
        mitigation_effectiveness=link.mitigation_effectiveness,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Framework control linked successfully")


@router.delete("/{risk_id}/link-framework-control/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_framework_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.id == link_id,
        RiskFrameworkControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/link-governance", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_governance(
    risk_id: int,
    link: RiskGovernanceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    governance_objective = db.query(GovernanceObjective).filter(
        GovernanceObjective.id == link.governance_objective_id,
        GovernanceObjective.tenant_id.in_(user_tenants)
    ).first()
    if not governance_objective:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance objective not found"
        )
    
    existing = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.risk_id == risk_id,
        RiskGovernanceLink.governance_objective_id == link.governance_objective_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskGovernanceLink(
        risk_id=risk_id,
        governance_objective_id=link.governance_objective_id,
        impact_level=link.impact_level
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Governance objective linked successfully")


@router.delete("/{risk_id}/link-governance/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_governance(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.id == link_id,
        RiskGovernanceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.delete("/{risk_id}/controls/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskControlLink).filter(
        RiskControlLink.id == link_id,
        RiskControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/assets", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_asset(
    risk_id: int,
    link: RiskAssetLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == link.asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    existing = db.query(RiskAssetLink).filter(
        RiskAssetLink.risk_id == risk_id,
        RiskAssetLink.asset_id == link.asset_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskAssetLink(risk_id=risk_id, asset_id=link.asset_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Asset linked successfully")


@router.delete("/{risk_id}/assets/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_asset(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskAssetLink).filter(
        RiskAssetLink.id == link_id,
        RiskAssetLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/evidence", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_evidence(
    risk_id: int,
    link: RiskEvidenceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    evidence = db.query(Evidence).filter(
        Evidence.id == link.evidence_id,
        Evidence.tenant_id.in_(user_tenants)
    ).first()
    if not evidence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence not found"
        )
    
    existing = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.risk_id == risk_id,
        RiskEvidenceLink.evidence_id == link.evidence_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskEvidenceLink(risk_id=risk_id, evidence_id=link.evidence_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Evidence linked successfully")


@router.delete("/{risk_id}/evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_evidence(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.id == link_id,
        RiskEvidenceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/upload")
async def upload_risk_register(
    file: UploadFile = File(...),
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        
        ws = None
        for sheet_name in ['Risk Assessment', 'Risks', 'Risk Register', 'Sheet1']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        
        headers = []
        header_row = 1
        header_keywords = ['asset name', 'threat', 'likelihood', 'impact', 'risk score']
        for row_num in range(1, 10):
            row_values = [cell.value for cell in ws[row_num]]
            row_str = ' '.join([str(v).lower() for v in row_values if v])
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches >= 3:
                headers = row_values
                header_row = row_num
                break
        
        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                header_map[str(h).lower().strip()] = idx
        
        def get_value(row, *possible_names):
            for name in possible_names:
                if name.lower() in header_map:
                    idx = header_map[name.lower()]
                    if idx < len(row):
                        return row[idx]
            return None
        
        def parse_int(val, default=1):
            if val is None:
                return default
            if isinstance(val, (int, float)):
                return max(1, min(5, int(val)))
            try:
                return max(1, min(5, int(float(str(val).strip()))))
            except:
                return default
        
        def parse_score(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and val.startswith('='):
                return None
            try:
                return float(val)
            except:
                return None
        
        def map_category(threat_or_category):
            if not threat_or_category:
                return 'operational'
            text = str(threat_or_category).lower()
            if any(w in text for w in ['strategic', 'business', 'market']):
                return 'strategic'
            if any(w in text for w in ['financial', 'money', 'cost', 'budget']):
                return 'financial'
            if any(w in text for w in ['compliance', 'regulatory', 'legal', 'pci', 'gdpr']):
                return 'compliance'
            if any(w in text for w in ['technology', 'system', 'network', 'cyber', 'malware', 'phishing', 'security']):
                return 'technology'
            if any(w in text for w in ['vendor', 'supplier', 'third', 'partner', 'outsourcing']):
                return 'third_party'
            return 'operational'
        
        def map_status(treatment_option, residual_score):
            if not treatment_option:
                return 'open'
            text = str(treatment_option).lower()
            if 'accept' in text:
                return 'accepted'
            if 'avoid' in text or 'close' in text:
                return 'closed'
            if 'mitigat' in text or 'reduc' in text or 'treat' in text:
                if residual_score and residual_score < 10:
                    return 'mitigated'
                return 'in_treatment'
            if 'transfer' in text:
                return 'in_treatment'
            return 'open'
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(row):
                continue
            
            ref = get_value(row, 'ref', 'ref.', 'id', 'risk id', 'risk_id')
            asset_name = get_value(row, 'asset name', 'asset', 'asset_name')
            threat = get_value(row, 'threat', 'threat description')
            vulnerability = get_value(row, 'vulnerabilities', 'vulnerability', 'vuln')
            
            if not asset_name and not threat and not vulnerability:
                skipped_count += 1
                continue
            
            title_parts = []
            if asset_name:
                title_parts.append(str(asset_name).strip())
            if threat:
                threat_clean = str(threat).strip().replace('\n', ' ')[:80]
                if threat_clean:
                    title_parts.append(threat_clean)
            
            if not title_parts:
                if vulnerability:
                    title_parts.append(str(vulnerability).strip()[:80])
                elif ref:
                    title_parts.append(str(ref))
            
            if not title_parts:
                skipped_count += 1
                continue
            
            title = " - ".join(title_parts)[:200]
            
            description_parts = []
            if threat:
                description_parts.append(f"Threat: {threat}")
            if vulnerability:
                description_parts.append(f"Vulnerability: {vulnerability}")
            gaps = get_value(row, 'gaps', 'gap')
            if gaps:
                description_parts.append(f"Gaps: {gaps}")
            recommendations = get_value(row, 'recommendations', 'recommendation')
            if recommendations:
                description_parts.append(f"Recommendations: {recommendations}")
            
            description = "\n\n".join(description_parts) if description_parts else None
            
            inherent_likelihood = parse_int(get_value(row, 'likelihood', 'inherent likelihood', 'probability'))
            inherent_impact = parse_int(get_value(row, 'impact', 'inherent impact', 'consequence'))
            inherent_score = parse_score(get_value(row, 'risk score', 'inherent score', 'inherent risk'))
            if inherent_score is None:
                inherent_score = inherent_likelihood * inherent_impact
            
            residual_likelihood = parse_int(get_value(row, 'post-treatment likelihood', 'residual likelihood'), default=None)
            residual_impact = parse_int(get_value(row, 'post-treatment impact', 'residual impact'), default=None)
            residual_score = parse_score(get_value(row, 'residual risk', 'residual score', 'post-treatment risk'))
            if residual_score is None and residual_likelihood and residual_impact:
                residual_score = residual_likelihood * residual_impact
            
            mitigating_controls = get_value(row, 'mitigating action controls', 'controls', 'existing controls', 'mitigating controls')
            action_plan = get_value(row, 'action plan', 'treatment plan', 'plan')
            treatment_parts = []
            if mitigating_controls:
                treatment_parts.append(f"Existing Controls: {mitigating_controls}")
            if action_plan:
                treatment_parts.append(f"Action Plan: {action_plan}")
            treatment_plan = "\n\n".join(treatment_parts) if treatment_parts else None
            
            treatment_option = get_value(row, 'risk treatment option', 'treatment option', 'treatment')
            category = map_category(threat)
            risk_status = map_status(treatment_option, residual_score)
            
            owner_name = get_value(row, 'responsibility', 'owner', 'risk owner')
            
            try:
                db_risk = Risk(
                    tenant_id=tenant_id,
                    title=title,
                    description=description,
                    category=category,
                    risk_category=category,
                    inherent_likelihood=inherent_likelihood,
                    inherent_impact=inherent_impact,
                    inherent_score=inherent_score,
                    residual_likelihood=residual_likelihood if residual_likelihood else None,
                    residual_impact=residual_impact if residual_impact else None,
                    residual_score=residual_score,
                    treatment_plan=treatment_plan,
                    status=risk_status,
                    owner_id=current_user.id
                )
                db.add(db_risk)
                created_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Successfully imported {created_count} risks",
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}"
        )


@router.get("/aging")
def get_risks_with_aging(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.order_by(Risk.created_at.asc()).all()
    
    now = datetime.utcnow()
    result = []
    for risk in risks:
        days_since_created = (now - risk.created_at).days if risk.created_at else 0
        days_since_updated = (now - risk.updated_at).days if risk.updated_at else days_since_created
        
        result.append({
            "id": risk.id,
            "title": risk.title,
            "category": risk.risk_category or risk.category,
            "status": risk.status,
            "inherent_score": risk.inherent_score,
            "residual_score": risk.residual_score,
            "created_at": risk.created_at.isoformat() if risk.created_at else None,
            "updated_at": risk.updated_at.isoformat() if risk.updated_at else None,
            "days_since_created": days_since_created,
            "days_since_updated": days_since_updated,
            "owner_id": risk.owner_id
        })
    
    return result


@router.post("/{risk_id}/close", response_model=RiskResponse)
def close_risk(
    risk_id: int,
    closure_notes: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    if risk.status == "closed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk is already closed"
        )
    
    risk.status = "closed"
    risk.closure_status = "closed"
    risk.closed_at = datetime.utcnow()
    risk.closed_by = current_user.id
    risk.closure_notes = closure_notes
    risk.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/reopen", response_model=RiskResponse)
def reopen_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    if risk.status != "closed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk is not closed"
        )
    
    risk.status = "open"
    risk.closure_status = None
    risk.closed_at = None
    risk.closed_by = None
    risk.closure_notes = None
    risk.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/audit-findings", response_model=RiskAuditFindingLinkResponse, status_code=status.HTTP_201_CREATED)
def link_audit_finding_to_risk(
    risk_id: int,
    link: RiskAuditFindingLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    issue = db.query(Issue).filter(
        Issue.id == link.issue_id,
        Issue.tenant_id.in_(user_tenants)
    ).first()
    
    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Audit finding/issue not found"
        )
    
    existing = db.query(RiskAuditFindingLink).filter(
        RiskAuditFindingLink.risk_id == risk_id,
        RiskAuditFindingLink.issue_id == link.issue_id
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskAuditFindingLink(
        risk_id=risk_id,
        issue_id=link.issue_id,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    db.refresh(db_link)
    
    response = RiskAuditFindingLinkResponse.model_validate(db_link)
    response.issue_title = issue.title
    response.issue_severity = issue.severity
    return response


@router.delete("/{risk_id}/audit-findings/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_audit_finding_from_risk(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskAuditFindingLink).filter(
        RiskAuditFindingLink.id == link_id,
        RiskAuditFindingLink.risk_id == risk_id
    ).first()
    
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Audit finding link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


class RiskAISuggestionRequest(BaseModel):
    name: str
    category: Optional[str] = None
    sub_category: Optional[str] = None
    description: Optional[str] = None


class RecommendedControl(BaseModel):
    control_id: int
    control_name: str
    control_code: Optional[str] = None
    relevance: str
    rationale: str


class RiskAISuggestionResponse(BaseModel):
    suggested_description: str
    suggested_causes: List[str]
    suggested_consequences: List[str]
    recommended_controls: List[RecommendedControl]
    suggested_likelihood: int
    suggested_impact: int
    risk_treatment_options: List[str]


def get_openai_client() -> OpenAI:
    api_key = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    is_modelfarm = base_url and "modelfarm" in base_url
    if not api_key and not is_modelfarm:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI features unavailable. OpenAI API key not configured."
        )
    if not is_modelfarm and (api_key.startswith("_DUMMY") or api_key == "your-api-key-here" or len(api_key) < 20):
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI features unavailable. OpenAI API key not configured."
        )
    return OpenAI(
        api_key=api_key,
        base_url=base_url
    )


def parse_ai_response(response_text: str) -> dict:
    try:
        cleaned = response_text.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        return json.loads(cleaned.strip())
    except json.JSONDecodeError:
        return {
            "suggested_description": "Unable to generate description",
            "suggested_causes": [],
            "suggested_consequences": [],
            "suggested_likelihood": 3,
            "suggested_impact": 3,
            "risk_treatment_options": ["Mitigate", "Accept", "Transfer"]
        }


def get_available_controls_for_matching(db: Session, user_tenants: List[int]) -> str:
    controls_info = []
    
    normalized_controls = db.query(NormalizedControl).limit(100).all()
    for ctrl in normalized_controls:
        controls_info.append(f"- ID:{ctrl.id} | Code:{ctrl.code} | Name:{ctrl.name}")
    
    for tenant_id in user_tenants:
        parsed_controls = db.query(ParsedFrameworkControl).join(
            UploadedFramework
        ).filter(
            (UploadedFramework.tenant_id == tenant_id) | 
            (UploadedFramework.is_shared == True),
            UploadedFramework.is_active == True
        ).limit(50).all()
        
        for ctrl in parsed_controls:
            ctrl_code = ctrl.original_reference or ctrl.control_id
            controls_info.append(f"- ID:{ctrl.id} | Code:{ctrl_code} | Title:{ctrl.title[:80] if ctrl.title else 'N/A'}")
    
    return "\n".join(controls_info[:100])


RISK_AI_SUGGESTION_PROMPT = """You are an expert Enterprise Risk Management (ERM) consultant with 20+ years of experience. Analyze the risk information provided and generate comprehensive suggestions.

RISK INFORMATION:
Name: {name}
Category: {category}
Sub-category: {sub_category}
Existing Description: {description}

AVAILABLE CONTROLS FOR RECOMMENDATION (select the most relevant ones):
{available_controls}

Based on this risk, provide suggestions in the following JSON format:
{{
    "suggested_description": "<A comprehensive 2-4 sentence professional risk description that explains what the risk is, its context, and potential business impact>",
    
    "suggested_causes": [
        "<Root cause 1 - specific and actionable>",
        "<Root cause 2 - specific and actionable>",
        "<Root cause 3 - specific and actionable>"
    ],
    
    "suggested_consequences": [
        "<Business consequence 1 - specific impact on operations, finances, reputation, or compliance>",
        "<Business consequence 2 - specific impact>",
        "<Business consequence 3 - specific impact>"
    ],
    
    "recommended_control_ids": [
        {{
            "control_id": <ID number from the available controls list>,
            "relevance": "<high|medium|low>",
            "rationale": "<Explain why this control helps mitigate this specific risk>"
        }}
    ],
    
    "suggested_likelihood": <1-5 scale where 1=Rare, 2=Unlikely, 3=Possible, 4=Likely, 5=Almost Certain>,
    
    "suggested_impact": <1-5 scale where 1=Negligible, 2=Minor, 3=Moderate, 4=Major, 5=Catastrophic>,
    
    "risk_treatment_options": [
        "<Primary treatment recommendation: Mitigate/Accept/Transfer/Avoid>",
        "<Alternative treatment option>",
        "<Supporting action>"
    ]
}}

GUIDELINES:
1. Base likelihood and impact on industry standards for the risk category
2. Select 2-4 most relevant controls from the provided list
3. Be specific and actionable in causes and consequences
4. Match the professional tone expected in enterprise GRC systems
5. ONLY recommend controls that exist in the provided list - use exact IDs

Return ONLY valid JSON, no additional text."""


@router.post("/ai-suggest", response_model=RiskAISuggestionResponse)
def get_risk_ai_suggestions(
    request: RiskAISuggestionRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if not request.name or len(request.name.strip()) < 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk name must be at least 3 characters"
        )
    
    user_tenants = get_user_tenants(current_user, db)
    
    try:
        client = get_openai_client()
    except HTTPException:
        return RiskAISuggestionResponse(
            suggested_description=f"Risk related to {request.name} in the {request.category or 'operational'} category.",
            suggested_causes=["Process failure", "Human error", "Inadequate controls"],
            suggested_consequences=["Operational disruption", "Financial loss", "Reputational damage"],
            recommended_controls=[],
            suggested_likelihood=3,
            suggested_impact=3,
            risk_treatment_options=["Mitigate", "Accept", "Transfer"]
        )
    
    available_controls = get_available_controls_for_matching(db, user_tenants)
    
    prompt = RISK_AI_SUGGESTION_PROMPT.format(
        name=request.name,
        category=request.category or "Not specified",
        sub_category=request.sub_category or "Not specified",
        description=request.description or "Not provided",
        available_controls=available_controls if available_controls else "No controls available"
    )
    
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an expert ERM consultant. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1500
        )
        
        ai_response = parse_ai_response(response.choices[0].message.content)
        
        recommended_controls = []
        ai_control_recs = ai_response.get("recommended_control_ids", [])
        
        for rec in ai_control_recs[:5]:
            control_id = rec.get("control_id")
            if not control_id:
                continue
            
            normalized_ctrl = db.query(NormalizedControl).filter(
                NormalizedControl.id == control_id
            ).first()
            
            if normalized_ctrl:
                recommended_controls.append(RecommendedControl(
                    control_id=normalized_ctrl.id,
                    control_name=normalized_ctrl.name,
                    control_code=normalized_ctrl.code,
                    relevance=rec.get("relevance", "medium"),
                    rationale=rec.get("rationale", "Relevant to this risk category")
                ))
            else:
                parsed_ctrl = db.query(ParsedFrameworkControl).filter(
                    ParsedFrameworkControl.id == control_id
                ).first()
                
                if parsed_ctrl:
                    recommended_controls.append(RecommendedControl(
                        control_id=parsed_ctrl.id,
                        control_name=parsed_ctrl.title or "Control",
                        control_code=parsed_ctrl.original_reference or parsed_ctrl.control_id,
                        relevance=rec.get("relevance", "medium"),
                        rationale=rec.get("rationale", "Relevant to this risk category")
                    ))
        
        return RiskAISuggestionResponse(
            suggested_description=ai_response.get("suggested_description", f"Risk related to {request.name}"),
            suggested_causes=ai_response.get("suggested_causes", [])[:5],
            suggested_consequences=ai_response.get("suggested_consequences", [])[:5],
            recommended_controls=recommended_controls,
            suggested_likelihood=min(5, max(1, ai_response.get("suggested_likelihood", 3))),
            suggested_impact=min(5, max(1, ai_response.get("suggested_impact", 3))),
            risk_treatment_options=ai_response.get("risk_treatment_options", ["Mitigate", "Accept", "Transfer"])[:4]
        )
        
    except Exception as e:
        logger.error(f"AI suggestion error: {str(e)}")
        return RiskAISuggestionResponse(
            suggested_description=f"Risk related to {request.name} requiring assessment and mitigation.",
            suggested_causes=["Process failure", "External factors", "Resource constraints"],
            suggested_consequences=["Operational impact", "Financial impact", "Compliance impact"],
            recommended_controls=[],
            suggested_likelihood=3,
            suggested_impact=3,
            risk_treatment_options=["Mitigate", "Accept", "Transfer"]
        )


class AITreatmentPlanResponse(BaseModel):
    treatment_plan: str


@router.post("/{risk_id}/ai-treatment-plan", response_model=AITreatmentPlanResponse)
def generate_ai_treatment_plan(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    risk = db.query(Risk).options(
        joinedload(Risk.control_links).joinedload(RiskControlLink.normalized_control),
        joinedload(Risk.framework_control_links).joinedload(RiskFrameworkControlLink.framework_control),
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()

    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )

    linked_controls = []
    for link in risk.control_links:
        if link.normalized_control:
            linked_controls.append(f"{link.normalized_control.code}: {link.normalized_control.name}")
    for link in risk.framework_control_links:
        if link.framework_control:
            linked_controls.append(f"{link.framework_control.code}: {link.framework_control.name}")

    controls_text = "\n".join(f"- {c}" for c in linked_controls) if linked_controls else "No controls currently linked."

    prompt = f"""You are an enterprise risk management expert. Generate a detailed, actionable treatment plan for the following risk.

Risk Title: {risk.title}
Description: {risk.description or 'N/A'}
Category: {risk.risk_category or risk.category or 'N/A'}
Inherent Likelihood: {risk.inherent_likelihood or 'N/A'}/5
Inherent Impact: {risk.inherent_impact or 'N/A'}/5
Inherent Score: {risk.inherent_score or 'N/A'}
Residual Likelihood: {risk.residual_likelihood or 'N/A'}/5
Residual Impact: {risk.residual_impact or 'N/A'}/5
Residual Score: {risk.residual_score or 'N/A'}
Current Status: {risk.status or 'N/A'}

Linked Controls:
{controls_text}

Generate a comprehensive treatment plan that includes:
1. Treatment Strategy (mitigate, transfer, accept, or avoid - with justification)
2. Specific Action Items (3-5 concrete steps with responsible parties and timelines)
3. Control Improvements (enhancements to existing controls or new controls needed)
4. Monitoring & Review (KPIs, review frequency, escalation triggers)
5. Expected Residual Risk (target likelihood and impact after treatment)

Write the plan in clear, professional language suitable for a risk committee. Return ONLY the treatment plan text, no JSON formatting."""

    try:
        client = get_openai_client()

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an enterprise risk management expert who generates detailed, actionable risk treatment plans."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1500
        )

        treatment_plan = response.choices[0].message.content.strip()
        return AITreatmentPlanResponse(treatment_plan=treatment_plan)

    except Exception as e:
        logger.error(f"AI treatment plan generation error: {str(e)}")
        category = risk.risk_category or risk.category or "operational"
        return AITreatmentPlanResponse(
            treatment_plan=f"""Treatment Plan for: {risk.title}

1. Treatment Strategy: Mitigate
   Reduce risk through enhanced controls and monitoring.

2. Action Items:
   - Conduct detailed risk assessment and root cause analysis (Week 1-2)
   - Implement additional preventive controls specific to {category} risks (Week 2-4)
   - Establish monitoring procedures and key risk indicators (Week 3-4)
   - Train relevant staff on updated procedures (Week 4-6)
   - Conduct effectiveness review (Week 8)

3. Control Improvements:
   - Review and strengthen existing control framework
   - Add detective controls for early warning
   - Implement automated monitoring where feasible

4. Monitoring & Review:
   - Monthly KRI reporting
   - Quarterly treatment plan review
   - Immediate escalation if risk materializes

5. Expected Residual Risk:
   - Target Likelihood: {max(1, (risk.inherent_likelihood or 3) - 1)}/5
   - Target Impact: {max(1, (risk.inherent_impact or 3) - 1)}/5"""
        )
