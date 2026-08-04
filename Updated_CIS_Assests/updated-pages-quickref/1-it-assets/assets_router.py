import random
import csv
import io
import os
import json
import re
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Request, Cookie
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text
from pydantic import BaseModel

from ..modules.compliance_plugins.services.os_detector import truncate_os_key_to_second_level
from ..models import (
    ITAsset, AssetControlLink, AssetInternalControlLink, AssetRiskAssessment, AssetFrameworkControlLink,
    AssetEvidenceLink, NormalizedControl, FrameworkControl, Evidence, Risk, InternalControl,
    Vulnerability, VulnerabilityAssetLink,
    GRCUser, Tenant, TenantUser, get_db
)
# AssetSecurityComplianceSelection import removed — the four security-compliance
# endpoints that read/wrote it were deleted (they were hardcoded to a static
# CIS_WS2012R2 seed JSON and never used by operators in practice; per-asset
# rule selection now lives only on /compliance/plugins/library).
from ..schemas import (
    ITAssetCreate, ITAssetUpdate, ITAssetResponse,
    AssetValuation, AssetControlLinkCreate, AssetRiskAssessmentResponse,
    AssetDashboard, AssetCoverage, MessageResponse,
    AssetFrameworkControlLinkCreate, AssetEvidenceLinkCreate,
    AssetDetailResponse, AssetCoverageAnalysis
)
from .auth_router import require_auth, get_user_tenants, get_user_primary_tenant, decode_token
from ..tenant_manager import get_tenant_session, IS_SQLITE
from ..tenant_models import TenantUser as TenantSchemaUser
from ..rich_audit import write_rich_audit_log, model_to_dict

router = APIRouter(prefix="/assets", tags=["IT Assets"])

# SECURITY_COMPLIANCE_BENCHMARK / SECURITY_COMPLIANCE_CONTROLS_FILE constants,
# `_security_compliance_sort_tokens`, `_load_security_compliance_controls`,
# and `SecurityComplianceSelectionRequest` were removed alongside the four
# endpoints they served (see the deletions further down this file). The
# CIS_WS2012R2 seed JSON they pointed at is kept on disk because the CIS
# PDF framework-upload path in compliance_assessments_router.py still
# reads it as a structured fallback when no rule extraction succeeds.


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


def _get_asset_for_user(asset_id: int, current_user: GRCUser, db: Session) -> ITAsset:
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found",
        )
    return asset


@router.get("/tenant-users")
def get_tenant_users(
    http_request: Request,
    token: Optional[str] = Cookie(None, alias="grc_auth_token"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Get all users from the tenant schema for owner selection"""
    try:
        auth_token = token
        if not auth_token:
            auth_header = http_request.headers.get("authorization", "")
            if auth_header.startswith("Bearer "):
                auth_token = auth_header[7:]
        
        if auth_token:
            payload = decode_token(auth_token)
            schema_name = payload.get("schema")
            if schema_name:
                TenantSession = get_tenant_session(schema_name)
                tenant_session = TenantSession()
                try:
                    tenant_session.execute(text(f'SET search_path TO \"{schema_name}\", public'))
                    tenant_users = tenant_session.query(TenantSchemaUser).filter(
                        TenantSchemaUser.is_active == True
                    ).all()
                    return [
                        {
                            "id": u.id,
                            "display_name": u.display_name or u.username,
                            "email": u.email
                        }
                        for u in tenant_users
                    ]
                finally:
                    tenant_session.close()
    except Exception as e:
        print(f"[tenant-users] Tenant schema lookup failed: {e}")
    
    # Fallback: return public GRCUser users
    users = db.query(GRCUser).filter(GRCUser.is_active == True).all()
    return [
        {
            "id": u.id,
            "display_name": u.display_name or u.username,
            "email": u.email
        }
        for u in users
    ]


class CIARecommendationRequest(BaseModel):
    name: str
    description: Optional[str] = None
    asset_type: str
    vendor: Optional[str] = None
    location: Optional[str] = None
    criticality: Optional[str] = None


@router.post("/cia-recommendation")
def get_cia_recommendation(
    request: CIARecommendationRequest,
    current_user: GRCUser = Depends(require_auth)
):
    """Get AI-driven CIA rating recommendations based on asset details"""
    
    # Check if OpenAI API key is available
    openai_key = os.getenv("OPENAI_API_KEY")
    
    if not openai_key:
        # Fallback to rule-based recommendations
        return get_rule_based_cia_recommendation(request)
    
    try:
        import openai
        openai.api_key = openai_key
        
        # Build prompt for AI
        prompt = f"""Based on the following IT asset information, recommend appropriate CIA (Confidentiality, Integrity, Availability) ratings on a scale of 1-5 where:
- 1 = Very Low
- 2 = Low  
- 3 = Medium
- 4 = High
- 5 = Very High/Critical

Asset Details:
- Name: {request.name}
- Type: {request.asset_type}
- Description: {request.description or 'Not provided'}
- Vendor: {request.vendor or 'Not provided'}
- Location: {request.location or 'Not provided'}
- Business Criticality: {request.criticality or 'Not specified'}

Provide a brief 2-line recommendation explaining the suggested CIA ratings (Confidentiality, Integrity, Availability) and the specific numeric ratings.
Format: First line with explanation, second line with ratings in format "Recommended: C=X, I=Y, A=Z"
"""
        
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an IT security expert specializing in asset risk assessment and CIA triad ratings."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=200
        )
        
        recommendation_text = response.choices[0].message.content.strip()
        
        # Parse ratings from response
        lines = recommendation_text.split('\n')
        rating_line = next((line for line in lines if 'C=' in line and 'I=' in line and 'A=' in line), '')
        
        conf_rating = 3
        int_rating = 3
        avail_rating = 3
        
        if rating_line:
            import re
            conf_match = re.search(r'C=(\d)', rating_line)
            int_match = re.search(r'I=(\d)', rating_line)
            avail_match = re.search(r'A=(\d)', rating_line)
            
            if conf_match:
                conf_rating = int(conf_match.group(1))
            if int_match:
                int_rating = int(int_match.group(1))
            if avail_match:
                avail_rating = int(avail_match.group(1))
        
        # Format recommendation as 2 lines
        recommendation_lines = [line for line in lines if line.strip() and not line.startswith('Recommended:')][:2]
        if len(recommendation_lines) < 2 and rating_line:
            recommendation_lines.append(rating_line)
        
        recommendation = '\n'.join(recommendation_lines[:2])
        
        return {
            "recommendation": recommendation,
            "confidentiality_rating": conf_rating,
            "integrity_rating": int_rating,
            "availability_rating": avail_rating
        }
        
    except Exception as e:
        # Fallback to rule-based on error
        return get_rule_based_cia_recommendation(request)


def get_rule_based_cia_recommendation(request: CIARecommendationRequest):
    """Rule-based CIA recommendations when AI is unavailable"""
    
    asset_type = request.asset_type.lower()
    criticality = (request.criticality or 'medium').lower()
    
    # Base ratings
    conf_rating = 3
    int_rating = 3
    avail_rating = 3
    
    # Adjust by asset type
    if asset_type == 'data':
        conf_rating = 5
        int_rating = 5
        avail_rating = 4
        recommendation = "Data assets require maximum protection for confidentiality and integrity to prevent unauthorized access and corruption.\nRecommended: C=5, I=5, A=4"
    
    elif asset_type == 'application':
        conf_rating = 4
        int_rating = 4
        avail_rating = 4
        recommendation = "Business applications need high protection across all CIA dimensions to ensure secure and reliable operations.\nRecommended: C=4, I=4, A=4"
    
    elif asset_type == 'infrastructure':
        conf_rating = 3
        int_rating = 4
        avail_rating = 5
        recommendation = "Infrastructure assets prioritize availability and integrity to maintain operational continuity and system reliability.\nRecommended: C=3, I=4, A=5"
    
    elif asset_type == 'cloud':
        conf_rating = 4
        int_rating = 4
        avail_rating = 5
        recommendation = "Cloud resources require high availability and strong confidentiality controls due to shared responsibility model.\nRecommended: C=4, I=4, A=5"
    
    elif asset_type == 'third_party':
        conf_rating = 4
        int_rating = 3
        avail_rating = 3
        recommendation = "Third-party systems need elevated confidentiality controls due to external access and data sharing requirements.\nRecommended: C=4, I=3, A=3"
    
    else:
        recommendation = "Standard protection levels recommended based on general IT asset best practices and industry standards.\nRecommended: C=3, I=3, A=3"
    
    # Adjust by criticality
    if criticality == 'critical':
        conf_rating = min(5, conf_rating + 1)
        int_rating = min(5, int_rating + 1)
        avail_rating = min(5, avail_rating + 1)
    elif criticality == 'high':
        avail_rating = min(5, avail_rating + 1)
    elif criticality == 'low':
        conf_rating = max(1, conf_rating - 1)
        int_rating = max(1, int_rating - 1)
    
    return {
        "recommendation": recommendation,
        "confidentiality_rating": conf_rating,
        "integrity_rating": int_rating,
        "availability_rating": avail_rating
    }


def _user_tenant_role_names(current_user: GRCUser, tenant_id: int, db: Session) -> List[str]:
    """Return the tenant-scoped role names for the current user.

    Roles live in `tenant_<slug>.roles` and join via `tenant_<slug>.users.id`,
    NOT `public.grc_users.id` — those id sequences are independent. So we
    match by email, which is unique in both tables.

    Returns an empty list if the user has no role in this tenant (which we
    treat as "no access" for scoping purposes).
    """
    tenant_row = db.execute(text("SELECT slug FROM grc_tenants WHERE id=:tid"),
                            {"tid": tenant_id}).fetchone()
    if not tenant_row:
        return []
    schema = "tenant_" + tenant_row[0].replace("-", "")
    try:
        rows = db.execute(
            text(
                f"SELECT r.name FROM {schema}.users u "
                f"JOIN {schema}.user_roles ur ON ur.user_id = u.id "
                f"JOIN {schema}.roles r ON r.id = ur.role_id "
                f"WHERE u.email = :email"
            ),
            {"email": current_user.email},
        ).fetchall()
        return [r[0] for r in rows]
    except Exception:
        # Tenant schema missing/misconfigured shouldn't crash the request; fall
        # back to "no roles" which scopes the list down. Safer than leaking.
        return []


# Roles that operate over the WHOLE tenant — they don't get data-scoped.
# Auditors get audit-independence and need to see everything; Admin and
# Scanning Admin own the day-to-day ops over the whole estate.
TENANT_WIDE_ROLES = {"Administrator", "Auditor", "Scanning Admin"}


def _should_scope_to_owner(current_user: GRCUser, tenant_id: int, db: Session) -> bool:
    """True if this user should only see resources they own (e.g. Banking User).

    A user is data-scoped iff NONE of their roles in this tenant grant
    tenant-wide visibility. New roles are scoped by default — safer to err
    on the side of less visible, more visible needs explicit role config.
    """
    roles = _user_tenant_role_names(current_user, tenant_id, db)
    if not roles:
        # No roles in this tenant — treat as scoped (won't see anything anyway,
        # but defensive in case role resolution is the only thing missing).
        return True
    return not any(r in TENANT_WIDE_ROLES for r in roles)


@router.get("", response_model=List[ITAssetResponse])
def list_assets(
    tenant_id: Optional[int] = None,
    asset_type: Optional[str] = None,
    criticality: Optional[str] = None,
    owner_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []

    query = db.query(ITAsset).filter(ITAsset.tenant_id.in_(user_tenants))

    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    if asset_type:
        query = query.filter(ITAsset.asset_type == asset_type)
    if criticality:
        query = query.filter(ITAsset.criticality == criticality)
    if owner_id:
        query = query.filter(ITAsset.owner_id == owner_id)
    if status_filter:
        query = query.filter(ITAsset.status == status_filter)

    # ─── Data scoping ─────────────────────────────────────────────────────
    # Banking User (and any other non-tenant-wide role) only sees assets they
    # own. Administrator / Scanning Admin / Auditor get the full tenant
    # estate. Owner = `owner_id` matches the GRCUser id, OR `custodian` /
    # `owner_name` matches the user's email/display_name (legacy assets
    # without a hard owner_id link). We default to id-only matching — the
    # safer, unambiguous path — and let an operator backfill owner_id if
    # they want a Banking User to see legacy rows.
    primary_tid = tenant_id or get_user_primary_tenant(current_user, db)
    if primary_tid and _should_scope_to_owner(current_user, primary_tid, db):
        query = query.filter(ITAsset.owner_id == current_user.id)

    assets = query.order_by(ITAsset.created_at.desc()).offset(skip).limit(limit).all()
    return assets


@router.post("", response_model=ITAssetResponse, status_code=status.HTTP_201_CREATED)
def create_asset(
    asset: ITAssetCreate,
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tenant not found"
        )
    
    # Duplicate prevention: block same name + IP within the same tenant
    if asset.ip_address:
        existing = (
            db.query(ITAsset)
            .filter(
                ITAsset.tenant_id == tenant_id,
                ITAsset.name == asset.name,
                ITAsset.ip_address == asset.ip_address,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Asset '{asset.name}' with IP {asset.ip_address} already exists in this tenant (id={existing.id}). Duplicate assets at the same IP are not allowed.",
            )

    db_asset = ITAsset(
        tenant_id=tenant_id,
        name=asset.name,
        description=asset.description,
        asset_type=asset.asset_type,
        owner_id=asset.owner_id,
        owner_name=asset.owner_name,
        custodian=asset.custodian,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        criticality=asset.criticality,
        vendor=asset.vendor,
        location=asset.location,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
        valuation=asset.valuation,
        cde_environment=asset.cde_environment,
        # Block D — accept OS fields on create so strict matcher resolves
        # immediately. If os_normalized is empty but os_version is provided,
        # we auto-classify below before commit (same path agents use).
        os_family=asset.os_family,
        os_version=asset.os_version,
        # SAFEGUARD: truncate any incoming os_normalized to the major.minor
        # convention. Manual create can carry "ubuntu-24.04.2" from operator
        # input; we collapse to "ubuntu-24.04" so the matcher resolves
        # deterministically against CIS benchmarks.
        os_normalized=truncate_os_key_to_second_level(asset.os_normalized) if asset.os_normalized else None,
        os_build=asset.os_build,
        os_edition=asset.os_edition,
    )

    # Auto-resolve owner_name from owner_id if not provided
    if asset.owner_id and not asset.owner_name:
        owner = db.query(GRCUser).filter(GRCUser.id == asset.owner_id).first()
        if owner:
            db_asset.owner_name = owner.display_name or owner.username

    # Gap A fix — auto-classify OS so manually-created assets resolve to
    # a benchmark without an extra "Re-detect OS" click. Same code path
    # the agent heartbeat + connect wizard already use, kept best-effort:
    # the OS normaliser hits OpenAI and may fail (no key, network, etc.).
    # A failure here MUST NOT block the asset create — the operator can
    # still classify manually later.
    if asset.os_version and not asset.os_normalized:
        # Regex FIRST: deterministic, instant, no API key needed. The
        # previous AI-only path left manually-created assets stuck at
        # "undetected" whenever OPENAI_API_KEY was absent or the call
        # failed — meaning no benchmark, no rules, dead asset.
        try:
            from grc.modules.compliance_plugins.services.os_detector import normalise_os_freetext
            regex_key = normalise_os_freetext(asset.os_version)
        except Exception:  # noqa: BLE001
            regex_key = None
        if regex_key:
            db_asset.os_normalized = truncate_os_key_to_second_level(regex_key)
            if not db_asset.os_family:
                db_asset.os_family = regex_key.split("-", 1)[0]
        else:
            # AI fallback for strings regex can't parse.
            try:
                from grc.modules.compliance_plugins.services.ai_os_normaliser import (
                    normalise_os_string,
                )
                normed = normalise_os_string(asset.os_version) or {}
                if normed.get("normalized"):
                    db_asset.os_normalized = truncate_os_key_to_second_level(normed["normalized"])
                    if not db_asset.os_family:
                        db_asset.os_family = normed["normalized"].split("-", 1)[0]
            except Exception:  # noqa: BLE001
                import logging
                logging.getLogger(__name__).warning(
                    "OS auto-classify failed for asset %r (os_version=%r); "
                    "operator can use 'Re-detect OS' later.",
                    asset.name, asset.os_version,
                )

    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)

    write_rich_audit_log(
        db=db,
        tenant_id=tenant_id,
        user_id=current_user.id,
        action="create",
        resource_type="assets",
        resource_id=db_asset.id,
        resource_name=db_asset.name,
        summary=f"Created Asset '{db_asset.name}' (type: {db_asset.asset_type}, criticality: {db_asset.criticality})",
        snapshot=model_to_dict(db_asset),
        resource_url=f"/assets/{db_asset.id}",
    )
    db.commit()

    return db_asset


@router.get("/dashboard", response_model=AssetDashboard)
def get_asset_dashboard(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return AssetDashboard(
            total_assets=0,
            by_type={},
            by_criticality={},
            by_status={},
            high_value_assets=0,
            assets_needing_assessment=0
        )
    
    query = db.query(ITAsset).filter(ITAsset.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    
    assets = query.all()
    total = len(assets)
    
    by_type = {}
    by_criticality = {}
    by_status = {}
    high_value_assets = 0
    assets_needing_assessment = 0
    
    for asset in assets:
        by_type[asset.asset_type] = by_type.get(asset.asset_type, 0) + 1
        by_criticality[asset.criticality] = by_criticality.get(asset.criticality, 0) + 1
        by_status[asset.status] = by_status.get(asset.status, 0) + 1
        
        if asset.criticality in ["high", "critical"]:
            high_value_assets += 1
        
        if not asset.risk_assessments:
            assets_needing_assessment += 1
    
    return AssetDashboard(
        total_assets=total,
        by_type=by_type,
        by_criticality=by_criticality,
        by_status=by_status,
        high_value_assets=high_value_assets,
        assets_needing_assessment=assets_needing_assessment
    )


@router.get("/coverage", response_model=AssetCoverage)
def get_asset_coverage(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return AssetCoverage(
            total_assets=0,
            assets_with_controls=0,
            coverage_percentage=0.0,
            by_criticality={}
        )
    
    query = db.query(ITAsset).options(joinedload(ITAsset.control_links)).filter(ITAsset.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    
    assets = query.all()
    total = len(assets)
    assets_with_controls = sum(1 for a in assets if a.control_links)
    
    coverage_by_criticality = {}
    for asset in assets:
        if asset.criticality not in coverage_by_criticality:
            coverage_by_criticality[asset.criticality] = {
                "total": 0,
                "with_controls": 0,
                "coverage_percentage": 0.0
            }
        coverage_by_criticality[asset.criticality]["total"] += 1
        if asset.control_links:
            coverage_by_criticality[asset.criticality]["with_controls"] += 1
    
    for crit, data in coverage_by_criticality.items():
        if data["total"] > 0:
            data["coverage_percentage"] = round(
                (data["with_controls"] / data["total"]) * 100, 2
            )
    
    return AssetCoverage(
        total_assets=total,
        assets_with_controls=assets_with_controls,
        coverage_percentage=round((assets_with_controls / total) * 100, 2) if total > 0 else 0.0,
        by_criticality=coverage_by_criticality
    )


ASSET_TEMPLATE_COLUMNS = [
    ("name", "Asset Name (Required)", "ERP System"),
    ("description", "Description", "Enterprise Resource Planning system for finance and operations"),
    ("asset_type", "Asset Type (Required: application/infrastructure/data/cloud/third_party)", "application"),
    ("criticality", "Criticality (low/medium/high/critical)", "high"),
    ("vendor", "Vendor Name", "SAP"),
    ("location", "Location", "Primary Data Center"),
    ("confidentiality_rating", "Confidentiality Rating (1-5)", "4"),
    ("integrity_rating", "Integrity Rating (1-5)", "5"),
    ("availability_rating", "Availability Rating (1-5)", "5"),
    ("valuation", "Valuation (USD)", "500000"),
    ("status", "Status (active/inactive/decommissioned)", "active"),
    ("cde_environment", "CDE Environment (true/false)", "false"),
    # OS fields — operators include these to pre-classify hosts. The
    # backend collapses os_normalized to level-2 (major.minor) on import
    # so "ubuntu-24.04" and "ubuntu-24.04.2" map to the same benchmark.
    ("host_name", "Host name / FQDN", "web-prod-01.bank.local"),
    ("ip_address", "IP address", "10.0.5.12"),
    ("os_family", "OS family (windows/linux/macos/cisco)", "linux"),
    ("os_version", "OS version (free text)", "Ubuntu 24.04.2 LTS"),
    ("os_normalized", "OS key — major.minor (e.g. ubuntu-24.04, windows-11-23H2)", "ubuntu-24.04"),
]


@router.get("/template/download")
def download_asset_template(
    current_user: GRCUser = Depends(require_auth)
):
    """Download CSV template for bulk asset import"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [col[0] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(headers)
    
    descriptions = [col[1] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(descriptions)
    
    examples = [col[2] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(examples)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=it_assets_template.csv"}
    )


@router.post("/import/upload")
async def upload_assets_file(
    file: UploadFile = File(...),
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Upload CSV or Excel file to bulk import IT assets"""
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided"
        )
    
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be CSV or Excel format (.csv, .xlsx, .xls)"
        )
    
    content = await file.read()
    rows = []
    
    try:
        if filename.endswith('.csv'):
            text_content = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            
            headers = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    continue
                if idx == 1:
                    first_cell = str(row[0]).lower() if row[0] else ""
                    if "required" in first_cell or "description" in first_cell or "name" in first_cell:
                        continue
                
                if not any(row):
                    continue
                    
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = cell
                rows.append(row_dict)
            wb.close()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error parsing file: {str(e)}"
        )
    
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file"
        )
    
    valid_asset_types = ["application", "infrastructure", "data", "cloud", "third_party"]
    valid_criticality = ["low", "medium", "high", "critical"]
    valid_status = ["active", "inactive", "decommissioned"]
    
    imported = []
    errors = []
    
    for idx, row in enumerate(rows, start=1):
        row_num = idx + 2
        
        name = str(row.get("name", "")).strip() if row.get("name") else ""
        if not name:
            errors.append({"row": row_num, "error": "Name is required"})
            continue
        
        asset_type = str(row.get("asset_type", "")).strip().lower() if row.get("asset_type") else ""
        if not asset_type:
            errors.append({"row": row_num, "error": "Asset type is required"})
            continue
        if asset_type not in valid_asset_types:
            errors.append({"row": row_num, "error": f"Invalid asset_type '{asset_type}'. Must be one of: {', '.join(valid_asset_types)}"})
            continue
        
        criticality = str(row.get("criticality", "medium")).strip().lower() if row.get("criticality") else "medium"
        if criticality not in valid_criticality:
            criticality = "medium"
        
        asset_status = str(row.get("status", "active")).strip().lower() if row.get("status") else "active"
        if asset_status not in valid_status:
            asset_status = "active"
        
        def parse_int(val, min_val=None, max_val=None):
            if val is None or val == "":
                return None
            try:
                v = int(float(str(val)))
                if min_val is not None and v < min_val:
                    return min_val
                if max_val is not None and v > max_val:
                    return max_val
                return v
            except:
                return None
        
        def parse_float(val):
            if val is None or val == "":
                return None
            try:
                return float(str(val).replace(",", ""))
            except:
                return None
        
        try:
            cde_raw = str(row.get("cde_environment", "")).strip().lower() if row.get("cde_environment") else ""
            cde_flag = cde_raw in ("true", "yes", "1", "y")

            # OS fields from CSV — operators include these to pre-classify
            # bulk-imported hosts. Both go through level-2 truncation so
            # the matcher resolves the same regardless of how precise the
            # CSV is ("ubuntu-24.04" or "ubuntu-24.04.2" both end up
            # "ubuntu-24.04").
            csv_os_version    = str(row.get("os_version", "")).strip() if row.get("os_version") else None
            csv_os_normalized = str(row.get("os_normalized", "")).strip() if row.get("os_normalized") else None
            csv_os_family     = str(row.get("os_family", "")).strip() if row.get("os_family") else None
            if csv_os_normalized:
                csv_os_normalized = truncate_os_key_to_second_level(csv_os_normalized)

            asset = ITAsset(
                tenant_id=tenant_id,
                name=name,
                description=str(row.get("description", "")).strip() if row.get("description") else None,
                asset_type=asset_type,
                criticality=criticality,
                vendor=str(row.get("vendor", "")).strip() if row.get("vendor") else None,
                location=str(row.get("location", "")).strip() if row.get("location") else None,
                confidentiality_rating=parse_int(row.get("confidentiality_rating"), 1, 5),
                integrity_rating=parse_int(row.get("integrity_rating"), 1, 5),
                availability_rating=parse_int(row.get("availability_rating"), 1, 5),
                valuation=parse_float(row.get("valuation")),
                status=asset_status,
                cde_environment=cde_flag,
                os_family=csv_os_family,
                os_version=csv_os_version,
                os_normalized=csv_os_normalized,
                host_name=str(row.get("host_name", "")).strip() if row.get("host_name") else None,
                ip_address=str(row.get("ip_address", "")).strip() if row.get("ip_address") else None,
            )
            db.add(asset)
            db.flush()
            imported.append({"id": asset.id, "name": asset.name})
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})
    
    db.commit()
    
    imported_count = len(imported)
    error_count = len(errors)
    error_messages = [f"Row {e['row']}: {e['error']}" for e in errors[:20]]
    
    return {
        "success": True,
        "imported": imported_count,
        "total_rows": len(rows),
        "errors": error_messages,
        "total_errors": error_count,
        "message": f"Successfully imported {imported_count} assets" + (f" with {error_count} errors" if error_count > 0 else "")
    }


# ─── Host → Application children ("room and chair") ──────────────────────

class PromoteSoftwareIn(BaseModel):
    """Selected inventory entries to promote into child assets."""
    software_keys: List[str]
    criticality: Optional[str] = None  # default: inherit parent's


def _tenant_asset_or_404(db: Session, current_user, asset_id: int) -> ITAsset:
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
    return asset


@router.get("/os-catalog")
def get_os_catalog(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """OS-level keys that have benchmarks, grouped by family — drives the
    structured Family → Version dropdowns on the Add Asset form. Built
    live from the plugin library so a newly ingested OS benchmark shows
    up as a pickable option with zero extra wiring.

    Replaces free-text-only OS entry: operators picking from a list can't
    typo their way into an unmatchable asset.
    """
    tenant_id = get_user_primary_tenant(current_user, db)
    rows = db.execute(text("""
        SELECT k AS os_key, COUNT(*) AS rule_count
        FROM grc_compliance_plugins,
             jsonb_array_elements_text(os_keys) AS k
        WHERE enabled = TRUE
          AND review_status IN ('approved', 'auto_approved')
          AND (tenant_id IS NULL OR tenant_id = :tid)
          AND jsonb_typeof(os_keys) = 'array'
        GROUP BY k
        ORDER BY k
    """), {"tid": tenant_id}).fetchall()

    def family_of(k: str):
        if k.startswith("windows-server"):
            return "Windows Server"
        if k.startswith("windows-"):
            return "Windows"
        if k.startswith("macos-"):
            return "macOS"
        if k.startswith(("ubuntu-", "debian-", "rhel-", "centos-", "almalinux-",
                         "rocky-", "oraclelinux-", "amazonlinux-", "sles-",
                         "suse-", "fedora-")) or k in ("freebsd", "aliyun",
                         "alibaba-linux", "chromeos", "google-cos",
                         "bottlerocket", "ibm-aix"):
            return "Linux / Unix"
        if k.startswith(("cisco-", "fortigate", "paloalto", "juniper",
                         "pfsense", "f5", "hpe-aruba")):
            return "Network / Firewall"
        if k.startswith("vmware-"):
            return "Hypervisor"
        if k == "ibm-zos":
            return "Mainframe"
        return None  # application/cloud keys — not OS options

    def prettify(k: str) -> str:
        # windows-server-2022 → "Windows Server 2022"; ubuntu-24.04 → "Ubuntu 24.04"
        words = []
        for part in k.split("-"):
            if re.match(r"^\d", part) or part.upper() in ("XE", "XR"):
                words.append(part.upper() if part.isalpha() else part)
            elif part in ("ios", "nx", "os", "asa"):
                words.append(part.upper())
            else:
                words.append(part.capitalize())
        return " ".join(words)

    groups: dict[str, list] = {}
    for key, n in rows:
        fam = family_of(key)
        if fam is None:
            continue
        groups.setdefault(fam, []).append({
            "key": key, "label": prettify(key), "rule_count": n,
        })
    order = ["Windows", "Windows Server", "Linux / Unix", "macOS",
             "Network / Firewall", "Hypervisor", "Mainframe"]
    return {
        "families": [
            {"family": f, "options": sorted(groups[f], key=lambda o: o["key"])}
            for f in order if f in groups
        ]
    }


@router.get("/software-catalog")
def get_software_catalog(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """All APPLICATION software_keys that have CIS benchmarks in the
    library — the manual-add dropdown. Excludes OS-level keys (those
    belong to the host itself, not to applications running on it).

    Derived live from the plugin library so newly ingested benchmarks
    (e.g. a fresh Tomcat PDF) appear here with zero extra wiring.
    """
    tenant_id = get_user_primary_tenant(current_user, db)
    rows = db.execute(text("""
        SELECT k AS software_key,
               MIN(benchmark) AS benchmark_name,
               COUNT(*) AS rule_count
        FROM grc_compliance_plugins,
             jsonb_array_elements_text(os_keys) AS k
        WHERE enabled = TRUE
          AND review_status IN ('approved', 'auto_approved')
          AND (tenant_id IS NULL OR tenant_id = :tid)
          AND jsonb_typeof(os_keys) = 'array'
        GROUP BY k
        ORDER BY k
    """), {"tid": tenant_id}).fetchall()
    # OS-level prefixes — these are host keys, not application keys.
    _OS_PREFIXES = (
        "windows-", "ubuntu-", "debian-", "rhel-", "centos-", "almalinux-",
        "rocky-", "oraclelinux-", "amazonlinux-", "sles-", "suse-", "aliyun",
        "alibaba-linux", "fedora-", "freebsd", "macos-", "cisco-", "chromeos",
        "ibm-aix", "ibm-zos", "aks-", "google-cos", "bottlerocket",
        "fortigate", "paloalto", "juniper", "pfsense", "f5", "hpe-aruba",
        "vmware-esxi", "aws", "azure", "gcp", "alibaba", "digitalocean",
        "ibm-cloud", "google-workspace", "microsoft-365", "github",
    )
    catalog = [
        {"software_key": r[0], "benchmark_name": r[1], "rule_count": r[2]}
        for r in rows
        if not any(r[0].startswith(p) for p in _OS_PREFIXES)
    ]
    return {"catalog": catalog}


class AddApplicationIn(BaseModel):
    """Manually attach an application child to a host — used when the
    agent didn't detect it (agentless host, portable install the scanner
    missed) or the operator simply knows it's there."""
    software_key: str
    display_name: Optional[str] = None
    criticality: Optional[str] = None


@router.post("/{asset_id}/add-application")
def add_application_manually(
    asset_id: int,
    body: AddApplicationIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Create an application child asset by hand. Same inheritance rules
    as promote-software; the software_key doubles as the match key so
    the benchmark resolves immediately. Duplicate keys on the same host
    are rejected (one SQL Server child per host — rule results would
    collide otherwise)."""
    asset = _tenant_asset_or_404(db, current_user, asset_id)
    key = (body.software_key or "").strip().lower()
    if not key:
        raise HTTPException(status_code=400, detail="software_key is required")
    dup = db.query(ITAsset).filter(
        ITAsset.parent_asset_id == asset.id,
        ITAsset.os_normalized == key,
    ).first()
    if dup:
        raise HTTPException(
            status_code=409,
            detail=f"This host already has a child asset for '{key}' (asset #{dup.id})",
        )
    label = body.display_name or key.replace("-", " ").title()
    child = ITAsset(
        tenant_id=asset.tenant_id,
        name=f"{label} @ {asset.host_name or asset.name}",
        description=f"Application asset added manually to host '{asset.name}'",
        asset_type="application",
        asset_role="application",
        parent_asset_id=asset.id,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        os_family=asset.os_family,
        os_version=label,
        os_normalized=key,
        criticality=body.criticality or asset.criticality,
        status="active",
        owner_id=asset.owner_id,
        owner_name=asset.owner_name,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
    )
    db.add(child)
    db.flush()
    # If this key was sitting in detected inventory, link it so the
    # "awaiting protection" list updates too.
    import copy as _copy
    inventory = _copy.deepcopy(asset.detected_software_json or [])
    changed = False
    for e in inventory:
        if e.get("software_key") == key and not e.get("promoted_asset_id"):
            e["promoted_asset_id"] = child.id
            changed = True
    if changed:
        asset.detected_software_json = inventory
    db.commit()
    return {"asset_id": child.id, "name": child.name, "software_key": key}


@router.get("/{asset_id}/detected-software")
def get_detected_software(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """The agent's enriched software inventory for this host: each entry
    carries software_key, benchmark availability, and (if already
    promoted) the child asset id. Drives the 'Detected on this server'
    panel."""
    asset = _tenant_asset_or_404(db, current_user, asset_id)
    inventory = asset.detected_software_json or []
    promotable = [e for e in inventory if e.get("benchmark_available") and not e.get("promoted_asset_id")]
    return {
        "asset_id": asset.id,
        "inventory": inventory,
        "counts": {
            "total": len(inventory),
            "promotable": len(promotable),
            "promoted": sum(1 for e in inventory if e.get("promoted_asset_id")),
            "no_benchmark": sum(1 for e in inventory if not e.get("benchmark_available")),
        },
    }


@router.post("/{asset_id}/promote-software")
def promote_software(
    asset_id: int,
    body: PromoteSoftwareIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Turn selected detected-software entries into child assets.

    Each promoted child gets:
      parent_asset_id = this host, asset_role = 'application',
      os_normalized   = the software_key (level-2 — matcher resolves the
                        application benchmark exactly like an OS),
      owner + criticality + business-impact inherited from the parent.
    Idempotent: an entry already promoted is skipped, not duplicated.
    """
    asset = _tenant_asset_or_404(db, current_user, asset_id)
    # DEEP copy, not list(): a shallow copy shares the entry dicts with
    # the ORM attribute's current value, so after in-place mutation the
    # old and new values compare equal and SQLAlchemy skips the UPDATE —
    # promotions would silently not persist.
    import copy as _copy
    inventory = _copy.deepcopy(asset.detected_software_json or [])
    by_key = {e.get("software_key"): e for e in inventory}
    created, skipped = [], []
    for key in body.software_keys:
        entry = by_key.get(key)
        if entry is None:
            skipped.append({"software_key": key, "reason": "not in inventory"})
            continue
        if entry.get("promoted_asset_id"):
            skipped.append({"software_key": key, "reason": "already promoted",
                            "asset_id": entry["promoted_asset_id"]})
            continue
        child = ITAsset(
            tenant_id=asset.tenant_id,
            name=f"{entry.get('name') or key} @ {asset.host_name or asset.name}",
            description=(f"Application asset promoted from detected software on "
                         f"'{asset.name}' (source: {entry.get('source')})"),
            asset_type="application",
            asset_role="application",
            parent_asset_id=asset.id,
            host_name=asset.host_name,
            ip_address=asset.ip_address,
            os_family=asset.os_family,
            os_version=entry.get("version") and f"{entry.get('name')} {entry.get('version')}" or entry.get("name"),
            os_normalized=key,  # software_key doubles as match key — same matcher path
            criticality=body.criticality or asset.criticality,
            status="active",
            owner_id=asset.owner_id,
            owner_name=asset.owner_name,
            # Inherit the room's business impact — the chair matters as
            # much as the room it's in (overridable on the child later).
            confidentiality_rating=asset.confidentiality_rating,
            integrity_rating=asset.integrity_rating,
            availability_rating=asset.availability_rating,
        )
        db.add(child)
        db.flush()
        entry["promoted_asset_id"] = child.id
        created.append({"software_key": key, "asset_id": child.id, "name": child.name})
    # JSON column: reassign so SQLAlchemy sees the mutation
    asset.detected_software_json = inventory
    db.commit()
    return {"created": created, "skipped": skipped}


# Composite scoring weights — the "room and chair" formula.
#   effective = W_SELF × own + W_CHILDREN × criticality-weighted child avg
#   − BROKEN_CHAIR_PENALTY  if any child scores < 50%
#   − UNKNOWN_CHAIR_PENALTY if promotable software is still unpromoted
#     (an unprotected chair we KNOW about must hurt the room's number,
#      otherwise the 89% lies to the operator)
_COMPOSITE_W_SELF = 0.6
_COMPOSITE_W_CHILDREN = 0.4
_BROKEN_CHAIR_PENALTY = 10.0   # percentage points
_UNKNOWN_CHAIR_PENALTY = 5.0
_CRIT_WEIGHT = {"low": 1.0, "medium": 2.0, "high": 3.0, "critical": 4.0}


def _get_tenant_crit_weights(db: Session, tid: int) -> dict:
    """Return the tenant's custom criticality weights, falling back to defaults."""
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    settings = tenant.settings or {} if tenant else {}
    return settings.get("composite_weights", _CRIT_WEIGHT)


class CompositeWeightsIn(BaseModel):
    low: float = 1.0
    medium: float = 2.0
    high: float = 3.0
    critical: float = 4.0


@router.get("/composite-weights")
def get_composite_weights(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    settings = tenant.settings or {} if tenant else {}
    is_custom = "composite_weights" in settings
    weights = settings.get("composite_weights", _CRIT_WEIGHT)
    return {"weights": weights, "is_custom": is_custom, "defaults": _CRIT_WEIGHT}


@router.put("/composite-weights")
def update_composite_weights(
    body: CompositeWeightsIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    weights = {"low": body.low, "medium": body.medium, "high": body.high, "critical": body.critical}
    if any(v <= 0 for v in weights.values()):
        raise HTTPException(400, "All weights must be positive numbers")
    settings = dict(tenant.settings or {})
    settings["composite_weights"] = weights
    tenant.settings = settings
    db.commit()
    return {"weights": weights, "is_custom": True, "defaults": _CRIT_WEIGHT}


@router.delete("/composite-weights")
def reset_composite_weights(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    settings = dict(tenant.settings or {})
    settings.pop("composite_weights", None)
    tenant.settings = settings
    db.commit()
    return {"weights": _CRIT_WEIGHT, "is_custom": False, "defaults": _CRIT_WEIGHT}


def _asset_own_compliance(db: Session, asset_id: int, tenant_id: int) -> Optional[float]:
    """Latest-run pass-rate for the asset's own rules (percent), or None
    when it has never been scanned. Uses each plugin's most recent
    non-leaked run, mirroring the asset compliance page."""
    rows = db.execute(text("""
        SELECT r.status FROM grc_compliance_plugin_runs r
        JOIN (
            SELECT plugin_id, MAX(id) AS max_id
            FROM grc_compliance_plugin_runs
            WHERE asset_id = :aid AND tenant_id = :tid
              AND is_leaked = false
              AND status IN ('passed', 'failed')
            GROUP BY plugin_id
        ) latest ON latest.max_id = r.id
    """), {"aid": asset_id, "tid": tenant_id}).fetchall()
    if not rows:
        return None
    passed = sum(1 for r in rows if r[0] == "passed")
    return round(100.0 * passed / len(rows), 1)


@router.get("/{asset_id}/children")
def get_asset_children(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Children tree + composite score for a host asset.

    Returns own_score, per-child scores, the blended effective_score, and
    which penalties applied — everything the composite header renders.
    """
    asset = _tenant_asset_or_404(db, current_user, asset_id)
    children = db.query(ITAsset).filter(
        ITAsset.parent_asset_id == asset.id,
        ITAsset.tenant_id == asset.tenant_id,
    ).order_by(ITAsset.id).all()

    own_score = _asset_own_compliance(db, asset.id, asset.tenant_id)
    crit_weights = _get_tenant_crit_weights(db, asset.tenant_id)
    child_rows = []
    weighted_sum, weight_total = 0.0, 0.0
    any_broken = False
    for c in children:
        score = _asset_own_compliance(db, c.id, c.tenant_id)
        w = crit_weights.get((c.criticality or "medium").lower(), 2.0)
        if score is not None:
            weighted_sum += score * w
            weight_total += w
            if score < 50.0:
                any_broken = True
        child_rows.append({
            "id": c.id, "name": c.name, "software_key": c.os_normalized,
            "criticality": c.criticality, "score": score,
            "never_scanned": score is None,
        })

    # Unpromoted-but-promotable software = known unprotected chairs
    inventory = asset.detected_software_json or []
    unpromoted = [e for e in inventory
                  if e.get("benchmark_available") and not e.get("promoted_asset_id")]

    effective = None
    penalties = []
    if own_score is not None:
        if weight_total > 0:
            children_avg = weighted_sum / weight_total
            effective = _COMPOSITE_W_SELF * own_score + _COMPOSITE_W_CHILDREN * children_avg
        else:
            effective = float(own_score)
        if any_broken:
            effective -= _BROKEN_CHAIR_PENALTY
            penalties.append({"reason": "child below 50% (broken chair)",
                              "points": _BROKEN_CHAIR_PENALTY})
        if unpromoted:
            effective -= _UNKNOWN_CHAIR_PENALTY
            penalties.append({"reason": f"{len(unpromoted)} detected application(s) not yet protected",
                              "points": _UNKNOWN_CHAIR_PENALTY})
        effective = round(max(0.0, min(100.0, effective)), 1)

    weakest = None
    scored = [c for c in child_rows if c["score"] is not None]
    if scored:
        weakest = min(scored, key=lambda c: c["score"])

    return {
        "asset_id": asset.id,
        "asset_role": asset.asset_role or "host",
        "parent_asset_id": asset.parent_asset_id,
        "own_score": own_score,
        "children": child_rows,
        "effective_score": effective,
        "weakest_child": weakest,
        "penalties": penalties,
        "unprotected_detected": [
            {"name": e.get("name"), "software_key": e.get("software_key"),
             "benchmark_name": e.get("benchmark_name")}
            for e in unpromoted
        ],
        "weights": {"self": _COMPOSITE_W_SELF, "children": _COMPOSITE_W_CHILDREN},
    }


@router.get("/{asset_id}/ip-peers")
def get_ip_peers(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """IP-based asset group: all assets co-located at the same IP address.

    Primary linking mechanism for inventory-imported assets (e.g. from the
    bank's third-party CMDB). Assets sharing an IP are assumed to run on the
    same physical/virtual host.

    Returns per-asset benchmark availability, individual compliance scores,
    and the composite IP-group score + formula.
    """
    from ..modules.compliance_plugins.services.strict_matcher import pick_benchmark_for_os
    from ..modules.compliance_plugins.services.software_normaliser import benchmark_for_software_key
    from ..models import CompliancePlugin

    asset = _tenant_asset_or_404(db, current_user, asset_id)
    tid = asset.tenant_id
    ip = asset.ip_address

    # Patterns that indicate an app-level software key (not an OS host)
    _APP_PATTERNS = (
        "sql", "postgres", "mysql", "mongo", "tomcat", "iis",
        "apache", "nginx", "redis", "oracle-db", "mssql", "mariadb",
        "jboss", "wildfly", "websphere", "weblogic",
    )

    def _benchmark_for(a: ITAsset):
        bname = None
        mapping = pick_benchmark_for_os(db, tid, a.os_normalized)
        if mapping:
            bname = mapping.benchmark_name
        if not bname and a.os_normalized:
            bname = benchmark_for_software_key(db, a.os_normalized)
        count = 0
        if bname:
            count = db.query(CompliancePlugin).filter(
                CompliancePlugin.benchmark == bname,
                CompliancePlugin.enabled.is_(True),
                (CompliancePlugin.tenant_id == tid) | (CompliancePlugin.tenant_id.is_(None)),
            ).count()
        return bname, count

    def _build(a: ITAsset):
        bname, rcount = _benchmark_for(a)
        score = _asset_own_compliance(db, a.id, tid)
        osk = (a.os_normalized or "").lower()
        is_host_os = (
            (a.asset_type or "") == "infrastructure" and
            not any(p in osk for p in _APP_PATTERNS)
        )
        return {
            "id": a.id,
            "name": a.name,
            "asset_type": a.asset_type or "infrastructure",
            "os_normalized": a.os_normalized,
            "criticality": a.criticality or "medium",
            "status": a.status,
            "benchmark_name": bname,
            "benchmark_available": bname is not None,
            "rule_count": rcount,
            "score": score,
            "never_scanned": score is None,
            "is_host_os": is_host_os,
            "is_self": a.id == asset_id,
        }

    crit_weights = _get_tenant_crit_weights(db, tid)

    if not ip:
        return {
            "asset_id": asset_id,
            "ip_address": None,
            "group": [_build(asset)],
            "composite": None,
            "formula": {
                "description": "Asset has no IP address — standalone, not part of an IP group.",
                "weights": {"host": _COMPOSITE_W_SELF, "applications": _COMPOSITE_W_CHILDREN},
                "criticality_weights": crit_weights,
            },
        }

    all_at_ip = db.query(ITAsset).filter(
        ITAsset.tenant_id == tid,
        ITAsset.ip_address == ip,
    ).order_by(ITAsset.asset_type, ITAsset.id).all()

    # Duplicate guard: deduplicate by id (shouldn't happen, but be safe)
    seen_ids = set()
    deduped = []
    for a in all_at_ip:
        if a.id not in seen_ids:
            seen_ids.add(a.id)
            deduped.append(a)

    group = [_build(a) for a in deduped]

    # ── Composite score (room-and-chair formula, IP-group edition) ────
    host_entry = next((g for g in group if g["is_host_os"]), None)
    app_entries = [g for g in group if not g["is_host_os"]]

    host_score = host_entry["score"] if host_entry else None
    app_contributions = []
    weighted_sum, weight_total = 0.0, 0.0
    any_broken = False
    for a in app_entries:
        w = crit_weights.get((a["criticality"] or "medium").lower(), 2.0)
        s = a["score"]
        app_contributions.append({
            "asset_id": a["id"], "name": a["name"],
            "score": s, "weight": w, "criticality": a["criticality"],
            "os_normalized": a.get("os_normalized"),
        })
        if s is not None:
            weighted_sum += s * w
            weight_total += w
            if s < 50.0:
                any_broken = True

    effective = None
    penalties = []
    if host_score is not None:
        if weight_total > 0:
            app_avg = weighted_sum / weight_total
            effective = _COMPOSITE_W_SELF * host_score + _COMPOSITE_W_CHILDREN * app_avg
        else:
            effective = float(host_score)
        if any_broken:
            effective -= _BROKEN_CHAIR_PENALTY
            penalties.append({"reason": "application scoring below 50%", "points": _BROKEN_CHAIR_PENALTY})
        effective = round(max(0.0, min(100.0, effective)), 1)
    elif app_entries:
        # No OS host in the group — equal-weight average of all apps
        scored = [a["score"] for a in app_entries if a["score"] is not None]
        if scored:
            effective = round(sum(scored) / len(scored), 1)

    weakest = None
    scored_all = [g for g in group if g["score"] is not None]
    if scored_all:
        weakest = min(scored_all, key=lambda g: g["score"])

    return {
        "asset_id": asset_id,
        "ip_address": ip,
        "group": group,
        "composite": {
            "host_id": host_entry["id"] if host_entry else None,
            "host_score": host_score,
            "app_contributions": app_contributions,
            "effective_score": effective,
            "weakest": weakest,
            "penalties": penalties,
        },
        "formula": {
            "description": (
                f"{int(_COMPOSITE_W_SELF * 100)}% host OS score "
                f"+ {int(_COMPOSITE_W_CHILDREN * 100)}% criticality-weighted application average"
            ),
            "weights": {"host": _COMPOSITE_W_SELF, "applications": _COMPOSITE_W_CHILDREN},
            "criticality_weights": crit_weights,
            "penalties": {
                "broken_app": f"−{_BROKEN_CHAIR_PENALTY} pts — any app scores below 50%",
            },
        },
    }


@router.get("/{asset_id}", response_model=dict)
def get_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.risk_links),
        joinedload(ITAsset.risk_assessments)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    latest_assessment = None
    if asset.risk_assessments:
        latest = sorted(asset.risk_assessments, key=lambda x: x.assessment_date, reverse=True)[0]
        latest_assessment = {
            "id": latest.id,
            "assessment_date": latest.assessment_date.isoformat(),
            "risk_score": latest.risk_score,
            "coverage_percentage": latest.coverage_percentage,
            "gaps": latest.gaps,
            "assessor_id": latest.assessor_id
        }
    
    return {
        "id": asset.id,
        "tenant_id": asset.tenant_id,
        "name": asset.name,
        "description": asset.description,
        "asset_type": asset.asset_type,
        "owner_id": asset.owner_id,
        "custodian": asset.custodian,
        "host_name": asset.host_name,
        "ip_address": asset.ip_address,
        "criticality": asset.criticality,
        "confidentiality_rating": asset.confidentiality_rating,
        "integrity_rating": asset.integrity_rating,
        "availability_rating": asset.availability_rating,
        "valuation": asset.valuation,
        "vendor": asset.vendor,
        "location": asset.location,
        "status": asset.status,
        "cde_environment": asset.cde_environment or False,
        "created_at": asset.created_at.isoformat(),
        "linked_controls": [link.normalized_control_id for link in asset.control_links],
        "linked_risks": [link.risk_id for link in asset.risk_links],
        "latest_assessment": latest_assessment
    }


@router.put("/{asset_id}", response_model=ITAssetResponse)
def update_asset(
    asset_id: int,
    asset_update: ITAssetUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    update_data = asset_update.model_dump(exclude_unset=True)
    _full_before = model_to_dict(asset)
    _before = {k: getattr(asset, k, None) for k in update_data}

    # SAFEGUARD: if the operator edits os_normalized via the form, truncate
    # to level-2 before storing (e.g. they paste "ubuntu-24.04.2" thinking
    # more precision = better; we collapse to "ubuntu-24.04" so the matcher
    # still resolves to the same CIS benchmark).
    if 'os_normalized' in update_data and update_data['os_normalized']:
        update_data['os_normalized'] = truncate_os_key_to_second_level(update_data['os_normalized'])

    for field, value in update_data.items():
        setattr(asset, field, value)
    
    # Auto-resolve owner_name when owner_id is updated without owner_name
    if 'owner_id' in update_data and 'owner_name' not in update_data and update_data.get('owner_id'):
        owner = db.query(GRCUser).filter(GRCUser.id == update_data['owner_id']).first()
        if owner:
            asset.owner_name = owner.display_name or owner.username
    
    db.commit()
    db.refresh(asset)

    _changed_fields = [k for k in update_data if _before.get(k) != update_data.get(k)]
    write_rich_audit_log(
        db=db,
        tenant_id=asset.tenant_id,
        user_id=current_user.id,
        action="update",
        resource_type="assets",
        resource_id=asset.id,
        resource_name=asset.name,
        summary=f"Updated Asset '{asset.name}': changed {', '.join(_changed_fields) or 'fields'}",
        before=_full_before,
        after=model_to_dict(asset),
        resource_url=f"/assets/{asset.id}",
    )
    db.commit()

    return asset


@router.delete("/{asset_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )

    _saved_id = asset.id
    _saved_name = asset.name
    _saved_tenant = asset.tenant_id
    _saved_snapshot = model_to_dict(asset)

    # ── Cascade cleanup ──────────────────────────────────────────────
    # The 11 child tables that reference grc_it_assets.id all use
    # NO ACTION on delete (set at migration time, never updated to
    # CASCADE). Without explicit cleanup here, the DELETE raises an
    # IntegrityError on any asset with linked records (agents, scan
    # runs, vulns, controls, etc.) — surfaces as a generic 500 to the
    # frontend with no useful detail.
    #
    # Rule of thumb per child table:
    #   - Pure link tables → DELETE the link, the linked-to entity stays
    #   - Asset-owned history → DELETE the rows (no point keeping orphaned
    #     scan runs whose asset is gone)
    #   - Agent rows → NULL the asset_id (agent itself stays — operator
    #     may want to re-link it; revoking is a separate action)
    from sqlalchemy import text as _sa_text
    from grc.models import (
        AssetControlLink, AssetEvidenceLink, AssetFrameworkControlLink,
        AssetInternalControlLink, AssetRiskAssessment, RiskAssetLink,
        VulnerabilityAssetLink, CompliancePluginRun,
    )
    # Link tables — delete the links only.
    for _tbl in (AssetControlLink, AssetEvidenceLink, AssetFrameworkControlLink,
                 AssetInternalControlLink, RiskAssetLink, VulnerabilityAssetLink):
        db.query(_tbl).filter(_tbl.asset_id == _saved_id).delete(synchronize_session=False)
    # Asset-owned history — delete.
    db.query(AssetRiskAssessment).filter(AssetRiskAssessment.asset_id == _saved_id).delete(synchronize_session=False)
    db.query(CompliancePluginRun).filter(CompliancePluginRun.asset_id == _saved_id).delete(synchronize_session=False)
    # Tables without a model export (less commonly imported) — use raw SQL.
    db.execute(_sa_text("DELETE FROM grc_asset_security_compliance_selections WHERE asset_id = :id"), {"id": _saved_id})
    db.execute(_sa_text("DELETE FROM grc_document_asset_links WHERE asset_id = :id"), {"id": _saved_id})
    # Agents — unlink rather than delete (operator can re-link or revoke).
    db.execute(_sa_text("UPDATE grc_compliance_agents SET asset_id = NULL WHERE asset_id = :id"), {"id": _saved_id})
    db.flush()

    db.delete(asset)
    write_rich_audit_log(
        db=db,
        tenant_id=_saved_tenant,
        user_id=current_user.id,
        action="delete",
        resource_type="assets",
        resource_id=_saved_id,
        resource_name=_saved_name,
        summary=f"Deleted Asset '{_saved_name}'",
        snapshot=_saved_snapshot,
        resource_url=f"/assets/{_saved_id}",
    )
    db.commit()
    return None


@router.post("/{asset_id}/valuation", response_model=ITAssetResponse)
def update_asset_valuation(
    asset_id: int,
    valuation: AssetValuation,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    asset.valuation = valuation.valuation
    if valuation.confidentiality_rating is not None:
        asset.confidentiality_rating = valuation.confidentiality_rating
    if valuation.integrity_rating is not None:
        asset.integrity_rating = valuation.integrity_rating
    if valuation.availability_rating is not None:
        asset.availability_rating = valuation.availability_rating
    
    db.commit()
    db.refresh(asset)
    return asset


@router.post("/{asset_id}/controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_control(
    asset_id: int,
    link: AssetControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    control = db.query(NormalizedControl).filter(
        NormalizedControl.id == link.normalized_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control not found"
        )
    
    existing = db.query(AssetControlLink).filter(
        AssetControlLink.asset_id == asset_id,
        AssetControlLink.normalized_control_id == link.normalized_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetControlLink(
        asset_id=asset_id,
        normalized_control_id=link.normalized_control_id
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Control linked successfully")


@router.post("/{asset_id}/assess", response_model=AssetRiskAssessmentResponse)
def assess_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    num_controls = len(asset.control_links) + len(asset.internal_control_links) + len(asset.framework_control_links)
    coverage = min(num_controls * 10, 100)
    
    base_risk = 5
    if asset.criticality == "critical":
        base_risk = 8
    elif asset.criticality == "high":
        base_risk = 6
    elif asset.criticality == "low":
        base_risk = 3
    
    risk_score = max(1, base_risk - (num_controls * 0.5))
    
    gaps = {
        "missing_controls": max(0, 10 - num_controls),
        "recommendations": []
    }
    if num_controls < 3:
        gaps["recommendations"].append("Add more controls to improve coverage")
    if asset.criticality in ["high", "critical"] and num_controls < 5:
        gaps["recommendations"].append("Critical assets should have at least 5 controls")
    
    assessment = AssetRiskAssessment(
        asset_id=asset_id,
        risk_score=round(risk_score, 2),
        coverage_percentage=float(coverage),
        gaps=gaps,
        assessor_id=current_user.id
    )
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    return assessment


@router.get("/{asset_id}/assessment", response_model=AssetRiskAssessmentResponse)
def get_latest_assessment(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    assessment = db.query(AssetRiskAssessment).filter(
        AssetRiskAssessment.asset_id == asset_id
    ).order_by(AssetRiskAssessment.assessment_date.desc()).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No assessment found for this asset"
        )
    
    return assessment


@router.get("/{asset_id}/detail", response_model=AssetDetailResponse)
def get_asset_detail(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.risk_links),
        joinedload(ITAsset.risk_assessments),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.evidence_links),
        joinedload(ITAsset.owner)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    linked_controls = []
    for link in asset.control_links:
        control = db.query(NormalizedControl).filter(NormalizedControl.id == link.normalized_control_id).first()
        if control:
            linked_controls.append({
                "id": link.id,
                "control_id": control.id,
                "code": control.code,
                "name": control.name
            })
    
    linked_internal_controls = []
    for link in asset.internal_control_links:
        ctrl = db.query(InternalControl).filter(InternalControl.id == link.internal_control_id).first()
        if ctrl:
            linked_internal_controls.append({
                "id": link.id,
                "internal_control_id": ctrl.id,
                "code": ctrl.control_id,
                "name": ctrl.name,
                "category": ctrl.category,
                "coverage_status": link.coverage_status
            })

    linked_framework_controls = []
    for link in asset.framework_control_links:
        fc = db.query(FrameworkControl).filter(FrameworkControl.id == link.framework_control_id).first()
        if fc:
            linked_framework_controls.append({
                "id": link.id,
                "framework_control_id": fc.id,
                "code": fc.code,
                "name": fc.name,
                "coverage_status": link.coverage_status,
                "notes": link.notes
            })
    
    linked_risks = []
    for link in asset.risk_links:
        risk = db.query(Risk).filter(
            Risk.id == link.risk_id,
            Risk.tenant_id.in_(user_tenants)
        ).first()
        if risk:
            linked_risks.append({
                "risk_id": risk.id,
                "title": risk.title,
                "status": risk.status,
                "inherent_score": risk.inherent_score,
                "residual_score": risk.residual_score,
            })
    
    linked_evidence = []
    for link in asset.evidence_links:
        ev = db.query(Evidence).filter(Evidence.id == link.evidence_id).first()
        if ev:
            linked_evidence.append({
                "id": link.id,
                "evidence_id": ev.id,
                "name": ev.name,
                "relationship_type": link.relationship_type
            })

    linked_vulnerabilities = []
    vuln_links = db.query(VulnerabilityAssetLink).options(
        joinedload(VulnerabilityAssetLink.vulnerability)
    ).join(Vulnerability).filter(
        VulnerabilityAssetLink.asset_id == asset_id,
        Vulnerability.tenant_id.in_(user_tenants)
    ).all()
    for link in vuln_links:
        vuln = link.vulnerability
        if vuln:
            linked_vulnerabilities.append({
                "link_id": link.id,
                "vulnerability_id": vuln.id,
                "vuln_id": vuln.vuln_id,
                "title": vuln.title,
                "severity": vuln.severity,
                "status": vuln.status,
                "impact_on_asset": link.impact_on_asset,
                "notes": link.notes,
                "created_at": link.created_at,
            })
    
    risk_assessments = []
    for assessment in asset.risk_assessments:
        risk_assessments.append({
            "id": assessment.id,
            "assessment_date": assessment.assessment_date.isoformat(),
            "risk_score": assessment.risk_score,
            "coverage_percentage": assessment.coverage_percentage,
            "gaps": assessment.gaps
        })
    
    total_controls = len(linked_controls) + len(linked_internal_controls) + len(linked_framework_controls)
    coverage = min(total_controls * 10, 100) if total_controls > 0 else 0
    
    return AssetDetailResponse(
        id=asset.id,
        tenant_id=asset.tenant_id,
        name=asset.name,
        description=asset.description,
        asset_type=asset.asset_type,
        owner_id=asset.owner_id,
        owner_name=asset.owner.display_name if asset.owner else None,
        custodian=asset.custodian,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        criticality=asset.criticality,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
        valuation=asset.valuation,
        vendor=asset.vendor,
        location=asset.location,
        status=asset.status,
        is_customer_facing=getattr(asset, "is_customer_facing", None),
        is_internet_facing=getattr(asset, "is_internet_facing", None),
        regulated_data_type=getattr(asset, "regulated_data_type", None),
        operational_dependency=getattr(asset, "operational_dependency", None),
        business_impact_notes=getattr(asset, "business_impact_notes", None),
        created_at=asset.created_at,
        asset_role=getattr(asset, "asset_role", "host") or "host",
        parent_asset_id=getattr(asset, "parent_asset_id", None),
        os_family=getattr(asset, "os_family", None),
        os_version=getattr(asset, "os_version", None),
        os_normalized=getattr(asset, "os_normalized", None),
        linked_controls=linked_controls,
        linked_internal_controls=linked_internal_controls,
        linked_framework_controls=linked_framework_controls,
        linked_risks=linked_risks,
        linked_evidence=linked_evidence,
        linked_vulnerabilities=linked_vulnerabilities,
        risk_assessments=risk_assessments,
        coverage_percentage=float(coverage)
    )


# ─── Removed: 4× /assets/{id}/security-compliance/* endpoints ────────────
# (list controls, get selections, add selections, remove selection)
#
# These were tied to a single static seed file (CIS_WS2012R2_Controls.json)
# and showed the same 363 Windows Server 2012 R2 controls on every asset
# regardless of OS — directly contradicting the strict OS→benchmark matcher
# wired into pick_benchmark_for_os(). The frontend tab that consumed them
# is also deleted (see assets/[id]/page.tsx). Existing selection rows with
# benchmark='CIS_WS2012R2' (currently zero in this DB) would be orphaned;
# the 3,351 rows with benchmark='CIS_PLUGIN' belong to a separate code
# path (the compliance-plugins library, written by run_service.py) and
# are left untouched.
#
# If we ever want a per-asset "show only the rules that apply to this OS"
# view, it should live on /compliance/plugins/library scoped by asset_id,
# NOT in its own asset-tab silo.


@router.post("/{asset_id}/link-framework-control", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_framework_control(
    asset_id: int,
    link: AssetFrameworkControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    control = db.query(FrameworkControl).filter(
        FrameworkControl.id == link.framework_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control not found"
        )
    
    existing = db.query(AssetFrameworkControlLink).filter(
        AssetFrameworkControlLink.asset_id == asset_id,
        AssetFrameworkControlLink.framework_control_id == link.framework_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetFrameworkControlLink(
        asset_id=asset_id,
        framework_control_id=link.framework_control_id,
        coverage_status=link.coverage_status,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Framework control linked successfully", id=db_link.id)


@router.delete("/{asset_id}/link-framework-control/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_framework_control_link(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    link = db.query(AssetFrameworkControlLink).filter(
        AssetFrameworkControlLink.id == link_id,
        AssetFrameworkControlLink.asset_id == asset_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{asset_id}/link-evidence", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_evidence(
    asset_id: int,
    link: AssetEvidenceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    evidence = db.query(Evidence).filter(
        Evidence.id == link.evidence_id,
        Evidence.tenant_id.in_(user_tenants)
    ).first()
    if not evidence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence not found"
        )
    
    existing = db.query(AssetEvidenceLink).filter(
        AssetEvidenceLink.asset_id == asset_id,
        AssetEvidenceLink.evidence_id == link.evidence_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetEvidenceLink(
        asset_id=asset_id,
        evidence_id=link.evidence_id,
        relationship_type=link.relationship_type
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Evidence linked successfully", id=db_link.id)


@router.delete("/{asset_id}/link-evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_evidence_link(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    link = db.query(AssetEvidenceLink).filter(
        AssetEvidenceLink.id == link_id,
        AssetEvidenceLink.asset_id == asset_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.get("/{asset_id}/coverage-analysis", response_model=AssetCoverageAnalysis)
def get_asset_coverage_analysis(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.risk_assessments)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    total_controls = len(asset.control_links) + len(asset.internal_control_links) + len(asset.framework_control_links)
    
    full_framework_coverage = sum(1 for link in asset.framework_control_links if link.coverage_status == "full")
    partial_framework_coverage = sum(1 for link in asset.framework_control_links if link.coverage_status == "partial")
    full_internal_coverage = sum(1 for link in asset.internal_control_links if link.coverage_status == "full")
    partial_internal_coverage = sum(1 for link in asset.internal_control_links if link.coverage_status == "partial")
    covered_controls = (
        len(asset.control_links)
        + full_internal_coverage
        + (partial_internal_coverage * 0.5)
        + full_framework_coverage
        + (partial_framework_coverage * 0.5)
    )
    
    expected_controls = 10
    coverage_percentage = min((covered_controls / expected_controls) * 100, 100) if expected_controls > 0 else 0
    
    gaps = []
    if total_controls < 3:
        gaps.append({"type": "insufficient_controls", "message": "Asset has fewer than 3 controls linked"})
    if asset.criticality in ["high", "critical"] and total_controls < 5:
        gaps.append({"type": "critical_asset_gap", "message": "Critical/high priority asset should have at least 5 controls"})
    if not asset.framework_control_links:
        gaps.append({"type": "no_framework_controls", "message": "No framework controls linked to this asset"})
    
    latest_assessment = None
    risk_score = None
    if asset.risk_assessments:
        latest_assessment = sorted(asset.risk_assessments, key=lambda x: x.assessment_date, reverse=True)[0]
        risk_score = latest_assessment.risk_score
    
    return AssetCoverageAnalysis(
        asset_id=asset.id,
        asset_name=asset.name,
        total_controls=total_controls,
        covered_controls=int(covered_controls),
        coverage_percentage=round(coverage_percentage, 2),
        gaps=gaps,
        risk_score=risk_score
    )


@router.post("/{asset_id}/assess-risk", response_model=AssetRiskAssessmentResponse)
def perform_risk_assessment(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.evidence_links)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    num_normalized_controls = len(asset.control_links)
    num_internal_controls = len(asset.internal_control_links)
    num_framework_controls = len(asset.framework_control_links)
    num_evidence = len(asset.evidence_links)
    
    total_controls = num_normalized_controls + num_internal_controls + num_framework_controls
    coverage = min(total_controls * 10, 100)
    
    base_risk = 5
    if asset.criticality == "critical":
        base_risk = 9
    elif asset.criticality == "high":
        base_risk = 7
    elif asset.criticality == "low":
        base_risk = 3
    
    control_reduction = total_controls * 0.4
    evidence_reduction = num_evidence * 0.2
    risk_score = max(1, base_risk - control_reduction - evidence_reduction)
    
    gaps = {
        "missing_controls": max(0, 10 - total_controls),
        "missing_evidence": max(0, 5 - num_evidence),
        "recommendations": []
    }
    if total_controls < 3:
        gaps["recommendations"].append("Add more controls to improve coverage")
    if asset.criticality in ["high", "critical"] and total_controls < 5:
        gaps["recommendations"].append("Critical assets should have at least 5 controls")
    if num_evidence < 2:
        gaps["recommendations"].append("Add more evidence documentation")
    if not asset.framework_control_links:
        gaps["recommendations"].append("Link to framework controls for better compliance tracking")
    
    assessment = AssetRiskAssessment(
        asset_id=asset_id,
        risk_score=round(risk_score, 2),
        coverage_percentage=float(coverage),
        gaps=gaps,
        assessor_id=current_user.id
    )
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    
    return assessment


class AssetInternalControlLinkCreate(BaseModel):
    internal_control_id: int
    coverage_status: Optional[str] = "partial"


@router.post("/{asset_id}/internal-controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_internal_control(
    asset_id: int,
    link: AssetInternalControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Link an ERM internal control to an asset"""
    user_tenants = get_user_tenants(current_user, db)

    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    ctrl = db.query(InternalControl).filter(
        InternalControl.id == link.internal_control_id,
        InternalControl.tenant_id.in_(user_tenants)
    ).first()
    if not ctrl:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Internal control not found")

    existing = db.query(AssetInternalControlLink).filter(
        AssetInternalControlLink.asset_id == asset_id,
        AssetInternalControlLink.internal_control_id == link.internal_control_id
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Link already exists")

    db_link = AssetInternalControlLink(
        asset_id=asset_id,
        internal_control_id=link.internal_control_id,
        coverage_status=link.coverage_status
    )
    db.add(db_link)
    db.commit()
    db.refresh(db_link)

    return MessageResponse(message="Internal control linked successfully", id=db_link.id)


@router.delete("/{asset_id}/internal-controls/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_asset_from_internal_control(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Unlink an ERM internal control from an asset"""
    user_tenants = get_user_tenants(current_user, db)

    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    db_link = db.query(AssetInternalControlLink).filter(
        AssetInternalControlLink.id == link_id,
        AssetInternalControlLink.asset_id == asset_id
    ).first()
    if not db_link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Link not found")

    db.delete(db_link)
    db.commit()
    return None
