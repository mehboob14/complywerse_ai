"""Build CompliverseAI pricing as a SINGLE-sheet workbook.

Five sections stacked vertically on one tab:
  1. Base Module Pricing
  2. Module-Based User Pricing
  3. AI Pricing (Claude Sonnet 4.5 + GPT-5)
  4. Implementation Fees
  5. Support Fees (Standard / Bronze / Premium)
Plus a quote calculator at the bottom that pulls from all five sections.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "pricing" / "CompliverseAI_Pricing_OneSheet.xlsx"
OUT.parent.mkdir(exist_ok=True)

NAVY = "1F4E78"
BAND = PatternFill("solid", fgColor=NAVY)
SUB = PatternFill("solid", fgColor="D9E1F2")
INP = PatternFill("solid", fgColor="FFF2CC")
DER = PatternFill("solid", fgColor="E2EFDA")
TOT = PatternFill("solid", fgColor="FCE4D6")
WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
NAVY_BOLD = Font(name="Calibri", size=12, bold=True, color=NAVY)
BOLD = Font(name="Calibri", size=11, bold=True)
TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def hdr(ws, row, cols, labels):
    for i, label in enumerate(labels):
        c = ws.cell(row, cols + i, label)
        c.fill, c.font, c.alignment, c.border = BAND, WHITE_BOLD, CENTER, BOX


def section_band(ws, row, n_cols, title):
    c = ws.cell(row, 1, title)
    c.fill, c.font, c.alignment = BAND, WHITE_BOLD, LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 22


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing"
    ws.sheet_properties.tabColor = NAVY
    N = 7  # working width

    # ----- Title --------------------------------------------------------------
    ws.cell(1, 1, "CompliverseAI GRC — Pricing Workbook (single sheet)").font = TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
    ws.cell(2, 1,
            "Yellow = editable input · Green = derived · Orange = totals. All currency USD. Edit any yellow cell — every total recalculates.").font = Font(italic=True, color="595959")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
    ws.row_dimensions[1].height = 24

    r = 4

    # ===== 1. Base Module Pricing ============================================
    section_band(ws, r, N, "1.  BASE MODULE PRICING  (per-tenant platform fee, monthly)")
    r += 1
    hdr(ws, r, 1, ["Module", "Description", "Included in tier", "Base $/tenant/mo", "Annual $/tenant"])
    r += 1
    base_modules = [
        ("Compliance Core",         "Frameworks, controls, evidence library",         "All tiers",          500),
        ("Risk Core (ERM)",         "Risk register, KRIs, RCSA",                      "All tiers",          500),
        ("Reporting & Dashboards",  "Standard reports, KPI dashboards",               "All tiers",          250),
        ("Vendor Risk Management",  "Vendor inventory, assessments, questionnaires",  "Growth +",           400),
        ("Governance & Policy",     "Policy lifecycle, attestations, doc mgmt",       "Growth +",           400),
        ("Vulnerability Management","Scanner integrations, remediation tracking",     "Growth +",           600),
        ("Internal Audit",          "Audit planning, fieldwork, findings",            "Enterprise +",       800),
        ("Asset Management (CMDB)", "Asset inventory, criticality, ownership",        "Enterprise +",       400),
        ("Workflow Engine",         "Visual workflow builder, automation",            "Enterprise +",       500),
        ("Integrations Hub",        "Pre-built connectors (50+) + iPaaS",             "Enterprise +",       600),
        ("AI Copilot",              "Embedded AI assistant (separate AI metering)",   "Enterprise + (or add-on)", 700),
    ]
    base_start = r
    for mod, desc, tiers, price in base_modules:
        ws.cell(r, 1, mod).font = BOLD; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, desc).border = BOX
        ws.cell(r, 3, tiers).border = BOX; ws.cell(r, 3).alignment = CENTER
        p = ws.cell(r, 4, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'
        a = ws.cell(r, 5, f"=D{r}*12"); a.fill = DER; a.border = BOX; a.number_format = '"$"#,##0'
        r += 1
    base_end = r - 1
    # totals
    ws.cell(r, 1, "TOTAL — all modules enabled").font = BOLD
    for col in (1, 2, 3): ws.cell(r, col).fill = TOT; ws.cell(r, col).border = BOX
    t1 = ws.cell(r, 4, f"=SUM(D{base_start}:D{base_end})"); t1.fill = TOT; t1.font = BOLD; t1.border = BOX; t1.number_format = '"$"#,##0'
    t2 = ws.cell(r, 5, f"=SUM(E{base_start}:E{base_end})"); t2.fill = TOT; t2.font = BOLD; t2.border = BOX; t2.number_format = '"$"#,##0'
    base_total_row = r
    r += 2

    # ===== 2. Module-Based User Pricing ======================================
    section_band(ws, r, N, "2.  MODULE-BASED USER PRICING  ($ per active user per month)")
    r += 1
    hdr(ws, r, 1, ["Module", "Starter $/user/mo", "Growth $/user/mo", "Enterprise $/user/mo", "Enterprise+ $/user/mo", "Notes"])
    r += 1
    user_pricing = [
        ("Compliance Core",         8,  6,  5,  4,  "Auditors, control owners"),
        ("Risk Core (ERM)",         8,  6,  5,  4,  "Risk owners, 2nd-line"),
        ("Reporting & Dashboards",  3,  2,  2,  1,  "View-only seats cheaper"),
        ("Vendor Risk Management",  None, 8,  6,  5,  "Not in Starter"),
        ("Governance & Policy",     None, 5,  4,  3,  "Not in Starter"),
        ("Vulnerability Management",None, 10, 8,  6,  "Not in Starter"),
        ("Internal Audit",          None, None, 10, 8,  "Enterprise+"),
        ("Asset Management (CMDB)", None, None, 4,  3,  "Enterprise+"),
        ("Workflow Engine",         None, None, 6,  5,  "Enterprise+"),
        ("Integrations Hub",        None, None, 4,  3,  "Enterprise+"),
        ("AI Copilot (per-user opt-in)", 5, 4, 0, 0,  "Free in Ent / Ent+ (unlimited)"),
    ]
    user_start = r
    for row in user_pricing:
        mod, s, g, e, ep, note = row
        ws.cell(r, 1, mod).font = BOLD; ws.cell(r, 1).border = BOX
        for col, v in enumerate([s, g, e, ep], start=2):
            c = ws.cell(r, col, v if v is not None else "—")
            c.fill = INP if v is not None else PatternFill("solid", fgColor="EEEEEE")
            c.border = BOX; c.alignment = CENTER
            if v is not None: c.number_format = '"$"#,##0.00'
        ws.cell(r, 6, note).border = BOX
        r += 1
    user_end = r - 1
    # Bundle totals (all-modules user)
    ws.cell(r, 1, "BUNDLE — all eligible modules per user").font = BOLD
    for col in range(1, 7): ws.cell(r, col).fill = TOT; ws.cell(r, col).border = BOX
    for col, letter in enumerate("BCDE", start=2):
        c = ws.cell(r, col, f'=SUMIF({letter}{user_start}:{letter}{user_end},"<>—")')
        c.fill, c.font, c.border, c.number_format = TOT, BOLD, BOX, '"$"#,##0.00'
    bundle_row = r
    r += 2

    # ===== 3. AI Pricing =====================================================
    section_band(ws, r, N, "3.  AI PRICING  (Claude Sonnet 4.5  +  GPT-5)")
    r += 1
    hdr(ws, r, 1, ["Item", "Unit", "Cost", "Customer price", "Markup", "Notes"])
    r += 1
    ai_rows = [
        ("Claude Sonnet 4.5 — input",   "$/M tokens", 3.00,  None, None, "Anthropic list"),
        ("Claude Sonnet 4.5 — output",  "$/M tokens", 15.00, None, None, "Anthropic list"),
        ("GPT-5 — input",               "$/M tokens", 5.00,  None, None, "OpenAI list"),
        ("GPT-5 — output",              "$/M tokens", 20.00, None, None, "OpenAI list"),
    ]
    ai_start = r
    for label, unit, cost, _p, _m, note in ai_rows:
        ws.cell(r, 1, label).border = BOX; ws.cell(r, 1).font = BOLD
        ws.cell(r, 2, unit).border = BOX; ws.cell(r, 2).alignment = CENTER
        c = ws.cell(r, 3, cost); c.fill = INP; c.border = BOX; c.number_format = '"$"#,##0.00'
        ws.cell(r, 4, "—").border = BOX; ws.cell(r, 4).alignment = CENTER
        ws.cell(r, 5, "—").border = BOX; ws.cell(r, 5).alignment = CENTER
        ws.cell(r, 6, note).border = BOX
        r += 1
    r += 1
    # Bucketed customer SKUs
    hdr(ws, r, 1, ["AI Bucket SKU", "Vendor", "Tokens included (M)", "Cost to deliver", "Customer price/mo", "Markup ×", "Notes"])
    r += 1
    # Blended cost per M tokens at 5:1 input:output ratio:
    # blended = input*(5/6) + output*(1/6)
    blended_claude = "(3*(5/6)+15*(1/6))"   # ~5.00
    blended_gpt5   = "(5*(5/6)+20*(1/6))"   # ~7.50
    bucket_rows = [
        ("AI Starter — Claude",     "Claude Sonnet 4.5", 5,   blended_claude, 50,    "Light usage; opt-in for Starter tier"),
        ("AI Starter — GPT-5",      "GPT-5",             3,   blended_gpt5,   50,    "Light usage; opt-in for Starter tier"),
        ("AI Growth — Claude",      "Claude Sonnet 4.5", 30,  blended_claude, 300,   "Included in Growth tier"),
        ("AI Growth — GPT-5",       "GPT-5",             20,  blended_gpt5,   300,   "Included in Growth tier"),
        ("AI Enterprise (Unlimited)","Claude + GPT-5",   None, None,          0,     "Bundled in Enterprise SaaS (fair-use)"),
        ("AI Enterprise+ (Dedicated)","Claude + GPT-5 dedicated", None, None, 0,     "Bundled in Enterprise+ SaaS; BYO key option"),
    ]
    bucket_start = r
    for sku, vendor, tokens, blended, price, note in bucket_rows:
        ws.cell(r, 1, sku).font = BOLD; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, vendor).border = BOX; ws.cell(r, 2).alignment = CENTER
        if tokens is not None:
            t = ws.cell(r, 3, tokens); t.fill = INP; t.border = BOX; t.number_format = "0"; t.alignment = CENTER
            cost_cell = ws.cell(r, 4, f"=C{r}*{blended}"); cost_cell.fill = DER; cost_cell.border = BOX; cost_cell.number_format = '"$"#,##0.00'
            p = ws.cell(r, 5, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'
            m = ws.cell(r, 6, f"=IFERROR(E{r}/D{r},0)"); m.fill = DER; m.border = BOX; m.number_format = "0.00\"×\""
        else:
            for col in (3, 4):
                ws.cell(r, col, "Bundled").border = BOX; ws.cell(r, col).alignment = CENTER
            p = ws.cell(r, 5, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'; p.alignment = CENTER
            ws.cell(r, 6, "n/a").border = BOX; ws.cell(r, 6).alignment = CENTER
        ws.cell(r, 7, note).border = BOX
        r += 1
    r += 2

    # ===== 4. Implementation Fees ============================================
    section_band(ws, r, N, "4.  IMPLEMENTATION FEES  (one-time, professional services)")
    r += 1
    hdr(ws, r, 1, ["Tier", "Cost to deliver", "Customer price", "Margin %", "Scope", "Timeline"])
    r += 1
    impl_rows = [
        ("Starter",        2000,  5000,   "Kickoff, 1 framework, 5 hrs training, basic config",                "2-4 weeks"),
        ("Growth",         6000,  15000,  "3 frameworks, 2 integrations, RCSA workshop, 20 hrs training",      "6-8 weeks"),
        ("Enterprise",     16000, 40000,  "Unlimited frameworks, custom workflows, 5 integrations, 40 hrs training", "10-14 weeks"),
        ("Enterprise Plus",40000, 100000, "Scoped SOW, dedicated PM, on-prem option, custom integrations",     "16-24 weeks"),
    ]
    for tier_name, cost, price, scope, timeline in impl_rows:
        ws.cell(r, 1, tier_name).font = BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).alignment = CENTER
        c = ws.cell(r, 2, cost); c.fill = INP; c.border = BOX; c.number_format = '"$"#,##0'
        p = ws.cell(r, 3, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'
        m = ws.cell(r, 4, f"=(C{r}-B{r})/C{r}"); m.fill = DER; m.border = BOX; m.number_format = "0.0%"; m.alignment = CENTER
        ws.cell(r, 5, scope).border = BOX
        ws.cell(r, 6, timeline).border = BOX; ws.cell(r, 6).alignment = CENTER
        r += 1
    r += 1
    # Optional add-ons
    hdr(ws, r, 1, ["Optional Implementation Add-ons", "Unit", "Cost", "Price", "Margin %", "Notes"])
    r += 1
    addons = [
        ("Custom integration build",       "per integration", 4000, 12000, "Beyond included integrations"),
        ("Data migration from legacy GRC", "fixed",           8000, 25000, "Up to 10 GB"),
        ("Custom framework load",          "per framework",   2000, 6000,  "Non-standard / regional"),
        ("Workshops / training (extra)",   "per day",         800,  2500,  "On-site +T&E"),
    ]
    for label, unit, cost, price, note in addons:
        ws.cell(r, 1, label).border = BOX
        ws.cell(r, 2, unit).border = BOX; ws.cell(r, 2).alignment = CENTER
        c = ws.cell(r, 3, cost); c.fill = INP; c.border = BOX; c.number_format = '"$"#,##0'
        p = ws.cell(r, 4, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'
        m = ws.cell(r, 5, f"=(D{r}-C{r})/D{r}"); m.fill = DER; m.border = BOX; m.number_format = "0.0%"; m.alignment = CENTER
        ws.cell(r, 6, note).border = BOX
        r += 1
    r += 2

    # ===== 5. Support Fees ===================================================
    section_band(ws, r, N, "5.  SUPPORT FEES  (Standard / Bronze / Premium)")
    r += 1
    hdr(ws, r, 1, ["Plan", "Response SLA", "Channels", "Cost/mo", "Customer $/mo", "Annual price"])
    r += 1
    support_rows = [
        ("Standard", "48h business",  "Email, in-app",                               50,   150,   "Default for Starter tier"),
        ("Bronze",   "8h business",   "Email, in-app, chat",                         150,  450,   "Recommended for Growth"),
        ("Premium",  "1h 24×5 + 15-min P1",
                                      "Email, chat, phone, dedicated CSM, quarterly business reviews",
                                                                                     500, 1500,   "Required for Enterprise / Ent+"),
    ]
    for plan, sla, ch, cost, price, note in support_rows:
        ws.cell(r, 1, plan).font = BOLD; ws.cell(r, 1).border = BOX; ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 2, sla).border = BOX; ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 3, ch).border = BOX
        c = ws.cell(r, 4, cost); c.fill = INP; c.border = BOX; c.number_format = '"$"#,##0'
        p = ws.cell(r, 5, price); p.fill = INP; p.border = BOX; p.number_format = '"$"#,##0'
        a = ws.cell(r, 6, f"=E{r}*12"); a.fill = DER; a.border = BOX; a.number_format = '"$"#,##0'
        r += 1
    sup_first_row = r - 3   # Standard
    sup_bronze_row = r - 2
    sup_premium_row = r - 1
    r += 2

    # ===== Quote Calculator ==================================================
    section_band(ws, r, N, "QUOTE CALCULATOR  (edit yellow inputs — totals recalculate)")
    r += 1
    inputs_start = r
    inputs = [
        ("Tier (Starter / Growth / Enterprise / Enterprise Plus)", "Growth"),
        ("Number of users",                                          50),
        ("Modules: # of base modules enabled (1-11)",                7),
        ("Avg modules per user (counted from Module User Pricing)",  5),
        ("AI plan (None / Starter Bucket / Growth Bucket / Unlimited / Dedicated)", "Growth Bucket"),
        ("Implementation tier",                                      "Growth"),
        ("Support plan (Standard / Bronze / Premium)",               "Bronze"),
        ("Annual prepay discount %",                                 0.10),
    ]
    for label, val in inputs:
        ws.cell(r, 1, label).border = BOX
        c = ws.cell(r, 2, val); c.fill = INP; c.border = BOX
        if isinstance(val, float): c.number_format = "0.0%"
        elif isinstance(val, int): c.number_format = "0"
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    tier_in = inputs_start
    users_in = inputs_start + 1
    nbase_in = inputs_start + 2
    nuser_modules_in = inputs_start + 3
    ai_in = inputs_start + 4
    impl_in = inputs_start + 5
    sup_in = inputs_start + 6
    prepay_in = inputs_start + 7
    r += 1

    # Outputs
    hdr(ws, r, 1, ["Line item", "Amount", "Formula reference", "", "", ""])
    r += 1

    def out(label, formula, fmt='"$"#,##0', bold_total=False):
        nonlocal r
        ws.cell(r, 1, label).border = BOX; ws.cell(r, 1).font = BOLD if bold_total else Font()
        c = ws.cell(r, 2, formula); c.border = BOX; c.number_format = fmt
        if bold_total: c.fill, c.font = TOT, Font(bold=True, color=NAVY, size=12)
        else: c.fill = DER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 5, "").border = BOX
        r += 1

    # Base subscription = avg base monthly per enabled module × #modules × 12
    out("Base modules subscription (annual)",
        f"=B{base_total_row}/COUNTA(A{base_start}:A{base_end})*B{nbase_in}*12*(1-B{prepay_in})")
    base_sub_row = r - 1

    # User subscription = (tier bundle $/user/mo, scaled by avg modules ratio) × users × 12
    out("Per-user subscription (annual)",
        f'=B{users_in}*12*(1-B{prepay_in})*'
        f'IF(B{tier_in}="Starter",B{bundle_row},'
        f'IF(B{tier_in}="Growth",C{bundle_row},'
        f'IF(B{tier_in}="Enterprise",D{bundle_row},E{bundle_row})))*'
        f'(B{nuser_modules_in}/COUNTA(A{user_start}:A{user_end}))')
    user_sub_row = r - 1

    # AI annual
    out("AI plan (annual)",
        f'=12*IF(B{ai_in}="None",0,'
        f'IF(B{ai_in}="Starter Bucket",E{bucket_start}+E{bucket_start+1},'
        f'IF(B{ai_in}="Growth Bucket",E{bucket_start+2}+E{bucket_start+3},'
        f'IF(B{ai_in}="Unlimited",E{bucket_start+4},E{bucket_start+5}))))')
    ai_sub_row = r - 1

    # Implementation (one-time) — locate impl rows: starter at "impl" section we built. Use IF
    # impl rows are 4 rows starting after "Implementation Fees" header row + hdr row.
    out("Implementation fee (one-time)",
        f'=IF(B{impl_in}="Starter",5000,IF(B{impl_in}="Growth",15000,IF(B{impl_in}="Enterprise",40000,100000)))')
    impl_out_row = r - 1

    # Support
    out("Support (annual)",
        f'=12*IF(B{sup_in}="Standard",E{sup_first_row},IF(B{sup_in}="Bronze",E{sup_bronze_row},E{sup_premium_row}))')
    sup_out_row = r - 1

    r += 1
    out("YEAR-1 TOTAL",
        f"=B{base_sub_row}+B{user_sub_row}+B{ai_sub_row}+B{impl_out_row}+B{sup_out_row}",
        bold_total=True)
    out("YEAR-2 RECURRING (no implementation)",
        f"=B{base_sub_row}+B{user_sub_row}+B{ai_sub_row}+B{sup_out_row}",
        bold_total=True)

    # Column widths
    widths = [44, 24, 22, 22, 22, 36, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    wb.save(OUT)
    print(f"[ok] {OUT}")


if __name__ == "__main__":
    main()
