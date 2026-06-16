"""Build the Module + Base + User pricing workbook AND a UBL-pre-filled
4-year TCO quote workbook.

Single-tier (Enterprise) pricing. No SMB / Mid-market scenarios.
Inputs:
  - Platform base fee  ($/month)
  - Per-user fee       ($/user/month)
  - Number of users    (100 baseline)
  - Support fees       (Standard, Premium only)
  - Implementation fee (one-time)
  - API costs broken out separately:
      Azure (hosting), GPT-5 (tokens), Claude (tokens),
      plus SMS / Cloud read / SSO / Storage line items.

Outputs:
  pricing/CompliverseAI_Pricing_Module_Plus_User.xlsx   (generic)
  pricing/CompliverseAI_UBL_Quote_v1.xlsx               (UBL-prefilled)
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

PRICING = Path(__file__).resolve().parent.parent / "pricing"
PRICING.mkdir(exist_ok=True)

GENERIC_PATH = PRICING / "CompliverseAI_Pricing_Module_Plus_User.xlsx"
UBL_PATH = PRICING / "CompliverseAI_UBL_Quote_v1.xlsx"

# ---------- styling ----------
BORDER = Side(border_style="thin", color="999999")
BOX = Border(left=BORDER, right=BORDER, top=BORDER, bottom=BORDER)
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="F4B084")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")


def hdr(c):
    c.font = WHITE_BOLD; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX


def sub(c):
    c.font = BOLD; c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = BOX


def inp(c): c.fill = INPUT_FILL; c.border = BOX
def der(c): c.fill = DERIVED_FILL; c.border = BOX
def tot(c): c.fill = TOTAL_FILL; c.border = BOX; c.font = BOLD


def autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_name(wb: Workbook, name: str, ref: str):
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


# ---------- assumption rows ----------
@dataclass
class Assumption:
    label: str
    value: Any = None
    name: str | None = None
    fmt: str = "General"
    kind: str = "input"        # input | derived | section
    note: str = ""


def assumptions_rows(*, ubl: bool) -> list[Assumption]:
    """All assumption rows. UBL toggles change a few defaults."""
    a = Assumption
    rows: list[Assumption] = [
        a("GLOBAL", kind="section"),
        a("Currency", "USD", fmt="@"),
        a("Annual prepay discount %", 0.15, "annual_discount", "0.0%"),
        a("Target gross margin %", 0.60, "target_margin", "0.0%"),
        a("Annual price uplift %", 0.05, "annual_uplift_pct", "0.0%"),
        a("TCO horizon (years)  — table is sized at build time; rebuild to extend",
          4, "tco_horizon_years", "0"),

        a("HOSTING MODEL  (toggle)", kind="section"),
        a('Mode  ("SaaS" or "SelfHost")', "SaaS", "hosting_mode", "@"),
        a("Azure SaaS hosting cost  $/tenant/month",
          500, "azure_hosting_per_tenant", '"$"#,##0',
          note="Pooled Azure cost — App Service, AKS, Postgres, blob, monitoring"),
        a("Self-host managed-services fee  $/tenant/month  (replaces Azure if SelfHost)",
          3500, "selfhost_managed_fee", '"$"#,##0'),
        a("Active hosting cost  $/tenant/month  (derived)",
          '=IF(EXACT(hosting_mode,"SelfHost"),selfhost_managed_fee,'
          'azure_hosting_per_tenant)',
          "hosting_active", '"$"#,##0', kind="derived"),

        a("USERS & PLATFORM PRICING", kind="section"),
        a("Number of users on the platform", 100, "users_count", "0"),
        a("Platform base fee  $/tenant/month",
          5500 if ubl else 4000, "platform_base_fee", '"$"#,##0'),
        a("Per-user fee  $/user/month",
          55 if ubl else 45, "per_user_fee", '"$"#,##0.00'),
        a("Minimum platform fee floor  $/tenant/month",
          6500 if ubl else 4500, "min_platform_fee_usd", '"$"#,##0',
          note="Hard floor — customer never pays less than this"),
        a("Per-user floor  $/user/month",
          45 if ubl else 35, "per_user_floor", '"$"#,##0.00',
          note="Floor coefficient applied alongside the platform fee floor"),

        a("LLM API COSTS  (heavy-usage assumption)", kind="section"),
        a("GPT-5 — input tokens per user per month",
          300_000, "gpt_input_tokens", "#,##0"),
        a("GPT-5 — output tokens per user per month",
          90_000, "gpt_output_tokens", "#,##0"),
        a("GPT-5 — input price  $/1M tokens",
          3.00, "gpt_input_price_per_1m", '"$"#,##0.00'),
        a("GPT-5 — output price $/1M tokens",
          15.00, "gpt_output_price_per_1m", '"$"#,##0.00'),
        a("GPT-5 cost per user per month  (derived)",
          "=(gpt_input_tokens*gpt_input_price_per_1m + "
          "gpt_output_tokens*gpt_output_price_per_1m)/1000000",
          "api_gpt5", '"$"#,##0.00', kind="derived"),

        a("Claude — input tokens per user per month",
          200_000, "claude_input_tokens", "#,##0"),
        a("Claude — output tokens per user per month",
          60_000, "claude_output_tokens", "#,##0"),
        a("Claude — input price  $/1M tokens",
          3.00, "claude_input_price_per_1m", '"$"#,##0.00'),
        a("Claude — output price $/1M tokens",
          15.00, "claude_output_price_per_1m", '"$"#,##0.00'),
        a("Claude cost per user per month  (derived)",
          "=(claude_input_tokens*claude_input_price_per_1m + "
          "claude_output_tokens*claude_output_price_per_1m)/1000000",
          "api_claude", '"$"#,##0.00', kind="derived"),

        a("Total LLM cost per user per month  (GPT-5 + Claude)",
          "=api_gpt5+api_claude", "api_llm_total", '"$"#,##0.00', kind="derived"),

        a("OTHER API COSTS  ($/user/month)", kind="section"),
        a("SMS / WhatsApp notifications", 0.10, "api_sms", '"$"#,##0.00'),
        a("Cloud read APIs (AWS/Azure/GCP)", 1.00, "api_cloud", '"$"#,##0.00'),
        a("SSO (Okta / Azure AD)", 0.25, "api_sso", '"$"#,##0.00'),
        a("Storage / CDN", 0.40, "api_storage", '"$"#,##0.00'),

        a("SUPPORT FEES  (per tenant per month)", kind="section"),
        a("Standard support — price",
          1200, "sup_std_price", '"$"#,##0',
          note="Business hours email + ticketing, 8h response"),
        a("Standard support — cost", 400, "sup_std_cost", '"$"#,##0'),
        a("Premium support — price",
          3500 if ubl else 2800, "sup_prem_price", '"$"#,##0',
          note="24x7 phone + chat + named CSM, 1h response P1"),
        a("Premium support — cost", 900, "sup_prem_cost", '"$"#,##0'),

        a("IMPLEMENTATION  (one-time)", kind="section"),
        a("Implementation fee  $/tenant",
          75_000 if ubl else 50_000, "impl_price", '"$"#,##0',
          note="Onboarding, data migration, training, 11-module config"),
        a("Implementation cost  $/tenant", 25_000, "impl_cost", '"$"#,##0'),
    ]
    return rows


def build_assumptions(wb: Workbook, *, ubl: bool) -> None:
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Pricing Assumptions  (single Enterprise tier)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:D1")

    headers = ["Assumption", "Value", "Named range", "Notes"]
    for col, h in enumerate(headers, start=1):
        ws.cell(2, col, h); hdr(ws.cell(2, col))

    r = 3
    for a in assumptions_rows(ubl=ubl):
        if a.kind == "section":
            ws.cell(r, 1, a.label); sub(ws.cell(r, 1)); ws.merge_cells(
                start_row=r, end_row=r, start_column=1, end_column=4)
            r += 1
            continue

        ws.cell(r, 1, a.label).border = BOX
        c = ws.cell(r, 2, a.value)
        c.number_format = a.fmt
        if a.kind == "derived":
            der(c)
        else:
            inp(c)
        ws.cell(r, 3, a.name or "").border = BOX
        ws.cell(r, 3).font = Font(italic=True, color="666666")
        ws.cell(r, 4, a.note).border = BOX
        ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="center")

        if a.name:
            add_name(wb, a.name, f"Assumptions!$B${r}")
        r += 1

    autosize(ws, [60, 18, 28, 70])
    ws.freeze_panes = "A3"


def build_cost_buildup(wb: Workbook) -> None:
    """Per-user variable cost (excl. hosting). Hosting is tenant-level, added
    once in pricing/TCO formulas, not double-counted here."""
    ws = wb.create_sheet("Cost Build-Up")
    ws["A1"] = "Per-user variable cost  (API stack only — hosting added separately at tenant level)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:C1")

    for col, h in enumerate(["Component", "Value", "Notes"], start=1):
        ws.cell(3, col, h); hdr(ws.cell(3, col))

    rows = [
        ("GPT-5 cost per user/month", "=api_gpt5",
         "Tokens × $/1M — see Assumptions"),
        ("Claude cost per user/month", "=api_claude",
         "Tokens × $/1M — see Assumptions"),
        ("LLM total per user/month  (GPT-5 + Claude)", "=api_llm_total",
         "Sum of above"),
        ("SMS / WhatsApp per user/month",  "=api_sms", ""),
        ("Cloud read APIs per user/month", "=api_cloud", ""),
        ("SSO per user/month",             "=api_sso", ""),
        ("Storage / CDN per user/month",   "=api_storage", ""),
        ("Per-user variable cost  (excl. hosting)",
         "=api_llm_total + api_sms + api_cloud + api_sso + api_storage",
         "API stack only — multiplied by user count in pricing/TCO formulas"),
    ]
    per_user_cost_row: int | None = None
    for i, (lbl, formula, note) in enumerate(rows, start=4):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, formula); c.number_format = '"$"#,##0.00'
        if "(excl. hosting)" in lbl:
            tot(c); per_user_cost_row = i
        else:
            der(c)
        ws.cell(i, 3, note).border = BOX

    assert per_user_cost_row is not None
    add_name(wb, "per_user_cost", f"'Cost Build-Up'!$B${per_user_cost_row}")

    # Hosting line item shown explicitly as a tenant-level fixed cost
    r = per_user_cost_row + 2
    ws.cell(r, 1, "Hosting  $/tenant/month  (tenant-level, not per user)").border = BOX
    h = ws.cell(r, 2, "=hosting_active"); h.number_format = '"$"#,##0'; tot(h)
    ws.cell(r, 3, "From Assumptions hosting toggle").border = BOX

    autosize(ws, [60, 16, 60])
    ws.freeze_panes = "A4"


def build_platform_pricing(wb: Workbook) -> None:
    """Single-tier Enterprise platform pricing. No multi-tier table."""
    ws = wb.create_sheet("Platform Pricing")
    ws["A1"] = "Platform Pricing  (single Enterprise tier — no SMB / Mid-market)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:C1")

    rows = [
        ("Number of users", "=users_count", "0"),
        ("Platform base fee  $/month", "=platform_base_fee", '"$"#,##0'),
        ("Per-user fee  $/user/month", "=per_user_fee", '"$"#,##0.00'),
        ("Linear monthly  (base + per_user × users)",
         "=platform_base_fee + per_user_fee*users_count", '"$"#,##0'),
        ("Floor monthly  (min_platform_fee + per_user_floor × users)",
         "=min_platform_fee_usd + per_user_floor*users_count", '"$"#,##0'),
        ("CUSTOMER MONTHLY PRICE  (MAX of floor / linear)",
         "=MAX(B7,B8)", '"$"#,##0'),
        ("Floor active?  (TRUE = floor is protecting margin)",
         "=B8>B7", "@"),
        ("", "", ""),
        ("Year-1 annual subscription  (with annual prepay discount)",
         "=B9*12*(1-annual_discount)", '"$"#,##0'),
        ("Year-1 cost  (per_user_cost × users × 12 + hosting × 12)",
         "=per_user_cost*users_count*12 + hosting_active*12", '"$"#,##0'),
        ("Year-1 gross margin %", "=IFERROR((B12-B13)/B12,0)", "0.0%"),
    ]

    ws.cell(3, 1, "Item"); hdr(ws.cell(3, 1))
    ws.cell(3, 2, "Value"); hdr(ws.cell(3, 2))

    for i, (lbl, formula, fmt) in enumerate(rows, start=4):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, formula); c.number_format = fmt
        if i == 9:                     # customer monthly price
            tot(c); c.font = Font(bold=True, size=12, color="C00000")
        elif i in (12, 13, 14):        # Y1 summary block (sub / cost / GM%)
            tot(c)
        else:
            der(c)

    # Conditional format Y1 GM% (B14) red below target
    ws.conditional_formatting.add(
        "B14:B14",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    autosize(ws, [56, 22, 4])
    ws.freeze_panes = "A4"


def build_tco(wb: Workbook) -> None:
    """4-Year TCO — years as rows, metrics as columns."""
    ws = wb.create_sheet("4-Year TCO")
    ws["A1"] = "4-Year Total Cost of Ownership  (single Enterprise scenario)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    ws.merge_cells("A1:H1")

    # Inputs panel
    ws["A3"] = "Quote inputs"; sub(ws["A3"]); ws.merge_cells("A3:H3")
    inputs = [
        ("Number of users", "=users_count", "0"),
        ("Support tier  (Standard / Premium)", "Premium", "@"),
        ("Hosting mode (mirrors Assumptions)", "=hosting_mode", "@"),
        ("Annual uplift %", "=annual_uplift_pct", "0.0%"),
        ("TCO horizon (years)", "=tco_horizon_years", "0"),
    ]
    for i, (lbl, val, fmt) in enumerate(inputs, start=4):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, val); c.number_format = fmt
        if isinstance(val, str) and val.startswith("="):
            der(c)
        else:
            inp(c)

    USERS = "B4"
    SUP = "B5"

    # Derived monthly subscription
    ws["A10"] = "Derived monthly subscription"; sub(ws["A10"]); ws.merge_cells("A10:H10")
    derived = [
        ("Platform base fee", "=platform_base_fee", '"$"#,##0'),
        ("Per-user fee",      "=per_user_fee",      '"$"#,##0.00'),
        ("Linear monthly  (base + per_user × users)",
         f"=platform_base_fee + per_user_fee*{USERS}", '"$"#,##0'),
        ("Floor monthly  (min_platform_fee + per_user_floor × users)",
         f"=min_platform_fee_usd + per_user_floor*{USERS}", '"$"#,##0'),
        ("Customer monthly  (MAX of floor / linear)",
         "=MAX(B13,B14)", '"$"#,##0'),
        ("Floor active?  (TRUE = floor protects margin)", "=B14>B13", "@"),
    ]
    for i, (lbl, formula, fmt) in enumerate(derived, start=11):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, formula); c.number_format = fmt
        if i == 15:
            tot(c)
        else:
            der(c)
    MONTHLY = "B15"

    # 4-Year TCO table — YEARS as rows, METRICS as columns
    ws["A18"] = "4-Year Total Cost of Ownership  (years as rows, metrics as columns)"
    sub(ws["A18"]); ws.merge_cells("A18:H18")
    metric_headers = ["Year", "Subscription", "Implementation", "Support",
                      "Total Customer Payment", "Total Cost",
                      "Gross Margin $", "Gross Margin %"]
    for c, h in enumerate(metric_headers, start=1):
        ws.cell(19, c, h); hdr(ws.cell(19, c))

    sup_price_yr = (
        f'12*IF({SUP}="Standard",sup_std_price,sup_prem_price)'
    )
    sup_cost_yr = (
        f'12*IF({SUP}="Standard",sup_std_cost,sup_prem_cost)'
    )
    base_op_cost_yr = (
        f"per_user_cost*{USERS}*12 + hosting_active*12"
    )

    for yr in range(1, 5):  # rows 20..23
        r = 19 + yr
        ws.cell(r, 1, f"Year {yr}").font = BOLD; ws.cell(r, 1).border = BOX
        factor = f"(1+annual_uplift_pct)^{yr - 1}"
        # Subscription
        sub_c = ws.cell(r, 2, f"={MONTHLY}*12*(1-annual_discount)*{factor}")
        sub_c.number_format = '"$"#,##0'; der(sub_c)
        # Implementation (Y1 only)
        if yr == 1:
            im = ws.cell(r, 3, "=impl_price")
        else:
            im = ws.cell(r, 3, 0)
        im.number_format = '"$"#,##0'; der(im)
        # Support (with uplift)
        sup_c = ws.cell(r, 4, f"={sup_price_yr}*{factor}")
        sup_c.number_format = '"$"#,##0'; der(sup_c)
        # Total Customer Payment
        tcp = ws.cell(r, 5, f"=B{r}+C{r}+D{r}")
        tcp.number_format = '"$"#,##0'; tot(tcp)
        # Total Cost
        if yr == 1:
            cost_f = f"=({base_op_cost_yr} + {sup_cost_yr})*{factor} + impl_cost"
        else:
            cost_f = f"=({base_op_cost_yr} + {sup_cost_yr})*{factor}"
        c = ws.cell(r, 6, cost_f); c.number_format = '"$"#,##0'; der(c)
        # GM $
        gm = ws.cell(r, 7, f"=E{r}-F{r}"); gm.number_format = '"$"#,##0'; tot(gm)
        # GM %
        gp = ws.cell(r, 8, f"=IFERROR((E{r}-F{r})/E{r},0)")
        gp.number_format = '0.0%'; der(gp)

    # 4-Year Total row
    tr = 24
    ws.cell(tr, 1, "4-Year Total").font = Font(bold=True, color="1F4E78")
    tot(ws.cell(tr, 1))
    for col in range(2, 7):
        L = get_column_letter(col)
        c = ws.cell(tr, col, f"=SUM({L}20:{L}23)")
        c.number_format = '"$"#,##0'; tot(c)
        c.font = Font(bold=True, size=11, color="C00000")
    gp_tot = ws.cell(tr, 8, f"=IFERROR((E{tr}-F{tr})/E{tr},0)")
    gp_tot.number_format = '0.0%'; tot(gp_tot)

    ws.conditional_formatting.add(
        "H20:H24",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    # Headline 4-year total reference (used by Cover sheet)
    ws.cell(26, 1, "Headline: 4-Year Total Customer Payment").font = Font(
        bold=True, size=12, color="0B3D91")
    ws.cell(26, 1).border = BOX
    h = ws.cell(26, 2, f"=E{tr}"); h.number_format = '"$"#,##0'; tot(h)
    h.font = Font(bold=True, size=12, color="C00000")
    add_name(wb, "tco_4year_total", f"'4-Year TCO'!$E${tr}")

    autosize(ws, [44, 18, 18, 18, 24, 18, 18, 14])
    ws.freeze_panes = "A4"


def build_ubl_cover(wb: Workbook) -> None:
    """One-page Cover sheet — placed first in the UBL workbook."""
    ws = wb.create_sheet("Cover", 0)

    ws.merge_cells("A1:F2")
    ws["A1"] = "CompliverseAI"
    ws["A1"].font = Font(bold=True, size=22, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:F4")
    ws["A3"] = "United Bank Limited — 4-Year Total Cost of Ownership"
    ws["A3"].font = Font(bold=True, size=16, color="0B3D91")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A5:F5")
    ws["A5"] = ("Single Enterprise platform tier  •  100 users  •  Azure SaaS  "
                "•  Premium support  •  USD")
    ws["A5"].font = Font(italic=True, size=11, color="555555")
    ws["A5"].alignment = Alignment(horizontal="center")

    rows = [
        ("Prepared for", "United Bank Limited"),
        ("Prepared by", "CompliverseAI Sales Engineering"),
        ("Quote validity", "30 days from issue"),
        ("Currency", "USD"),
        ("Hosting model", "Azure SaaS  (toggle on Assumptions sheet)"),
        ("Users", "100  (single platform pricing)"),
        ("Support tier", "Premium  (24x7, 1h response P1)"),
        ("Implementation", "One-time onboarding, migration, training"),
        ("LLM stack", "GPT-5 + Claude  (token-based, heavy usage)"),
    ]
    for i, (k, v) in enumerate(rows, start=8):
        a = ws.cell(i, 2, k); a.font = BOLD; a.border = BOX
        a.alignment = Alignment(horizontal="left")
        b = ws.cell(i, 3, v); b.border = BOX
        b.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.merge_cells(start_row=i, end_row=i, start_column=3, end_column=5)

    headline_row = 8 + len(rows) + 2
    ws.cell(headline_row, 2, "Headline 4-Year Total").font = Font(
        bold=True, size=14, color="0B3D91")
    h = ws.cell(headline_row, 3, "=tco_4year_total")
    h.number_format = '"$"#,##0'
    h.font = Font(bold=True, size=18, color="C00000")
    h.fill = PatternFill("solid", fgColor="FFF2CC")
    h.border = BOX
    h.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=headline_row, end_row=headline_row,
                   start_column=3, end_column=5)

    autosize(ws, [4, 22, 18, 18, 18, 4])


# ---------- builder ----------
def build_workbook(*, ubl: bool, path: Path, label: str) -> None:
    wb = Workbook()
    # remove the default empty sheet
    default = wb.active
    wb.remove(default)

    build_assumptions(wb, ubl=ubl)
    build_cost_buildup(wb)
    build_platform_pricing(wb)
    build_tco(wb)

    if ubl:
        build_ubl_cover(wb)

    # Sheet order: (Cover for UBL,) Assumptions, Cost Build-Up,
    # Platform Pricing, 4-Year TCO
    desired = (
        ["Cover", "Assumptions", "Cost Build-Up", "Platform Pricing", "4-Year TCO"]
        if ubl else
        ["Assumptions", "Cost Build-Up", "Platform Pricing", "4-Year TCO"]
    )
    wb._sheets = [wb[name] for name in desired]

    wb.save(path)
    print(f"[ok] {label}: {path}")


def main() -> None:
    build_workbook(ubl=False, path=GENERIC_PATH,
                   label="Generic Workbook 2 (single tier)")
    build_workbook(ubl=True, path=UBL_PATH,
                   label="UBL Quote v1 (single tier)")


if __name__ == "__main__":
    main()
