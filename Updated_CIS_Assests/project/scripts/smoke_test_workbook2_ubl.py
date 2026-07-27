"""Structural / formula sanity test for the single-Enterprise-tier
pricing workbook + UBL quote.

Asserts the model the user actually asked for:
  - One platform tier (no Starter / Pro / SMB / Mid-market sheets)
  - Fixed at users_count (default 100)
  - Support fees = Standard + Premium only
  - One implementation fee
  - API costs broken out: Azure (hosting) + GPT-5 + Claude
  - Hard floor protects margin if users drop
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

PRICING = Path(__file__).resolve().parent.parent / "pricing"
GENERIC = PRICING / "CompliverseAI_Pricing_Module_Plus_User.xlsx"
UBL     = PRICING / "CompliverseAI_UBL_Quote_v1.xlsx"

REQUIRED_NAMES = [
    "annual_discount", "target_margin", "annual_uplift_pct", "tco_horizon_years",
    "hosting_mode", "azure_hosting_per_tenant", "selfhost_managed_fee",
    "hosting_active",
    "users_count", "platform_base_fee", "per_user_fee",
    "min_platform_fee_usd", "per_user_floor",
    "gpt_input_tokens", "gpt_output_tokens",
    "gpt_input_price_per_1m", "gpt_output_price_per_1m", "api_gpt5",
    "claude_input_tokens", "claude_output_tokens",
    "claude_input_price_per_1m", "claude_output_price_per_1m", "api_claude",
    "api_llm_total",
    "api_sms", "api_cloud", "api_sso", "api_storage",
    "sup_std_price", "sup_std_cost", "sup_prem_price", "sup_prem_cost",
    "impl_price", "impl_cost",
    "per_user_cost", "tco_4year_total",
]
REQUIRED_SHEETS = ["Assumptions", "Cost Build-Up", "Platform Pricing", "4-Year TCO"]

# Sheets that MUST NOT exist anymore (multi-tier was removed)
FORBIDDEN_SHEETS = ["Pricing Tiers", "Customer Scenarios"]


def check(path: Path, label: str, *, expect_cover: bool) -> None:
    print(f"\n=== {label} ===  ({path.name})")
    assert path.exists(), f"missing file: {path}"
    wb = load_workbook(path, data_only=False)

    # Sheet presence + absence
    for s in REQUIRED_SHEETS:
        assert s in wb.sheetnames, f"missing sheet: {s}"
    for s in FORBIDDEN_SHEETS:
        assert s not in wb.sheetnames, (
            f"forbidden sheet still present: {s} — multi-tier was removed")
    if expect_cover:
        assert wb.sheetnames[0] == "Cover", \
            f"Cover must be the first sheet, got {wb.sheetnames!r}"

    # Named ranges
    for n in REQUIRED_NAMES:
        assert n in wb.defined_names, f"missing named range: {n}"

    asm = wb["Assumptions"]

    def name_value(n: str):
        ref = wb.defined_names[n].attr_text  # Assumptions!$B$10
        cell = ref.split("!", 1)[1].replace("$", "")
        return asm[cell].value

    # Inputs
    discount = name_value("annual_discount")
    uplift   = name_value("annual_uplift_pct")
    horizon  = name_value("tco_horizon_years")
    mode     = name_value("hosting_mode")
    azure    = name_value("azure_hosting_per_tenant")
    selfhost = name_value("selfhost_managed_fee")
    users    = name_value("users_count")
    base     = name_value("platform_base_fee")
    per_user = name_value("per_user_fee")
    floor_min = name_value("min_platform_fee_usd")
    floor_per = name_value("per_user_floor")

    gpt_in   = name_value("gpt_input_tokens")
    gpt_out  = name_value("gpt_output_tokens")
    gpt_in_p = name_value("gpt_input_price_per_1m")
    gpt_out_p = name_value("gpt_output_price_per_1m")
    cl_in    = name_value("claude_input_tokens")
    cl_out   = name_value("claude_output_tokens")
    cl_in_p  = name_value("claude_input_price_per_1m")
    cl_out_p = name_value("claude_output_price_per_1m")

    sms     = name_value("api_sms")
    cloud   = name_value("api_cloud")
    sso     = name_value("api_sso")
    storage = name_value("api_storage")

    sup_std_price  = name_value("sup_std_price")
    sup_prem_price = name_value("sup_prem_price")
    impl_price     = name_value("impl_price")

    print(f"  Mode={mode!r}  Users={users}  Base=${base}/mo  "
          f"Per-user=${per_user}/user/mo")
    print(f"  Floor: ${floor_min}/mo + ${floor_per}/user/mo")
    print(f"  Azure SaaS=${azure}/mo   SelfHost=${selfhost}/mo")
    print(f"  Support: Standard=${sup_std_price}, Premium=${sup_prem_price}")
    print(f"  Implementation: ${impl_price:,} one-time")

    # Defaults
    assert users == 100, f"Number of users must default to 100, got {users}"
    assert horizon == 4, "TCO horizon default must be 4"
    assert abs(uplift - 0.05) < 1e-9, "Annual uplift must default to 5%"

    # Recompute API stack
    gpt5   = (gpt_in * gpt_in_p + gpt_out * gpt_out_p) / 1_000_000
    claude = (cl_in * cl_in_p + cl_out * cl_out_p) / 1_000_000
    api_llm = gpt5 + claude
    per_user_cost = api_llm + sms + cloud + sso + storage
    hosting_active = azure if mode == "SaaS" else selfhost
    print(f"  GPT-5=${gpt5:.2f}/user/mo   Claude=${claude:.2f}/user/mo  "
          f"LLM total=${api_llm:.2f}/user/mo")
    print(f"  per_user_cost=${per_user_cost:.2f}/user/mo  (excl. hosting)  "
          f"hosting_active=${hosting_active}/tenant/mo")

    # Single-tier Platform Pricing math
    linear = base + per_user * users
    floor = floor_min + floor_per * users
    monthly = max(linear, floor)
    y1_rev = monthly * 12 * (1 - discount)
    y1_cost = per_user_cost * users * 12 + hosting_active * 12
    y1_gm = (y1_rev - y1_cost) / y1_rev if y1_rev else 0
    print(f"  Platform: linear=${linear:,.0f}  floor=${floor:,.0f}  "
          f"monthly=${monthly:,.0f}")
    print(f"  Y1 rev=${y1_rev:,.0f}  Y1 cost=${y1_cost:,.0f}  Y1 GM={y1_gm:.1%}")

    # Margin check at the chosen user count
    target = name_value("target_margin")
    assert y1_gm >= target - 1e-6, \
        f"Y1 GM {y1_gm:.1%} must be ≥ target_margin {target:.1%}"

    # Floor protection: at 25 users (deep drop) the floor must strictly bind
    linear_25 = base + per_user * 25
    floor_25 = floor_min + floor_per * 25
    assert floor_25 > linear_25, (
        f"Floor must strictly bind at 25 users (deep drop): "
        f"floor=${floor_25:,.0f} vs linear=${linear_25:,.0f}")
    # And at 50 users (mild drop) the floor must at least match the linear
    linear_50 = base + per_user * 50
    floor_50 = floor_min + floor_per * 50
    assert floor_50 >= linear_50, (
        f"Floor must be ≥ linear at 50 users: "
        f"floor=${floor_50:,.0f} vs linear=${linear_50:,.0f}")
    print(f"  Floor protection: 25 users → floor=${floor_25:,.0f} "
          f"> linear=${linear_25:,.0f};  50 users → floor=${floor_50:,.0f} "
          f"≥ linear=${linear_50:,.0f}  ✓")

    # Year-4 uplift
    y4_factor = (1 + uplift) ** 3
    print(f"  Year-4 / Year-1 ratio = {y4_factor:.4f}  (expected 1.1576)")
    assert abs(y4_factor - 1.157625) < 1e-6

    # Hosting toggle is meaningful
    assert azure != selfhost, "SaaS vs SelfHost values must differ"

    # Platform Pricing sheet structure
    pp = wb["Platform Pricing"]
    assert pp.cell(9, 1).value.startswith("CUSTOMER MONTHLY PRICE"), (
        f"Platform Pricing A9 should be the customer monthly price label, "
        f"got {pp.cell(9, 1).value!r}")
    assert pp.cell(9, 2).value == "=MAX(B7,B8)", (
        f"Customer monthly should be MAX(linear, floor), "
        f"got {pp.cell(9, 2).value!r}")
    assert pp.cell(10, 2).value == "=B8>B7", (
        f"Floor-active indicator should be boolean =B8>B7, "
        f"got {pp.cell(10, 2).value!r}")
    # Y1 summary block: sub @ B12, cost @ B13, GM% @ B14
    assert pp.cell(12, 2).value == "=B9*12*(1-annual_discount)", (
        f"Y1 sub formula should be at B12, got {pp.cell(12, 2).value!r}")
    assert pp.cell(13, 2).value == "=per_user_cost*users_count*12 + hosting_active*12", (
        f"Y1 cost formula should be at B13, got {pp.cell(13, 2).value!r}")
    assert pp.cell(14, 2).value == "=IFERROR((B12-B13)/B12,0)", (
        f"Y1 GM% must be (B12-B13)/B12 at B14, got {pp.cell(14, 2).value!r}")
    # Conditional format must target B14 (the GM% cell), not B13 (cost)
    cf_ranges = [str(r) for r in pp.conditional_formatting._cf_rules.keys()]
    assert any("B14" in r for r in cf_ranges), (
        f"Conditional format must target B14 (GM% cell), got ranges={cf_ranges}")
    assert not any(r == "B13" or r == "B13:B13" for r in cf_ranges), (
        f"Conditional format must NOT target the cost cell B13, "
        f"got ranges={cf_ranges}")

    # 4-Year TCO layout
    tco = wb["4-Year TCO"]
    assert tco.cell(19, 1).value == "Year"
    assert tco.cell(19, 2).value == "Subscription"
    assert tco.cell(20, 1).value == "Year 1"
    assert tco.cell(23, 1).value == "Year 4"
    assert tco.cell(24, 1).value == "4-Year Total"
    tco_total_ref = wb.defined_names["tco_4year_total"].attr_text
    assert "$E$24" in tco_total_ref, (
        f"tco_4year_total should point at E24, got {tco_total_ref}")
    print(f"  4-Year TCO years-as-rows layout; "
          f"tco_4year_total -> {tco_total_ref}")

    if expect_cover:
        cov = wb["Cover"]
        found = any(
            isinstance(c.value, str) and c.value == "=tco_4year_total"
            for row in cov.iter_rows() for c in row
        )
        assert found, "Cover sheet must reference =tco_4year_total"
        print("  Cover sheet headline references tco_4year_total")
        # UBL business intent: Premium support default
        assert tco.cell(5, 2).value == "Premium", (
            f"UBL TCO default support tier should be Premium, "
            f"got {tco.cell(5, 2).value!r}")
        print("  UBL default support tier confirmed: Premium")

    print(f"  PASS — {label}")


def main() -> int:
    check(GENERIC, "Generic Workbook 2 (single tier)", expect_cover=False)
    check(UBL,     "UBL Quote v1 (single tier)",       expect_cover=True)
    print("\nAll smoke checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
