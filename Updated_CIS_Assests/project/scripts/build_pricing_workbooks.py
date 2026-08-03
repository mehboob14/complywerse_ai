"""Build CompliverseAI GRC pricing workbooks (USD).

Generates two .xlsx files in pricing/:
  1. CompliverseAI_Pricing_UserBased.xlsx
  2. CompliverseAI_Pricing_Module_Plus_User.xlsx

Every derived cell is an Excel formula referencing named ranges on the
Assumptions sheet, so changing one assumption recalculates everything.
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT_DIR = Path(__file__).resolve().parent.parent / "pricing"
OUT_DIR.mkdir(exist_ok=True)

# ---- Styling helpers ---------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
BOLD = Font(name="Calibri", size=11, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(cell):
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BOX


def sub(cell):
    cell.fill = SUB_FILL
    cell.font = SUB_FONT
    cell.border = BOX


def input_cell(cell):
    cell.fill = INPUT_FILL
    cell.border = BOX


def derived(cell):
    cell.fill = DERIVED_FILL
    cell.border = BOX


def total(cell):
    cell.fill = TOTAL_FILL
    cell.font = BOLD
    cell.border = BOX


def add_name(wb, name, ref):
    """Workbook-scoped defined name."""
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Assumptions sheet (shared) ---------------------------------------------
def build_assumptions(wb: Workbook):
    ws = wb.create_sheet("Assumptions", 0)
    ws.sheet_properties.tabColor = "FFC000"

    ws["A1"] = "CompliverseAI GRC — Pricing Assumptions (USD)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Yellow cells = editable inputs. Green cells = derived. Change inputs and the whole workbook recalculates."
    ws["A3"].font = Font(italic=True, color="595959")
    ws.merge_cells("A3:D3")

    rows = [
        # (label, value, name, format, section)
        ("GLOBAL", None, None, None, "section"),
        ("Currency", "USD", None, "@", "input"),
        ("FX rate (USD = 1)", 1, "fx_rate", "0.0000", "input"),
        ("Annual prepay discount %", 0.15, "annual_discount", "0.0%", "input"),
        ("Target gross margin %", 0.60, "target_margin", "0.0%", "input"),

        ("HOSTING", None, None, None, "section"),
        ("Hosting cost $/tenant/month", 500, "hosting_per_tenant", '"$"#,##0', "input"),
        ("Avg users per tenant (for per-user allocation)", 100, "avg_users", "0", "input"),

        ("API COSTS  ($/user/month)", None, None, None, "section"),
        ("LLM (OpenAI/Anthropic/Gemini)", 0.50, "api_llm", '"$"#,##0.00', "input"),
        ("SMS / WhatsApp notifications", 0.10, "api_sms", '"$"#,##0.00', "input"),
        ("Cloud read APIs (AWS/Azure/GCP)", 1.00, "api_cloud", '"$"#,##0.00', "input"),
        ("Identity / SSO (Okta/Azure AD)", 0.30, "api_sso", '"$"#,##0.00', "input"),
        ("Storage / CDN", 0.20, "api_storage", '"$"#,##0.00', "input"),
        ("Total API cost $/user/month", "=api_llm+api_sms+api_cloud+api_sso+api_storage", "api_total", '"$"#,##0.00', "derived"),

        ("SUPPORT  (cost vs price per customer per month)", None, None, None, "section"),
        ("Basic — cost", 50, "sup_basic_cost", '"$"#,##0', "input"),
        ("Basic — price", 150, "sup_basic_price", '"$"#,##0', "input"),
        ("Standard — cost", 200, "sup_std_cost", '"$"#,##0', "input"),
        ("Standard — price", 600, "sup_std_price", '"$"#,##0', "input"),
        ("Premium — cost", 500, "sup_prem_cost", '"$"#,##0', "input"),
        ("Premium — price", 1500, "sup_prem_price", '"$"#,##0', "input"),
        ("24x7 — cost", 1500, "sup_247_cost", '"$"#,##0', "input"),
        ("24x7 — price", 4500, "sup_247_price", '"$"#,##0', "input"),

        ("IMPLEMENTATION  (one-time fee)", None, None, None, "section"),
        ("Starter — cost (delivery)", 2000, "impl_starter_cost", '"$"#,##0', "input"),
        ("Starter — price", 5000, "impl_starter_price", '"$"#,##0', "input"),
        ("Pro — cost", 6000, "impl_pro_cost", '"$"#,##0', "input"),
        ("Pro — price", 15000, "impl_pro_price", '"$"#,##0', "input"),
        ("Enterprise — cost", 16000, "impl_ent_cost", '"$"#,##0', "input"),
        ("Enterprise — price", 40000, "impl_ent_price", '"$"#,##0', "input"),

        ("VOLUME DISCOUNT  (applied to per-user list price by user count)", None, None, None, "section"),
        ("Bracket 1 max users", 25, "vol_b1_max", "0", "input"),
        ("Bracket 1 discount %", 0.00, "vol_b1_disc", "0.0%", "input"),
        ("Bracket 2 max users", 100, "vol_b2_max", "0", "input"),
        ("Bracket 2 discount %", 0.10, "vol_b2_disc", "0.0%", "input"),
        ("Bracket 3 max users", 500, "vol_b3_max", "0", "input"),
        ("Bracket 3 discount %", 0.20, "vol_b3_disc", "0.0%", "input"),
        ("Bracket 4 (above bracket 3) discount %", 0.30, "vol_b4_disc", "0.0%", "input"),

        ("PARTNER / RESELLER DISCOUNT", None, None, None, "section"),
        ("Bronze %", 0.10, "partner_bronze", "0.0%", "input"),
        ("Silver %", 0.15, "partner_silver", "0.0%", "input"),
        ("Gold %", 0.20, "partner_gold", "0.0%", "input"),
        ("Platinum %", 0.25, "partner_platinum", "0.0%", "input"),

        ("CUSTOMER PERSONAS  (used in scenarios)", None, None, None, "section"),
        ("SMB — users", 25, "p_smb_users", "0", "input"),
        ("Mid-Market — users", 150, "p_mid_users", "0", "input"),
        ("Enterprise — users", 750, "p_ent_users", "0", "input"),
    ]

    r = 5
    for label, val, name, fmt, kind in rows:
        if kind == "section":
            ws.cell(r, 1, label)
            sub(ws.cell(r, 1))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1
            continue
        ws.cell(r, 1, label).border = BOX
        c = ws.cell(r, 2, val)
        if fmt:
            c.number_format = fmt
        if kind == "input":
            input_cell(c)
        else:
            derived(c)
        if name:
            add_name(wb, name, f"Assumptions!${get_column_letter(2)}${r}")
        r += 1

    autosize(ws, [55, 18, 4, 4])
    ws.freeze_panes = "A4"
    return ws


# ---- Cost build-up sheet -----------------------------------------------------
def build_cost_buildup_userbased(wb: Workbook):
    ws = wb.create_sheet("Cost Build-Up")
    ws.sheet_properties.tabColor = "00B050"

    ws["A1"] = "Per-User Cost Build-Up (Monthly)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:C1")

    headers = ["Cost component", "Value", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)
        hdr(ws.cell(3, c))

    rows = [
        ("Hosting allocation per user", "=hosting_per_tenant/avg_users", "Flat tenant hosting / avg users"),
        ("LLM API per user", "=api_llm", ""),
        ("SMS / WhatsApp per user", "=api_sms", ""),
        ("Cloud read APIs per user", "=api_cloud", ""),
        ("SSO per user", "=api_sso", ""),
        ("Storage / CDN per user", "=api_storage", ""),
        ("Subtotal per-user variable cost", "=SUM(B5:B10)", "Excludes support; support is per-customer"),
    ]
    for i, (lbl, formula, note) in enumerate(rows, start=4):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, formula)
        c.number_format = '"$"#,##0.00'
        if i == 10:
            total(c)
        else:
            derived(c)
        ws.cell(i, 3, note).border = BOX

    autosize(ws, [42, 16, 50])
    ws.freeze_panes = "A4"


def build_cost_buildup_module(wb: Workbook):
    ws = wb.create_sheet("Cost Build-Up")
    ws.sheet_properties.tabColor = "00B050"

    ws["A1"] = "Per-User + Per-Module Cost Build-Up (Monthly)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:C1")

    # Reuse user costs
    headers = ["Cost component", "Value", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h)
        hdr(ws.cell(3, c))

    user_rows = [
        ("Hosting allocation per user", "=hosting_per_tenant/avg_users"),
        ("LLM API per user", "=api_llm"),
        ("SMS / WhatsApp per user", "=api_sms"),
        ("Cloud read APIs per user", "=api_cloud"),
        ("SSO per user", "=api_sso"),
        ("Storage / CDN per user", "=api_storage"),
        ("Subtotal per-user variable cost", "=SUM(B4:B9)"),
    ]
    for i, (lbl, formula) in enumerate(user_rows, start=4):
        ws.cell(i, 1, lbl).border = BOX
        c = ws.cell(i, 2, formula)
        c.number_format = '"$"#,##0.00'
        if i == 10:
            total(c)
        else:
            derived(c)
    add_name(wb, "per_user_cost", "'Cost Build-Up'!$B$10")

    # Module costs (engineering + maintenance amortized monthly)
    ws["A12"] = "Per-Module Allocated Cost (engineering + ops, $/customer/month)"
    sub(ws["A12"])
    ws.merge_cells("A12:C12")

    modules = [
        ("Compliance", 80),
        ("Risk / ERM", 80),
        ("Governance", 60),
        ("Vendor Risk", 60),
        ("Vulnerability Management", 90),
        ("IT Asset Management", 50),
        ("Workflow Engine / Automation", 70),
        ("AI Copilot", 100),
        ("Integrations Hub", 70),
        ("Plugin Library / CIS", 60),
        ("Reporting & Dashboards", 50),
    ]
    ws.cell(13, 1, "Module"); hdr(ws.cell(13, 1))
    ws.cell(13, 2, "Cost $/cust/mo"); hdr(ws.cell(13, 2))
    ws.cell(13, 3, "Notes"); hdr(ws.cell(13, 3))
    for i, (m, c) in enumerate(modules, start=14):
        ws.cell(i, 1, m).border = BOX
        v = ws.cell(i, 2, c); v.number_format = '"$"#,##0'; input_cell(v)
        ws.cell(i, 3, "").border = BOX
    last = 13 + len(modules)
    ws.cell(last + 1, 1, "Total all 11 modules"); total(ws.cell(last + 1, 1))
    tc = ws.cell(last + 1, 2, f"=SUM(B14:B{last})"); tc.number_format = '"$"#,##0'; total(tc)

    autosize(ws, [38, 18, 50])
    ws.freeze_panes = "A4"


# ---- Pricing Tiers sheet -----------------------------------------------------
def build_pricing_tiers_userbased(wb: Workbook):
    ws = wb.create_sheet("Pricing Tiers")
    ws.sheet_properties.tabColor = "4472C4"

    ws["A1"] = "User-Based Pricing Tiers (per user / month, USD)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")

    headers = ["Tier", "List $/user/mo", "Annual $/user/mo (after discount)",
               "Per-user cost", "Gross margin %", "Margin $ (mo)",
               "Implementation (one-time price)", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h); hdr(ws.cell(3, c))

    tiers = [
        ("Starter", 25, "impl_starter_price", "Up to 25 users; Compliance + Risk only; Basic support"),
        ("Pro",     50, "impl_pro_price",     "Up to 250 users; all modules; Standard support"),
        ("Enterprise", 90, "impl_ent_price",  "Unlimited users; SSO, advanced workflows, Premium/24x7 support"),
    ]
    for i, (name, price, impl_name, note) in enumerate(tiers, start=4):
        ws.cell(i, 1, name).border = BOX; ws.cell(i, 1).font = BOLD
        p = ws.cell(i, 2, price); p.number_format = '"$"#,##0.00'; input_cell(p)
        a = ws.cell(i, 3, f"=B{i}*(1-annual_discount)"); a.number_format = '"$"#,##0.00'; derived(a)
        c = ws.cell(i, 4, "='Cost Build-Up'!$B$10"); c.number_format = '"$"#,##0.00'; derived(c)
        m = ws.cell(i, 5, f"=IFERROR((B{i}-D{i})/B{i},0)"); m.number_format = '0.0%'; derived(m)
        md = ws.cell(i, 6, f"=B{i}-D{i}"); md.number_format = '"$"#,##0.00'; derived(md)
        impl = ws.cell(i, 7, f"={impl_name}"); impl.number_format = '"$"#,##0'; derived(impl)
        ws.cell(i, 8, note).border = BOX

    # Conditional format on margin column
    ws.conditional_formatting.add(
        f"E4:E{3 + len(tiers)}",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    # Volume discount lookup helper
    ws["A8"] = "Volume Discount Reference"
    sub(ws["A8"]); ws.merge_cells("A8:H8")
    vol_hdr = ["Up to N users", "Discount %"]
    for c, h in enumerate(vol_hdr, start=1):
        ws.cell(9, c, h); hdr(ws.cell(9, c))
    vol_rows = [
        ("=vol_b1_max", "=vol_b1_disc"),
        ("=vol_b2_max", "=vol_b2_disc"),
        ("=vol_b3_max", "=vol_b3_disc"),
        ('"500+"',      "=vol_b4_disc"),
    ]
    for i, (n, d) in enumerate(vol_rows, start=10):
        nv = ws.cell(i, 1, n if n.startswith("=") else f"={n}"); derived(nv)
        if not n.startswith("="):
            nv.value = "500+"
        dv = ws.cell(i, 2, d); dv.number_format = '0.0%'; derived(dv)

    autosize(ws, [14, 16, 24, 16, 16, 16, 26, 60])
    ws.freeze_panes = "A4"


def build_pricing_tiers_module(wb: Workbook):
    ws = wb.create_sheet("Pricing Tiers")
    ws.sheet_properties.tabColor = "4472C4"

    ws["A1"] = "Bundled Tiers (Base + Modules + Users)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")

    headers = ["Tier", "Base platform $/mo", "Per-user $/mo",
               "Modules included", "Annual $/user (after discount)",
               "Implementation price", "Margin % (per user, modules+base allocated)", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h); hdr(ws.cell(3, c))

    tiers = [
        ("Starter",     500, 15, "3 (Compliance, Risk, Reporting)", "impl_starter_price",
         "Small teams getting started; basic compliance"),
        ("Pro",        1500, 30, "8 modules (everything except AI Copilot, CIS Plugins, Integrations Hub)",
         "impl_pro_price", "Mid-market all-in-one"),
        ("Enterprise", 3500, 50, "All 11 modules", "impl_ent_price",
         "Unlimited users, SSO, AI Copilot, dedicated CSM"),
    ]
    for i, (name, base, per_user, mods, impl_name, note) in enumerate(tiers, start=4):
        ws.cell(i, 1, name).border = BOX; ws.cell(i, 1).font = BOLD
        b = ws.cell(i, 2, base); b.number_format = '"$"#,##0'; input_cell(b)
        u = ws.cell(i, 3, per_user); u.number_format = '"$"#,##0'; input_cell(u)
        ws.cell(i, 4, mods).border = BOX
        a = ws.cell(i, 5, f"=C{i}*(1-annual_discount)"); a.number_format = '"$"#,##0.00'; derived(a)
        impl = ws.cell(i, 6, f"={impl_name}"); impl.number_format = '"$"#,##0'; derived(impl)
        # Margin per user assuming 100 users + base allocated + modules cost allocated
        # cost_per_user = per_user_cost + (base_alloc/100) + (modules_total/100)
        m = ws.cell(
            i, 7,
            f"=IFERROR((C{i}-(per_user_cost + (hosting_per_tenant/avg_users) + ('Cost Build-Up'!$B${14 + 11}/avg_users)))/C{i},0)"
        )
        m.number_format = '0.0%'; derived(m)
        ws.cell(i, 8, note).border = BOX

    ws.conditional_formatting.add(
        f"G4:G{3 + len(tiers)}",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    # Module a-la-carte pricing
    ws["A8"] = "Module à-la-carte Pricing (per customer per month)"
    sub(ws["A8"]); ws.merge_cells("A8:H8")
    mh = ["Module", "Cost $/cust/mo", "List price $/mo", "Annual $/mo", "Margin %"]
    for c, h in enumerate(mh, start=1):
        ws.cell(9, c, h); hdr(ws.cell(9, c))
    modules_pricing = [
        ("Compliance",                 200),
        ("Risk / ERM",                 200),
        ("Governance",                 150),
        ("Vendor Risk",                150),
        ("Vulnerability Management",   250),
        ("IT Asset Management",        120),
        ("Workflow Engine / Automation", 180),
        ("AI Copilot",                 300),
        ("Integrations Hub",           180),
        ("Plugin Library / CIS",       150),
        ("Reporting & Dashboards",     120),
    ]
    for i, (m, price) in enumerate(modules_pricing, start=10):
        ws.cell(i, 1, m).border = BOX
        # Cost references the corresponding row in Cost Build-Up (rows 14..24)
        cost = ws.cell(i, 2, f"='Cost Build-Up'!$B${14 + (i - 10)}")
        cost.number_format = '"$"#,##0'; derived(cost)
        p = ws.cell(i, 3, price); p.number_format = '"$"#,##0'; input_cell(p)
        a = ws.cell(i, 4, f"=C{i}*(1-annual_discount)*12"); a.number_format = '"$"#,##0'; derived(a)
        marg = ws.cell(i, 5, f"=IFERROR((C{i}-B{i})/C{i},0)"); marg.number_format = '0.0%'; derived(marg)

    ws.conditional_formatting.add(
        "E10:E20",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    autosize(ws, [14, 18, 16, 38, 22, 18, 28, 50])
    ws.freeze_panes = "A4"


# ---- Customer Scenarios sheet -----------------------------------------------
def build_scenarios_userbased(wb: Workbook):
    ws = wb.create_sheet("Customer Scenarios")
    ws.sheet_properties.tabColor = "7030A0"

    ws["A1"] = "Customer-Size Scenarios (User-Based Model)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:H1")

    headers = ["Persona", "Users", "Tier", "$/user/mo (list)",
               "Volume disc %", "Net $/user/mo", "Monthly ACV", "Annual ACV"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h); hdr(ws.cell(3, c))

    personas = [
        ("SMB",         "p_smb_users", "Starter",    "='Pricing Tiers'!$B$4"),
        ("Mid-Market",  "p_mid_users", "Pro",        "='Pricing Tiers'!$B$5"),
        ("Enterprise",  "p_ent_users", "Enterprise", "='Pricing Tiers'!$B$6"),
    ]
    for i, (name, users_ref, tier, price_ref) in enumerate(personas, start=4):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        u = ws.cell(i, 2, f"={users_ref}"); u.number_format = '0'; derived(u)
        ws.cell(i, 3, tier).border = BOX
        p = ws.cell(i, 4, price_ref); p.number_format = '"$"#,##0.00'; derived(p)
        d = ws.cell(
            i, 5,
            f"=IF(B{i}<=vol_b1_max,vol_b1_disc,IF(B{i}<=vol_b2_max,vol_b2_disc,IF(B{i}<=vol_b3_max,vol_b3_disc,vol_b4_disc)))"
        )
        d.number_format = '0.0%'; derived(d)
        n = ws.cell(i, 6, f"=D{i}*(1-E{i})"); n.number_format = '"$"#,##0.00'; derived(n)
        macv = ws.cell(i, 7, f"=B{i}*F{i}"); macv.number_format = '"$"#,##0'; derived(macv)
        aacv = ws.cell(i, 8, f"=G{i}*12*(1-annual_discount)"); aacv.number_format = '"$"#,##0'; total(aacv)

    # Year-1 vs Year-2 economics
    ws["A8"] = "Year-1 vs Year-2 Economics"
    sub(ws["A8"]); ws.merge_cells("A8:H8")
    sub_h = ["Persona", "Annual ACV", "Implementation (one-time)", "Support tier", "Support $/yr",
             "Year-1 Total", "Year-2 Recurring", "Y1 Gross Margin %"]
    for c, h in enumerate(sub_h, start=1):
        ws.cell(9, c, h); hdr(ws.cell(9, c))

    sup_rows = [
        ("SMB",        "=H4", "=impl_starter_price", "Basic",    "=sup_basic_price*12", "=sup_basic_cost*12", "=impl_starter_cost"),
        ("Mid-Market", "=H5", "=impl_pro_price",     "Standard", "=sup_std_price*12",   "=sup_std_cost*12",   "=impl_pro_cost"),
        ("Enterprise", "=H6", "=impl_ent_price",     "Premium",  "=sup_prem_price*12",  "=sup_prem_cost*12",  "=impl_ent_cost"),
    ]
    for i, (name, acv, impl, st, sup_yr, sup_cost_yr, impl_cost) in enumerate(sup_rows, start=10):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        a = ws.cell(i, 2, acv); a.number_format = '"$"#,##0'; derived(a)
        im = ws.cell(i, 3, impl); im.number_format = '"$"#,##0'; derived(im)
        ws.cell(i, 4, st).border = BOX
        s = ws.cell(i, 5, sup_yr); s.number_format = '"$"#,##0'; derived(s)
        # Year-1 total = ACV + Implementation + Support
        y1 = ws.cell(i, 6, f"=B{i}+C{i}+E{i}"); y1.number_format = '"$"#,##0'; total(y1)
        # Year-2 recurring = ACV + Support (no impl)
        y2 = ws.cell(i, 7, f"=B{i}+E{i}"); y2.number_format = '"$"#,##0'; total(y2)
        # GM% Year 1: revenue minus (per-user cost * users * 12 + hosting*12 + impl_cost + support_cost)
        users_ref = f"B{i - 6}"  # rows 4..6
        gm = ws.cell(
            i, 8,
            f"=IFERROR((F{i} - (per_user_cost*{users_ref}*12 + hosting_per_tenant*12 + {impl_cost} + {sup_cost_yr}))/F{i},0)"
        )
        gm.number_format = '0.0%'; derived(gm)

    ws.conditional_formatting.add(
        "H10:H12",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    autosize(ws, [14, 10, 14, 18, 18, 22, 22, 24])
    ws.freeze_panes = "A4"


def build_scenarios_module(wb: Workbook):
    ws = wb.create_sheet("Customer Scenarios")
    ws.sheet_properties.tabColor = "7030A0"

    ws["A1"] = "Customer-Size Scenarios (Base + Modules + Users)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:I1")

    headers = ["Persona", "Users", "Tier", "Base $/mo", "Per-user $/mo",
               "Volume disc %", "Net per-user $/mo", "Monthly ACV", "Annual ACV"]
    for c, h in enumerate(headers, start=1):
        ws.cell(3, c, h); hdr(ws.cell(3, c))

    personas = [
        ("SMB",        "p_smb_users", "Starter",    "='Pricing Tiers'!$B$4", "='Pricing Tiers'!$C$4"),
        ("Mid-Market", "p_mid_users", "Pro",        "='Pricing Tiers'!$B$5", "='Pricing Tiers'!$C$5"),
        ("Enterprise", "p_ent_users", "Enterprise", "='Pricing Tiers'!$B$6", "='Pricing Tiers'!$C$6"),
    ]
    for i, (name, users_ref, tier, base_ref, user_ref) in enumerate(personas, start=4):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        u = ws.cell(i, 2, f"={users_ref}"); u.number_format = '0'; derived(u)
        ws.cell(i, 3, tier).border = BOX
        b = ws.cell(i, 4, base_ref); b.number_format = '"$"#,##0'; derived(b)
        p = ws.cell(i, 5, user_ref); p.number_format = '"$"#,##0.00'; derived(p)
        d = ws.cell(
            i, 6,
            f"=IF(B{i}<=vol_b1_max,vol_b1_disc,IF(B{i}<=vol_b2_max,vol_b2_disc,IF(B{i}<=vol_b3_max,vol_b3_disc,vol_b4_disc)))"
        )
        d.number_format = '0.0%'; derived(d)
        n = ws.cell(i, 7, f"=E{i}*(1-F{i})"); n.number_format = '"$"#,##0.00'; derived(n)
        macv = ws.cell(i, 8, f"=D{i}+B{i}*G{i}"); macv.number_format = '"$"#,##0'; derived(macv)
        aacv = ws.cell(i, 9, f"=H{i}*12*(1-annual_discount)"); aacv.number_format = '"$"#,##0'; total(aacv)

    ws["A8"] = "Year-1 vs Year-2 Economics"
    sub(ws["A8"]); ws.merge_cells("A8:I8")
    sub_h = ["Persona", "Annual ACV", "Implementation", "Support tier", "Support $/yr",
             "Year-1 Total", "Year-2 Recurring", "Y1 Gross Margin %", "Notes"]
    for c, h in enumerate(sub_h, start=1):
        ws.cell(9, c, h); hdr(ws.cell(9, c))

    sup_rows = [
        ("SMB",        "=I4", "=impl_starter_price", "Basic",    "=sup_basic_price*12", "=sup_basic_cost*12", "=impl_starter_cost", "3 modules included"),
        ("Mid-Market", "=I5", "=impl_pro_price",     "Standard", "=sup_std_price*12",   "=sup_std_cost*12",   "=impl_pro_cost",     "8 modules included"),
        ("Enterprise", "=I6", "=impl_ent_price",     "Premium",  "=sup_prem_price*12",  "=sup_prem_cost*12",  "=impl_ent_cost",     "All 11 modules"),
    ]
    for i, (name, acv, impl, st, sup_yr, sup_cost_yr, impl_cost, note) in enumerate(sup_rows, start=10):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        a = ws.cell(i, 2, acv); a.number_format = '"$"#,##0'; derived(a)
        im = ws.cell(i, 3, impl); im.number_format = '"$"#,##0'; derived(im)
        ws.cell(i, 4, st).border = BOX
        s = ws.cell(i, 5, sup_yr); s.number_format = '"$"#,##0'; derived(s)
        y1 = ws.cell(i, 6, f"=B{i}+C{i}+E{i}"); y1.number_format = '"$"#,##0'; total(y1)
        y2 = ws.cell(i, 7, f"=B{i}+E{i}"); y2.number_format = '"$"#,##0'; total(y2)
        users_ref = f"B{i - 6}"
        # Cost: per_user_cost * users * 12 + hosting * 12 + module_total_cost*12 + impl_cost + support
        # Use total module cost from Cost Build-Up cell B25 (sum row)
        gm = ws.cell(
            i, 8,
            f"=IFERROR((F{i} - (per_user_cost*{users_ref}*12 + hosting_per_tenant*12 + 'Cost Build-Up'!$B$25*12 + {impl_cost} + {sup_cost_yr}))/F{i},0)"
        )
        gm.number_format = '0.0%'; derived(gm)
        ws.cell(i, 9, note).border = BOX

    ws.conditional_formatting.add(
        "H10:H12",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    autosize(ws, [14, 10, 14, 14, 16, 16, 20, 20, 22, 28])
    ws.freeze_panes = "A4"


# ---- Competitor Benchmark sheet ---------------------------------------------
def build_competitor(wb: Workbook):
    ws = wb.create_sheet("Competitor Benchmark")
    ws.sheet_properties.tabColor = "C00000"

    ws["A1"] = "Competitor Pricing Benchmark (USD, public list pricing as of 2025-2026)"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:G1")

    ws["A3"] = "Note: most enterprise GRC vendors do NOT publish list pricing — figures below are from public reviews, G2/Capterra reports, customer-disclosed quotes, and analyst research. Treat as directional, not authoritative."
    ws["A3"].font = Font(italic=True, color="595959")
    ws.merge_cells("A3:G3")
    ws.row_dimensions[3].height = 32

    headers = ["Competitor", "Pricing Model", "Entry / Starter (annual)",
               "Mid-Tier (annual)", "Enterprise (annual)", "Source", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(5, c, h); hdr(ws.cell(5, c))

    rows = [
        ("Vanta",          "Per-framework + per-employee", 11000, 28000, 60000,
         "https://www.vanta.com/pricing",
         "SOC2-only starter ~$11k; multi-framework + advanced ~$28k+; enterprise custom"),
        ("Drata",          "Per-employee (50/100/250+ tiers)", 12000, 30000, 75000,
         "https://drata.com/pricing",
         "Bundled controls automation; pricing scales with headcount"),
        ("AuditBoard",     "Module bundles + per-user, enterprise quote", 50000, 100000, 250000,
         "https://www.auditboard.com",
         "Sales-led; Internal audit / SOX-heavy; enterprise-only"),
        ("ServiceNow GRC", "Subscription per user + module", 75000, 150000, 400000,
         "https://www.servicenow.com/products/governance-risk-and-compliance.html",
         "Requires ServiceNow platform license; very enterprise"),
        ("Hyperproof",     "Per-framework + per-user", 15000, 35000, 80000,
         "https://hyperproof.io/pricing/",
         "Continuous compliance; mid-market focus"),
        ("OneTrust GRC",   "Module + per-user, enterprise quote", 30000, 75000, 200000,
         "https://www.onetrust.com/products/grc/",
         "Privacy heritage; broad GRC suite"),
        ("MetricStream",   "Module + per-user, enterprise quote", 60000, 140000, 350000,
         "https://www.metricstream.com",
         "Legacy enterprise GRC; on-prem option"),
        ("Sprinto",        "Per-employee, framework-led", 8000, 20000, 50000,
         "https://sprinto.com/pricing/",
         "SMB-focused; SOC2/ISO/HIPAA"),
        ("Secureframe",    "Per-framework + per-employee", 12000, 30000, 70000,
         "https://secureframe.com/pricing",
         "Vanta/Drata competitor; mid-market"),
    ]
    for i, (comp, model, e, m, ent, src, note) in enumerate(rows, start=6):
        ws.cell(i, 1, comp).font = BOLD; ws.cell(i, 1).border = BOX
        ws.cell(i, 2, model).border = BOX
        for col, v in [(3, e), (4, m), (5, ent)]:
            c = ws.cell(i, col, v); c.number_format = '"$"#,##0'; c.border = BOX
        ws.cell(i, 6, src).border = BOX
        ws.cell(i, 7, note).border = BOX

    # CompliverseAI row pulled from Customer Scenarios annual ACV
    last = 5 + len(rows) + 1
    ws.cell(last, 1, "CompliverseAI (this model)").font = Font(bold=True, color="1F4E78")
    total(ws.cell(last, 1))
    ws.cell(last, 2, "See Customer Scenarios sheet").border = BOX
    e1 = ws.cell(last, 3, "='Customer Scenarios'!H4"); e1.number_format = '"$"#,##0'; total(e1)
    e2 = ws.cell(last, 4, "='Customer Scenarios'!H5"); e2.number_format = '"$"#,##0'; total(e2)
    e3 = ws.cell(last, 5, "='Customer Scenarios'!H6"); e3.number_format = '"$"#,##0'; total(e3)
    ws.cell(last, 6, "Internal").border = BOX
    ws.cell(last, 7, "Annual ACV at SMB / Mid / Enterprise persona sizes").border = BOX

    autosize(ws, [22, 36, 22, 22, 22, 50, 60])
    ws.freeze_panes = "A6"


def build_competitor_module(wb: Workbook):
    """Module-version pulls from Customer Scenarios col I (Annual ACV in module wb)."""
    build_competitor(wb)
    ws = wb["Competitor Benchmark"]
    # Override the CompliverseAI row references — col I in module workbook
    last = 5 + 9 + 1
    ws.cell(last, 3).value = "='Customer Scenarios'!I4"
    ws.cell(last, 4).value = "='Customer Scenarios'!I5"
    ws.cell(last, 5).value = "='Customer Scenarios'!I6"


# ---- Partner / Reseller sheet -----------------------------------------------
def build_partner(wb: Workbook, scenario_acv_col: str):
    ws = wb.create_sheet("Partner_Reseller")
    ws.sheet_properties.tabColor = "ED7D31"

    ws["A1"] = "Partner / Reseller Discount Model"
    ws["A1"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A1:G1")

    ws["A3"] = "Discount applied to list ACV. Partner sells at list, keeps the discount as margin. CompliverseAI nets (list − discount). Set partner discounts on Assumptions sheet."
    ws["A3"].font = Font(italic=True, color="595959")
    ws.merge_cells("A3:G3")

    headers = ["Tier", "Discount %", "SMB net to us", "Mid-Market net to us",
               "Enterprise net to us", "Partner margin (Mid-Market)", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(5, c, h); hdr(ws.cell(5, c))

    tiers = [
        ("Bronze",   "partner_bronze",   "1-3 deals/year, basic enablement"),
        ("Silver",   "partner_silver",   "4-9 deals/year, co-marketing"),
        ("Gold",     "partner_gold",     "10-24 deals/year, dedicated PAM"),
        ("Platinum", "partner_platinum", "25+ deals/year, joint roadmap, MDF"),
    ]
    for i, (name, disc, note) in enumerate(tiers, start=6):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        d = ws.cell(i, 2, f"={disc}"); d.number_format = '0.0%'; derived(d)
        smb = ws.cell(i, 3, f"='Customer Scenarios'!{scenario_acv_col}4*(1-B{i})"); smb.number_format = '"$"#,##0'; derived(smb)
        mid = ws.cell(i, 4, f"='Customer Scenarios'!{scenario_acv_col}5*(1-B{i})"); mid.number_format = '"$"#,##0'; derived(mid)
        ent = ws.cell(i, 5, f"='Customer Scenarios'!{scenario_acv_col}6*(1-B{i})"); ent.number_format = '"$"#,##0'; derived(ent)
        margin = ws.cell(i, 6, f"='Customer Scenarios'!{scenario_acv_col}5*B{i}"); margin.number_format = '"$"#,##0'; total(margin)
        ws.cell(i, 7, note).border = BOX

    autosize(ws, [12, 14, 18, 22, 22, 24, 50])
    ws.freeze_panes = "A6"


# ---- Summary sheet ----------------------------------------------------------
def build_summary(wb: Workbook, scenario_acv_col: str, model_label: str):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "1F4E78"

    ws["A1"] = f"Executive Summary — {model_label}"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:E1")

    ws["A3"] = "Snapshot — 3 customer-size personas"
    sub(ws["A3"]); ws.merge_cells("A3:E3")
    headers = ["Persona", "Users", "Year-1 Total", "Year-2 Recurring", "Y1 Gross Margin %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h); hdr(ws.cell(4, c))
    for i, persona in enumerate(["SMB", "Mid-Market", "Enterprise"], start=5):
        ws.cell(i, 1, persona).font = BOLD; ws.cell(i, 1).border = BOX
        scn_row = i - 1  # Scenarios row 4..6
        u = ws.cell(i, 2, f"='Customer Scenarios'!B{scn_row}"); u.number_format = '0'; derived(u)
        # Year-1 total = Scenarios col F at row 10..12
        sub_row = scn_row + 6  # 10..12
        y1 = ws.cell(i, 3, f"='Customer Scenarios'!F{sub_row}"); y1.number_format = '"$"#,##0'; total(y1)
        y2 = ws.cell(i, 4, f"='Customer Scenarios'!G{sub_row}"); y2.number_format = '"$"#,##0'; total(y2)
        gm = ws.cell(i, 5, f"='Customer Scenarios'!H{sub_row}"); gm.number_format = '0.0%'; derived(gm)

    ws.conditional_formatting.add(
        "E5:E7",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    # Quote calculator
    ws["A10"] = "Quick Quote Calculator  (edit yellow cells)"
    sub(ws["A10"]); ws.merge_cells("A10:E10")

    calc_rows = [
        ("User count",                  100,  None,   "0",            "input"),
        ("Tier",                        "Pro", None,  "@",            "input"),
        ("Annual prepay? (TRUE/FALSE)", True,  None,  "@",            "input"),
        ("Support tier (Basic/Standard/Premium/24x7)", "Standard", None, "@", "input"),
        ("Partner level (None/Bronze/Silver/Gold/Platinum)", "None", None, "@", "input"),
    ]
    if model_label.startswith("Module"):
        calc_rows.insert(2, ("Modules included (Starter=3 / Pro=8 / Enterprise=11)", "All 11", None, "@", "input"))

    r = 11
    for label, val, _, fmt, _ in calc_rows:
        ws.cell(r, 1, label).border = BOX
        c = ws.cell(r, 2, val); c.number_format = fmt; input_cell(c)
        r += 1
    calc_end = r - 1

    # Outputs
    ws.cell(r, 1, "List $/user/mo").font = BOLD; ws.cell(r, 1).border = BOX
    ws.cell(r, 2, f'=IF(B{calc_end-3}="Starter",\'Pricing Tiers\'!$B$4,IF(B{calc_end-3}="Pro",\'Pricing Tiers\'!$B$5,\'Pricing Tiers\'!$B$6))').number_format = '"$"#,##0.00'
    derived(ws.cell(r, 2))
    list_row = r
    r += 1

    ws.cell(r, 1, "Volume discount %").font = BOLD; ws.cell(r, 1).border = BOX
    user_row = calc_end - 4 if model_label.startswith("Module") else calc_end - 4
    # User count is the FIRST input — row 11
    ws.cell(r, 2, "=IF(B11<=vol_b1_max,vol_b1_disc,IF(B11<=vol_b2_max,vol_b2_disc,IF(B11<=vol_b3_max,vol_b3_disc,vol_b4_disc)))").number_format = '0.0%'
    derived(ws.cell(r, 2))
    vol_row = r
    r += 1

    ws.cell(r, 1, "Annual prepay discount %").font = BOLD; ws.cell(r, 1).border = BOX
    prepay_row_in_inputs = 13  # third input
    ws.cell(r, 2, f"=IF(B{prepay_row_in_inputs}=TRUE,annual_discount,0)").number_format = '0.0%'
    derived(ws.cell(r, 2))
    prepay_row = r
    r += 1

    ws.cell(r, 1, "Net $/user/mo").font = BOLD; ws.cell(r, 1).border = BOX
    ws.cell(r, 2, f"=B{list_row}*(1-B{vol_row})*(1-B{prepay_row})").number_format = '"$"#,##0.00'
    derived(ws.cell(r, 2))
    net_row = r
    r += 1

    if model_label.startswith("Module"):
        ws.cell(r, 1, "Base $/mo").font = BOLD; ws.cell(r, 1).border = BOX
        ws.cell(r, 2, f'=IF(B{calc_end-4}="Starter",\'Pricing Tiers\'!$B$4,IF(B{calc_end-4}="Pro",\'Pricing Tiers\'!$B$5,\'Pricing Tiers\'!$B$6))').number_format = '"$"#,##0'
        derived(ws.cell(r, 2))
        base_row = r
        r += 1
        ws.cell(r, 1, "Monthly subscription").font = BOLD; ws.cell(r, 1).border = BOX
        m = ws.cell(r, 2, f"=B{base_row}+B11*B{net_row}"); m.number_format = '"$"#,##0'; total(m)
        mo_row = r
        r += 1
    else:
        ws.cell(r, 1, "Monthly subscription").font = BOLD; ws.cell(r, 1).border = BOX
        m = ws.cell(r, 2, f"=B11*B{net_row}"); m.number_format = '"$"#,##0'; total(m)
        mo_row = r
        r += 1

    ws.cell(r, 1, "Annual subscription").font = BOLD; ws.cell(r, 1).border = BOX
    a = ws.cell(r, 2, f"=B{mo_row}*12"); a.number_format = '"$"#,##0'; total(a)
    ann_row = r
    r += 1

    sup_input_row = calc_end - 1
    ws.cell(r, 1, "Support $/yr").font = BOLD; ws.cell(r, 1).border = BOX
    ws.cell(
        r, 2,
        f'=12*IF(B{sup_input_row}="Basic",sup_basic_price,IF(B{sup_input_row}="Standard",sup_std_price,IF(B{sup_input_row}="Premium",sup_prem_price,sup_247_price)))'
    ).number_format = '"$"#,##0'
    derived(ws.cell(r, 2))
    sup_row = r
    r += 1

    ws.cell(r, 1, "Implementation (one-time)").font = BOLD; ws.cell(r, 1).border = BOX
    tier_input_row = 12 if not model_label.startswith("Module") else 12
    ws.cell(r, 2, f'=IF(B{tier_input_row}="Starter",impl_starter_price,IF(B{tier_input_row}="Pro",impl_pro_price,impl_ent_price))').number_format = '"$"#,##0'
    derived(ws.cell(r, 2))
    impl_row = r
    r += 1

    partner_input_row = calc_end
    ws.cell(r, 1, "Partner discount %").font = BOLD; ws.cell(r, 1).border = BOX
    ws.cell(
        r, 2,
        f'=IF(B{partner_input_row}="Bronze",partner_bronze,IF(B{partner_input_row}="Silver",partner_silver,IF(B{partner_input_row}="Gold",partner_gold,IF(B{partner_input_row}="Platinum",partner_platinum,0))))'
    ).number_format = '0.0%'
    derived(ws.cell(r, 2))
    pdisc_row = r
    r += 1

    ws.cell(r, 1, "YEAR-1 TOTAL TO COMPLIVERSEAI (after partner)").font = Font(bold=True, color="1F4E78")
    total(ws.cell(r, 1))
    f = ws.cell(r, 2, f"=(B{ann_row}+B{sup_row}+B{impl_row})*(1-B{pdisc_row})"); f.number_format = '"$"#,##0'; total(f)
    f.font = Font(bold=True, size=12, color="C00000")
    r += 1
    ws.cell(r, 1, "YEAR-2 RECURRING TO COMPLIVERSEAI").font = Font(bold=True, color="1F4E78")
    total(ws.cell(r, 1))
    f2 = ws.cell(r, 2, f"=(B{ann_row}+B{sup_row})*(1-B{pdisc_row})"); f2.number_format = '"$"#,##0'; total(f2)
    f2.font = Font(bold=True, size=12, color="C00000")

    autosize(ws, [50, 22, 4, 4, 4])
    ws.freeze_panes = "A4"


# ---- Build workbooks --------------------------------------------------------
def build_user_based():
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb)
    build_cost_buildup_userbased(wb)
    build_pricing_tiers_userbased(wb)
    build_scenarios_userbased(wb)
    build_competitor(wb)
    build_partner(wb, scenario_acv_col="H")
    build_summary(wb, scenario_acv_col="H", model_label="User-Based Pricing")
    out = OUT_DIR / "CompliverseAI_Pricing_UserBased.xlsx"
    wb.save(out)
    return out


def build_module_plus_user():
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb)
    build_cost_buildup_module(wb)
    build_pricing_tiers_module(wb)
    build_scenarios_module(wb)
    build_competitor_module(wb)
    build_partner(wb, scenario_acv_col="I")
    build_summary(wb, scenario_acv_col="I", model_label="Module + User Pricing")
    out = OUT_DIR / "CompliverseAI_Pricing_Module_Plus_User.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    f1 = build_user_based()
    print(f"[ok] {f1}")
    print("[note] Workbook 2 (Module + Base + User) is now built by "
          "scripts/build_pricing_workbook2_ubl.py — it adds the hosting "
          "toggle, margin floor, token-based LLM cost, 4-Year TCO sheet, "
          "and a UBL pre-filled quote copy. Run that script for Workbook 2.")
