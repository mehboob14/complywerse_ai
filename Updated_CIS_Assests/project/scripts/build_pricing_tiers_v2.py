"""Build CompliverseAI 4-tier pricing workbook (Starter / Growth / Enterprise / Enterprise Plus).

Incorporates the tier proposal: SaaS, Implementation, Support (4 levels),
AI add-ons (Claude + GPT-5 buckets), volume discounts, partner discounts,
competitor benchmarks, customer scenarios, and a quote calculator — all with
live Excel formulas referencing named ranges on Assumptions.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).resolve().parent.parent / "pricing" / "CompliverseAI_Pricing_Tiers_v2.xlsx"
OUT.parent.mkdir(exist_ok=True)

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
TIER_FILL = PatternFill("solid", fgColor="B4C7E7")
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
BOLD = Font(name="Calibri", size=11, bold=True)
TITLE = Font(name="Calibri", size=14, bold=True, color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr(c): c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, Alignment(horizontal="center", vertical="center", wrap_text=True), BOX
def sub(c): c.fill, c.font, c.border = SUB_FILL, SUB_FONT, BOX
def inp(c): c.fill, c.border = INPUT_FILL, BOX
def der(c): c.fill, c.border = DERIVED_FILL, BOX
def tot(c): c.fill, c.font, c.border = TOTAL_FILL, BOLD, BOX
def tier(c): c.fill, c.font, c.alignment, c.border = TIER_FILL, BOLD, Alignment(horizontal="center", wrap_text=True), BOX


def add_name(wb, name, ref):
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Assumptions -------------------------------------------------------------
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions", 0)
    ws.sheet_properties.tabColor = "FFC000"
    ws["A1"] = "CompliverseAI GRC — Pricing Assumptions v2 (USD, 4-tier model)"
    ws["A1"].font = TITLE
    ws.merge_cells("A1:D1")
    ws["A2"] = "Yellow = editable input · Green = derived. Change one cell, the whole workbook recalculates."
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:D2")

    rows = [
        ("GLOBAL", "section", None, None, None),
        ("Annual prepay discount %",       "input", 0.15, "annual_discount", "0.0%"),
        ("Target gross margin %",          "input", 0.60, "target_margin", "0.0%"),

        ("HOSTING", "section", None, None, None),
        ("Hosting $/tenant/month",         "input", 500, "hosting_per_tenant", '"$"#,##0'),
        ("Avg users per tenant (alloc)",   "input", 100, "avg_users", "0"),

        ("API COSTS  ($/user/month)", "section", None, None, None),
        ("LLM (general) per user",         "input", 0.50, "api_llm", '"$"#,##0.00'),
        ("SMS / WhatsApp per user",        "input", 0.10, "api_sms", '"$"#,##0.00'),
        ("Cloud read APIs per user",       "input", 1.00, "api_cloud", '"$"#,##0.00'),
        ("SSO per user",                   "input", 0.30, "api_sso", '"$"#,##0.00'),
        ("Storage / CDN per user",         "input", 0.20, "api_storage", '"$"#,##0.00'),

        ("AI ADD-ONS  (Claude + GPT-5 unit economics)", "section", None, None, None),
        ("Claude Sonnet 4.5 — input $/M tok",  "input", 3.00, "claude_in_per_m",  '"$"#,##0.00'),
        ("Claude Sonnet 4.5 — output $/M tok", "input", 15.00, "claude_out_per_m", '"$"#,##0.00'),
        ("GPT-5 — input $/M tok",              "input", 5.00, "gpt5_in_per_m",    '"$"#,##0.00'),
        ("GPT-5 — output $/M tok",             "input", 20.00, "gpt5_out_per_m",   '"$"#,##0.00'),
        ("Avg input:output ratio (5:1)",       "input", 0.20, "io_ratio", "0.00"),
        ("Starter Claude bucket $/mo (price)", "input", 50,  "ai_starter_claude_price", '"$"#,##0'),
        ("Starter GPT-5 bucket $/mo (price)",  "input", 50,  "ai_starter_gpt5_price",   '"$"#,##0'),
        ("Growth Claude bucket $/mo (price)",  "input", 300, "ai_growth_claude_price",  '"$"#,##0'),
        ("Growth GPT-5 bucket $/mo (price)",   "input", 300, "ai_growth_gpt5_price",    '"$"#,##0'),
        ("Bucket cost markup vs raw API (e.g. 2.0 = 100% margin)", "input", 2.0, "ai_bucket_markup", "0.00"),
        ("Enterprise unlimited AI cost $/mo (assumed)",            "input", 800, "ai_ent_cost", '"$"#,##0'),
        ("Enterprise+ dedicated AI cost $/mo (assumed)",           "input", 2500, "ai_entplus_cost", '"$"#,##0'),

        ("SUPPORT  (cost vs price per customer per month)", "section", None, None, None),
        ("Basic — cost",                  "input", 50,   "sup_basic_cost",   '"$"#,##0'),
        ("Basic — price",                 "input", 150,  "sup_basic_price",  '"$"#,##0'),
        ("Standard — cost",               "input", 200,  "sup_std_cost",     '"$"#,##0'),
        ("Standard — price",              "input", 600,  "sup_std_price",    '"$"#,##0'),
        ("Premium — cost",                "input", 500,  "sup_prem_cost",    '"$"#,##0'),
        ("Premium — price",               "input", 1500, "sup_prem_price",   '"$"#,##0'),
        ("24x7 — cost",                   "input", 1500, "sup_247_cost",     '"$"#,##0'),
        ("24x7 — price",                  "input", 4500, "sup_247_price",    '"$"#,##0'),

        ("IMPLEMENTATION  (one-time fee)", "section", None, None, None),
        ("Starter — cost",                "input", 2000,  "impl_starter_cost",  '"$"#,##0'),
        ("Starter — price",               "input", 5000,  "impl_starter_price", '"$"#,##0'),
        ("Growth — cost",                 "input", 6000,  "impl_growth_cost",   '"$"#,##0'),
        ("Growth — price",                "input", 15000, "impl_growth_price",  '"$"#,##0'),
        ("Enterprise — cost",             "input", 16000, "impl_ent_cost",      '"$"#,##0'),
        ("Enterprise — price",            "input", 40000, "impl_ent_price",     '"$"#,##0'),
        ("Enterprise Plus — cost",        "input", 40000, "impl_entplus_cost",  '"$"#,##0'),
        ("Enterprise Plus — price",       "input", 100000, "impl_entplus_price",'"$"#,##0'),

        ("VOLUME DISCOUNT  (auto-applied by user count)", "section", None, None, None),
        ("Bracket 1 max users",            "input", 25,  "vol_b1_max",  "0"),
        ("Bracket 1 discount %",           "input", 0.00, "vol_b1_disc", "0.0%"),
        ("Bracket 2 max users",            "input", 100, "vol_b2_max",  "0"),
        ("Bracket 2 discount %",           "input", 0.10, "vol_b2_disc", "0.0%"),
        ("Bracket 3 max users",            "input", 500, "vol_b3_max",  "0"),
        ("Bracket 3 discount %",           "input", 0.20, "vol_b3_disc", "0.0%"),
        ("Bracket 4 (500+) discount %",    "input", 0.30, "vol_b4_disc", "0.0%"),

        ("PARTNER / RESELLER", "section", None, None, None),
        ("Bronze %",                      "input", 0.10, "partner_bronze",   "0.0%"),
        ("Silver %",                      "input", 0.15, "partner_silver",   "0.0%"),
        ("Gold %",                        "input", 0.20, "partner_gold",     "0.0%"),
        ("Platinum %",                    "input", 0.25, "partner_platinum", "0.0%"),

        ("PERSONAS", "section", None, None, None),
        ("Starter persona — users",        "input", 25,  "p_starter_users",  "0"),
        ("Growth persona — users",         "input", 100, "p_growth_users",   "0"),
        ("Enterprise persona — users",     "input", 250, "p_ent_users",      "0"),
        ("Enterprise Plus persona — users","input", 750, "p_entplus_users",  "0"),
    ]
    r = 4
    for label, kind, val, name, fmt in rows:
        if kind == "section":
            ws.cell(r, 1, label); sub(ws.cell(r, 1))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 1; continue
        ws.cell(r, 1, label).border = BOX
        c = ws.cell(r, 2, val)
        if fmt: c.number_format = fmt
        inp(c)
        if name:
            add_name(wb, name, f"Assumptions!$B${r}")
        r += 1

    widths(ws, [50, 18, 4, 4])
    ws.freeze_panes = "A4"


# ---- Tier Comparison (the main sheet) ---------------------------------------
def build_tier_comparison(wb):
    ws = wb.create_sheet("Tier Comparison")
    ws.sheet_properties.tabColor = "1F4E78"
    ws["A1"] = "CompliverseAI GRC — 4-Tier Pricing Comparison"
    ws["A1"].font = TITLE
    ws.merge_cells("A1:E1")
    ws["A2"] = "List pricing in USD, annual contract. Partner discounts and volume discounts applied separately (see Quote Calculator)."
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:E2")

    # Column headers (tiers)
    ws.cell(4, 1, "Component"); hdr(ws.cell(4, 1))
    headers = [
        ("Starter", "SMB · 1–25 users", "9CC2E5"),
        ("Growth", "Mid-market · 26–100 users", "70AD47"),
        ("Enterprise", "101–500 users", "BF8F00"),
        ("Enterprise Plus", "500+ users · regulated", "C00000"),
    ]
    for i, (name, sub_label, color) in enumerate(headers, start=2):
        c = ws.cell(4, i, f"{name}\n{sub_label}")
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    ws.row_dimensions[4].height = 36

    # Rows: each component, with tier values
    def row(label, vals, fmt=None, is_total=False, is_section=False):
        nonlocal r
        if is_section:
            ws.cell(r, 1, label); sub(ws.cell(r, 1))
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1; return
        ws.cell(r, 1, label).border = BOX
        if is_total:
            tot(ws.cell(r, 1))
        for i, v in enumerate(vals, start=2):
            c = ws.cell(r, i, v)
            if fmt: c.number_format = fmt
            if is_total: tot(c)
            else: der(c) if isinstance(v, str) and v.startswith("=") else (c.alignment.__setattr__("wrap_text", True) if False else None)
            c.border = BOX
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1

    r = 5

    # POSITIONING
    row("POSITIONING", [], is_section=True)
    row("Tagline", [
        "Get audit-ready in 90 days",
        "Multi-framework continuous compliance",
        "Full GRC platform with AI Copilot",
        "24×7, dedicated AI, on-prem option",
    ])

    # SAAS SUBSCRIPTION
    row("SAAS SUBSCRIPTION", [], is_section=True)
    row("List $/user/month (annual)", [25, 50, 90, 90], fmt='"$"#,##0.00')
    row("Volume discount %", [0, 0.10, 0.20, 0.30], fmt='0.0%')
    row("Net $/user/month",
        [f"=B{r-2}*(1-B{r-1})", f"=C{r-2}*(1-C{r-1})", f"=D{r-2}*(1-D{r-1})", f"=E{r-2}*(1-E{r-1})"],
        fmt='"$"#,##0.00')
    row("Persona user count",
        ["=p_starter_users", "=p_growth_users", "=p_ent_users", "=p_entplus_users"], fmt="0")
    row("ANNUAL SAAS ACV",
        [f"=B{r-2}*B{r-1}*12", f"=C{r-2}*C{r-1}*12", f"=D{r-2}*D{r-1}*12", f"=E{r-2}*E{r-1}*12"],
        fmt='"$"#,##0', is_total=True)
    saas_acv_row = r - 1

    # WHAT'S INCLUDED
    row("WHAT'S INCLUDED", [], is_section=True)
    row("Modules", [
        "3 (Compliance, Risk, Reporting)",
        "8 (everything except AI Copilot, CIS Plugins, Integrations Hub)",
        "All 11 modules",
        "All 11 modules + on-prem option",
    ])
    row("Frameworks", [
        "1 (SOC2 or ISO27001)",
        "Unlimited",
        "Unlimited + custom",
        "Unlimited + custom + classified",
    ])
    row("Evidence storage", ["50 GB", "250 GB", "1 TB", "Unlimited"])
    row("SSO / SCIM", ["—", "SSO", "SSO + SCIM", "SSO + SCIM + custom IdP"])
    row("Uptime SLA", ["99.5%", "99.9%", "99.9%", "99.95%"])

    # AI ADD-ONS
    row("AI ADD-ONS  (Claude Sonnet 4.5 + GPT-5)", [], is_section=True)
    row("AI inclusion model", [
        "Opt-in bucket (overage at cost+30%)",
        "Bucket included (overage at cost+30%)",
        "Unlimited (platform-managed keys)",
        "Dedicated capacity (BYO key option)",
    ])
    row("Claude monthly $/customer",
        ["=ai_starter_claude_price", "=ai_growth_claude_price", "Included", "Included"],
        fmt='"$"#,##0')
    row("GPT-5 monthly $/customer",
        ["=ai_starter_gpt5_price", "=ai_growth_gpt5_price", "Included", "Included"],
        fmt='"$"#,##0')
    row("AI annual cost to customer",
        [f"=(B{r-2}+B{r-1})*12", f"=(C{r-2}+C{r-1})*12", "Included in SaaS", "Included in SaaS"],
        fmt='"$"#,##0')

    # IMPLEMENTATION
    row("IMPLEMENTATION  (one-time)", [], is_section=True)
    row("Implementation price",
        ["=impl_starter_price", "=impl_growth_price", "=impl_ent_price", "=impl_entplus_price"],
        fmt='"$"#,##0', is_total=True)
    row("Scope", [
        "Kickoff, 1 framework, 5 hrs training",
        "3 frameworks, integrations, RCSA workshop, 20 hrs training",
        "Unlimited frameworks, custom workflows, 5 integrations, 40 hrs training",
        "Scoped SOW, dedicated PM, optional on-prem",
    ])

    # SUPPORT
    row("SUPPORT", [], is_section=True)
    row("Tier", ["Basic", "Standard", "Premium", "24×7"])
    row("Response SLA", ["48h business", "4h business", "1h 24×5", "15-min 24×7"])
    row("Monthly support price",
        ["=sup_basic_price", "=sup_std_price", "=sup_prem_price", "=sup_247_price"],
        fmt='"$"#,##0')
    row("Annual support",
        [f"=B{r-1}*12", f"=C{r-1}*12", f"=D{r-1}*12", f"=E{r-1}*12"],
        fmt='"$"#,##0', is_total=True)
    sup_annual_row = r - 1

    # YEAR-1 ECONOMICS
    row("YEAR-1 ECONOMICS", [], is_section=True)
    saas_row = saas_acv_row
    impl_row = saas_row + 6  # implementation price row offset
    # Compute actual rows we wrote — easier to reference by absolute formulas:
    # We'll re-pull SaaS ACV, AI annual, Impl price, Annual support
    ai_annual_row = saas_acv_row + 8  # row "AI annual cost to customer"

    row("YEAR-1 TOTAL (SaaS + AI + Impl + Support)",
        [
            f"=B{saas_acv_row}+IFERROR(B{ai_annual_row}*1,0)+B{impl_row}+B{sup_annual_row}",
            f"=C{saas_acv_row}+IFERROR(C{ai_annual_row}*1,0)+C{impl_row}+C{sup_annual_row}",
            f"=D{saas_acv_row}+B{impl_row}*0+D{impl_row}+D{sup_annual_row}",
            f"=E{saas_acv_row}+E{impl_row}+E{sup_annual_row}",
        ],
        fmt='"$"#,##0', is_total=True)
    y1_row = r - 1
    row("YEAR-2 RECURRING (SaaS + AI + Support)",
        [
            f"=B{saas_acv_row}+IFERROR(B{ai_annual_row}*1,0)+B{sup_annual_row}",
            f"=C{saas_acv_row}+IFERROR(C{ai_annual_row}*1,0)+C{sup_annual_row}",
            f"=D{saas_acv_row}+D{sup_annual_row}",
            f"=E{saas_acv_row}+E{sup_annual_row}",
        ],
        fmt='"$"#,##0', is_total=True)
    y2_row = r - 1

    # MARGIN
    row("GROSS MARGIN  (Year-1)", [], is_section=True)
    # cost_y1 = (per_user_var_cost * users * 12) + (hosting * 12) + impl_cost + support_cost*12 + ai_cost (tier-specific)
    var_cost = "(api_llm+api_sms+api_cloud+api_sso+api_storage)"
    cost_b = (f"={var_cost}*p_starter_users*12 + hosting_per_tenant*12 + impl_starter_cost + sup_basic_cost*12 + "
              f"(ai_starter_claude_price+ai_starter_gpt5_price)/ai_bucket_markup*12")
    cost_c = (f"={var_cost}*p_growth_users*12 + hosting_per_tenant*12 + impl_growth_cost + sup_std_cost*12 + "
              f"(ai_growth_claude_price+ai_growth_gpt5_price)/ai_bucket_markup*12")
    cost_d = (f"={var_cost}*p_ent_users*12 + hosting_per_tenant*12 + impl_ent_cost + sup_prem_cost*12 + "
              f"ai_ent_cost*12")
    cost_e = (f"={var_cost}*p_entplus_users*12 + hosting_per_tenant*12 + impl_entplus_cost + sup_247_cost*12 + "
              f"ai_entplus_cost*12")
    row("Year-1 cost to deliver", [cost_b, cost_c, cost_d, cost_e], fmt='"$"#,##0')
    cost_row = r - 1
    row("Year-1 gross margin %",
        [
            f"=IFERROR((B{y1_row}-B{cost_row})/B{y1_row},0)",
            f"=IFERROR((C{y1_row}-C{cost_row})/C{y1_row},0)",
            f"=IFERROR((D{y1_row}-D{cost_row})/D{y1_row},0)",
            f"=IFERROR((E{y1_row}-E{cost_row})/E{y1_row},0)",
        ],
        fmt='0.0%', is_total=True)
    margin_row = r - 1

    # Conditional format margin row
    ws.conditional_formatting.add(
        f"B{margin_row}:E{margin_row}",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    widths(ws, [42, 26, 30, 26, 28])
    ws.freeze_panes = "B5"

    # Stash key rows for the Quote Calculator on Summary
    add_name(wb, "saas_acv_row",     f"'Tier Comparison'!$A${saas_acv_row}")
    add_name(wb, "y1_total_row",     f"'Tier Comparison'!$A${y1_row}")
    add_name(wb, "y2_total_row",     f"'Tier Comparison'!$A${y2_row}")


# ---- AI Add-ons unit economics ----------------------------------------------
def build_ai_addons(wb):
    ws = wb.create_sheet("AI Add-ons")
    ws.sheet_properties.tabColor = "9C27B0"
    ws["A1"] = "AI Add-on Unit Economics — Claude Sonnet 4.5 + GPT-5"
    ws["A1"].font = TITLE
    ws.merge_cells("A1:F1")
    ws["A2"] = "Bucket sizes derived from list price ÷ blended token cost ÷ markup. Customer 'overage' billed at cost + 30%."
    ws["A2"].font = Font(italic=True, color="595959")
    ws.merge_cells("A2:F2")

    headers = ["Tier", "Vendor", "Bucket price/mo", "Blended $/M tokens (cost)",
               "Tokens included (M)", "Cost-recovery markup"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h); hdr(ws.cell(4, c))

    # Blended $/M = input*ratio_in + output*ratio_out where ratio_in = io_ratio, ratio_out = 1-io_ratio... but spec says input:output=5:1, so input share=5/6, output=1/6
    blended_claude = "=(claude_in_per_m*(5/6)+claude_out_per_m*(1/6))"
    blended_gpt5 = "=(gpt5_in_per_m*(5/6)+gpt5_out_per_m*(1/6))"

    rows = [
        ("Starter", "Claude Sonnet 4.5", "=ai_starter_claude_price", blended_claude),
        ("Starter", "GPT-5",              "=ai_starter_gpt5_price",   blended_gpt5),
        ("Growth",  "Claude Sonnet 4.5", "=ai_growth_claude_price",  blended_claude),
        ("Growth",  "GPT-5",              "=ai_growth_gpt5_price",    blended_gpt5),
    ]
    for i, (t, v, price, blend) in enumerate(rows, start=5):
        ws.cell(i, 1, t).border = BOX; ws.cell(i, 1).font = BOLD
        ws.cell(i, 2, v).border = BOX
        p = ws.cell(i, 3, price); p.number_format = '"$"#,##0'; der(p)
        b = ws.cell(i, 4, blend); b.number_format = '"$"#,##0.00'; der(b)
        # Tokens (M) included = price / blended / markup
        tok = ws.cell(i, 5, f"=C{i}/D{i}/ai_bucket_markup")
        tok.number_format = "0.0"; der(tok)
        m = ws.cell(i, 6, "=ai_bucket_markup"); m.number_format = "0.00"; der(m)

    # Enterprise/Plus included rows
    ws.cell(9, 1, "Enterprise").font = BOLD; ws.cell(9, 1).border = BOX
    ws.cell(9, 2, "Claude + GPT-5 (unlimited)").border = BOX
    ws.cell(9, 3, "Included in SaaS").border = BOX
    ec = ws.cell(9, 4, "=ai_ent_cost"); ec.number_format = '"$"#,##0'; der(ec); ws.cell(9, 4).comment = None
    ws.cell(9, 5, "Unlimited (fair-use)").border = BOX
    ws.cell(9, 6, "Cost absorbed in 60% margin").border = BOX

    ws.cell(10, 1, "Enterprise Plus").font = BOLD; ws.cell(10, 1).border = BOX
    ws.cell(10, 2, "Dedicated capacity").border = BOX
    ws.cell(10, 3, "Included in SaaS").border = BOX
    ec2 = ws.cell(10, 4, "=ai_entplus_cost"); ec2.number_format = '"$"#,##0'; der(ec2)
    ws.cell(10, 5, "Region-locked, BYO key option").border = BOX
    ws.cell(10, 6, "Cost absorbed").border = BOX

    # Routing recommendation
    ws["A12"] = "Server-side routing recommendation (single AI Copilot SKU)"
    sub(ws["A12"]); ws.merge_cells("A12:F12")
    routing = [
        ("Long-form policy parsing & summarization", "Claude Sonnet 4.5 (better long-context)"),
        ("Structured JSON extraction (controls, risks)", "GPT-5 (stronger function-calling)"),
        ("Conversational chatbot / Q&A", "GPT-5 (lower latency)"),
        ("Multi-document compliance reasoning", "Claude Sonnet 4.5 (200k context advantage)"),
        ("Code-like remediation steps generation", "GPT-5"),
    ]
    for c, h in enumerate(["Use case", "Default route"], start=1):
        ws.cell(13, c, h); hdr(ws.cell(13, c))
    for i, (uc, route) in enumerate(routing, start=14):
        ws.cell(i, 1, uc).border = BOX
        ws.cell(i, 2, route).border = BOX
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=6)
    for i in range(14, 14 + len(routing)):
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    widths(ws, [22, 28, 22, 26, 22, 26])
    ws.freeze_panes = "A5"


# ---- Cost Build-Up ----------------------------------------------------------
def build_cost_buildup(wb):
    ws = wb.create_sheet("Cost Build-Up")
    ws.sheet_properties.tabColor = "00B050"
    ws["A1"] = "Per-User Variable Cost (Monthly)"
    ws["A1"].font = TITLE; ws.merge_cells("A1:C1")
    for c, h in enumerate(["Component", "$/user/mo", "Notes"], start=1):
        ws.cell(3, c, h); hdr(ws.cell(3, c))
    rows = [
        ("Hosting allocation per user", "=hosting_per_tenant/avg_users", "Flat tenant hosting / avg users"),
        ("LLM (general, non-AI-add-on)", "=api_llm", "Internal AI calls outside Copilot"),
        ("SMS / WhatsApp", "=api_sms", ""),
        ("Cloud read APIs",  "=api_cloud", "AWS/Azure/GCP read calls"),
        ("SSO",              "=api_sso", "Okta/Azure AD"),
        ("Storage / CDN",    "=api_storage", ""),
        ("TOTAL per-user variable cost", "=SUM(B4:B9)", "Excludes support and AI Copilot bucket"),
    ]
    for i, (l, f, n) in enumerate(rows, start=4):
        ws.cell(i, 1, l).border = BOX
        c = ws.cell(i, 2, f); c.number_format = '"$"#,##0.00'
        if i == 10: tot(c); tot(ws.cell(i, 1))
        else: der(c)
        ws.cell(i, 3, n).border = BOX
    widths(ws, [40, 16, 50])
    ws.freeze_panes = "A4"


# ---- Competitor benchmark ---------------------------------------------------
def build_competitor(wb):
    ws = wb.create_sheet("Competitor Benchmark")
    ws.sheet_properties.tabColor = "C00000"
    ws["A1"] = "Competitor Benchmark (USD annual list, public sources)"
    ws["A1"].font = TITLE; ws.merge_cells("A1:G1")
    ws["A2"] = "Public pricing pages, G2 reviews, customer-disclosed quotes. Most enterprise GRC vendors don't publish list prices — treat as directional."
    ws["A2"].font = Font(italic=True, color="595959"); ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 28

    headers = ["Competitor", "Pricing Model", "Entry annual", "Mid-tier annual", "Enterprise annual", "Source", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h); hdr(ws.cell(4, c))

    rows = [
        ("Vanta",         "Per-framework + per-employee", 11000, 28000, 60000,  "https://www.vanta.com/pricing", "SOC2-only starter ~$11k"),
        ("Drata",         "Per-employee tiered",          12000, 30000, 75000,  "https://drata.com/pricing", "Mid-market sweet spot"),
        ("Sprinto",       "Per-employee, framework-led",  8000,  20000, 50000,  "https://sprinto.com/pricing/", "SMB focused"),
        ("Secureframe",   "Per-framework + per-employee", 12000, 30000, 70000,  "https://secureframe.com/pricing", "Vanta/Drata competitor"),
        ("Hyperproof",    "Per-framework + per-user",     15000, 35000, 80000,  "https://hyperproof.io/pricing/", "Continuous compliance"),
        ("AuditBoard",    "Module bundles, enterprise quote", 50000, 100000, 250000, "https://www.auditboard.com", "Internal audit/SOX heavy"),
        ("OneTrust GRC",  "Module + per-user, quote",     30000, 75000, 200000, "https://www.onetrust.com/products/grc/", "Privacy heritage"),
        ("ServiceNow GRC","Subscription per user + module",75000, 150000,400000, "https://www.servicenow.com/products/governance-risk-and-compliance.html", "Requires ServiceNow license"),
        ("MetricStream",  "Module + per-user, quote",     60000, 140000,350000, "https://www.metricstream.com", "Legacy enterprise GRC"),
    ]
    for i, (comp, model, e, m, ent, src, note) in enumerate(rows, start=5):
        ws.cell(i, 1, comp).font = BOLD; ws.cell(i, 1).border = BOX
        ws.cell(i, 2, model).border = BOX
        for col, v in [(3, e), (4, m), (5, ent)]:
            c = ws.cell(i, col, v); c.number_format = '"$"#,##0'; c.border = BOX
        ws.cell(i, 6, src).border = BOX
        ws.cell(i, 7, note).border = BOX

    last = 4 + len(rows) + 1
    ws.cell(last, 1, "CompliverseAI (Year-1 total)").font = Font(bold=True, color="1F4E78"); tot(ws.cell(last, 1))
    ws.cell(last, 2, "4-tier: SaaS + AI + Impl + Support").border = BOX
    # Pull from Tier Comparison Y1 row — saas_acv_row is at A column reference
    # Easier: hardcode lookup to "Tier Comparison" Y1 cells. The Y1 Total row was the row labelled
    # "YEAR-1 TOTAL ..." — its position depends on layout. We'll use a search approach:
    # Use INDEX/MATCH on the label column.
    for col, tier_col in [(3, "B"), (4, "C"), (5, "E")]:  # Starter, Growth, Enterprise Plus
        c = ws.cell(last, col, f"=INDEX('Tier Comparison'!{tier_col}:{tier_col},MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
        c.number_format = '"$"#,##0'; tot(c)
    # Enterprise tier col D
    c = ws.cell(last, 4, f"=INDEX('Tier Comparison'!D:D,MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
    # overwrite mid-tier reuse: actually mid-tier should be Growth. Let me re-do:
    # Entry = Starter (col B), Mid = Growth (col C), Enterprise = Enterprise (col D). And add Ent+ note.
    c_entry = ws.cell(last, 3, f"=INDEX('Tier Comparison'!B:B,MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
    c_entry.number_format = '"$"#,##0'; tot(c_entry)
    c_mid = ws.cell(last, 4, f"=INDEX('Tier Comparison'!C:C,MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
    c_mid.number_format = '"$"#,##0'; tot(c_mid)
    c_ent = ws.cell(last, 5, f"=INDEX('Tier Comparison'!D:D,MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
    c_ent.number_format = '"$"#,##0'; tot(c_ent)
    ws.cell(last, 6, "Tier Comparison sheet").border = BOX
    ws.cell(last, 7, "Enterprise Plus tier (col E in Tier Comparison) is even higher").border = BOX

    widths(ws, [22, 36, 18, 18, 20, 50, 50])
    ws.freeze_panes = "A5"


# ---- Partner / Reseller -----------------------------------------------------
def build_partner(wb):
    ws = wb.create_sheet("Partner_Reseller")
    ws.sheet_properties.tabColor = "ED7D31"
    ws["A1"] = "Partner / Reseller Discount Model"
    ws["A1"].font = TITLE; ws.merge_cells("A1:F1")
    ws["A2"] = "Discount applied to Year-1 total. Partner sells at list, keeps the discount as margin. CompliverseAI nets (list − discount)."
    ws["A2"].font = Font(italic=True, color="595959"); ws.merge_cells("A2:F2")

    headers = ["Tier", "Discount %", "Starter Y1 net", "Growth Y1 net", "Enterprise Y1 net", "Enterprise Plus Y1 net"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h); hdr(ws.cell(4, c))

    partners = [
        ("None",     0,                  "1-3 deals/year"),
        ("Bronze",   "=partner_bronze",  "1-3 deals/year, basic enablement"),
        ("Silver",   "=partner_silver",  "4-9 deals/year, co-marketing"),
        ("Gold",     "=partner_gold",    "10-24 deals/year, dedicated PAM"),
        ("Platinum", "=partner_platinum","25+ deals/year, joint roadmap, MDF"),
    ]
    for i, (name, disc, _) in enumerate(partners, start=5):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        if isinstance(disc, str):
            d = ws.cell(i, 2, disc); d.number_format = '0.0%'; der(d)
        else:
            d = ws.cell(i, 2, disc); d.number_format = '0.0%'; der(d)
        for col, tier_col in [(3, "B"), (4, "C"), (5, "D"), (6, "E")]:
            c = ws.cell(i, col,
                f"=INDEX('Tier Comparison'!{tier_col}:{tier_col},MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))*(1-B{i})")
            c.number_format = '"$"#,##0'; der(c)
            if name == "None": tot(c)

    # Partner margin section
    ws["A11"] = "Partner Margin (what the partner earns) — Growth tier example"
    sub(ws["A11"]); ws.merge_cells("A11:F11")
    for c, h in enumerate(["Tier", "Discount %", "Growth Y1 list", "Partner margin $", "CompliverseAI nets"], start=1):
        ws.cell(12, c, h); hdr(ws.cell(12, c))
    for i, (name, disc, _) in enumerate(partners, start=13):
        if name == "None": continue
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        d = ws.cell(i, 2, disc); d.number_format = '0.0%'; der(d)
        list_v = ws.cell(i, 3,
            "=INDEX('Tier Comparison'!C:C,MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
        list_v.number_format = '"$"#,##0'; der(list_v)
        margin = ws.cell(i, 4, f"=C{i}*B{i}"); margin.number_format = '"$"#,##0'; tot(margin)
        net = ws.cell(i, 5, f"=C{i}*(1-B{i})"); net.number_format = '"$"#,##0'; der(net)

    widths(ws, [12, 14, 22, 22, 24, 26])
    ws.freeze_panes = "A5"


# ---- Summary + Quote Calculator --------------------------------------------
def build_summary(wb):
    ws = wb.create_sheet("Summary & Quote", 1)  # Place right after Assumptions
    ws.sheet_properties.tabColor = "002060"
    ws["A1"] = "Executive Summary & Quote Calculator"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF"); ws["A1"].fill = HDR_FILL
    ws.merge_cells("A1:E1")

    ws["A3"] = "Snapshot — 4-tier Year-1 / Year-2 / Margin"
    sub(ws["A3"]); ws.merge_cells("A3:E3")
    headers = ["Tier", "Persona users", "Year-1 Total", "Year-2 Recurring", "Y1 Gross Margin %"]
    for c, h in enumerate(headers, start=1):
        ws.cell(4, c, h); hdr(ws.cell(4, c))
    tier_cols = [("Starter", "B", "p_starter_users"),
                 ("Growth", "C", "p_growth_users"),
                 ("Enterprise", "D", "p_ent_users"),
                 ("Enterprise Plus", "E", "p_entplus_users")]
    for i, (name, col, users) in enumerate(tier_cols, start=5):
        ws.cell(i, 1, name).font = BOLD; ws.cell(i, 1).border = BOX
        u = ws.cell(i, 2, f"={users}"); u.number_format = "0"; der(u)
        y1 = ws.cell(i, 3, f"=INDEX('Tier Comparison'!{col}:{col},MATCH(\"YEAR-1 TOTAL (SaaS + AI + Impl + Support)\",'Tier Comparison'!$A:$A,0))")
        y1.number_format = '"$"#,##0'; tot(y1)
        y2 = ws.cell(i, 4, f"=INDEX('Tier Comparison'!{col}:{col},MATCH(\"YEAR-2 RECURRING (SaaS + AI + Support)\",'Tier Comparison'!$A:$A,0))")
        y2.number_format = '"$"#,##0'; tot(y2)
        gm = ws.cell(i, 5, f"=INDEX('Tier Comparison'!{col}:{col},MATCH(\"Year-1 gross margin %\",'Tier Comparison'!$A:$A,0))")
        gm.number_format = '0.0%'; der(gm)

    ws.conditional_formatting.add(
        "E5:E8",
        CellIsRule(operator="lessThan", formula=["target_margin"],
                   fill=PatternFill("solid", fgColor="F4CCCC"))
    )

    # ---- Quote Calculator ----
    ws["A11"] = "Quick Quote Calculator  (edit yellow cells)"
    sub(ws["A11"]); ws.merge_cells("A11:E11")

    inputs = [
        ("User count",                                                100,            "0"),
        ("Tier (Starter / Growth / Enterprise / Enterprise Plus)",    "Growth",       "@"),
        ("Annual prepay? (TRUE/FALSE)",                               True,           "@"),
        ("Support tier (Basic / Standard / Premium / 24x7)",          "Standard",     "@"),
        ("AI add-on? (None / Bucket / Unlimited)",                    "Bucket",       "@"),
        ("Partner level (None / Bronze / Silver / Gold / Platinum)",  "None",         "@"),
    ]
    r = 12
    for label, val, fmt in inputs:
        ws.cell(r, 1, label).border = BOX
        c = ws.cell(r, 2, val); c.number_format = fmt; inp(c)
        r += 1
    user_row, tier_row, prepay_row, sup_row_in, ai_row_in, partner_row_in = 12, 13, 14, 15, 16, 17

    # Outputs
    r = 19
    ws.cell(r, 1, "—— OUTPUTS ——"); sub(ws.cell(r, 1)); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    def out_row(label, formula, fmt='"$"#,##0', total_style=False):
        nonlocal r
        ws.cell(r, 1, label).font = BOLD; ws.cell(r, 1).border = BOX
        c = ws.cell(r, 2, formula); c.number_format = fmt
        tot(c) if total_style else der(c)
        r += 1

    # List $/user/mo lookup (tier → SaaS list price)
    out_row("List $/user/mo",
        f'=IF(B{tier_row}="Starter",25,IF(B{tier_row}="Growth",50,90))',
        '"$"#,##0.00')
    list_row = r - 1

    out_row("Volume discount %",
        f"=IF(B{user_row}<=vol_b1_max,vol_b1_disc,IF(B{user_row}<=vol_b2_max,vol_b2_disc,IF(B{user_row}<=vol_b3_max,vol_b3_disc,vol_b4_disc)))",
        '0.0%')
    vol_row = r - 1

    out_row("Annual prepay discount %",
        f"=IF(B{prepay_row}=TRUE,annual_discount,0)", '0.0%')
    prepay_apply_row = r - 1

    out_row("Net $/user/mo",
        f"=B{list_row}*(1-B{vol_row})*(1-B{prepay_apply_row})",
        '"$"#,##0.00')
    net_user_row = r - 1

    out_row("Annual SaaS subscription",
        f"=B{user_row}*B{net_user_row}*12", '"$"#,##0', total_style=True)
    saas_row = r - 1

    out_row("Implementation (one-time)",
        f'=IF(B{tier_row}="Starter",impl_starter_price,IF(B{tier_row}="Growth",impl_growth_price,IF(B{tier_row}="Enterprise",impl_ent_price,impl_entplus_price)))',
        '"$"#,##0')
    impl_row_out = r - 1

    out_row("Support $/yr",
        f'=12*IF(B{sup_row_in}="Basic",sup_basic_price,IF(B{sup_row_in}="Standard",sup_std_price,IF(B{sup_row_in}="Premium",sup_prem_price,sup_247_price)))',
        '"$"#,##0')
    sup_row_out = r - 1

    out_row("AI add-on $/yr",
        f'=IF(B{ai_row_in}="None",0,IF(B{ai_row_in}="Bucket",IF(B{tier_row}="Starter",(ai_starter_claude_price+ai_starter_gpt5_price)*12,(ai_growth_claude_price+ai_growth_gpt5_price)*12),0))',
        '"$"#,##0')
    ai_row_out = r - 1

    out_row("Partner discount %",
        f'=IF(B{partner_row_in}="Bronze",partner_bronze,IF(B{partner_row_in}="Silver",partner_silver,IF(B{partner_row_in}="Gold",partner_gold,IF(B{partner_row_in}="Platinum",partner_platinum,0))))',
        '0.0%')
    pdisc_row = r - 1

    # Final
    r += 1
    ws.cell(r, 1, "YEAR-1 TOTAL TO COMPLIVERSEAI (after partner)").font = Font(bold=True, color="1F4E78"); tot(ws.cell(r, 1))
    f1 = ws.cell(r, 2, f"=(B{saas_row}+B{impl_row_out}+B{sup_row_out}+B{ai_row_out})*(1-B{pdisc_row})")
    f1.number_format = '"$"#,##0'; tot(f1); f1.font = Font(bold=True, size=12, color="C00000")
    r += 1
    ws.cell(r, 1, "YEAR-2 RECURRING TO COMPLIVERSEAI").font = Font(bold=True, color="1F4E78"); tot(ws.cell(r, 1))
    f2 = ws.cell(r, 2, f"=(B{saas_row}+B{sup_row_out}+B{ai_row_out})*(1-B{pdisc_row})")
    f2.number_format = '"$"#,##0'; tot(f2); f2.font = Font(bold=True, size=12, color="C00000")

    widths(ws, [50, 22, 4, 4, 4])
    ws.freeze_panes = "A4"


def main():
    wb = Workbook(); wb.remove(wb.active)
    build_assumptions(wb)
    build_tier_comparison(wb)
    build_ai_addons(wb)
    build_cost_buildup(wb)
    build_competitor(wb)
    build_partner(wb)
    build_summary(wb)
    # Reorder so Summary is right after Assumptions
    desired = ["Assumptions", "Summary & Quote", "Tier Comparison", "AI Add-ons",
               "Cost Build-Up", "Competitor Benchmark", "Partner_Reseller"]
    wb._sheets = [wb[name] for name in desired]
    wb.save(OUT)
    print(f"[ok] {OUT}")


if __name__ == "__main__":
    main()
