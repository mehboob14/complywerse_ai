import os
import uuid
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
import openpyxl
import csv
import io

from ....models import (
    VulnerabilityReport, Vulnerability, VulnerabilitySLAConfig,
    GRCUser, get_db
)
from ....schemas import (
    VulnerabilityReportCreate, VulnerabilityReportUpdate, VulnerabilityReportResponse,
    MessageResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(tags=["Vulnerability Reports"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "uploads", "reports")
os.makedirs(UPLOAD_DIR, exist_ok=True)


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


def get_report_or_404(report_id: int, user_tenants: List[int], db: Session) -> VulnerabilityReport:
    report = db.query(VulnerabilityReport).filter(
        VulnerabilityReport.id == report_id,
        VulnerabilityReport.tenant_id.in_(user_tenants)
    ).first()
    if not report:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vulnerability report not found"
        )
    return report


def get_sla_days(tenant_id: int, severity: str, db: Session) -> int:
    sla = db.query(VulnerabilitySLAConfig).filter(
        VulnerabilitySLAConfig.tenant_id == tenant_id,
        VulnerabilitySLAConfig.severity == severity,
        VulnerabilitySLAConfig.is_active == True
    ).first()
    if sla:
        return sla.remediation_days
    defaults = {"critical": 7, "high": 30, "medium": 90, "low": 180, "info": 365}
    return defaults.get(severity, 90)


def generate_vuln_id(tenant_id: int, db: Session) -> str:
    count = db.query(Vulnerability).filter(Vulnerability.tenant_id == tenant_id).count()
    return f"VULN-{count + 1:05d}"


@router.post("/reports", response_model=VulnerabilityReportResponse, status_code=status.HTTP_201_CREATED)
async def upload_report(
    file: UploadFile = File(...),
    name: str = Query(...),
    report_type: str = Query("vulnerability_scan"),
    description: Optional[str] = Query(None),
    scan_tool: Optional[str] = Query(None),
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="User not associated with any tenant")
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            tenant_id = user_tenants[0]
    
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in [".xlsx", ".xls", ".csv"]:
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx, .xls) and CSV files are supported")
    
    file_id = str(uuid.uuid4())
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_ext}")
    
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    report = VulnerabilityReport(
        tenant_id=tenant_id,
        name=name,
        description=description,
        report_type=report_type,
        file_path=file_path,
        file_name=file.filename,
        file_type=file_ext[1:],
        scan_tool=scan_tool,
        scan_date=datetime.utcnow(),
        status="uploaded",
        uploaded_by=current_user.id,
        uploaded_at=datetime.utcnow()
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    
    vulnerabilities = []
    try:
        if file_ext in [".xlsx", ".xls"]:
            vulnerabilities = parse_excel_report(content, tenant_id, report.id, db)
        elif file_ext == ".csv":
            vulnerabilities = parse_csv_report(content, tenant_id, report.id, db)
        
        for vuln_data in vulnerabilities:
            sla_days = get_sla_days(tenant_id, vuln_data.get("severity", "medium"), db)
            from datetime import timedelta
            due_date = datetime.utcnow() + timedelta(days=sla_days)
            
            vuln = Vulnerability(
                tenant_id=tenant_id,
                report_id=report.id,
                vuln_id=generate_vuln_id(tenant_id, db),
                title=vuln_data.get("title", "Untitled Vulnerability"),
                description=vuln_data.get("description"),
                severity=vuln_data.get("severity", "medium"),
                cvss_score=vuln_data.get("cvss_score"),
                cve_id=vuln_data.get("cve_id"),
                cwe_id=vuln_data.get("cwe_id"),
                affected_component=vuln_data.get("affected_component"),
                affected_host=vuln_data.get("affected_host"),
                affected_port=vuln_data.get("affected_port"),
                affected_url=vuln_data.get("affected_url"),
                evidence=vuln_data.get("evidence"),
                recommendation=vuln_data.get("recommendation"),
                status="open",
                due_date=due_date,
                discovered_at=datetime.utcnow()
            )
            db.add(vuln)
        
        db.commit()
        
        report.total_vulnerabilities = len(vulnerabilities)
        report.critical_count = sum(1 for v in vulnerabilities if v.get("severity") == "critical")
        report.high_count = sum(1 for v in vulnerabilities if v.get("severity") == "high")
        report.medium_count = sum(1 for v in vulnerabilities if v.get("severity") == "medium")
        report.low_count = sum(1 for v in vulnerabilities if v.get("severity") == "low")
        report.info_count = sum(1 for v in vulnerabilities if v.get("severity") == "info")
        report.status = "parsed"
        db.commit()
        
    except Exception as e:
        report.status = "uploaded"
        db.commit()
    
    db.refresh(report)
    
    return VulnerabilityReportResponse(
        id=report.id,
        tenant_id=report.tenant_id,
        name=report.name,
        description=report.description,
        report_type=report.report_type,
        file_path=report.file_path,
        file_name=report.file_name,
        file_type=report.file_type,
        scan_tool=report.scan_tool,
        scan_date=report.scan_date,
        scan_scope=report.scan_scope,
        asset_scope_ids=report.asset_scope_ids or [],
        total_vulnerabilities=report.total_vulnerabilities,
        critical_count=report.critical_count,
        high_count=report.high_count,
        medium_count=report.medium_count,
        low_count=report.low_count,
        info_count=report.info_count,
        status=report.status,
        uploaded_by=report.uploaded_by,
        uploaded_at=report.uploaded_at,
        created_at=report.created_at,
        updated_at=report.updated_at,
        uploader_name=report.uploader.display_name if report.uploader else None
    )


def parse_excel_report(content: bytes, tenant_id: int, report_id: int, db: Session) -> List[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    vulnerabilities = []
    headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
    
    severity_map = {
        "critical": "critical", "high": "high", "medium": "medium", 
        "low": "low", "info": "info", "informational": "info",
        "4": "critical", "3": "high", "2": "medium", "1": "low", "0": "info"
    }
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        row_dict = dict(zip(headers, row))
        
        title = row_dict.get("title") or row_dict.get("vulnerability") or row_dict.get("name") or row_dict.get("finding")
        if not title:
            continue
        
        severity_raw = str(row_dict.get("severity") or row_dict.get("risk") or "medium").lower()
        severity = severity_map.get(severity_raw, "medium")
        
        cvss = row_dict.get("cvss") or row_dict.get("cvss_score") or row_dict.get("cvss score")
        try:
            cvss_score = float(cvss) if cvss else None
        except:
            cvss_score = None
        
        port = row_dict.get("port") or row_dict.get("affected_port")
        try:
            port_int = int(port) if port else None
        except:
            port_int = None
        
        vulnerabilities.append({
            "title": str(title),
            "description": str(row_dict.get("description") or row_dict.get("details") or ""),
            "severity": severity,
            "cvss_score": cvss_score,
            "cve_id": row_dict.get("cve") or row_dict.get("cve_id") or row_dict.get("cve id"),
            "cwe_id": row_dict.get("cwe") or row_dict.get("cwe_id") or row_dict.get("cwe id"),
            "affected_component": row_dict.get("component") or row_dict.get("affected_component") or row_dict.get("application"),
            "affected_host": row_dict.get("host") or row_dict.get("affected_host") or row_dict.get("ip") or row_dict.get("target"),
            "affected_port": port_int,
            "affected_url": row_dict.get("url") or row_dict.get("affected_url"),
            "evidence": row_dict.get("evidence") or row_dict.get("proof"),
            "recommendation": row_dict.get("recommendation") or row_dict.get("remediation") or row_dict.get("solution")
        })
    
    return vulnerabilities


def parse_csv_report(content: bytes, tenant_id: int, report_id: int, db: Session) -> List[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    
    vulnerabilities = []
    severity_map = {
        "critical": "critical", "high": "high", "medium": "medium", 
        "low": "low", "info": "info", "informational": "info"
    }
    
    for row in reader:
        row = {k.lower(): v for k, v in row.items()}
        
        title = row.get("title") or row.get("vulnerability") or row.get("name") or row.get("finding")
        if not title:
            continue
        
        severity_raw = str(row.get("severity") or row.get("risk") or "medium").lower()
        severity = severity_map.get(severity_raw, "medium")
        
        cvss = row.get("cvss") or row.get("cvss_score") or row.get("cvss score")
        try:
            cvss_score = float(cvss) if cvss else None
        except:
            cvss_score = None
        
        port = row.get("port") or row.get("affected_port")
        try:
            port_int = int(port) if port else None
        except:
            port_int = None
        
        vulnerabilities.append({
            "title": str(title),
            "description": row.get("description") or row.get("details") or "",
            "severity": severity,
            "cvss_score": cvss_score,
            "cve_id": row.get("cve") or row.get("cve_id") or row.get("cve id"),
            "cwe_id": row.get("cwe") or row.get("cwe_id") or row.get("cwe id"),
            "affected_component": row.get("component") or row.get("affected_component") or row.get("application"),
            "affected_host": row.get("host") or row.get("affected_host") or row.get("ip") or row.get("target"),
            "affected_port": port_int,
            "affected_url": row.get("url") or row.get("affected_url"),
            "evidence": row.get("evidence") or row.get("proof"),
            "recommendation": row.get("recommendation") or row.get("remediation") or row.get("solution")
        })
    
    return vulnerabilities


@router.get("/reports", response_model=List[VulnerabilityReportResponse])
def list_reports(
    tenant_id: Optional[int] = None,
    report_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    scan_tool: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(VulnerabilityReport).options(
        joinedload(VulnerabilityReport.uploader)
    ).filter(VulnerabilityReport.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(VulnerabilityReport.tenant_id == tenant_id)
    if report_type:
        query = query.filter(VulnerabilityReport.report_type == report_type)
    if status_filter:
        query = query.filter(VulnerabilityReport.status == status_filter)
    if scan_tool:
        query = query.filter(VulnerabilityReport.scan_tool == scan_tool)
    
    reports = query.order_by(VulnerabilityReport.created_at.desc()).offset(skip).limit(limit).all()
    
    return [
        VulnerabilityReportResponse(
            id=r.id,
            tenant_id=r.tenant_id,
            name=r.name,
            description=r.description,
            report_type=r.report_type,
            file_path=r.file_path,
            file_name=r.file_name,
            file_type=r.file_type,
            scan_tool=r.scan_tool,
            scan_date=r.scan_date,
            scan_scope=r.scan_scope,
            asset_scope_ids=r.asset_scope_ids or [],
            total_vulnerabilities=r.total_vulnerabilities,
            critical_count=r.critical_count,
            high_count=r.high_count,
            medium_count=r.medium_count,
            low_count=r.low_count,
            info_count=r.info_count,
            status=r.status,
            uploaded_by=r.uploaded_by,
            uploaded_at=r.uploaded_at,
            created_at=r.created_at,
            updated_at=r.updated_at,
            uploader_name=r.uploader.display_name if r.uploader else None
        )
        for r in reports
    ]


@router.get("/reports/{report_id}", response_model=VulnerabilityReportResponse)
def get_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="User not associated with any tenant")
    
    report = db.query(VulnerabilityReport).options(
        joinedload(VulnerabilityReport.uploader)
    ).filter(
        VulnerabilityReport.id == report_id,
        VulnerabilityReport.tenant_id.in_(user_tenants)
    ).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    return VulnerabilityReportResponse(
        id=report.id,
        tenant_id=report.tenant_id,
        name=report.name,
        description=report.description,
        report_type=report.report_type,
        file_path=report.file_path,
        file_name=report.file_name,
        file_type=report.file_type,
        scan_tool=report.scan_tool,
        scan_date=report.scan_date,
        scan_scope=report.scan_scope,
        asset_scope_ids=report.asset_scope_ids or [],
        total_vulnerabilities=report.total_vulnerabilities,
        critical_count=report.critical_count,
        high_count=report.high_count,
        medium_count=report.medium_count,
        low_count=report.low_count,
        info_count=report.info_count,
        status=report.status,
        uploaded_by=report.uploaded_by,
        uploaded_at=report.uploaded_at,
        created_at=report.created_at,
        updated_at=report.updated_at,
        uploader_name=report.uploader.display_name if report.uploader else None
    )


@router.delete("/reports/{report_id}", response_model=MessageResponse)
def delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    report = get_report_or_404(report_id, user_tenants, db)
    
    if report.file_path and os.path.exists(report.file_path):
        try:
            os.remove(report.file_path)
        except:
            pass
    
    db.delete(report)
    db.commit()
    
    return MessageResponse(message="Report deleted successfully")
