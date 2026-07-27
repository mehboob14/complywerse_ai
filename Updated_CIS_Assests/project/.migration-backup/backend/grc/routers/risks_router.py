from typing import List, Optional
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
import openpyxl

from ..models import (
    Risk, RiskControlLink, RiskAssetLink, RiskEvidenceLink,
    RiskFrameworkControlLink, RiskGovernanceLink,
    NormalizedControl, FrameworkControl, ITAsset, Evidence,
    GovernanceObjective, GRCUser, Tenant, get_db
)
from ..schemas import (
    RiskCreate, RiskUpdate, RiskResponse,
    RiskAssessment, RiskTreatment,
    RiskControlLinkCreate, RiskAssetLinkCreate, RiskEvidenceLinkCreate,
    RiskFrameworkControlLinkCreate, RiskGovernanceLinkCreate,
    RiskDetailResponse, RiskHeatmapData, MessageResponse
)
from .auth_router import require_auth, require_tenant_permission, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/risks", tags=["Risk Management"])


def calculate_risk_score(likelihood: int, impact: int) -> float:
    return likelihood * impact


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


@router.get("", response_model=List[RiskResponse])
def list_risks(
    tenant_id: Optional[int] = None,
    category: Optional[str] = None,
    register_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    min_score: Optional[float] = None,
    max_score: Optional[float] = None,
    skip: int = 0,
    limit: int = 100,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:view")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    if category:
        query = query.filter(Risk.category == category)
    if register_type:
        query = query.filter(Risk.register_type == register_type)
    if status_filter:
        query = query.filter(Risk.status == status_filter)
    if min_score is not None:
        query = query.filter(Risk.inherent_score >= min_score)
    if max_score is not None:
        query = query.filter(Risk.inherent_score <= max_score)
    
    risks = query.order_by(Risk.created_at.desc()).offset(skip).limit(limit).all()
    return risks


@router.post("", response_model=RiskResponse, status_code=status.HTTP_201_CREATED)
def create_risk(
    risk: RiskCreate,
    tenant_id: Optional[int] = Query(None),
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:create")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tenant not found"
        )
    
    resolved_category = risk.category or getattr(risk, 'risk_category', None) or "Operational"
    # Pre-existing bug fix: create_risk used to drop scores/status/likelihood
    # silently (only saved title/description/category/owner_id). That broke
    # Risk Posture aggregation because residual_score was always NULL.
    # Now we persist every field RiskCreate exposes.
    db_risk = Risk(
        tenant_id=tenant_id,
        title=risk.title,
        description=risk.description,
        category=resolved_category,
        risk_category=getattr(risk, 'risk_category', None) or resolved_category,
        risk_sub_category=getattr(risk, 'risk_sub_category', None),
        register_type=getattr(risk, 'register_type', None),
        owner_id=risk.owner_id,
        business_owner_id=getattr(risk, 'business_owner_id', None),
        inherent_likelihood=getattr(risk, 'inherent_likelihood', None),
        inherent_impact=getattr(risk, 'inherent_impact', None),
        inherent_score=getattr(risk, 'inherent_score', None),
        residual_likelihood=getattr(risk, 'residual_likelihood', None),
        residual_impact=getattr(risk, 'residual_impact', None),
        residual_score=getattr(risk, 'residual_score', None),
        risk_appetite=getattr(risk, 'risk_appetite', None),
        status=getattr(risk, 'status', None) or "open",
        treatment_plan=getattr(risk, 'treatment_plan', None),
    )
    db.add(db_risk)
    db.commit()
    db.refresh(db_risk)
    return db_risk


@router.get("/dashboard")
def get_risk_dashboard(
    tenant_id: Optional[int] = None,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:view")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {
            "total_risks": 0,
            "by_category": {},
            "by_status": {},
            "by_score_range": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "avg_inherent_score": 0,
            "avg_residual_score": 0,
            "open_risks": 0,
            "risks_needing_review": 0
        }
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    by_category = {}
    by_status = {}
    by_score_range = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    total_inherent_score = 0
    total_residual_score = 0
    risks_with_score = 0
    
    for risk in risks:
        cat = risk.risk_category or risk.category or "operational"
        by_category[cat] = by_category.get(cat, 0) + 1
        
        status_val = risk.status or "open"
        by_status[status_val] = by_status.get(status_val, 0) + 1
        
        score = risk.residual_score or risk.inherent_score or 0
        if score >= 20:
            by_score_range["critical"] += 1
        elif score >= 12:
            by_score_range["high"] += 1
        elif score >= 6:
            by_score_range["medium"] += 1
        else:
            by_score_range["low"] += 1
        
        if risk.inherent_score:
            total_inherent_score += risk.inherent_score
            risks_with_score += 1
        if risk.residual_score:
            total_residual_score += risk.residual_score
    
    return {
        "total_risks": len(risks),
        "by_category": by_category,
        "by_status": by_status,
        "by_score_range": by_score_range,
        "avg_inherent_score": round(total_inherent_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "avg_residual_score": round(total_residual_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "open_risks": by_status.get("open", 0),
        "risks_needing_review": sum(1 for r in risks if r.review_date and r.review_date < datetime.utcnow())
    }


@router.get("/heatmap")
def get_risk_heatmap(
    risk_type: Optional[str] = None,
    tenant_id: Optional[int] = None,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:view")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    heatmap = {}
    risk_type_prefix = risk_type if risk_type in ["inherent", "residual"] else "inherent"
    
    for risk in risks:
        likelihood = getattr(risk, f"{risk_type_prefix}_likelihood") or 0
        impact = getattr(risk, f"{risk_type_prefix}_impact") or 0
        
        if likelihood > 0 and impact > 0:
            key = f"{likelihood}-{impact}"
            if key not in heatmap:
                heatmap[key] = {"likelihood": likelihood, "impact": impact, "count": 0, "risks": []}
            heatmap[key]["count"] += 1
            heatmap[key]["risks"].append({
                "id": risk.id,
                "title": risk.title,
                "score": getattr(risk, f"{risk_type_prefix}_score")
            })
    
    return list(heatmap.values())


@router.get("/{risk_id}/detail")
def get_risk_detail(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links).joinedload(RiskControlLink.normalized_control),
        joinedload(Risk.asset_links).joinedload(RiskAssetLink.asset),
        joinedload(Risk.evidence_links).joinedload(RiskEvidenceLink.evidence),
        joinedload(Risk.framework_control_links).joinedload(RiskFrameworkControlLink.framework_control),
        joinedload(Risk.governance_links).joinedload(RiskGovernanceLink.governance_objective),
        joinedload(Risk.owner)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    linked_controls = []
    for link in risk.control_links:
        if link.normalized_control:
            linked_controls.append({
                "id": link.id,
                "control_id": link.normalized_control.id,
                "code": link.normalized_control.code,
                "name": link.normalized_control.name
            })
    
    linked_framework_controls = []
    for link in risk.framework_control_links:
        if link.framework_control:
            linked_framework_controls.append({
                "id": link.id,
                "framework_control_id": link.framework_control.id,
                "code": link.framework_control.code,
                "name": link.framework_control.name,
                "mitigation_effectiveness": link.mitigation_effectiveness,
                "notes": link.notes
            })
    
    linked_assets = []
    for link in risk.asset_links:
        if link.asset:
            linked_assets.append({
                "id": link.id,
                "asset_id": link.asset.id,
                "name": link.asset.name,
                "asset_type": link.asset.asset_type
            })
    
    linked_evidence = []
    for link in risk.evidence_links:
        if link.evidence:
            linked_evidence.append({
                "id": link.id,
                "evidence_id": link.evidence.id,
                "name": link.evidence.name,
                "status": link.evidence.status
            })
    
    linked_governance = []
    for link in risk.governance_links:
        if link.governance_objective:
            linked_governance.append({
                "id": link.id,
                "governance_objective_id": link.governance_objective.id,
                "name": link.governance_objective.name,
                "impact_level": link.impact_level
            })
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "owner_id": risk.owner_id,
        "owner_name": risk.owner.display_name if risk.owner else None,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "due_date": risk.due_date.isoformat() if risk.due_date else None,
        "review_date": risk.review_date.isoformat() if risk.review_date else None,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": linked_controls,
        "linked_framework_controls": linked_framework_controls,
        "linked_assets": linked_assets,
        "linked_evidence": linked_evidence,
        "linked_governance": linked_governance
    }


@router.get("/{risk_id}", response_model=dict)
def get_risk(
    risk_id: int,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:view")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links),
        joinedload(Risk.asset_links),
        joinedload(Risk.evidence_links)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "owner_id": risk.owner_id,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": [link.normalized_control_id for link in risk.control_links],
        "linked_assets": [link.asset_id for link in risk.asset_links],
        "linked_evidence": [link.evidence_id for link in risk.evidence_links]
    }


@router.put("/{risk_id}", response_model=RiskResponse)
def update_risk(
    risk_id: int,
    risk_update: RiskUpdate,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:edit")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    update_data = risk_update.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(risk, field, value)
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.delete("/{risk_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_risk(
    risk_id: int,
    _perm: bool = Depends(require_tenant_permission("risks:risk_register:delete")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    db.delete(risk)
    db.commit()
    return None


@router.post("/{risk_id}/assess", response_model=RiskResponse)
def assess_risk(
    risk_id: int,
    assessment: RiskAssessment,
    _perm: bool = Depends(require_tenant_permission("risks:risk_assessment:edit")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.inherent_likelihood = assessment.inherent_likelihood
    risk.inherent_impact = assessment.inherent_impact
    risk.inherent_score = calculate_risk_score(
        assessment.inherent_likelihood,
        assessment.inherent_impact
    )
    
    if assessment.residual_likelihood and assessment.residual_impact:
        risk.residual_likelihood = assessment.residual_likelihood
        risk.residual_impact = assessment.residual_impact
        risk.residual_score = calculate_risk_score(
            assessment.residual_likelihood,
            assessment.residual_impact
        )
    
    if assessment.risk_appetite:
        risk.risk_appetite = assessment.risk_appetite
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/treatment", response_model=RiskResponse)
def add_treatment_plan(
    risk_id: int,
    treatment: RiskTreatment,
    _perm: bool = Depends(require_tenant_permission("risks:risk_treatment:edit")),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.treatment_plan = treatment.treatment_plan
    risk.status = "mitigating"
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_control(
    risk_id: int,
    link: RiskControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    control = db.query(NormalizedControl).filter(
        NormalizedControl.id == link.normalized_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control not found"
        )
    
    existing = db.query(RiskControlLink).filter(
        RiskControlLink.risk_id == risk_id,
        RiskControlLink.normalized_control_id == link.normalized_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskControlLink(
        risk_id=risk_id,
        normalized_control_id=link.normalized_control_id
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Control linked successfully")


@router.post("/{risk_id}/link-framework-control", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_framework_control(
    risk_id: int,
    link: RiskFrameworkControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    framework_control = db.query(FrameworkControl).filter(
        FrameworkControl.id == link.framework_control_id
    ).first()
    if not framework_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control not found"
        )
    
    existing = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.risk_id == risk_id,
        RiskFrameworkControlLink.framework_control_id == link.framework_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskFrameworkControlLink(
        risk_id=risk_id,
        framework_control_id=link.framework_control_id,
        mitigation_effectiveness=link.mitigation_effectiveness,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Framework control linked successfully")


@router.delete("/{risk_id}/link-framework-control/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_framework_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.id == link_id,
        RiskFrameworkControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/link-governance", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_governance(
    risk_id: int,
    link: RiskGovernanceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    governance_objective = db.query(GovernanceObjective).filter(
        GovernanceObjective.id == link.governance_objective_id,
        GovernanceObjective.tenant_id.in_(user_tenants)
    ).first()
    if not governance_objective:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance objective not found"
        )
    
    existing = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.risk_id == risk_id,
        RiskGovernanceLink.governance_objective_id == link.governance_objective_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskGovernanceLink(
        risk_id=risk_id,
        governance_objective_id=link.governance_objective_id,
        impact_level=link.impact_level
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Governance objective linked successfully")


@router.delete("/{risk_id}/link-governance/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_governance(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.id == link_id,
        RiskGovernanceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.delete("/{risk_id}/controls/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskControlLink).filter(
        RiskControlLink.id == link_id,
        RiskControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/assets", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_asset(
    risk_id: int,
    link: RiskAssetLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == link.asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    existing = db.query(RiskAssetLink).filter(
        RiskAssetLink.risk_id == risk_id,
        RiskAssetLink.asset_id == link.asset_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskAssetLink(risk_id=risk_id, asset_id=link.asset_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Asset linked successfully")


@router.delete("/{risk_id}/assets/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_asset(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskAssetLink).filter(
        RiskAssetLink.id == link_id,
        RiskAssetLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/evidence", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_evidence(
    risk_id: int,
    link: RiskEvidenceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    evidence = db.query(Evidence).filter(
        Evidence.id == link.evidence_id,
        Evidence.tenant_id.in_(user_tenants)
    ).first()
    if not evidence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence not found"
        )
    
    existing = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.risk_id == risk_id,
        RiskEvidenceLink.evidence_id == link.evidence_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskEvidenceLink(risk_id=risk_id, evidence_id=link.evidence_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Evidence linked successfully")


@router.delete("/{risk_id}/evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_evidence(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.id == link_id,
        RiskEvidenceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/upload")
async def upload_risk_register(
    file: UploadFile = File(...),
    tenant_id: Optional[int] = Query(None),
    register_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        
        ws = None
        for sheet_name in ['Risk Assessment', 'Risks', 'Risk Register', 'Sheet1']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        
        headers = []
        header_row = 1
        header_keywords = [
            'risk', 'risk name', 'risk title', 'risk description',
            'risk category', 'category', 'likelihood', 'impact',
            'risk score', 'inherent', 'residual', 'threat', 'vulnerability'
        ]
        for row_num in range(1, 10):
            row_values = [cell.value for cell in ws[row_num]]
            row_str = ' '.join([str(v).lower() for v in row_values if v])
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches >= 3:
                headers = row_values
                header_row = row_num
                break
        if not headers:
            for row_num in range(1, 6):
                row_values = [cell.value for cell in ws[row_num]]
                if any(row_values):
                    headers = row_values
                    header_row = row_num
                    break
        
        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                header_map[str(h).lower().strip()] = idx
        
        def get_value(row, *possible_names):
            for name in possible_names:
                if name.lower() in header_map:
                    idx = header_map[name.lower()]
                    if idx < len(row):
                        return row[idx]
            return None
        
        def parse_int(val, default=1):
            if val is None:
                return default
            if isinstance(val, (int, float)):
                return max(1, min(5, int(val)))
            try:
                return max(1, min(5, int(float(str(val).strip()))))
            except:
                return default
        
        def parse_score(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and val.startswith('='):
                return None
            try:
                return float(val)
            except:
                return None
        
        def map_category(threat_or_category):
            if not threat_or_category:
                return 'operational'
            text = str(threat_or_category).lower()
            if any(w in text for w in ['internal control', 'internal controls', 'internal']):
                return 'internal'
            if any(w in text for w in ['strategic', 'business', 'market']):
                return 'strategic'
            if any(w in text for w in ['financial', 'money', 'cost', 'budget']):
                return 'financial'
            if any(w in text for w in ['compliance', 'regulatory', 'legal', 'pci', 'gdpr']):
                return 'compliance'
            if any(w in text for w in ['technology', 'system', 'network', 'cyber', 'malware', 'phishing', 'security']):
                return 'technology'
            if any(w in text for w in ['vendor', 'supplier', 'third', 'partner', 'outsourcing']):
                return 'third_party'
            return 'operational'
        
        def map_status(treatment_option, residual_score):
            if not treatment_option:
                return 'open'
            text = str(treatment_option).lower()
            if 'accept' in text:
                return 'accepted'
            if 'avoid' in text or 'close' in text:
                return 'closed'
            if 'mitigat' in text or 'reduc' in text or 'treat' in text:
                if residual_score and residual_score < 10:
                    return 'mitigated'
                return 'in_treatment'
            if 'transfer' in text:
                return 'in_treatment'
            return 'open'
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(row):
                continue
            
            ref = get_value(row, 'ref', 'ref.', 'id', 'risk id', 'risk_id')
            asset_name = get_value(row, 'asset name', 'asset', 'asset_name')
            threat = get_value(row, 'threat', 'threat description')
            vulnerability = get_value(row, 'vulnerabilities', 'vulnerability', 'vuln')
            risk_title = get_value(
                row,
                'risk title', 'risk name', 'risk',
                'risk description', 'risk statement',
                'event', 'issue', 'scenario', 'description'
            )
            explicit_category = get_value(row, 'risk category', 'category', 'risk_category', 'risk type', 'risk_type', 'type')
            sub_category = get_value(row, 'risk sub category', 'sub category', 'subcategory', 'sub_category', 'sub type', 'sub-type')
            row_register_type = get_value(row, 'register type', 'register_type', 'register', 'register category')
            
            # More lenient skip logic - accept row if it has ANY identifying information
            has_any_info = any([ref, asset_name, threat, vulnerability, risk_title, explicit_category])
            if not has_any_info:
                skipped_count += 1
                continue
            
            # Build title from available fields with more flexibility
            title_parts = []
            
            # Try to use risk title first (most descriptive)
            if risk_title:
                risk_title_clean = str(risk_title).strip().replace('\n', ' ')[:150]
                if risk_title_clean:
                    title_parts.append(risk_title_clean)
            
            # If no risk title, build from other fields
            if not title_parts:
                if asset_name:
                    title_parts.append(str(asset_name).strip())
                if threat:
                    threat_clean = str(threat).strip().replace('\n', ' ')[:100]
                    if threat_clean:
                        title_parts.append(threat_clean)
                if vulnerability and not title_parts:
                    vuln_clean = str(vulnerability).strip().replace('\n', ' ')[:100]
                    if vuln_clean:
                        title_parts.append(vuln_clean)
            
            # Last resort: use ref or explicit_category
            if not title_parts:
                if ref:
                    title_parts.append(f"Risk {ref}")
                elif explicit_category:
                    title_parts.append(f"{explicit_category} Risk")
            
            # If still no title after all attempts, skip
            if not title_parts:
                skipped_count += 1
                continue
            
            title = " - ".join(title_parts)[:200]
            
            description_parts = []
            if threat:
                description_parts.append(f"Threat: {threat}")
            if vulnerability:
                description_parts.append(f"Vulnerability: {vulnerability}")
            gaps = get_value(row, 'gaps', 'gap')
            if gaps:
                description_parts.append(f"Gaps: {gaps}")
            recommendations = get_value(row, 'recommendations', 'recommendation')
            if recommendations:
                description_parts.append(f"Recommendations: {recommendations}")
            
            description = "\n\n".join(description_parts) if description_parts else None
            
            inherent_likelihood = parse_int(get_value(row, 'likelihood', 'inherent likelihood', 'probability'))
            inherent_impact = parse_int(get_value(row, 'impact', 'inherent impact', 'consequence'))
            inherent_score = parse_score(get_value(row, 'risk score', 'inherent score', 'inherent risk'))
            if inherent_score is None:
                inherent_score = inherent_likelihood * inherent_impact
            
            residual_likelihood = parse_int(get_value(row, 'post-treatment likelihood', 'residual likelihood'), default=None)
            residual_impact = parse_int(get_value(row, 'post-treatment impact', 'residual impact'), default=None)
            residual_score = parse_score(get_value(row, 'residual risk', 'residual score', 'post-treatment risk'))
            if residual_score is None and residual_likelihood and residual_impact:
                residual_score = residual_likelihood * residual_impact
            
            mitigating_controls = get_value(row, 'mitigating action controls', 'controls', 'existing controls', 'mitigating controls')
            action_plan = get_value(row, 'action plan', 'treatment plan', 'plan')
            treatment_parts = []
            if mitigating_controls:
                treatment_parts.append(f"Existing Controls: {mitigating_controls}")
            if action_plan:
                treatment_parts.append(f"Action Plan: {action_plan}")
            treatment_plan = "\n\n".join(treatment_parts) if treatment_parts else None
            
            treatment_option = get_value(row, 'risk treatment option', 'treatment option', 'treatment')
            category = map_category(explicit_category or threat or risk_title)
            risk_status = map_status(treatment_option, residual_score)
            register_type_value = row_register_type or register_type
            if not register_type_value and category == 'internal':
                register_type_value = 'internal'
            
            owner_name = get_value(row, 'responsibility', 'owner', 'risk owner')
            
            try:
                db_risk = Risk(
                    tenant_id=tenant_id,
                    title=title,
                    description=description,
                    category=category,
                    risk_category=category,
                    risk_sub_category=sub_category if sub_category else None,
                    register_type=register_type_value if register_type_value else None,
                    inherent_likelihood=inherent_likelihood,
                    inherent_impact=inherent_impact,
                    inherent_score=inherent_score,
                    residual_likelihood=residual_likelihood if residual_likelihood else None,
                    residual_impact=residual_impact if residual_impact else None,
                    residual_score=residual_score,
                    treatment_plan=treatment_plan,
                    status=risk_status,
                    owner_id=current_user.id
                )
                db.add(db_risk)
                created_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Successfully imported {created_count} risks",
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}"
        )
