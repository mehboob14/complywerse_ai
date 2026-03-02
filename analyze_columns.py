import openpyxl
import os

templates = {
    'PCI DSS': 'backend/risks_templates/PCI DSS Risk Template.xlsm',
    'Internal': 'backend/risks_templates/Internal_Risk_Template.xlsx'
}

for name, path in templates.items():
    if os.path.exists(path):
        print(f"\n{'='*70}")
        print(f"TEMPLATE: {name}")
        print('='*70)
        
        wb = openpyxl.load_workbook(path)
        
        # Check Risk Assessment/Risk Register sheet
        target_sheets = ['Risk Assessment', 'Risk Register']
        
        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            
            try:
                ws = wb[sheet_name]
                print(f"\nSheet: {sheet_name}")
                
                # Find and display first 15 rows
                print("\nFirst 15 rows:")
                for row_num in range(1, 16):
                    try:
                        row_values = [cell.value for cell in ws[row_num]]
                        # Only show non-empty rows
                        if any(row_values):
                            # Show first 6 columns
                            cols = [str(v)[:20] if v else "." for v in row_values[:6]]
                            print(f"  Row {row_num:2d}: {' | '.join(cols)}")
                        else:
                            print(f"  Row {row_num:2d}: (empty)")
                    except Exception as e:
                        print(f"  Row {row_num:2d}: Error - {str(e)[:50]}")
                        break
                
                # Find data rows (first non-header row with data)
                print("\nData Pattern:")
                for row_num in range(1, 20):
                    try:
                        row_values = [cell.value for cell in ws[row_num]]
                        row_str = ' '.join([str(v).lower() if v else '' for v in row_values if v])
                        
                        # Look for actual data (not header indicators)
                        has_numbers = any(isinstance(v, (int, float)) and v not in [1, 2, 3, 4, 5] for v in row_values)
                        is_header = any(kw in row_str for kw in ['risk', 'threat', 'likelihood', 'impact', 'title', 'id', 'category'])
                        
                        if row_str and (is_header or has_numbers or (any(row_values) and row_num > 7)):
                            marker = "[HEADER]" if is_header else "[DATA]" if has_numbers else "[?]"
                            print(f"  Row {row_num}: {marker} {row_str[:80]}")
                    except Exception as e:
                        break
            except Exception as e:
                print(f"  Error accessing sheet {sheet_name}: {str(e)[:100]}")
    else:
        print(f"{path} not found")
