from ..config import get_openai_api_key, get_openai_model

import os
import uuid
import io
import json
import re
import logging
import traceback
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_
from sqlalchemy.orm.attributes import flag_modified
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from pydantic import BaseModel
from openai import OpenAI

logger = logging.getLogger(__name__)

from ..models import (
    ComplianceAssessmentDocument, ComplianceAssessmentDocumentItem,
    AssessmentEvidenceApprovalWorkflow, AssessmentEvidenceApprovalTier,
    AssessmentItemEvidence, AssessmentEvidenceApprovalHistory,
    Evidence, GRCUser, Tenant, get_db,
    EvidenceControlMapping, ParsedFrameworkControl, ComplianceSlaPolicy
)
from .auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/compliance/assessments", tags=["Compliance Assessments"])

UPLOAD_DIR = "backend/grc/uploads/compliance_assessments"
os.makedirs(UPLOAD_DIR, exist_ok=True)

AI_INTEGRATIONS_OPENAI_API_KEY = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY")
AI_INTEGRATIONS_OPENAI_BASE_URL = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")

EVIDENCE_UPLOAD_DIR = "backend/grc/uploads/assessment_evidence"
os.makedirs(EVIDENCE_UPLOAD_DIR, exist_ok=True)

CIS_WS2012R2_CONTROLS_JSON = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "seed_data",
        "CIS",
        "CIS_WS2012R2_Controls.json",
    )
)

EVIDENCE_RECOMMENDATION_PROMPT = """Analyze this assessment item/control requirement and recommend what specific evidence would demonstrate compliance or completion.

Assessment: {assessment_name}
Assessment Type: {assessment_type}
Item Reference: {item_number}
Area/Domain: {area_domain}
Control/Requirement: {control_description}
Current Status: {compliance_status}
Gaps Identified: {gaps_identified}

Based on this requirement, provide specific evidence recommendations in JSON format:
{{
    "recommendations": [
        {{
            "evidence_type": "<specific type e.g., Policy Document, Audit Log, Screenshot, Report>",
            "description": "<detailed description of what this evidence should contain>",
            "priority": "<high|medium|low>",
            "example_files": ["<example1.pdf>", "<example2.xlsx>"]
        }}
    ],
    "summary": "<brief summary of why these evidence types are appropriate>"
}}

Provide 2-5 relevant evidence types prioritized by importance."""


class WorkflowCreate(BaseModel):
    name: str
    description: Optional[str] = None
    is_default: bool = False
    tiers: List[dict] = []


class WorkflowUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    is_default: Optional[bool] = None
    is_active: Optional[bool] = None
    tiers: Optional[List[dict]] = None


class ApprovalActionRequest(BaseModel):
    action: str
    comments: Optional[str] = None
    delegated_to: Optional[int] = None


class LinkExistingEvidenceRequest(BaseModel):
    evidence_id: int
    workflow_id: Optional[int] = None


class AssessmentContextRequest(BaseModel):
    name: str
    assessment_type: str
    source: Optional[str] = None
    notes: Optional[str] = None


class XlsxScoreUpdateRequest(BaseModel):
    sheet: str
    row_index: int
    target_score: Optional[float] = None
    policy_score: Optional[float] = None
    practice_score: Optional[float] = None
    policy_maturity: Optional[float] = None
    practice_maturity: Optional[float] = None
    # Per-row metadata persisted into the JSONB blob alongside the maturity
    # scores. Sentinel "" / null means "clear this field" — distinguish from
    # `None` (i.e. field not in payload) so a partial PUT doesn't wipe state.
    remarks: Optional[str] = None
    assigned_to_id: Optional[int] = None
    assigned_to_name: Optional[str] = None
    due_date: Optional[str] = None
    # Free-form fields (mirrors what page.tsx exposes for non-UBL items).
    gaps_identified: Optional[str] = None
    proposed_solution: Optional[str] = None
    # Framework-structural fields stored on each detail row. Editable per
    # the user's "all fields editable except numbering and assessment text"
    # rule — `subcategory` (the actual control text) and its code stay
    # read-only at the UI layer.
    function: Optional[str] = None
    category: Optional[str] = None
    references: Optional[List[str]] = None

COLUMN_MAPPINGS = {
    "item_number": ["sr", "sr.", "sr#", "s#", "s.no", "s.no.", "no", "no.", "item", "item no", "item number", "control id", "control #", "id", "ref", "reference", "#"],
    "area_domain": ["area", "domain", "category", "section", "control area", "control domain", "control category", "subject area", "topic", "area / domain", "area/domain"],
    "control_description": ["control", "control description", "description", "question", "requirement", "control statement", "control text", "control requirement", "security control", "checklist item", "audit question", "control measure", "control measure / activity", "question/parameter", "parameter"],
    "compliance_status": ["status", "compliance status", "compliance", "assessment status", "result", "response", "finding", "compliant", "complied", "compliance status(y/n)"],
    "gaps_identified": ["gap", "gaps", "gaps identified", "gap identified", "finding", "findings", "observations", "observation", "issues", "issue", "remarks", "audit remarks", "information security remarks", "internal audit"],
    "proposed_solution": ["solution", "proposed solution", "remediation", "recommendation", "recommendations", "action", "corrective action", "proposed action", "mitigation", "proposed solution for compliance"],
    "responsible_party": ["responsible", "responsibility", "responsible party", "owner", "responsible person", "assigned to", "assignee", "department", "itg comments"],
    "timeline": ["timeline", "due date", "target date", "deadline", "completion date", "expected date", "target", "date", "timeline for compliance"],
    "priority": ["priority", "severity", "criticality", "risk level", "importance", "risk rating", "risk"],
    "evidence_reference": ["evidence", "evidence reference", "evidence ref", "documentation", "doc reference", "proof", "supporting evidence"],
    "remarks": ["remarks", "comments", "notes", "additional comments", "additional notes", "remark", "itg comments"]
}

COLUMN_KEYWORDS_PRIORITY = [
    ("compliance_status", ["status", "compliance", "compliant"]),
    ("control_description", ["control measure", "control description", "question/parameter", "checklist"]),
    ("area_domain", ["area / domain", "area/domain", "domain", "category", "section"]),
    ("gaps_identified", ["gaps identified", "gap assessment", "finding", "observation"]),
    ("proposed_solution", ["proposed solution", "solution for compliance", "remediation", "recommendation"]),
    ("responsible_party", ["responsible", "owner", "assigned", "assignee", "itg comment"]),
    ("timeline", ["timeline", "due date", "deadline", "target date"]),
    ("item_number", ["sr", "s#", "s.no", "no.", "item"]),
    ("priority", ["priority", "severity", "criticality", "risk level"]),
    ("evidence_reference", ["evidence", "documentation"]),
    ("remarks", ["remark", "comment", "note"]),
]

STATUS_MAPPINGS = {
    "complied": ["complied", "yes", "y", "complete", "completed", "done", "met", "satisfied", "in place", "implemented", "fully complied", "fully implemented", "pass", "passed", "conform", "conforms", "conforming"],
    "partially_complied": ["partial", "partially", "partially complied", "partially compliant", "partially_compliant", "partial compliant", "partial_compliant", "partial compliance", "in progress", "wip", "work in progress", "partially met", "partially implemented", "partially done", "some"],
    "not_complied": ["not complied", "no", "n", "not met", "not satisfied", "not implemented", "non-compliant", "non compliant", "fail", "failed", "missing", "absent", "not in place", "gap", "not done"],
    "in_progress": ["in progress", "ongoing", "pending", "wip", "work in progress", "under development", "being implemented"],
    "na": ["na", "n/a", "not applicable", "not relevant", "n.a.", "n.a", "-", "none"]
}


def _normalize_status_token(value: str) -> str:
    return re.sub(r"[\s\-_]+", " ", str(value).lower()).strip()


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


def normalize_status(value: str) -> str:
    if not value:
        return "in_progress"
    value_lower = str(value).lower().strip()
    value_token = _normalize_status_token(value_lower)
    for status_key, variants in STATUS_MAPPINGS.items():
        if value_lower in variants or value_token in {_normalize_status_token(v) for v in variants}:
            return status_key
    if value_token == "partially compliant":
        return "partially_complied"
    return "in_progress"


def find_column_mapping(header: str) -> Optional[str]:
    if not header:
        return None
    header_lower = header.lower().strip()
    header_clean = ''.join(c for c in header_lower if c.isalnum() or c in ' /')
    
    for field, variants in COLUMN_MAPPINGS.items():
        if header_lower in variants or header_clean in variants:
            return field
    
    for field, keywords in COLUMN_KEYWORDS_PRIORITY:
        for keyword in keywords:
            if keyword in header_lower:
                return field
    
    return None


def parse_excel_file(file_content: bytes, file_name: str) -> tuple[List[dict], dict]:
    items = []
    column_map = {}
    current_area = None
    
    try:
        if file_name.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
            headers = df.columns.tolist()
            for idx, header in enumerate(headers):
                field = find_column_mapping(str(header))
                if field:
                    column_map[field] = header
            
            for _, row in df.iterrows():
                item = extract_row_data(row, column_map, current_area)
                if item.get("control_description"):
                    if item.get("area_domain"):
                        current_area = item["area_domain"]
                    elif current_area:
                        item["area_domain"] = current_area
                    items.append(item)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            
            headers = []
            header_row_idx = 1
            for row_idx in range(1, min(10, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_idx] if cell.value]
                if len(row_values) >= 3:
                    potential_headers = [cell.value for cell in ws[row_idx]]
                    matches = sum(1 for h in potential_headers if h and find_column_mapping(str(h)))
                    if matches >= 2:
                        headers = potential_headers
                        header_row_idx = row_idx
                        break
            
            if not headers:
                headers = [cell.value for cell in ws[1]]
                header_row_idx = 1
            
            for col_idx, header in enumerate(headers):
                if header:
                    field = find_column_mapping(str(header))
                    if field:
                        column_map[field] = col_idx
            
            merged_ranges = list(ws.merged_cells.ranges)
            
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                row_data = {}
                row_values = [cell.value for cell in ws[row_idx]]
                
                for cell in ws[row_idx]:
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            if top_left_cell.value:
                                for field, col_idx in column_map.items():
                                    if col_idx == cell.column - 1:
                                        if field == "area_domain":
                                            current_area = str(top_left_cell.value).strip()
                
                for field, col_idx in column_map.items():
                    if col_idx < len(row_values):
                        value = row_values[col_idx]
                        if value is not None:
                            row_data[field] = str(value).strip() if value else None
                
                if not row_data.get("area_domain") and current_area:
                    row_data["area_domain"] = current_area
                elif row_data.get("area_domain"):
                    current_area = row_data["area_domain"]
                
                if row_data.get("control_description"):
                    if row_data.get("compliance_status"):
                        row_data["compliance_status"] = normalize_status(row_data["compliance_status"])
                    items.append(row_data)
    
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {str(e)}"
        )
    
    return items, column_map


def extract_row_data(row, column_map: dict, current_area: str) -> dict:
    item = {}
    for field, col_name in column_map.items():
        value = row.get(col_name)
        if pd.notna(value):
            item[field] = str(value).strip()
    
    if item.get("compliance_status"):
        item["compliance_status"] = normalize_status(item["compliance_status"])
    
    if not item.get("area_domain") and current_area:
        item["area_domain"] = current_area
    
    return item


# --------------------------------------------------------------------------- #
#  XLSX Maturity Tool — detection & parsing
# --------------------------------------------------------------------------- #

XLSX_MATURITY_SHEET_NAMES = {"Introduction", "CSF Summary", "Maturity Levels", "NIST CSF Details", "References"}


def detect_xlsx_maturity_format(wb) -> bool:
    """Return True when the workbook contains the NIST CSF Maturity Tool sheets."""
    return XLSX_MATURITY_SHEET_NAMES.issubset(set(wb.sheetnames))


def _safe_float(v):
    try:
        return round(float(v), 3) if v is not None else None
    except (ValueError, TypeError):
        return None


def _normalize_sheet_label(value: Optional[str]) -> str:
    if not value:
        return ""
    return ''.join(ch.lower() for ch in str(value) if ch.isalnum())


def _recalculate_csf_summary_overall(data: dict) -> None:
    sheets = data.get("sheets", {})
    summary = sheets.get("csf_summary") or {}
    categories = summary.get("categories") or []

    target_vals = [c.get("target_score") for c in categories if c.get("target_score") is not None]
    policy_vals = [c.get("policy_score") for c in categories if c.get("policy_score") is not None]
    practice_vals = [c.get("practice_score") for c in categories if c.get("practice_score") is not None]

    summary["overall"] = {
        "target_score": round(sum(target_vals) / len(target_vals), 3) if target_vals else None,
        "policy_score": round(sum(policy_vals) / len(policy_vals), 3) if policy_vals else None,
        "practice_score": round(sum(practice_vals) / len(practice_vals), 3) if practice_vals else None,
    }


def _recalculate_csf_summary_from_details(data: dict) -> None:
    sheets = data.get("sheets", {})
    summary = sheets.get("csf_summary") or {}
    details = sheets.get("details") or []
    categories = summary.get("categories") or []

    if not categories or not details:
        _recalculate_csf_summary_overall(data)
        return

    details_by_category = {}
    for detail in details:
        key = _normalize_sheet_label(detail.get("category"))
        if not key:
            continue
        details_by_category.setdefault(key, []).append(detail)

    for category in categories:
        key = _normalize_sheet_label(category.get("category"))
        if not key or key not in details_by_category:
            continue

        category_details = details_by_category[key]
        policy_vals = [d.get("policy_maturity") for d in category_details if d.get("policy_maturity") is not None]
        practice_vals = [d.get("practice_maturity") for d in category_details if d.get("practice_maturity") is not None]

        category["policy_score"] = round(sum(policy_vals) / len(policy_vals), 3) if policy_vals else None
        category["practice_score"] = round(sum(practice_vals) / len(practice_vals), 3) if practice_vals else None

    _recalculate_csf_summary_overall(data)


def parse_xlsx_maturity_tool(file_content: bytes) -> dict:
    """Parse a NIST CSF Maturity Tool workbook into a structured JSON dict."""
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    result: dict = {"format": "xlsx_maturity", "framework_name": "NIST CSF", "sheets": {}}

    # ------------------------------------------------------------------ #
    # Sheet 1 – Introduction
    # ------------------------------------------------------------------ #
    if "Introduction" in wb.sheetnames:
        ws = wb["Introduction"]
        texts = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val and isinstance(val, str) and val.strip():
                    texts.append(val.strip())
        result["sheets"]["introduction"] = {"text": "\n\n".join(texts)}

    # ------------------------------------------------------------------ #
    # Sheet 2 – CSF Summary  (uses workbook headers exactly)
    # ------------------------------------------------------------------ #
    if "CSF Summary" in wb.sheetnames:
        ws = wb["CSF Summary"]
        rows_list = [tuple(c.value for c in row) for row in ws.iter_rows()]

        # Detect year from any cell containing a 4-digit year
        year = 2018
        for row in rows_list[:5]:
            for v in row:
                if isinstance(v, (int, float)) and 2000 <= int(v) <= 2100:
                    year = int(v)
                    break

        # Find header row: contains both 'target' and 'policy' (case-insensitive)
        header_row_idx = None
        for i, row in enumerate(rows_list):
            row_lower = [str(v).lower() if v else "" for v in row]
            if any("target" in v for v in row_lower) and any("policy" in v for v in row_lower):
                header_row_idx = i
                break

        categories = []
        category_header = "NIST 2018 CSF Categories"
        target_header = "Target Score"
        policy_header = "Policy Score"
        practice_header = "Practice Score"
        if header_row_idx is not None:
            raw_header = [str(v).strip() if v else "" for v in rows_list[header_row_idx]]
            header = [h.lower() for h in raw_header]

            # Exact workbook layout for this template: B=Category, C=Target, D=Policy, E=Practice, A=Function markers
            col_func = 0
            col_cat = 1
            col_target = 2
            col_policy = 3
            col_practice = 4

            # Preserve header labels from workbook row (when available)
            if len(raw_header) > 1 and raw_header[1]:
                category_header = raw_header[1]
            if len(raw_header) > 2 and raw_header[2]:
                target_header = raw_header[2]
            if len(raw_header) > 3 and raw_header[3]:
                policy_header = raw_header[3]
            if len(raw_header) > 4 and raw_header[4]:
                practice_header = raw_header[4]

            current_func = ""
            overall_row = None
            for row in rows_list[header_row_idx + 1:]:
                if not row or not any(v for v in row):
                    continue
                f_val = row[col_func] if col_func < len(row) else None
                c_val = row[col_cat] if col_cat < len(row) else None
                t_val = row[col_target] if col_target < len(row) else None
                p_val = row[col_policy] if col_policy < len(row) else None
                pr_val = row[col_practice] if col_practice < len(row) else None

                if f_val and str(f_val).strip():
                    current_func = str(f_val).strip()
                if not c_val:
                    continue
                target = _safe_float(t_val)
                policy = _safe_float(p_val)
                practice = _safe_float(pr_val)
                if target is None and policy is None:
                    continue
                category_name = str(c_val).strip()
                if category_name.lower() == "overall":
                    overall_row = {
                        "target_score": target,
                        "policy_score": policy,
                        "practice_score": practice,
                    }
                    continue

                categories.append({
                    "function": current_func,
                    "category": category_name,
                    "target_score": target,
                    "policy_score": policy,
                    "practice_score": practice,
                })

        # Compute overall averages
        policy_vals = [c["policy_score"] for c in categories if c["policy_score"] is not None]
        practice_vals = [c["practice_score"] for c in categories if c["practice_score"] is not None]
        target_vals = [c["target_score"] for c in categories if c["target_score"] is not None]
        overall = overall_row if 'overall_row' in locals() and overall_row else {
            "target_score": round(sum(target_vals) / len(target_vals), 3) if target_vals else None,
            "policy_score": round(sum(policy_vals) / len(policy_vals), 3) if policy_vals else None,
            "practice_score": round(sum(practice_vals) / len(practice_vals), 3) if practice_vals else None,
        }
        result["sheets"]["csf_summary"] = {
            "year": year,
            "headers": {
                "category": category_header,
                "target_score": target_header,
                "policy_score": policy_header,
                "practice_score": practice_header,
            },
            "overall": overall,
            "categories": categories,
        }

    # ------------------------------------------------------------------ #
    # Sheet 3 – Maturity Levels (text labels like "Level 1 - Initial")
    # ------------------------------------------------------------------ #
    if "Maturity Levels" in wb.sheetnames:
        ws = wb["Maturity Levels"]
        maturity_levels = []
        rows = list(ws.iter_rows(values_only=True))

        level_header = "Maturity Level"
        policy_header = "Expectation of Policy Maturity Level"
        process_header = "Expectation of Process Maturity Level"
        if rows and rows[0]:
            if rows[0][0]:
                level_header = str(rows[0][0]).replace("\xa0", " ").strip()
            if len(rows[0]) > 1 and rows[0][1]:
                policy_header = str(rows[0][1]).replace("\xa0", " ").strip()
            if len(rows[0]) > 2 and rows[0][2]:
                process_header = str(rows[0][2]).replace("\xa0", " ").strip()

        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            lv = row[0]

            lv_text = str(lv).strip().replace("\xa0", " ")
            if not lv_text.lower().startswith("level"):
                continue

            # Parse "Level 1 - Initial" -> level=1, name=Initial
            lv_int = None
            lv_name = ""
            try:
                right = lv_text.split(" ", 1)[1]  # "1 - Initial"
                lv_int = int(right.split("-")[0].strip())
                if "-" in right:
                    lv_name = right.split("-", 1)[1].strip()
            except Exception:
                continue

            if not (1 <= lv_int <= 10):
                continue
            maturity_levels.append({
                "level": lv_int,
                "name": lv_name,
                "policy_expectation": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "process_expectation": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            })
        result["sheets"]["maturity_levels"] = {
            "headers": {
                "level": level_header,
                "policy_expectation": policy_header,
                "process_expectation": process_header,
            },
            "rows": sorted(maturity_levels, key=lambda x: x["level"]),
        }

    # ------------------------------------------------------------------ #
    # Sheet 4 – NIST CSF Details  (Function | Category | Subcategory | References | Policy | Practice)
    # ------------------------------------------------------------------ #
    if "NIST CSF Details" in wb.sheetnames:
        ws = wb["NIST CSF Details"]
        rows_list = [tuple(c.value for c in row) for row in ws.iter_rows()]

        # Detect header row
        header_row_idx = None
        for i, row in enumerate(rows_list[:15]):
            low = [str(v).lower().strip() if v else "" for v in row]
            if any("function" in v for v in low) and any("subcategory" in v or "sub-category" in v for v in low):
                header_row_idx = i
                break

        details = []
        if header_row_idx is not None:
            header = [str(v).lower().strip() if v else "" for v in rows_list[header_row_idx]]
            col = {}
            for j, h in enumerate(header):
                if "function" in h and "col" not in h and "col" not in h:
                    col.setdefault("function", j)
                elif "subcategory" in h or "sub-category" in h:
                    col.setdefault("subcategory", j)
                elif "category" in h:
                    col.setdefault("category", j)
                elif "reference" in h or "informative" in h:
                    col.setdefault("references", j)
                elif "policy" in h and "maturity" in h:
                    col.setdefault("policy_maturity", j)
                elif "practice" in h and "maturity" in h:
                    col.setdefault("practice_maturity", j)

            # Defaults if not found
            col = {
                "function": col.get("function", 0),
                "category": col.get("category", 1),
                "subcategory": col.get("subcategory", 2),
                "references": col.get("references", 3),
                "policy_maturity": col.get("policy_maturity", 4),
                "practice_maturity": col.get("practice_maturity", 5),
            }

            # Group rows by subcategory, accumulating references; maturity from first row
            current_sub = current_func = current_cat = ""
            current_refs: list = []
            current_policy = current_practice = None

            def _flush():
                if current_sub:
                    details.append({
                        "function": current_func,
                        "category": current_cat,
                        "subcategory": current_sub,
                        "references": list(current_refs),
                        "policy_maturity": current_policy,
                        "practice_maturity": current_practice,
                    })

            for row in rows_list[header_row_idx + 1:]:
                if not row or not any(v for v in row):
                    continue
                get = lambda c: row[col[c]] if col[c] < len(row) else None  # noqa: E731

                sub_val = get("subcategory")
                sub_str = str(sub_val).strip() if sub_val else ""

                if sub_str and sub_str != current_sub:
                    _flush()
                    current_refs = []
                    current_policy = current_practice = None
                    current_sub = sub_str
                    f_v = get("function")
                    c_v = get("category")
                    if f_v:
                        current_func = str(f_v).strip()
                    if c_v:
                        current_cat = str(c_v).strip()

                ref_val = get("references")
                if ref_val:
                    ref_str = str(ref_val).strip()
                    if ref_str and ref_str not in current_refs:
                        current_refs.append(ref_str)

                if current_policy is None:
                    current_policy = _safe_float(get("policy_maturity"))
                if current_practice is None:
                    current_practice = _safe_float(get("practice_maturity"))

            _flush()

        result["sheets"]["details"] = details

    # ------------------------------------------------------------------ #
    # Sheet 5 – References
    # ------------------------------------------------------------------ #
    if "References" in wb.sheetnames:
        ws = wb["References"]
        refs = []
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            doc_str = str(row[0]).strip()
            if not doc_str:
                continue
            link_str = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            refs.append({"document": doc_str, "link": link_str})
        result["sheets"]["references"] = refs

    return result


def _normalized_header(value: object) -> str:
    return ''.join(ch.lower() for ch in str(value or "").strip() if ch.isalnum())


def _strip_reference_links(text: object) -> str:
    """Strip embedded reference links / URLs from a requirement string so the
    dashboards show clean control text. The ASVS export appends OWASP
    Proactive-Controls references like " ([C1](https://owasp.org/...))" to each
    requirement; those are noise in the UI."""
    t = str(text or "")
    # Drop a trailing parenthetical made up of markdown links, e.g.
    # " ([C1](url), [C2](url))".
    t = re.sub(r"\s*\(\s*(?:\[[^\]]*\]\([^)]*\)[,;\s]*)+\)", "", t)
    # Drop any remaining markdown links and bare URLs.
    t = re.sub(r"\[[^\]]*\]\([^)]*\)", "", t)
    t = re.sub(r"https?://\S+", "", t)
    # Tidy leftover empty brackets/parens and doubled whitespace.
    t = re.sub(r"\(\s*[,;]?\s*\)", "", t)
    t = re.sub(r"\s{2,}", " ", t).strip()
    return t


# Friendly names for the workbook templates, used in upload/reupload guard
# messages so a mismatch reads "This is a Mobile App Security assessment …".
_FORMAT_LABELS = {
    "asvs_checklist": "OWASP ASVS",
    "mobile_app_security": "Mobile App Security (OWASP MASVS)",
    "owasp_v4_testing_checklist": "OWASP Testing",
    "csir_maturity": "CSIR Maturity",
    "cti_maturity": "CTI Maturity",
    "itsecops_maturity": "IT Security Operations Maturity",
    "incident_maturity": "Incident Management Maturity",
    "kpi_report": "Cyber Security KPI Report",
    "dpia_pia": "DPIA / PIA",
    "nca_vuln_register": "NCA Vulnerability Register",
    "nca_audit_register": "NCA Cybersecurity Audit Plan",
    "nca_risk_register": "NCA Cybersecurity Risk Register",
    "nca_dcc_tool": "NCA DCC-1:2022 Assessment",
    "pdpl_assessment_toolkit": "Saudi PDPL",
    "nca_container": "NCA",
    "digital_ops_maturity": "Digital Operations Maturity",
    "ubl_audit_master_tracking": "Internal Audit",
    "xlsx_maturity": "Maturity Model",
    "standard": "generic / unrecognised",
}


def _format_label(fmt: str | None) -> str:
    return _FORMAT_LABELS.get((fmt or "").strip(), (fmt or "unknown").replace("_", " ").title())


def _status_from_asvs_valid(value: object) -> str:
    if value is None:
        return "in_progress"
    if isinstance(value, bool):
        return "complied" if value else "not_complied"
    if isinstance(value, (int, float)):
        return "complied" if float(value) > 0 else "not_complied"

    text = str(value).strip().lower()
    if not text:
        return "in_progress"
    if text in {"yes", "y", "valid", "true", "pass", "passed", "done", "implemented", "complied"}:
        return "complied"
    if text in {"partial", "partially", "partially complied"}:
        return "partially_complied"
    if text in {"no", "n", "invalid", "false", "fail", "failed", "not valid", "not complied"}:
        return "not_complied"
    if text in {"na", "n/a", "not applicable"}:
        return "na"
    if text in {"not started", "pending", "todo", "to do"}:
        return "in_progress"
    return normalize_status(text)


def detect_asvs_checklist_format(wb) -> bool:
    if "ASVS Results" not in wb.sheetnames:
        return False
    for sheet_name in wb.sheetnames:
        if sheet_name == "ASVS Results":
            continue
        ws = wb[sheet_name]
        headers = [_normalized_header(cell.value) for cell in ws[1]]
        if "verificationrequirement" in headers and "#" in [str(cell.value).strip() for cell in ws[1] if cell.value is not None]:
            return True
    return False


def parse_asvs_checklist_workbook(wb) -> tuple[List[dict], dict]:
    items: List[dict] = []
    parsed_sheets: List[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "ASVS Results":
            continue
        ws = wb[sheet_name]
        raw_headers = [cell.value for cell in ws[1]]
        normalized_headers = [_normalized_header(v) for v in raw_headers]

        try:
            id_idx = next(i for i, val in enumerate(raw_headers) if str(val or "").strip() == "#")
        except StopIteration:
            id_idx = -1

        index_map = {
            "area": normalized_headers.index("area") if "area" in normalized_headers else -1,
            "requirement": normalized_headers.index("verificationrequirement") if "verificationrequirement" in normalized_headers else -1,
            "valid": normalized_headers.index("valid") if "valid" in normalized_headers else -1,
            "source": normalized_headers.index("sourcecodereference") if "sourcecodereference" in normalized_headers else -1,
            "comment": normalized_headers.index("comment") if "comment" in normalized_headers else -1,
            "tool": normalized_headers.index("toolused") if "toolused" in normalized_headers else -1,
            "level": normalized_headers.index("asvslevel") if "asvslevel" in normalized_headers else -1,
            "cwe": normalized_headers.index("cwe") if "cwe" in normalized_headers else -1,
            "nist": normalized_headers.index("nist") if "nist" in normalized_headers else -1,
        }
        if id_idx < 0 or index_map["requirement"] < 0:
            continue

        current_area = sheet_name
        parsed_sheets.append(sheet_name)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(cell not in (None, "") for cell in row):
                continue

            control_id_raw = row[id_idx] if id_idx < len(row) else None
            requirement_raw = row[index_map["requirement"]] if index_map["requirement"] < len(row) else None
            area_raw = row[index_map["area"]] if index_map["area"] >= 0 and index_map["area"] < len(row) else None

            if area_raw:
                current_area = str(area_raw).strip()

            control_id = str(control_id_raw or "").strip()
            requirement = _strip_reference_links(requirement_raw)
            if not control_id and not requirement:
                continue
            if not requirement:
                continue

            level_val = str(row[index_map["level"]] or "").strip() if index_map["level"] >= 0 and index_map["level"] < len(row) else ""
            cwe_val = str(row[index_map["cwe"]] or "").strip() if index_map["cwe"] >= 0 and index_map["cwe"] < len(row) else ""
            nist_val = str(row[index_map["nist"]] or "").strip() if index_map["nist"] >= 0 and index_map["nist"] < len(row) else ""
            source_ref = str(row[index_map["source"]] or "").strip() if index_map["source"] >= 0 and index_map["source"] < len(row) else ""
            comment_val = str(row[index_map["comment"]] or "").strip() if index_map["comment"] >= 0 and index_map["comment"] < len(row) else ""
            tool_val = str(row[index_map["tool"]] or "").strip() if index_map["tool"] >= 0 and index_map["tool"] < len(row) else ""

            remark_parts = []
            if level_val:
                remark_parts.append(f"ASVS Level: {level_val}")
            if cwe_val:
                remark_parts.append(f"CWE: {cwe_val}")
            if nist_val:
                remark_parts.append(f"NIST: {nist_val}")
            if comment_val:
                remark_parts.append(f"Comment: {comment_val}")
            if tool_val:
                remark_parts.append(f"Tool Used: {tool_val}")

            level_num = None
            try:
                level_num = int(float(level_val))
            except Exception:
                level_num = None

            priority = None
            if level_num is not None:
                if level_num >= 3:
                    priority = "high"
                elif level_num == 2:
                    priority = "medium"
                elif level_num == 1:
                    priority = "low"

            valid_raw = row[index_map["valid"]] if index_map["valid"] >= 0 and index_map["valid"] < len(row) else None
            status_val = _status_from_asvs_valid(valid_raw)
            items.append({
                "item_number": control_id or str(len(items) + 1),
                # Group by ASVS category (the sheet) so the dashboard matches the
                # "ASVS Results" summary; keep the finer "Area" as the subdomain.
                "area_domain": sheet_name,
                "subdomain_name": current_area if current_area and current_area != sheet_name else None,
                "control_description": requirement,
                "compliance_status": status_val,
                "evidence_reference": source_ref or None,
                "remarks": " | ".join(remark_parts) if remark_parts else None,
                "priority": priority,
            })

    metadata = {
        "assessment_format": "asvs_checklist",
        "detected_format": "asvs_checklist",
        "sheets_parsed": parsed_sheets,
        "columns_detected": [
            "item_number",
            "area_domain",
            "control_description",
            "compliance_status",
            "evidence_reference",
            "remarks",
            "priority",
        ],
    }
    return items, metadata


# --------------------------------------------------------------------------- #
# OWASP MASVS – Mobile Application Security Checklist
# Two platforms (Android + iOS), each with a "Security Requirements" sheet
# (Level 1 / Level 2 columns) and an "Anti-RE" sheet (Resilience "R" column).
# The natural scope dimension is the platform, so platform is tagged per item
# and the dedicated dashboard toggles Android vs iOS.
# --------------------------------------------------------------------------- #

def _norm_mstg_id(value: object) -> str:
    """MSTG-IDs use non-breaking / figure hyphens (U+2011, U+2010, en/em dash)
    inconsistently; normalise them all to a plain ASCII hyphen."""
    text = str(value or "").strip()
    for ch in ("‐", "‑", "‒", "–", "—", "−"):
        text = text.replace(ch, "-")
    return text


def _masvs_header_row(ws) -> tuple[int, list]:
    """Locate the header row of a MASVS sheet (the row that carries 'MSTG-ID').
    Returns (row_index_0based, normalized_headers) or (-1, [])."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True)):
        norm = [_normalized_header(v) for v in row]
        if "mstgid" in norm and "status" in norm:
            return i, norm
    return -1, []


def detect_mobile_app_security_format(wb) -> bool:
    names = wb.sheetnames
    has_summary = any(n.strip().lower() == "management summary" for n in names)
    has_req_sheet = any(n.strip().lower().startswith("security requirements") for n in names)
    if not (has_summary and has_req_sheet):
        return False
    # Confirm the MASVS column shape on at least one requirements sheet.
    for n in names:
        if n.strip().lower().startswith("security requirements"):
            idx, norm = _masvs_header_row(wb[n])
            if idx >= 0 and ("level1" in norm or "level2" in norm):
                return True
    return False


def parse_mobile_app_security_workbook(wb) -> tuple[List[dict], dict]:
    items: List[dict] = []
    parsed_sheets: List[str] = []
    _LEGEND = {"legend", "symbol", "pass", "fail", "n/a", "na", "definition"}

    for sheet_name in wb.sheetnames:
        low = sheet_name.strip().lower()
        is_security = low.startswith("security requirements")
        is_antire = low.startswith("anti-re")
        if not (is_security or is_antire):
            continue

        platform = "iOS" if "ios" in low else ("Android" if "android" in low else "General")
        ws = wb[sheet_name]
        header_idx, norm = _masvs_header_row(ws)
        if header_idx < 0:
            continue

        def col(name: str) -> int:
            return norm.index(name) if name in norm else -1

        c_id = col("id")
        c_mstg = col("mstgid")
        c_req = col("detailedverificationrequirement")
        if c_req < 0:
            c_req = col("resiliencyagainstreverseengineeringrequirements")
        c_l1 = col("level1")
        c_l2 = col("level2")
        c_r = col("r")
        c_status = col("status")
        c_test = col("testingprocedures")
        if c_test < 0:
            c_test = col("testingprocedure")
        c_comment = col("comment")

        parsed_sheets.append(sheet_name)
        # Anti-RE sheets are all one MASVS category (V8 Resilience).
        current_category = "V8: Resiliency Against Reverse Engineering" if is_antire else None
        current_subgroup = None

        def cell(row, idx):
            return str(row[idx] or "").strip() if 0 <= idx < len(row) else ""

        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            if not row or not any(v not in (None, "") for v in row):
                continue
            id_val = cell(row, c_id)
            mstg_val = _norm_mstg_id(cell(row, c_mstg))
            req_val = cell(row, c_req)

            if id_val.lower() in _LEGEND or req_val.lower() in _LEGEND:
                break  # legend block sits below the data

            # Category header (security sheets): "V1" ... "V7" with no MSTG-ID.
            if id_val and id_val.upper().startswith("V") and not mstg_val:
                current_category = f"{id_val}: {req_val}" if req_val else id_val
                current_subgroup = None
                continue
            # Sub-group header (Anti-RE): no ID, no MSTG-ID, just a heading.
            if not id_val and not mstg_val and req_val:
                current_subgroup = req_val
                continue
            # Requirement row.
            if not mstg_val or not req_val:
                continue

            has_l1 = cell(row, c_l1) not in ("", "0")
            has_l2 = cell(row, c_l2) not in ("", "0")
            has_r = cell(row, c_r) not in ("", "0")
            levels = []
            if has_l1:
                levels.append("L1")
            if has_l2:
                levels.append("L2")
            if has_r:
                levels.append("R")

            status_val = _status_from_asvs_valid(cell(row, c_status))
            test_proc = cell(row, c_test)
            comment_val = cell(row, c_comment)

            remark_parts = [f"Platform: {platform}"]
            if levels:
                remark_parts.append("MASVS: " + ",".join(levels))
            if mstg_val:
                remark_parts.append(f"MSTG: {mstg_val}")
            if test_proc and test_proc != "-":
                remark_parts.append(f"Testing: {test_proc}")
            if comment_val:
                remark_parts.append(f"Ref: {comment_val}")

            priority = "high" if has_l1 else ("medium" if has_l2 else "low")

            items.append({
                "item_number": f"{platform[:3].upper()}-{id_val}" if id_val else f"{platform[:3].upper()}-{len(items) + 1}",
                "area_domain": current_category or "General",
                "subdomain_name": current_subgroup,
                "control_description": req_val,
                "compliance_status": status_val,
                "remarks": " | ".join(remark_parts),
                "priority": priority,
            })

    metadata = {
        "assessment_format": "mobile_app_security",
        "detected_format": "mobile_app_security",
        "sheets_parsed": parsed_sheets,
        "platforms": sorted({("iOS" if "ios" in s.lower() else "Android") for s in parsed_sheets if ("ios" in s.lower() or "android" in s.lower())}),
        "columns_detected": [
            "item_number",
            "area_domain",
            "subdomain_name",
            "control_description",
            "compliance_status",
            "remarks",
            "priority",
        ],
    }
    return items, metadata


def _status_from_owasp_result(value: object) -> str:
    if value is None:
        return "in_progress"
    text = str(value).strip().lower()
    if not text:
        return "in_progress"
    if text in {"pass", "passed", "complete", "completed", "done", "ok", "success"}:
        return "complied"
    if text in {"partial", "partially complete", "partially completed", "partially"}:
        return "partially_complied"
    if text in {"fail", "failed", "non-compliant", "not complied", "not compliant"}:
        return "not_complied"
    if text in {"na", "n/a", "not applicable"}:
        return "na"
    if text in {"not started", "in progress", "pending", "open", "under test", "testing"}:
        return "in_progress"
    return normalize_status(text)


def _status_from_ubl_audit_tracking(value: object) -> str:
    if value is None:
        return "in_progress"
    text = str(value).strip().lower()
    if not text:
        return "in_progress"
    if any(token in text for token in ["closed", "complete", "completed", "done", "resolved"]):
        return "complied"
    if any(token in text for token in ["in-progress", "in progress", "ongoing", "open", "started"]):
        return "in_progress"
    if any(token in text for token in ["to be started", "not started", "pending"]):
        return "in_progress"
    if any(token in text for token in ["rejected", "failed", "not complied", "non-compliant"]):
        return "not_complied"
    if text in {"na", "n/a", "not applicable"}:
        return "na"
    return normalize_status(text)


def _cell_to_text(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    return text or None


def _find_header_row_with_keys(ws, expected_keys: List[str], max_scan_rows: int = 12) -> Optional[int]:
    expected = {_normalized_header(key) for key in expected_keys if key}
    if not expected:
        return None
    max_row = min(ws.max_row, max_scan_rows)
    for row_idx in range(1, max_row + 1):
        row_norm = {_normalized_header(cell.value) for cell in ws[row_idx] if cell.value is not None}
        if len(expected.intersection(row_norm)) >= max(2, min(3, len(expected))):
            return row_idx
    return None


def detect_ubl_audit_master_tracking_format(wb) -> bool:
    normalized_sheets = {_normalize_sheet_label(s) for s in wb.sheetnames}
    signal_sheets = {"auditospoints", "internalaudit", "mlintl", "mldomestic", "sbp"}
    matches = len(signal_sheets.intersection(normalized_sheets))
    return matches >= 3 or (matches >= 2 and "dashboard" in normalized_sheets)


def parse_ubl_audit_master_tracking_workbook(wb) -> tuple[List[dict], dict]:
    items: List[dict] = []
    parsed_sheets: List[str] = []

    def _add_item(
        *,
        sheet: str,
        item_number: Optional[str],
        area_domain: Optional[str],
        control_description: Optional[str],
        compliance_status: Optional[str],
        gaps_identified: Optional[str] = None,
        proposed_solution: Optional[str] = None,
        responsible_party: Optional[str] = None,
        timeline: Optional[str] = None,
        priority: Optional[str] = None,
        evidence_reference: Optional[str] = None,
        remarks: Optional[str] = None,
    ) -> None:
        description = (control_description or "").strip()
        if not description:
            return
        item = {
            "item_number": (item_number or str(len(items) + 1))[:50],
            "area_domain": (area_domain or sheet)[:255],
            "control_description": description[:8000],
            "compliance_status": compliance_status or "in_progress",
            "gaps_identified": gaps_identified[:5000] if gaps_identified else None,
            "proposed_solution": proposed_solution[:5000] if proposed_solution else None,
            "responsible_party": responsible_party[:255] if responsible_party else None,
            "timeline": timeline[:255] if timeline else None,
            "priority": priority[:100].lower() if priority else None,
            "evidence_reference": evidence_reference[:255] if evidence_reference else None,
            "remarks": remarks[:5000] if remarks else None,
        }
        items.append(item)

    for sheet_name in wb.sheetnames:
        normalized_name = _normalize_sheet_label(sheet_name)
        if normalized_name in {"dashboard", "sheet1"}:
            continue
        ws = wb[sheet_name]

        # -------------------------------------------------------------- #
        # Audit OS Points
        # -------------------------------------------------------------- #
        if normalized_name == "auditospoints":
            header_row_idx = _find_header_row_with_keys(
                ws, ["S#", "Description", "Source", "Raised By:", "Responsible", "Target Completion date", "Status"]
            )
            if not header_row_idx:
                continue
            headers = [cell.value for cell in ws[header_row_idx]]
            header_map = {_normalized_header(h): idx for idx, h in enumerate(headers) if h is not None}

            def g(row, key: str) -> Optional[str]:
                idx = header_map.get(_normalized_header(key))
                if idx is None or idx >= len(row):
                    return None
                return _cell_to_text(row[idx])

            parsed_sheets.append(sheet_name)
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                description = g(row, "Description")
                if not description:
                    continue
                status_raw = g(row, "Status")
                _add_item(
                    sheet=sheet_name,
                    item_number=g(row, "S#"),
                    area_domain=f"{sheet_name} - {(g(row, 'Raised By:') or 'Audit')}",
                    control_description=description,
                    compliance_status=_status_from_ubl_audit_tracking(status_raw),
                    proposed_solution=g(row, "Next action Item"),
                    responsible_party=g(row, "Responsible"),
                    timeline=g(row, "Target Completion date"),
                    evidence_reference=g(row, "Source"),
                    remarks=f"Raised By: {g(row, 'Raised By:')}" if g(row, "Raised By:") else None,
                )
            continue

        # -------------------------------------------------------------- #
        # Internal Audit
        # -------------------------------------------------------------- #
        if normalized_name == "internalaudit":
            header_row_idx = _find_header_row_with_keys(
                ws, ["Year", "Branch / Unit", "Brief Detail of Exception", "Management Response", "Revised Target Date", "Status"]
            )
            if not header_row_idx:
                continue
            headers = [cell.value for cell in ws[header_row_idx]]
            header_map = {_normalized_header(h): idx for idx, h in enumerate(headers) if h is not None}

            def g(row, key: str) -> Optional[str]:
                idx = header_map.get(_normalized_header(key))
                if idx is None or idx >= len(row):
                    return None
                return _cell_to_text(row[idx])

            parsed_sheets.append(sheet_name)
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=1):
                exception_text = g(row, "Brief Detail of Exception")
                if not exception_text:
                    continue
                branch = g(row, "Branch / Unit") or "Internal Audit"
                year = g(row, "Year")
                status_raw = g(row, "Status")
                area = branch
                if year:
                    area = f"{branch} ({year})"
                remarks_parts = []
                intl_domestic = g(row, "International / Domestic")
                tagged = g(row, "Tagged")
                if intl_domestic:
                    remarks_parts.append(f"Scope: {intl_domestic}")
                if tagged:
                    remarks_parts.append(f"Tagged: {tagged}")
                _add_item(
                    sheet=sheet_name,
                    item_number=f"IA-{row_idx}",
                    area_domain=area,
                    control_description=exception_text,
                    compliance_status=_status_from_ubl_audit_tracking(status_raw),
                    gaps_identified=exception_text,
                    proposed_solution=g(row, "Management Response"),
                    responsible_party=tagged or branch,
                    timeline=g(row, "Revised Target Date"),
                    priority=g(row, "Risk"),
                    remarks=" | ".join(remarks_parts) if remarks_parts else None,
                )
            continue

        # -------------------------------------------------------------- #
        # ML International / Domestic
        # -------------------------------------------------------------- #
        if normalized_name in {"mlintl", "mldomestic"}:
            header_row_idx = _find_header_row_with_keys(
                ws, ["Point No.", "Title", "Observation", "Recommendation", "Rating", "Status"]
            )
            if not header_row_idx:
                continue
            headers = [cell.value for cell in ws[header_row_idx]]
            header_map = {_normalized_header(h): idx for idx, h in enumerate(headers) if h is not None}

            def g(row, key: str) -> Optional[str]:
                idx = header_map.get(_normalized_header(key))
                if idx is None or idx >= len(row):
                    return None
                return _cell_to_text(row[idx])

            parsed_sheets.append(sheet_name)
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                title = g(row, "Title")
                observation = g(row, "Observation")
                recommendation = g(row, "Recommendation")
                if not title and not observation:
                    continue
                description = title or "Audit Observation"
                if observation:
                    description = f"{description}. {observation}" if title else observation
                owner = g(row, "Owner/Action") or g(row, "Owner")
                mgmt_resp = g(row, "Management Response - Sep'22")
                proposed_solution = recommendation
                if mgmt_resp:
                    proposed_solution = f"{recommendation}\n\nManagement Response: {mgmt_resp}" if recommendation else mgmt_resp
                _add_item(
                    sheet=sheet_name,
                    item_number=g(row, "Point No."),
                    area_domain=sheet_name,
                    control_description=description,
                    compliance_status=_status_from_ubl_audit_tracking(g(row, "Status")),
                    gaps_identified=observation,
                    proposed_solution=proposed_solution,
                    responsible_party=owner,
                    timeline=g(row, "Revised Target Date"),
                    priority=g(row, "Rating"),
                    remarks=f"Year: {g(row, 'Year')}" if g(row, "Year") else None,
                )
            continue

        # -------------------------------------------------------------- #
        # SBP observations
        # -------------------------------------------------------------- #
        if normalized_name == "sbp":
            header_row_idx = _find_header_row_with_keys(
                ws, ["S. No", "CG Ref. No.", "SBP Observation", "Pertain to", "Management Response", "Target Date", "Status 1"]
            )
            if not header_row_idx:
                continue
            headers = [cell.value for cell in ws[header_row_idx]]
            header_map = {_normalized_header(h): idx for idx, h in enumerate(headers) if h is not None}

            def g(row, key: str) -> Optional[str]:
                idx = header_map.get(_normalized_header(key))
                if idx is None or idx >= len(row):
                    return None
                return _cell_to_text(row[idx])

            parsed_sheets.append(sheet_name)
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                observation = g(row, "SBP Observation")
                if not observation:
                    continue
                report_name = g(row, "Report Name") or "SBP"
                pertain_to = g(row, "Pertain to")
                area = report_name
                if pertain_to:
                    area = f"{report_name} - {pertain_to}"
                remarks_parts = []
                cg_ref = g(row, "CG Ref. No.")
                ia_comments = g(row, "IA Comments (14-Sep-2022)")
                initial_timeline = g(row, "Initial Timeline")
                groups_revised = g(row, "Groups Revised")
                if cg_ref:
                    remarks_parts.append(f"CG Ref: {cg_ref}")
                if groups_revised:
                    remarks_parts.append(f"Groups Revised: {groups_revised}")
                if initial_timeline:
                    remarks_parts.append(f"Initial Timeline: {initial_timeline}")
                if ia_comments:
                    remarks_parts.append(f"IA Comments: {ia_comments}")
                _add_item(
                    sheet=sheet_name,
                    item_number=g(row, "S. No") or cg_ref,
                    area_domain=area,
                    control_description=observation,
                    compliance_status=_status_from_ubl_audit_tracking(g(row, "Status 1")),
                    proposed_solution=g(row, "Management Response"),
                    responsible_party=pertain_to or g(row, "Group"),
                    timeline=g(row, "Target Date") or initial_timeline,
                    remarks=" | ".join(remarks_parts) if remarks_parts else None,
                )
            continue

    metadata = {
        "assessment_format": "ubl_audit_master_tracking",
        "detected_format": "ubl_audit_master_tracking",
        "sheets_parsed": parsed_sheets,
        "columns_detected": [
            "item_number",
            "area_domain",
            "control_description",
            "compliance_status",
            "gaps_identified",
            "proposed_solution",
            "responsible_party",
            "timeline",
            "priority",
            "evidence_reference",
            "remarks",
        ],
    }
    return items, metadata


# ─────────────────────────────────────────────────────────────────────────────
# Saudi PDPL Compliance Assessment Toolkit
# ─────────────────────────────────────────────────────────────────────────────

def _status_from_pdpl(maturity: object, status_text: object) -> str:
    """Derive compliance status for a PDPL control.

    The toolkit's Read Me defines the rule: maturity 0-5 drives status, where
    3-5 = Compliant. We map 0 -> not_complied, 1-2 -> partially_complied,
    3-5 -> complied. When maturity is blank we fall back to the sheet's textual
    'Compliance Status' (e.g. 'Not Assessed' -> in_progress).
    """
    m = None
    if maturity is not None and str(maturity).strip() != "":
        try:
            m = float(str(maturity).strip())
        except (TypeError, ValueError):
            m = None
    if m is not None:
        if m >= 3:
            return "complied"
        if m <= 0:
            return "not_complied"
        return "partially_complied"  # 1-2 = initial/developing
    text = str(status_text or "").strip().lower()
    if not text or "not assessed" in text:
        return "in_progress"
    if "partial" in text:
        return "partially_complied"
    if "non" in text or "not complied" in text or "not compliant" in text:
        return "not_complied"
    if "compliant" in text or "complied" in text:
        return "complied"
    if text in {"na", "n/a", "not applicable"}:
        return "na"
    return normalize_status(text)


# The Assessment + Client Profile + Remediation Plan trio is unique to the PDPL
# toolkit; no other supported template carries all three.
# Distinctive sheet signature: an Assessment + Remediation Plan pair. (The
# toolkit ships in variants — some have a separate 'Client Profile' sheet, the
# bundled reference uses a 'Read Me' instead — so we don't require Client
# Profile. The unique 'PDPL Ref.' column checked below is the real signature.)
PDPL_SIGNAL_SHEETS = {"assessment", "remediationplan"}
PDPL_ASSESSMENT_KEYS = ["Control ID", "Domain", "PDPL Ref.", "Maturity (0-5)", "Compliance Status"]


def _pdpl_assessment_sheet_name(wb) -> Optional[str]:
    return next((s for s in wb.sheetnames if _normalize_sheet_label(s) == "assessment"), None)


def detect_pdpl_assessment_format(wb) -> bool:
    """Detect the Saudi PDPL Compliance Assessment Toolkit.

    Two-part signature so a look-alike workbook can't false-match: (1) the
    distinctive Assessment + Client Profile + Remediation Plan sheet trio, and
    (2) the Assessment sheet header row exposing the unique 'PDPL Ref.' column.
    """
    normalized_sheets = {_normalize_sheet_label(s) for s in wb.sheetnames}
    if not PDPL_SIGNAL_SHEETS.issubset(normalized_sheets):
        return False
    sheet_name = _pdpl_assessment_sheet_name(wb)
    if not sheet_name:
        return False
    ws = wb[sheet_name]
    header_row = _find_header_row_with_keys(ws, PDPL_ASSESSMENT_KEYS)
    if not header_row:
        return False
    header_norms = {_normalized_header(c.value) for c in ws[header_row] if c.value is not None}
    return "pdplref" in header_norms


def parse_pdpl_assessment_workbook(wb) -> tuple[List[dict], dict]:
    """Parse the PDPL toolkit's 'Assessment' sheet into assessment items.

    Maps the 14 PDPL columns onto the assessment-item shape. Maturity (0-5)
    drives compliance_status; PDPL Ref. / maturity / risk rating / assessment
    question are preserved in `remarks` since the item table has no dedicated
    columns for them (so nothing from the sheet is lost on import).
    """
    sheet_name = _pdpl_assessment_sheet_name(wb)
    ws = wb[sheet_name]
    header_row = _find_header_row_with_keys(ws, PDPL_ASSESSMENT_KEYS) or 1
    headers = [cell.value for cell in ws[header_row]]
    header_map = {_normalized_header(h): idx for idx, h in enumerate(headers) if h is not None}

    def g(row, key: str) -> Optional[str]:
        idx = header_map.get(_normalized_header(key))
        if idx is None or idx >= len(row):
            return None
        return _cell_to_text(row[idx])

    items: List[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        control_id = g(row, "Control ID")
        requirement = g(row, "Control Requirement")
        question = g(row, "Assessment Question")
        if not control_id and not requirement:
            continue  # skip blank / spacer rows
        maturity = g(row, "Maturity (0-5)")
        maturity_score = None
        if maturity is not None and str(maturity).strip() != "":
            try:
                maturity_score = int(float(str(maturity).strip()))
            except (TypeError, ValueError):
                maturity_score = None
        pdpl_ref = g(row, "PDPL Ref.")
        risk = g(row, "Risk Rating")
        remarks_parts = []
        if pdpl_ref:
            remarks_parts.append(f"PDPL Ref: {pdpl_ref}")
        if question:
            remarks_parts.append(f"Q: {question}")
        priority = (g(row, "Priority") or "").lower().strip()
        items.append({
            "item_number": (control_id or str(len(items) + 1))[:50],
            "area_domain": (g(row, "Domain") or "PDPL")[:500],
            "control_description": (requirement or question or "")[:8000],
            "compliance_status": _status_from_pdpl(maturity, g(row, "Compliance Status")),
            "maturity_score": maturity_score,
            "risk_rating": (risk or None) and risk[:50],
            "gaps_identified": g(row, "Findings / Gap"),
            "proposed_solution": g(row, "Remediation Action"),
            "responsible_party": (g(row, "Owner") or None) and g(row, "Owner")[:255],
            "timeline": (g(row, "Target Date") or None) and g(row, "Target Date")[:255],
            "priority": priority[:50] or None,
            "evidence_reference": g(row, "Evidence to Request"),
            "remarks": " | ".join(remarks_parts) if remarks_parts else None,
        })

    metadata = {
        "assessment_format": "pdpl_assessment_toolkit",
        "detected_format": "pdpl_assessment_toolkit",
        "framework": "Saudi PDPL",
        "columns_detected": [
            "item_number", "area_domain", "control_description", "compliance_status",
            "gaps_identified", "proposed_solution", "responsible_party", "timeline",
            "priority", "evidence_reference", "remarks",
        ],
    }
    return items, metadata


def detect_owasp_v4_checklist_format(wb) -> bool:
    if "Testing Checklist" not in wb.sheetnames:
        return False
    ws = wb["Testing Checklist"]
    for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=True):
        normalized = [_normalized_header(v) for v in row]
        if "testname" in normalized and "result" in normalized and "description" in normalized:
            return True
    return False


def parse_owasp_v4_checklist_workbook(wb) -> tuple[List[dict], dict]:
    ws = wb["Testing Checklist"]
    header_row_idx = None
    header_values = None

    for row_idx in range(1, min(25, ws.max_row + 1)):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [_normalized_header(v) for v in values]
        if "testname" in normalized and "result" in normalized and "description" in normalized:
            header_row_idx = row_idx
            header_values = values
            break

    if header_row_idx is None or header_values is None:
        return [], {
            "assessment_format": "owasp_v4_testing_checklist",
            "detected_format": "owasp_v4_testing_checklist",
            "columns_detected": [],
        }

    normalized_headers = [_normalized_header(v) for v in header_values]
    name_idx = normalized_headers.index("testname") if "testname" in normalized_headers else 1
    desc_idx = normalized_headers.index("description") if "description" in normalized_headers else 2
    tools_idx = normalized_headers.index("tools") if "tools" in normalized_headers else 3
    result_idx = normalized_headers.index("result") if "result" in normalized_headers else 4
    remark_idx = normalized_headers.index("remark") if "remark" in normalized_headers else 5
    code_idx = 0

    items: List[dict] = []
    current_area = "OWASP Testing"
    code_pattern = re.compile(r"^OTG-[A-Z0-9-]+$", re.IGNORECASE)

    # Iterate from the top: each WSTG category is introduced by a row whose first
    # cell is the category name and whose "Test Name" column literally reads
    # "Test Name" (a repeated sub-header). Test rows carry an OTG-… code.
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not any(cell not in (None, "") for cell in row):
            continue

        code_val = str(row[code_idx] or "").strip() if code_idx < len(row) else ""
        test_name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""

        # Category header row (e.g. "Authentication Testing" | "Test Name" | …).
        if code_val and not code_pattern.match(code_val) and _normalized_header(test_name) == "testname":
            current_area = code_val
            continue
        # Only OTG-coded rows are real tests; skip title/blank/summary rows.
        if not code_pattern.match(code_val):
            continue

        description = str(row[desc_idx] or "").strip() if desc_idx < len(row) else ""
        tools = str(row[tools_idx] or "").strip() if tools_idx < len(row) else ""
        result = row[result_idx] if result_idx < len(row) else None
        remark = str(row[remark_idx] or "").strip() if remark_idx < len(row) else ""

        # Keep the fields separate so the dashboard can render tools as chips and
        # the description as subtext. Desc goes last (it may contain '|').
        remark_parts = []
        if tools:
            remark_parts.append(f"Tools: {tools}")
        if remark:
            remark_parts.append(f"Note: {remark}")
        if description:
            remark_parts.append(f"Desc: {description}")

        items.append({
            "item_number": code_val,
            "area_domain": current_area,
            "control_description": test_name or description or code_val,
            "compliance_status": _status_from_owasp_result(result),
            "evidence_reference": tools or None,
            "remarks": " | ".join(remark_parts) if remark_parts else None,
        })

    metadata = {
        "assessment_format": "owasp_v4_testing_checklist",
        "detected_format": "owasp_v4_testing_checklist",
        "columns_detected": [
            "item_number",
            "area_domain",
            "control_description",
            "compliance_status",
            "evidence_reference",
            "remarks",
        ],
    }
    return items, metadata


# =========================================================================== #
# Cyber Security maturity assessment tools (CREST / MMAT family)
# Four distinct workbooks — CSIR (Incident Response high-level), CTI (Threat
# Intelligence), IT Security Operations, and Incident Management (detailed) —
# all rate a hierarchy of questions/capabilities on a 1-5 CMMI maturity scale.
# Each is detected as its own format so it stays an independent assessment and
# maps to its own nav tab, but all normalise to the same maturity item shape
# and render through the shared MaturityAssessmentTab dashboard.
# =========================================================================== #

def _cell(row, idx):
    return str(row[idx] or "").strip() if row is not None and 0 <= idx < len(row) else ""


def _maturity_int(value) -> Optional[int]:
    try:
        n = int(float(str(value).strip()))
        return n if 1 <= n <= 5 else None
    except Exception:
        return None


def detect_maturity_kind(wb) -> Optional[str]:
    """Return 'csir' | 'cti' | 'itsecops' | 'incident' for a recognised CREST
    maturity workbook, else None. Distinguished by sheet names + the tool title
    so each of the four stays a separate assessment/tab."""
    try:
        sheets = set(wb.sheetnames)
    except Exception:
        return None
    if "1. Current and Target States" in sheets:
        return "itsecops"
    if {"Assessment - Phase 1", "Content"} & sheets and any(s.startswith("Assessment - Phase") for s in sheets):
        return "incident"
    if any(s.startswith("Assess ") for s in sheets) and "content" in {s.lower() for s in sheets}:
        return "cti"
    if "Assessment" in sheets and "MMAT ref" in sheets:
        return "csir"
    return None


def detect_cyber_maturity_format(wb) -> bool:
    return detect_maturity_kind(wb) is not None


def _parse_csir(wb) -> List[dict]:
    ws = wb["Assessment"]
    rows = list(ws.iter_rows(values_only=True))
    # header row carries "Statement" + "Level of maturity"
    hdr = None
    for i, r in enumerate(rows[:8]):
        norm = [_normalized_header(v) for v in (r or [])]
        if "statement" in norm and "levelofmaturity" in norm:
            hdr = i
            break
    if hdr is None:
        return []
    items: List[dict] = []
    current = "General"
    for r in rows[hdr + 1:]:
        stmt = _cell(r, 4)
        if not stmt:
            continue
        low = stmt.lower()
        if low.startswith("phase"):
            current = stmt
            continue
        if low.startswith("step"):
            ref = _cell(r, 13) or str(len(items) + 1)
            level = _maturity_int(r[5] if len(r) > 5 else None)
            weight = _cell(r, 6)
            parts = []
            if weight:
                parts.append(f"Weighting: {weight}")
            items.append({
                "item_number": ref,
                "area_domain": current,
                "control_description": stmt,
                "maturity_score": level,
                "compliance_status": "in_progress",
                "remarks": " | ".join(parts) or None,
            })
    return items


def _parse_content_style(wb, kind: str) -> List[dict]:
    """CTI 'content' and Incident 'Content' share a master-question layout:
    Order | FullQ | Stage/Phase | Step | Q | subQ | Text | Weighting | …
    Domain headers are single-letter (CTI stage) or plain-number (Incident
    phase) FullQ rows; questions carry the deepest FullQ / a Text value."""
    sheet = "content" if kind == "cti" else "Content"
    if sheet not in wb.sheetnames:
        sheet = next((s for s in wb.sheetnames if s.lower() == "content"), None)
        if not sheet:
            return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    norm = [_normalized_header(v) for v in rows[0]]
    def cix(name, default):
        return norm.index(name) if name in norm else default
    fq_i = cix("fullq", 1)
    txt_i = cix("text", 6)
    w_i = cix("weighting", 7)
    # The scored questions are the ones with a DEEP FullQ reference (≥2 dots,
    # e.g. "A.1.01" for CTI or "1.1.01"/"1.1.02a" for Incident). Rows with a
    # shallower ref are domain (0 dots: "A"/"1") or step headers (1 dot: "A.1"),
    # and rows with a BLANK FullQ are supporting detail / guidance — which the
    # Summary-Level tool does NOT score (verified against its Assess A-D forms).
    stage_re = re.compile(r"^[A-Za-z]$")
    items: List[dict] = []
    domain = "General"
    step = None
    for r in rows[1:]:
        fq = _cell(r, fq_i)
        text = _cell(r, txt_i)
        if not text:
            continue
        if not fq:
            continue  # detail / guidance rows have no reference — skip
        ndots = fq.count(".")
        if ndots == 0:
            domain = f"{fq}. {text}" if stage_re.match(fq) else text
            step = None
            continue
        if ndots == 1:
            step = text
            continue
        # ndots >= 2 → an assessable (scored) question.
        weight = _cell(r, w_i)
        parts = []
        if weight and weight.lower() != "n/a":
            parts.append(f"Weighting: {weight}")
        items.append({
            "item_number": fq,
            "area_domain": domain,
            "subdomain_name": step,
            "control_description": text,
            "maturity_score": None,
            "compliance_status": "in_progress",
            "remarks": " | ".join(parts) or None,
        })
    return items


def _parse_itsecops(wb) -> List[dict]:
    ws = wb["1. Current and Target States"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for i, r in enumerate(rows[:20]):
        norm = [_normalized_header(v) for v in (r or [])]
        if "promptid" in norm and "securitycapability" in norm:
            hdr = i
            hnorm = norm
            break
    if hdr is None:
        return []
    pid_i = hnorm.index("promptid")
    fn_i = hnorm.index("securityfunction") if "securityfunction" in hnorm else pid_i + 1
    cap_i = hnorm.index("securitycapability")
    cur_i = hnorm.index("currentstate") if "currentstate" in hnorm else cap_i + 1
    tgt_i = hnorm.index("targetstate12months") if "targetstate12months" in hnorm else None
    dim_i = fn_i + 1  # People/Process/Technology/Policy dimension sits between
    items: List[dict] = []
    current_fn = "Security Operations"
    for r in rows[hdr + 1:]:
        pid = _cell(r, pid_i)
        cap = _cell(r, cap_i)
        if not pid.upper().startswith("PID") or not cap:
            continue
        fn = _cell(r, fn_i)
        if fn:
            current_fn = fn
        dim = _cell(r, dim_i)
        cur = _maturity_int(r[cur_i] if len(r) > cur_i else None)
        tgt = _maturity_int(r[tgt_i] if (tgt_i is not None and len(r) > tgt_i) else None)
        parts = []
        if tgt is not None:
            parts.append(f"Target: {tgt}")
        if dim:
            parts.append(f"Dimension: {dim}")
        items.append({
            "item_number": pid,
            "area_domain": current_fn,
            "subdomain_name": dim or None,
            "control_description": cap,
            "maturity_score": cur,
            "compliance_status": "in_progress",
            "remarks": " | ".join(parts) or None,
        })
    return items


_MATURITY_LEVELS = [
    "Level 1 - Initial", "Level 2 - Established", "Level 3 - Business Enabling",
    "Level 4 - Quantitatively Managed", "Level 5 - Optimised",
]
_MATURITY_TITLES = {
    "csir": "Cyber Security Incident Response (High-level)",
    "cti": "Cyber Threat Intelligence",
    "itsecops": "IT Security Operations",
    "incident": "Incident Management (Detailed)",
}


def parse_cyber_maturity_workbook(wb, kind: str) -> tuple[List[dict], dict]:
    if kind == "csir":
        items = _parse_csir(wb)
    elif kind == "itsecops":
        items = _parse_itsecops(wb)
    else:
        items = _parse_content_style(wb, kind)
    fmt = f"{kind}_maturity"
    metadata = {
        "assessment_format": fmt,
        "detected_format": fmt,
        "maturity_kind": kind,
        "maturity_title": _MATURITY_TITLES.get(kind, "Cyber Maturity"),
        "maturity_levels": _MATURITY_LEVELS,
        "columns_detected": ["item_number", "area_domain", "subdomain_name", "control_description", "maturity_score", "remarks"],
    }
    return items, metadata


# --------------------------------------------------------------------------- #
# Cyber Security KPI Report — quarterly target-vs-actual metrics by domain.
# --------------------------------------------------------------------------- #

def detect_kpi_report_format(wb) -> bool:
    try:
        sheets = set(wb.sheetnames)
    except Exception:
        return False
    if not ({"KPI", "Measurement table"} <= sheets):
        return False
    ws = wb["Measurement table"]
    for r in ws.iter_rows(min_row=1, max_row=14, values_only=True):
        norm = [_normalized_header(v) for v in (r or [])]
        if "kpiid" in norm and "cybersecuritydomain" in norm:
            return True
    return False


def _kpi_definition_map(wb) -> dict:
    """kpi_id -> {name, description, type, frequency, source} from the 'KPI'
    sheet, so the dashboard can show each KPI's topic, cadence and data source
    (the Measurement table only carries the definition + numbers)."""
    out: dict = {}
    if "KPI" not in wb.sheetnames:
        return out
    ws = wb["KPI"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for i, r in enumerate(rows[:16]):
        norm = [_normalized_header(v) for v in (r or [])]
        if "kpiid" in norm and "kpitype" in norm:
            hdr, hn = i, norm
            break
    if hdr is None:
        return out
    def ix(name):
        return hn.index(name) if name in hn else -1
    def starts(pfx):
        return next((j for j, h in enumerate(hn) if h.startswith(pfx)), -1)
    id_i = ix("kpiid")
    name_i = ix("keyperformanceindicatorkpi")
    desc_i = ix("kpidescription")
    type_i = ix("kpitype")
    freq_i = starts("frequency")
    src_i = starts("datasource")
    for r in rows[hdr + 1:]:
        kid = _cell(r, id_i)
        if not kid:
            continue
        out[kid] = {
            "name": _cell(r, name_i),
            "description": _cell(r, desc_i),
            "type": _cell(r, type_i),
            "frequency": _cell(r, freq_i),
            "source": _cell(r, src_i),
        }
    return out


def parse_kpi_report_workbook(wb) -> tuple[List[dict], dict]:
    meta_map = _kpi_definition_map(wb)
    ws = wb["Measurement table"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = None
    for i, r in enumerate(rows[:16]):
        norm = [_normalized_header(v) for v in (r or [])]
        if "kpiid" in norm and "cybersecuritydomain" in norm:
            hdr, hn = i, norm
            break
    if hdr is None:
        return [], {"assessment_format": "kpi_report", "columns_detected": []}

    def ix(name):
        return hn.index(name) if name in hn else -1
    id_i = ix("kpiid")
    dom_i = ix("cybersecuritydomain")
    def_i = ix("kpidefinition")
    type_i = ix("kpitype")
    prior_i = ix("prioryearq4actual")
    # Per-quarter Target/Actual/Notes column indices (Q1..Q4).
    quarters = []
    for q in range(1, 5):
        quarters.append({
            "q": q,
            "t": ix(f"targetq{q}"),
            "a": ix(f"actualq{q}"),
            "n": ix(f"notesq{q}"),
        })

    def num(v):
        try:
            return float(v)
        except Exception:
            return None
    def fmt(v, is_pct):
        n = num(v)
        if n is None:
            return ""
        return f"{round(n * 100, 1)}%" if (is_pct and abs(n) <= 1.5) else (f"{round(n, 2)}" if isinstance(n, float) else str(n))

    items: List[dict] = []
    for r in rows[hdr + 1:]:
        kid = _cell(r, id_i)
        definition = _cell(r, def_i) if def_i >= 0 else ""
        if not (kid and definition):
            continue
        info = meta_map.get(kid, {})
        ktype = info.get("type") or (_cell(r, type_i) if type_i >= 0 else "")
        is_pct = "percent" in (ktype or "").lower()

        parts = []
        if info.get("name"):
            parts.append(f"Topic: {info['name']}")
        parts.append(f"Type: {ktype or 'Percentage'}")
        if info.get("frequency"):
            parts.append(f"Freq: {info['frequency']}")
        if info.get("source"):
            parts.append(f"Source: {info['source']}")
        parts.append(f"Def: {definition}")
        if prior_i >= 0:
            p = fmt(r[prior_i] if len(r) > prior_i else None, is_pct)
            if p:
                parts.append(f"Prior: {p}")
        note_parts = []
        for qc in quarters:
            t = fmt(r[qc["t"]] if (qc["t"] >= 0 and len(r) > qc["t"]) else None, is_pct)
            a = fmt(r[qc["a"]] if (qc["a"] >= 0 and len(r) > qc["a"]) else None, is_pct)
            if t or a:
                parts.append(f"Q{qc['q']}: {t or '-'}/{a or '-'}")
            n = _cell(r, qc["n"]) if qc["n"] >= 0 else ""
            if n and n != "-":
                note_parts.append(f"Q{qc['q']}Note: {n}")
        parts.extend(note_parts)

        items.append({
            "item_number": kid,
            "area_domain": _cell(r, dom_i) or "General",
            "control_description": info.get("name") or definition,
            "compliance_status": "in_progress",
            "remarks": " | ".join(parts) or None,
        })

    metadata = {
        "assessment_format": "kpi_report",
        "detected_format": "kpi_report",
        "columns_detected": ["item_number", "area_domain", "control_description", "remarks"],
    }
    return items, metadata


# --------------------------------------------------------------------------- #
# DPIA / PIA — Data Protection / Privacy Impact Assessment.
# A risk-assessment workbook: Screening (threshold Yes/No), an Assessment
# narrative, and a Risk Register (risks scored Likelihood×Impact with inherent +
# residual ratings). Parsed into one item list tagged by section so the
# dedicated dashboard can render the screening verdict, the risk heat-map and
# the risk register. Risk band: score = L×I → 1-4 Low, 5-9 Medium, 10-14 High,
# 15-25 Critical.
# --------------------------------------------------------------------------- #

def detect_dpia_format(wb) -> bool:
    try:
        sheets = wb.sheetnames
    except Exception:
        return False
    low = [s.lower() for s in sheets]
    has_screen = any("screening" in s for s in low)
    has_risk = any("risk register" in s for s in low)
    return has_screen and has_risk


def _dpia_sheet(wb, needle):
    for s in wb.sheetnames:
        if needle in s.lower():
            return wb[s]
    return None


def parse_dpia_workbook(wb) -> tuple[List[dict], dict]:
    items: List[dict] = []

    # 1) Screening — threshold questions (# | question | Yes/No | Notes).
    ws = _dpia_sheet(wb, "screening")
    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        for i, r in enumerate(rows[:8]):
            norm = [_normalized_header(v) for v in (r or [])]
            if "screeningquestion" in norm:
                hdr, hn = i, norm
                break
        if hdr is not None:
            q_i = hn.index("screeningquestion")
            ans_i = next((j for j, h in enumerate(hn) if h in ("yesno", "yes", "answer")), q_i + 1)
            note_i = next((j for j, h in enumerate(hn) if h.startswith("note")), ans_i + 1)
            n = 0
            for r in rows[hdr + 1:]:
                q = _cell(r, q_i)
                if not q or _normalized_header(q) == "screeningquestion":
                    continue
                n += 1
                ans = _cell(r, ans_i)
                note = _cell(r, note_i)
                parts = ["Section: screening"]
                if ans:
                    parts.append(f"Answer: {ans}")
                if note:
                    parts.append(f"Notes: {note}")
                items.append({
                    "item_number": f"S-{n:02d}",
                    "area_domain": "Screening",
                    "control_description": q,
                    "compliance_status": _status_from_asvs_valid(ans),
                    "remarks": " | ".join(parts),
                })

    # 2) Risk Register — the core deliverable.
    ws = _dpia_sheet(wb, "risk register")
    if ws is not None:
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        for i, r in enumerate(rows[:6]):
            if any(_normalized_header(v) == "riskid" for v in (r or [])):
                hdr, hn = i, [_normalized_header(v) for v in r]
                break
        if hdr is not None:
            def ix(*names):
                for nm in names:
                    if nm in hn:
                        return hn.index(nm)
                return -1
            c_id = ix("riskid")
            c_cat = ix("riskcategory")
            c_desc = ix("riskdescription")
            c_subj = ix("affecteddatasubjects")
            c_l = ix("likelihood15", "likelihood")
            c_i = ix("impact15", "impact")
            c_inh = ix("inherentscore")
            c_inhr = ix("inherentrating")
            c_ctrl = ix("existingplannedcontrols", "existingcontrols")
            c_owner = ix("controlowner")
            c_rl = ix("residlikelihood", "residuallikelihood")
            c_ri = ix("residimpact", "residualimpact")
            c_res = ix("residualscore")
            c_resr = ix("residualrating")
            c_fw = ix("frameworkreference")
            c_status = ix("status")
            c_target = ix("targetdate")
            for r in rows[hdr + 1:]:
                rid = _cell(r, c_id)
                desc = _cell(r, c_desc) if c_desc >= 0 else ""
                if not rid or not desc:
                    continue
                parts = ["Section: risk"]
                for label, idx in [("Subjects", c_subj), ("L", c_l), ("I", c_i),
                                   ("Inherent", c_inh), ("InherentRating", c_inhr),
                                   ("Controls", c_ctrl), ("Owner", c_owner),
                                   ("ResL", c_rl), ("ResI", c_ri), ("Residual", c_res),
                                   ("ResidualRating", c_resr), ("Framework", c_fw),
                                   ("Target", c_target)]:
                    v = _cell(r, idx) if idx >= 0 else ""
                    if v:
                        parts.append(f"{label}: {v}")
                items.append({
                    "item_number": rid,
                    "area_domain": _cell(r, c_cat) or "Uncategorised",
                    "control_description": desc,
                    "compliance_status": _status_from_asvs_valid(_cell(r, c_status) if c_status >= 0 else ""),
                    "risk_rating": (_cell(r, c_inhr) or None) if c_inhr >= 0 else None,
                    "remarks": " | ".join(parts),
                })

    # 3) Assessment narrative — field labels + guidance (context; often blank in
    # a fresh template but captured so nothing is dropped).
    ws = _dpia_sheet(wb, "assessment")
    if ws is not None:
        section = None
        n = 0
        for r in ws.iter_rows(values_only=True):
            vals = [_cell(r, j) for j in range(min(4, len(r)))]
            label = next((v for v in vals if v), "")
            if not label:
                continue
            # Section headers look like "A. Project & ownership".
            if re.match(r"^[A-Z]\.\s", label):
                section = label
                continue
            if label.upper().startswith("DPIA") or _normalized_header(label) == "":
                continue
            guidance = vals[1] if len(vals) > 1 and vals[1] and vals[1] != label else ""
            n += 1
            parts = ["Section: assessment"]
            if guidance:
                parts.append(f"Guidance: {guidance}")
            items.append({
                "item_number": f"A-{n:02d}",
                "area_domain": "Assessment",
                "subdomain_name": section,
                "control_description": label,
                "compliance_status": "in_progress",
                "remarks": " | ".join(parts),
            })

    metadata = {
        "assessment_format": "dpia_pia",
        "detected_format": "dpia_pia",
        "risk_bands": {"Low": [1, 4], "Medium": [5, 9], "High": [10, 14], "Critical": [15, 25]},
        "columns_detected": ["item_number", "area_domain", "control_description", "compliance_status", "remarks"],
    }
    return items, metadata


# --------------------------------------------------------------------------- #
# NCA register templates — Vulnerability Register, Cybersecurity Audit Plan and
# Cybersecurity Risk Register. Each is a tabular log (Legend sheet + a data
# sheet + Heat-map/Summary). They share a flexible parser: locate the data
# sheet + its header row, then capture every column of every row into the item
# (the full row is preserved as JSON in `remarks` so nothing is dropped, while
# the key fields are also mapped onto dedicated item columns for the dashboard).
# --------------------------------------------------------------------------- #

_NCA_REGISTER_CONF = {
    "vuln": {
        "sheet": "vulnerability register", "legend": "vulnerability register legend",
        "id": ["vulnerabilityid"], "name": ["title"], "desc": ["vulnerabilitydescription"],
        "group": ["risklevel", "riskseverity"], "status": ["status"],
        "sev": ["risklevel"], "owner": ["owner"], "due": ["duedate"],
    },
    "audit": {
        "sheet": "audit plan", "legend": "audit plan legend",
        "id": ["auditid"], "name": ["auditname"], "desc": ["scopeofaudit"],
        "group": ["typeofaudit", "teamresponsible"], "status": ["status"],
        "sev": [], "owner": ["leadauditor"], "due": ["auditend"],
    },
    "risk": {
        "sheet": "risk register", "legend": "risk register legend",
        "id": ["riskidentifier"], "name": ["descriptionoftherisk"], "desc": ["descriptionoftherisk"],
        "group": ["riskareascopeofrisk", "riskarea"], "status": ["status"],
        "sev": ["overallinherentriskrating", "updatedoverallinherentriskrating"],
        "owner": ["riskowner"], "due": ["deadlineforaction"],
    },
}


def detect_nca_register_kind(wb) -> Optional[str]:
    try:
        low = {s.strip().lower() for s in wb.sheetnames}
    except Exception:
        return None
    for kind, cfg in _NCA_REGISTER_CONF.items():
        if cfg["sheet"] in low and cfg["legend"] in low:
            return kind
    return None


def detect_nca_register_format(wb) -> bool:
    return detect_nca_register_kind(wb) is not None


def parse_nca_register_workbook(wb, kind: str) -> tuple[List[dict], dict]:
    cfg = _NCA_REGISTER_CONF[kind]
    ws = next((wb[s] for s in wb.sheetnames if s.strip().lower() == cfg["sheet"]), None)
    if ws is None:
        return [], {"assessment_format": f"nca_{kind}_register"}
    rows = list(ws.iter_rows(values_only=True))
    # Header row = first row with >= 4 non-empty text cells.
    hdr = None
    for i, r in enumerate(rows[:30]):
        if sum(1 for x in r if str(x or "").strip()) >= 4:
            hdr = i
            break
    if hdr is None:
        return [], {"assessment_format": f"nca_{kind}_register"}
    headers = [str(v or "").strip() for v in rows[hdr]]
    norm = [_normalized_header(v) for v in headers]

    def col(cands):
        for cand in cands:
            if cand in norm:
                return norm.index(cand)
        return -1
    id_i = col(cfg["id"])
    name_i = col(cfg["name"])
    desc_i = col(cfg["desc"])
    group_i = col(cfg["group"])
    status_i = col(cfg["status"])
    sev_i = col(cfg["sev"])
    owner_i = col(cfg["owner"])
    due_i = col(cfg["due"])
    if id_i < 0:
        id_i = 0

    items: List[dict] = []
    for r in rows[hdr + 1:]:
        if sum(1 for x in r if str(x or "").strip()) < 2:
            continue
        rid = _cell_to_text(r[id_i]) if id_i < len(r) else None
        if not rid or _normalized_header(rid) in norm:
            continue
        name = _cell_to_text(r[name_i]) if 0 <= name_i < len(r) else None
        desc = _cell_to_text(r[desc_i]) if 0 <= desc_i < len(r) else None
        # Skip blank template placeholder rows. These NCA templates ship with a
        # few worked examples plus several empty rows (just an ID) for the user
        # to fill. Keep a row that has a title/description OR enough real data
        # (e.g. a scoring-only vulnerability) — drop pure-ID placeholders.
        filled = sum(1 for x in r if str(x or "").strip())
        if not (name or desc) and filled < 4:
            continue
        # Full row → {header: value} JSON so no column is lost.
        rowdict = {}
        for j, h in enumerate(headers):
            if not h or j >= len(r):
                continue
            v = _cell_to_text(r[j])
            if v:
                rowdict[h] = v
        items.append({
            "item_number": rid[:50],
            "area_domain": (_cell_to_text(r[group_i]) if 0 <= group_i < len(r) else None) or "General",
            "control_description": (name or desc or rid)[:8000],
            "compliance_status": _status_from_nca_register_status(_cell_to_text(r[status_i]) if 0 <= status_i < len(r) else None),
            "priority": (_cell_to_text(r[sev_i]) if 0 <= sev_i < len(r) else None),
            "risk_rating": (_cell_to_text(r[sev_i]) if 0 <= sev_i < len(r) else None),
            "responsible_party": (_cell_to_text(r[owner_i]) if 0 <= owner_i < len(r) else None),
            "timeline": (_cell_to_text(r[due_i]) if 0 <= due_i < len(r) else None),
            "remarks": json.dumps(rowdict, ensure_ascii=False),
        })

    fmt = f"nca_{kind}_register"
    metadata = {
        "assessment_format": fmt,
        "detected_format": fmt,
        "nca_register_kind": kind,
        "columns_detected": ["item_number", "area_domain", "control_description", "compliance_status", "priority", "remarks"],
    }
    return items, metadata


def _status_from_nca_register_status(value) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return "in_progress"
    if any(t in text for t in ["resolved", "closed", "done", "complete", "mitigated"]):
        return "complied"
    if any(t in text for t in ["open", "in progress", "on hold", "planned", "new"]):
        return "in_progress"
    if any(t in text for t in ["accepted", "n/a", "not applicable"]):
        return "na"
    return "in_progress"


# --------------------------------------------------------------------------- #
# NCA DCC-1:2022 Data Cybersecurity Controls — the bilingual (Arabic/English)
# Assessment & Compliance Excel tool. Hierarchical controls (Main Domain →
# Subdomain → control ref "1-1-2" / subcontrol "1-2-1-1") with a bilingual
# requirement text, a control type (أساسي Essential / فرعي Sub) and a compliance
# status (Implemented / Partially / Not Implemented / Not Applicable).
# --------------------------------------------------------------------------- #

_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def _dcc_norm(value) -> str:
    t = str(value or "").strip().translate(_ARABIC_DIGITS)
    for d in ("–", "—", "‑", "−", "­", "ـ"):
        t = t.replace(d, "-")
    return t


def _dcc_controls_sheet(wb):
    return next((wb[s] for s in wb.sheetnames if "الالتزام بالضوابط" in s), None)


def detect_dcc_tool_format(wb) -> bool:
    try:
        sheets = wb.sheetnames
    except Exception:
        return False
    return any("الالتزام بالضوابط" in s for s in sheets) and "tbl_choices" in sheets


def _dcc_status(value) -> str:
    t = str(value or "").strip().lower()
    if not t:
        return "in_progress"
    if "partial" in t or "جزئ" in t:
        return "partially_complied"
    if "not implemented" in t or "غير مطبق" in t:
        return "not_complied"
    if "not applicable" in t or "لاينطبق" in t or "لا ينطبق" in t:
        return "na"
    if "implement" in t or "مطبق" in t:
        return "complied"
    return "in_progress"


def parse_dcc_tool_workbook(wb) -> tuple[List[dict], dict]:
    ws = _dcc_controls_sheet(wb)
    if ws is None:
        return [], {"assessment_format": "nca_dcc_tool"}
    rows = [[str(x).strip() if x is not None else "" for x in r] for r in ws.iter_rows(values_only=True)]

    dom_re = re.compile(r"^(\d+)-\s*\D")         # "1- Cybersecurity Governance"
    sub_re = re.compile(r"^(\d+-\d+)(\s|$)")     # "1-1" (subdomain, 2 parts)
    ctl_re = re.compile(r"^(\d+-\d+-\d+(?:-\d+)*)")  # "1-1-2" / "1-2-1-1" (control)
    clean = lambda s: re.sub(r"\s+", " ", s).strip()

    # Single pass with CARRY-FORWARD grouping. A control's real domain is the
    # main-domain header it physically sits under (Governance / Defense /
    # Third-Party) — NOT the first digit of its ref (the DCC ref numbering does
    # not correspond to the domain, e.g. a "5-1-6-2" control lives under
    # Defense). We track the last domain / subdomain header seen and assign it.
    items: List[dict] = []
    current_domain = None
    current_subdomain = None
    for r in rows:
        cells = [(raw, _dcc_norm(raw)) for raw in r if raw]
        if not cells:
            continue
        # Main-domain header ("N- <name>"), not a ref cell.
        dcell = next((raw for raw, nv in cells if dom_re.match(nv) and not ctl_re.match(nv) and len(raw) > 8), None)
        if dcell:
            current_domain = clean(dcell)
            current_subdomain = None
        # Subdomain header (a bare "N-M" ref + its name).
        subm = next(((raw, nv) for raw, nv in cells if sub_re.match(nv) and not ctl_re.match(nv)), None)
        if subm:
            name = max((o for o, onv in cells if onv != subm[1] and not ctl_re.match(onv) and not dom_re.match(onv)), key=len, default="")
            current_subdomain = clean(f"{subm[1]} {name}") if name else subm[1]
        # Control row (3+ part ref).
        ref_m = next((ctl_re.match(nv) for _, nv in cells if ctl_re.match(nv)), None)
        if not ref_m:
            continue
        ref = ref_m.group(1)
        text = max((raw for raw, nv in cells
                    if nv != ref and raw not in ("أساسي", "فرعي") and not ctl_re.match(nv) and len(raw) > 8),
                   key=len, default="")
        ctrl_type = next((raw for raw, _ in cells if raw in ("أساسي", "فرعي")), "")
        type_label = "Essential" if ctrl_type == "أساسي" else ("Sub" if ctrl_type == "فرعي" else "")
        remark_parts = [f"Ref: {ref}"]
        if type_label:
            remark_parts.append(f"Type: {type_label}")
        items.append({
            "item_number": ref[:50],
            "area_domain": current_domain or "DCC",
            "subdomain_name": current_subdomain,
            "control_description": clean(text or ref)[:8000],
            "compliance_status": "in_progress",
            "priority": ("high" if ctrl_type == "أساسي" else "medium" if ctrl_type == "فرعي" else None),
            "remarks": " | ".join(remark_parts),
        })

    metadata = {
        "assessment_format": "nca_dcc_tool",
        "detected_format": "nca_dcc_tool",
        "framework": "Saudi NCA DCC-1:2022",
        "columns_detected": ["item_number", "area_domain", "subdomain_name", "control_description", "compliance_status", "priority", "remarks"],
    }
    return items, metadata


def _normalize_pdf_signature_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def _contains_pdf_signature(probe: str, signatures: List[str]) -> bool:
    normalized_probe = _normalize_pdf_signature_text(probe)
    return any(_normalize_pdf_signature_text(signature) in normalized_probe for signature in signatures)


def _extract_cloud_controls_from_lines(lines: List[str]) -> List[dict]:
    control_pattern = re.compile(r"^\s*(\d{1,2}-\d{1,2}-[PT]-\d{1,2}(?:-\d{1,2})?)\s*(.*)$", re.IGNORECASE)
    provider_tenant_map = {"P": "CSP", "T": "CST"}
    domain_map = {
        "1": "Cybersecurity Governance",
        "2": "Cybersecurity Defense",
        "3": "Cybersecurity Resilience",
        "4": "Third-Party Cybersecurity",
    }

    controls: List[dict] = []
    seen_ids = set()
    current_id: Optional[str] = None
    current_desc_parts: List[str] = []

    def flush_current() -> None:
        nonlocal current_id, current_desc_parts
        if not current_id or current_id in seen_ids:
            current_id = None
            current_desc_parts = []
            return

        parts = [part.strip() for part in current_desc_parts if part and part.strip()]
        merged_description = " ".join(parts).strip()
        if not merged_description:
            merged_description = f"Control requirement for {current_id}"

        id_parts = current_id.split("-")
        main_domain_no = id_parts[0] if id_parts else ""
        party_code = id_parts[2].upper() if len(id_parts) >= 3 else ""
        party_label = provider_tenant_map.get(party_code)
        domain_label = domain_map.get(main_domain_no, f"Domain {main_domain_no}" if main_domain_no else "Cloud Cybersecurity Controls")
        if party_label:
            domain_label = f"{domain_label} ({party_label})"

        controls.append({
            "item_number": current_id,
            "area_domain": domain_label,
            "control_description": merged_description,
            "compliance_status": "in_progress",
            "priority": "medium",
            # Remarks left blank — assessors fill their own notes here. The
            # framework banner that used to occupy this field was redundant
            # with the assessment-level metadata.
            "remarks": None,
        })
        seen_ids.add(current_id)
        current_id = None
        current_desc_parts = []

    for raw_line in lines:
        line = re.sub(r"\s+", " ", (raw_line or "")).strip()
        if not line:
            continue
        if len(line) <= 2:
            continue
        if line.lower().startswith(("document classification", "table of contents", "list of ", "figure ", "table ")):
            continue

        match = control_pattern.match(line)
        if match:
            flush_current()
            current_id = match.group(1).upper()
            tail = (match.group(2) or "").strip(" :-")
            if tail:
                current_desc_parts.append(tail)
            continue

        if current_id:
            # Keep short continuation lines for wrapped control text.
            if len(line) < 220:
                current_desc_parts.append(line)

    flush_current()
    return controls


def _ccc_sort_key(code: str) -> Tuple[int, int, int, int, int]:
    """Order CCC items deterministically — domain, subdomain, party (P before
    T), main control, subcontrol. Codes that fall outside this shape sort
    last so we never crash on a stray identifier."""
    parts = (code or "").split("-")
    def _i(value: str, fallback: int = 999) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return fallback
    party_order = {"P": 0, "T": 1}
    domain = _i(parts[0]) if len(parts) > 0 else 999
    subdomain = _i(parts[1]) if len(parts) > 1 else 999
    party = party_order.get(parts[2].upper(), 9) if len(parts) > 2 else 9
    main_ctrl = _i(parts[3]) if len(parts) > 3 else 0
    sub_ctrl = _i(parts[4]) if len(parts) > 4 else 0
    return (domain, subdomain, party, main_ctrl, sub_ctrl)


def _apply_curated_ccc_text(items: List[dict]) -> List[dict]:
    """Override extracted CCC descriptions with curated official text and
    inject any curated codes the PDF extractor missed entirely.

    Curated text wins over the line-by-line PyPDF extraction because the
    extractor regularly produces fragments / partial sentences on the
    CCC-2:2024 PDF — the source document is dense with bullet lists that
    don't survive plain-text extraction cleanly. Codes not in the curated
    file keep whatever the extractor produced (no regression for items
    we haven't transcribed yet).
    """
    curated = _get_ccc_curated_text()
    if not curated:
        return items

    seen_codes = {item.get("item_number") for item in items if item.get("item_number")}

    for item in items:
        code = (item.get("item_number") or "").strip().upper()
        # Re-key against the curated map's casing convention (P/T are upper).
        if code in curated:
            item["control_description"] = curated[code]
        elif code.upper() in curated:
            item["control_description"] = curated[code.upper()]

    provider_tenant_map = {"P": "CSP", "T": "CST"}
    domain_map = {
        "1": "Cybersecurity Governance",
        "2": "Cybersecurity Defense",
        "3": "Cybersecurity Resilience",
        "4": "Third-Party Cybersecurity",
    }
    for code, text in curated.items():
        if code in seen_codes:
            continue
        id_parts = code.split("-")
        main_domain_no = id_parts[0] if id_parts else ""
        party_code = id_parts[2].upper() if len(id_parts) >= 3 else ""
        party_label = provider_tenant_map.get(party_code)
        domain_label = domain_map.get(main_domain_no, f"Domain {main_domain_no}" if main_domain_no else "Cloud Cybersecurity Controls")
        if party_label:
            domain_label = f"{domain_label} ({party_label})"
        items.append({
            "item_number": code,
            "area_domain": domain_label,
            "control_description": text,
            "compliance_status": "in_progress",
            "priority": "medium",
            # See note above: leave remarks blank for assessor input.
            "remarks": None,
        })
        seen_codes.add(code)

    items.sort(key=lambda it: _ccc_sort_key(it.get("item_number") or ""))
    return items


DCC_DOMAIN_LABELS: Dict[int, str] = {
    1: "Cybersecurity Governance",
    2: "Cybersecurity Defense",
    3: "Third-Party and Cloud Computing Cybersecurity",
}


# ---------------------------------------------------------------------------
# Curated NCA control text
# ---------------------------------------------------------------------------
# The DCC and CCC PDFs render their control clauses as flowing text with
# nested bullets, so PyPDF2 text extraction (and even OCR) routinely produces
# misaligned fragments. Rather than depend on that, we ship hand-transcribed
# clauses from the official NCA documents and use them as the *primary* text
# source. The PDF/OCR path remains as a fallback, so anything we haven't
# transcribed yet still gets the best-effort extraction it had before.
#
# Files are read once per process and cached.
# ---------------------------------------------------------------------------

_NCA_TEXT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "seed_data", "nca_control_text",
)
_DCC_CONTROL_TEXT_PATH = os.path.join(_NCA_TEXT_DIR, "dcc_control_text.json")
_CCC_CONTROL_TEXT_PATH = os.path.join(_NCA_TEXT_DIR, "ccc_control_text.json")

_DCC_CONTROL_TEXT_CACHE: Optional[Dict[str, str]] = None
_CCC_CONTROL_TEXT_CACHE: Optional[Dict[str, str]] = None


def _load_curated_control_text(path: str) -> Dict[str, str]:
    """Read the `{code: text}` map from a curated NCA JSON file.

    Returns an empty dict on any failure so the upload path falls through
    cleanly to OCR/text-extraction rather than 500ing.
    """
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh) or {}
    except Exception:
        return {}
    controls = payload.get("controls") if isinstance(payload, dict) else None
    if not isinstance(controls, dict):
        return {}
    out: Dict[str, str] = {}
    for code, text in controls.items():
        if isinstance(code, str) and isinstance(text, str) and text.strip():
            out[code.strip()] = text.strip()
    return out


def _get_dcc_curated_text() -> Dict[str, str]:
    global _DCC_CONTROL_TEXT_CACHE
    if _DCC_CONTROL_TEXT_CACHE is None:
        _DCC_CONTROL_TEXT_CACHE = _load_curated_control_text(_DCC_CONTROL_TEXT_PATH)
    return _DCC_CONTROL_TEXT_CACHE


def _get_ccc_curated_text() -> Dict[str, str]:
    global _CCC_CONTROL_TEXT_CACHE
    if _CCC_CONTROL_TEXT_CACHE is None:
        _CCC_CONTROL_TEXT_CACHE = _load_curated_control_text(_CCC_CONTROL_TEXT_PATH)
    return _CCC_CONTROL_TEXT_CACHE


def _dcc_code_sort_key(code: str) -> Tuple[int, int, int, int]:
    parts = [int(part) for part in code.split("-")]
    while len(parts) < 4:
        parts.append(0)
    return tuple(parts[:4])


def _build_dcc_expected_code_catalog() -> List[Tuple[str, int]]:
    entries: List[Tuple[str, int]] = []

    def add_control(domain: int, subdomain: int, control: int, max_subcontrols: Optional[int] = None) -> None:
        entries.append((f"{domain}-{subdomain}-{control}", domain))
        if max_subcontrols:
            for subcontrol in range(1, max_subcontrols + 1):
                entries.append((f"{domain}-{subdomain}-{control}-{subcontrol}", domain))

    # Domain 1: Cybersecurity Governance
    add_control(1, 1, 1)
    add_control(1, 1, 2)
    add_control(1, 2, 1, 2)
    add_control(1, 3, 1, 7)

    # Domain 2: Cybersecurity Defense
    add_control(2, 1, 1, 2)
    add_control(2, 1, 2)
    add_control(2, 1, 3)
    add_control(2, 2, 1, 4)
    add_control(2, 3, 1, 2)
    add_control(2, 4, 1, 4)
    add_control(2, 5, 1, 2)
    add_control(2, 6, 1, 5)
    add_control(2, 6, 2)
    add_control(2, 7, 1)
    add_control(2, 7, 2)
    add_control(2, 7, 3, 5)
    add_control(2, 7, 4)

    # Domain 3: Third-Party and Cloud Computing Cybersecurity
    add_control(3, 1, 1, 6)
    add_control(3, 1, 2, 8)
    add_control(3, 2, 1, 7)

    return sorted(entries, key=lambda item: _dcc_code_sort_key(item[0]))


def _resolve_tesseract_cmd() -> Optional[str]:
    candidates = [
        os.environ.get("TESSERACT_CMD"),
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _clean_dcc_ocr_fragment(value: str) -> str:
    text = re.sub(r"[\|\u2013\u2014]+", " ", value or "")
    text = re.sub(r"\b[vV](?:\s*[vV])+\b", " ", text)
    text = re.sub(r"\s{2,}", " ", text).strip(" -:;,.")
    if not text:
        return ""
    if re.fullmatch(r"[vV\W\d_]+", text):
        return ""
    return text


def _extract_dcc_control_descriptions_from_ocr(file_content: bytes) -> Dict[str, str]:
    try:
        import fitz
        from PIL import Image, ImageOps
        import pytesseract
    except Exception:
        return {}

    tesseract_cmd = _resolve_tesseract_cmd()
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    expected_codes = {code for code, _ in _build_dcc_expected_code_catalog()}
    descriptions: Dict[str, List[str]] = {}
    line_pattern = re.compile(r"^\s*[^0-9]{0,4}(?P<code>[1-3]-\d{1,2}-\d{1,2}(?:-\d{1,2})?)\b(?P<tail>.*)$")

    try:
        doc = fitz.open(stream=file_content, filetype="pdf")
    except Exception:
        return {}

    # DCC controls are in the middle pages of the provided template.
    start_page = 10 if doc.page_count > 15 else 0
    end_page = min(doc.page_count, 20 if doc.page_count > 20 else doc.page_count)

    for page_index in range(start_page, end_page):
        try:
            page = doc[page_index]
            pix = page.get_pixmap(matrix=fitz.Matrix(2.6, 2.6), alpha=False)
            image = Image.open(io.BytesIO(pix.tobytes("png")))
            image = ImageOps.grayscale(image).point(lambda p: 255 if p > 175 else 0)
            text = pytesseract.image_to_string(image, lang="eng", config="--oem 1 --psm 6")
        except Exception:
            continue

        current_code: Optional[str] = None
        for raw_line in text.splitlines():
            line = re.sub(r"\s+", " ", (raw_line or "")).strip()
            if not line:
                continue
            lowered = line.lower()

            matched = line_pattern.match(line)
            if matched:
                code = matched.group("code")
                current_code = code if code in expected_codes else None
                if current_code:
                    tail = _clean_dcc_ocr_fragment(matched.group("tail"))
                    if tail and "ecc control" not in lowered and "ecc subcontrol" not in lowered:
                        descriptions.setdefault(current_code, []).append(tail)
                continue

            if not current_code:
                continue
            if lowered.startswith(("document classification", "data cybersecurity controls", "appendix", "figure", "table ")):
                continue
            if len(line) > 220:
                continue
            fragment = _clean_dcc_ocr_fragment(line)
            if fragment and "ecc control" not in lowered and "ecc subcontrol" not in lowered:
                descriptions.setdefault(current_code, []).append(fragment)

    merged: Dict[str, str] = {}
    for code, fragments in descriptions.items():
        unique_fragments: List[str] = []
        seen = set()
        for fragment in fragments:
            normalized = fragment.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            unique_fragments.append(fragment)
        merged_text = " ".join(unique_fragments[:3]).strip()
        merged_text = re.sub(r"\s{2,}", " ", merged_text)
        if merged_text:
            merged[code] = merged_text

    return merged


def _build_dcc_full_control_items(file_content: bytes) -> List[dict]:
    expected_catalog = _build_dcc_expected_code_catalog()
    # Curated text from the official PDF — primary source. OCR remains as a
    # secondary signal so any code the curated file doesn't cover yet still
    # gets best-effort text rather than only a placeholder.
    curated = _get_dcc_curated_text()
    ocr_map = _extract_dcc_control_descriptions_from_ocr(file_content) if not curated else {}
    items: List[dict] = []

    for code, domain in expected_catalog:
        description = curated.get(code) or ocr_map.get(code)
        if not description:
            if len(code.split("-")) == 3:
                description = f"Saudi NCA DCC control requirement {code}."
            else:
                description = f"Saudi NCA DCC subcontrol requirement {code}."

        items.append({
            "item_number": code,
            "area_domain": DCC_DOMAIN_LABELS.get(domain, f"DCC Domain {domain}"),
            "control_description": description,
            "compliance_status": "in_progress",
            "priority": "medium",
            # Remarks intentionally left blank so the assessor's own notes
            # land here when they fill the row in; the framework banner that
            # used to pre-fill this field was just visual noise.
            "remarks": None,
        })

    return items


def parse_cis_windows_server_2012_r2_pdf(file_content: bytes, file_name: str) -> tuple[List[dict], dict]:
    from PyPDF2 import PdfReader

    try:
        reader = PdfReader(io.BytesIO(file_content))
        all_page_text = [(page.extract_text() or "") for page in reader.pages]
        sample_text = " ".join(all_page_text[: min(8, len(all_page_text))])
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unable to read PDF: {exc}",
        ) from exc

    probe_text = f"{file_name or ''} {sample_text}"

    cis_signatures = [
        "cis microsoft windows server 2012 r2 benchmark",
        "windows server 2012 r2 benchmark",
    ]
    nca_cloud_signatures = [
        "saudi nca cloud cybersecurity controls",
        "cloud cybersecurity controls",
        "ccc 2 2024",
        "ccc-2:2024",
    ]
    nca_data_signatures = [
        "saudi nca data cybersecurity controls",
        "data cybersecurity controls",
        "dcc 1 2022",
        "dcc-1:2022",
    ]

    if _contains_pdf_signature(probe_text, cis_signatures):
        if not os.path.exists(CIS_WS2012R2_CONTROLS_JSON):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"CIS controls seed not found at {CIS_WS2012R2_CONTROLS_JSON}",
            )

        try:
            with open(CIS_WS2012R2_CONTROLS_JSON, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to load CIS controls seed: {exc}",
            ) from exc

        controls = payload.get("controls") or []
        if not controls:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No controls found in CIS controls seed for this PDF handler.",
            )

        items: List[dict] = []
        for index, control in enumerate(controls):
            control_id = str(control.get("ControlID") or "").strip() or str(index + 1)
            section = str(control.get("Section") or "").strip() or "CIS Benchmark"
            title = str(control.get("Title") or "").strip()
            description = str(control.get("Description") or "").strip()
            expected_value = str(control.get("ExpectedValue") or "").strip()
            level = str(control.get("Level") or "").strip().upper()

            control_description = title or f"CIS Control {control_id}"
            if description:
                control_description = f"{control_description}\n\n{description}"
            if expected_value:
                control_description = f"{control_description}\n\nExpected Value: {expected_value}"

            remarks_parts = []
            assessment = str(control.get("Assessment") or "").strip()
            scan_type = str(control.get("ScanType") or "").strip()
            if level:
                remarks_parts.append(f"Level: {level}")
            if assessment:
                remarks_parts.append(f"Assessment: {assessment}")
            if scan_type:
                remarks_parts.append(f"Scan Type: {scan_type}")

            priority = "high" if level == "L1" else "medium" if level == "L2" else None
            items.append({
                "item_number": control_id,
                "area_domain": section,
                "control_description": control_description,
                "compliance_status": "in_progress",
                "priority": priority,
                "remarks": " | ".join(remarks_parts) if remarks_parts else None,
            })

        metadata = {
            "assessment_format": "cis_windows_server_2012_r2_pdf",
            "detected_format": "cis_windows_server_2012_r2_pdf",
            "benchmark": payload.get("benchmark"),
            "version": payload.get("version"),
            "columns_detected": [
                "item_number",
                "area_domain",
                "control_description",
                "compliance_status",
                "priority",
                "remarks",
            ],
        }
        return items, metadata

    all_lines = []
    for chunk in all_page_text:
        all_lines.extend(chunk.splitlines())

    if _contains_pdf_signature(probe_text, nca_cloud_signatures):
        cloud_items = _extract_cloud_controls_from_lines(all_lines)
        # Override per-control descriptions with curated official text and
        # backfill any codes the line-by-line extractor missed. Falls through
        # cleanly to the raw extraction when the curated JSON is absent.
        cloud_items = _apply_curated_ccc_text(cloud_items or [])
        if not cloud_items:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not extract controls from Saudi NCA Cloud Cybersecurity Controls PDF.",
            )
        metadata = {
            "assessment_format": "nca_cloud_cybersecurity_controls_pdf",
            "detected_format": "nca_cloud_cybersecurity_controls_pdf",
            "framework": "Saudi NCA Cloud Cybersecurity Controls",
            "version": "CCC-2:2024",
            "columns_detected": [
                "item_number",
                "area_domain",
                "control_description",
                "compliance_status",
                "priority",
                "remarks",
            ],
        }
        return cloud_items, metadata

    if _contains_pdf_signature(probe_text, nca_data_signatures):
        dcc_items = _build_dcc_full_control_items(file_content)
        if not dcc_items:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not extract control-level items from Saudi NCA Data Cybersecurity Controls PDF.",
            )
        metadata = {
            "assessment_format": "nca_data_cybersecurity_controls_pdf",
            "detected_format": "nca_data_cybersecurity_controls_pdf",
            "framework": "Saudi NCA Data Cybersecurity Controls",
            "version": "DCC-1:2022",
            "coverage": "full_control_level",
            "total_expected_controls": len(dcc_items),
            "columns_detected": [
                "item_number",
                "area_domain",
                "control_description",
                "compliance_status",
                "priority",
                "remarks",
            ],
        }
        return dcc_items, metadata

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            "PDF upload for assessments currently supports CIS Windows Server 2012 R2 Benchmark, "
            "Saudi NCA Cloud Cybersecurity Controls, and Saudi NCA Data Cybersecurity Controls formats."
        ),
    )


def calculate_assessment_stats(items: List[ComplianceAssessmentDocumentItem]) -> dict:
    stats = {
        "total": len(items),
        "complied": 0,
        "partially_complied": 0,
        "not_complied": 0,
        "in_progress": 0,
        "na": 0,
        "overall_score": 0.0
    }
    
    for item in items:
        status_val = item.compliance_status or "in_progress"
        if status_val in stats:
            stats[status_val] += 1
    
    applicable_items = stats["total"] - stats["na"]
    if applicable_items > 0:
        stats["overall_score"] = round(
            (stats["complied"] + (stats["partially_complied"] * 0.5)) / applicable_items * 100, 2
        )
    
    return stats


def _fallback_assessment_context(payload: AssessmentContextRequest) -> dict:
    assessment_type_label = payload.assessment_type.replace("_", " ").title()
    source_text = f" Source: {payload.source}." if payload.source else ""
    notes_text = f" Notes: {payload.notes}." if payload.notes else ""
    summary = (
        f"{payload.name} is a {assessment_type_label} focused assessment to evaluate current control effectiveness, "
        f"identify non-compliance gaps, and prioritize remediation actions.{source_text}{notes_text}"
    )
    risk_perspective = (
        "From a risk perspective, this assessment improves visibility of control weaknesses, supports risk scoring updates, "
        "and helps reduce residual risk through targeted corrective actions."
    )
    compliance_perspective = (
        "From a compliance perspective, this assessment provides audit-ready evidence of due diligence, tracks requirement-level "
        "conformance, and supports ongoing regulatory and framework adherence."
    )
    return {
        "summary": summary,
        "risk_perspective": risk_perspective,
        "compliance_perspective": compliance_perspective,
        "generated_by": "fallback",
    }


@router.post("/ai-context")
def generate_assessment_ai_context(
    request: AssessmentContextRequest,
    current_user: GRCUser = Depends(require_auth),
):
    api_key = AI_INTEGRATIONS_OPENAI_API_KEY or os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return _fallback_assessment_context(request)

    try:
        client_kwargs = {"api_key": api_key}
        if AI_INTEGRATIONS_OPENAI_BASE_URL:
            client_kwargs["base_url"] = AI_INTEGRATIONS_OPENAI_BASE_URL
        client = OpenAI(**client_kwargs)

        prompt = f"""
You are a senior GRC advisor.
Generate a concise assessment context in JSON for this compliance assessment upload:

Assessment Name: {request.name}
Assessment Type: {request.assessment_type}
Source: {request.source or 'N/A'}
Notes: {request.notes or 'N/A'}

Return strict JSON with keys:
- summary (2-3 sentences describing what this assessment is about)
- risk_perspective (1-2 sentences on risk-management value)
- compliance_perspective (1-2 sentences on compliance/audit value)
"""

        response = client.chat.completions.create(
            model=get_openai_model(),
            temperature=0.3,
            messages=[
                {"role": "system", "content": "You produce concise, practical GRC guidance as valid JSON."},
                {"role": "user", "content": prompt},
            ],
        )
        content = response.choices[0].message.content or "{}"
        parsed = json.loads(content)

        return {
            "summary": parsed.get("summary") or _fallback_assessment_context(request)["summary"],
            "risk_perspective": parsed.get("risk_perspective") or _fallback_assessment_context(request)["risk_perspective"],
            "compliance_perspective": parsed.get("compliance_perspective") or _fallback_assessment_context(request)["compliance_perspective"],
            "generated_by": "ai",
        }
    except Exception:
        logger.warning("AI context generation failed, using fallback", exc_info=True)
        return _fallback_assessment_context(request)


@router.post("/upload")
async def upload_assessment(
    name: str = Form(...),
    assessment_type: str = Form(...),
    source: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    assessor: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    tenant_id: Optional[int] = Form(None),
    expected_format: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    logger.info(f"Assessment upload started: name={name}, file={file.filename}, expected_format={expected_format}")
    file_path = None
    
    try:
        if tenant_id:
            validate_tenant_access(current_user, tenant_id, db)
        else:
            tenant_id = get_user_primary_tenant(current_user, db)
            if not tenant_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="User is not assigned to any tenant"
                )
        
        tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        if not tenant:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Tenant not found"
            )
        
        lower_file_name = (file.filename or "").lower()
        if not lower_file_name.endswith(('.xlsx', '.xls', '.xlsm', '.csv', '.pdf')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be an Excel (.xlsx, .xls), CSV, or supported PDF file"
            )
        
        logger.info(f"Reading file content: {file.filename}")
        file_content = await file.read()
        file_size = len(file_content)
        logger.info(f"File read complete: {file_size} bytes")
        
        file_ext = os.path.splitext(file.filename)[1]
        file_id = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_ext}")
        
        with open(file_path, "wb") as f:
            f.write(file_content)
        logger.info(f"File saved to: {file_path}")

        # ---------------------------------------------------------- #
        # Detect known workbook templates before generic parsing
        # ---------------------------------------------------------- #
        is_maturity_format = False
        is_asvs_checklist = False
        is_owasp_checklist = False
        is_ubl_audit_master = False
        is_pdpl = False
        is_doma = False
        is_mobile_app_security = False
        maturity_kind = None
        is_kpi_report = False
        is_dpia = False
        nca_reg_kind = None
        is_dcc_tool = False
        if lower_file_name.endswith(('.xlsx', '.xls', '.xlsm')):
            _wb_check = None
            try:
                _wb_check = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                is_dcc_tool = detect_dcc_tool_format(_wb_check)
                is_doma = (not is_dcc_tool) and detect_doma_format(_wb_check)
                # CREST cyber-maturity tools (CSIR/CTI/IT-SecOps/Incident), the
                # KPI report, the DPIA/PIA and the NCA registers are detected
                # BEFORE the generic maturity detector so they get their
                # dedicated parsers.
                if not is_doma:
                    maturity_kind = detect_maturity_kind(_wb_check)
                if not is_doma and not maturity_kind:
                    is_kpi_report = detect_kpi_report_format(_wb_check)
                if not is_doma and not maturity_kind and not is_kpi_report:
                    is_dpia = detect_dpia_format(_wb_check)
                if not is_doma and not maturity_kind and not is_kpi_report and not is_dpia:
                    nca_reg_kind = detect_nca_register_kind(_wb_check)
                _pre = is_dcc_tool or is_doma or bool(maturity_kind) or is_kpi_report or is_dpia or bool(nca_reg_kind)
                if not _pre:
                    is_maturity_format = detect_xlsx_maturity_format(_wb_check)
                if not _pre and not is_maturity_format:
                    is_pdpl = detect_pdpl_assessment_format(_wb_check)
                if not _pre and not is_maturity_format and not is_pdpl:
                    is_ubl_audit_master = detect_ubl_audit_master_tracking_format(_wb_check)
                if not _pre and not is_maturity_format and not is_pdpl and not is_ubl_audit_master:
                    is_mobile_app_security = detect_mobile_app_security_format(_wb_check)
                if not _pre and not is_maturity_format and not is_pdpl and not is_ubl_audit_master and not is_mobile_app_security:
                    is_asvs_checklist = detect_asvs_checklist_format(_wb_check)
                if not _pre and not is_maturity_format and not is_pdpl and not is_asvs_checklist and not is_ubl_audit_master and not is_mobile_app_security:
                    is_owasp_checklist = detect_owasp_v4_checklist_format(_wb_check)
            except Exception:
                pass
            finally:
                try:
                    if _wb_check:
                        _wb_check.close()
                except Exception:
                    pass

        # Guard: when a dedicated tab (ASVS, Mobile App Security, PDPL, …) opens
        # its own "Upload" button it passes expected_format. Reject a workbook
        # that doesn't match that tab so the wrong Excel can't land there.
        if expected_format and expected_format != "standard":
            detected = (
                "digital_ops_maturity" if is_doma else
                f"{maturity_kind}_maturity" if maturity_kind else
                "kpi_report" if is_kpi_report else
                "dpia_pia" if is_dpia else
                f"nca_{nca_reg_kind}_register" if nca_reg_kind else
                "nca_dcc_tool" if is_dcc_tool else
                "xlsx_maturity" if is_maturity_format else
                "pdpl_assessment_toolkit" if is_pdpl else
                "ubl_audit_master_tracking" if is_ubl_audit_master else
                "mobile_app_security" if is_mobile_app_security else
                "asvs_checklist" if is_asvs_checklist else
                "owasp_v4_testing_checklist" if is_owasp_checklist else
                "standard"
            )
            if detected != expected_format:
                if file_path and os.path.exists(file_path):
                    os.remove(file_path)
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Please upload the {_format_label(expected_format)} workbook here. "
                        f"The file you selected was recognised as {_format_label(detected)}."
                    ),
                )

        if is_maturity_format:
            logger.info("Detected XLSX Maturity Tool format – using maturity parser")
            xlsx_data = parse_xlsx_maturity_tool(file_content)

            # Extract framework name for the assessment name if not provided
            framework_name = xlsx_data.get("framework_name", "Maturity Assessment")

            # Build summary items from the details sheet for compatibility
            details = xlsx_data.get("sheets", {}).get("details", [])
            items_data = [
                {
                    "item_number": d.get("subcategory", "")[:50] if d.get("subcategory") else str(i + 1),
                    "area_domain": f"{d.get('function', '')} - {d.get('category', '')}".strip(" -"),
                    "control_description": d.get("subcategory", ""),
                    "compliance_status": "in_progress",
                    "remarks": (
                        f"Policy Maturity: {d['policy_maturity']}, Practice Maturity: {d['practice_maturity']}"
                        if d.get("policy_maturity") is not None else None
                    ),
                }
                for i, d in enumerate(details)
            ]
            logger.info(f"Maturity tool parsed: {len(items_data)} subcategories")

            parsed_due_date = None
            if due_date:
                try:
                    parsed_due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
                except Exception:
                    pass

            db_assessment = ComplianceAssessmentDocument(
                tenant_id=tenant_id,
                name=name or framework_name,
                assessment_type=assessment_type,
                source=source,
                file_name=file.filename,
                file_path=file_path,
                due_date=parsed_due_date,
                assessor=assessor,
                notes=notes,
                status="draft",
                created_by=current_user.id,
                total_items=len(items_data),
                assessment_format="xlsx_maturity",
                xlsx_data=xlsx_data,
            )
            db.add(db_assessment)
            db.flush()

            for idx, item_data in enumerate(items_data):
                db_item = ComplianceAssessmentDocumentItem(
                    assessment_id=db_assessment.id,
                    tenant_id=tenant_id,
                    item_number=item_data.get("item_number") or str(idx + 1),
                    area_domain=item_data.get("area_domain"),
                    control_description=item_data.get("control_description"),
                    compliance_status=item_data.get("compliance_status", "in_progress"),
                    remarks=item_data.get("remarks"),
                )
                db.add(db_item)

            db.flush()
            items_objs = db.query(ComplianceAssessmentDocumentItem).filter(
                ComplianceAssessmentDocumentItem.assessment_id == db_assessment.id
            ).all()
            stats = calculate_assessment_stats(items_objs)
            db_assessment.complied_count = stats["complied"]
            db_assessment.partially_complied_count = stats["partially_complied"]
            db_assessment.not_complied_count = stats["not_complied"]
            db_assessment.in_progress_count = stats["in_progress"]
            db_assessment.na_count = stats["na"]
            db_assessment.overall_score = stats["overall_score"]

            db.commit()
            db.refresh(db_assessment)
            logger.info(f"Maturity tool upload complete: ID={db_assessment.id}")

            csf_summary = xlsx_data.get("sheets", {}).get("csf_summary", {})
            overall_scores = csf_summary.get("overall", {})
            return {
                "id": db_assessment.id,
                "name": db_assessment.name,
                "assessment_type": db_assessment.assessment_type,
                "assessment_format": "xlsx_maturity",
                "source": db_assessment.source,
                "file_name": db_assessment.file_name,
                "status": db_assessment.status,
                "total_items": db_assessment.total_items,
                "overall_scores": overall_scores,
                "framework_name": xlsx_data.get("framework_name"),
                "message": f"Successfully uploaded maturity assessment with {len(details)} subcategories across {len(set(d.get('function','') for d in details))} functions",
            }
        
        parser_metadata = {"assessment_format": "standard", "columns_detected": []}
        if is_doma:
            logger.info("Detected Digital Operations Maturity (DOMA) format")
            items_data, parser_metadata = parse_doma_workbook(file_content)
        elif maturity_kind:
            logger.info(f"Detected CREST cyber-maturity tool: {maturity_kind}")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_cyber_maturity_workbook(wb, maturity_kind)
            finally:
                wb.close()
        elif is_kpi_report:
            logger.info("Detected Cyber Security KPI Report format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_kpi_report_workbook(wb)
            finally:
                wb.close()
        elif is_dpia:
            logger.info("Detected DPIA / PIA format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_dpia_workbook(wb)
            finally:
                wb.close()
        elif nca_reg_kind:
            logger.info(f"Detected NCA {nca_reg_kind} register format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_nca_register_workbook(wb, nca_reg_kind)
            finally:
                wb.close()
        elif is_dcc_tool:
            logger.info("Detected NCA DCC-1:2022 bilingual Excel tool")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_dcc_tool_workbook(wb)
            finally:
                wb.close()
        elif lower_file_name.endswith('.pdf'):
            logger.info("Detected PDF upload format")
            items_data, parser_metadata = parse_cis_windows_server_2012_r2_pdf(file_content, file.filename or "")
        elif is_pdpl:
            logger.info("Detected Saudi PDPL Assessment Toolkit format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_pdpl_assessment_workbook(wb)
            finally:
                wb.close()
        elif is_ubl_audit_master:
            logger.info("Detected UBL Audit Master Tracking format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_ubl_audit_master_tracking_workbook(wb)
            finally:
                wb.close()
        elif is_mobile_app_security:
            logger.info("Detected OWASP MASVS Mobile App Security checklist format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_mobile_app_security_workbook(wb)
            finally:
                wb.close()
        elif is_asvs_checklist:
            logger.info("Detected ASVS checklist format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_asvs_checklist_workbook(wb)
            finally:
                wb.close()
        elif is_owasp_checklist:
            logger.info("Detected OWASP v4 testing checklist format")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            try:
                items_data, parser_metadata = parse_owasp_v4_checklist_workbook(wb)
            finally:
                wb.close()
        else:
            logger.info("Parsing with generic Excel/CSV parser...")
            items_data, column_map = parse_excel_file(file_content, file.filename)
            parser_metadata = {
                "assessment_format": "standard",
                "columns_detected": list(column_map.keys()),
            }
        logger.info(
            f"Parsed {len(items_data)} items, format={parser_metadata.get('assessment_format')}, "
            f"columns={parser_metadata.get('columns_detected')}"
        )
        
        if not items_data:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid assessment items found in the file. Please check the column headers."
            )
        
        parsed_due_date = None
        if due_date:
            try:
                parsed_due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
            except:
                pass
        
        logger.info("Creating assessment document in database...")
        db_assessment = ComplianceAssessmentDocument(
            tenant_id=tenant_id,
            name=name,
            assessment_type=assessment_type,
            source=source,
            file_name=file.filename,
            file_path=file_path,
            due_date=parsed_due_date,
            assessor=assessor,
            notes=notes,
            status="draft",
            created_by=current_user.id,
            total_items=len(items_data),
            assessment_format=parser_metadata.get("assessment_format", "standard")
        )
        db.add(db_assessment)
        db.flush()
        logger.info(f"Assessment document created with ID: {db_assessment.id}")
        
        logger.info(f"Creating {len(items_data)} assessment items...")
        for idx, item_data in enumerate(items_data):
            db_item = ComplianceAssessmentDocumentItem(
                assessment_id=db_assessment.id,
                tenant_id=tenant_id,
                item_number=item_data.get("item_number") or str(idx + 1),
                area_domain=item_data.get("area_domain"),
                control_description=item_data.get("control_description"),
                compliance_status=item_data.get("compliance_status", "in_progress"),
                gaps_identified=item_data.get("gaps_identified"),
                proposed_solution=item_data.get("proposed_solution"),
                responsible_party=item_data.get("responsible_party"),
                timeline=item_data.get("timeline"),
                priority=item_data.get("priority"),
                evidence_reference=item_data.get("evidence_reference"),
                remarks=item_data.get("remarks"),
                maturity_score=item_data.get("maturity_score"),
                risk_rating=item_data.get("risk_rating"),
                subdomain_name=item_data.get("subdomain_name"),
            )
            db.add(db_item)

        db.flush()
        logger.info("Assessment items created, calculating stats...")
        
        items = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == db_assessment.id
        ).all()
        stats = calculate_assessment_stats(items)
        
        db_assessment.complied_count = stats["complied"]
        db_assessment.partially_complied_count = stats["partially_complied"]
        db_assessment.not_complied_count = stats["not_complied"]
        db_assessment.in_progress_count = stats["in_progress"]
        db_assessment.na_count = stats["na"]
        db_assessment.overall_score = stats["overall_score"]
        
        db.commit()
        db.refresh(db_assessment)
        logger.info(f"Assessment upload complete: ID={db_assessment.id}, items={db_assessment.total_items}")
        
        return {
            "id": db_assessment.id,
            "name": db_assessment.name,
            "assessment_type": db_assessment.assessment_type,
            "assessment_format": db_assessment.assessment_format or "standard",
            "source": db_assessment.source,
            "file_name": db_assessment.file_name,
            "status": db_assessment.status,
            "total_items": db_assessment.total_items,
            "complied_count": db_assessment.complied_count,
            "partially_complied_count": db_assessment.partially_complied_count,
            "not_complied_count": db_assessment.not_complied_count,
            "in_progress_count": db_assessment.in_progress_count,
            "na_count": db_assessment.na_count,
            "overall_score": db_assessment.overall_score,
            "columns_detected": parser_metadata.get("columns_detected", []),
            "detected_format": parser_metadata.get("detected_format", "standard"),
            "message": f"Successfully uploaded assessment with {len(items_data)} items"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Assessment upload failed: {str(e)}")
        logger.error(traceback.format_exc())
        db.rollback()
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process assessment file: {str(e)}"
        )


# ── Digital Operations Maturity (DOMA) ───────────────────────────────────────
# A multi-module maturity questionnaire: 9 capability modules (sheets), each row
# a Question grouped by a Capability, scored across 5 named maturity levels
# (Basic → Emerging → Advanced → Differentiated → Best-in-Class). We map each
# question to an assessment item: area_domain = module, subdomain_name =
# capability, control_description = question, and preserve the 5 level
# descriptions in `remarks` so nothing is lost. Assessors set maturity_score
# 1-5 (= the chosen level). Vendor branding is intentionally not stored.
_DOMA_MODULES = [
    'Data Analytics & Prescriptive', 'Digital Ops Strategy', 'Digital PD & PLM',
    'Integrated Planning', 'IT Architecture and Systems', 'Procurement 4.0',
    'Smart Factory', 'Smart Logistics', 'Smart Warehousing',
    'Transparency and Visibility',
]
_DOMA_LEVELS = ['Basic', 'Emerging', 'Advanced', 'Differentiated', 'Best In Class']


def detect_doma_format(wb) -> bool:
    try:
        sheets = set(wb.sheetnames)
    except Exception:
        return False
    return len(set(_DOMA_MODULES) & sheets) >= 4


def parse_doma_workbook(file_content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    items = []
    gseq = 0
    try:
        for name in _DOMA_MODULES:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            it = ws.iter_rows(values_only=True)
            hi = None
            for row in it:
                vals = [('' if x is None else str(x)).strip() for x in row]
                if 'Question' in vals and 'Basic' in vals:
                    hi = {v: i for i, v in enumerate(vals) if v}
                    break
            if not hi:
                continue
            qi = hi.get('Question'); capi = hi.get('Capability'); rapi = hi.get('Rapid DOMA')
            empties = 0
            for row in it:
                vals = [('' if x is None else str(x)).strip() for x in row]
                q = vals[qi] if (qi is not None and qi < len(vals)) else ''
                if not q:
                    empties += 1
                    if empties > 40:
                        break
                    continue
                empties = 0
                gseq += 1
                cap = vals[capi] if (capi is not None and capi < len(vals)) else ''
                rubric = []
                for lv in _DOMA_LEVELS:
                    idx = hi.get(lv)
                    if idx is not None and idx < len(vals) and vals[idx]:
                        rubric.append(f"{lv}: {vals[idx]}")
                rapid = bool(vals[rapi]) if (rapi is not None and rapi < len(vals) and vals[rapi]) else False
                remarks = ("[Rapid DOMA] " if rapid else "") + "  ||  ".join(rubric)
                items.append({
                    "item_number": str(gseq),
                    "area_domain": name,
                    "subdomain_name": cap or None,
                    "control_description": q,
                    "compliance_status": "in_progress",
                    "remarks": remarks or None,
                })
    finally:
        wb.close()
    return items, {"assessment_format": "digital_ops_maturity"}


def _parse_assessment_file(file_content: bytes, filename: str):
    """Detect the workbook template and parse it into (items_data, format,
    xlsx_data) — the same detection/parsers used by /upload, reused by the
    re-upload endpoint so an updated workbook refreshes an existing assessment."""
    lower = (filename or "").lower()
    is_maturity = is_pdpl = is_ubl = is_asvs = is_owasp = is_doma = is_mobile = is_kpi = is_dpia = False
    maturity_kind = None
    nca_reg_kind = None
    is_dcc_tool = False
    if lower.endswith(('.xlsx', '.xls', '.xlsm')):
        _wb = None
        try:
            _wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            is_dcc_tool = detect_dcc_tool_format(_wb)
            is_doma = (not is_dcc_tool) and detect_doma_format(_wb)
            if not is_doma:
                maturity_kind = detect_maturity_kind(_wb)
            if not is_doma and not maturity_kind:
                is_kpi = detect_kpi_report_format(_wb)
            if not is_doma and not maturity_kind and not is_kpi:
                is_dpia = detect_dpia_format(_wb)
            if not is_doma and not maturity_kind and not is_kpi and not is_dpia:
                nca_reg_kind = detect_nca_register_kind(_wb)
            _pre = is_dcc_tool or is_doma or bool(maturity_kind) or is_kpi or is_dpia or bool(nca_reg_kind)
            if not _pre:
                is_maturity = detect_xlsx_maturity_format(_wb)
            if not _pre and not is_maturity:
                is_pdpl = detect_pdpl_assessment_format(_wb)
            if not _pre and not (is_maturity or is_pdpl):
                is_ubl = detect_ubl_audit_master_tracking_format(_wb)
            if not _pre and not (is_maturity or is_pdpl or is_ubl):
                is_mobile = detect_mobile_app_security_format(_wb)
            if not _pre and not (is_maturity or is_pdpl or is_ubl or is_mobile):
                is_asvs = detect_asvs_checklist_format(_wb)
            if not _pre and not (is_maturity or is_pdpl or is_ubl or is_asvs or is_mobile):
                is_owasp = detect_owasp_v4_checklist_format(_wb)
        except Exception:
            pass
        finally:
            try:
                if _wb:
                    _wb.close()
            except Exception:
                pass

    if is_doma:
        items_data, meta = parse_doma_workbook(file_content)
        return items_data, meta.get("assessment_format", "digital_ops_maturity"), None

    if maturity_kind:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_cyber_maturity_workbook(wb, maturity_kind)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", f"{maturity_kind}_maturity"), None

    if is_kpi:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_kpi_report_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "kpi_report"), None

    if is_dpia:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_dpia_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "dpia_pia"), None

    if nca_reg_kind:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_nca_register_workbook(wb, nca_reg_kind)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", f"nca_{nca_reg_kind}_register"), None

    if is_dcc_tool:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_dcc_tool_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "nca_dcc_tool"), None

    if is_maturity:
        xlsx_data = parse_xlsx_maturity_tool(file_content)
        details = xlsx_data.get("sheets", {}).get("details", [])
        items_data = [
            {
                "item_number": d.get("subcategory", "")[:50] if d.get("subcategory") else str(i + 1),
                "area_domain": f"{d.get('function', '')} - {d.get('category', '')}".strip(" -"),
                "control_description": d.get("subcategory", ""),
                "compliance_status": "in_progress",
                "remarks": (
                    f"Policy Maturity: {d['policy_maturity']}, Practice Maturity: {d['practice_maturity']}"
                    if d.get("policy_maturity") is not None else None
                ),
            }
            for i, d in enumerate(details)
        ]
        return items_data, "xlsx_maturity", xlsx_data

    if lower.endswith('.pdf'):
        items_data, meta = parse_cis_windows_server_2012_r2_pdf(file_content, filename or "")
        return items_data, meta.get("assessment_format", "standard"), None
    if is_pdpl:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_pdpl_assessment_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "pdpl_assessment_toolkit"), None
    if is_ubl:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_ubl_audit_master_tracking_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "ubl_audit_master_tracking"), None
    if is_mobile:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_mobile_app_security_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "mobile_app_security"), None
    if is_asvs:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_asvs_checklist_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "asvs_checklist"), None
    if is_owasp:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        try:
            items_data, meta = parse_owasp_v4_checklist_workbook(wb)
        finally:
            wb.close()
        return items_data, meta.get("assessment_format", "owasp_v4_testing_checklist"), None

    items_data, _column_map = parse_excel_file(file_content, filename)
    return items_data, "standard", None


@router.post("/{assessment_id}/reupload")
async def reupload_assessment(
    assessment_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Refresh an EXISTING assessment from an updated workbook with the same
    structure. Items are matched by item_number and updated in place (new rows
    added) so linked evidence on unchanged controls is preserved."""
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
    ).first()
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")

    lower_file_name = (file.filename or "").lower()
    if not lower_file_name.endswith(('.xlsx', '.xls', '.xlsm', '.csv', '.pdf')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must be an Excel (.xlsx, .xls), CSV, or supported PDF file")

    file_path = None
    try:
        file_content = await file.read()
        file_ext = os.path.splitext(file.filename)[1]
        file_id = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_ext}")
        with open(file_path, "wb") as f:
            f.write(file_content)

        items_data, detected_format, xlsx_data = _parse_assessment_file(file_content, file.filename)
        if not items_data:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid assessment items found in the file. Please check the column headers.")

        # Guard: an assessment may only be refreshed with ITS OWN kind of
        # workbook. Each dedicated tab (ASVS, Mobile App Security, PDPL, …) has
        # its own upload button; this stops the wrong template overwriting a
        # typed assessment. A "standard" assessment accepts any file.
        current_fmt = (assessment.assessment_format or "").strip()
        if current_fmt and current_fmt != "standard" and detected_format != current_fmt:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"This is a {_format_label(current_fmt)} assessment — please upload the matching "
                    f"{_format_label(current_fmt)} workbook. The file you uploaded was recognised as "
                    f"{_format_label(detected_format)}."
                ),
            )

        existing = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == assessment.id
        ).all()
        by_number = {(it.item_number or "").strip(): it for it in existing}
        FIELDS = [
            "area_domain", "subdomain_name", "control_description", "compliance_status", "gaps_identified",
            "proposed_solution", "responsible_party", "timeline", "priority",
            "evidence_reference", "remarks", "maturity_score", "risk_rating",
        ]
        updated = 0
        added = 0
        for idx, item_data in enumerate(items_data):
            number = (item_data.get("item_number") or str(idx + 1)).strip()
            target = by_number.get(number)
            if target is None:
                target = ComplianceAssessmentDocumentItem(
                    assessment_id=assessment.id, tenant_id=assessment.tenant_id, item_number=number,
                )
                db.add(target)
                added += 1
            else:
                updated += 1
            for fld in FIELDS:
                if fld not in item_data:
                    continue
                val = item_data.get(fld)
                # Only overwrite when the workbook actually carries a value for this
                # field. A blank cell leaves the existing (possibly hand-entered)
                # value untouched — so re-uploading updated data overrides what the
                # file changed and preserves everything the file left blank.
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                setattr(target, fld, val)
            if not getattr(target, "compliance_status", None):
                target.compliance_status = item_data.get("compliance_status", "in_progress")

        db.flush()
        items = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == assessment.id
        ).all()
        stats = calculate_assessment_stats(items)
        assessment.complied_count = stats["complied"]
        assessment.partially_complied_count = stats["partially_complied"]
        assessment.not_complied_count = stats["not_complied"]
        assessment.in_progress_count = stats["in_progress"]
        assessment.na_count = stats["na"]
        assessment.overall_score = stats["overall_score"]
        assessment.total_items = len(items)
        assessment.file_name = file.filename
        assessment.file_path = file_path
        if xlsx_data is not None and hasattr(assessment, "xlsx_data"):
            assessment.xlsx_data = xlsx_data

        db.commit()
        db.refresh(assessment)
        return {
            "id": assessment.id,
            "name": assessment.name,
            "assessment_format": assessment.assessment_format or "standard",
            "total_items": assessment.total_items,
            "overall_score": assessment.overall_score,
            "updated_count": updated,
            "added_count": added,
            "message": f"Refreshed assessment from {file.filename}: {updated} updated, {added} added",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Assessment re-upload failed: {str(e)}")
        logger.error(traceback.format_exc())
        db.rollback()
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception:
                pass
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to refresh assessment: {str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# Remediation Plan — the client-facing action log of assessment gaps.
# A gap is any item assessed below the compliance bar (PDPL rule: maturity < 3
# => not_complied / partially_complied). Each gap carries an independent
# open -> in_progress -> closed status tracked to closure.
# ─────────────────────────────────────────────────────────────────────────────

# compliance_status values that represent an open gap needing remediation.
_REMEDIATION_GAP_STATUSES = ["not_complied", "partially_complied"]
_REMEDIATION_STATUSES = {"open", "in_progress", "closed"}


def _parse_remarks_kv(remarks: Optional[str]) -> dict:
    """Pull the 'Key: value | Key: value' pairs the PDPL parser stores in
    remarks (e.g. 'PDPL Ref: Art. 31 | Risk: High') back into a dict."""
    out: dict = {}
    if not remarks:
        return out
    for part in str(remarks).split("|"):
        if ":" in part:
            key, _, val = part.partition(":")
            out[key.strip().lower()] = val.strip()
    return out


@router.get("/remediation-plan")
def get_remediation_plan(
    tenant_id: Optional[int] = None,
    assessment_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Return every open gap across the tenant's assessments as a remediation
    action log. Spans all formats; for PDPL these are the controls scored < 3."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"items": [], "total": 0, "summary": {}}

    q = (
        db.query(ComplianceAssessmentDocumentItem, ComplianceAssessmentDocument)
        .join(
            ComplianceAssessmentDocument,
            ComplianceAssessmentDocumentItem.assessment_id == ComplianceAssessmentDocument.id,
        )
        .filter(
            ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants),
            ComplianceAssessmentDocumentItem.compliance_status.in_(_REMEDIATION_GAP_STATUSES),
        )
    )
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        q = q.filter(ComplianceAssessmentDocumentItem.tenant_id == tenant_id)
    if assessment_id:
        q = q.filter(ComplianceAssessmentDocumentItem.assessment_id == assessment_id)

    rows = q.order_by(
        ComplianceAssessmentDocument.id, ComplianceAssessmentDocumentItem.id
    ).all()

    items = []
    counts = {"open": 0, "in_progress": 0, "closed": 0}
    for item, doc in rows:
        kv = _parse_remarks_kv(item.remarks)
        rem_status = (item.remediation_status or "open").lower()
        counts[rem_status] = counts.get(rem_status, 0) + 1
        items.append({
            "id": item.id,
            "assessment_id": doc.id,
            "assessment_name": doc.name,
            "assessment_format": doc.assessment_format,
            "control_id": item.item_number,
            "domain": item.area_domain,
            "pdpl_ref": kv.get("pdpl ref"),
            "risk": item.risk_rating or kv.get("risk"),
            "gap": item.gaps_identified,
            "remediation_action": item.proposed_solution,
            "priority": item.priority,
            "owner": item.responsible_party,
            "target_date": item.timeline,
            "compliance_status": item.compliance_status,
            "remediation_status": rem_status,
        })

    total = len(items)
    closed = counts.get("closed", 0)
    summary = {
        "total": total,
        "open": counts.get("open", 0),
        "in_progress": counts.get("in_progress", 0),
        "closed": closed,
        "closure_pct": round((closed / total) * 100, 1) if total else 0.0,
    }
    return {"items": items, "total": total, "summary": summary}


class RemediationStatusUpdate(BaseModel):
    remediation_status: str


@router.patch("/remediation-items/{item_id}")
def update_remediation_item(
    item_id: int,
    body: RemediationStatusUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Update a single gap's remediation status (open / in_progress / closed)."""
    user_tenants = get_user_tenants(current_user, db) or [-1]
    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants),
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Remediation item not found")
    new_status = (body.remediation_status or "").strip().lower()
    if new_status not in _REMEDIATION_STATUSES:
        raise HTTPException(
            status_code=400,
            detail="Invalid status. Use open, in_progress, or closed.",
        )
    item.remediation_status = new_status
    db.commit()
    db.refresh(item)
    return {"id": item.id, "remediation_status": item.remediation_status}


@router.get("/evidence/{evidence_id}/controls")
def get_evidence_linked_controls(
    evidence_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Reverse lookup: which assessment controls is this evidence linked to?

    Auto-populated from AssessmentItemEvidence — the evidence-library side never
    links manually; it just reflects links made when attaching evidence to a
    control (e.g. on the PDPL Assessment page).
    """
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"evidence_id": evidence_id, "total": 0, "controls": []}
    rows = (
        db.query(AssessmentItemEvidence, ComplianceAssessmentDocumentItem, ComplianceAssessmentDocument)
        .join(ComplianceAssessmentDocumentItem, AssessmentItemEvidence.assessment_item_id == ComplianceAssessmentDocumentItem.id)
        .join(ComplianceAssessmentDocument, ComplianceAssessmentDocumentItem.assessment_id == ComplianceAssessmentDocument.id)
        .filter(
            AssessmentItemEvidence.evidence_id == evidence_id,
            ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
        )
        .order_by(ComplianceAssessmentDocument.id, ComplianceAssessmentDocumentItem.id)
        .all()
    )
    controls = [
        {
            "link_id": link.id,
            "assessment_id": doc.id,
            "assessment_name": doc.name,
            "assessment_format": doc.assessment_format,
            "item_id": it.id,
            "control_id": it.item_number,
            "domain": it.area_domain,
            "control_description": it.control_description,
            "compliance_status": it.compliance_status,
        }
        for link, it, doc in rows
    ]
    return {"evidence_id": evidence_id, "total": len(controls), "controls": controls}


@router.get("")
def list_assessments(
    tenant_id: Optional[int] = None,
    assessment_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    source: Optional[str] = None,
    assessment_format: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"assessments": [], "total": 0, "summary": {}}

    query = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
    )
    if assessment_format:
        # Explicit format request (e.g. the dedicated PDPL Assessment tab):
        # return ONLY that format and bypass the default hide-list, since the
        # caller is a dedicated surface that owns that format.
        query = query.filter(ComplianceAssessmentDocument.assessment_format == assessment_format)
    else:
        # Default list hides formats that have their own top-level tab: the NCA
        # singleton container and PDPL toolkit uploads. They are surfaced only
        # on their dedicated tabs, never in the generic Assessment list.
        query = query.filter(
            (ComplianceAssessmentDocument.assessment_format.notin_(["nca_container", "pdpl_assessment_toolkit"]))
            | (ComplianceAssessmentDocument.assessment_format.is_(None))
        )

    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ComplianceAssessmentDocument.tenant_id == tenant_id)
    if assessment_type:
        query = query.filter(ComplianceAssessmentDocument.assessment_type == assessment_type)
    if status_filter:
        query = query.filter(ComplianceAssessmentDocument.status == status_filter)
    if source:
        query = query.filter(ComplianceAssessmentDocument.source == source)
    
    total = query.count()
    summary_assessments = query.all()
    assessments = query.order_by(ComplianceAssessmentDocument.created_at.desc()).offset(skip).limit(limit).all()

    total_items = sum(a.total_items or 0 for a in summary_assessments)
    total_complied = sum(a.complied_count or 0 for a in summary_assessments)
    total_partial = sum(a.partially_complied_count or 0 for a in summary_assessments)
    total_not_complied = sum(a.not_complied_count or 0 for a in summary_assessments)
    total_in_progress = sum(a.in_progress_count or 0 for a in summary_assessments)
    total_na = sum(a.na_count or 0 for a in summary_assessments)

    by_type: dict = {}
    by_status: dict = {}
    by_format: dict = {}
    scored = [float(a.overall_score) for a in summary_assessments if a.overall_score is not None]
    average_overall_score = round(sum(scored) / len(scored), 2) if scored else None

    overdue_count = 0
    due_soon_count = 0
    now = datetime.utcnow()
    for a in summary_assessments:
        t = a.assessment_type or "other"
        s = a.status or "draft"
        fmt = getattr(a, "assessment_format", "standard") or "standard"
        by_type[t] = by_type.get(t, 0) + 1
        by_status[s] = by_status.get(s, 0) + 1
        by_format[fmt] = by_format.get(fmt, 0) + 1
        if a.due_date:
            try:
                diff_days = (a.due_date - now).days
                if diff_days < 0:
                    overdue_count += 1
                elif diff_days <= 30:
                    due_soon_count += 1
            except Exception:
                continue
    
    return {
        "assessments": [
            {
                "id": a.id,
                "name": a.name,
                "assessment_type": a.assessment_type,
                "assessment_format": getattr(a, "assessment_format", "standard") or "standard",
                "source": a.source,
                "file_name": a.file_name,
                "status": a.status,
                "due_date": a.due_date.isoformat() if a.due_date else None,
                "assessor": a.assessor,
                "overall_score": a.overall_score,
                "total_items": a.total_items,
                "complied_count": a.complied_count,
                "partially_complied_count": a.partially_complied_count,
                "not_complied_count": a.not_complied_count,
                "in_progress_count": a.in_progress_count,
                "na_count": a.na_count,
                "created_at": a.created_at.isoformat(),
                "updated_at": a.updated_at.isoformat() if a.updated_at else None
            }
            for a in assessments
        ],
        "total": total,
        "summary": {
            "total_assessments": total,
            "total_items": total_items,
            "total_complied": total_complied,
            "total_partial": total_partial,
            "total_not_complied": total_not_complied,
            "total_in_progress": total_in_progress,
            "total_na": total_na,
            "average_overall_score": average_overall_score,
            "overdue_count": overdue_count,
            "due_soon_count": due_soon_count,
            "by_type": by_type,
            "by_status": by_status,
            "by_format": by_format,
        }
    }


@router.get("/workflows")
def list_workflows(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    workflows = db.query(AssessmentEvidenceApprovalWorkflow).filter(
        AssessmentEvidenceApprovalWorkflow.tenant_id.in_(user_tenants)
    ).order_by(AssessmentEvidenceApprovalWorkflow.created_at.desc()).all()
    
    return {
        "workflows": [
            {
                "id": w.id,
                "name": w.name,
                "description": w.description,
                "is_default": w.is_default,
                "is_active": w.is_active,
                "created_at": w.created_at.isoformat() if w.created_at else None,
                "tier_count": len(w.tiers) if w.tiers else 0
            }
            for w in workflows
        ]
    }


@router.post("/workflows")
def create_workflow(
    request: WorkflowCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User not assigned to any tenant")
    
    if request.is_default:
        db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.tenant_id == tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_default == True
        ).update({"is_default": False})
    
    workflow = AssessmentEvidenceApprovalWorkflow(
        tenant_id=tenant_id,
        name=request.name,
        description=request.description,
        is_default=request.is_default,
        is_active=True,
        created_by=current_user.id
    )
    db.add(workflow)
    db.flush()
    
    for tier_data in request.tiers:
        tier = AssessmentEvidenceApprovalTier(
            workflow_id=workflow.id,
            tier_order=tier_data.get("tier_order", 1),
            tier_name=tier_data.get("tier_name", "Reviewer"),
            approver_type=tier_data.get("approver_type", "user"),
            approver_role_id=tier_data.get("approver_role_id"),
            approver_user_id=tier_data.get("approver_user_id"),
            can_delegate=tier_data.get("can_delegate", True),
            auto_approve_days=tier_data.get("auto_approve_days")
        )
        db.add(tier)
    
    db.commit()
    db.refresh(workflow)
    
    return {
        "id": workflow.id,
        "name": workflow.name,
        "description": workflow.description,
        "is_default": workflow.is_default,
        "is_active": workflow.is_active,
        "tiers": [
            {
                "id": t.id,
                "tier_order": t.tier_order,
                "tier_name": t.tier_name,
                "approver_type": t.approver_type
            }
            for t in workflow.tiers
        ],
        "message": "Workflow created successfully"
    }


@router.get("/workflows/{workflow_id}")
def get_workflow(
    workflow_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    workflow = db.query(AssessmentEvidenceApprovalWorkflow).options(
        joinedload(AssessmentEvidenceApprovalWorkflow.tiers)
    ).filter(
        AssessmentEvidenceApprovalWorkflow.id == workflow_id,
        AssessmentEvidenceApprovalWorkflow.tenant_id.in_(user_tenants)
    ).first()
    
    if not workflow:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workflow not found")
    
    return {
        "id": workflow.id,
        "name": workflow.name,
        "description": workflow.description,
        "is_default": workflow.is_default,
        "is_active": workflow.is_active,
        "created_at": workflow.created_at.isoformat() if workflow.created_at else None,
        "tiers": sorted([
            {
                "id": t.id,
                "tier_order": t.tier_order,
                "tier_name": t.tier_name,
                "approver_type": t.approver_type,
                "approver_role_id": t.approver_role_id,
                "approver_user_id": t.approver_user_id,
                "can_delegate": t.can_delegate,
                "auto_approve_days": t.auto_approve_days
            }
            for t in workflow.tiers
        ], key=lambda x: x["tier_order"])
    }


@router.put("/workflows/{workflow_id}")
def update_workflow(
    workflow_id: int,
    request: WorkflowUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
        AssessmentEvidenceApprovalWorkflow.id == workflow_id,
        AssessmentEvidenceApprovalWorkflow.tenant_id.in_(user_tenants)
    ).first()
    
    if not workflow:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workflow not found")
    
    if request.is_default:
        db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.tenant_id == workflow.tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_default == True,
            AssessmentEvidenceApprovalWorkflow.id != workflow_id
        ).update({"is_default": False})
    
    if request.name is not None:
        workflow.name = request.name
    if request.description is not None:
        workflow.description = request.description
    if request.is_default is not None:
        workflow.is_default = request.is_default
    if request.is_active is not None:
        workflow.is_active = request.is_active
    
    if request.tiers is not None:
        db.query(AssessmentEvidenceApprovalTier).filter(
            AssessmentEvidenceApprovalTier.workflow_id == workflow_id
        ).delete()
        
        for tier_data in request.tiers:
            tier = AssessmentEvidenceApprovalTier(
                workflow_id=workflow.id,
                tier_order=tier_data.get("tier_order", 1),
                tier_name=tier_data.get("tier_name", "Reviewer"),
                approver_type=tier_data.get("approver_type", "user"),
                approver_role_id=tier_data.get("approver_role_id"),
                approver_user_id=tier_data.get("approver_user_id"),
                can_delegate=tier_data.get("can_delegate", True),
                auto_approve_days=tier_data.get("auto_approve_days")
            )
            db.add(tier)
    
    db.commit()
    db.refresh(workflow)
    
    return {
        "id": workflow.id,
        "name": workflow.name,
        "description": workflow.description,
        "is_default": workflow.is_default,
        "is_active": workflow.is_active,
        "message": "Workflow updated successfully"
    }


@router.delete("/workflows/{workflow_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_workflow(
    workflow_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
        AssessmentEvidenceApprovalWorkflow.id == workflow_id,
        AssessmentEvidenceApprovalWorkflow.tenant_id.in_(user_tenants)
    ).first()
    
    if not workflow:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workflow not found")
    
    linked_evidence = db.query(AssessmentItemEvidence).filter(
        AssessmentItemEvidence.workflow_id == workflow_id
    ).count()
    
    if linked_evidence > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete workflow. {linked_evidence} evidence items are using this workflow."
        )
    
    db.delete(workflow)
    db.commit()
    return None


@router.get("/pending-approvals")
def get_pending_approvals(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    pending_items = db.query(AssessmentItemEvidence).options(
        joinedload(AssessmentItemEvidence.evidence),
        joinedload(AssessmentItemEvidence.assessment_item).joinedload(ComplianceAssessmentDocumentItem.assessment),
        joinedload(AssessmentItemEvidence.workflow).joinedload(AssessmentEvidenceApprovalWorkflow.tiers)
    ).filter(
        AssessmentItemEvidence.tenant_id.in_(user_tenants),
        AssessmentItemEvidence.status.in_(["pending_review", "pending_tier_1", "pending_tier_2", "pending_tier_3"])
    ).all()
    
    result = []
    for item in pending_items:
        can_approve = False
        if item.workflow:
            current_tier_obj = None
            for tier in item.workflow.tiers:
                if tier.tier_order == item.current_tier:
                    current_tier_obj = tier
                    break
            
            if current_tier_obj:
                if current_tier_obj.approver_type == "user" and current_tier_obj.approver_user_id == current_user.id:
                    can_approve = True
        
        result.append({
            "id": item.id,
            "evidence_id": item.evidence_id,
            "evidence_name": item.evidence.name if item.evidence else None,
            "evidence_file_name": item.evidence.file_name if item.evidence else None,
            "assessment_id": item.assessment_item.assessment_id if item.assessment_item else None,
            "assessment_name": item.assessment_item.assessment.name if item.assessment_item and item.assessment_item.assessment else None,
            "item_number": item.assessment_item.item_number if item.assessment_item else None,
            "control_description": item.assessment_item.control_description if item.assessment_item else None,
            "current_tier": item.current_tier,
            "status": item.status,
            "workflow_name": item.workflow.name if item.workflow else None,
            "can_approve": can_approve,
            "submitted_at": item.submitted_at.isoformat() if item.submitted_at else None
        })
    
    return {
        "pending_approvals": result,
        "total": len(result)
    }


@router.post("/evidence/{evidence_link_id}/approval")
def perform_approval_action(
    evidence_link_id: int,
    request: ApprovalActionRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    evidence_link = db.query(AssessmentItemEvidence).options(
        joinedload(AssessmentItemEvidence.workflow).joinedload(AssessmentEvidenceApprovalWorkflow.tiers),
        joinedload(AssessmentItemEvidence.evidence)
    ).filter(
        AssessmentItemEvidence.id == evidence_link_id,
        AssessmentItemEvidence.tenant_id.in_(user_tenants)
    ).first()
    
    if not evidence_link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence link not found")
    
    action = request.action.lower()
    
    if action == "submit":
        evidence_link.status = "pending_review"
        evidence_link.current_tier = 1
        evidence_link.submitted_by = current_user.id
        evidence_link.submitted_at = datetime.utcnow()
        
        history = AssessmentEvidenceApprovalHistory(
            assessment_item_evidence_id=evidence_link.id,
            tier_id=None,
            action="submitted",
            tier_number=0,
            performed_by=current_user.id,
            comments=request.comments
        )
        db.add(history)
        
    elif action == "approve":
        current_tier = evidence_link.current_tier or 1
        workflow = evidence_link.workflow
        
        tier_obj = None
        if workflow:
            tier_obj = db.query(AssessmentEvidenceApprovalTier).filter(
                AssessmentEvidenceApprovalTier.workflow_id == workflow.id,
                AssessmentEvidenceApprovalTier.tier_order == current_tier
            ).first()
            
            max_tier = db.query(func.max(AssessmentEvidenceApprovalTier.tier_order)).filter(
                AssessmentEvidenceApprovalTier.workflow_id == workflow.id
            ).scalar() or 1
            
            if current_tier < max_tier:
                evidence_link.current_tier = current_tier + 1
                evidence_link.status = f"pending_tier_{current_tier + 1}"
            else:
                evidence_link.status = "approved"
                if evidence_link.evidence:
                    evidence_link.evidence.status = "approved"
                    evidence_link.evidence.approved_at = datetime.utcnow()
        else:
            evidence_link.status = "approved"
            if evidence_link.evidence:
                evidence_link.evidence.status = "approved"
                evidence_link.evidence.approved_at = datetime.utcnow()
        
        history = AssessmentEvidenceApprovalHistory(
            assessment_item_evidence_id=evidence_link.id,
            tier_id=tier_obj.id if tier_obj else None,
            action="approved",
            tier_number=current_tier,
            performed_by=current_user.id,
            comments=request.comments
        )
        db.add(history)
        
    elif action == "reject":
        current_tier = evidence_link.current_tier or 1
        workflow = evidence_link.workflow
        
        tier_obj = None
        if workflow:
            tier_obj = db.query(AssessmentEvidenceApprovalTier).filter(
                AssessmentEvidenceApprovalTier.workflow_id == workflow.id,
                AssessmentEvidenceApprovalTier.tier_order == current_tier
            ).first()
        
        evidence_link.status = "rejected"
        if evidence_link.evidence:
            evidence_link.evidence.status = "rejected"
        
        history = AssessmentEvidenceApprovalHistory(
            assessment_item_evidence_id=evidence_link.id,
            tier_id=tier_obj.id if tier_obj else None,
            action="rejected",
            tier_number=current_tier,
            performed_by=current_user.id,
            comments=request.comments
        )
        db.add(history)
        
    elif action == "return":
        current_tier = evidence_link.current_tier or 1
        workflow = evidence_link.workflow
        
        tier_obj = None
        if workflow:
            tier_obj = db.query(AssessmentEvidenceApprovalTier).filter(
                AssessmentEvidenceApprovalTier.workflow_id == workflow.id,
                AssessmentEvidenceApprovalTier.tier_order == current_tier
            ).first()
        
        evidence_link.status = "returned"
        if evidence_link.evidence:
            evidence_link.evidence.status = "draft"
        
        history = AssessmentEvidenceApprovalHistory(
            assessment_item_evidence_id=evidence_link.id,
            tier_id=tier_obj.id if tier_obj else None,
            action="returned",
            tier_number=current_tier,
            performed_by=current_user.id,
            comments=request.comments,
            delegated_to=request.delegated_to
        )
        db.add(history)
        
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid action. Valid actions: submit, approve, reject, return"
        )
    
    evidence_link.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(evidence_link)
    
    return {
        "id": evidence_link.id,
        "status": evidence_link.status,
        "current_tier": evidence_link.current_tier,
        "action_performed": action,
        "message": f"Evidence {action} successfully"
    }


@router.get("/evidence/{evidence_link_id}/approval-history")
def get_approval_history(
    evidence_link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    evidence_link = db.query(AssessmentItemEvidence).filter(
        AssessmentItemEvidence.id == evidence_link_id,
        AssessmentItemEvidence.tenant_id.in_(user_tenants)
    ).first()
    
    if not evidence_link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence link not found")
    
    history = db.query(AssessmentEvidenceApprovalHistory).options(
        joinedload(AssessmentEvidenceApprovalHistory.performer),
        joinedload(AssessmentEvidenceApprovalHistory.tier)
    ).filter(
        AssessmentEvidenceApprovalHistory.assessment_item_evidence_id == evidence_link_id
    ).order_by(AssessmentEvidenceApprovalHistory.performed_at.desc()).all()
    
    return {
        "evidence_link_id": evidence_link_id,
        "current_status": evidence_link.status,
        "current_tier": evidence_link.current_tier,
        "history": [
            {
                "id": h.id,
                "action": h.action,
                "tier_number": h.tier_number,
                "tier_name": h.tier.tier_name if h.tier else None,
                "performed_by": h.performer.display_name or h.performer.username if h.performer else None,
                "performed_by_id": h.performed_by,
                "comments": h.comments,
                "delegated_to": h.delegated_to,
                "performed_at": h.performed_at.isoformat() if h.performed_at else None
            }
            for h in history
        ]
    }


@router.get("/{assessment_id}/xlsx-data")
def get_assessment_xlsx_data(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Return the full parsed multi-sheet JSON for an xlsx_maturity assessment."""
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    if getattr(assessment, "assessment_format", "standard") != "xlsx_maturity":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not in xlsx_maturity format")
    
    raw = getattr(assessment, "xlsx_data", None)
    if raw is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No xlsx data stored for this assessment")
    
    import json as _json
    data = raw if isinstance(raw, dict) else _json.loads(raw)

    # Backward-compatibility: re-parse older uploads if summary/maturity sections are missing.
    sheets = data.get("sheets", {}) if isinstance(data, dict) else {}
    csf_categories = sheets.get("csf_summary", {}).get("categories", [])
    maturity_section = sheets.get("maturity_levels", {})
    maturity_rows = maturity_section.get("rows", []) if isinstance(maturity_section, dict) else maturity_section

    needs_reparse = (len(csf_categories) == 0) or (len(maturity_rows) == 0)
    if needs_reparse and assessment.file_path and os.path.exists(assessment.file_path):
        try:
            with open(assessment.file_path, "rb") as fh:
                reparsed = parse_xlsx_maturity_tool(fh.read())
            data = reparsed
            assessment.xlsx_data = reparsed
            db.commit()
        except Exception:
            # Keep serving stored data if re-parse fails.
            pass

    return data


@router.put("/{assessment_id}/xlsx-data")
def update_assessment_xlsx_data(
    assessment_id: int,
    request: XlsxScoreUpdateRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()

    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    if getattr(assessment, "assessment_format", "standard") != "xlsx_maturity":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not in xlsx_maturity format")

    raw = getattr(assessment, "xlsx_data", None)
    if raw is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No xlsx data stored for this assessment")

    data = raw if isinstance(raw, dict) else json.loads(raw)
    sheets = data.setdefault("sheets", {})

    if request.sheet == "details":
        details = sheets.get("details") or []
        if request.row_index < 0 or request.row_index >= len(details):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid details row index")

        row = details[request.row_index]
        if request.policy_maturity is not None:
            row["policy_maturity"] = _safe_float(request.policy_maturity)
        if request.practice_maturity is not None:
            row["practice_maturity"] = _safe_float(request.practice_maturity)

        # Free-form per-row metadata. Read with `__contains__` against the
        # request's set fields so that an explicit empty string clears the
        # field but an absent key leaves the existing value alone.
        sent_fields = request.model_fields_set if hasattr(request, "model_fields_set") else set()
        if "remarks" in sent_fields:
            row["remarks"] = (request.remarks or "").strip() or None
        if "assigned_to_id" in sent_fields:
            row["assigned_to_id"] = request.assigned_to_id
        if "assigned_to_name" in sent_fields:
            row["assigned_to_name"] = (request.assigned_to_name or "").strip() or None
        if "due_date" in sent_fields:
            row["due_date"] = (request.due_date or "").strip() or None
        if "gaps_identified" in sent_fields:
            row["gaps_identified"] = (request.gaps_identified or "").strip() or None
        if "proposed_solution" in sent_fields:
            row["proposed_solution"] = (request.proposed_solution or "").strip() or None
        if "function" in sent_fields:
            row["function"] = (request.function or "").strip()
        if "category" in sent_fields:
            row["category"] = (request.category or "").strip()
        if "references" in sent_fields:
            cleaned_refs = [r.strip() for r in (request.references or []) if r and r.strip()]
            row["references"] = cleaned_refs

        _recalculate_csf_summary_from_details(data)

    elif request.sheet == "csf_summary":
        summary = sheets.get("csf_summary") or {}
        categories = summary.get("categories") or []
        if request.row_index < 0 or request.row_index >= len(categories):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid summary row index")

        row = categories[request.row_index]
        if request.target_score is not None:
            row["target_score"] = _safe_float(request.target_score)
        if request.policy_score is not None:
            row["policy_score"] = _safe_float(request.policy_score)
        if request.practice_score is not None:
            row["practice_score"] = _safe_float(request.practice_score)

        _recalculate_csf_summary_overall(data)

    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported xlsx sheet update target")

    assessment.xlsx_data = data
    flag_modified(assessment, "xlsx_data")
    assessment.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(assessment)

    return data


@router.post("/{assessment_id}/xlsx-data/details/{row_index}/ai-recommendation")
def generate_xlsx_detail_ai_recommendation(
    assessment_id: int,
    row_index: int,
    force: bool = Query(False, description="Regenerate recommendation even if already present"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()

    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    if getattr(assessment, "assessment_format", "standard") != "xlsx_maturity":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not in xlsx_maturity format")

    raw = getattr(assessment, "xlsx_data", None)
    if raw is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No xlsx data stored for this assessment")

    data = raw if isinstance(raw, dict) else json.loads(raw)
    sheets = data.get("sheets") or {}
    details = sheets.get("details") or []
    if not isinstance(details, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid details structure in xlsx data")
    if row_index < 0 or row_index >= len(details):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid details row index")

    row = details[row_index]
    if not isinstance(row, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid details row data")

    existing = row.get("recommended_evidence")
    if not force and isinstance(existing, list) and existing:
        return {
            "assessment_id": assessment_id,
            "row_index": row_index,
            "recommendations": existing[:5],
            "generated_at": row.get("recommended_evidence_generated_at"),
            "cached": True,
        }

    framework_name = str(data.get("framework_name") or "NIST CSF").strip() or "NIST CSF"
    payload = [{
        "row_index": row_index,
        "function": str(row.get("function") or ""),
        "category": str(row.get("category") or ""),
        "subcategory": str(row.get("subcategory") or ""),
        "references": [str(r).strip() for r in (row.get("references") or []) if str(r).strip()][:4],
        "policy_maturity": row.get("policy_maturity"),
        "practice_maturity": row.get("practice_maturity"),
    }]

    ai_map = _generate_detail_recommendations_batch_ai(framework_name, payload)
    recommendations = ai_map.get(row_index) or _fallback_detail_recommendations(row, framework_name)
    recommendations = recommendations[:5]
    generated_at = datetime.utcnow().isoformat()

    row["recommended_evidence"] = recommendations
    row["recommended_evidence_generated_at"] = generated_at

    assessment.xlsx_data = data
    flag_modified(assessment, "xlsx_data")
    assessment.updated_at = datetime.utcnow()
    db.commit()

    return {
        "assessment_id": assessment_id,
        "row_index": row_index,
        "recommendations": recommendations,
        "generated_at": generated_at,
        "cached": False,
    }


# ── SLA policy (tenant-level) ────────────────────────────────────────────────
# Days allowed per priority tier + the default "due soon" horizon. Drives the
# dynamic-SLA closure board: a point's target date = raised date + tier days
# (unless it carries an explicit target_date). One row per tenant. Registered
# BEFORE the "/{assessment_id}" route so the static path isn't shadowed.
_SLA_DEFAULTS = {
    "critical_days": 30, "high_days": 60, "medium_days": 90, "low_days": 180, "due_soon_days": 30,
    "score_closed_ontime": 100, "score_closed_late": 70, "score_on_track": 40,
    "score_due_soon": 20, "score_overdue": 0, "score_no_date": 30,
}


def _sla_policy_dict(p: "ComplianceSlaPolicy") -> dict:
    return {
        "critical_days": p.critical_days,
        "high_days": p.high_days,
        "medium_days": p.medium_days,
        "low_days": p.low_days,
        "due_soon_days": p.due_soon_days,
        "score_closed_ontime": getattr(p, "score_closed_ontime", 100),
        "score_closed_late": getattr(p, "score_closed_late", 70),
        "score_on_track": getattr(p, "score_on_track", 40),
        "score_due_soon": getattr(p, "score_due_soon", 20),
        "score_overdue": getattr(p, "score_overdue", 0),
        "score_no_date": getattr(p, "score_no_date", 30),
    }


@router.get("/sla-policy")
def get_sla_policy(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    p = db.query(ComplianceSlaPolicy).filter(ComplianceSlaPolicy.tenant_id == tenant_id).first()
    return _sla_policy_dict(p) if p else dict(_SLA_DEFAULTS)


@router.put("/sla-policy")
def update_sla_policy(
    critical_days: Optional[int] = None,
    high_days: Optional[int] = None,
    medium_days: Optional[int] = None,
    low_days: Optional[int] = None,
    due_soon_days: Optional[int] = None,
    score_closed_ontime: Optional[int] = None,
    score_closed_late: Optional[int] = None,
    score_on_track: Optional[int] = None,
    score_due_soon: Optional[int] = None,
    score_overdue: Optional[int] = None,
    score_no_date: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    p = db.query(ComplianceSlaPolicy).filter(ComplianceSlaPolicy.tenant_id == tenant_id).first()
    if not p:
        p = ComplianceSlaPolicy(tenant_id=tenant_id, **_SLA_DEFAULTS)
        db.add(p)
    # Day fields must be positive; score weights may be 0-100 (0 is valid).
    for field, val in (
        ("critical_days", critical_days), ("high_days", high_days),
        ("medium_days", medium_days), ("low_days", low_days),
        ("due_soon_days", due_soon_days),
    ):
        if val is not None and val > 0:
            setattr(p, field, val)
    for field, val in (
        ("score_closed_ontime", score_closed_ontime), ("score_closed_late", score_closed_late),
        ("score_on_track", score_on_track), ("score_due_soon", score_due_soon),
        ("score_overdue", score_overdue), ("score_no_date", score_no_date),
    ):
        if val is not None and 0 <= val <= 100:
            setattr(p, field, val)
    db.commit()
    db.refresh(p)
    return _sla_policy_dict(p)


@router.get("/points")
def list_all_points(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Flat list of every assessment point across the tenant, with the fields
    the dynamic-SLA closure board needs. One query (join), no per-assessment
    round-trips. Each point keeps its own dates — the board never blends them."""
    from grc.modules.assessments.scoring import EXCLUDED_FORMATS
    user_tenants = get_user_tenants(current_user, db)
    rows = (
        db.query(ComplianceAssessmentDocumentItem, ComplianceAssessmentDocument)
        .join(
            ComplianceAssessmentDocument,
            ComplianceAssessmentDocumentItem.assessment_id == ComplianceAssessmentDocument.id,
        )
        .filter(ComplianceAssessmentDocument.tenant_id.in_(user_tenants))
        .all()
    )
    points = []
    for item, a in rows:
        # Internal Audit (UBL) + KPI Report + container/standard are not part of
        # the Assessments module — keep them out of the SLA / open-by-assessment feed.
        if (getattr(a, "assessment_format", "standard") or "standard") in EXCLUDED_FORMATS:
            continue
        points.append({
            "id": item.id,
            "assessment_id": a.id,
            "assessment_name": a.name,
            "assessment_type": a.assessment_type,
            "assessment_format": getattr(a, "assessment_format", "standard") or "standard",
            "item_number": item.item_number,
            "area_domain": item.area_domain,
            "control_description": item.control_description,
            "priority": item.priority,
            "compliance_status": item.compliance_status,
            "remediation_status": item.remediation_status,
            "timeline": item.timeline,
            "target_date": item.target_date.isoformat() if getattr(item, "target_date", None) else None,
            "closed_at": item.closed_at.isoformat() if getattr(item, "closed_at", None) else None,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        })
    return {"points": points}


@router.get("/by-asset/{asset_id}")
def assessments_for_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Reverse view: every assessment scoped to this IT asset (i.e. whose
    linked_asset_ids contains it), with a compact status/validity summary. Feeds
    the 'Assessments' section on the asset detail page."""
    user_tenants = get_user_tenants(current_user, db)
    rows = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).all()
    out = []
    for a in rows:
        ids = getattr(a, "linked_asset_ids", None) or []
        try:
            linked = int(asset_id) in [int(x) for x in ids]
        except Exception:
            linked = False
        if not linked:
            continue
        total = a.total_items or 0
        complied = a.complied_count or 0
        na = a.na_count or 0
        denom = max(1, total - na)
        out.append({
            "id": a.id,
            "name": a.name,
            "assessment_type": a.assessment_type,
            "assessment_format": getattr(a, "assessment_format", "standard") or "standard",
            "status": a.status,
            "total_items": total,
            "complied_count": complied,
            "not_complied_count": a.not_complied_count or 0,
            "in_progress_count": a.in_progress_count or 0,
            "validity_pct": round((complied / denom) * 100),
            "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        })
    return {"asset_id": asset_id, "assessments": out}


@router.get("/overview")
def get_assessments_overview(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Board overview: every assessment scored by its family formula, grouped by
    category, with a module rollup + a universal SLA closure summary. Powers the
    redesigned graphical Assessments overview. (Registered before /{assessment_id}
    so the literal path wins.)"""
    from datetime import datetime as _dt
    from grc.modules.assessments.scoring import score_assessment, EXCLUDED_FORMATS
    from grc.models import (
        ComplianceAssessmentDocument as _Doc,
        ComplianceAssessmentDocumentItem as _Item,
        ComplianceSlaPolicy as _Policy,
        AssessmentItemEvidence as _Ev,
    )
    tids = get_user_tenants(current_user, db)
    now = _dt.utcnow()
    if not tids:
        return {"as_of": now.isoformat(), "categories": [],
                "performance": {"score": None, "grade": None, "assessments": 0},
                "sla": {}, "attention": {}}

    policy = db.query(_Policy).filter(_Policy.tenant_id.in_(tids)).first()
    docs = db.query(_Doc).filter(_Doc.tenant_id.in_(tids)).all()
    EV_FORMATS = {"asvs_checklist", "mobile_app_security", "owasp_v4_testing_checklist", "nca_dcc_tool"}

    cats: dict = {}
    sla_tot = {"gaps": 0, "closed": 0, "open": 0, "overdue": 0}
    unmapped = 0  # recognized (non-excluded) assessments with no family scorer
    for d in docs:
        fmt = getattr(d, "assessment_format", "standard") or "standard"
        if fmt in EXCLUDED_FORMATS:
            continue
        items = db.query(_Item).filter(_Item.assessment_id == d.id).all()
        ev = {}
        if fmt in EV_FORMATS:
            for it in items:
                ev[it.id] = db.query(_Ev).filter(_Ev.assessment_item_id == it.id).count()
        res = score_assessment(d, items, ev, policy, now)
        if res is None:
            # A non-excluded format we don't yet score — surface it instead of
            # silently dropping it from the board.
            unmapped += 1
            continue
        entry = {
            "id": d.id, "name": d.name, "format": fmt, "family": res["family"],
            "item_noun": res["item_noun"], "status": d.status, "n_items": len(items),
            "content": res["content"]["score"], "sla": res["sla"]["score"],
            "level_achieved": res["content"].get("level_achieved"),
            "metrics": res["content"]["metrics"], "sla_metrics": res["sla"]["metrics"],
            "counts": res["content"].get("counts", {}),
            "sla_counts": res["sla"].get("counts", {}),
            "weights": res["content"].get("weights", {}),
            "by_dimension": res["content"].get("by_dimension"),
            "by_domain": res["content"].get("by_domain"),
            "by_platform": res["content"].get("by_platform"),
        }
        cats.setdefault(res["category"], []).append(entry)
        sc = res["sla"].get("counts", {})
        for k in sla_tot:
            sla_tot[k] += sc.get(k, 0)

    # Rollups are ITEM-COUNT WEIGHTED so a 5-item toy assessment no longer weighs
    # the same as a 500-control register. Weight is capped (200) so one very large
    # assessment can't swamp the board.
    def _cap(n):
        return max(1, min(int(n or 1), 200))

    def _wmean(pairs):
        num = sum(s * w for s, w in pairs)
        den = sum(w for _, w in pairs)
        return round(num / den, 1) if den else None

    CAT_ORDER = ["Cyber Security", "NCA", "Digital Operations", "Privacy & Data"]
    categories = []
    for cat in CAT_ORDER + [c for c in cats if c not in CAT_ORDER]:
        rows = cats.get(cat)
        if not rows:
            continue
        scored = [(r["content"], _cap(r.get("n_items"))) for r in rows if r["content"] is not None]
        slas = [r["sla"] for r in rows if r["sla"] is not None]
        categories.append({
            "category": cat,
            "score": _wmean(scored),
            "sla": round(sum(slas) / len(slas), 1) if slas else None,
            "count": len(rows),
            "assessments": rows,
        })

    all_content = [(r["content"], _cap(r.get("n_items")))
                   for rows in cats.values() for r in rows if r["content"] is not None]
    all_sla = [r["sla"] for rows in cats.values() for r in rows if r["sla"] is not None]
    perf = _wmean(all_content)
    grade = (None if perf is None else "excellent" if perf >= 85 else "good" if perf >= 70
             else "fair" if perf >= 50 else "poor")
    module_sla = round(sum(all_sla) / len(all_sla), 1) if all_sla else None
    return {
        "as_of": now.isoformat(),
        "performance": {"score": perf, "grade": grade,
                        "assessments": sum(len(v) for v in cats.values())},
        "sla": {"score": module_sla, "gaps": sla_tot["gaps"], "closed": sla_tot["closed"],
                "open": sla_tot["open"], "overdue": sla_tot["overdue"]},
        "attention": {
            "overdue_gaps": sla_tot["overdue"],
            "open_gaps": sla_tot["open"],
            "not_started": sum(1 for rows in cats.values() for r in rows if r["content"] is None),
            "unmapped_formats": unmapped,
        },
        "categories": categories,
    }


@router.get("/kpi-live")
def kpi_live_metrics(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Live-computed Cyber Security KPI actuals from REAL in-platform modules.

    Only KPIs that genuinely map to data this platform owns are computed here —
    each with a real numerator/denominator/formula. Everything else the caller
    flags as 'external source' (no fabricated number). Nothing static.
    """
    from datetime import datetime
    from grc.modules.assessments.kpi_live import compute_kpi_metrics
    now = datetime.utcnow()
    tids = get_user_tenants(current_user, db) or []
    metrics = compute_kpi_metrics(db, now)

    # Snapshot today's value per KPI, then read the trend history so the dashboard
    # can draw a real line over time (grc_metric_snapshot; idempotent daily upsert).
    if tids:
        try:
            from grc.services import metric_snapshots as ms
            ms.ensure_table(db)
            today = now.date()
            for key, m in metrics.items():
                if m.get("actual") is not None:
                    ms.upsert(db, tids[0], f"kpi_{key}", today, m["actual"])
            db.commit()
            for key, m in metrics.items():
                m["history"] = ms.read_trend(db, tids, f"kpi_{key}", days=400)
        except Exception:
            db.rollback()
            for m in metrics.values():
                m.setdefault("history", [])

    return {"as_of": now.isoformat(), "metrics": metrics}


@router.get("/{assessment_id}")
def get_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):

    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(ComplianceAssessmentDocument).options(
        joinedload(ComplianceAssessmentDocument.items)
    ).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment not found"
        )
    
    # Real linked-evidence count per item (single grouped query, no N+1) so the
    # controls table can show an accurate evidence badge.
    item_ids = [it.id for it in assessment.items]
    evidence_counts = {}
    if item_ids:
        for row in (
            db.query(
                AssessmentItemEvidence.assessment_item_id,
                func.count(AssessmentItemEvidence.id),
            )
            .filter(AssessmentItemEvidence.assessment_item_id.in_(item_ids))
            .group_by(AssessmentItemEvidence.assessment_item_id)
            .all()
        ):
            evidence_counts[row[0]] = row[1]

    items_by_domain = {}
    for item in assessment.items:
        domain = item.area_domain or "Uncategorized"
        if domain not in items_by_domain:
            items_by_domain[domain] = []
        items_by_domain[domain].append({
            "evidence_count": evidence_counts.get(item.id, 0),
            "id": item.id,
            "item_number": item.item_number,
            "area_domain": item.area_domain,
            "control_description": item.control_description,
            "compliance_status": item.compliance_status,
            "maturity_score": item.maturity_score,
            "risk_rating": item.risk_rating,
            "gaps_identified": item.gaps_identified,
            "proposed_solution": item.proposed_solution,
            "responsible_party": item.responsible_party,
            "timeline": item.timeline,
            "priority": item.priority,
            "evidence_reference": item.evidence_reference,
            "remarks": item.remarks,
            "remediation_status": item.remediation_status,
            "asset_status": getattr(item, "asset_status", None) or {},
            "ai_evidence_recommendation": item.ai_evidence_recommendation,
            "ai_recommendation_generated_at": item.ai_recommendation_generated_at.isoformat() if item.ai_recommendation_generated_at else None,
            "target_date": item.target_date.isoformat() if getattr(item, "target_date", None) else None,
            "closed_at": item.closed_at.isoformat() if getattr(item, "closed_at", None) else None,
            "created_at": item.created_at.isoformat(),
            "updated_at": item.updated_at.isoformat() if item.updated_at else None
        })
    
    return {
        "id": assessment.id,
        "tenant_id": assessment.tenant_id,
        "name": assessment.name,
        "assessment_type": assessment.assessment_type,
        "assessment_format": getattr(assessment, "assessment_format", "standard") or "standard",
        "source": assessment.source,
        "file_name": assessment.file_name,
        "status": assessment.status,
        "due_date": assessment.due_date.isoformat() if assessment.due_date else None,
        "assessor": assessment.assessor,
        "overall_score": assessment.overall_score,
        "total_items": assessment.total_items,
        "complied_count": assessment.complied_count,
        "partially_complied_count": assessment.partially_complied_count,
        "not_complied_count": assessment.not_complied_count,
        "in_progress_count": assessment.in_progress_count,
        "na_count": assessment.na_count,
        "notes": assessment.notes,
        "linked_asset_ids": getattr(assessment, "linked_asset_ids", None) or [],
        "asset_levels": getattr(assessment, "asset_levels", None) or {},
        "created_at": assessment.created_at.isoformat(),
        "updated_at": assessment.updated_at.isoformat() if assessment.updated_at else None,
        "items": [
            {
                "id": item.id,
                "evidence_count": evidence_counts.get(item.id, 0),
                "item_number": item.item_number,
                "area_domain": item.area_domain,
                "control_description": item.control_description,
                "compliance_status": item.compliance_status,
                "maturity_score": item.maturity_score,
                "risk_rating": item.risk_rating,
                "gaps_identified": item.gaps_identified,
                "proposed_solution": item.proposed_solution,
                "responsible_party": item.responsible_party,
                "timeline": item.timeline,
                "priority": item.priority,
                "evidence_reference": item.evidence_reference,
                "remarks": item.remarks,
                "remediation_status": item.remediation_status,
                "asset_status": getattr(item, "asset_status", None) or {},
                "ai_evidence_recommendation": item.ai_evidence_recommendation,
                "ai_recommendation_generated_at": item.ai_recommendation_generated_at.isoformat() if item.ai_recommendation_generated_at else None,
                "target_date": item.target_date.isoformat() if getattr(item, "target_date", None) else None,
                "closed_at": item.closed_at.isoformat() if getattr(item, "closed_at", None) else None,
                "created_at": item.created_at.isoformat(),
                "updated_at": item.updated_at.isoformat() if item.updated_at else None
            }
            for item in assessment.items
        ],
        "items_by_domain": items_by_domain
    }


class _AssetScopeRequest(BaseModel):
    asset_ids: List[int] = []


class _AssetLevelRequest(BaseModel):
    asset_id: int
    level: int


@router.put("/{assessment_id}/asset-level")
def set_assessment_asset_level(
    assessment_id: int,
    payload: _AssetLevelRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Set the target ASVS level for one in-scope asset (user override)."""
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
    ).first()
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    lvl = payload.level if payload.level in (1, 2, 3) else 1
    levels = dict(getattr(assessment, "asset_levels", None) or {})
    levels[str(payload.asset_id)] = lvl
    assessment.asset_levels = levels
    assessment.updated_at = datetime.utcnow()
    db.commit()
    return {"asset_levels": assessment.asset_levels}


@router.put("/{assessment_id}/assets")
def set_assessment_assets(
    assessment_id: int,
    payload: _AssetScopeRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Set the IT Assets an assessment is scoped to (the app(s) it verifies)."""
    user_tenants = get_user_tenants(current_user, db)
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
    ).first()
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    assessment.linked_asset_ids = [int(a) for a in (payload.asset_ids or [])]
    assessment.updated_at = datetime.utcnow()
    db.commit()
    return {"linked_asset_ids": assessment.linked_asset_ids}


@router.put("/{assessment_id}")
def update_assessment(
    assessment_id: int,
    name: Optional[str] = None,
    assessment_type: Optional[str] = None,
    source: Optional[str] = None,
    status: Optional[str] = None,
    due_date: Optional[str] = None,
    assessor: Optional[str] = None,
    notes: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment not found"
        )
    
    if name is not None:
        assessment.name = name
    if assessment_type is not None:
        assessment.assessment_type = assessment_type
    if source is not None:
        assessment.source = source
    if status is not None:
        assessment.status = status
    if assessor is not None:
        assessment.assessor = assessor
    if notes is not None:
        assessment.notes = notes
    if due_date is not None:
        try:
            assessment.due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        except:
            pass
    
    assessment.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(assessment)
    
    return {
        "id": assessment.id,
        "name": assessment.name,
        "assessment_type": assessment.assessment_type,
        "source": assessment.source,
        "status": assessment.status,
        "due_date": assessment.due_date.isoformat() if assessment.due_date else None,
        "assessor": assessment.assessor,
        "notes": assessment.notes,
        "updated_at": assessment.updated_at.isoformat()
    }


@router.put("/items/{item_id}")
def update_assessment_item(
    item_id: int,
    compliance_status: Optional[str] = None,
    control_description: Optional[str] = None,
    area_domain: Optional[str] = None,
    gaps_identified: Optional[str] = None,
    proposed_solution: Optional[str] = None,
    responsible_party: Optional[str] = None,
    timeline: Optional[str] = None,
    priority: Optional[str] = None,
    evidence_reference: Optional[str] = None,
    remarks: Optional[str] = None,
    maturity_score: Optional[int] = None,
    risk_rating: Optional[str] = None,
    remediation_status: Optional[str] = None,
    target_date: Optional[str] = None,
    asset_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()
    
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment item not found"
        )
    
    if compliance_status is not None:
        if asset_id is not None:
            # Per-asset verification: write into asset_status[asset_id], leave the
            # global compliance_status untouched. Reassign the dict so SQLAlchemy
            # sees the JSON mutation.
            statuses = dict(getattr(item, "asset_status", None) or {})
            statuses[str(asset_id)] = normalize_status(compliance_status)
            item.asset_status = statuses
        else:
            item.compliance_status = normalize_status(compliance_status)
    if control_description is not None:
        item.control_description = control_description
    if remediation_status is not None:
        item.remediation_status = remediation_status or None
    if area_domain is not None:
        item.area_domain = area_domain
    if gaps_identified is not None:
        item.gaps_identified = gaps_identified
    if proposed_solution is not None:
        item.proposed_solution = proposed_solution
    if responsible_party is not None:
        item.responsible_party = responsible_party
    if timeline is not None:
        item.timeline = timeline
    if priority is not None:
        item.priority = priority
    if evidence_reference is not None:
        item.evidence_reference = evidence_reference
    if remarks is not None:
        item.remarks = remarks
    if risk_rating is not None:
        item.risk_rating = risk_rating or None
    # Maturity (0-5) is the PDPL assessment driver: setting it auto-derives the
    # compliance status (3-5 compliant). A negative value clears the score back
    # to "Not Assessed" (in_progress). Applied last so it wins over any
    # explicit compliance_status passed in the same call.
    if maturity_score is not None:
        if maturity_score < 0:
            item.maturity_score = None
            item.compliance_status = "in_progress"
        else:
            item.maturity_score = maturity_score
            item.compliance_status = _status_from_pdpl(maturity_score, None)

    # Per-point SLA deadline. Empty string clears it (back to policy-derived).
    if target_date is not None:
        if not target_date.strip():
            item.target_date = None
        else:
            try:
                item.target_date = datetime.fromisoformat(target_date.replace("Z", "+00:00"))
            except Exception:
                pass
    # Stamp / clear closed_at from the effective closed state so the closure /
    # aging math always has a real close date, even for points whose workbook
    # never carried one. A point is "closed" when its remediation is closed or
    # its compliance is fully complied.
    _is_closed = (item.remediation_status == "closed") or (item.compliance_status == "complied")
    if _is_closed and not item.closed_at:
        item.closed_at = datetime.utcnow()
    elif not _is_closed and item.closed_at:
        item.closed_at = None

    item.updated_at = datetime.utcnow()
    db.commit()
    
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == item.assessment_id
    ).first()
    
    if assessment:
        items = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == assessment.id
        ).all()
        stats = calculate_assessment_stats(items)
        
        assessment.complied_count = stats["complied"]
        assessment.partially_complied_count = stats["partially_complied"]
        assessment.not_complied_count = stats["not_complied"]
        assessment.in_progress_count = stats["in_progress"]
        assessment.na_count = stats["na"]
        assessment.overall_score = stats["overall_score"]
        assessment.updated_at = datetime.utcnow()
        db.commit()
    
    db.refresh(item)

    return {
        "id": item.id,
        "item_number": item.item_number,
        "area_domain": item.area_domain,
        "control_description": item.control_description,
        "compliance_status": item.compliance_status,
        "gaps_identified": item.gaps_identified,
        "proposed_solution": item.proposed_solution,
        "responsible_party": item.responsible_party,
        "timeline": item.timeline,
        "priority": item.priority,
        "evidence_reference": item.evidence_reference,
        "remarks": item.remarks,
        "target_date": item.target_date.isoformat() if item.target_date else None,
        "closed_at": item.closed_at.isoformat() if item.closed_at else None,
        "updated_at": item.updated_at.isoformat()
    }


# ── Item add / delete ────────────────────────────────────────────────────────
# Bring add/delete CRUD to parity with criticality assessments. Works for every
# compliance assessment type since they all share ComplianceAssessmentDocumentItem
# (DCC items use the same table with control_source='dcc' marker).

class _CreateAssessmentItemRequest(BaseModel):
    item_number: Optional[str] = None
    area_domain: Optional[str] = None
    control_description: str
    compliance_status: Optional[str] = "in_progress"
    gaps_identified: Optional[str] = None
    proposed_solution: Optional[str] = None
    responsible_party: Optional[str] = None
    timeline: Optional[str] = None
    priority: Optional[str] = "medium"
    evidence_reference: Optional[str] = None
    remarks: Optional[str] = None


@router.post("/{assessment_id}/items", status_code=status.HTTP_201_CREATED)
def create_assessment_item(
    assessment_id: int,
    payload: _CreateAssessmentItemRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)

    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants),
    ).first()
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment not found",
        )

    # Item number autoincrement when blank.
    item_number = (payload.item_number or "").strip()
    if not item_number:
        existing_count = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ).count()
        item_number = str(existing_count + 1)

    item = ComplianceAssessmentDocumentItem(
        tenant_id=assessment.tenant_id,
        assessment_id=assessment_id,
        item_number=item_number,
        area_domain=(payload.area_domain or "").strip() or None,
        control_description=(payload.control_description or "").strip(),
        compliance_status=normalize_status(payload.compliance_status or "in_progress"),
        gaps_identified=(payload.gaps_identified or "").strip() or None,
        proposed_solution=(payload.proposed_solution or "").strip() or None,
        responsible_party=(payload.responsible_party or "").strip() or None,
        timeline=(payload.timeline or "").strip() or None,
        priority=(payload.priority or "medium").strip().lower(),
        evidence_reference=(payload.evidence_reference or "").strip() or None,
        remarks=(payload.remarks or "").strip() or None,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(item)
    db.flush()

    # Recompute aggregate stats on the parent assessment.
    items = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.assessment_id == assessment.id
    ).all()
    stats = calculate_assessment_stats(items)
    assessment.total_items = len(items)
    assessment.complied_count = stats["complied"]
    assessment.partially_complied_count = stats["partially_complied"]
    assessment.not_complied_count = stats["not_complied"]
    assessment.in_progress_count = stats["in_progress"]
    assessment.na_count = stats["na"]
    assessment.overall_score = stats["overall_score"]
    assessment.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(item)

    return {
        "id": item.id,
        "assessment_id": item.assessment_id,
        "item_number": item.item_number,
        "area_domain": item.area_domain,
        "control_description": item.control_description,
        "compliance_status": item.compliance_status,
        "gaps_identified": item.gaps_identified,
        "proposed_solution": item.proposed_solution,
        "responsible_party": item.responsible_party,
        "timeline": item.timeline,
        "priority": item.priority,
        "evidence_reference": item.evidence_reference,
        "remarks": item.remarks,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


@router.delete("/items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_assessment_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)

    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants),
    ).first()
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment item not found",
        )

    parent_id = item.assessment_id
    db.delete(item)
    db.flush()

    # Recompute aggregate stats on the parent assessment.
    parent = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == parent_id
    ).first()
    if parent:
        items = db.query(ComplianceAssessmentDocumentItem).filter(
            ComplianceAssessmentDocumentItem.assessment_id == parent.id
        ).all()
        stats = calculate_assessment_stats(items)
        parent.total_items = len(items)
        parent.complied_count = stats["complied"]
        parent.partially_complied_count = stats["partially_complied"]
        parent.not_complied_count = stats["not_complied"]
        parent.in_progress_count = stats["in_progress"]
        parent.na_count = stats["na"]
        parent.overall_score = stats["overall_score"]
        parent.updated_at = datetime.utcnow()

    db.commit()
    return None


@router.delete("/{assessment_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(ComplianceAssessmentDocument).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment not found"
        )
    
    if assessment.file_path and os.path.exists(assessment.file_path):
        try:
            os.remove(assessment.file_path)
        except:
            pass
    
    db.delete(assessment)
    db.commit()
    return None


@router.get("/{assessment_id}/export")
def export_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(ComplianceAssessmentDocument).options(
        joinedload(ComplianceAssessmentDocument.items)
    ).filter(
        ComplianceAssessmentDocument.id == assessment_id,
        ComplianceAssessmentDocument.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Assessment not found"
        )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Items"
    
    headers = [
        "Item #", "Area/Domain", "Control Description", "Compliance Status",
        "Gaps Identified", "Proposed Solution", "Responsible Party",
        "Timeline", "Priority", "Evidence Reference", "Remarks"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    
    for item in assessment.items:
        ws.append([
            item.item_number,
            item.area_domain,
            item.control_description,
            item.compliance_status,
            item.gaps_identified,
            item.proposed_solution,
            item.responsible_party,
            item.timeline,
            item.priority,
            item.evidence_reference,
            item.remarks
        ])
    
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Assessment Summary"])
    summary_ws.append([])
    summary_ws.append(["Name", assessment.name])
    summary_ws.append(["Type", assessment.assessment_type])
    summary_ws.append(["Source", assessment.source or ""])
    summary_ws.append(["Status", assessment.status])
    summary_ws.append(["Assessor", assessment.assessor or ""])
    summary_ws.append(["Due Date", assessment.due_date.strftime("%Y-%m-%d") if assessment.due_date else ""])
    summary_ws.append([])
    summary_ws.append(["Compliance Statistics"])
    summary_ws.append(["Total Items", assessment.total_items])
    summary_ws.append(["Complied", assessment.complied_count])
    summary_ws.append(["Partially Complied", assessment.partially_complied_count])
    summary_ws.append(["Not Complied", assessment.not_complied_count])
    summary_ws.append(["In Progress", assessment.in_progress_count])
    summary_ws.append(["N/A", assessment.na_count])
    summary_ws.append(["Overall Score", f"{assessment.overall_score or 0}%"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    safe_name = "".join(c for c in assessment.name if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"{safe_name}_export.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/types/list")
def get_assessment_types():
    return {
        "assessment_types": [
            {"value": "gap_assessment", "label": "Gap Assessment"},
            {"value": "security_checklist", "label": "Security Checklist"},
            {"value": "internal_audit", "label": "Internal Audit"},
            {"value": "external_audit", "label": "External Audit"},
            {"value": "regulatory_assessment", "label": "Regulatory Assessment"},
            {"value": "vendor_assessment", "label": "Vendor Assessment"},
            {"value": "self_assessment", "label": "Self Assessment"},
            {"value": "maturity_assessment", "label": "Maturity Assessment"}
        ],
        "sources": [
            {"value": "SBP", "label": "State Bank of Pakistan"},
            {"value": "Internal", "label": "Internal"},
            {"value": "External Auditor", "label": "External Auditor"},
            {"value": "Regulator", "label": "Regulator"},
            {"value": "Vendor", "label": "Vendor"},
            {"value": "Other", "label": "Other"}
        ],
        "compliance_statuses": [
            {"value": "complied", "label": "Complied"},
            {"value": "partially_complied", "label": "Partially Complied"},
            {"value": "not_complied", "label": "Not Complied"},
            {"value": "in_progress", "label": "In Progress"},
            {"value": "na", "label": "N/A"}
        ],
        "priorities": [
            {"value": "critical", "label": "Critical"},
            {"value": "high", "label": "High"},
            {"value": "medium", "label": "Medium"},
            {"value": "low", "label": "Low"}
        ]
    }


def check_ai_available() -> bool:
    """Check if OpenAI API key is configured."""
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    if base_url and "modelfarm" in base_url:
        return True
    api_key = get_openai_api_key()
    if not api_key:
        return False
    if api_key.startswith("_DUMMY") or api_key == "your-api-key-here" or len(api_key) < 20:
        return False
    return True


def get_openai_client() -> OpenAI:
    if not check_ai_available():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI features unavailable. OpenAI API key not configured."
        )
    api_key = get_openai_api_key()
    # An empty AI_INTEGRATIONS_OPENAI_BASE_URL ("" in .env) must become None, or
    # the OpenAI client treats "" as the endpoint and fails with a bare
    # "Connection error". `or None` collapses empty string → default api.openai.com.
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL") or None
    return OpenAI(api_key=api_key, base_url=base_url)


def parse_ai_response(response_text: str) -> dict:
    try:
        cleaned = response_text.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        return json.loads(cleaned.strip())
    except json.JSONDecodeError:
        return {"recommendations": [], "summary": "Failed to parse AI response"}


def _iter_openai_clients_for_xlsx() -> List[OpenAI]:
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    is_modelfarm = bool(base_url and "modelfarm" in base_url)
    clients: List[OpenAI] = []
    seen = set()

    for key_name in ("OPENAI_API_KEY", "AI_INTEGRATIONS_OPENAI_API_KEY"):
        key = (os.environ.get(key_name) or "").strip()
        if not key:
            continue
        if key.startswith("_DUMMY") or key == "your-api-key-here" or len(key) < 20:
            continue
        if key in seen:
            continue
        seen.add(key)
        clients.append(OpenAI(api_key=key, base_url=base_url))

    if is_modelfarm and not clients:
        clients.append(OpenAI(api_key=None, base_url=base_url))

    return clients


def _normalize_recommendations(raw_recommendations) -> List[dict]:
    if not isinstance(raw_recommendations, list):
        return []

    cleaned = []
    for rec in raw_recommendations:
        if not isinstance(rec, dict):
            continue
        evidence_type = str(rec.get("evidence_type") or rec.get("type") or "").strip()
        if not evidence_type:
            continue
        cleaned.append({
            "evidence_type": evidence_type,
            "artifact_name": str(rec.get("artifact_name") or rec.get("artifact") or "").strip(),
            "why_auditable": str(rec.get("why_auditable") or rec.get("description") or "").strip(),
            "priority": str(rec.get("priority") or "medium").strip().lower() or "medium",
            "verification_checks": [
                str(x).strip() for x in (rec.get("verification_checks") or [])
                if str(x).strip()
            ][:3],
        })
        if len(cleaned) >= 5:
            break

    return cleaned


def _fallback_detail_recommendations(detail: dict, framework_name: str) -> List[dict]:
    subcategory = str(detail.get("subcategory") or "this control").strip()
    category = str(detail.get("category") or "control domain").strip()
    references = detail.get("references") or []
    refs = [str(r).strip() for r in references if str(r).strip()][:3]
    ref_text = ", ".join(refs) if refs else f"{framework_name} {subcategory}"

    return [
        {
            "evidence_type": "Approved Policy/Standard",
            "artifact_name": f"{subcategory} Policy Standard",
            "why_auditable": f"Demonstrates formal management approval, scope, ownership, and review cadence for {subcategory} within {category}.",
            "priority": "high",
            "verification_checks": ["Approval signatures", "Version history", "Effective/review dates"],
        },
        {
            "evidence_type": "Control Procedure / SOP",
            "artifact_name": f"{subcategory} SOP",
            "why_auditable": f"Shows how the control is executed in practice, including responsible roles and step-by-step activities mapped to {ref_text}.",
            "priority": "high",
            "verification_checks": ["RACI/owner", "Execution steps", "Exception handling"],
        },
        {
            "evidence_type": "System Configuration Evidence",
            "artifact_name": f"{subcategory} Configuration Baseline and Screenshots",
            "why_auditable": "Provides objective proof that required control settings are enabled and consistently configured in production systems.",
            "priority": "high",
            "verification_checks": ["Timestamped screenshots", "System identifiers", "Configuration values"],
        },
        {
            "evidence_type": "Operational Logs / Reports",
            "artifact_name": f"{subcategory} Monitoring Logs and Monthly Exception Report",
            "why_auditable": "Demonstrates the control is operating over time, not just documented, and captures incidents or deviations.",
            "priority": "medium",
            "verification_checks": ["Date range coverage", "Exception counts", "Reviewer sign-off"],
        },
        {
            "evidence_type": "Independent Review / Test Result",
            "artifact_name": f"{subcategory} Internal Audit or Control Testing Result",
            "why_auditable": "Confirms effectiveness through independent validation and documents gaps with remediation tracking.",
            "priority": "medium",
            "verification_checks": ["Test scope", "Findings and ratings", "Remediation actions"],
        },
    ]


def _generate_detail_recommendations_batch_ai(framework_name: str, detail_batch: List[dict]) -> dict:
    clients = _iter_openai_clients_for_xlsx()
    if not clients:
        return {}

    payload = {
        "framework_name": framework_name,
        "items": detail_batch,
        "instructions": (
            "For EACH item, generate up to 5 specific, auditable evidence recommendations. "
            "Recommendations must be tailored to the exact subcategory/control context, not generic. "
            "Return ONLY JSON object format: {\"items\":[{\"row_index\":<number>,\"recommendations\":[{\"evidence_type\":\"...\",\"artifact_name\":\"...\",\"why_auditable\":\"...\",\"priority\":\"high|medium|low\",\"verification_checks\":[\"...\"]}]}]}."
        ),
    }

    for client in clients:
        try:
            response = client.chat.completions.create(
                model=os.environ.get("AI_INTEGRATIONS_OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a senior GRC auditor. You provide precise, item-specific, audit-defensible evidence requirements. "
                            "Output valid JSON only."
                        ),
                    },
                    {"role": "user", "content": json.dumps(payload)},
                ],
                response_format={"type": "json_object"},
                temperature=0.2,
                max_tokens=2500,
            )
            parsed = parse_ai_response(response.choices[0].message.content or "{}")
            items = parsed.get("items") if isinstance(parsed, dict) else None
            if not isinstance(items, list):
                continue

            out = {}
            for item in items:
                if not isinstance(item, dict):
                    continue
                row_index = item.get("row_index")
                if not isinstance(row_index, int):
                    continue
                normalized = _normalize_recommendations(item.get("recommendations"))
                if normalized:
                    out[row_index] = normalized
            return out
        except Exception:
            continue

    return {}


def _ensure_xlsx_detail_recommendations(data: dict) -> bool:
    if not isinstance(data, dict):
        return False
    sheets = data.get("sheets") or {}
    details = sheets.get("details") or []
    if not isinstance(details, list) or not details:
        return False

    framework_name = str(data.get("framework_name") or "NIST CSF").strip() or "NIST CSF"
    missing_indices = [
        idx for idx, row in enumerate(details)
        if isinstance(row, dict) and not row.get("recommended_evidence")
    ]
    if not missing_indices:
        return False

    changed = False
    chunk_size = 12
    for start in range(0, len(missing_indices), chunk_size):
        chunk = missing_indices[start:start + chunk_size]
        batch_payload = []
        for idx in chunk:
            row = details[idx]
            batch_payload.append({
                "row_index": idx,
                "function": str(row.get("function") or ""),
                "category": str(row.get("category") or ""),
                "subcategory": str(row.get("subcategory") or ""),
                "references": [str(r).strip() for r in (row.get("references") or []) if str(r).strip()][:4],
                "policy_maturity": row.get("policy_maturity"),
                "practice_maturity": row.get("practice_maturity"),
            })

        ai_map = _generate_detail_recommendations_batch_ai(framework_name, batch_payload)

        for idx in chunk:
            row = details[idx]
            recommendations = ai_map.get(idx) or _fallback_detail_recommendations(row, framework_name)
            row["recommended_evidence"] = recommendations[:5]
            changed = True

    return changed


@router.get("/{assessment_id}/items/{item_id}/evidence")
def get_item_evidence(
    assessment_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()
    
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")
    
    evidence_links = db.query(AssessmentItemEvidence).options(
        joinedload(AssessmentItemEvidence.evidence),
        joinedload(AssessmentItemEvidence.workflow)
    ).filter(
        AssessmentItemEvidence.assessment_item_id == item_id
    ).all()

    # Track evidence IDs already covered by AssessmentItemEvidence to avoid duplicates
    linked_evidence_ids = {link.evidence_id for link in evidence_links if link.evidence_id}

    result = [
        {
            "id": link.id,
            "evidence_id": link.evidence_id,
            "evidence_name": link.evidence.name if link.evidence else None,
            "evidence_file_name": link.evidence.file_name if link.evidence else None,
            "evidence_file_type": link.evidence.file_type if link.evidence else None,
            "evidence_status": link.evidence.status if link.evidence else None,
            "evidence_uploaded_at": link.evidence.uploaded_at.isoformat() if link.evidence and link.evidence.uploaded_at else None,
            "workflow_id": link.workflow_id,
            "workflow_name": link.workflow.name if link.workflow else None,
            "current_tier": link.current_tier,
            "approval_status": link.status,
            "submitted_at": link.submitted_at.isoformat() if link.submitted_at else None,
            "created_at": link.created_at.isoformat(),
            "source": "assessment_upload",
        }
        for link in evidence_links
    ]

    # Also pull evidence linked via the Evidence module's AI "Link to Requirement" feature.
    # That feature creates EvidenceControlMapping entries (ParsedFrameworkControl → Evidence)
    # which are separate from AssessmentItemEvidence. We match them here by comparing the
    # assessment item's item_number against the parsed control's control_id / original_reference.
    if item.item_number:
        item_number_lower = item.item_number.strip().lower()
        framework_links = (
            db.query(EvidenceControlMapping)
            .join(ParsedFrameworkControl, EvidenceControlMapping.parsed_control_id == ParsedFrameworkControl.id)
            .join(Evidence, EvidenceControlMapping.evidence_id == Evidence.id)
            .filter(
                Evidence.tenant_id == item.tenant_id,
                or_(
                    func.lower(ParsedFrameworkControl.control_id) == item_number_lower,
                    func.lower(ParsedFrameworkControl.original_reference) == item_number_lower,
                )
            )
            .all()
        )

        for mapping in framework_links:
            if mapping.evidence_id in linked_evidence_ids:
                continue  # Already included via AssessmentItemEvidence
            if not mapping.evidence:
                continue
            ev = mapping.evidence
            linked_evidence_ids.add(ev.id)
            result.append({
                "id": f"ecm-{mapping.id}",  # Synthetic ID to distinguish from AssessmentItemEvidence
                "evidence_id": ev.id,
                "evidence_name": ev.name,
                "evidence_file_name": ev.file_name,
                "evidence_file_type": ev.file_type,
                "evidence_status": ev.status,
                "evidence_uploaded_at": ev.uploaded_at.isoformat() if ev.uploaded_at else None,
                "workflow_id": None,
                "workflow_name": None,
                "current_tier": 0,
                "approval_status": "framework_linked",
                "submitted_at": None,
                "created_at": mapping.created_at.isoformat() if mapping.created_at else "",
                "source": "framework_link",
                "framework_name": mapping.framework_name,
                "control_code": mapping.control_code,
                "confidence_score": mapping.confidence_score,
                "matching_rationale": mapping.matching_rationale,
            })

    return {
        "item_id": item_id,
        "assessment_id": assessment_id,
        "evidence": result,
    }


@router.post("/{assessment_id}/items/{item_id}/evidence/link")
def link_existing_item_evidence(
    assessment_id: int,
    item_id: int,
    payload: LinkExistingEvidenceRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()

    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")

    evidence = db.query(Evidence).filter(
        Evidence.id == payload.evidence_id,
        Evidence.tenant_id == item.tenant_id
    ).first()
    if not evidence:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence not found")

    existing_link = db.query(AssessmentItemEvidence).filter(
        AssessmentItemEvidence.assessment_item_id == item_id,
        AssessmentItemEvidence.evidence_id == evidence.id
    ).first()
    if existing_link:
        return {
            "id": existing_link.id,
            "evidence_id": evidence.id,
            "evidence_name": evidence.name,
            "evidence_file_name": evidence.file_name,
            "evidence_file_type": evidence.file_type,
            "evidence_status": evidence.status,
            "evidence_uploaded_at": evidence.uploaded_at.isoformat() if evidence.uploaded_at else None,
            "workflow_id": existing_link.workflow_id,
            "current_tier": existing_link.current_tier,
            "approval_status": existing_link.status,
            "already_linked": True,
            "message": "Evidence is already linked to this assessment item"
        }

    workflow = None
    if payload.workflow_id:
        workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.id == payload.workflow_id,
            AssessmentEvidenceApprovalWorkflow.tenant_id == item.tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_active == True
        ).first()
    else:
        workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.tenant_id == item.tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_default == True,
            AssessmentEvidenceApprovalWorkflow.is_active == True
        ).first()

    evidence_link = AssessmentItemEvidence(
        assessment_item_id=item_id,
        evidence_id=evidence.id,
        tenant_id=item.tenant_id,
        workflow_id=workflow.id if workflow else None,
        current_tier=0,
        status="draft"
    )
    db.add(evidence_link)
    db.commit()
    db.refresh(evidence_link)

    return {
        "id": evidence_link.id,
        "evidence_id": evidence.id,
        "evidence_name": evidence.name,
        "evidence_file_name": evidence.file_name,
        "evidence_file_type": evidence.file_type,
        "evidence_status": evidence.status,
        "evidence_uploaded_at": evidence.uploaded_at.isoformat() if evidence.uploaded_at else None,
        "workflow_id": evidence_link.workflow_id,
        "current_tier": evidence_link.current_tier,
        "approval_status": evidence_link.status,
        "already_linked": False,
        "message": "Evidence linked successfully"
    }


@router.post("/{assessment_id}/items/{item_id}/evidence/upload")
async def upload_item_evidence(
    assessment_id: int,
    item_id: int,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    file: UploadFile = File(...),
    workflow_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()
    
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")
    
    file_content = await file.read()
    file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
    file_id = str(uuid.uuid4())
    file_path = os.path.join(EVIDENCE_UPLOAD_DIR, f"{file_id}{file_ext}")
    
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    evidence = Evidence(
        tenant_id=item.tenant_id,
        name=name,
        description=description,
        file_path=file_path,
        file_name=file.filename,
        file_type=file.content_type,
        status="draft",
        uploaded_by=current_user.id,
        uploaded_at=datetime.utcnow()
    )
    db.add(evidence)
    db.flush()
    
    if workflow_id:
        workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.id == workflow_id,
            AssessmentEvidenceApprovalWorkflow.tenant_id == item.tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_active == True
        ).first()
    else:
        workflow = db.query(AssessmentEvidenceApprovalWorkflow).filter(
            AssessmentEvidenceApprovalWorkflow.tenant_id == item.tenant_id,
            AssessmentEvidenceApprovalWorkflow.is_default == True,
            AssessmentEvidenceApprovalWorkflow.is_active == True
        ).first()
    
    evidence_link = AssessmentItemEvidence(
        assessment_item_id=item_id,
        evidence_id=evidence.id,
        tenant_id=item.tenant_id,
        workflow_id=workflow.id if workflow else None,
        current_tier=0,
        status="draft"
    )
    db.add(evidence_link)
    db.commit()
    db.refresh(evidence_link)
    
    return {
        "id": evidence_link.id,
        "evidence_id": evidence.id,
        "evidence_name": evidence.name,
        "evidence_file_name": evidence.file_name,
        "evidence_file_type": evidence.file_type,
        "evidence_status": evidence.status,
        "workflow_id": evidence_link.workflow_id,
        "current_tier": evidence_link.current_tier,
        "approval_status": evidence_link.status,
        "message": "Evidence uploaded successfully"
    }


@router.post("/{assessment_id}/items/{item_id}/ai-recommendation")
def generate_ai_recommendation(
    assessment_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    item = db.query(ComplianceAssessmentDocumentItem).options(
        joinedload(ComplianceAssessmentDocumentItem.assessment)
    ).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()
    
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")
    
    try:
        client = get_openai_client()
        
        assessment = item.assessment
        prompt = EVIDENCE_RECOMMENDATION_PROMPT.format(
            assessment_name=assessment.name if assessment else "Unknown",
            assessment_type=assessment.assessment_type if assessment else "Unknown",
            item_number=item.item_number or "N/A",
            area_domain=item.area_domain or "General",
            control_description=item.control_description or "No description",
            compliance_status=item.compliance_status or "Unknown",
            gaps_identified=item.gaps_identified or "None specified"
        )
        
        response = client.chat.completions.create(
            model=get_openai_model(),
            messages=[
                {
                    "role": "system",
                    "content": "You are a compliance expert recommending evidence for assessment items. Respond only with valid JSON."
                },
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=2000,
            temperature=0.3
        )
        
        result = parse_ai_response(response.choices[0].message.content or '{"recommendations": []}')
        
        item.ai_evidence_recommendation = json.dumps(result)
        item.ai_recommendation_generated_at = datetime.utcnow()
        db.commit()
        
        return {
            "item_id": item_id,
            "assessment_id": assessment_id,
            "recommendation": result,
            "generated_at": item.ai_recommendation_generated_at.isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI recommendation generation failed: {str(e)}"
        )


@router.post("/{assessment_id}/items/{item_id}/ai-assess")
def ai_assess_item(
    assessment_id: int,
    item_id: int,
    gap: str = "",
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Draft a full PDPL assessment for one control: likely findings/gap, a
    recommended remediation action, a risk rating and a priority. Maturity is
    intentionally NOT generated — it reflects the organisation's actual
    implementation and must be judged by the assessor (it drives the score)."""
    user_tenants = get_user_tenants(current_user, db)
    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants),
    ).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")
    try:
        client = get_openai_client()
        kv = _parse_remarks_kv(item.remarks)
        prompt = (
            "You are a Saudi PDPL (Personal Data Protection Law) compliance assessor. "
            "For the control below, produce STRICT JSON with exactly these keys:\n"
            "  how_to_assess — string: a clear, practical explanation (4-6 sentences) for an "
            "assessor UNFAMILIAR with this control — what to examine, the specific questions to ask, "
            "where to look, and what a compliant ('good') implementation looks like in practice.\n"
            "  what_good_looks_like — string: 2-3 sentences describing the ideal end-state.\n"
            "  evidence_examples — array of 4-7 strings: specific documents/records/screens that "
            "would demonstrate compliance for this control.\n"
            "  findings   — string: the likely gap/finding when this control is typically unmet\n"
            "  remediation — string: the concrete remediation action to satisfy the control\n"
            "  risk_rating — one of: Low, Medium, High, Critical\n"
            "  priority    — one of: low, medium, high, critical\n"
            "  suggested_maturity — integer 0-5 (0=Absent,1=Initial,2=Developing,3=Defined,"
            "4=Managed,5=Optimised): a conservative starting estimate of likely current maturity\n\n"
            f"Domain: {item.area_domain or 'General'}\n"
            f"PDPL Reference: {kv.get('pdpl ref') or 'N/A'}\n"
            f"Control: {item.control_description or ''}\n"
            f"Assessment question: {kv.get('q') or ''}\n"
        )
        gap_text = (gap or "").strip()
        if gap_text:
            prompt += (
                f"\nThe assessor has ALREADY identified this specific gap for this control:\n"
                f"\"{gap_text}\"\n"
                "Set `findings` to this exact gap (you may lightly tidy the wording, but do not "
                "invent a different gap). Make `remediation` a concrete, step-by-step action that "
                "directly closes THIS gap, and set `risk_rating`/`priority` consistent with it.\n"
            )
        response = client.chat.completions.create(
            model=get_openai_model(),
            messages=[
                {"role": "system", "content": "You are a PDPL compliance expert. Respond ONLY with valid JSON."},
                {"role": "user", "content": prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=800,
            temperature=0.3,
        )
        result = parse_ai_response(response.choices[0].message.content or "{}")
        return {
            "item_id": item_id,
            "assessment_id": assessment_id,
            "draft": {
                "how_to_assess": result.get("how_to_assess"),
                "what_good_looks_like": result.get("what_good_looks_like"),
                "evidence_examples": result.get("evidence_examples") or [],
                "findings": result.get("findings"),
                "remediation": result.get("remediation"),
                "risk_rating": result.get("risk_rating"),
                "priority": (result.get("priority") or "").lower() or None,
                "suggested_maturity": result.get("suggested_maturity"),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"AI assessment draft failed: {str(e)}")


@router.get("/{assessment_id}/items/{item_id}/ai-recommendation")
def get_ai_recommendation(
    assessment_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    item = db.query(ComplianceAssessmentDocumentItem).filter(
        ComplianceAssessmentDocumentItem.id == item_id,
        ComplianceAssessmentDocumentItem.assessment_id == assessment_id,
        ComplianceAssessmentDocumentItem.tenant_id.in_(user_tenants)
    ).first()
    
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment item not found")
    
    recommendation = None
    if item.ai_evidence_recommendation:
        try:
            recommendation = json.loads(item.ai_evidence_recommendation)
        except json.JSONDecodeError:
            recommendation = {"raw": item.ai_evidence_recommendation}
    
    return {
        "item_id": item_id,
        "assessment_id": assessment_id,
        "recommendation": recommendation,
        "generated_at": item.ai_recommendation_generated_at.isoformat() if item.ai_recommendation_generated_at else None
    }
