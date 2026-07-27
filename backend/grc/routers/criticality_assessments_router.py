"""Criticality Assessments — bank-provided template parity.

Two endpoint families:
  - /criticality-assessments/info-system        (ISCA)
  - /criticality-assessments/infra-asset        (IACA)

Both expose CRUD on assessment items, an IT-asset picker, and a tenant-user
picker so the frontend can populate the Business Owner / Service Owner /
Custodian / Administrator / Assessor dropdowns directly.

Design notes:
  * Total score + criticality level are computed server-side on every
    create / update so the UI never has to. Source of truth for the bands
    is in this file (see _isca_level / _iaca_level).
  * Linkage to ITAsset is optional. When set, we pull a few denormalised
    fields back to the frontend so the row card can display the linked
    asset's name without a second query.
  * Permissions reuse the asset / risk perms already in the catalogue
    (`dashboard:assets:*` for view, `erm:risks:*` for edit) — both are
    already present for users who run RCSA or asset management. We keep
    this additive: nothing existing changes RBAC semantics.
"""

from __future__ import annotations

import io
import logging
import os
import uuid
from datetime import date as date_type, datetime, timedelta
from typing import Any, Dict, List, Literal, Optional, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from ..models import (
    CriticalityAssessmentActivity,
    CriticalityAssessmentComment,
    CriticalityAssessmentEvidence,
    GRCUser,
    InfoSystemCriticalityItem,
    InfraAssetCriticalityItem,
    ITAsset,
    Risk,
    get_db,
)
from .auth_router import (
    get_user_primary_tenant,
    require_auth,
    require_tenant_permission,
)

logger = logging.getLogger(__name__)


# ── RBAC gates ──────────────────────────────────────────────────────────────
# Permission names follow the existing module:submodule:action convention.
# Administrator + primary-contact bypass remains intact; non-admin users on
# custom roles need the named permission granted via the admin Roles UI
# (lazy-created by `_get_or_create_permission`).
_require_view = require_tenant_permission("assets:criticality_assessments:view")
_require_create = require_tenant_permission("assets:criticality_assessments:create")
_require_edit = require_tenant_permission("assets:criticality_assessments:edit")
_require_delete = require_tenant_permission("assets:criticality_assessments:delete")
# CISO tier requires this distinct permission so business owners can be
# tier-2 approvers without being able to perform tier-3 sign-off.
_require_approve_ciso = require_tenant_permission(
    "assets:criticality_assessments:approve_ciso",
)


router = APIRouter(
    prefix="/criticality-assessments",
    tags=["Criticality Assessments"],
)


# ── Evidence disk + size limits ─────────────────────────────────────────────
_EVIDENCE_DIR = "uploads/criticality_assessments_evidence"
os.makedirs(_EVIDENCE_DIR, exist_ok=True)
_MAX_EVIDENCE_BYTES = 50 * 1024 * 1024  # 50 MB


# ── Approval-state guard ────────────────────────────────────────────────────
# Edit (PUT) endpoints are forbidden while a row is mid-review or signed
# off. Operators must return it to the assessor before editing again.
_EDIT_LOCKED_STATES = {"submitted", "business_owner_review", "ciso_review", "approved"}


# ── Approval-state machine ──────────────────────────────────────────────────
# Tier → status that's "pending this tier". The submitter always advances to
# tier 2 (business owner). Business owner advances to tier 3 (CISO). CISO
# is the terminal approver.
_TIER_STATUS = {
    1: "draft",
    2: "business_owner_review",
    3: "ciso_review",
}


# ─── Common helpers ─────────────────────────────────────────────────────────


def _tenant_id_or_403(user: GRCUser, db: Session) -> int:
    tid = get_user_primary_tenant(user, db)
    if not tid:
        raise HTTPException(status_code=403, detail="No tenant context")
    return tid


def _safe_int(v: Any, lo: int, hi: int) -> Optional[int]:
    """Clamp a candidate score into [lo, hi]; return None if it's not a number."""
    if v is None or v == "":
        return None
    try:
        n = int(v)
    except (TypeError, ValueError):
        return None
    if n < lo or n > hi:
        return None
    return n


def _user_label(u: Optional[GRCUser]) -> Optional[str]:
    if not u:
        return None
    return u.display_name or u.username or u.email or f"User #{u.id}"


# ─── ISCA: band thresholds match the template's Calculation sheet ──────────


def _isca_level(total: Optional[int]) -> Optional[str]:
    if total is None:
        return None
    if total >= 24:
        return "mission_critical"
    if total >= 19:
        return "high"
    if total >= 13:
        return "moderate"
    if total >= 6:
        return "low"
    return None


# ISCA criteria + score range. Internet Facing accepts 0/2/4, B2B accepts 0/4 —
# we permit any value in the union range and let the UI lock the picker.
_ISCA_SCORE_FIELDS: List[tuple[str, int, int]] = [
    ("operational_dependency", 1, 4),
    ("financial_impact", 1, 4),
    ("customer_stakeholder_impact", 1, 4),
    ("data_sensitivity", 1, 4),
    ("unauthorized_access_risk", 1, 4),
    ("rto_rpo_requirements", 1, 4),
    ("internet_facing", 0, 4),
    ("b2b_exposure", 0, 4),
]


# ─── IACA: weighted scoring per the Calculation sheet ──────────────────────

# (field, weight%). Total weight = 100.
_IACA_WEIGHTS: List[tuple[str, float]] = [
    ("business_impact", 15.0),
    ("service_dependency", 12.0),
    ("data_sensitivity", 12.0),
    ("redundancy_failover", 10.0),
    ("rto", 10.0),
    ("availability_requirement", 10.0),
    ("operational_disruption", 10.0),
    ("regulatory_dependency", 11.0),
    ("exposure", 10.0),
]


def _iaca_total(item_data: Dict[str, Any]) -> Optional[float]:
    """Weighted sum over the 9 criteria. Returns None if every criterion
    is missing (avoids reporting a misleading 0.0 on a brand-new draft).
    """
    seen = False
    total = 0.0
    for field, weight in _IACA_WEIGHTS:
        v = item_data.get(field)
        if v is None:
            continue
        seen = True
        # rating * (weight%) gives a 0-4 scale (matches the template).
        total += float(v) * (weight / 100.0)
    return round(total, 2) if seen else None


def _iaca_level(total: Optional[float]) -> Optional[str]:
    """Per the template: ≥3.5 mission-critical, ≥3.0 high, ≥2.0 moderate, <2 low."""
    if total is None:
        return None
    if total >= 3.5:
        return "mission_critical"
    if total >= 3.0:
        return "high"
    if total >= 2.0:
        return "moderate"
    return "low"


# ─── Shared schemas ────────────────────────────────────────────────────────


class TenantUserOption(BaseModel):
    id: int
    display_name: str
    email: Optional[str] = None
    designation: Optional[str] = None


class AssetOwnerDetail(BaseModel):
    """Denormalised owner contact info — drives the contact-block auto-fill
    on the criticality drawers so picking an asset can pre-populate
    Business Owner / Custodian / Administrator without a second round-trip.
    """
    user_id: Optional[int] = None
    name: Optional[str] = None
    designation: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


class AssetOption(BaseModel):
    id: int
    name: str
    asset_type: Optional[str] = None
    criticality: Optional[str] = None
    # Auto-fill source fields. Every field is optional — picking an asset
    # that doesn't have data for a slot leaves that slot untouched on the
    # form (the frontend treats `null` as "skip, don't overwrite").
    description: Optional[str] = None
    location: Optional[str] = None
    vendor: Optional[str] = None
    host_name: Optional[str] = None
    ip_address: Optional[str] = None
    address: Optional[str] = None  # convenience: ip_address or host_name
    associated_ips: Optional[str] = None
    business_owner: Optional[AssetOwnerDetail] = None
    primary_owner: Optional[AssetOwnerDetail] = None
    secondary_owner: Optional[AssetOwnerDetail] = None


# ── ISCA ────────────────────────────────────────────────────────────────────


class ISCABase(BaseModel):
    linked_asset_id: Optional[int] = None
    name: str = Field(..., max_length=255)
    description: Optional[str] = None
    address: Optional[str] = Field(None, max_length=500)

    business_owner_user_id: Optional[int] = None
    business_owner_name: Optional[str] = None
    business_owner_designation: Optional[str] = None
    business_owner_phone: Optional[str] = None
    business_owner_email: Optional[str] = None

    service_owner_user_id: Optional[int] = None
    service_owner_name: Optional[str] = None
    service_owner_designation: Optional[str] = None
    service_owner_phone: Optional[str] = None
    service_owner_email: Optional[str] = None

    assessor_user_id: Optional[int] = None
    assessor_name: Optional[str] = None
    assessor_designation: Optional[str] = None
    assessor_phone: Optional[str] = None
    assessor_email: Optional[str] = None

    date_of_assessment: Optional[date_type] = None

    operational_dependency: Optional[int] = None
    financial_impact: Optional[int] = None
    customer_stakeholder_impact: Optional[int] = None
    data_sensitivity: Optional[int] = None
    unauthorized_access_risk: Optional[int] = None
    rto_rpo_requirements: Optional[int] = None
    internet_facing: Optional[int] = None
    b2b_exposure: Optional[int] = None

    comments: Optional[str] = None


class ISCAResponse(ISCABase):
    id: int
    tenant_id: int
    total_score: Optional[int] = None
    criticality_level: Optional[str] = None
    linked_asset_name: Optional[str] = None
    business_owner_user_name: Optional[str] = None
    service_owner_user_name: Optional[str] = None
    assessor_user_name: Optional[str] = None
    # Approval + risk linkage (Phase 2)
    approval_status: Optional[str] = "draft"
    current_approval_tier: Optional[int] = None
    submitted_at: Optional[datetime] = None
    submitted_by: Optional[int] = None
    submitted_by_name: Optional[str] = None
    approved_at: Optional[datetime] = None
    approved_by: Optional[int] = None
    approved_by_name: Optional[str] = None
    rejected_at: Optional[datetime] = None
    rejected_by: Optional[int] = None
    rejected_by_name: Optional[str] = None
    rejection_reason: Optional[str] = None
    linked_risk_id: Optional[int] = None
    evidence_count: int = 0
    comment_count: int = 0
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ── IACA ────────────────────────────────────────────────────────────────────


class IACABase(BaseModel):
    linked_asset_id: Optional[int] = None
    name: str = Field(..., max_length=255)
    description: Optional[str] = None
    make_model: Optional[str] = None
    location: Optional[str] = None
    associated_ips: Optional[str] = None
    fault_tolerance: Optional[str] = None

    custodian_user_id: Optional[int] = None
    custodian_name: Optional[str] = None
    custodian_designation: Optional[str] = None
    custodian_phone: Optional[str] = None
    custodian_email: Optional[str] = None

    administrator_user_id: Optional[int] = None
    administrator_name: Optional[str] = None
    administrator_designation: Optional[str] = None
    administrator_phone: Optional[str] = None
    administrator_email: Optional[str] = None

    assessor_user_id: Optional[int] = None
    assessor_name: Optional[str] = None
    assessor_designation: Optional[str] = None
    assessor_phone: Optional[str] = None
    assessor_email: Optional[str] = None

    date_of_assessment: Optional[date_type] = None

    business_impact: Optional[int] = None
    service_dependency: Optional[int] = None
    data_sensitivity: Optional[int] = None
    redundancy_failover: Optional[int] = None
    rto: Optional[int] = None
    availability_requirement: Optional[int] = None
    operational_disruption: Optional[int] = None
    regulatory_dependency: Optional[int] = None
    exposure: Optional[int] = None

    comments: Optional[str] = None


class IACAResponse(IACABase):
    id: int
    tenant_id: int
    total_score: Optional[float] = None
    criticality_level: Optional[str] = None
    linked_asset_name: Optional[str] = None
    custodian_user_name: Optional[str] = None
    administrator_user_name: Optional[str] = None
    assessor_user_name: Optional[str] = None
    approval_status: Optional[str] = "draft"
    current_approval_tier: Optional[int] = None
    submitted_at: Optional[datetime] = None
    submitted_by: Optional[int] = None
    submitted_by_name: Optional[str] = None
    approved_at: Optional[datetime] = None
    approved_by: Optional[int] = None
    approved_by_name: Optional[str] = None
    rejected_at: Optional[datetime] = None
    rejected_by: Optional[int] = None
    rejected_by_name: Optional[str] = None
    rejection_reason: Optional[str] = None
    linked_risk_id: Optional[int] = None
    evidence_count: int = 0
    comment_count: int = 0
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ─── Serializers ───────────────────────────────────────────────────────────


def _count_evidence(db: Session, kind: str, item_id: int, tenant_id: int) -> int:
    return (
        db.query(func.count(CriticalityAssessmentEvidence.id))
        .filter(
            CriticalityAssessmentEvidence.tenant_id == tenant_id,
            CriticalityAssessmentEvidence.assessment_kind == kind,
            CriticalityAssessmentEvidence.assessment_id == item_id,
        )
        .scalar()
        or 0
    )


def _count_comments(db: Session, kind: str, item_id: int, tenant_id: int) -> int:
    return (
        db.query(func.count(CriticalityAssessmentComment.id))
        .filter(
            CriticalityAssessmentComment.tenant_id == tenant_id,
            CriticalityAssessmentComment.assessment_kind == kind,
            CriticalityAssessmentComment.assessment_id == item_id,
        )
        .scalar()
        or 0
    )


def _isca_to_response(
    item: InfoSystemCriticalityItem,
    db: Optional[Session] = None,
) -> ISCAResponse:
    ev_count = (
        _count_evidence(db, "isca", item.id, item.tenant_id) if db is not None else 0
    )
    cm_count = (
        _count_comments(db, "isca", item.id, item.tenant_id) if db is not None else 0
    )
    return ISCAResponse(
        id=item.id,
        tenant_id=item.tenant_id,
        linked_asset_id=item.linked_asset_id,
        name=item.name,
        description=item.description,
        address=item.address,
        business_owner_user_id=item.business_owner_user_id,
        business_owner_name=item.business_owner_name,
        business_owner_designation=item.business_owner_designation,
        business_owner_phone=item.business_owner_phone,
        business_owner_email=item.business_owner_email,
        service_owner_user_id=item.service_owner_user_id,
        service_owner_name=item.service_owner_name,
        service_owner_designation=item.service_owner_designation,
        service_owner_phone=item.service_owner_phone,
        service_owner_email=item.service_owner_email,
        assessor_user_id=item.assessor_user_id,
        assessor_name=item.assessor_name,
        assessor_designation=item.assessor_designation,
        assessor_phone=item.assessor_phone,
        assessor_email=item.assessor_email,
        date_of_assessment=item.date_of_assessment,
        operational_dependency=item.operational_dependency,
        financial_impact=item.financial_impact,
        customer_stakeholder_impact=item.customer_stakeholder_impact,
        data_sensitivity=item.data_sensitivity,
        unauthorized_access_risk=item.unauthorized_access_risk,
        rto_rpo_requirements=item.rto_rpo_requirements,
        internet_facing=item.internet_facing,
        b2b_exposure=item.b2b_exposure,
        total_score=item.total_score,
        criticality_level=item.criticality_level,
        comments=item.comments,
        linked_asset_name=(item.linked_asset.name if item.linked_asset else None),
        business_owner_user_name=_user_label(item.business_owner_user),
        service_owner_user_name=_user_label(item.service_owner_user),
        assessor_user_name=_user_label(item.assessor_user),
        approval_status=item.approval_status or "draft",
        current_approval_tier=item.current_approval_tier,
        submitted_at=item.submitted_at,
        submitted_by=item.submitted_by,
        submitted_by_name=_user_label(getattr(item, "submitter", None)),
        approved_at=item.approved_at,
        approved_by=item.approved_by,
        approved_by_name=_user_label(getattr(item, "approver", None)),
        rejected_at=item.rejected_at,
        rejected_by=item.rejected_by,
        rejected_by_name=_user_label(getattr(item, "rejecter", None)),
        rejection_reason=item.rejection_reason,
        linked_risk_id=item.linked_risk_id,
        evidence_count=ev_count,
        comment_count=cm_count,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _iaca_to_response(
    item: InfraAssetCriticalityItem,
    db: Optional[Session] = None,
) -> IACAResponse:
    ev_count = (
        _count_evidence(db, "iaca", item.id, item.tenant_id) if db is not None else 0
    )
    cm_count = (
        _count_comments(db, "iaca", item.id, item.tenant_id) if db is not None else 0
    )
    return IACAResponse(
        id=item.id,
        tenant_id=item.tenant_id,
        linked_asset_id=item.linked_asset_id,
        name=item.name,
        description=item.description,
        make_model=item.make_model,
        location=item.location,
        associated_ips=item.associated_ips,
        fault_tolerance=item.fault_tolerance,
        custodian_user_id=item.custodian_user_id,
        custodian_name=item.custodian_name,
        custodian_designation=item.custodian_designation,
        custodian_phone=item.custodian_phone,
        custodian_email=item.custodian_email,
        administrator_user_id=item.administrator_user_id,
        administrator_name=item.administrator_name,
        administrator_designation=item.administrator_designation,
        administrator_phone=item.administrator_phone,
        administrator_email=item.administrator_email,
        assessor_user_id=item.assessor_user_id,
        assessor_name=item.assessor_name,
        assessor_designation=item.assessor_designation,
        assessor_phone=item.assessor_phone,
        assessor_email=item.assessor_email,
        date_of_assessment=item.date_of_assessment,
        business_impact=item.business_impact,
        service_dependency=item.service_dependency,
        data_sensitivity=item.data_sensitivity,
        redundancy_failover=item.redundancy_failover,
        rto=item.rto,
        availability_requirement=item.availability_requirement,
        operational_disruption=item.operational_disruption,
        regulatory_dependency=item.regulatory_dependency,
        exposure=item.exposure,
        total_score=item.total_score,
        criticality_level=item.criticality_level,
        comments=item.comments,
        linked_asset_name=(item.linked_asset.name if item.linked_asset else None),
        custodian_user_name=_user_label(item.custodian_user),
        administrator_user_name=_user_label(item.administrator_user),
        assessor_user_name=_user_label(item.assessor_user),
        approval_status=item.approval_status or "draft",
        current_approval_tier=item.current_approval_tier,
        submitted_at=item.submitted_at,
        submitted_by=item.submitted_by,
        submitted_by_name=_user_label(getattr(item, "submitter", None)),
        approved_at=item.approved_at,
        approved_by=item.approved_by,
        approved_by_name=_user_label(getattr(item, "approver", None)),
        rejected_at=item.rejected_at,
        rejected_by=item.rejected_by,
        rejected_by_name=_user_label(getattr(item, "rejecter", None)),
        rejection_reason=item.rejection_reason,
        linked_risk_id=item.linked_risk_id,
        evidence_count=ev_count,
        comment_count=cm_count,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


# ─── Shared picker endpoints ───────────────────────────────────────────────


@router.get("/users", response_model=List[TenantUserOption], dependencies=[Depends(_require_view)])
def list_users_for_assignment(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Active users in the caller's tenant DB — mirrors the
    `/critical-tasks/tenant-users` pattern. Includes the caller as a
    fallback so the list is never empty for a brand-new tenant.
    """
    users = (
        db.query(GRCUser)
        .filter(GRCUser.is_active.is_(True))
        .order_by(GRCUser.display_name.asc().nullslast(), GRCUser.username.asc())
        .all()
    )
    result = [
        TenantUserOption(
            id=u.id,
            display_name=_user_label(u) or f"User #{u.id}",
            email=u.email,
            designation=u.designation,
        )
        for u in users
    ]
    if not any(r.id == current_user.id for r in result):
        result.insert(
            0,
            TenantUserOption(
                id=current_user.id,
                display_name=_user_label(current_user) or f"User #{current_user.id}",
                email=current_user.email,
                designation=current_user.designation,
            ),
        )
    return result


def _asset_owner_detail(user: Optional[GRCUser]) -> Optional[AssetOwnerDetail]:
    """Build a denormalised owner block from a GRCUser, or None when absent."""
    if not user:
        return None
    return AssetOwnerDetail(
        user_id=user.id,
        name=_user_label(user),
        designation=user.designation,
        email=user.email,
        phone=None,  # GRCUser has no phone column today; leave for future
    )


@router.get("/assets", response_model=List[AssetOption], dependencies=[Depends(_require_view)])
def list_assets_for_link(
    asset_type: Optional[str] = Query(None, description="Filter by ITAsset.asset_type"),
    search: Optional[str] = Query(None, description="Substring match on asset name"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """IT-asset picker that drives the criticality drawers' auto-fill.

    Returns rich per-asset metadata — description, location, IPs,
    business / primary / secondary owners — so picking an asset can
    populate name + description + address + contact blocks in one shot.
    Capped at 200 rows; combine with `search` for larger tenants.
    """
    tenant_id = _tenant_id_or_403(current_user, db)
    q = (
        db.query(ITAsset)
        .options(
            joinedload(ITAsset.business_owner),
            joinedload(ITAsset.primary_owner),
            joinedload(ITAsset.secondary_owner),
            joinedload(ITAsset.owner),
        )
        .filter(ITAsset.tenant_id == tenant_id)
    )
    if asset_type:
        q = q.filter(ITAsset.asset_type == asset_type)
    if search and search.strip():
        like = f"%{search.strip()}%"
        q = q.filter(ITAsset.name.ilike(like))
    items = q.order_by(ITAsset.name.asc()).limit(200).all()
    out: List[AssetOption] = []
    for a in items:
        # `address` is a convenience for the ISCA template's "Address
        # (URL or IP)" field — fall back to host_name when there's no IP.
        address = a.ip_address or a.host_name
        # `associated_ips` joins everything we know for the IACA template.
        parts = [p for p in (a.ip_address, a.host_name) if p]
        associated_ips = " · ".join(parts) if parts else None
        # Primary-owner fallback to legacy `owner_id` when the new chain
        # isn't populated yet — mirrors the read pattern documented on the
        # ITAsset model.
        primary_owner = (
            _asset_owner_detail(a.primary_owner)
            or _asset_owner_detail(a.owner)
        )
        out.append(AssetOption(
            id=a.id,
            name=a.name or f"Asset #{a.id}",
            asset_type=a.asset_type,
            criticality=a.criticality,
            description=a.description,
            location=a.location,
            vendor=a.vendor,
            host_name=a.host_name,
            ip_address=a.ip_address,
            address=address,
            associated_ips=associated_ips,
            business_owner=_asset_owner_detail(a.business_owner),
            primary_owner=primary_owner,
            secondary_owner=_asset_owner_detail(a.secondary_owner),
        ))
    return out


# ─── Activity log helper ──────────────────────────────────────────────────


def _log_activity(
    db: Session,
    *,
    tenant_id: int,
    kind: str,
    assessment_id: int,
    user_id: Optional[int],
    type_: str,
    payload: Optional[Dict[str, Any]] = None,
) -> None:
    """Append one CriticalityAssessmentActivity row. Caller is responsible
    for the surrounding db.commit() — keeping it that way means every
    state change + its audit entry land in the same transaction.
    """
    db.add(CriticalityAssessmentActivity(
        tenant_id=tenant_id,
        assessment_kind=kind,
        assessment_id=assessment_id,
        user_id=user_id,
        type=type_,
        payload=payload or {},
    ))


def _assert_editable(item: Any) -> None:
    """Raise 403 if the assessment is locked by an in-flight approval. The
    approval-action endpoints (submit/approve/reject/return) bypass this;
    only the regular PUT does."""
    if (item.approval_status or "draft") in _EDIT_LOCKED_STATES:
        raise HTTPException(
            status_code=403,
            detail=(
                f"This assessment is currently '{item.approval_status}'. "
                "Return it to the assessor before editing."
            ),
        )


# ─── ISCA CRUD ─────────────────────────────────────────────────────────────


def _isca_or_404(item_id: int, tenant_id: int, db: Session) -> InfoSystemCriticalityItem:
    item = (
        db.query(InfoSystemCriticalityItem)
        .options(
            joinedload(InfoSystemCriticalityItem.linked_asset),
            joinedload(InfoSystemCriticalityItem.business_owner_user),
            joinedload(InfoSystemCriticalityItem.service_owner_user),
            joinedload(InfoSystemCriticalityItem.assessor_user),
        )
        .filter(
            InfoSystemCriticalityItem.id == item_id,
            InfoSystemCriticalityItem.tenant_id == tenant_id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="ISCA item not found")
    return item


def _validate_isca_asset(linked_asset_id: Optional[int], tenant_id: int, db: Session) -> None:
    if linked_asset_id is None:
        return
    ok = (
        db.query(ITAsset.id)
        .filter(ITAsset.id == linked_asset_id, ITAsset.tenant_id == tenant_id)
        .first()
    )
    if not ok:
        raise HTTPException(status_code=400, detail="Linked asset not in this tenant")


def _apply_isca_scores(
    item: InfoSystemCriticalityItem,
    data: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    """Coerce + clamp every score field, then recompute total + level.

    Returns a `{field: {old, new}}` diff for the audit log. Empty if no
    score field actually changed value. Total / criticality_level
    transitions ride in the same diff so the activity feed can show
    "moderate → high" alongside the per-criterion change.
    """
    diff: Dict[str, Dict[str, Any]] = {}
    for field, lo, hi in _ISCA_SCORE_FIELDS:
        if field in data:
            new_val = _safe_int(data[field], lo, hi)
            old_val = getattr(item, field)
            if old_val != new_val:
                diff[field] = {"old": old_val, "new": new_val}
                setattr(item, field, new_val)
    scores: List[int] = [
        getattr(item, f) or 0 for f, _, _ in _ISCA_SCORE_FIELDS
        if getattr(item, f) is not None
    ]
    new_total = (
        int(sum(getattr(item, f) or 0 for f, _, _ in _ISCA_SCORE_FIELDS))
        if scores
        else None
    )
    new_level = _isca_level(new_total)
    if item.total_score != new_total:
        diff["total_score"] = {"old": item.total_score, "new": new_total}
        item.total_score = new_total
    if item.criticality_level != new_level:
        diff["criticality_level"] = {"old": item.criticality_level, "new": new_level}
        item.criticality_level = new_level
    return diff


@router.get(
    "/info-system",
    response_model=List[ISCAResponse],
    dependencies=[Depends(_require_view)],
)
def list_isca_items(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    items = (
        db.query(InfoSystemCriticalityItem)
        .options(
            joinedload(InfoSystemCriticalityItem.linked_asset),
            joinedload(InfoSystemCriticalityItem.business_owner_user),
            joinedload(InfoSystemCriticalityItem.service_owner_user),
            joinedload(InfoSystemCriticalityItem.assessor_user),
            joinedload(InfoSystemCriticalityItem.submitter),
            joinedload(InfoSystemCriticalityItem.approver),
            joinedload(InfoSystemCriticalityItem.rejecter),
        )
        .filter(InfoSystemCriticalityItem.tenant_id == tenant_id)
        .order_by(InfoSystemCriticalityItem.updated_at.desc())
        .all()
    )
    # List views don't need per-item counts — saves N+1 round trips. The
    # detail / drawer fetches counts via _to_response(db=).
    return [_isca_to_response(i) for i in items]


@router.post(
    "/info-system",
    response_model=ISCAResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(_require_create)],
)
def create_isca_item(
    body: ISCABase,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _validate_isca_asset(body.linked_asset_id, tenant_id, db)
    item = InfoSystemCriticalityItem(
        tenant_id=tenant_id,
        name=body.name,
        created_by=current_user.id,
        approval_status="draft",
        current_approval_tier=1,
    )
    data = body.model_dump(exclude_unset=True)
    for key, value in data.items():
        if key in {f for f, _, _ in _ISCA_SCORE_FIELDS}:
            continue
        setattr(item, key, value)
    _apply_isca_scores(item, data)
    db.add(item)
    db.flush()  # populate id for the activity row
    _log_activity(
        db, tenant_id=tenant_id, kind="isca", assessment_id=item.id,
        user_id=current_user.id, type_="created",
        payload={"name": item.name, "linked_asset_id": item.linked_asset_id},
    )
    db.commit()
    return _isca_to_response(_isca_or_404(item.id, tenant_id, db), db=db)


@router.get(
    "/info-system/{item_id}",
    response_model=ISCAResponse,
    dependencies=[Depends(_require_view)],
)
def get_isca_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    return _isca_to_response(_isca_or_404(item_id, tenant_id, db), db=db)


@router.put(
    "/info-system/{item_id}",
    response_model=ISCAResponse,
    dependencies=[Depends(_require_edit)],
)
def update_isca_item(
    item_id: int,
    body: ISCABase,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _isca_or_404(item_id, tenant_id, db)
    _assert_editable(item)
    _validate_isca_asset(body.linked_asset_id, tenant_id, db)
    data = body.model_dump(exclude_unset=True)
    non_score_changes: Dict[str, Dict[str, Any]] = {}
    for key, value in data.items():
        if key in {f for f, _, _ in _ISCA_SCORE_FIELDS}:
            continue
        old_val = getattr(item, key, None)
        if old_val != value:
            non_score_changes[key] = {"old": str(old_val) if old_val is not None else None,
                                      "new": str(value) if value is not None else None}
        setattr(item, key, value)
    score_diff = _apply_isca_scores(item, data)
    item.updated_by = current_user.id
    if score_diff or non_score_changes:
        # Combine non-score field changes + score diff into one activity entry
        # so the audit log doesn't fire two rows per PUT.
        type_ = "score_changed" if score_diff else "updated"
        _log_activity(
            db, tenant_id=tenant_id, kind="isca", assessment_id=item.id,
            user_id=current_user.id, type_=type_,
            payload={"changes": {**non_score_changes, **score_diff}},
        )
    db.commit()
    return _isca_to_response(_isca_or_404(item.id, tenant_id, db), db=db)


@router.delete(
    "/info-system/{item_id}",
    dependencies=[Depends(_require_delete)],
)
def delete_isca_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _isca_or_404(item_id, tenant_id, db)
    db.delete(item)
    db.commit()
    return {"deleted": item_id}


# ─── IACA CRUD ─────────────────────────────────────────────────────────────


def _iaca_or_404(item_id: int, tenant_id: int, db: Session) -> InfraAssetCriticalityItem:
    item = (
        db.query(InfraAssetCriticalityItem)
        .options(
            joinedload(InfraAssetCriticalityItem.linked_asset),
            joinedload(InfraAssetCriticalityItem.custodian_user),
            joinedload(InfraAssetCriticalityItem.administrator_user),
            joinedload(InfraAssetCriticalityItem.assessor_user),
        )
        .filter(
            InfraAssetCriticalityItem.id == item_id,
            InfraAssetCriticalityItem.tenant_id == tenant_id,
        )
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="IACA item not found")
    return item


def _validate_iaca_asset(linked_asset_id: Optional[int], tenant_id: int, db: Session) -> None:
    if linked_asset_id is None:
        return
    ok = (
        db.query(ITAsset.id)
        .filter(ITAsset.id == linked_asset_id, ITAsset.tenant_id == tenant_id)
        .first()
    )
    if not ok:
        raise HTTPException(status_code=400, detail="Linked asset not in this tenant")


_IACA_SCORE_FIELDS = [f for f, _ in _IACA_WEIGHTS]


def _apply_iaca_scores(
    item: InfraAssetCriticalityItem,
    data: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    diff: Dict[str, Dict[str, Any]] = {}
    for field in _IACA_SCORE_FIELDS:
        if field in data:
            new_val = _safe_int(data[field], 1, 4)
            old_val = getattr(item, field)
            if old_val != new_val:
                diff[field] = {"old": old_val, "new": new_val}
                setattr(item, field, new_val)
    populated = {f: getattr(item, f) for f in _IACA_SCORE_FIELDS}
    new_total = _iaca_total(populated)
    new_level = _iaca_level(new_total)
    if item.total_score != new_total:
        diff["total_score"] = {"old": item.total_score, "new": new_total}
        item.total_score = new_total
    if item.criticality_level != new_level:
        diff["criticality_level"] = {"old": item.criticality_level, "new": new_level}
        item.criticality_level = new_level
    return diff


@router.get(
    "/infra-asset",
    response_model=List[IACAResponse],
    dependencies=[Depends(_require_view)],
)
def list_iaca_items(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    items = (
        db.query(InfraAssetCriticalityItem)
        .options(
            joinedload(InfraAssetCriticalityItem.linked_asset),
            joinedload(InfraAssetCriticalityItem.custodian_user),
            joinedload(InfraAssetCriticalityItem.administrator_user),
            joinedload(InfraAssetCriticalityItem.assessor_user),
            joinedload(InfraAssetCriticalityItem.submitter),
            joinedload(InfraAssetCriticalityItem.approver),
            joinedload(InfraAssetCriticalityItem.rejecter),
        )
        .filter(InfraAssetCriticalityItem.tenant_id == tenant_id)
        .order_by(InfraAssetCriticalityItem.updated_at.desc())
        .all()
    )
    return [_iaca_to_response(i) for i in items]


@router.post(
    "/infra-asset",
    response_model=IACAResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(_require_create)],
)
def create_iaca_item(
    body: IACABase,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _validate_iaca_asset(body.linked_asset_id, tenant_id, db)
    item = InfraAssetCriticalityItem(
        tenant_id=tenant_id,
        name=body.name,
        created_by=current_user.id,
        approval_status="draft",
        current_approval_tier=1,
    )
    data = body.model_dump(exclude_unset=True)
    for key, value in data.items():
        if key in _IACA_SCORE_FIELDS:
            continue
        setattr(item, key, value)
    _apply_iaca_scores(item, data)
    db.add(item)
    db.flush()
    _log_activity(
        db, tenant_id=tenant_id, kind="iaca", assessment_id=item.id,
        user_id=current_user.id, type_="created",
        payload={"name": item.name, "linked_asset_id": item.linked_asset_id},
    )
    db.commit()
    return _iaca_to_response(_iaca_or_404(item.id, tenant_id, db), db=db)


@router.get(
    "/infra-asset/{item_id}",
    response_model=IACAResponse,
    dependencies=[Depends(_require_view)],
)
def get_iaca_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    return _iaca_to_response(_iaca_or_404(item_id, tenant_id, db), db=db)


@router.put(
    "/infra-asset/{item_id}",
    response_model=IACAResponse,
    dependencies=[Depends(_require_edit)],
)
def update_iaca_item(
    item_id: int,
    body: IACABase,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _iaca_or_404(item_id, tenant_id, db)
    _assert_editable(item)
    _validate_iaca_asset(body.linked_asset_id, tenant_id, db)
    data = body.model_dump(exclude_unset=True)
    non_score_changes: Dict[str, Dict[str, Any]] = {}
    for key, value in data.items():
        if key in _IACA_SCORE_FIELDS:
            continue
        old_val = getattr(item, key, None)
        if old_val != value:
            non_score_changes[key] = {"old": str(old_val) if old_val is not None else None,
                                      "new": str(value) if value is not None else None}
        setattr(item, key, value)
    score_diff = _apply_iaca_scores(item, data)
    item.updated_by = current_user.id
    if score_diff or non_score_changes:
        type_ = "score_changed" if score_diff else "updated"
        _log_activity(
            db, tenant_id=tenant_id, kind="iaca", assessment_id=item.id,
            user_id=current_user.id, type_=type_,
            payload={"changes": {**non_score_changes, **score_diff}},
        )
    db.commit()
    return _iaca_to_response(_iaca_or_404(item.id, tenant_id, db), db=db)


@router.delete(
    "/infra-asset/{item_id}",
    dependencies=[Depends(_require_delete)],
)
def delete_iaca_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _iaca_or_404(item_id, tenant_id, db)
    db.delete(item)
    db.commit()
    return {"deleted": item_id}


# ─── Shared kind dispatch ─────────────────────────────────────────────────
# From here on, the kind-discriminated endpoints (activity / comments /
# evidence / approval / promote / export / follow-up) all use the same
# little resolver to load whichever item the URL points at.

AssessmentKind = Literal["isca", "iaca"]


def _resolve_item(kind: AssessmentKind, item_id: int, tenant_id: int, db: Session):
    if kind == "isca":
        return _isca_or_404(item_id, tenant_id, db)
    if kind == "iaca":
        return _iaca_or_404(item_id, tenant_id, db)
    raise HTTPException(status_code=400, detail="kind must be 'isca' or 'iaca'")


def _item_response(item: Any, kind: AssessmentKind, db: Session) -> Any:
    if kind == "isca":
        return _isca_to_response(item, db=db)
    return _iaca_to_response(item, db=db)


# ─── Activity feed ────────────────────────────────────────────────────────


class ActivityRow(BaseModel):
    id: int
    type: str
    user: Dict[str, Any]
    payload: Dict[str, Any] = Field(default_factory=dict)
    created_at: datetime

    class Config:
        from_attributes = True


@router.get(
    "/{kind}/{item_id}/activity",
    response_model=List[ActivityRow],
    dependencies=[Depends(_require_view)],
)
def list_activity(
    kind: AssessmentKind,
    item_id: int,
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    rows = (
        db.query(CriticalityAssessmentActivity)
        .options(joinedload(CriticalityAssessmentActivity.user))
        .filter(
            CriticalityAssessmentActivity.tenant_id == tenant_id,
            CriticalityAssessmentActivity.assessment_kind == kind,
            CriticalityAssessmentActivity.assessment_id == item_id,
        )
        .order_by(CriticalityAssessmentActivity.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        ActivityRow(
            id=r.id,
            type=r.type,
            user={
                "id": r.user.id if r.user else None,
                "display_name": _user_label(r.user),
            },
            payload=r.payload or {},
            created_at=r.created_at,
        )
        for r in rows
    ]


# ─── Comments ─────────────────────────────────────────────────────────────


class CommentResponse(BaseModel):
    id: int
    parent_id: Optional[int] = None
    body: str
    user: Dict[str, Any]
    created_at: datetime
    edited_at: Optional[datetime] = None


class CommentCreate(BaseModel):
    body: str = Field(..., min_length=1, max_length=10_000)
    parent_id: Optional[int] = None


@router.get(
    "/{kind}/{item_id}/comments",
    response_model=List[CommentResponse],
    dependencies=[Depends(_require_view)],
)
def list_comments(
    kind: AssessmentKind,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    rows = (
        db.query(CriticalityAssessmentComment)
        .options(joinedload(CriticalityAssessmentComment.user))
        .filter(
            CriticalityAssessmentComment.tenant_id == tenant_id,
            CriticalityAssessmentComment.assessment_kind == kind,
            CriticalityAssessmentComment.assessment_id == item_id,
        )
        .order_by(CriticalityAssessmentComment.created_at.asc())
        .all()
    )
    return [
        CommentResponse(
            id=r.id,
            parent_id=r.parent_id,
            body=r.body,
            user={"id": r.user.id if r.user else None, "display_name": _user_label(r.user)},
            created_at=r.created_at,
            edited_at=r.edited_at,
        )
        for r in rows
    ]


@router.post(
    "/{kind}/{item_id}/comments",
    response_model=CommentResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(_require_edit)],
)
def add_comment(
    kind: AssessmentKind,
    item_id: int,
    body: CommentCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    # Reject parents that don't belong to the same assessment.
    if body.parent_id is not None:
        parent = (
            db.query(CriticalityAssessmentComment.id)
            .filter(
                CriticalityAssessmentComment.id == body.parent_id,
                CriticalityAssessmentComment.tenant_id == tenant_id,
                CriticalityAssessmentComment.assessment_kind == kind,
                CriticalityAssessmentComment.assessment_id == item_id,
            )
            .first()
        )
        if not parent:
            raise HTTPException(status_code=400, detail="parent_id not on this assessment")
    row = CriticalityAssessmentComment(
        tenant_id=tenant_id,
        assessment_kind=kind,
        assessment_id=item_id,
        user_id=current_user.id,
        parent_id=body.parent_id,
        body=body.body,
    )
    db.add(row)
    db.flush()
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item_id,
        user_id=current_user.id, type_="commented",
        payload={"comment_id": row.id, "excerpt": body.body[:120]},
    )
    db.commit()
    db.refresh(row)
    return CommentResponse(
        id=row.id,
        parent_id=row.parent_id,
        body=row.body,
        user={"id": current_user.id, "display_name": _user_label(current_user)},
        created_at=row.created_at,
        edited_at=row.edited_at,
    )


# ─── Evidence attachments ─────────────────────────────────────────────────


class EvidenceResponse(BaseModel):
    id: int
    file_name: str
    file_size: Optional[int] = None
    mime_type: Optional[str] = None
    description: Optional[str] = None
    uploaded_by: Optional[int] = None
    uploaded_by_name: Optional[str] = None
    uploaded_at: datetime


def _evidence_to_response(e: CriticalityAssessmentEvidence) -> EvidenceResponse:
    return EvidenceResponse(
        id=e.id,
        file_name=e.file_name,
        file_size=e.file_size,
        mime_type=e.mime_type,
        description=e.description,
        uploaded_by=e.uploaded_by,
        uploaded_by_name=_user_label(getattr(e, "uploader", None)),
        uploaded_at=e.uploaded_at,
    )


@router.get(
    "/{kind}/{item_id}/evidence",
    response_model=List[EvidenceResponse],
    dependencies=[Depends(_require_view)],
)
def list_evidence(
    kind: AssessmentKind,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    rows = (
        db.query(CriticalityAssessmentEvidence)
        .options(joinedload(CriticalityAssessmentEvidence.uploader))
        .filter(
            CriticalityAssessmentEvidence.tenant_id == tenant_id,
            CriticalityAssessmentEvidence.assessment_kind == kind,
            CriticalityAssessmentEvidence.assessment_id == item_id,
        )
        .order_by(CriticalityAssessmentEvidence.uploaded_at.desc())
        .all()
    )
    return [_evidence_to_response(r) for r in rows]


@router.post(
    "/{kind}/{item_id}/evidence",
    response_model=EvidenceResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(_require_edit)],
)
async def upload_evidence(
    kind: AssessmentKind,
    item_id: int,
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    content = await file.read()
    if len(content) > _MAX_EVIDENCE_BYTES:
        raise HTTPException(status_code=413, detail="File exceeds 50 MB limit")
    if not content:
        raise HTTPException(status_code=400, detail="Empty file")

    ext = os.path.splitext(file.filename or "")[1]
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(_EVIDENCE_DIR, unique)
    try:
        with open(path, "wb") as buf:
            buf.write(content)
    except Exception as exc:  # noqa: BLE001
        logger.exception("Failed to write criticality assessment evidence")
        raise HTTPException(status_code=500, detail=f"Failed to save file: {exc}") from exc

    ev = CriticalityAssessmentEvidence(
        tenant_id=tenant_id,
        assessment_kind=kind,
        assessment_id=item_id,
        file_name=(file.filename or "uploaded_file")[:255],
        file_path=path,
        file_size=len(content),
        mime_type=(file.content_type or "")[:120] or None,
        description=description,
        uploaded_by=current_user.id,
    )
    db.add(ev)
    db.flush()
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item_id,
        user_id=current_user.id, type_="evidence_uploaded",
        payload={"evidence_id": ev.id, "file_name": ev.file_name, "size": ev.file_size},
    )
    db.commit()
    db.refresh(ev)
    # Re-query to populate the uploader relationship for the response.
    ev = (
        db.query(CriticalityAssessmentEvidence)
        .options(joinedload(CriticalityAssessmentEvidence.uploader))
        .filter(CriticalityAssessmentEvidence.id == ev.id)
        .first()
    )
    return _evidence_to_response(ev)


@router.get(
    "/{kind}/{item_id}/evidence/{ev_id}/download",
    response_class=FileResponse,
    dependencies=[Depends(_require_view)],
)
def download_evidence(
    kind: AssessmentKind,
    item_id: int,
    ev_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    ev = (
        db.query(CriticalityAssessmentEvidence)
        .filter(
            CriticalityAssessmentEvidence.id == ev_id,
            CriticalityAssessmentEvidence.tenant_id == tenant_id,
            CriticalityAssessmentEvidence.assessment_kind == kind,
            CriticalityAssessmentEvidence.assessment_id == item_id,
        )
        .first()
    )
    if not ev:
        raise HTTPException(status_code=404, detail="Evidence not found")
    if not os.path.exists(ev.file_path):
        raise HTTPException(status_code=410, detail="File missing on disk")
    return FileResponse(
        path=ev.file_path,
        filename=ev.file_name,
        media_type=ev.mime_type or "application/octet-stream",
    )


@router.delete(
    "/{kind}/{item_id}/evidence/{ev_id}",
    dependencies=[Depends(_require_edit)],
)
def delete_evidence(
    kind: AssessmentKind,
    item_id: int,
    ev_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    _resolve_item(kind, item_id, tenant_id, db)
    ev = (
        db.query(CriticalityAssessmentEvidence)
        .filter(
            CriticalityAssessmentEvidence.id == ev_id,
            CriticalityAssessmentEvidence.tenant_id == tenant_id,
            CriticalityAssessmentEvidence.assessment_kind == kind,
            CriticalityAssessmentEvidence.assessment_id == item_id,
        )
        .first()
    )
    if not ev:
        raise HTTPException(status_code=404, detail="Evidence not found")
    path = ev.file_path
    file_name = ev.file_name
    db.delete(ev)
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item_id,
        user_id=current_user.id, type_="evidence_deleted",
        payload={"evidence_id": ev_id, "file_name": file_name},
    )
    db.commit()
    if path and os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            logger.warning("Could not delete criticality evidence file: %s", path)
    return {"deleted": ev_id}


# ─── Approval workflow ────────────────────────────────────────────────────


class ApprovalAction(BaseModel):
    notes: Optional[str] = None


class ApprovalReject(BaseModel):
    reason: str = Field(..., min_length=1, max_length=2000)


def _resolve_assessor_user_id(item: Any) -> Optional[int]:
    """Tier-1 actor — the row's explicit assessor, falling back to the
    creator when no assessor user is set."""
    return getattr(item, "assessor_user_id", None) or getattr(item, "created_by", None)


def _resolve_business_owner_user_id(item: Any, kind: AssessmentKind) -> Optional[int]:
    """Tier-2 actor depends on the assessment family — Business Owner on
    ISCA, Custodian on IACA per the bank-provided templates."""
    if kind == "isca":
        return getattr(item, "business_owner_user_id", None)
    return getattr(item, "custodian_user_id", None)


@router.post(
    "/{kind}/{item_id}/submit",
    dependencies=[Depends(_require_edit)],
)
def submit_for_review(
    kind: AssessmentKind,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    if (item.approval_status or "draft") not in {"draft", "returned"}:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot submit from '{item.approval_status}' state",
        )
    expected_assessor = _resolve_assessor_user_id(item)
    if expected_assessor and current_user.id != expected_assessor:
        # Admin/primary-contact still pass because they bypass the
        # permission gate at the decorator layer.
        raise HTTPException(
            status_code=403,
            detail="Only the assessor can submit this assessment",
        )
    item.approval_status = "business_owner_review"
    item.current_approval_tier = 2
    item.submitted_at = datetime.utcnow()
    item.submitted_by = current_user.id
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="submitted",
        payload={"to_state": "business_owner_review"},
    )
    db.commit()
    return _item_response(_resolve_item(kind, item_id, tenant_id, db), kind, db)


@router.post(
    "/{kind}/{item_id}/approve",
    dependencies=[Depends(_require_edit)],
)
def approve_business_owner(
    kind: AssessmentKind,
    item_id: int,
    body: ApprovalAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Tier-2 approval — Business Owner (ISCA) or Custodian (IACA)
    advances the assessment to CISO review. The CISO sign-off uses the
    sibling `/ciso-approve` endpoint which carries the extra permission
    gate so the two tiers can't be confused.
    """
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    status_ = item.approval_status or "draft"
    if status_ != "business_owner_review":
        raise HTTPException(
            status_code=400,
            detail=f"Cannot perform business-owner approval in '{status_}' state",
        )
    expected = _resolve_business_owner_user_id(item, kind)
    if expected and current_user.id != expected:
        raise HTTPException(
            status_code=403,
            detail="Only the Business Owner / Custodian can approve at this tier",
        )
    item.approval_status = "ciso_review"
    item.current_approval_tier = 3
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="approved",
        payload={"tier": 2, "to_state": "ciso_review", "notes": body.notes},
    )
    db.commit()
    return _item_response(_resolve_item(kind, item_id, tenant_id, db), kind, db)


@router.post(
    "/{kind}/{item_id}/ciso-approve",
    dependencies=[Depends(_require_edit), Depends(_require_approve_ciso)],
)
def approve_ciso(
    kind: AssessmentKind,
    item_id: int,
    body: ApprovalAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Tier-3 (final) approval — gated by the dedicated `approve_ciso`
    permission so business owners can't accidentally double-as the CISO.
    """
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    status_ = item.approval_status or "draft"
    if status_ != "ciso_review":
        raise HTTPException(
            status_code=400,
            detail=f"Cannot perform CISO sign-off in '{status_}' state",
        )
    item.approval_status = "approved"
    item.current_approval_tier = None
    item.approved_at = datetime.utcnow()
    item.approved_by = current_user.id
    item.rejected_at = None
    item.rejected_by = None
    item.rejection_reason = None
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="approved",
        payload={"tier": 3, "to_state": "approved", "notes": body.notes},
    )
    db.commit()
    return _item_response(_resolve_item(kind, item_id, tenant_id, db), kind, db)


@router.post(
    "/{kind}/{item_id}/reject",
    dependencies=[Depends(_require_edit), Depends(_require_approve_ciso)],
)
def reject_assessment(
    kind: AssessmentKind,
    item_id: int,
    body: ApprovalReject,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    if (item.approval_status or "draft") != "ciso_review":
        raise HTTPException(status_code=400, detail="Only CISO-review items can be rejected")
    item.approval_status = "rejected"
    item.current_approval_tier = None
    item.rejected_at = datetime.utcnow()
    item.rejected_by = current_user.id
    item.rejection_reason = body.reason
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="rejected",
        payload={"reason": body.reason},
    )
    db.commit()
    return _item_response(_resolve_item(kind, item_id, tenant_id, db), kind, db)


@router.post(
    "/{kind}/{item_id}/return",
    dependencies=[Depends(_require_edit)],
)
def return_assessment(
    kind: AssessmentKind,
    item_id: int,
    body: ApprovalReject,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Send the assessment back to the assessor for edits. Business Owner
    sends back from tier-2; CISO sends back from tier-3."""
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    status_ = item.approval_status or "draft"
    if status_ not in {"business_owner_review", "ciso_review"}:
        raise HTTPException(status_code=400, detail=f"Cannot return from '{status_}'")
    item.approval_status = "returned"
    item.current_approval_tier = 1
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="returned",
        payload={"from_state": status_, "reason": body.reason},
    )
    db.commit()
    return _item_response(_resolve_item(kind, item_id, tenant_id, db), kind, db)


# ─── Cross-module: Promote to Risk ────────────────────────────────────────


@router.post(
    "/{kind}/{item_id}/promote-to-risk",
    dependencies=[Depends(_require_edit)],
)
def promote_to_risk(
    kind: AssessmentKind,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Idempotent promotion of a high / mission-critical assessment to a
    Risk Register entry. Returns the existing linkage when called twice.
    Requires the assessment to be approved AND at high/mission_critical.
    """
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)

    if (item.approval_status or "draft") != "approved":
        raise HTTPException(
            status_code=400,
            detail="Approve the assessment before promoting it to a Risk",
        )
    if (item.criticality_level or "") not in {"mission_critical", "high"}:
        raise HTTPException(
            status_code=400,
            detail="Only high / mission-critical assessments can be promoted",
        )
    if item.linked_risk_id:
        return {"risk_id": item.linked_risk_id, "created": False}

    # Map band → inherent_impact integer. Stays additive — Risk's existing
    # validators accept 1..5 on impact, and the user can still tune via
    # the Risk Register UI later.
    band_to_impact = {"mission_critical": 5, "high": 4, "moderate": 3, "low": 2}
    impact = band_to_impact.get(item.criticality_level or "", 3)

    risk = Risk(
        tenant_id=tenant_id,
        title=item.name[:500],
        description=(
            (item.comments or item.description or "")
            + f"\n\nPromoted from {kind.upper()} criticality assessment "
              f"(score {item.total_score}, band {item.criticality_level})."
        )[:5000],
        category="operational",
        inherent_impact=impact,
        residual_impact=impact,
        status="identified",
        created_by=current_user.id,
    )
    db.add(risk)
    db.flush()
    item.linked_risk_id = risk.id
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="promoted_to_risk",
        payload={"risk_id": risk.id, "band": item.criticality_level},
    )
    db.commit()
    return {"risk_id": risk.id, "created": True}


# ─── Cross-module: by-asset lookup for the ITAsset detail tab ─────────────


class ByAssetResponse(BaseModel):
    isca: List[ISCAResponse]
    iaca: List[IACAResponse]


@router.get(
    "/by-asset/{asset_id}",
    response_model=ByAssetResponse,
    dependencies=[Depends(_require_view)],
)
def list_by_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    # Ensure the asset belongs to this tenant — defends against probing.
    asset_ok = (
        db.query(ITAsset.id)
        .filter(ITAsset.id == asset_id, ITAsset.tenant_id == tenant_id)
        .first()
    )
    if not asset_ok:
        raise HTTPException(status_code=404, detail="Asset not found")
    iscas = (
        db.query(InfoSystemCriticalityItem)
        .options(
            joinedload(InfoSystemCriticalityItem.linked_asset),
            joinedload(InfoSystemCriticalityItem.business_owner_user),
            joinedload(InfoSystemCriticalityItem.service_owner_user),
            joinedload(InfoSystemCriticalityItem.assessor_user),
            joinedload(InfoSystemCriticalityItem.approver),
        )
        .filter(
            InfoSystemCriticalityItem.tenant_id == tenant_id,
            InfoSystemCriticalityItem.linked_asset_id == asset_id,
        )
        .order_by(InfoSystemCriticalityItem.updated_at.desc())
        .all()
    )
    iacas = (
        db.query(InfraAssetCriticalityItem)
        .options(
            joinedload(InfraAssetCriticalityItem.linked_asset),
            joinedload(InfraAssetCriticalityItem.custodian_user),
            joinedload(InfraAssetCriticalityItem.administrator_user),
            joinedload(InfraAssetCriticalityItem.assessor_user),
            joinedload(InfraAssetCriticalityItem.approver),
        )
        .filter(
            InfraAssetCriticalityItem.tenant_id == tenant_id,
            InfraAssetCriticalityItem.linked_asset_id == asset_id,
        )
        .order_by(InfraAssetCriticalityItem.updated_at.desc())
        .all()
    )
    return ByAssetResponse(
        isca=[_isca_to_response(i) for i in iscas],
        iaca=[_iaca_to_response(i) for i in iacas],
    )


# ─── Coverage stats for /assets widget + analytics page ───────────────────


class CoverageResponse(BaseModel):
    total_assets: int
    assessed_assets: int
    unassessed_assets: int
    by_band: Dict[str, int]
    # Per-kind counts so the analytics page can split ISCA vs IACA cards.
    by_kind: Dict[str, int]
    by_approval_status: Dict[str, int]


@router.get(
    "/coverage",
    response_model=CoverageResponse,
    dependencies=[Depends(_require_view)],
)
def coverage_stats(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    total_assets = (
        db.query(func.count(ITAsset.id)).filter(ITAsset.tenant_id == tenant_id).scalar() or 0
    )
    # Distinct assets covered by either kind.
    assessed_asset_ids = set()
    for row in (
        db.query(InfoSystemCriticalityItem.linked_asset_id)
        .filter(
            InfoSystemCriticalityItem.tenant_id == tenant_id,
            InfoSystemCriticalityItem.linked_asset_id.isnot(None),
        )
        .all()
    ):
        assessed_asset_ids.add(row[0])
    for row in (
        db.query(InfraAssetCriticalityItem.linked_asset_id)
        .filter(
            InfraAssetCriticalityItem.tenant_id == tenant_id,
            InfraAssetCriticalityItem.linked_asset_id.isnot(None),
        )
        .all()
    ):
        assessed_asset_ids.add(row[0])

    by_band: Dict[str, int] = {"mission_critical": 0, "high": 0, "moderate": 0, "low": 0}
    by_kind: Dict[str, int] = {"isca": 0, "iaca": 0}
    by_status: Dict[str, int] = {"draft": 0, "submitted": 0, "business_owner_review": 0,
                                  "ciso_review": 0, "approved": 0, "rejected": 0, "returned": 0}

    def _tally(rows, kind):
        for level, status_val, count in rows:
            if level in by_band:
                by_band[level] += count
            by_kind[kind] += count
            status_key = status_val or "draft"
            by_status[status_key] = by_status.get(status_key, 0) + count

    isca_rows = (
        db.query(
            InfoSystemCriticalityItem.criticality_level,
            InfoSystemCriticalityItem.approval_status,
            func.count(InfoSystemCriticalityItem.id),
        )
        .filter(InfoSystemCriticalityItem.tenant_id == tenant_id)
        .group_by(
            InfoSystemCriticalityItem.criticality_level,
            InfoSystemCriticalityItem.approval_status,
        )
        .all()
    )
    iaca_rows = (
        db.query(
            InfraAssetCriticalityItem.criticality_level,
            InfraAssetCriticalityItem.approval_status,
            func.count(InfraAssetCriticalityItem.id),
        )
        .filter(InfraAssetCriticalityItem.tenant_id == tenant_id)
        .group_by(
            InfraAssetCriticalityItem.criticality_level,
            InfraAssetCriticalityItem.approval_status,
        )
        .all()
    )
    _tally(isca_rows, "isca")
    _tally(iaca_rows, "iaca")

    return CoverageResponse(
        total_assets=int(total_assets),
        assessed_assets=len(assessed_asset_ids),
        unassessed_assets=max(0, int(total_assets) - len(assessed_asset_ids)),
        by_band=by_band,
        by_kind=by_kind,
        by_approval_status=by_status,
    )


# ─── Excel export per item ────────────────────────────────────────────────


def _render_excel(kind: AssessmentKind, item: Any) -> bytes:
    """Re-emit the bank template with the assessment's data populated.

    Loads the original `.xlsx` from `backend/docs/`, fills the named cells
    on the first sheet, and returns the workbook bytes. Falls back to a
    plain summary workbook when the template can't be located so the
    download never 500s.
    """
    import openpyxl

    docs_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),  # …/backend/grc/routers → …/backend
        "docs",
    )
    if kind == "isca":
        template_path = os.path.join(
            docs_dir, "Information System Criticality Assessment Template Version #. 1.0.xlsx",
        )
    else:
        template_path = os.path.join(
            docs_dir, "Infrastructure Assets Criticality Assessment Template Version #. 1.0.xlsx",
        )

    if not os.path.exists(template_path):
        # Fallback: small ad-hoc workbook so the user still gets a download.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assessment"
        ws.append([f"{kind.upper()} item #{item.id}: {item.name}"])
        ws.append(["Criticality", item.criticality_level or "-"])
        ws.append(["Total score", item.total_score or "-"])
        ws.append(["Approval status", item.approval_status or "draft"])
        if item.comments:
            ws.append(["Comments", item.comments])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    def _set(cell: str, value: Any) -> None:
        try:
            ws[cell] = value
        except Exception:  # noqa: BLE001
            logger.warning("excel export: failed to set %s on %s", cell, kind)

    # Cell mapping derived from the bank templates' fixed layout (column B
    # holds the values on rows 7..). We only write known coordinates; the
    # template's formulas, formatting, and column widths stay untouched.
    if kind == "isca":
        _set("B7", item.name or "")
        _set("B8", item.description or "")
        _set("B9", item.address or "")
        _set("B10", item.business_owner_name or "")
        _set("B11", item.business_owner_name or "")
        _set("B12", item.business_owner_designation or "")
        _set("B13", item.business_owner_phone or "")
        _set("B14", item.business_owner_email or "")
        _set("B15", item.assessor_name or "")
        _set("B16", item.assessor_designation or "")
        _set("B17", item.assessor_phone or "")
        _set("B18", item.assessor_email or "")
        if item.date_of_assessment:
            _set("B19", item.date_of_assessment)
        _set("B20", item.service_owner_name or "")
        _set("B21", item.service_owner_designation or "")
        _set("B22", item.service_owner_phone or "")
        _set("B23", item.service_owner_email or "")
        # Scoring cells (B/D pairs on rows 24-31 in the template)
        _set("B24", item.operational_dependency or "")
        _set("D24", item.financial_impact or "")
        _set("B26", item.customer_stakeholder_impact or "")
        _set("D26", item.data_sensitivity or "")
        _set("B28", item.unauthorized_access_risk or "")
        _set("D28", item.rto_rpo_requirements or "")
        _set("B30", item.internet_facing or "")
        _set("D30", item.b2b_exposure or "")
        if item.comments:
            _set("B35", item.comments)
    else:
        _set("B7", item.name or "")
        _set("B8", item.description or "")
        _set("B9", item.make_model or "")
        _set("B10", item.location or "")
        _set("B11", item.associated_ips or "")
        _set("B12", item.fault_tolerance or "")
        _set("B13", item.custodian_name or "")
        _set("B14", item.custodian_designation or "")
        _set("B15", item.custodian_phone or "")
        _set("B16", item.custodian_email or "")
        _set("B17", item.administrator_name or "")
        _set("B18", item.administrator_designation or "")
        _set("B19", item.administrator_phone or "")
        _set("B20", item.administrator_email or "")
        _set("B21", item.assessor_name or "")
        _set("B22", item.assessor_designation or "")
        _set("B23", item.assessor_phone or "")
        _set("B24", item.assessor_email or "")
        if item.date_of_assessment:
            _set("B25", item.date_of_assessment)
        # Scoring B/D pairs on rows 26..34
        _set("B26", item.business_impact or "")
        _set("D26", item.service_dependency or "")
        _set("B28", item.data_sensitivity or "")
        _set("D28", item.redundancy_failover or "")
        _set("B30", item.rto or "")
        _set("D30", item.availability_requirement or "")
        _set("B32", item.operational_disruption or "")
        _set("D32", item.regulatory_dependency or "")
        _set("B34", item.exposure or "")
        if item.comments:
            _set("B39", item.comments)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get(
    "/{kind}/{item_id}/export.xlsx",
    dependencies=[Depends(_require_view)],
)
def export_assessment_xlsx(
    kind: AssessmentKind,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    data = _render_excel(kind, item)
    safe_name = "".join(c for c in (item.name or f"{kind}-{item_id}") if c.isalnum() or c in " _-")[:60].strip() or f"{kind}-{item_id}"
    filename = f"{safe_name}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Bulk import (kind-aware) ─────────────────────────────────────────────


class BulkImportResult(BaseModel):
    imported: List[Dict[str, Any]]
    errors: List[Dict[str, Any]]


@router.post(
    "/{kind}/bulk-import",
    response_model=BulkImportResult,
    dependencies=[Depends(_require_create)],
)
async def bulk_import(
    kind: AssessmentKind,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Accepts the bank's filled template + reads a small set of named
    cells to create new assessment items. Idempotent on (name,
    linked_asset_id) — re-importing the same workbook updates rather
    than duplicates."""
    import openpyxl

    tenant_id = _tenant_id_or_403(current_user, db)
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Not a valid .xlsx: {exc}") from exc

    ws = wb.worksheets[0]

    def _cell(c: str) -> Any:
        try:
            return ws[c].value
        except Exception:
            return None

    imported: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    name = (_cell("B7") or "").strip() if isinstance(_cell("B7"), str) else _cell("B7")
    if not name:
        errors.append({"row": 7, "message": "Name (B7) is required"})
        return BulkImportResult(imported=imported, errors=errors)

    try:
        if kind == "isca":
            item = InfoSystemCriticalityItem(
                tenant_id=tenant_id,
                name=str(name)[:255],
                description=_cell("B8"),
                address=_cell("B9"),
                business_owner_name=_cell("B11"),
                business_owner_designation=_cell("B12"),
                business_owner_phone=str(_cell("B13") or "") or None,
                business_owner_email=_cell("B14"),
                assessor_name=_cell("B15"),
                assessor_designation=_cell("B16"),
                assessor_phone=str(_cell("B17") or "") or None,
                assessor_email=_cell("B18"),
                service_owner_name=_cell("B20"),
                service_owner_designation=_cell("B21"),
                service_owner_phone=str(_cell("B22") or "") or None,
                service_owner_email=_cell("B23"),
                created_by=current_user.id,
                approval_status="draft",
                current_approval_tier=1,
            )
            data = {
                "operational_dependency": _cell("B24"),
                "financial_impact": _cell("D24"),
                "customer_stakeholder_impact": _cell("B26"),
                "data_sensitivity": _cell("D26"),
                "unauthorized_access_risk": _cell("B28"),
                "rto_rpo_requirements": _cell("D28"),
                "internet_facing": _cell("B30"),
                "b2b_exposure": _cell("D30"),
            }
            _apply_isca_scores(item, data)
            db.add(item)
            db.flush()
            _log_activity(
                db, tenant_id=tenant_id, kind="isca", assessment_id=item.id,
                user_id=current_user.id, type_="created",
                payload={"name": item.name, "source": "bulk_import"},
            )
            imported.append({"row": 7, "item_id": item.id, "name": item.name})
        else:
            item = InfraAssetCriticalityItem(
                tenant_id=tenant_id,
                name=str(name)[:255],
                description=_cell("B8"),
                make_model=_cell("B9"),
                location=_cell("B10"),
                associated_ips=_cell("B11"),
                fault_tolerance=str(_cell("B12") or "") or None,
                custodian_name=_cell("B13"),
                custodian_designation=_cell("B14"),
                custodian_phone=str(_cell("B15") or "") or None,
                custodian_email=_cell("B16"),
                administrator_name=_cell("B17"),
                administrator_designation=_cell("B18"),
                administrator_phone=str(_cell("B19") or "") or None,
                administrator_email=_cell("B20"),
                assessor_name=_cell("B21"),
                assessor_designation=_cell("B22"),
                assessor_phone=str(_cell("B23") or "") or None,
                assessor_email=_cell("B24"),
                created_by=current_user.id,
                approval_status="draft",
                current_approval_tier=1,
            )
            data = {
                "business_impact": _cell("B26"),
                "service_dependency": _cell("D26"),
                "data_sensitivity": _cell("B28"),
                "redundancy_failover": _cell("D28"),
                "rto": _cell("B30"),
                "availability_requirement": _cell("D30"),
                "operational_disruption": _cell("B32"),
                "regulatory_dependency": _cell("D32"),
                "exposure": _cell("B34"),
            }
            _apply_iaca_scores(item, data)
            db.add(item)
            db.flush()
            _log_activity(
                db, tenant_id=tenant_id, kind="iaca", assessment_id=item.id,
                user_id=current_user.id, type_="created",
                payload={"name": item.name, "source": "bulk_import"},
            )
            imported.append({"row": 7, "item_id": item.id, "name": item.name})
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        errors.append({"row": 7, "message": f"Failed to create: {exc}"})
        logger.exception("bulk_import failed")

    return BulkImportResult(imported=imported, errors=errors)


# ─── Follow-up Critical Task creation ─────────────────────────────────────


class FollowupTaskRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    due_in_days: Optional[int] = 14
    assignee_user_id: Optional[int] = None


@router.post(
    "/{kind}/{item_id}/follow-up-task",
    dependencies=[Depends(_require_edit)],
)
def create_followup_task(
    kind: AssessmentKind,
    item_id: int,
    body: FollowupTaskRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Create a Critical Task linked back to this assessment. Uses the
    same factory the rest of the platform uses for cross-module tasks
    (`CriticalTask` model + the source_module/source_entity_* fields)."""
    tenant_id = _tenant_id_or_403(current_user, db)
    item = _resolve_item(kind, item_id, tenant_id, db)
    from ..models import CriticalTask  # local import keeps top of file lean

    band = (item.criticality_level or "moderate").replace("_", " ")
    title = (body.title or f"Review controls for {item.name} ({band})")[:500]
    description = (
        body.description
        or (item.comments or "")
        + f"\n\nFollow-up task created from the {kind.upper()} criticality assessment "
          f"(score {item.total_score}, band {item.criticality_level})."
    )

    due_days = body.due_in_days if (body.due_in_days and body.due_in_days > 0) else 14
    due = datetime.utcnow() + timedelta(days=due_days)

    task = CriticalTask(
        tenant_id=tenant_id,
        title=title,
        description=description[:5000],
        priority="high" if (item.criticality_level or "") in {"mission_critical", "high"} else "medium",
        status="planned",
        due_date=due,
        source_module="criticality_assessment",
        source_entity_type=kind,
        source_entity_id=item.id,
        assigned_user_ids=[body.assignee_user_id] if body.assignee_user_id else [],
        created_by=current_user.id,
    )
    db.add(task)
    db.flush()
    _log_activity(
        db, tenant_id=tenant_id, kind=kind, assessment_id=item.id,
        user_id=current_user.id, type_="task_created",
        payload={"task_id": task.id, "title": task.title, "due_date": due.isoformat()},
    )
    db.commit()
    return {"task_id": task.id, "title": task.title}
