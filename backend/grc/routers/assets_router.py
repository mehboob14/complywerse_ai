from ..config import get_openai_model
import random
import csv
import io
import os
import json
import re
from functools import lru_cache
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Request, Cookie, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text, func
from pydantic import BaseModel

from ..models import (
    ITAsset, AssetControlLink, AssetInternalControlLink, AssetRiskAssessment, AssetFrameworkControlLink,
    AssetEvidenceLink, NormalizedControl, FrameworkControl, Evidence, Risk, InternalControl,
    Vulnerability, VulnerabilityAssetLink, AssetSecurityComplianceSelection,
    SoftwareIdentifier,
    # Trajectory-map endpoint pulls these in to traverse Asset → Vuln → Control → Risk
    VulnerabilityControlLink, ParsedFrameworkControl,
    RiskAssetLink, RiskControlLink, RiskFrameworkControlLink,
    Framework,
    GRCUser, Tenant, TenantUser, get_db
)
from ..schemas import (
    ITAssetCreate, ITAssetUpdate, ITAssetResponse,
    AssetValuation, AssetControlLinkCreate, AssetRiskAssessmentResponse,
    AssetDashboard, AssetCoverage, MessageResponse,
    AssetFrameworkControlLinkCreate, AssetEvidenceLinkCreate,
    AssetDetailResponse, AssetCoverageAnalysis,
    LifecycleTransitionRequest,
)
from .auth_router import require_auth, get_user_tenants, get_user_primary_tenant, decode_token
# Phase 5.4 + 5.3 helpers — keep all logic out of the router body.
from ..services.asset_criticality import recompute_for_asset as recompute_asset_criticality
from ..services import asset_lifecycle
from ..services.asset_control_recommender import recommend_for_asset

import logging

logger = logging.getLogger(__name__)
# Per-database-per-tenant: the tenant DB *is* the active session, so tenant
# users are just GRCUser rows in the current request's session.

router = APIRouter(prefix="/assets", tags=["IT Assets"])

SECURITY_COMPLIANCE_BENCHMARK = "CIS_WS2012R2"
SECURITY_COMPLIANCE_CONTROLS_FILE = os.path.abspath(
    os.path.join(
        os.path.dirname(__file__),
        "..",
        "seed_data",
        "CIS",
        "CIS_WS2012R2_Controls.json",
    )
)


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


def _security_compliance_sort_tokens(value: str) -> Tuple[Any, ...]:
    """Natural sort for control IDs with sub-points like 1.2.10."""
    chunks = re.findall(r"\d+|[^\d]+", str(value or ""))
    tokens: List[Any] = []
    for chunk in chunks:
        if chunk.isdigit():
            tokens.append(int(chunk))
        else:
            tokens.append(chunk.lower())
    return tuple(tokens)


@lru_cache(maxsize=1)
def _load_security_compliance_controls() -> Dict[str, Any]:
    if not os.path.exists(SECURITY_COMPLIANCE_CONTROLS_FILE):
        raise FileNotFoundError(
            f"Security compliance controls file not found: {SECURITY_COMPLIANCE_CONTROLS_FILE}"
        )
    with open(SECURITY_COMPLIANCE_CONTROLS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_asset_for_user(asset_id: int, current_user: GRCUser, db: Session) -> ITAsset:
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found",
        )
    return asset


class SecurityComplianceSelectionRequest(BaseModel):
    control_ids: List[str]


@router.get("/tenant-users")
def get_tenant_users(
    http_request: Request,
    token: Optional[str] = Cookie(None, alias="grc_auth_token"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Get all users in the current tenant DB for owner selection."""
    users = db.query(GRCUser).filter(GRCUser.is_active == True).all()
    return [
        {
            "id": u.id,
            "display_name": u.display_name or u.username,
            "email": u.email
        }
        for u in users
    ]


class CIARecommendationRequest(BaseModel):
    name: str
    description: Optional[str] = None
    asset_type: str
    vendor: Optional[str] = None
    location: Optional[str] = None
    criticality: Optional[str] = None


@router.post("/cia-recommendation")
def get_cia_recommendation(
    request: CIARecommendationRequest,
    current_user: GRCUser = Depends(require_auth)
):
    """Get AI-driven CIA rating recommendations based on asset details"""
    
    # Check if OpenAI API key is available
    openai_key = os.getenv("OPENAI_API_KEY")
    
    if not openai_key:
        # Fallback to rule-based recommendations
        return get_rule_based_cia_recommendation(request)
    
    try:
        import openai
        openai.api_key = openai_key
        
        # Build prompt for AI
        prompt = f"""Based on the following IT asset information, recommend appropriate CIA (Confidentiality, Integrity, Availability) ratings on a scale of 1-5 where:
- 1 = Very Low
- 2 = Low  
- 3 = Medium
- 4 = High
- 5 = Very High/Critical

Asset Details:
- Name: {request.name}
- Type: {request.asset_type}
- Description: {request.description or 'Not provided'}
- Vendor: {request.vendor or 'Not provided'}
- Location: {request.location or 'Not provided'}
- Business Criticality: {request.criticality or 'Not specified'}

Provide a brief 2-line recommendation explaining the suggested CIA ratings (Confidentiality, Integrity, Availability) and the specific numeric ratings.
Format: First line with explanation, second line with ratings in format "Recommended: C=X, I=Y, A=Z"
"""
        
        response = openai.ChatCompletion.create(
            model=get_openai_model(),
            messages=[
                {"role": "system", "content": "You are an IT security expert specializing in asset risk assessment and CIA triad ratings."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=200
        )
        
        recommendation_text = response.choices[0].message.content.strip()
        
        # Parse ratings from response
        lines = recommendation_text.split('\n')
        rating_line = next((line for line in lines if 'C=' in line and 'I=' in line and 'A=' in line), '')
        
        conf_rating = 3
        int_rating = 3
        avail_rating = 3
        
        if rating_line:
            import re
            conf_match = re.search(r'C=(\d)', rating_line)
            int_match = re.search(r'I=(\d)', rating_line)
            avail_match = re.search(r'A=(\d)', rating_line)
            
            if conf_match:
                conf_rating = int(conf_match.group(1))
            if int_match:
                int_rating = int(int_match.group(1))
            if avail_match:
                avail_rating = int(avail_match.group(1))
        
        # Format recommendation as 2 lines
        recommendation_lines = [line for line in lines if line.strip() and not line.startswith('Recommended:')][:2]
        if len(recommendation_lines) < 2 and rating_line:
            recommendation_lines.append(rating_line)
        
        recommendation = '\n'.join(recommendation_lines[:2])
        
        return {
            "recommendation": recommendation,
            "confidentiality_rating": conf_rating,
            "integrity_rating": int_rating,
            "availability_rating": avail_rating
        }
        
    except Exception as e:
        # Fallback to rule-based on error
        return get_rule_based_cia_recommendation(request)


def get_rule_based_cia_recommendation(request: CIARecommendationRequest):
    """Rule-based CIA recommendations when AI is unavailable"""
    
    asset_type = request.asset_type.lower()
    criticality = (request.criticality or 'medium').lower()
    
    # Base ratings
    conf_rating = 3
    int_rating = 3
    avail_rating = 3
    
    # Adjust by asset type
    if asset_type == 'data':
        conf_rating = 5
        int_rating = 5
        avail_rating = 4
        recommendation = "Data assets require maximum protection for confidentiality and integrity to prevent unauthorized access and corruption.\nRecommended: C=5, I=5, A=4"
    
    elif asset_type == 'application':
        conf_rating = 4
        int_rating = 4
        avail_rating = 4
        recommendation = "Business applications need high protection across all CIA dimensions to ensure secure and reliable operations.\nRecommended: C=4, I=4, A=4"
    
    elif asset_type == 'infrastructure':
        conf_rating = 3
        int_rating = 4
        avail_rating = 5
        recommendation = "Infrastructure assets prioritize availability and integrity to maintain operational continuity and system reliability.\nRecommended: C=3, I=4, A=5"
    
    elif asset_type == 'cloud':
        conf_rating = 4
        int_rating = 4
        avail_rating = 5
        recommendation = "Cloud resources require high availability and strong confidentiality controls due to shared responsibility model.\nRecommended: C=4, I=4, A=5"
    
    elif asset_type == 'third_party':
        conf_rating = 4
        int_rating = 3
        avail_rating = 3
        recommendation = "Third-party systems need elevated confidentiality controls due to external access and data sharing requirements.\nRecommended: C=4, I=3, A=3"
    
    else:
        recommendation = "Standard protection levels recommended based on general IT asset best practices and industry standards.\nRecommended: C=3, I=3, A=3"
    
    # Adjust by criticality
    if criticality == 'critical':
        conf_rating = min(5, conf_rating + 1)
        int_rating = min(5, int_rating + 1)
        avail_rating = min(5, avail_rating + 1)
    elif criticality == 'high':
        avail_rating = min(5, avail_rating + 1)
    elif criticality == 'low':
        conf_rating = max(1, conf_rating - 1)
        int_rating = max(1, int_rating - 1)
    
    return {
        "recommendation": recommendation,
        "confidentiality_rating": conf_rating,
        "integrity_rating": int_rating,
        "availability_rating": avail_rating
    }


@router.get("", response_model=List[ITAssetResponse])
def list_assets(
    tenant_id: Optional[int] = None,
    asset_type: Optional[str] = None,
    criticality: Optional[str] = None,
    owner_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    # Phase 5 list filters. All optional, default behavior unchanged.
    lifecycle_state: Optional[str] = None,
    data_classification: Optional[str] = None,
    internet_facing: Optional[bool] = None,
    stale_only: bool = Query(False, description="When true, return assets where last_seen_at is older than `stale_days` (default 30)."),
    stale_days: int = Query(30, ge=1, le=365),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []

    query = db.query(ITAsset).filter(ITAsset.tenant_id.in_(user_tenants))

    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    if asset_type:
        query = query.filter(ITAsset.asset_type == asset_type)
    if criticality:
        query = query.filter(ITAsset.criticality == criticality)
    if owner_id:
        # Match the legacy `owner_id` OR the new Phase 5.2 `primary_owner_id`
        # so callers don't have to know which one is populated.
        query = query.filter(
            (ITAsset.owner_id == owner_id) | (ITAsset.primary_owner_id == owner_id)
        )
    if status_filter:
        query = query.filter(ITAsset.status == status_filter)
    if lifecycle_state:
        query = query.filter(ITAsset.lifecycle_state == lifecycle_state)
    if data_classification:
        query = query.filter(ITAsset.data_classification == data_classification)
    if internet_facing is not None:
        query = query.filter(ITAsset.internet_facing == internet_facing)
    if stale_only:
        from datetime import timedelta
        cutoff = datetime.utcnow() - timedelta(days=stale_days)
        # NULL last_seen_at is treated as "never seen", which is also stale —
        # match the UI semantics for the stale filter.
        query = query.filter(
            (ITAsset.last_seen_at == None) | (ITAsset.last_seen_at < cutoff)  # noqa: E711
        )

    assets = query.order_by(ITAsset.created_at.desc()).offset(skip).limit(limit).all()
    return assets


# ── Inventory toolbar: facets + bulk operations (Register view) ──────────────
# These static paths MUST stay ABOVE the `/{asset_id}` route below so "facets",
# "bulk" and "bulk-delete" are never captured as an asset id.

class _BulkUpdatePayload(BaseModel):
    asset_ids: List[int]
    patch: Dict[str, Any]


class _BulkDeletePayload(BaseModel):
    asset_ids: List[int]


# Fields the bulk editor may set — whitelisted so a client can't patch arbitrary
# columns. Aliases map the UI's short names onto the real ITAsset columns.
_BULK_FIELD_ALIASES = {"type": "asset_type", "lifecycle": "lifecycle_state"}
_BULK_ALLOWED_FIELDS = {
    "criticality", "status", "lifecycle_state", "asset_type", "environment",
    "department", "data_classification", "owner_id", "primary_owner_id",
    "internet_facing",
}


@router.get("/facets")
def asset_facets(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Per-value counts for the inventory filter facets (Type / Criticality /
    Status / Lifecycle / Environment …). Returns {facet_key: {value: count}}.
    A column that doesn't exist is simply omitted, so the UI degrades gracefully."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {}

    def _counts(col_name: str) -> Dict[str, int]:
        col = getattr(ITAsset, col_name, None)
        if col is None:
            return {}
        rows = (
            db.query(col, func.count(ITAsset.id))
            .filter(ITAsset.tenant_id.in_(user_tenants))
            .group_by(col)
            .all()
        )
        return {str(v): int(c) for v, c in rows if v is not None and v != ""}

    return {
        "type": _counts("asset_type"),
        "criticality": _counts("criticality"),
        "status": _counts("status"),
        "lifecycle": _counts("lifecycle_state"),
        "environment": _counts("environment"),
        "data_classification": _counts("data_classification"),
    }


@router.patch("/bulk")
def bulk_update_assets(
    payload: _BulkUpdatePayload,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Apply one `patch` to many assets (the register's "Set …" actions). Only
    whitelisted fields are applied, and only to assets in the caller's tenant(s).
    Returns {updated, fields}."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants or not payload.asset_ids:
        return {"updated": 0, "fields": []}

    clean: Dict[str, Any] = {}
    for key, value in (payload.patch or {}).items():
        col = _BULK_FIELD_ALIASES.get(key, key)
        if col in _BULK_ALLOWED_FIELDS and hasattr(ITAsset, col):
            clean[col] = value
    if not clean:
        return {"updated": 0, "fields": []}

    assets = (
        db.query(ITAsset)
        .filter(ITAsset.id.in_(payload.asset_ids), ITAsset.tenant_id.in_(user_tenants))
        .all()
    )
    for asset in assets:
        for col, value in clean.items():
            setattr(asset, col, value)
    db.commit()
    return {"updated": len(assets), "fields": sorted(clean.keys())}


@router.post("/bulk-delete")
def bulk_delete_assets(
    payload: _BulkDeletePayload,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Delete many assets at once (the register's "Delete selected" action).
    Scoped to the caller's tenant(s); relies on the same cascade as the single
    delete. Returns {deleted}."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants or not payload.asset_ids:
        return {"deleted": 0}

    assets = (
        db.query(ITAsset)
        .filter(ITAsset.id.in_(payload.asset_ids), ITAsset.tenant_id.in_(user_tenants))
        .all()
    )
    for asset in assets:
        db.delete(asset)
    db.commit()
    return {"deleted": len(assets)}


@router.post("", response_model=ITAssetResponse, status_code=status.HTTP_201_CREATED)
def create_asset(
    asset: ITAssetCreate,
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tenant not found"
        )
    
    # Validate any manual criticality override up-front so we don't
    # persist a half-valid row. Override requires both a valid bucket
    # AND a reason; the recompute call below will keep the override
    # bucket but still write the derived `criticality_score` for audit.
    from ..services.asset_criticality import is_valid_bucket
    manual_override = bool(asset.criticality_manual_override) and bool(asset.criticality)
    if manual_override:
        if not is_valid_bucket(asset.criticality):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="criticality override must be one of low/medium/high/critical",
            )
        if not (asset.criticality_override_reason or "").strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="criticality_override_reason is required when overriding the derived criticality",
            )

    # Derive a canonical os_normalized key from whatever OS text the
    # operator supplied (os_version / os_family), so a manually-added asset
    # resolves to a CIS benchmark just like scanner/wizard assets — instead
    # of landing with a vague family or NULL key. An explicit os_normalized
    # from the caller always wins.
    _os_family = getattr(asset, "os_family", None)
    _os_version = getattr(asset, "os_version", None)
    _os_normalized = getattr(asset, "os_normalized", None)
    _os_build = getattr(asset, "os_build", None)
    _os_edition = getattr(asset, "os_edition", None)
    if not _os_normalized and (_os_version or _os_family):
        try:
            from ..modules.compliance_plugins.services.os_detector import normalize_os_string
            _fam, _norm, _build, _edition = normalize_os_string(_os_version or _os_family)
            _os_normalized = _norm or _os_family
            _os_family = _os_family or _fam
            _os_build = _os_build or _build
            _os_edition = _os_edition or _edition
        except Exception:  # noqa: BLE001 — never block asset creation
            _os_normalized = _os_normalized or _os_family

    db_asset = ITAsset(
        tenant_id=tenant_id,
        name=asset.name,
        description=asset.description,
        asset_type=asset.asset_type,
        owner_id=asset.owner_id,
        owner_name=asset.owner_name,
        custodian=asset.custodian,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        # `criticality` is set here only when the user is overriding the
        # derived value. Otherwise it's left blank and the recompute
        # below fills it from the CIA + exposure inputs.
        criticality=(asset.criticality if manual_override else None),
        criticality_manual_override=manual_override,
        criticality_override_reason=asset.criticality_override_reason if manual_override else None,
        vendor=asset.vendor,
        location=asset.location,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
        valuation=asset.valuation,
        cde_environment=asset.cde_environment,
        pci_dss=asset.pci_dss,
        # Phase 5.1 — Exposure metadata. None values are stored as NULL.
        internet_facing=bool(asset.internet_facing) if asset.internet_facing is not None else False,
        network_segment=asset.network_segment,
        data_classification=asset.data_classification,
        business_function=asset.business_function,
        compliance_scope=asset.compliance_scope if asset.compliance_scope is not None else [],
        # Phase 5.2 — Ownership chain.
        primary_owner_id=asset.primary_owner_id,
        secondary_owner_id=asset.secondary_owner_id,
        owning_team=asset.owning_team,
        owning_team_id=asset.owning_team_id,
        escalation_contact_id=asset.escalation_contact_id,
        business_owner_id=asset.business_owner_id,
        # Phase 5.3 — Lifecycle state (only the starting state on create;
        # subsequent moves go through /lifecycle-transition so the machine
        # validates them).
        lifecycle_state=(asset.lifecycle_state or "active"),
        # CIS Compliance tab — operator-supplied OS profile so the strict
        # matcher resolves a benchmark immediately without needing a
        # Connect Wizard handshake first.
        os_family=_os_family,
        os_version=_os_version,
        os_normalized=_os_normalized,
        os_build=_os_build,
        os_edition=_os_edition,
    )

    # Auto-resolve owner_name from owner_id if not provided
    if asset.owner_id and not asset.owner_name:
        owner = db.query(GRCUser).filter(GRCUser.id == asset.owner_id).first()
        if owner:
            db_asset.owner_name = owner.display_name or owner.username

    # ISO 27005 derived criticality (bucket + 0-10 score). When the row
    # carries `criticality_manual_override=True`, the bucket is preserved
    # but the score is still re-derived from the inputs so the audit log
    # shows what the system would have chosen.
    recompute_asset_criticality(db_asset)

    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)
    return db_asset


@router.get("/dashboard", response_model=AssetDashboard)
def get_asset_dashboard(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return AssetDashboard(
            total_assets=0,
            by_type={},
            by_criticality={},
            by_status={},
            high_value_assets=0,
            assets_needing_assessment=0
        )
    
    query = db.query(ITAsset).filter(ITAsset.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    
    assets = query.all()
    total = len(assets)
    
    by_type = {}
    by_criticality = {}
    by_status = {}
    high_value_assets = 0
    assets_needing_assessment = 0
    
    for asset in assets:
        by_type[asset.asset_type] = by_type.get(asset.asset_type, 0) + 1
        by_criticality[asset.criticality] = by_criticality.get(asset.criticality, 0) + 1
        by_status[asset.status] = by_status.get(asset.status, 0) + 1
        
        if asset.criticality in ["high", "critical"]:
            high_value_assets += 1
        
        if not asset.risk_assessments:
            assets_needing_assessment += 1
    
    return AssetDashboard(
        total_assets=total,
        by_type=by_type,
        by_criticality=by_criticality,
        by_status=by_status,
        high_value_assets=high_value_assets,
        assets_needing_assessment=assets_needing_assessment
    )


@router.get("/coverage", response_model=AssetCoverage)
def get_asset_coverage(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return AssetCoverage(
            total_assets=0,
            assets_with_controls=0,
            coverage_percentage=0.0,
            by_criticality={}
        )
    
    query = db.query(ITAsset).options(joinedload(ITAsset.control_links)).filter(ITAsset.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(ITAsset.tenant_id == tenant_id)
    
    assets = query.all()
    total = len(assets)
    assets_with_controls = sum(1 for a in assets if a.control_links)
    
    coverage_by_criticality = {}
    for asset in assets:
        if asset.criticality not in coverage_by_criticality:
            coverage_by_criticality[asset.criticality] = {
                "total": 0,
                "with_controls": 0,
                "coverage_percentage": 0.0
            }
        coverage_by_criticality[asset.criticality]["total"] += 1
        if asset.control_links:
            coverage_by_criticality[asset.criticality]["with_controls"] += 1
    
    for crit, data in coverage_by_criticality.items():
        if data["total"] > 0:
            data["coverage_percentage"] = round(
                (data["with_controls"] / data["total"]) * 100, 2
            )
    
    return AssetCoverage(
        total_assets=total,
        assets_with_controls=assets_with_controls,
        coverage_percentage=round((assets_with_controls / total) * 100, 2) if total > 0 else 0.0,
        by_criticality=coverage_by_criticality
    )


@router.get("/inventory-overview")
def get_inventory_overview(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Board-level SCORED inventory overview: five sections (inventory hygiene,
    criticality coverage, vulnerability exposure, scan/monitoring, lifecycle &
    exposure) blended into one inventory score + an attention queue — every metric
    a numerator/denominator over real asset/vuln/criticality records. Registered
    before /{asset_id} so the literal path wins."""
    from ..modules.it_assets.scoring import score_inventory
    tids = get_user_tenants(current_user, db)
    if not tids:
        return {"as_of": None, "counts": {}, "sections": {},
                "performance": {"score": None, "grade": None, "components": []},
                "attention_queue": {}}
    return score_inventory(db, tids)


@router.get("/scorecard-config")
def get_inventory_scorecard_config(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Current section weights + target (built-in defaults merged with tenant overrides)."""
    from ..services import scorecard_config as sc_cfg
    tids = get_user_tenants(current_user, db)
    if not tids:
        return {"module": "assets", "sections": [], "target": 85, "default_target": 85, "customized": False}
    return sc_cfg.merged(db, tids[0], "assets")


@router.put("/scorecard-config")
def put_inventory_scorecard_config(
    body: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Save section-weight, metric-weight and/or target overrides for this tenant.
    Any field omitted is left unchanged; weights renormalize to 100%."""
    from ..services import scorecard_config as sc_cfg
    tids = get_user_tenants(current_user, db)
    if not tids:
        return {"ok": False}
    cfg = sc_cfg.save_config(
        db, tids[0], "assets",
        section_weights=body.get("weights"),
        metric_weights=body.get("metric_weights"),
        target=body.get("target"),
        updated_by=getattr(current_user, "id", None),
    )
    return {"ok": True, "config": cfg}


@router.delete("/scorecard-config")
def reset_inventory_scorecard_config(
    section: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Reset scorecard tuning to defaults — the whole module, or (with ?section=)
    just one section's metric weights."""
    from ..services import scorecard_config as sc_cfg
    tids = get_user_tenants(current_user, db)
    if tids:
        sc_cfg.reset_config(db, tids[0], "assets", section=section)
    return {"ok": True}


# ─── Criticality helper endpoints ─────────────────────────────────
# Two helpers powering the new IT-Asset form:
#   * `/criticality/business-functions` — catalogue of categories the
#     form's "Business function" dropdown should render. Grouped, with
#     each entry's `high_impact` flag exposed so the UI can show a
#     small "boosts criticality" tag next to those options.
#   * `/criticality/preview` — pure compute. Lets the form show the
#     calculated bucket + 0-10 score live as the user fills the
#     inputs, before they save. No database writes.

@router.get("/criticality/business-functions")
def list_business_functions(
    current_user: GRCUser = Depends(require_auth),
):
    from ..services.asset_criticality import list_business_function_categories
    return {"items": list_business_function_categories()}


class CriticalityPreviewRequest(BaseModel):
    confidentiality_rating: Optional[int] = None
    integrity_rating: Optional[int] = None
    availability_rating: Optional[int] = None
    data_classification: Optional[str] = None
    internet_facing: Optional[bool] = None
    business_function: Optional[str] = None


@router.post("/criticality/preview")
def preview_criticality(
    payload: CriticalityPreviewRequest,
    current_user: GRCUser = Depends(require_auth),
):
    from ..services.asset_criticality import compute_criticality_score, score_to_bucket
    score = compute_criticality_score(
        confidentiality_rating=payload.confidentiality_rating,
        integrity_rating=payload.integrity_rating,
        availability_rating=payload.availability_rating,
        data_classification=payload.data_classification,
        internet_facing=payload.internet_facing,
        business_function=payload.business_function,
    )
    return {
        "score": score,
        "bucket": score_to_bucket(score),
    }


ASSET_TEMPLATE_COLUMNS = [
    # ── Identity ────────────────────────────────────────────────────
    ("name",                   "Asset Name (Required)", "ERP System"),
    ("description",            "Description", "Enterprise Resource Planning system for finance and operations"),
    ("asset_type",             "Asset Type (Required: application/infrastructure/data/cloud/third_party)", "application"),
    ("host_name",              "Host Name / FQDN", "erp-prod-01.example.com"),
    ("ip_address",             "IP Address", "10.20.30.40"),
    ("vendor",                 "Vendor", "SAP"),
    ("location",               "Location", "Primary Data Center"),
    # ── OS (drives CIS rule matching) ───────────────────────────────
    ("operating_system",       "Operating System (e.g. 'Microsoft Windows Server 2019', 'Ubuntu 22.04', 'VMware ESXi 7.0') — used to match CIS benchmark rules", "Microsoft Windows Server 2019"),
    # ── CIA ratings (drive criticality) ─────────────────────────────
    ("confidentiality_rating", "Confidentiality Rating (1-5)", "4"),
    ("integrity_rating",       "Integrity Rating (1-5)", "5"),
    ("availability_rating",    "Availability Rating (1-5)", "5"),
    # ── Exposure metadata (drive criticality) ───────────────────────
    ("data_classification",    "Data Classification (public/internal/confidential/restricted)", "confidential"),
    ("internet_facing",        "Internet Facing (true/false)", "true"),
    ("business_function",      "Business Function (category id — see catalogue, e.g. payment_processing / authentication_iam / customer_data / other)", "payment_processing"),
    ("network_segment",        "Network Segment", "dmz"),
    ("compliance_scope",       "Compliance Scope (comma-separated framework short codes, e.g. PCI-DSS,SWIFT)", "PCI-DSS,SWIFT"),
    # ── Ownership ───────────────────────────────────────────────────
    ("owner_name",             "Owner Display Name", "Jane Doe"),
    ("owning_team",            "Owning Team Name", "Payments Engineering"),
    # ── Business + lifecycle ────────────────────────────────────────
    ("valuation",              "Valuation (USD)", "500000"),
    ("status",                 "Status (active/inactive/decommissioned)", "active"),
    ("lifecycle_state",        "Lifecycle State (planned/active/maintenance/decommissioned/retired)", "active"),
    ("cde_environment",        "CDE Environment (true/false)", "false"),
    # ── Optional criticality override ───────────────────────────────
    ("criticality",            "Criticality OVERRIDE (low/medium/high/critical) — LEAVE BLANK to let the system calculate from CIA + exposure inputs above", ""),
    ("criticality_override_reason", "Reason for override (required if Criticality column is filled)", ""),
]


@router.get("/template/download")
def download_asset_template(
    current_user: GRCUser = Depends(require_auth)
):
    """Download CSV template for bulk asset import"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [col[0] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(headers)
    
    descriptions = [col[1] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(descriptions)
    
    examples = [col[2] for col in ASSET_TEMPLATE_COLUMNS]
    writer.writerow(examples)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=it_assets_template.csv"}
    )


@router.post("/import/upload")
async def upload_assets_file(
    file: UploadFile = File(...),
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Upload CSV or Excel file to bulk import IT assets"""
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided"
        )
    
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be CSV or Excel format (.csv, .xlsx, .xls)"
        )
    
    content = await file.read()
    rows = []
    
    try:
        if filename.endswith('.csv'):
            text_content = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text_content))
            rows = list(reader)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            
            headers = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    continue
                if idx == 1:
                    first_cell = str(row[0]).lower() if row[0] else ""
                    if "required" in first_cell or "description" in first_cell or "name" in first_cell:
                        continue
                
                if not any(row):
                    continue
                    
                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = cell
                rows.append(row_dict)
            wb.close()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error parsing file: {str(e)}"
        )
    
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in file"
        )
    
    valid_asset_types = ["application", "infrastructure", "data", "cloud", "third_party"]
    valid_criticality = ["low", "medium", "high", "critical"]
    valid_status = ["active", "inactive", "decommissioned"]
    valid_data_class = ["public", "internal", "confidential", "restricted"]
    valid_lifecycle = ["planned", "active", "maintenance", "decommissioned", "retired"]

    from ..services.asset_criticality import recompute_for_asset, is_valid_bucket
    try:
        from ..modules.compliance_plugins.services.os_detector import normalize_os_string
    except Exception:  # noqa: BLE001
        normalize_os_string = None  # type: ignore

    imported = []
    errors = []

    def _bool(val) -> bool:
        if val is None:
            return False
        return str(val).strip().lower() in ("true", "yes", "1", "y")

    def parse_int(val, min_val=None, max_val=None):
        if val is None or val == "":
            return None
        try:
            v = int(float(str(val)))
            if min_val is not None and v < min_val:
                return min_val
            if max_val is not None and v > max_val:
                return max_val
            return v
        except Exception:
            return None

    def parse_float(val):
        if val is None or val == "":
            return None
        try:
            return float(str(val).replace(",", ""))
        except Exception:
            return None

    def parse_str(val) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        return s or None

    def parse_csv_list(val):
        if val is None:
            return None
        raw = str(val).strip()
        if not raw:
            return None
        return [item.strip() for item in raw.split(",") if item.strip()]

    for idx, row in enumerate(rows, start=1):
        row_num = idx + 2

        name = str(row.get("name", "")).strip() if row.get("name") else ""
        if not name:
            errors.append({"row": row_num, "error": "Name is required"})
            continue

        asset_type = str(row.get("asset_type", "")).strip().lower() if row.get("asset_type") else ""
        if not asset_type:
            errors.append({"row": row_num, "error": "Asset type is required"})
            continue
        if asset_type not in valid_asset_types:
            errors.append({"row": row_num, "error": f"Invalid asset_type '{asset_type}'. Must be one of: {', '.join(valid_asset_types)}"})
            continue

        # Optional manual override of the derived criticality bucket. If the
        # cell is empty, we let `recompute_for_asset` derive it from the CIA
        # ratings + exposure metadata.
        override_raw = parse_str(row.get("criticality"))
        override_reason = parse_str(row.get("criticality_override_reason"))
        manual_override = False
        criticality_value: Optional[str] = None
        if override_raw:
            override_norm = override_raw.lower()
            if override_norm not in valid_criticality:
                errors.append({
                    "row": row_num,
                    "error": (
                        f"Invalid criticality override '{override_raw}'. Must be one "
                        f"of {valid_criticality} or blank to let the system compute it."
                    ),
                })
                continue
            if not override_reason:
                errors.append({
                    "row": row_num,
                    "error": "When 'criticality' is set the 'criticality_override_reason' column is required.",
                })
                continue
            manual_override = True
            criticality_value = override_norm

        asset_status_raw = parse_str(row.get("status"))
        asset_status = asset_status_raw.lower() if asset_status_raw else "active"
        if asset_status not in valid_status:
            asset_status = "active"

        data_classification_raw = parse_str(row.get("data_classification"))
        data_classification = data_classification_raw.lower() if data_classification_raw else None
        if data_classification and data_classification not in valid_data_class:
            errors.append({
                "row": row_num,
                "error": f"Invalid data_classification '{data_classification_raw}'. Must be one of {valid_data_class}.",
            })
            continue

        lifecycle_raw = parse_str(row.get("lifecycle_state"))
        lifecycle_state = lifecycle_raw.lower() if lifecycle_raw else None
        if lifecycle_state and lifecycle_state not in valid_lifecycle:
            errors.append({
                "row": row_num,
                "error": f"Invalid lifecycle_state '{lifecycle_raw}'. Must be one of {valid_lifecycle}.",
            })
            continue

        # Normalise the free-form OS string into a canonical key so imported
        # assets match CIS benchmarks (instead of landing with NULL OS).
        os_raw = parse_str(row.get("operating_system"))
        os_family = os_normalized = os_build = os_edition = None
        if os_raw and normalize_os_string is not None:
            try:
                os_family, os_normalized, os_build, os_edition = normalize_os_string(os_raw)
            except Exception:  # noqa: BLE001
                pass

        try:
            asset = ITAsset(
                tenant_id=tenant_id,
                name=name,
                description=parse_str(row.get("description")),
                asset_type=asset_type,
                host_name=parse_str(row.get("host_name")),
                ip_address=parse_str(row.get("ip_address")),
                vendor=parse_str(row.get("vendor")),
                location=parse_str(row.get("location")),
                os_family=os_family,
                os_version=os_raw,
                os_normalized=os_normalized,
                os_build=os_build,
                os_edition=os_edition,
                owner_name=parse_str(row.get("owner_name")),
                owning_team=parse_str(row.get("owning_team")),
                # CIA — drive the derived criticality.
                confidentiality_rating=parse_int(row.get("confidentiality_rating"), 1, 5),
                integrity_rating=parse_int(row.get("integrity_rating"), 1, 5),
                availability_rating=parse_int(row.get("availability_rating"), 1, 5),
                # Exposure metadata — also drive the derived criticality.
                data_classification=data_classification,
                internet_facing=_bool(row.get("internet_facing")) if parse_str(row.get("internet_facing")) is not None else None,
                business_function=parse_str(row.get("business_function")),
                network_segment=parse_str(row.get("network_segment")),
                compliance_scope=parse_csv_list(row.get("compliance_scope")) or [],
                # Business + lifecycle.
                valuation=parse_float(row.get("valuation")),
                status=asset_status,
                lifecycle_state=lifecycle_state or "active",
                cde_environment=_bool(row.get("cde_environment")),
                # Override (if present, validated above).
                criticality=criticality_value,
                criticality_manual_override=manual_override,
                criticality_override_reason=override_reason if manual_override else None,
            )
            # Always run the derived-criticality calculation. When
            # `criticality_manual_override` is True, the service keeps
            # the user's bucket but still writes the audit score; when
            # False, both the bucket and score are system-computed.
            recompute_for_asset(asset)
            db.add(asset)
            db.flush()
            imported.append({"id": asset.id, "name": asset.name})
        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    db.commit()
    
    imported_count = len(imported)
    error_count = len(errors)
    error_messages = [f"Row {e['row']}: {e['error']}" for e in errors[:20]]
    
    return {
        "success": True,
        "imported": imported_count,
        "total_rows": len(rows),
        "errors": error_messages,
        "total_errors": error_count,
        "message": f"Successfully imported {imported_count} assets" + (f" with {error_count} errors" if error_count > 0 else "")
    }


@router.get("/{asset_id}", response_model=dict)
def get_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.risk_links),
        joinedload(ITAsset.risk_assessments)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    latest_assessment = None
    if asset.risk_assessments:
        latest = sorted(asset.risk_assessments, key=lambda x: x.assessment_date, reverse=True)[0]
        latest_assessment = {
            "id": latest.id,
            "assessment_date": latest.assessment_date.isoformat(),
            "risk_score": latest.risk_score,
            "coverage_percentage": latest.coverage_percentage,
            "gaps": latest.gaps,
            "assessor_id": latest.assessor_id
        }
    
    return {
        "id": asset.id,
        "tenant_id": asset.tenant_id,
        "name": asset.name,
        "description": asset.description,
        "asset_type": asset.asset_type,
        "owner_id": asset.owner_id,
        "custodian": asset.custodian,
        "host_name": asset.host_name,
        "ip_address": asset.ip_address,
        "criticality": asset.criticality,
        "confidentiality_rating": asset.confidentiality_rating,
        "integrity_rating": asset.integrity_rating,
        "availability_rating": asset.availability_rating,
        "valuation": asset.valuation,
        "vendor": asset.vendor,
        "location": asset.location,
        "status": asset.status,
        "cde_environment": asset.cde_environment or False,
        "created_at": asset.created_at.isoformat(),
        # Phase 5 operational context fields.
        "internet_facing": bool(asset.internet_facing) if asset.internet_facing is not None else False,
        # `is_internet_facing` is a legacy wire alias (the Risk Posture page still uses
        # this key). It now mirrors the CANONICAL `internet_facing` column so the page's
        # checkbox reflects — and its Save round-trips through — the same value the asset
        # form, the composite score and the vuln reachability engine all use.
        "is_internet_facing": bool(asset.internet_facing) if asset.internet_facing is not None else False,
        "network_segment": asset.network_segment,
        "data_classification": asset.data_classification,
        "business_function": asset.business_function,
        "compliance_scope": asset.compliance_scope or [],
        "primary_owner_id": asset.primary_owner_id,
        "secondary_owner_id": asset.secondary_owner_id,
        "owning_team": asset.owning_team,
        "escalation_contact_id": asset.escalation_contact_id,
        "business_owner_id": asset.business_owner_id,
        "lifecycle_state": asset.lifecycle_state,
        "decommissioned_at": asset.decommissioned_at.isoformat() if asset.decommissioned_at else None,
        "retirement_reason": asset.retirement_reason,
        "replacement_asset_id": asset.replacement_asset_id,
        "criticality_score": asset.criticality_score,
        "last_seen_at": asset.last_seen_at.isoformat() if asset.last_seen_at else None,
        "last_seen_source": asset.last_seen_source,
        "linked_controls": [link.normalized_control_id for link in asset.control_links],
        "linked_risks": [link.risk_id for link in asset.risk_links],
        "latest_assessment": latest_assessment
    }


@router.put("/{asset_id}", response_model=ITAssetResponse)
def update_asset(
    asset_id: int,
    asset_update: ITAssetUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    update_data = asset_update.model_dump(exclude_unset=True)

    # Validate any criticality override before applying any fields so we
    # don't half-update the row.
    from ..services.asset_criticality import is_valid_bucket
    if update_data.get("criticality_manual_override"):
        override_bucket = update_data.get("criticality") or asset.criticality
        override_reason = (
            update_data.get("criticality_override_reason")
            or asset.criticality_override_reason
            or ""
        )
        if not is_valid_bucket(override_bucket):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="criticality override must be one of low/medium/high/critical",
            )
        if not str(override_reason).strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="criticality_override_reason is required when overriding the derived criticality",
            )

    for field, value in update_data.items():
        # ── Risk Posture v2 column rename translation ───────────────────
        # The wire field `operational_dependency` (v2 spec) maps to the
        # asset column `op_dep_business_impact` (renamed to avoid
        # collision with the Criticality-Assessments Integer column of
        # the same name). Without this mapping, the Save button on the
        # Risk Posture asset detail page would silently no-op.
        if field == "operational_dependency":
            setattr(asset, "op_dep_business_impact", value)
            continue
        # Internet-exposure has a single canonical column, `internet_facing` (the asset
        # form, CSV import and discovery all write it, and the vuln reachability engine
        # reads it). The Risk Posture page's wire field is still `is_internet_facing`;
        # map it to the canonical column so its Save persists where everything else
        # reads, instead of the retired duplicate column that nothing reads anymore.
        if field == "is_internet_facing":
            setattr(asset, "internet_facing", value)
            continue
        setattr(asset, field, value)

    # Auto-derive os_normalized from os_family when the caller set the
    # family but didn't supply a normalized key. Lets the operator pick
    # "windows" from a dropdown without having to know the canonical
    # normalized_key. The family-fallback BenchmarkOsMapping (e.g.
    # pattern='windows' → Win 11 v5.0.1) picks it up from there.
    if (
        "os_family" in update_data
        and update_data.get("os_family")
        and not asset.os_normalized
    ):
        asset.os_normalized = update_data["os_family"]

    # When the caller explicitly turns the override OFF, clear the
    # reason so we don't keep stale audit text around. The recompute
    # below will overwrite `criticality` with the derived bucket.
    if update_data.get("criticality_manual_override") is False:
        asset.criticality_override_reason = None

    # Auto-resolve owner_name when owner_id is updated without owner_name
    if 'owner_id' in update_data and 'owner_name' not in update_data and update_data.get('owner_id'):
        owner = db.query(GRCUser).filter(GRCUser.id == update_data['owner_id']).first()
        if owner:
            asset.owner_name = owner.display_name or owner.username

    # Recompute derived criticality whenever any of its inputs changed
    # OR the override flag was toggled. The recompute respects
    # `criticality_manual_override` so it never clobbers an active override.
    _criticality_inputs = {
        "confidentiality_rating", "integrity_rating", "availability_rating",
        "data_classification", "internet_facing", "business_function",
        "criticality_manual_override", "criticality",
    }
    if _criticality_inputs.intersection(update_data.keys()):
        recompute_asset_criticality(asset)
        # An explicitly-submitted criticality is a HUMAN DECISION and must
        # survive the recompute. The trigger set above includes "criticality"
        # itself, so editing only that field ran the derivation and then
        # overwrote the chosen bucket with whatever CIA/classification implied
        # — picking "high" and saving silently stored the derived value
        # instead. Re-apply the operator's choice and record it as a manual
        # override so the next recompute (and the UI) knows it is deliberate.
        _chosen = update_data.get("criticality")
        if _chosen and not update_data.get("criticality_manual_override"):
            if not is_valid_bucket(_chosen):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="criticality must be one of low/medium/high/critical",
                )
            asset.criticality = str(_chosen).lower().strip()
            asset.criticality_manual_override = True

    db.commit()
    db.refresh(asset)
    return asset


@router.post("/{asset_id}/lifecycle-transition", response_model=dict)
def transition_asset_lifecycle(
    asset_id: int,
    payload: LifecycleTransitionRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Phase 5.3 — Move an asset through the lifecycle state machine.

    States:  planned → active ↔ maintenance → decommissioned → retired.
    Decommissioning or retiring auto-closes the asset's open vulnerabilities
    (best-effort; the transition still succeeds if the close fails).
    """
    user_tenants = get_user_tenants(current_user, db)

    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found",
        )

    # If a replacement is named, ensure it exists in a tenant the user can see.
    if payload.replacement_asset_id is not None:
        replacement = db.query(ITAsset).filter(
            ITAsset.id == payload.replacement_asset_id,
            ITAsset.tenant_id.in_(user_tenants),
        ).first()
        if not replacement:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Replacement asset not found",
            )
        if replacement.id == asset.id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="An asset cannot be its own replacement",
            )

    try:
        summary = asset_lifecycle.transition(
            db,
            asset,
            payload.to_state,
            reason=payload.reason,
            replacement_asset_id=payload.replacement_asset_id,
            actor_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        ) from exc

    db.commit()
    db.refresh(asset)

    # Stamp into the response for the UI; include the new asset state so the
    # frontend can refresh without a follow-up GET.
    summary["asset_id"] = asset.id
    summary["lifecycle_state"] = asset.lifecycle_state
    summary["retirement_reason"] = asset.retirement_reason
    summary["replacement_asset_id"] = asset.replacement_asset_id
    return summary


@router.delete("/{asset_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    # An asset is referenced by ~23 tables. A bare db.delete() raises a
    # ForeignKeyViolation on the first of them (vulnerability links, in
    # practice), the request 500s, and the row silently stays — which is what
    # "the delete button does nothing" looked like from the UI.
    #
    # Two different treatments, chosen per table by what the row MEANS:
    #   DETACH  — the row is evidence in its own right and outlives the asset
    #             (a CIS scan happened; a discovery observation was recorded).
    #             Its asset reference is nulled so history survives.
    #   DELETE  — the row is a pure association that cannot exist without the
    #             asset (asset↔control link, asset↔vulnerability link).
    from sqlalchemy import text as _sql

    detach = [
        ("grc_compliance_plugin_runs", "asset_id"),        # scan evidence
        ("grc_discovery_observations", "resolved_asset_id"),  # sweep evidence
        ("grc_info_system_criticality_items", "linked_asset_id"),
        ("grc_infra_asset_criticality_items", "linked_asset_id"),
    ]
    purge = [
        ("grc_vulnerability_asset_links", "asset_id"),
        ("grc_asset_control_links", "asset_id"),
        ("grc_asset_framework_control_links", "asset_id"),
        ("grc_asset_internal_control_links", "asset_id"),
        ("grc_asset_evidence_links", "asset_id"),
        ("grc_asset_risk_assessments", "asset_id"),
        ("grc_asset_external_identities", "asset_id"),
        ("grc_asset_security_compliance_selections", "asset_id"),
        ("grc_asset_alert_states", "asset_id"),
        ("grc_risk_asset_links", "asset_id"),
        ("grc_document_asset_links", "asset_id"),
        ("grc_incident_asset_links", "asset_id"),
        ("grc_issue_asset_links", "asset_id"),
        ("grc_software_identifiers", "asset_id"),
        ("grc_compliance_agents", "asset_id"),
        ("grc_asset_relationships", "source_asset_id"),
        ("grc_asset_relationships", "target_asset_id"),
    ]
    # Discovery-state cleanup. An asset promoted from discovery leaves behind a
    # DiscoveryObservation (resolution='created') and an agentless
    # IntegrationConnection keyed to its host. If we only null the observation's
    # FK, the row keeps resolution='created' and the connection stays active —
    # so the Connect queue shows the (now-deleted) device as "In inventory" with
    # a Disconnect button that targets a gone asset and does nothing. Revert the
    # observation to 'unclaimed' so the device becomes connectable again, and
    # deactivate the host's agentless connection.
    _host = (asset.host_name or asset.ip_address or "").strip()
    try:
        for table, col in detach:
            db.execute(_sql(f"UPDATE {table} SET {col} = NULL WHERE {col} = :aid"), {"aid": asset_id})
        for table, col in purge:
            db.execute(_sql(f"DELETE FROM {table} WHERE {col} = :aid"), {"aid": asset_id})
        # Observation(s) that resolved to this asset → back to unclaimed.
        db.execute(_sql(
            "UPDATE grc_discovery_observations "
            "SET resolution='unclaimed', resolved_asset_id=NULL, "
            "    resolution_note='asset deleted — found on the network, needs a login' "
            "WHERE resolved_asset_id = :aid OR "
            "      (resolution='created' AND resolved_asset_id IS NULL AND "
            "       (host_name = :h OR ip_address = :ip))"),
            {"aid": asset_id, "h": asset.host_name, "ip": asset.ip_address})
        # Deactivate agentless connections pinned to this host (they can't scan a
        # deleted asset; a fresh connect re-creates one).
        if _host:
            db.execute(_sql(
                "UPDATE grc_integration_connections SET is_active=false, status='disconnected' "
                "WHERE console_url = :h AND integration_type IN "
                "('windows_winrm','linux_ssh','netdev_ssh')"), {"h": _host})
        # Self-references: promoted applications point at this host as parent,
        # and other assets may name it as their replacement. Detach rather than
        # cascade — deleting a laptop must not silently delete the PostgreSQL
        # asset someone promoted from it, along with its own scan history.
        db.execute(_sql("UPDATE grc_it_assets SET parent_asset_id = NULL WHERE parent_asset_id = :aid"),
                   {"aid": asset_id})
        db.execute(_sql("UPDATE grc_it_assets SET replacement_asset_id = NULL WHERE replacement_asset_id = :aid"),
                   {"aid": asset_id})
        db.delete(asset)
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        # Never fail silently again: say which constraint blocked it.
        logger.exception("asset %s delete failed", asset_id)
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Could not delete this asset — it is still referenced: {str(exc)[:300]}",
        )
    return None


@router.post("/{asset_id}/valuation", response_model=ITAssetResponse)
def update_asset_valuation(
    asset_id: int,
    valuation: AssetValuation,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    asset.valuation = valuation.valuation
    if valuation.confidentiality_rating is not None:
        asset.confidentiality_rating = valuation.confidentiality_rating
    if valuation.integrity_rating is not None:
        asset.integrity_rating = valuation.integrity_rating
    if valuation.availability_rating is not None:
        asset.availability_rating = valuation.availability_rating
    
    db.commit()
    db.refresh(asset)
    return asset


@router.post("/{asset_id}/controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_control(
    asset_id: int,
    link: AssetControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    control = db.query(NormalizedControl).filter(
        NormalizedControl.id == link.normalized_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control not found"
        )
    
    existing = db.query(AssetControlLink).filter(
        AssetControlLink.asset_id == asset_id,
        AssetControlLink.normalized_control_id == link.normalized_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetControlLink(
        asset_id=asset_id,
        normalized_control_id=link.normalized_control_id
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Control linked successfully")


@router.post("/{asset_id}/assess", response_model=AssetRiskAssessmentResponse)
def assess_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    num_controls = len(asset.control_links) + len(asset.internal_control_links) + len(asset.framework_control_links)
    coverage = min(num_controls * 10, 100)
    
    base_risk = 5
    if asset.criticality == "critical":
        base_risk = 8
    elif asset.criticality == "high":
        base_risk = 6
    elif asset.criticality == "low":
        base_risk = 3
    
    risk_score = max(1, base_risk - (num_controls * 0.5))
    
    gaps = {
        "missing_controls": max(0, 10 - num_controls),
        "recommendations": []
    }
    if num_controls < 3:
        gaps["recommendations"].append("Add more controls to improve coverage")
    if asset.criticality in ["high", "critical"] and num_controls < 5:
        gaps["recommendations"].append("Critical assets should have at least 5 controls")
    
    assessment = AssetRiskAssessment(
        asset_id=asset_id,
        risk_score=round(risk_score, 2),
        coverage_percentage=float(coverage),
        gaps=gaps,
        assessor_id=current_user.id
    )
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    return assessment


@router.get("/{asset_id}/assessment", response_model=AssetRiskAssessmentResponse)
def get_latest_assessment(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    assessment = db.query(AssetRiskAssessment).filter(
        AssetRiskAssessment.asset_id == asset_id
    ).order_by(AssetRiskAssessment.assessment_date.desc()).first()
    
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No assessment found for this asset"
        )
    
    return assessment


@router.get("/{asset_id}/detail", response_model=AssetDetailResponse)
def get_asset_detail(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.risk_links),
        joinedload(ITAsset.risk_assessments),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.evidence_links),
        joinedload(ITAsset.owner)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    linked_controls = []
    for link in asset.control_links:
        control = db.query(NormalizedControl).filter(NormalizedControl.id == link.normalized_control_id).first()
        if control:
            linked_controls.append({
                "id": link.id,
                "control_id": control.id,
                "code": control.code,
                "name": control.name
            })
    
    linked_internal_controls = []
    for link in asset.internal_control_links:
        ctrl = db.query(InternalControl).filter(InternalControl.id == link.internal_control_id).first()
        if ctrl:
            linked_internal_controls.append({
                "id": link.id,
                "internal_control_id": ctrl.id,
                "code": ctrl.control_id,
                "name": ctrl.name,
                "category": ctrl.category,
                "coverage_status": link.coverage_status
            })

    linked_framework_controls = []
    for link in asset.framework_control_links:
        fc = db.query(FrameworkControl).filter(FrameworkControl.id == link.framework_control_id).first()
        if fc:
            linked_framework_controls.append({
                "id": link.id,
                "framework_control_id": fc.id,
                "code": fc.code,
                "name": fc.name,
                "coverage_status": link.coverage_status,
                "notes": link.notes
            })
    
    linked_risks = []
    for link in asset.risk_links:
        risk = db.query(Risk).filter(
            Risk.id == link.risk_id,
            Risk.tenant_id.in_(user_tenants)
        ).first()
        if risk:
            linked_risks.append({
                "risk_id": risk.id,
                "title": risk.title,
                "status": risk.status,
                "inherent_score": risk.inherent_score,
                "residual_score": risk.residual_score,
            })
    
    linked_evidence = []
    for link in asset.evidence_links:
        ev = db.query(Evidence).filter(Evidence.id == link.evidence_id).first()
        if ev:
            linked_evidence.append({
                "id": link.id,
                "evidence_id": ev.id,
                "name": ev.name,
                "relationship_type": link.relationship_type
            })

    linked_vulnerabilities = []
    vuln_links = db.query(VulnerabilityAssetLink).options(
        joinedload(VulnerabilityAssetLink.vulnerability)
    ).join(Vulnerability).filter(
        VulnerabilityAssetLink.asset_id == asset_id,
        Vulnerability.tenant_id.in_(user_tenants)
    ).all()
    for link in vuln_links:
        vuln = link.vulnerability
        if vuln:
            linked_vulnerabilities.append({
                "link_id": link.id,
                "vulnerability_id": vuln.id,
                "vuln_id": vuln.vuln_id,
                "title": vuln.title,
                "severity": vuln.severity,
                "status": vuln.status,
                "impact_on_asset": link.impact_on_asset,
                "notes": link.notes,
                "created_at": link.created_at,
                # Provenance — Auto badge + source chip on the UI.
                "link_source": getattr(link, "link_source", "manual") or "manual",
                "auto_linked": bool(getattr(link, "auto_linked", False)),
            })
    
    risk_assessments = []
    for assessment in asset.risk_assessments:
        risk_assessments.append({
            "id": assessment.id,
            "assessment_date": assessment.assessment_date.isoformat() if assessment.assessment_date else None,
            "risk_score": assessment.risk_score,
            "coverage_percentage": assessment.coverage_percentage,
            "gaps": assessment.gaps
        })
    
    total_controls = len(linked_controls) + len(linked_internal_controls) + len(linked_framework_controls)
    coverage = min(total_controls * 10, 100) if total_controls > 0 else 0

    # Phase 5.2 — Resolve owner-chain display names with a single batch query
    # so the response carries human-readable labels for the UI.
    owner_ids = {
        i for i in (
            asset.primary_owner_id, asset.secondary_owner_id,
            asset.escalation_contact_id, asset.business_owner_id,
        ) if i is not None
    }
    owner_names: Dict[int, str] = {}
    if owner_ids:
        for u in db.query(GRCUser).filter(GRCUser.id.in_(owner_ids)).all():
            owner_names[u.id] = u.display_name or u.username

    replacement_name: Optional[str] = None
    if asset.replacement_asset_id is not None:
        rep = db.query(ITAsset.name).filter(ITAsset.id == asset.replacement_asset_id).first()
        replacement_name = rep[0] if rep else None

    # Owning team name — prefer FK, fall back to the legacy free-text field
    # so old rows still render something useful.
    owning_team_name: Optional[str] = asset.owning_team
    if asset.owning_team_id is not None:
        try:
            from ..models import Team
            t = db.query(Team.name).filter(Team.id == asset.owning_team_id).first()
            if t and t[0]:
                owning_team_name = t[0]
        except Exception:
            pass

    # Defensive getattr: production DBs that haven't picked up ITAM /
    # business-impact columns yet must still serve detail (nulls), not 500.
    def _g(name, default=None):
        return getattr(asset, name, default)

    return AssetDetailResponse(
        id=asset.id,
        tenant_id=asset.tenant_id,
        name=asset.name,
        description=asset.description,
        asset_type=asset.asset_type,
        owner_id=asset.owner_id,
        owner_name=asset.owner.display_name if asset.owner else None,
        custodian=asset.custodian,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        # Pass NULL through as NULL — "unrated" is a real state the UI must
        # be able to show. Coercing it to "medium" here re-created the same
        # fabricated rating the model default used to produce.
        criticality=asset.criticality,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
        valuation=asset.valuation,
        vendor=asset.vendor,
        location=asset.location,
        status=asset.status or "active",
        created_at=asset.created_at,
        linked_controls=linked_controls,
        linked_internal_controls=linked_internal_controls,
        linked_framework_controls=linked_framework_controls,
        linked_risks=linked_risks,
        linked_evidence=linked_evidence,
        linked_vulnerabilities=linked_vulnerabilities,
        risk_assessments=risk_assessments,
        coverage_percentage=float(coverage),
        # Business-impact inputs (risk-posture panel). Wire name
        # `operational_dependency` maps from column `op_dep_business_impact`.
        is_customer_facing=_g("is_customer_facing"),
        regulated_data_type=_g("regulated_data_type"),
        operational_dependency=_g("op_dep_business_impact"),
        business_impact_notes=_g("business_impact_notes"),
        # Phase 5 fields.
        internet_facing=bool(_g("internet_facing")) if _g("internet_facing") is not None else False,
        network_segment=_g("network_segment"),
        data_classification=_g("data_classification"),
        business_function=_g("business_function"),
        compliance_scope=_g("compliance_scope") or [],
        primary_owner_id=_g("primary_owner_id"),
        primary_owner_name=owner_names.get(asset.primary_owner_id) if _g("primary_owner_id") else None,
        secondary_owner_id=_g("secondary_owner_id"),
        secondary_owner_name=owner_names.get(asset.secondary_owner_id) if _g("secondary_owner_id") else None,
        owning_team=_g("owning_team"),
        owning_team_id=_g("owning_team_id"),
        owning_team_name=owning_team_name,
        escalation_contact_id=_g("escalation_contact_id"),
        escalation_contact_name=owner_names.get(asset.escalation_contact_id) if _g("escalation_contact_id") else None,
        business_owner_id=_g("business_owner_id"),
        business_owner_name=owner_names.get(asset.business_owner_id) if _g("business_owner_id") else None,
        lifecycle_state=_g("lifecycle_state"),
        decommissioned_at=_g("decommissioned_at"),
        retirement_reason=_g("retirement_reason"),
        replacement_asset_id=_g("replacement_asset_id"),
        replacement_asset_name=replacement_name,
        criticality_score=_g("criticality_score"),
        last_seen_at=_g("last_seen_at"),
        last_seen_source=_g("last_seen_source"),
        # ITAM parity — only present after column migration; never 500 if missing.
        os_family=_g("os_family"),
        os_version=_g("os_version"),
        os_normalized=_g("os_normalized"),
        cpu_cores=_g("cpu_cores"),
        memory_gb=_g("memory_gb"),
        storage_gb=_g("storage_gb"),
        agent_version=_g("agent_version"),
        # Machine-derived (read-only) — stored by the collector all along, but
        # never served, so the UI could not display what the scan found.
        os_build=_g("os_build"),
        os_edition=_g("os_edition"),
        fqdn=_g("fqdn"),
        primary_mac=_g("primary_mac"),
        detected_software_json=_g("detected_software_json"),
        security_posture=_g("security_posture"),
        source_system=_g("source_system"),
        discovery_state=_g("discovery_state"),
        first_seen_at=_g("first_seen_at"),
        asset_role=_g("asset_role"),
        app_attributes_json=_g("app_attributes_json"),
        parent_asset_id=_g("parent_asset_id"),
        cloud_resource_id=_g("cloud_resource_id"),
        cde_environment=_g("cde_environment"),
        pci_dss=_g("pci_dss"),
        manufacturer=_g("manufacturer"),
        model=_g("model"),
        serial_number=_g("serial_number"),
        department=_g("department"),
        assigned_user=_g("assigned_user"),
        purchase_cost=_g("purchase_cost"),
        purchase_date=_g("purchase_date"),
        warranty_expiry=_g("warranty_expiry"),
        eol_date=_g("eol_date"),
        environment=_g("environment"),
    )


@router.get("/{asset_id}/trajectory")
def get_asset_trajectory(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Asset → Vulnerability → linked Risk trajectory map data.

    Three-column flow for the asset detail page diagram:

       Asset ──(VulnerabilityAssetLink)──▶ Vulnerabilities
       Vuln  ──(VulnerabilityControlLink × RiskFrameworkControlLink|RiskControlLink)──▶ Risk
       Asset ──(RiskAssetLink, direct)──▶ Risk

    Controls sit *between* a vuln and a risk in the data model (and we
    auto-CWE-map them), but the user-facing graph collapses that bridge into
    a single edge with the bridge-control list attached as edge metadata.
    Two reasons: (a) auto-mapped vulns point at `parsed_framework_control_id`
    rows while risks can only link to `framework_control_id` / `normalized_
    control_id`, so middle-column control nodes would frequently dead-end
    with no outgoing edge; (b) the user's mental model is "this vuln drives
    this risk" — controls are the mechanism, not the destination.
    """
    asset = _get_asset_for_user(asset_id, current_user, db)

    # ── 1. Vulnerabilities affecting this asset ──────────────────────────
    val_rows = db.query(VulnerabilityAssetLink).options(
        joinedload(VulnerabilityAssetLink.vulnerability),
    ).filter(
        VulnerabilityAssetLink.asset_id == asset_id,
    ).all()

    open_vals = [
        l for l in val_rows
        if l.vulnerability is not None and l.vulnerability.status in ("open", "in_progress")
    ]
    vuln_ids = [l.vulnerability_id for l in open_vals]

    vulnerabilities_payload = []
    asset_vuln_edges = []
    kev_count = 0
    for link in open_vals:
        v = link.vulnerability
        if v.kev_flag:
            kev_count += 1
        vulnerabilities_payload.append({
            "id": v.id,
            "vuln_id": v.vuln_id,
            "title": v.title,
            "severity": v.severity,
            "cvss_score": v.cvss_score,
            "composite_priority": v.composite_priority,
            "kev_flag": bool(v.kev_flag),
            "epss_score": v.epss_score,
            "status": v.status,
            "cve_id": v.cve_id,
            "cwe_id": v.cwe_id,
        })
        asset_vuln_edges.append({
            "from": "asset",
            "to": f"vuln:{v.id}",
            "kind": "affects",
            "link_source": link.link_source or "manual",
            "auto_linked": bool(link.auto_linked),
            "impact": link.impact_on_asset,
            "severity": v.severity,
        })

    # ── 2. Resolve the vuln-control-risk bridge ──────────────────────────
    # We don't emit control nodes; we use the controls solely to *join*
    # vulnerabilities to risks. The result is a per-(vuln, risk) summary
    # with the list of bridge controls and their mitigation effectiveness.
    vcl_rows: List[VulnerabilityControlLink] = []
    if vuln_ids:
        vcl_rows = db.query(VulnerabilityControlLink).filter(
            VulnerabilityControlLink.vulnerability_id.in_(vuln_ids)
        ).all()

    # Pre-fetch the control tables we need by id list. Parsed controls are
    # included so we can show the user *why* a vuln auto-mapped, even though
    # parsed controls themselves cannot connect to risks in the data model.
    parsed_ids = sorted({r.parsed_framework_control_id for r in vcl_rows if r.parsed_framework_control_id})
    framework_ids = sorted({r.framework_control_id for r in vcl_rows if r.framework_control_id})
    normalized_ids = sorted({r.normalized_control_id for r in vcl_rows if r.normalized_control_id})
    internal_ids = sorted({r.internal_control_id for r in vcl_rows if r.internal_control_id})

    parsed_map: Dict[int, ParsedFrameworkControl] = {}
    if parsed_ids:
        for c in db.query(ParsedFrameworkControl).filter(ParsedFrameworkControl.id.in_(parsed_ids)).all():
            parsed_map[c.id] = c

    framework_map: Dict[int, FrameworkControl] = {}
    framework_short_codes: Dict[int, str] = {}
    if framework_ids:
        for c in db.query(FrameworkControl).filter(FrameworkControl.id.in_(framework_ids)).all():
            framework_map[c.id] = c
        try:
            fw_pairs = db.query(FrameworkControl.id, Framework.short_code).join(
                Framework, Framework.id == FrameworkControl.framework_id, isouter=True
            ).filter(FrameworkControl.id.in_(framework_ids)).all()
            for cid, sc in fw_pairs:
                if sc:
                    framework_short_codes[cid] = sc
        except Exception:
            pass

    normalized_map: Dict[int, NormalizedControl] = {}
    if normalized_ids:
        for c in db.query(NormalizedControl).filter(NormalizedControl.id.in_(normalized_ids)).all():
            normalized_map[c.id] = c

    internal_map: Dict[int, InternalControl] = {}
    if internal_ids:
        for c in db.query(InternalControl).filter(InternalControl.id.in_(internal_ids)).all():
            internal_map[c.id] = c

    def _classify_source(notes: Optional[str]) -> Tuple[str, Optional[str]]:
        if notes and notes.startswith("auto:cwe:"):
            cwe = notes.split(":", 2)[-1].strip() or None
            return ("auto_cwe", cwe)
        return ("manual", None)

    # Map vuln_id → list of bridge-control descriptors (only those that can
    # *actually* reach a risk: framework_control_id + normalized_control_id).
    # Parsed/internal entries are captured for tooltip context but don't
    # contribute to risk linkage in the current data model.
    BridgeCtrl = Dict[str, Any]  # {target_type, target_id, code, name, framework_short_code, source, auto_cwe}
    vuln_bridges: Dict[int, List[BridgeCtrl]] = {}
    vuln_total_ctrls: Dict[int, int] = {}  # total VCL links per vuln (all target types) for context chips

    for vcl in vcl_rows:
        source, auto_cwe = _classify_source(vcl.notes)
        target_type: Optional[str] = None
        target_id: Optional[int] = None
        code = name = None
        framework_short_code: Optional[str] = None

        if vcl.parsed_framework_control_id and vcl.parsed_framework_control_id in parsed_map:
            target_type, target_id = "parsed", vcl.parsed_framework_control_id
            c = parsed_map[target_id]
            code, name = c.control_id, c.title
        elif vcl.framework_control_id and vcl.framework_control_id in framework_map:
            target_type, target_id = "framework", vcl.framework_control_id
            c = framework_map[target_id]
            code, name = c.code, c.name
            framework_short_code = framework_short_codes.get(target_id)
        elif vcl.normalized_control_id and vcl.normalized_control_id in normalized_map:
            target_type, target_id = "normalized", vcl.normalized_control_id
            c = normalized_map[target_id]
            code = getattr(c, "code", None) or getattr(c, "control_id", None)
            name = getattr(c, "name", None) or getattr(c, "title", None)
        elif vcl.internal_control_id and vcl.internal_control_id in internal_map:
            target_type, target_id = "internal", vcl.internal_control_id
            c = internal_map[target_id]
            code, name = c.control_id, c.name
        else:
            continue

        vuln_total_ctrls[vcl.vulnerability_id] = vuln_total_ctrls.get(vcl.vulnerability_id, 0) + 1
        vuln_bridges.setdefault(vcl.vulnerability_id, []).append({
            "target_type": target_type,
            "target_id": target_id,
            "code": code,
            "name": name,
            "framework_short_code": framework_short_code,
            "source": source,
            "auto_cwe": auto_cwe,
        })

    # ── 3. Bridge controls → risks ───────────────────────────────────────
    # ctrl_risk_map[(target_type, target_id)] → list of (risk_id, effectiveness)
    ctrl_risk_map: Dict[Tuple[str, int], List[Tuple[int, Optional[str]]]] = {}
    risk_objects: Dict[int, Risk] = {}

    if framework_ids:
        rfcl = db.query(RiskFrameworkControlLink).options(
            joinedload(RiskFrameworkControlLink.risk),
        ).filter(RiskFrameworkControlLink.framework_control_id.in_(framework_ids)).all()
        for link in rfcl:
            r = link.risk
            if r is None:
                continue
            risk_objects[r.id] = r
            ctrl_risk_map.setdefault(("framework", link.framework_control_id), []).append(
                (r.id, link.mitigation_effectiveness),
            )

    if normalized_ids:
        rcl = db.query(RiskControlLink).options(
            joinedload(RiskControlLink.risk),
        ).filter(RiskControlLink.normalized_control_id.in_(normalized_ids)).all()
        for link in rcl:
            r = link.risk
            if r is None:
                continue
            risk_objects[r.id] = r
            ctrl_risk_map.setdefault(("normalized", link.normalized_control_id), []).append(
                (r.id, None),
            )

    # ── 4. Compose vuln → risk edges with bridge_controls metadata ───────
    # For each vuln, for each of its bridge controls, for each risk that
    # control links to → record the edge. Deduplicate by (vuln_id, risk_id),
    # accumulating bridge controls so the UI can show "via 2 controls".
    EFFECTIVENESS_RANK = {"full": 3, "partial": 2, "minimal": 1, "none": 0}

    def _weakest_effectiveness(bridges: List[Dict[str, Any]]) -> Optional[str]:
        # Lowest mitigation wins (worst case for residual risk).
        ranked = [b.get("effectiveness") for b in bridges if b.get("effectiveness")]
        if not ranked:
            return None
        return min(ranked, key=lambda x: EFFECTIVENESS_RANK.get(x, 99))

    vuln_risk_edges_map: Dict[Tuple[int, int], Dict[str, Any]] = {}
    risks_from_chain: Dict[int, Risk] = {}

    for vuln_id, bridges in vuln_bridges.items():
        for b in bridges:
            risks_via = ctrl_risk_map.get((b["target_type"], b["target_id"]), [])
            for rid, effectiveness in risks_via:
                key = (vuln_id, rid)
                entry = vuln_risk_edges_map.setdefault(key, {
                    "from": f"vuln:{vuln_id}",
                    "to": f"risk:{rid}",
                    "kind": "via_control",
                    "bridge_controls": [],
                })
                # Attach this bridge to the edge metadata
                entry["bridge_controls"].append({
                    "code": b["code"],
                    "name": b["name"],
                    "framework_short_code": b["framework_short_code"],
                    "target_type": b["target_type"],
                    "source": b["source"],
                    "auto_cwe": b["auto_cwe"],
                    "effectiveness": effectiveness,
                })
                risks_from_chain[rid] = risk_objects[rid]

    # Recompute weakest effectiveness per edge after accumulation
    vuln_risk_edges = list(vuln_risk_edges_map.values())
    for e in vuln_risk_edges:
        e["weakest_effectiveness"] = _weakest_effectiveness(e["bridge_controls"])

    # ── 5. Direct asset → risk linkages ──────────────────────────────────
    direct_links = db.query(RiskAssetLink).options(
        joinedload(RiskAssetLink.risk),
    ).filter(RiskAssetLink.asset_id == asset_id).all()

    direct_risk_ids: set[int] = set()
    asset_risk_edges: List[Dict[str, Any]] = []
    risks_direct: Dict[int, Risk] = {}
    for rl in direct_links:
        r = rl.risk
        if r is None:
            continue
        direct_risk_ids.add(r.id)
        risks_direct[r.id] = r
        asset_risk_edges.append({
            "from": "asset",
            "to": f"risk:{r.id}",
            "kind": "direct",
        })

    # ── 6. Union of risks (direct + via_control). Direct provenance wins
    # for the "source" tag on the node, but the via_control edges remain
    # so the user can see the chain too.
    all_risks: Dict[int, Tuple[Risk, str]] = {}
    for rid, r in risks_from_chain.items():
        all_risks[rid] = (r, "via_control")
    for rid, r in risks_direct.items():
        all_risks[rid] = (r, "direct")  # overwrite — direct wins

    def _risk_tier(score: Optional[float]) -> str:
        if score is None: return "unknown"
        if score >= 15: return "critical"
        if score >= 10: return "high"
        if score >= 5:  return "medium"
        return "low"

    risks_payload = []
    for rid, (r, source_tag) in all_risks.items():
        inherent = getattr(r, "inherent_score", None) or getattr(r, "inherent_risk_score", None)
        residual = getattr(r, "residual_score", None) or getattr(r, "residual_risk_score", None)
        risks_payload.append({
            "id": r.id,
            "title": r.title,
            "status": r.status,
            "inherent_score": inherent,
            "residual_score": residual,
            "tier": _risk_tier(residual if residual is not None else inherent),
            "source": source_tag,
        })
    risks_payload.sort(key=lambda r: (r["residual_score"] or 0), reverse=True)

    edges = asset_vuln_edges + vuln_risk_edges + asset_risk_edges

    stats = {
        "open_vulns": len(vulnerabilities_payload),
        "kev_count": kev_count,
        "vulns_with_risk_path": len({e["from"] for e in vuln_risk_edges}),
        "bridge_controls_total": sum(len(b) for b in vuln_bridges.values()),
        "risks_direct": sum(1 for r in risks_payload if r["source"] == "direct"),
        "risks_transitive": sum(1 for r in risks_payload if r["source"] == "via_control"),
        "max_residual": max(
            (r["residual_score"] for r in risks_payload if r["residual_score"] is not None),
            default=0,
        ),
    }

    return {
        "asset": {
            "id": asset.id,
            "name": asset.name,
            "type": asset.asset_type,
            "criticality": asset.criticality,
            "criticality_score": asset.criticality_score,
            "internet_facing": bool(getattr(asset, "internet_facing", False)),
            "confidentiality_rating": asset.confidentiality_rating,
            "integrity_rating": asset.integrity_rating,
            "availability_rating": asset.availability_rating,
        },
        "vulnerabilities": vulnerabilities_payload,
        "risks": risks_payload,
        "edges": edges,
        "stats": stats,
    }


@router.get("/{asset_id}/security-compliance/controls")
def list_security_compliance_controls(
    asset_id: int,
    search: Optional[str] = Query(None),
    sort_by: str = Query("control_id"),
    sort_order: str = Query("asc"),
    level: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    selected_only: bool = Query(False),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    asset = _get_asset_for_user(asset_id, current_user, db)
    if sort_by not in {"control_id", "title", "level", "section"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid sort_by")
    if sort_order not in {"asc", "desc"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid sort_order")

    try:
        payload = _load_security_compliance_controls()
    except FileNotFoundError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e),
        ) from e

    controls = payload.get("controls") or []
    if not isinstance(controls, list):
        controls = []

    selected_control_ids = {
        row.control_id
        for row in db.query(AssetSecurityComplianceSelection).filter(
            AssetSecurityComplianceSelection.asset_id == asset.id,
            AssetSecurityComplianceSelection.benchmark == SECURITY_COMPLIANCE_BENCHMARK,
        ).all()
    }

    search_term = (search or "").strip().lower()
    level_term = (level or "").strip().lower()
    section_term = (section or "").strip().lower()

    filtered_controls: List[Dict[str, Any]] = []
    for entry in controls:
        if not isinstance(entry, dict):
            continue

        control_id = str(entry.get("ControlID") or "").strip()
        if not control_id:
            continue

        title = str(entry.get("Title") or "")
        section_value = str(entry.get("Section") or "")
        level_value = str(entry.get("Level") or "")
        description = str(entry.get("Description") or "")

        if search_term:
            haystack = " ".join([control_id, title, section_value, level_value, description]).lower()
            if search_term not in haystack:
                continue

        if level_term and level_term != level_value.lower():
            continue
        if section_term and section_term not in section_value.lower():
            continue

        is_selected = control_id in selected_control_ids
        if selected_only and not is_selected:
            continue

        enriched = dict(entry)
        enriched["control_id"] = control_id
        enriched["selected"] = is_selected
        filtered_controls.append(enriched)

    reverse = sort_order == "desc"
    if sort_by == "control_id":
        filtered_controls.sort(
            key=lambda item: _security_compliance_sort_tokens(item.get("ControlID", "")),
            reverse=reverse,
        )
    elif sort_by == "title":
        filtered_controls.sort(key=lambda item: str(item.get("Title") or "").lower(), reverse=reverse)
    elif sort_by == "level":
        filtered_controls.sort(key=lambda item: str(item.get("Level") or "").lower(), reverse=reverse)
    else:
        filtered_controls.sort(key=lambda item: str(item.get("Section") or "").lower(), reverse=reverse)

    total_filtered = len(filtered_controls)
    paged_items = filtered_controls[skip: skip + limit]
    return {
        "benchmark": payload.get("benchmark") or SECURITY_COMPLIANCE_BENCHMARK,
        "version": payload.get("version"),
        "published": payload.get("published"),
        "total_controls_in_source": payload.get("total_controls") or len(controls),
        "total": total_filtered,
        "skip": skip,
        "limit": limit,
        "selected_count": len(selected_control_ids),
        "controls": paged_items,
    }


@router.get("/{asset_id}/security-compliance/selections")
def get_security_compliance_selections(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _get_asset_for_user(asset_id, current_user, db)
    selections = db.query(AssetSecurityComplianceSelection).filter(
        AssetSecurityComplianceSelection.asset_id == asset_id,
        AssetSecurityComplianceSelection.benchmark == SECURITY_COMPLIANCE_BENCHMARK,
    ).order_by(AssetSecurityComplianceSelection.control_id.asc()).all()

    control_ids = [row.control_id for row in selections]
    return {
        "benchmark": SECURITY_COMPLIANCE_BENCHMARK,
        "count": len(control_ids),
        "control_ids": control_ids,
    }


@router.post(
    "/{asset_id}/security-compliance/selections",
    response_model=MessageResponse,
    status_code=status.HTTP_201_CREATED,
)
def add_security_compliance_selections(
    asset_id: int,
    payload: SecurityComplianceSelectionRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _get_asset_for_user(asset_id, current_user, db)

    requested_ids = sorted({str(control_id).strip() for control_id in payload.control_ids if str(control_id).strip()})
    if not requested_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No control IDs provided")

    try:
        source_payload = _load_security_compliance_controls()
    except FileNotFoundError as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e),
        ) from e

    valid_ids = {
        str(item.get("ControlID") or "").strip()
        for item in (source_payload.get("controls") or [])
        if isinstance(item, dict) and str(item.get("ControlID") or "").strip()
    }
    invalid_ids = [control_id for control_id in requested_ids if control_id not in valid_ids]
    if invalid_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unknown control IDs: {', '.join(invalid_ids[:10])}",
        )

    existing_ids = {
        row.control_id
        for row in db.query(AssetSecurityComplianceSelection).filter(
            AssetSecurityComplianceSelection.asset_id == asset_id,
            AssetSecurityComplianceSelection.benchmark == SECURITY_COMPLIANCE_BENCHMARK,
            AssetSecurityComplianceSelection.control_id.in_(requested_ids),
        ).all()
    }

    created_count = 0
    for control_id in requested_ids:
        if control_id in existing_ids:
            continue
        db.add(
            AssetSecurityComplianceSelection(
                asset_id=asset_id,
                benchmark=SECURITY_COMPLIANCE_BENCHMARK,
                control_id=control_id,
                selected_by=current_user.id,
            )
        )
        created_count += 1

    db.commit()
    return MessageResponse(message=f"{created_count} control(s) selected", id=created_count)


@router.delete(
    "/{asset_id}/security-compliance/selections/{control_id}",
    response_model=MessageResponse,
)
def remove_security_compliance_selection(
    asset_id: int,
    control_id: str,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _get_asset_for_user(asset_id, current_user, db)

    db_row = db.query(AssetSecurityComplianceSelection).filter(
        AssetSecurityComplianceSelection.asset_id == asset_id,
        AssetSecurityComplianceSelection.benchmark == SECURITY_COMPLIANCE_BENCHMARK,
        AssetSecurityComplianceSelection.control_id == control_id,
    ).first()
    if not db_row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Selected control not found for this asset",
        )

    db.delete(db_row)
    db.commit()
    return MessageResponse(message="Control unselected")


@router.post("/{asset_id}/link-framework-control", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_framework_control(
    asset_id: int,
    link: AssetFrameworkControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    control = db.query(FrameworkControl).filter(
        FrameworkControl.id == link.framework_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control not found"
        )
    
    existing = db.query(AssetFrameworkControlLink).filter(
        AssetFrameworkControlLink.asset_id == asset_id,
        AssetFrameworkControlLink.framework_control_id == link.framework_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetFrameworkControlLink(
        asset_id=asset_id,
        framework_control_id=link.framework_control_id,
        coverage_status=link.coverage_status,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Framework control linked successfully", id=db_link.id)


@router.delete("/{asset_id}/link-framework-control/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_framework_control_link(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    link = db.query(AssetFrameworkControlLink).filter(
        AssetFrameworkControlLink.id == link_id,
        AssetFrameworkControlLink.asset_id == asset_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


# ── Mapping recommendations ───────────────────────────────────────────────────
# Regex-driven recommender that scores every framework control against this
# asset's profile (OS family, asset type, network exposure, data class,
# criticality, business function, vendor). See
# services/asset_control_recommender.py for the signal library.

class _AcceptRecommendationsRequest(BaseModel):
    framework_control_ids: List[int]
    coverage_status: str = "partial"
    notes: Optional[str] = None


@router.get("/{asset_id}/mapping-recommendations")
def get_mapping_recommendations(
    asset_id: int,
    framework_id: Optional[int] = Query(None, description="Restrict to a single framework"),
    min_score: int = Query(1, ge=1, le=20, description="Minimum score to surface (default 1 — surfaces universal-only matches for sparse-profile assets)"),
    limit: int = Query(100, ge=1, le=500),
    include_linked: bool = Query(False, description="Include controls already linked"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    result = recommend_for_asset(
        db,
        asset,
        min_score=min_score,
        limit=limit,
        framework_id=framework_id,
        include_linked=include_linked,
    )

    return {
        "recommendations": [
            {
                "framework_control_id": r.framework_control_id,
                "framework_id": r.framework_id,
                "framework_name": r.framework_name,
                "framework_short_code": r.framework_short_code,
                "code": r.code,
                "name": r.name,
                "statement": r.statement,
                "score": r.score,
                "confidence": r.confidence,
                "matched_signals": [
                    {"key": s.key, "label": s.label, "weight": s.weight}
                    for s in r.matched_signals
                ],
                "negative_notes": r.negative_notes,
            }
            for r in result.recommendations
        ],
        "total_controls_scanned": result.total_controls_scanned,
        "total_already_linked": result.total_already_linked,
        "asset_profile": result.asset_profile,
    }


@router.post("/{asset_id}/mapping-recommendations/accept", status_code=status.HTTP_200_OK)
def accept_mapping_recommendations(
    asset_id: int,
    payload: _AcceptRecommendationsRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    coverage = (payload.coverage_status or "partial").lower()
    if coverage not in {"partial", "full", "minimal"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="coverage_status must be one of: partial, full, minimal",
        )

    ids = list({int(fc_id) for fc_id in payload.framework_control_ids if int(fc_id) > 0})
    if not ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="framework_control_ids is required",
        )

    existing_ids = {
        row.framework_control_id
        for row in db.query(AssetFrameworkControlLink.framework_control_id).filter(
            AssetFrameworkControlLink.asset_id == asset_id,
            AssetFrameworkControlLink.framework_control_id.in_(ids),
        )
    }
    valid_ids = {
        row.id
        for row in db.query(FrameworkControl.id).filter(FrameworkControl.id.in_(ids))
    }

    linked: List[int] = []
    skipped_existing = 0
    skipped_missing = 0
    for fc_id in ids:
        if fc_id not in valid_ids:
            skipped_missing += 1
            continue
        if fc_id in existing_ids:
            skipped_existing += 1
            continue
        db_link = AssetFrameworkControlLink(
            asset_id=asset_id,
            framework_control_id=fc_id,
            coverage_status=coverage,
            notes=payload.notes,
        )
        db.add(db_link)
        db.flush()
        linked.append(db_link.id)

    db.commit()
    return {
        "linked": len(linked),
        "link_ids": linked,
        "skipped_existing": skipped_existing,
        "skipped_missing": skipped_missing,
    }


@router.post("/{asset_id}/link-evidence", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_evidence(
    asset_id: int,
    link: AssetEvidenceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    evidence = db.query(Evidence).filter(
        Evidence.id == link.evidence_id,
        Evidence.tenant_id.in_(user_tenants)
    ).first()
    if not evidence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence not found"
        )
    
    existing = db.query(AssetEvidenceLink).filter(
        AssetEvidenceLink.asset_id == asset_id,
        AssetEvidenceLink.evidence_id == link.evidence_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = AssetEvidenceLink(
        asset_id=asset_id,
        evidence_id=link.evidence_id,
        relationship_type=link.relationship_type
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Evidence linked successfully", id=db_link.id)


@router.delete("/{asset_id}/link-evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_evidence_link(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    link = db.query(AssetEvidenceLink).filter(
        AssetEvidenceLink.id == link_id,
        AssetEvidenceLink.asset_id == asset_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.get("/{asset_id}/coverage-analysis", response_model=AssetCoverageAnalysis)
def get_asset_coverage_analysis(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.risk_assessments)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    total_controls = len(asset.control_links) + len(asset.internal_control_links) + len(asset.framework_control_links)
    
    full_framework_coverage = sum(1 for link in asset.framework_control_links if link.coverage_status == "full")
    partial_framework_coverage = sum(1 for link in asset.framework_control_links if link.coverage_status == "partial")
    full_internal_coverage = sum(1 for link in asset.internal_control_links if link.coverage_status == "full")
    partial_internal_coverage = sum(1 for link in asset.internal_control_links if link.coverage_status == "partial")
    covered_controls = (
        len(asset.control_links)
        + full_internal_coverage
        + (partial_internal_coverage * 0.5)
        + full_framework_coverage
        + (partial_framework_coverage * 0.5)
    )
    
    expected_controls = 10
    coverage_percentage = min((covered_controls / expected_controls) * 100, 100) if expected_controls > 0 else 0
    
    gaps = []
    if total_controls < 3:
        gaps.append({"type": "insufficient_controls", "message": "Asset has fewer than 3 controls linked"})
    if asset.criticality in ["high", "critical"] and total_controls < 5:
        gaps.append({"type": "critical_asset_gap", "message": "Critical/high priority asset should have at least 5 controls"})
    if not asset.framework_control_links:
        gaps.append({"type": "no_framework_controls", "message": "No framework controls linked to this asset"})
    
    latest_assessment = None
    risk_score = None
    if asset.risk_assessments:
        latest_assessment = sorted(asset.risk_assessments, key=lambda x: x.assessment_date, reverse=True)[0]
        risk_score = latest_assessment.risk_score
    
    return AssetCoverageAnalysis(
        asset_id=asset.id,
        asset_name=asset.name,
        total_controls=total_controls,
        covered_controls=int(covered_controls),
        coverage_percentage=round(coverage_percentage, 2),
        gaps=gaps,
        risk_score=risk_score
    )


@router.post("/{asset_id}/assess-risk", response_model=AssetRiskAssessmentResponse)
def perform_risk_assessment(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    asset = db.query(ITAsset).options(
        joinedload(ITAsset.control_links),
        joinedload(ITAsset.internal_control_links),
        joinedload(ITAsset.framework_control_links),
        joinedload(ITAsset.evidence_links)
    ).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    num_normalized_controls = len(asset.control_links)
    num_internal_controls = len(asset.internal_control_links)
    num_framework_controls = len(asset.framework_control_links)
    num_evidence = len(asset.evidence_links)
    
    total_controls = num_normalized_controls + num_internal_controls + num_framework_controls
    coverage = min(total_controls * 10, 100)
    
    base_risk = 5
    if asset.criticality == "critical":
        base_risk = 9
    elif asset.criticality == "high":
        base_risk = 7
    elif asset.criticality == "low":
        base_risk = 3
    
    control_reduction = total_controls * 0.4
    evidence_reduction = num_evidence * 0.2
    risk_score = max(1, base_risk - control_reduction - evidence_reduction)
    
    gaps = {
        "missing_controls": max(0, 10 - total_controls),
        "missing_evidence": max(0, 5 - num_evidence),
        "recommendations": []
    }
    if total_controls < 3:
        gaps["recommendations"].append("Add more controls to improve coverage")
    if asset.criticality in ["high", "critical"] and total_controls < 5:
        gaps["recommendations"].append("Critical assets should have at least 5 controls")
    if num_evidence < 2:
        gaps["recommendations"].append("Add more evidence documentation")
    if not asset.framework_control_links:
        gaps["recommendations"].append("Link to framework controls for better compliance tracking")
    
    assessment = AssetRiskAssessment(
        asset_id=asset_id,
        risk_score=round(risk_score, 2),
        coverage_percentage=float(coverage),
        gaps=gaps,
        assessor_id=current_user.id
    )
    db.add(assessment)
    db.commit()
    db.refresh(assessment)
    
    return assessment


class AssetInternalControlLinkCreate(BaseModel):
    internal_control_id: int
    coverage_status: Optional[str] = "partial"


@router.post("/{asset_id}/internal-controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_asset_to_internal_control(
    asset_id: int,
    link: AssetInternalControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Link an ERM internal control to an asset"""
    user_tenants = get_user_tenants(current_user, db)

    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    ctrl = db.query(InternalControl).filter(
        InternalControl.id == link.internal_control_id,
        InternalControl.tenant_id.in_(user_tenants)
    ).first()
    if not ctrl:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Internal control not found")

    existing = db.query(AssetInternalControlLink).filter(
        AssetInternalControlLink.asset_id == asset_id,
        AssetInternalControlLink.internal_control_id == link.internal_control_id
    ).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Link already exists")

    db_link = AssetInternalControlLink(
        asset_id=asset_id,
        internal_control_id=link.internal_control_id,
        coverage_status=link.coverage_status
    )
    db.add(db_link)
    db.commit()
    db.refresh(db_link)

    return MessageResponse(message="Internal control linked successfully", id=db_link.id)


@router.delete("/{asset_id}/internal-controls/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_asset_from_internal_control(
    asset_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Unlink an ERM internal control from an asset"""
    user_tenants = get_user_tenants(current_user, db)

    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")

    db_link = db.query(AssetInternalControlLink).filter(
        AssetInternalControlLink.id == link_id,
        AssetInternalControlLink.asset_id == asset_id
    ).first()
    if not db_link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Link not found")

    db.delete(db_link)
    db.commit()
    return None


# ── Phase 4: Software inventory (CPE / PURL) per asset ──────────────────────
# Feeds the CPE matcher (services/cpe_matcher.py). When an enrichment run
# pulls a CVE's affected_configurations, the matcher walks every
# SoftwareIdentifier row in the tenant and auto-creates the linked vuln rows.

class SoftwareIdentifierCreate(BaseModel):
    identifier_type: str  # "cpe" or "purl"
    identifier: str
    vendor: Optional[str] = None
    product: Optional[str] = None
    version: Optional[str] = None
    source: Optional[str] = "manual"


@router.get("/{asset_id}/software-identifiers")
def list_software_identifiers(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    asset = _get_asset_for_user(asset_id, current_user, db)
    rows = (
        db.query(SoftwareIdentifier)
        .filter(SoftwareIdentifier.asset_id == asset.id)
        .order_by(SoftwareIdentifier.identifier_type.asc(), SoftwareIdentifier.identifier.asc())
        .all()
    )
    return [
        {
            "id": r.id,
            "identifier_type": r.identifier_type,
            "identifier": r.identifier,
            "vendor": r.vendor,
            "product": r.product,
            "version": r.version,
            "source": r.source,
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/{asset_id}/software-identifiers", status_code=status.HTTP_201_CREATED)
def add_software_identifier(
    asset_id: int,
    payload: SoftwareIdentifierCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    asset = _get_asset_for_user(asset_id, current_user, db)
    ident_type = (payload.identifier_type or "").strip().lower()
    if ident_type not in ("cpe", "purl"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="identifier_type must be 'cpe' or 'purl'.",
        )
    ident = (payload.identifier or "").strip()
    if not ident:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="identifier is required.",
        )
    # Parse to extract vendor/product/version when the caller didn't supply
    # them — saves the matcher work later and keeps the row searchable.
    from ..services.cpe_matcher import parse_cpe, parse_purl
    parsed = parse_cpe(ident) if ident_type == "cpe" else parse_purl(ident)
    vendor = payload.vendor or (parsed.vendor if parsed else None)
    product = payload.product or (parsed.product if parsed else None)
    version = payload.version or (parsed.version if parsed and parsed.version != "*" else None)

    row = SoftwareIdentifier(
        tenant_id=asset.tenant_id,
        asset_id=asset.id,
        identifier_type=ident_type,
        identifier=ident,
        vendor=(vendor or None),
        product=(product or None),
        version=(version or None),
        source=(payload.source or "manual"),
    )
    try:
        db.add(row)
        db.commit()
        db.refresh(row)
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not create identifier: {exc.__class__.__name__} (possibly duplicate).",
        )
    return {"id": row.id, "identifier": row.identifier}


@router.delete("/{asset_id}/software-identifiers/{identifier_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_software_identifier(
    asset_id: int,
    identifier_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    asset = _get_asset_for_user(asset_id, current_user, db)
    row = (
        db.query(SoftwareIdentifier)
        .filter(SoftwareIdentifier.id == identifier_id)
        .filter(SoftwareIdentifier.asset_id == asset.id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Identifier not found.")
    db.delete(row)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Detected-software inventory + promote to child assets
# ─────────────────────────────────────────────────────────────────────────────

class PromoteSoftwareIn(BaseModel):
    """Body for promote-software. Caller picks software_keys from the
    detected inventory; criticality defaults to the parent's value."""
    software_keys: List[str]
    criticality: Optional[str] = None


@router.get("/{asset_id}/detected-software")
def get_detected_software(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """The agent's enriched software inventory for this host: each entry
    carries software_key, benchmark availability, and (if already promoted)
    the child asset id. Drives the "Detected on this server" panel."""
    asset = _tenant_asset_or_404(db, current_user, asset_id)
    inventory = list(asset.detected_software_json or [])
    # Enrich each entry with live benchmark + executable rule count (same
    # exclusions as the scanner — never advertise hollow expect:any rules).
    try:
        from sqlalchemy import cast as _cast, String as _String
        from ..models import CompliancePlugin
        from ..modules.compliance_plugins.services.software_normaliser import (
            benchmark_for_software_key,
        )
        tid = asset.tenant_id
        enriched = []
        for e in inventory:
            row = dict(e) if isinstance(e, dict) else {"name": str(e)}
            key = row.get("software_key")
            bname = row.get("benchmark_name")
            if not bname and key:
                bname = benchmark_for_software_key(db, key)
            count = 0
            if bname:
                count = db.query(CompliancePlugin).filter(
                    CompliancePlugin.benchmark == bname,
                    CompliancePlugin.enabled.is_(True),
                    (CompliancePlugin.tenant_id == tid) | (CompliancePlugin.tenant_id.is_(None)),
                    CompliancePlugin.runner_type != "manual",
                    ~_cast(CompliancePlugin.check_definition, _String).ilike("%TODO%"),
                    ~_cast(CompliancePlugin.check_definition, _String).ilike('%"kind": "any"%'),
                    ~_cast(CompliancePlugin.check_definition, _String).ilike('%"kind":"any"%'),
                ).count()
            # "set up" only when the matched benchmark has runnable rules —
            # hollow / zero-rule hits fall through to track-anyway.
            if bname and count > 0:
                row["benchmark_name"] = bname
                row["benchmark_available"] = True
            else:
                row["benchmark_name"] = None
                row["benchmark_available"] = False
            row["rule_count"] = count
            enriched.append(row)
        inventory = enriched
    except Exception:  # noqa: BLE001
        logger.exception("detected-software enrichment failed (non-fatal)")
    promotable = [e for e in inventory if e.get("benchmark_available") and not e.get("promoted_asset_id")]
    # Security posture: prefer the stored value (computed at collection time),
    # but recompute on the fly for assets last inventoried before the posture
    # layer existed so the card is never blank when software IS present.
    posture = getattr(asset, "security_posture", None)
    if posture is None and inventory:
        try:
            from ..modules.compliance_plugins.services.security_classifier import summarize_posture
            posture = summarize_posture(inventory)
        except Exception:
            posture = None
    return {
        "asset_id": asset.id,
        "inventory": inventory,
        "security_posture": posture,
        "counts": {
            "total": len(inventory),
            "promotable": len(promotable),
            "promoted": sum(1 for e in inventory if e.get("promoted_asset_id")),
            "no_benchmark": sum(1 for e in inventory if not e.get("benchmark_available")),
        },
    }


@router.post("/{asset_id}/promote-software")
def promote_software(
    asset_id: int,
    body: PromoteSoftwareIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Turn selected detected-software entries into child assets.

    Each promoted child gets parent_asset_id = this host, asset_role =
    'application', os_normalized = the software_key (the matcher resolves
    the application benchmark exactly like an OS), and inherits the
    parent's owner + criticality + CIA ratings.

    Idempotent: an entry already promoted is skipped, not duplicated.

    Benchmarked software can NO LONGER be created here — that path must go
    through POST /setup-software (validated credential or explicit track).
    Calling promote on a CIS-ready key returns it in ``skipped`` with reason
    ``requires_setup_scan`` so old UI buttons fail closed instead of creating
    unscannable shell assets.
    """
    import copy as _copy

    asset = _tenant_asset_or_404(db, current_user, asset_id)
    # Deep copy so SQLAlchemy sees the JSON column as mutated. A shallow
    # list() shares the inner dicts with the ORM attribute, and after an
    # in-place mutation the comparison returns equal — the UPDATE is then
    # silently skipped and the promotions don't persist.
    inventory = _copy.deepcopy(asset.detected_software_json or [])
    by_key = {e.get("software_key"): e for e in inventory}
    created, skipped = [], []
    for key in body.software_keys:
        entry = by_key.get(key)
        if entry is None:
            skipped.append({"software_key": key, "reason": "not in inventory"})
            continue
        if entry.get("promoted_asset_id"):
            skipped.append({
                "software_key": key,
                "reason": "already promoted",
                "asset_id": entry["promoted_asset_id"],
            })
            continue
        if entry.get("benchmark_available"):
            skipped.append({
                "software_key": key,
                "reason": "requires_setup_scan",
                "hint": "Use POST /assets/{id}/setup-software with mode=scan (or host_connection). "
                        "Benchmarked software cannot be promoted without a validated credential.",
            })
            continue
        child = _create_child_from_software_entry(
            db, asset, entry, criticality=body.criticality,
        )
        entry["promoted_asset_id"] = child.id
        created.append({"software_key": key, "asset_id": child.id, "name": child.name})
    asset.detected_software_json = inventory
    db.commit()
    return {"created": created, "skipped": skipped}


class SetupSoftwareIn(BaseModel):
    """Bring detected software into inventory — only after reachability is proven
    for scannable products. Modes:
      scan            — SQL product: validate DB creds, create child + connection
      host_connection — non-SQL with benchmark: require parent host connection,
                        create child (room-scan model)
      track           — no benchmark: inventory only, no scan connection
    """
    mode: str  # scan | host_connection | track
    software_key: str
    hostname: Optional[str] = None
    display_label: Optional[str] = None
    port: Optional[int] = None
    database: Optional[str] = None
    oracle_sid: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None


def _software_platform(software_key: str) -> Optional[str]:
    k = (software_key or "").lower()
    if k.startswith("postgres") or k.startswith("postgresql"):
        return "postgres"
    if k.startswith("mssql") or k.startswith("sql-server"):
        return "mssql"
    if k.startswith("mysql") or k.startswith("mariadb"):
        return "mysql"
    if k.startswith("oracle"):
        return "oracle"
    return None


def _create_child_from_software_entry(
    db: Session,
    asset: ITAsset,
    entry: dict,
    *,
    criticality: Optional[str] = None,
) -> ITAsset:
    """Shared child-asset constructor used by promote + setup-software."""
    key = entry.get("software_key") or "application"
    child = ITAsset(
        tenant_id=asset.tenant_id,
        name=f"{entry.get('name') or key} @ {asset.host_name or asset.name}",
        description=(
            f"Application asset from detected software on "
            f"'{asset.name}' (source: {entry.get('source')})"
        ),
        asset_type="application",
        asset_role="application",
        parent_asset_id=asset.id,
        host_name=asset.host_name,
        ip_address=asset.ip_address,
        os_family=asset.os_family,
        os_version=(
            f"{entry.get('name')} {entry.get('version')}"
            if entry.get("version") else entry.get("name")
        ),
        os_normalized=key,
        criticality=criticality or asset.criticality,
        status=asset.status,
        owner_id=asset.owner_id,
        owner_name=asset.owner_name,
        confidentiality_rating=asset.confidentiality_rating,
        integrity_rating=asset.integrity_rating,
        availability_rating=asset.availability_rating,
        fqdn=asset.fqdn,
        primary_mac=asset.primary_mac,
        network_segment=asset.network_segment,
        internet_facing=asset.internet_facing,
        is_internet_facing=getattr(asset, "is_internet_facing", None) or asset.internet_facing,
        data_classification=asset.data_classification,
        business_function=asset.business_function,
        last_seen_at=asset.last_seen_at,
        last_seen_source=asset.last_seen_source,
        first_seen_at=asset.first_seen_at,
        source_system=asset.source_system,
    )
    if isinstance(entry.get("attributes"), dict) and entry["attributes"]:
        child.app_attributes_json = entry["attributes"]
    else:
        try:
            from ..modules.compliance_plugins.services.software_profiler import (
                profile_software, probes_for,
            )
            from ..models import IntegrationConnection as _IC
            fam = (asset.os_family or "").lower()
            transport = "windows" if fam.startswith("windows") else "linux"
            if probes_for(key, transport):
                from ..modules.compliance_plugins.services.credentials import (
                    resolve_credentials_for_connection,
                )
                conn = (
                    db.query(_IC)
                    .filter(
                        _IC.tenant_id == asset.tenant_id,
                        _IC.console_url == (asset.host_name or asset.ip_address),
                        _IC.is_active.is_(True),
                    )
                    .first()
                )
                if conn is not None:
                    creds = resolve_credentials_for_connection(conn)

                    def _run(shell: str, command: str):
                        from ..modules.compliance_plugins.runners.winrm_runner import (
                            windows_winrm_runner,
                        )
                        from ..modules.compliance_plugins.runners.ssh_runner import (
                            linux_ssh_runner,
                        )
                        cd = {"shell": shell, "command": command, "expect": {"kind": "exit_zero"}}
                        r = (windows_winrm_runner if transport == "windows" else linux_ssh_runner)(cd, creds)
                        raw = r.raw_output or {}
                        return raw.get("stdout", ""), raw.get("exit_status", 1)

                    attrs = profile_software(_run, key, transport)
                    if attrs:
                        child.app_attributes_json = attrs
        except Exception:  # noqa: BLE001
            logger.exception("software profiling failed for %s (non-fatal)", key)
    db.add(child)
    db.flush()
    return child


@router.post("/{asset_id}/setup-software")
def setup_software(
    asset_id: int,
    body: SetupSoftwareIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Create a child application asset only when the chosen mode's gate passes.

    scan:            live DB preflight must succeed → child + IntegrationConnection
    host_connection: parent must already have an active OS connection → child only
    track:           no credential / no scan — inventory tracking only
    """
    import copy as _copy
    from ..crypto import encrypt_secret
    from ..models import IntegrationConnection
    from ..modules.compliance_plugins.services.preflight import preflight_check

    mode = (body.mode or "").strip().lower()
    if mode == "scan_via_host":
        mode = "host_connection"
    if mode not in ("scan", "host_connection", "track"):
        raise HTTPException(400, f"Unsupported mode {body.mode!r}")

    asset = _tenant_asset_or_404(db, current_user, asset_id)
    inventory = _copy.deepcopy(asset.detected_software_json or [])
    entry = next((e for e in inventory if e.get("software_key") == body.software_key), None)
    if entry is None:
        raise HTTPException(404, f"software_key {body.software_key!r} not in detected inventory")
    if entry.get("promoted_asset_id"):
        return {
            "created": False,
            "asset_id": entry["promoted_asset_id"],
            "reason": "already_promoted",
            "mode": mode,
        }

    platform = _software_platform(body.software_key)
    connection_id = None

    if mode == "scan":
        if not platform:
            raise HTTPException(
                400,
                "mode=scan requires a SQL software_key (postgres/mssql/mysql/oracle); "
                "use mode=host_connection for OS-scanned apps",
            )
        host = (body.hostname or asset.ip_address or asset.host_name or "").strip()
        user = (body.username or "").strip()
        pw = (body.password or "").strip()
        if not host or not user or not pw:
            raise HTTPException(400, "hostname, username, and password are required for mode=scan")
        default_port = {"postgres": 5432, "mssql": 1433, "mysql": 3306, "oracle": 1521}[platform]
        port = int(body.port or default_port)
        dbname = (body.database or "").strip() or (
            "postgres" if platform == "postgres" else
            "master" if platform == "mssql" else
            "information_schema" if platform == "mysql" else (body.database or "")
        )
        itype = {
            "postgres": "postgres_sql",
            "mssql": "mssql_sql",
            "mysql": "mysql_sql",
            "oracle": "oracle_sql",
        }[platform]
        prefix = {"postgres": "postgres", "mssql": "mssql", "mysql": "mysql", "oracle": "oracle"}[platform]
        test_creds = {
            f"{prefix}_host": host,
            f"{prefix}_port": port,
            f"{prefix}_username": user,
            f"{prefix}_password": pw,
            f"{prefix}_database": dbname or None,
        }
        if platform == "oracle":
            test_creds["oracle_service_name"] = dbname or None
            test_creds["oracle_sid"] = (body.oracle_sid or "").strip() or None
        # Oracle preflight may not exist yet — fall back to TCP-only then auth via runner later.
        pf = preflight_check(itype, test_creds)
        if not pf.ok and itype != "oracle_sql":
            raise HTTPException(
                status_code=400,
                detail={
                    "preflight_failed": True,
                    "code": pf.code,
                    "message": pf.detail,
                    "hint": "Fix credentials / reachability, then retry. No asset was created.",
                },
            )
        # console_url must match the child asset's host_name for scan/probe
        # lookup (multiple connections can share a host). The real DB endpoint
        # lives in credentials_extra_json[{prefix}_host].
        match_url = (asset.host_name or host).strip()
        label = (body.display_label or entry.get("name") or body.software_key).strip()
        conn_name = f"{label} @ {host}:{port}"
        extra = {
            f"{prefix}_host": host,
            f"{prefix}_port": port,
            f"{prefix}_username": user,
            f"{prefix}_password": encrypt_secret(pw),
            f"{prefix}_database": dbname or None,
        }
        if platform == "oracle":
            extra["oracle_service_name"] = dbname or None
            extra["oracle_sid"] = (body.oracle_sid or "").strip() or None
        existing = (
            db.query(IntegrationConnection)
            .filter(
                IntegrationConnection.tenant_id == asset.tenant_id,
                IntegrationConnection.connection_name == conn_name,
            )
            .first()
        )
        if existing:
            conn = existing
            conn.console_url = match_url
            conn.console_port = port
            conn.username = user
            conn.password = encrypt_secret(pw)
            conn.credentials_extra_json = extra
            conn.status = "connected"
            conn.is_active = True
            conn.integration_type = itype
            if hasattr(conn, "category") and not conn.category:
                conn.category = "compliance_db"
        else:
            conn = IntegrationConnection(
                tenant_id=asset.tenant_id,
                connection_name=conn_name,
                integration_type=itype,
                category="compliance_db",
                console_url=match_url,
                console_port=port,
                username=user,
                password=encrypt_secret(pw),
                credentials_extra_json=extra,
                status="connected",
                is_active=True,
                auth_method="basic",
            )
            db.add(conn)
            db.flush()
        connection_id = conn.id


    elif mode == "host_connection":
        from ..models import IntegrationConnection as _IC
        host_key = (asset.host_name or asset.ip_address or "").strip()
        parent_conn = None
        if host_key:
            parent_conn = (
                db.query(_IC)
                .filter(
                    _IC.tenant_id == asset.tenant_id,
                    _IC.is_active.is_(True),
                    _IC.console_url == host_key,
                )
                .first()
            )
        if parent_conn is None and asset.ip_address:
            parent_conn = (
                db.query(_IC)
                .filter(
                    _IC.tenant_id == asset.tenant_id,
                    _IC.is_active.is_(True),
                    _IC.console_url == asset.ip_address,
                )
                .first()
            )
        if parent_conn is None:
            raise HTTPException(
                400,
                detail={
                    "preflight_failed": True,
                    "code": "no_host_connection",
                    "message": (
                        f"Host '{asset.name}' has no active OS connection. "
                        "Connect the host first, then add this software."
                    ),
                    "hint": "Open Admin → Integrations → Connect, or Asset Discovery → Connect.",
                },
            )
        connection_id = parent_conn.id

    # track / scan / host_connection all create the child only after gates above
    child = _create_child_from_software_entry(db, asset, entry)
    # Stamp vendor so OS-detect / preferred-runner pick the SQL runner,
    # not the parent's Windows/Linux connection.
    if platform == "postgres":
        child.vendor = "postgresql"
    elif platform == "mssql":
        child.vendor = "microsoft"
    elif platform == "mysql":
        child.vendor = "mysql"
    elif platform == "oracle":
        child.vendor = "oracle"
    if mode == "scan" and connection_id:
        conn = db.query(IntegrationConnection).get(connection_id)
        if conn is not None and hasattr(conn, "asset_id"):
            conn.asset_id = child.id
    entry["promoted_asset_id"] = child.id
    asset.detected_software_json = inventory
    db.commit()
    return {
        "created": True,
        "asset_id": child.id,
        "name": child.name,
        "mode": mode,
        "connection_id": connection_id,
        "software_key": body.software_key,
    }


# ─────────────────────────────────────────────────────────────────────────────
# IP-group composite scoring ("room-and-chair" model)
# ─────────────────────────────────────────────────────────────────────────────
# Brought over from the Updated_CIS_Assests reference. The composite score is
# 60% host-OS plus 40% criticality-weighted application average. Apps scoring
# below 50% trigger a 10-point penalty against the room. Per-tenant weight
# overrides live on Tenant.settings.composite_weights.

_COMPOSITE_W_SELF = 0.6
_COMPOSITE_W_CHILDREN = 0.4
_BROKEN_CHAIR_PENALTY = 10.0   # percentage points
_UNKNOWN_CHAIR_PENALTY = 5.0
_CRIT_WEIGHT = {"low": 1.0, "medium": 2.0, "high": 3.0, "critical": 4.0}


def _get_tenant_crit_weights(db: Session, tid: int) -> dict:
    """Return the tenant's custom criticality weights, falling back to defaults."""
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    settings = tenant.settings or {} if tenant else {}
    return settings.get("composite_weights", _CRIT_WEIGHT)


def _tenant_asset_or_404(db: Session, current_user, asset_id: int) -> ITAsset:
    """Tenant scoped fetch for an asset id. 404 when not found in any of the
    user's tenants."""
    user_tenants = get_user_tenants(current_user, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Asset not found")
    return asset


def _asset_own_compliance(db: Session, asset_id: int, tenant_id: int) -> Optional[float]:
    """Latest-run pass-rate for the asset's own rules (percent), or None
    when it has never been scanned. Uses each plugin's most recent
    non-leaked run, mirroring the asset compliance page."""
    rows = db.execute(text("""
        SELECT r.status FROM grc_compliance_plugin_runs r
        JOIN (
            SELECT plugin_id, MAX(id) AS max_id
            FROM grc_compliance_plugin_runs
            WHERE asset_id = :aid AND tenant_id = :tid
              AND is_leaked = false
              AND status IN ('passed', 'failed')
            GROUP BY plugin_id
        ) latest ON latest.max_id = r.id
    """), {"aid": asset_id, "tid": tenant_id}).fetchall()
    if not rows:
        return None
    passed = sum(1 for r in rows if r[0] == "passed")
    return round(100.0 * passed / len(rows), 1)


class _CompositeWeightsIn(BaseModel):
    low: float = 1.0
    medium: float = 2.0
    high: float = 3.0
    critical: float = 4.0


@router.get("/composite-weights")
def get_composite_weights(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    settings = tenant.settings or {} if tenant else {}
    is_custom = "composite_weights" in settings
    weights = settings.get("composite_weights", _CRIT_WEIGHT)
    return {"weights": weights, "is_custom": is_custom, "defaults": _CRIT_WEIGHT}


@router.put("/composite-weights")
def update_composite_weights(
    body: _CompositeWeightsIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    weights = {"low": body.low, "medium": body.medium, "high": body.high, "critical": body.critical}
    if any(v <= 0 for v in weights.values()):
        raise HTTPException(400, "All weights must be positive numbers")
    settings = dict(tenant.settings or {})
    settings["composite_weights"] = weights
    tenant.settings = settings
    db.commit()
    return {"weights": weights, "is_custom": True, "defaults": _CRIT_WEIGHT}


@router.delete("/composite-weights")
def reset_composite_weights(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tid = get_user_primary_tenant(current_user, db)
    tenant = db.query(Tenant).filter(Tenant.id == tid).first()
    if not tenant:
        raise HTTPException(404, "Tenant not found")
    settings = dict(tenant.settings or {})
    settings.pop("composite_weights", None)
    tenant.settings = settings
    db.commit()
    return {"weights": _CRIT_WEIGHT, "is_custom": False, "defaults": _CRIT_WEIGHT}


@router.get("/{asset_id}/ip-peers")
def get_ip_peers(
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """IP-based asset group: all assets co-located at the same IP address.

    Primary linking mechanism for inventory-imported assets (e.g. from the
    bank's third-party CMDB). Assets sharing an IP are assumed to run on the
    same physical/virtual host.

    Returns per-asset benchmark availability, individual compliance scores,
    and the composite IP-group score + formula.
    """
    # Late imports keep the module import graph from growing at startup;
    # only loaded when the composite scoring endpoint is actually called.
    from ..modules.compliance_plugins.services.strict_matcher import pick_benchmark_for_os
    from ..modules.compliance_plugins.services.software_normaliser import benchmark_for_software_key
    from ..models import CompliancePlugin

    asset = _tenant_asset_or_404(db, current_user, asset_id)
    tid = asset.tenant_id
    ip = asset.ip_address

    # Patterns that indicate an app-level software key (not an OS host)
    _APP_PATTERNS = (
        "sql", "postgres", "mysql", "mongo", "tomcat", "iis",
        "apache", "nginx", "redis", "oracle-db", "mssql", "mariadb",
        "jboss", "wildfly", "websphere", "weblogic",
    )

    def _benchmark_for(a: ITAsset):
        bname = None
        mapping = pick_benchmark_for_os(db, tid, a.os_normalized)
        if mapping:
            bname = mapping.benchmark_name
        if not bname and a.os_normalized:
            bname = benchmark_for_software_key(db, a.os_normalized)
        count = 0
        if bname:
            from sqlalchemy import cast as _cast, String as _String
            # Count only rules that would ACTUALLY run — same exclusions as the
            # scanner. Otherwise the panel advertised "70 rules" for PostgreSQL
            # when 66 were unauthored auto-pass placeholders; the operator saw a
            # benchmark that looked complete and, if scanned, reported ~91%
            # compliant while checking nothing.
            count = db.query(CompliancePlugin).filter(
                CompliancePlugin.benchmark == bname,
                CompliancePlugin.enabled.is_(True),
                (CompliancePlugin.tenant_id == tid) | (CompliancePlugin.tenant_id.is_(None)),
                CompliancePlugin.runner_type != "manual",
                ~_cast(CompliancePlugin.check_definition, _String).ilike('%TODO%'),
                ~_cast(CompliancePlugin.check_definition, _String).ilike('%"kind": "any"%'),
                ~_cast(CompliancePlugin.check_definition, _String).ilike('%"kind":"any"%'),
            ).count()
        return bname, count

    # Pre-load every active integration connection in this tenant once so the
    # per-asset is_connected flag is a dict lookup, not N+1 queries.
    from ..models import IntegrationConnection as _IC
    _active_console_urls: set[str] = {
        (c.console_url or "").lower().strip()
        for c in db.query(_IC).filter(
            _IC.tenant_id == tid,
            _IC.is_active.is_(True),
        ).all()
        if (c.console_url or "").strip()
    }

    def _build(a: ITAsset):
        bname, rcount = _benchmark_for(a)
        score = _asset_own_compliance(db, a.id, tid)
        osk = (a.os_normalized or "").lower()
        is_host_os = (
            (a.asset_type or "") == "infrastructure" and
            not any(p in osk for p in _APP_PATTERNS)
        )
        # An asset is "connected" when an active integration connection exists
        # pinned to its host_name (matches the lookup _do_scan_all does). Drives
        # the room-scan checkbox eligibility — a peer that has no own connection
        # AND no host in its IP group with one can't actually be scanned.
        own_host = (a.host_name or "").lower().strip()
        is_connected = bool(own_host) and own_host in _active_console_urls
        return {
            "id": a.id,
            "name": a.name,
            "asset_type": a.asset_type or "infrastructure",
            "os_normalized": a.os_normalized,
            "criticality": a.criticality,
            "status": a.status,
            "benchmark_name": bname if rcount > 0 else None,
            "benchmark_available": bool(bname) and rcount > 0,
            "rule_count": rcount,
            "score": score,
            "never_scanned": score is None,
            "is_host_os": is_host_os,
            "is_self": a.id == asset_id,
            "is_connected": is_connected,
        }

    crit_weights = _get_tenant_crit_weights(db, tid)

    if not ip:
        return {
            "asset_id": asset_id,
            "ip_address": None,
            "group": [_build(asset)],
            # Standalone asset has no peers, so "group connection availability"
            # collapses to "is the opened asset itself connected?".
            "connection_available": _build(asset)["is_connected"],
            "composite": None,
            "formula": {
                "description": "Asset has no IP address. Standalone, not part of an IP group.",
                "weights": {"host": _COMPOSITE_W_SELF, "applications": _COMPOSITE_W_CHILDREN},
                "criticality_weights": crit_weights,
            },
        }

    all_at_ip = db.query(ITAsset).filter(
        ITAsset.tenant_id == tid,
        ITAsset.ip_address == ip,
    ).order_by(ITAsset.asset_type, ITAsset.id).all()

    # Duplicate guard: deduplicate by id (shouldn't happen, but be safe)
    seen_ids = set()
    deduped = []
    for a in all_at_ip:
        if a.id not in seen_ids:
            seen_ids.add(a.id)
            deduped.append(a)

    group = [_build(a) for a in deduped]

    # Composite score (room-and-chair formula, IP-group edition)
    host_entry = next((g for g in group if g["is_host_os"]), None)
    app_entries = [g for g in group if not g["is_host_os"]]

    host_score = host_entry["score"] if host_entry else None
    app_contributions = []
    weighted_sum, weight_total = 0.0, 0.0
    any_broken = False
    for a in app_entries:
        w = crit_weights.get((a["criticality"] or "medium").lower(), 2.0)
        s = a["score"]
        app_contributions.append({
            "asset_id": a["id"], "name": a["name"],
            "score": s, "weight": w, "criticality": a["criticality"],
            "os_normalized": a.get("os_normalized"),
        })
        if s is not None:
            weighted_sum += s * w
            weight_total += w
            if s < 50.0:
                any_broken = True

    effective = None
    penalties = []
    if host_score is not None:
        if weight_total > 0:
            app_avg = weighted_sum / weight_total
            effective = _COMPOSITE_W_SELF * host_score + _COMPOSITE_W_CHILDREN * app_avg
        else:
            effective = float(host_score)
        if any_broken:
            effective -= _BROKEN_CHAIR_PENALTY
            penalties.append({"reason": "application scoring below 50%", "points": _BROKEN_CHAIR_PENALTY})
        effective = round(max(0.0, min(100.0, effective)), 1)
    elif app_entries:
        # No OS host in the group, equal-weight average of all apps
        scored = [a["score"] for a in app_entries if a["score"] is not None]
        if scored:
            effective = round(sum(scored) / len(scored), 1)

    weakest = None
    scored_all = [g for g in group if g["score"] is not None]
    if scored_all:
        weakest = min(scored_all, key=lambda g: g["score"])

    # Group-level connection availability — true when ANY member of the IP
    # group has an active integration connection. Room-scan execution piggy-
    # backs off that connection, so this is the right gate for the panel
    # checkboxes: peers in a group with at least one connected member can be
    # ticked even if they have no connection of their own.
    connection_available = any(g.get("is_connected") for g in group)

    return {
        "asset_id": asset_id,
        "ip_address": ip,
        "group": group,
        "connection_available": connection_available,
        "composite": {
            "host_id": host_entry["id"] if host_entry else None,
            "host_score": host_score,
            "app_contributions": app_contributions,
            "effective_score": effective,
            "weakest": weakest,
            "penalties": penalties,
        },
        "formula": {
            "description": (
                f"{int(_COMPOSITE_W_SELF * 100)}% host OS score "
                f"plus {int(_COMPOSITE_W_CHILDREN * 100)}% criticality-weighted application average"
            ),
            "weights": {"host": _COMPOSITE_W_SELF, "applications": _COMPOSITE_W_CHILDREN},
            "criticality_weights": crit_weights,
            "penalties": {
                "broken_app": f"minus {_BROKEN_CHAIR_PENALTY} pts when any app scores below 50%",
            },
        },
    }
