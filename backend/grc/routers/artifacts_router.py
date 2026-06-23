import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session
from io import BytesIO

from ..models import (
    ArtifactCatalogItem,
    TenantArtifact,
    GRCUser,
    ITAsset,
    Risk,
    UploadedFramework,
    ParsedFrameworkControl,
    get_db,
)
from .auth_router import require_auth, get_user_primary_tenant

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/artifacts", tags=["Artifacts"])

# ---------------------------------------------------------------------------
# Catalog seed path
# ---------------------------------------------------------------------------
_CATALOG_PATH = Path(__file__).parent.parent / "seed_data" / "artifact_catalog.json"

# Map assessment_type strings → catalog framework_key
ASSESSMENT_TYPE_MAP: dict[str, str] = {
    # ISO 27001
    "iso_27001": "iso_27001_2022",
    "iso27001": "iso_27001_2022",
    "iso_iec_27001": "iso_27001_2022",
    "iso 27001": "iso_27001_2022",
    "iso/iec 27001": "iso_27001_2022",
    "iso/iec 27001:2022": "iso_27001_2022",
    "iso 27001:2022": "iso_27001_2022",
    # ISO 41001
    "iso_41001": "iso_41001_2018",
    "iso41001": "iso_41001_2018",
    "iso 41001": "iso_41001_2018",
    "iso 41001:2018": "iso_41001_2018",
    # PCI DSS
    "pci_dss": "pci_dss_v4",
    "pci dss": "pci_dss_v4",
    "pci": "pci_dss_v4",
    "pci dss v4": "pci_dss_v4",
    "pci dss v4.0": "pci_dss_v4",
    "pci dss v4.0.1": "pci_dss_v4",
    # SWIFT
    "swift": "swift_cscf",
    "swift_cscf": "swift_cscf",
    "swift cscf": "swift_cscf",
    # COBIT
    "cobit": "cobit_2019",
    "cobit_2019": "cobit_2019",
    "cobit 2019": "cobit_2019",
    # DORA
    "dora": "dora",
    # HIPAA
    "hipaa": "hipaa",
    # GDPR
    "gdpr": "gdpr",
    # PDPL / KSA
    "pdpl": "pdpl_ksa",
    "pdpl_ksa": "pdpl_ksa",
    "saudi_pdpl": "pdpl_ksa",
    "ksa pdpl": "pdpl_ksa",
    # CIS Controls
    "cis": "cis_v8",
    "cis_controls": "cis_v8",
    "cis_v8": "cis_v8",
    "cis controls": "cis_v8",
    "cis controls v8": "cis_v8",
    # SOX
    "sox": "sox_itgc",
    "sox_itgc": "sox_itgc",
    "sox itgc": "sox_itgc",
    # SOC 2
    "soc2": "soc2",
    "soc_2": "soc2",
    "soc 2": "soc2",
    "soc 2 type ii": "soc2",
    # NIST CSF
    "nist_csf": "nist_csf_2",
    "nist csf": "nist_csf_2",
    "nist_csf_2.0": "nist_csf_2",
    "nist_csf_2": "nist_csf_2",
    "nist csf 2.0": "nist_csf_2",
    # ISO 22301
    "iso_22301": "iso_22301_2019",
    "iso22301": "iso_22301_2019",
    "bcm": "iso_22301_2019",
    "iso 22301": "iso_22301_2019",
    "iso 22301:2019": "iso_22301_2019",
    "iso/iec 22301": "iso_22301_2019",
    "iso/iec 22301:2019": "iso_22301_2019",
    # NIS 2
    "nis2": "nis2",
    "nis_2": "nis2",
    "nis 2": "nis2",
    "nis2 directive": "nis2",
    # ADHICS
    "adhics": "adhics",
    "abu dhabi healthcare": "adhics",
    "abu dhabi healthcare information and cyber security standard": "adhics",
    "abu dhabi healthcare information and cyber security": "adhics",
    # ISO 42001 (AI Management) — distinct from ISO 41001 (Facility Mgmt)
    "iso_42001": "iso_42001_2023",
    "iso42001": "iso_42001_2023",
    "iso 42001": "iso_42001_2023",
    "iso 42001:2023": "iso_42001_2023",
    "iso/iec 42001": "iso_42001_2023",
    "iso/iec 42001:2023": "iso_42001_2023",
    "iso/iec 42001:2023 ai management system": "iso_42001_2023",
    # NIST AI RMF
    "nist ai rmf": "nist_ai_rmf",
    "nist ai rmf 1.0": "nist_ai_rmf",
    "ai rmf": "nist_ai_rmf",
    "nist artificial intelligence risk management framework": "nist_ai_rmf",
    "nist artificial intelligence risk management framework (ai rmf 1.0)": "nist_ai_rmf",
    # NIST SP 800-53 Rev 5
    "nist sp 800-53": "nist_sp_800_53_r5",
    "nist 800-53": "nist_sp_800_53_r5",
    "sp 800-53": "nist_sp_800_53_r5",
    "nist sp 800-53 rev 5": "nist_sp_800_53_r5",
    # HITRUST CSF
    "hitrust": "hitrust_csf",
    "hitrust csf": "hitrust_csf",
    "hitrust common security framework": "hitrust_csf",
    "hitrust common security framework (csf)": "hitrust_csf",
    # ARAMCO
    "aramco": "aramco_csc",
    "aramco csc": "aramco_csc",
    "aramco cybersecurity compliance certification": "aramco_csc",
    # ADHIE
    "adhie": "adhie",
    "abu dhabi health information exchange": "adhie",
    "doh policy on the abu dhabi health information exchange (adhie)": "adhie",
    # KSA NDMO
    "ndmo": "ksa_ndmo",
    "ksa ndmo": "ksa_ndmo",
    "ksa national data management": "ksa_ndmo",
    "ksa national data management and personal data protection standards": "ksa_ndmo",
    # MAS TRM
    "mas trm": "mas_trm",
    "mas technology risk management": "mas_trm",
    "mas technology risk management guidelines": "mas_trm",
    # Qatar Central Bank
    "qcb": "qatar_cb",
    "qatar central bank": "qatar_cb",
    "qatar central bank technology risks circular": "qatar_cb",
    # SABIC CyberTrust
    "sabic": "sabic_cybertrust",
    "sabic cybertrust": "sabic_cybertrust",
    "sabic cybertrust guidelines": "sabic_cybertrust",
    # SAMA
    "sama": "sama_csf",
    "sama csf": "sama_csf",
    "sama cyber security framework": "sama_csf",
    # SBP variants
    "sbp cloud outsourcing": "sbp_cloud_outsourcing",
    "sbp cloud outsourcing framework": "sbp_cloud_outsourcing",
    "sbp etgrmf": "sbp_etgrmf",
    "sbp internet banking": "sbp_internet_banking",
    "sbp internet banking framework": "sbp_internet_banking",
    # Sri Lanka BSS
    "sri lanka bss": "sri_lanka_bss",
    "sri lanka baseline security standard": "sri_lanka_bss",
    "sri lanka baseline security standard (bss)": "sri_lanka_bss",
    # KSA Personal Data Transfer Regulation (PDPL implementing regulation)
    "ksa data transfer": "ksa_data_transfer",
    "personal data transfer": "ksa_data_transfer",
    "data transfer outside kingdom": "ksa_data_transfer",
    "regulation on personal data transfer outside the kingdom": "ksa_data_transfer",
    "regulation on personal data transfer outside ksa": "ksa_data_transfer",
}

# Artifact names that map to real platform data
PLATFORM_NATIVE_PATTERNS = {
    "risk register": "risk_register",
    "risk_register": "risk_register",
    "asset inventory": "asset_inventory",
    "asset register": "asset_inventory",
    "it asset": "asset_inventory",
    "enterprise risk register": "risk_register",
    "supply chain risk register": "risk_register",
}


def _get_platform_data_type(artifact_name: str) -> Optional[str]:
    name_lower = artifact_name.lower()
    for pattern, dtype in PLATFORM_NATIVE_PATTERNS.items():
        if pattern in name_lower:
            return dtype
    return None


def _ensure_catalog_seeded(db: Session) -> None:
    """Ensure artifact tables exist and seed ArtifactCatalogItem table once if empty."""
    # Create tables if they don't exist yet (existing tenant DBs may be missing them)
    try:
        from ..models import Base
        bind = db.get_bind()
        engine = getattr(bind, "engine", bind)
        Base.metadata.create_all(engine, checkfirst=True)
    except Exception:
        logger.exception("Failed to auto-create artifact tables")

    if not _CATALOG_PATH.exists():
        logger.warning("Artifact catalog JSON not found at %s", _CATALOG_PATH)
        return
    with open(_CATALOG_PATH, encoding="utf-8") as f:
        catalog: dict = json.load(f)

    # Self-heal: seed ONLY framework_keys not already present, so frameworks
    # added to the JSON after a tenant DB was first seeded still land in that
    # DB. (Previously this returned early whenever the table was non-empty, so
    # newly added frameworks never appeared for existing tenants.)
    try:
        existing_keys = {row[0] for row in db.query(ArtifactCatalogItem.framework_key).distinct().all()}
    except Exception:
        logger.exception("Cannot query artifact catalog table")
        return
    missing_keys = [k for k in catalog if k not in existing_keys]
    if not missing_keys:
        return

    items: list[ArtifactCatalogItem] = []
    for fw_key in missing_keys:
        fw_data = catalog[fw_key]
        fw_name = fw_data["name"]
        for art in fw_data["artifacts"]:
            stage_str: str = art.get("stage", "")
            stage_num = None
            if stage_str.startswith("Stage "):
                try:
                    stage_num = int(stage_str.split(" ")[1].rstrip(":"))
                except (IndexError, ValueError):
                    pass
            platform_type = _get_platform_data_type(art.get("name", ""))
            items.append(
                ArtifactCatalogItem(
                    framework_key=fw_key,
                    framework_name=fw_name,
                    artifact_id=art.get("artifact_id", ""),
                    stage=stage_str,
                    stage_number=stage_num,
                    name=art.get("name", ""),
                    artifact_type=art.get("type", ""),
                    control_ref=art.get("control_ref") or None,
                    mandatory=art.get("mandatory", False),
                    description=art.get("description") or None,
                    format=art.get("format") or None,
                    owner=art.get("owner") or None,
                    is_platform_native=platform_type is not None,
                    platform_data_type=platform_type,
                )
            )
    db.bulk_save_objects(items)
    db.commit()
    logger.info("Seeded %d artifact catalog items across %d new framework(s)", len(items), len(missing_keys))


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class ArtifactCatalogOut(BaseModel):
    id: int
    framework_key: str
    framework_name: str
    artifact_id: str
    stage: str
    stage_number: Optional[int]
    name: str
    artifact_type: str
    control_ref: Optional[str]
    mandatory: bool
    description: Optional[str]
    format: Optional[str]
    owner: Optional[str]
    is_platform_native: bool
    platform_data_type: Optional[str]

    class Config:
        from_attributes = True


class TenantArtifactOut(BaseModel):
    id: int
    tenant_id: int
    catalog_item_id: Optional[int]
    assessment_id: Optional[int]
    framework_key: str
    name: str
    artifact_type: str
    stage: Optional[str]
    control_ref: Optional[str]
    description: Optional[str]
    format: Optional[str]
    content: Optional[str]
    file_name: Optional[str]
    file_size: Optional[int]
    status: str
    assigned_to_id: Optional[int]
    assigned_to_name: Optional[str] = None
    created_by_id: Optional[int]
    created_by_name: Optional[str] = None
    is_platform_native: bool
    platform_data_type: Optional[str]
    platform_record_count: Optional[int]
    created_at: datetime
    updated_at: Optional[datetime]

    class Config:
        from_attributes = True


class TenantArtifactCreate(BaseModel):
    catalog_item_id: Optional[int] = None
    assessment_id: Optional[int] = None
    framework_key: str
    name: str
    artifact_type: str
    stage: Optional[str] = None
    control_ref: Optional[str] = None
    description: Optional[str] = None
    format: Optional[str] = None
    content: Optional[str] = None
    assigned_to_id: Optional[int] = None
    is_platform_native: bool = False
    platform_data_type: Optional[str] = None


class TenantArtifactUpdate(BaseModel):
    name: Optional[str] = None
    artifact_type: Optional[str] = None
    stage: Optional[str] = None
    description: Optional[str] = None
    content: Optional[str] = None
    status: Optional[str] = None
    assigned_to_id: Optional[int] = None
    format: Optional[str] = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _slugify_framework(raw: str) -> str:
    """Last-resort slug for an unmapped framework. Stable, deterministic,
    safe to use as a `framework_key` for storing/retrieving tenant artifacts.

    Example: "ISO/IEC 42001:2023 AI Management System"
          -> "iso_iec_42001_2023_ai_management_system"
    """
    s = (raw or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s[:64] or "framework"


def _resolve_framework_key(raw: str) -> Optional[str]:
    """Resolve a free-form framework label to a stable `framework_key`.

    Always returns a non-empty key when `raw` is non-empty — falls back to
    a slugified form of the input when no map entry hits. That guarantee
    matters because the frontend gates its Create-Artifact modal on
    `data.framework_key` being truthy; if we returned None for unmapped
    frameworks, the Create button silently no-opped.
    """
    if not raw:
        return None
    normalized = raw.lower().strip()
    # Exact match first
    hit = ASSESSMENT_TYPE_MAP.get(normalized)
    if hit:
        return hit
    # Substring fallback: find the longest key that appears in the normalized name
    best = None
    best_len = 0
    for key, val in ASSESSMENT_TYPE_MAP.items():
        if key in normalized and len(key) > best_len:
            best = val
            best_len = len(key)
    if best:
        return best
    # No curated mapping — derive a deterministic slug so the framework
    # still has a usable key for storing tenant artifacts.
    return _slugify_framework(raw)


# ---------------------------------------------------------------------------
# Control-ref fuzzy matching
# ---------------------------------------------------------------------------
# Catalog/tenant `control_ref` values are free-form strings ("Cl. 5.1",
# "A.5.9", "Cl. 6.1.2 / 8.2"). To surface a catalog item next to a parsed
# framework requirement, we compare normalized identifiers token-by-token.
#
# Examples handled correctly:
#   catalog "Cl. 5.1" + requirement "5.1"        → match
#   catalog "Cl. 6.1.2 / 8.2" + requirement "8.2" → match (split on / and ,)
#   catalog "A.5.9" + requirement "A.5.9"          → match
#   catalog "5.1" + requirement "5.10"            → NO match (token-equal,
#                                                    not substring — guards
#                                                    against false positives)
# ---------------------------------------------------------------------------

_REF_PREFIX_RE = re.compile(r'^(cl\.|clause|cl|c\.|control)\s*', re.IGNORECASE)


def _normalize_ref(s: Optional[str]) -> str:
    if not s:
        return ""
    out = s.strip().lower()
    out = _REF_PREFIX_RE.sub("", out)
    out = re.sub(r"\s+", "", out)
    return out


def _split_catalog_ref(s: Optional[str]) -> List[str]:
    """A single catalog control_ref may bundle multiple references with
    `/` or `,` separators. Return the normalized tokens for each.
    """
    if not s:
        return []
    parts = re.split(r"[/,]\s*", s)
    return [n for n in (_normalize_ref(p) for p in parts) if n]


def _ref_matches_any(catalog_ref: Optional[str], requirement_codes: List[str]) -> bool:
    catalog_tokens = set(_split_catalog_ref(catalog_ref))
    if not catalog_tokens:
        return False
    for code in requirement_codes:
        if _normalize_ref(code) in catalog_tokens:
            return True
    return False


def _infer_artifact_type_from_evidence(name: str, filetype: Optional[str]) -> str:
    """Derive a logical artifact type (Policy / Register / ...) from an
    evidence requirement's name + filetype. Used when no seeded catalog
    entry exists, so the dynamic top-level catalog still gets sensible
    type badges — matching the inline per-requirement view's behaviour.
    """
    combined = (name or "").lower()
    if "policy" in combined:
        return "Policy"
    if "procedure" in combined:
        return "Procedure"
    if "register" in combined or "inventory" in combined:
        return "Register"
    if "matrix" in combined:
        return "Matrix"
    if "plan" in combined:
        return "Plan"
    if "contract" in combined or "agreement" in combined or "nda" in combined:
        return "Contract"
    if "attest" in combined or "acknowledg" in combined:
        return "Attestation"
    if "config" in combined:
        return "Configuration"
    if "screenshot" in combined:
        return "Screenshot"
    if "training" in combined or "awareness" in combined:
        return "Training Record"
    if "audit" in combined or "assessment" in combined:
        return "Assessment"
    if "report" in combined:
        return "Report"
    if "log" in combined or "minute" in combined or "email" in combined:
        return "Record/Log"
    if "certificate" in combined:
        return "Evidence"
    ft = (filetype or "").upper()
    if ft == "XLSX":
        return "Register"
    if ft == "EML":
        return "Record/Log"
    return "Evidence"


def _derive_catalog_from_framework(
    db: Session,
    *,
    tenant_id: int,
    assessment_type: str,
) -> Optional[dict]:
    """Build a virtual catalog from a tenant's uploaded framework when no
    seeded `ArtifactCatalogItem` rows exist for the framework.

    Aggregates the `evidence_requirements` JSON across every parsed control
    on the matching `UploadedFramework`, deduplicates by name (case-folded),
    and tags each row with the controls (and their domains) where it was
    asked for. Lets the user see and create the same set of artifacts from
    the top-level Artifacts tab that they'd otherwise have to dig into each
    requirement to find.

    Returns ``None`` when no framework matches the caller's
    `assessment_type` — the catalog endpoint then falls back to the empty
    seed-catalog response so the UI message stays explicit.
    """
    if not assessment_type:
        return None
    needle = assessment_type.strip()
    if not needle:
        return None

    # Match by exact name first, then case-insensitive contains. Restrict to
    # this tenant's frameworks plus shared library frameworks (tenant_id is
    # NULL on shared rows in some deployments).
    fw_q = db.query(UploadedFramework).filter(
        UploadedFramework.is_active.is_(True),
        ((UploadedFramework.tenant_id == tenant_id) | (UploadedFramework.tenant_id.is_(None))),
    )
    framework = fw_q.filter(UploadedFramework.name == needle).first()
    if not framework:
        framework = fw_q.filter(UploadedFramework.name.ilike(f"%{needle}%")).first()
    if not framework:
        return None

    controls = (
        db.query(ParsedFrameworkControl)
        .filter(ParsedFrameworkControl.uploaded_framework_id == framework.id)
        .all()
    )
    if not controls:
        return None

    # Aggregate by lowercased name so HR/IT/etc. domains don't each produce
    # their own copy of "Roles and responsibilities matrix".
    by_key: dict[str, dict] = {}
    for c in controls:
        evs = c.evidence_requirements or []
        if not isinstance(evs, list):
            continue
        for ev in evs:
            if not isinstance(ev, dict):
                continue
            ev_name = (ev.get("name") or ev.get("title") or "").strip()
            if not ev_name:
                continue
            key = ev_name.lower()
            file_type = (ev.get("filetype") or ev.get("format") or "DOCX").upper()
            entry = by_key.get(key)
            ref = c.original_reference or c.control_id or ""
            domain = c.domain or "General"
            if entry is None:
                by_key[key] = {
                    "name": ev_name,
                    "description": (ev.get("description") or "").strip(),
                    "format": file_type,
                    "artifact_type": _infer_artifact_type_from_evidence(ev_name, file_type),
                    "control_refs": {ref} if ref else set(),
                    "domains": {domain},
                    "mandatory": bool(c.is_mandatory),
                }
            else:
                if ref:
                    entry["control_refs"].add(ref)
                entry["domains"].add(domain)
                # Keep the longest non-empty description we see.
                desc = (ev.get("description") or "").strip()
                if desc and len(desc) > len(entry["description"]):
                    entry["description"] = desc
                # Promote to mandatory if any control flags it so.
                if c.is_mandatory:
                    entry["mandatory"] = True

    if not by_key:
        return None

    # Domain → stage mapping so the catalog renders grouped tabs even
    # though the framework JSON itself isn't staged. The stage label is the
    # control's `domain` field with a stable index for sorting.
    domain_order: dict[str, int] = {}
    for c in controls:
        d = c.domain or "General"
        if d not in domain_order:
            domain_order[d] = len(domain_order) + 1

    items: List[dict] = []
    # Stable iteration order: by domain index, then name.
    for idx, (key, entry) in enumerate(
        sorted(by_key.items(), key=lambda kv: (min(domain_order.get(d, 999) for d in kv[1]["domains"]), kv[0])),
        start=1,
    ):
        primary_domain = sorted(entry["domains"], key=lambda d: domain_order.get(d, 999))[0]
        stage_num = domain_order.get(primary_domain, idx)
        items.append({
            "id": -idx,  # negative ids — virtual, no DB row
            "artifact_id": f"virtual_{idx:03d}",
            "stage": f"Stage {stage_num}: {primary_domain}",
            "stage_number": stage_num,
            "name": entry["name"],
            "artifact_type": entry["artifact_type"],
            "control_ref": ", ".join(sorted(entry["control_refs"])) if entry["control_refs"] else None,
            "mandatory": entry["mandatory"],
            "description": entry["description"] or None,
            "format": entry["format"],
            "owner": None,
            "is_platform_native": False,
            "platform_data_type": None,
        })

    stages = sorted({i["stage"] for i in items}, key=lambda s: int(s.split(" ")[1].rstrip(":")) if s.startswith("Stage ") else 999)
    return {
        "framework_key": (framework.name or assessment_type).lower().replace(" ", "_")[:64],
        "framework_name": framework.name or assessment_type,
        "items": items,
        "stages": stages,
    }


def _artifact_out(artifact: TenantArtifact) -> dict:
    assigned_name = None
    if artifact.assigned_to:
        u = artifact.assigned_to
        assigned_name = getattr(u, "display_name", None) or getattr(u, "username", None) or getattr(u, "email", None)
    created_name = None
    if artifact.created_by:
        u = artifact.created_by
        created_name = getattr(u, "display_name", None) or getattr(u, "username", None) or getattr(u, "email", None)
    return {
        "id": artifact.id,
        "tenant_id": artifact.tenant_id,
        "catalog_item_id": artifact.catalog_item_id,
        "assessment_id": artifact.assessment_id,
        "framework_key": artifact.framework_key,
        "name": artifact.name,
        "artifact_type": artifact.artifact_type,
        "stage": artifact.stage,
        "control_ref": artifact.control_ref,
        "description": artifact.description,
        "format": artifact.format,
        "content": artifact.content,
        "file_name": artifact.file_name,
        "file_size": artifact.file_size,
        "status": artifact.status,
        "assigned_to_id": artifact.assigned_to_id,
        "assigned_to_name": assigned_name,
        "created_by_id": artifact.created_by_id,
        "created_by_name": created_name,
        "is_platform_native": artifact.is_platform_native or False,
        "platform_data_type": artifact.platform_data_type,
        "platform_record_count": artifact.platform_record_count,
        "created_at": artifact.created_at,
        "updated_at": artifact.updated_at,
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/catalog")
def get_catalog(
    framework_key: Optional[str] = Query(None),
    assessment_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """List catalog items for a framework. Accept either framework_key or assessment_type."""
    _ensure_catalog_seeded(db)
    tenant_id = get_user_primary_tenant(current_user, db)
    resolved_key = framework_key
    if not resolved_key and assessment_type:
        resolved_key = _resolve_framework_key(assessment_type)

    # If the framework is mapped, return its seeded catalog rows.
    items: list = []
    if resolved_key:
        items = (
            db.query(ArtifactCatalogItem)
            .filter(ArtifactCatalogItem.framework_key == resolved_key)
            .order_by(ArtifactCatalogItem.stage_number, ArtifactCatalogItem.artifact_id)
            .all()
        )

    # Dynamic fallback: when no seeded rows are returned (either because the
    # framework isn't in `ASSESSMENT_TYPE_MAP`, or because it is mapped but
    # `artifact_catalog.json` doesn't curate it yet — ADHICS is the
    # canonical example), build a virtual catalog from the tenant's parsed
    # framework controls so the top-level Artifacts tab matches what the
    # inline per-requirement compliance-artifacts section already shows.
    if not items and assessment_type:
        derived = _derive_catalog_from_framework(
            db, tenant_id=tenant_id, assessment_type=assessment_type,
        )
        if derived:
            return derived

    if not resolved_key:
        return {"framework_key": None, "framework_name": None, "items": [], "stages": []}

    stages = sorted({i.stage for i in items}, key=lambda s: (s.split(" ")[1] if s.startswith("Stage ") else s))
    # Prefer the framework_name from a seeded catalog row; otherwise echo the
    # caller's `assessment_type` so the UI displays "Abu Dhabi Healthcare ..."
    # instead of the bare slug ("adhics") when a framework is mapped but has
    # no curated catalog entries yet.
    if items:
        fw_name = items[0].framework_name
    elif assessment_type:
        fw_name = assessment_type
    else:
        fw_name = resolved_key
    return {
        "framework_key": resolved_key,
        "framework_name": fw_name,
        "items": [
            {
                "id": i.id,
                "artifact_id": i.artifact_id,
                "stage": i.stage,
                "stage_number": i.stage_number,
                "name": i.name,
                "artifact_type": i.artifact_type,
                "control_ref": i.control_ref,
                "mandatory": i.mandatory,
                "description": i.description,
                "format": i.format,
                "owner": i.owner,
                "is_platform_native": i.is_platform_native,
                "platform_data_type": i.platform_data_type,
            }
            for i in items
        ],
        "stages": stages,
    }


@router.get("/by-control")
def list_artifacts_by_control(
    framework_key: Optional[str] = Query(None),
    assessment_type: Optional[str] = Query(None),
    control_ref: str = Query(..., description="Comma-separated requirement reference(s) — e.g. '5.1' or 'A.5.9' or '5.1,5.2'"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Return catalog items + tenant artifacts whose `control_ref` matches
    the provided requirement reference(s). Used to render an inline
    "Compliance Artifacts" section per requirement on the framework detail
    page, so users can see and create artifacts without leaving the row.

    Empty result (`catalog: []`) is normal — most parsed-framework
    requirements don't have a corresponding artifact catalog entry, and the
    UI hides the section in that case.
    """
    _ensure_catalog_seeded(db)
    tenant_id = get_user_primary_tenant(current_user, db)

    resolved_key = framework_key
    if not resolved_key and assessment_type:
        resolved_key = _resolve_framework_key(assessment_type)
    if not resolved_key:
        return {
            "framework_key": None,
            "control_ref": control_ref,
            "catalog": [],
            "artifacts": [],
        }

    requirement_codes = [c.strip() for c in (control_ref or "").split(",") if c.strip()]
    if not requirement_codes:
        return {
            "framework_key": resolved_key,
            "control_ref": control_ref,
            "catalog": [],
            "artifacts": [],
        }

    catalog_items = (
        db.query(ArtifactCatalogItem)
        .filter(ArtifactCatalogItem.framework_key == resolved_key)
        .all()
    )
    matched_catalog = [
        i for i in catalog_items
        if _ref_matches_any(i.control_ref, requirement_codes)
    ]

    tenant_artifacts = (
        db.query(TenantArtifact)
        .filter(
            TenantArtifact.tenant_id == tenant_id,
            TenantArtifact.framework_key == resolved_key,
        )
        .all()
    )
    matched_tenant = [
        a for a in tenant_artifacts
        if _ref_matches_any(a.control_ref, requirement_codes)
    ]

    return {
        "framework_key": resolved_key,
        "control_ref": control_ref,
        "catalog": [
            {
                "id": i.id,
                "artifact_id": i.artifact_id,
                "stage": i.stage,
                "stage_number": i.stage_number,
                "name": i.name,
                "artifact_type": i.artifact_type,
                "control_ref": i.control_ref,
                "mandatory": i.mandatory,
                "description": i.description,
                "format": i.format,
                "owner": i.owner,
                "is_platform_native": i.is_platform_native,
                "platform_data_type": i.platform_data_type,
            }
            for i in matched_catalog
        ],
        "artifacts": [_artifact_out(a) for a in matched_tenant],
    }


@router.get("/platform-data/risk-register")
def get_platform_risk_register(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    risks = (
        db.query(Risk)
        .filter(Risk.tenant_id == tenant_id)
        .order_by(Risk.created_at.desc())
        .limit(200)
        .all()
    )
    return {
        "total": db.query(func.count(Risk.id)).filter(Risk.tenant_id == tenant_id).scalar(),
        "items": [
            {
                "id": r.id,
                "title": r.title,
                "category": r.category,
                "status": r.status,
                "inherent_score": r.inherent_score,
                "residual_score": r.residual_score,
                "created_at": r.created_at,
            }
            for r in risks
        ],
    }


@router.get("/platform-data/asset-inventory")
def get_platform_asset_inventory(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    assets = (
        db.query(ITAsset)
        .filter(ITAsset.tenant_id == tenant_id)
        .order_by(ITAsset.created_at.desc())
        .limit(200)
        .all()
    )
    return {
        "total": db.query(func.count(ITAsset.id)).filter(ITAsset.tenant_id == tenant_id).scalar(),
        "items": [
            {
                "id": a.id,
                "name": a.name,
                "asset_type": a.asset_type,
                "criticality": a.criticality,
                "status": a.status,
                "owner_name": a.owner_name,
                "created_at": a.created_at,
            }
            for a in assets
        ],
    }


def _safe_filename(name: str) -> str:
    """Make a filename safe for Content-Disposition headers."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "artifact").strip())
    return cleaned[:80] or "artifact"


def _build_risk_register_xlsx(db: Session, tenant_id: int, title: str) -> BytesIO:
    """Generate an XLSX workbook containing the tenant's full risk register.

    Pulls every Risk row for the tenant (not just the latest 200 like the
    preview endpoint) so the exported file is a complete snapshot suitable
    for evidence / audit consumption.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = [
        "ID", "Title", "Description", "Category", "Sub-Category",
        "Register Type", "Status", "Owner", "Business Owner",
        "Inherent Likelihood", "Inherent Impact", "Inherent Score",
        "Residual Likelihood", "Residual Impact", "Residual Score",
        "Risk Appetite", "Treatment Plan", "Source Type",
        "Due Date", "Review Date", "Created At", "Updated At",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    risks = (
        db.query(Risk)
        .filter(Risk.tenant_id == tenant_id)
        .order_by(Risk.id.asc())
        .all()
    )
    for r in risks:
        owner_name = getattr(getattr(r, "owner", None), "display_name", None) \
            or getattr(getattr(r, "owner", None), "username", None) or ""
        bo_name = getattr(getattr(r, "business_owner", None), "display_name", None) \
            or getattr(getattr(r, "business_owner", None), "username", None) or ""
        ws.append([
            r.id, r.title, r.description, r.risk_category or r.category,
            r.risk_sub_category, r.register_type, r.status, owner_name, bo_name,
            r.inherent_likelihood, r.inherent_impact, r.inherent_score,
            r.residual_likelihood, r.residual_impact, r.residual_score,
            r.risk_appetite, r.treatment_plan, getattr(r, "source_type", None),
            r.due_date.isoformat() if r.due_date else None,
            r.review_date.isoformat() if r.review_date else None,
            r.created_at.isoformat() if r.created_at else None,
            r.updated_at.isoformat() if r.updated_at else None,
        ])

    # Reasonable default column widths so the file is readable as-shipped.
    widths = [6, 40, 60, 16, 22, 18, 12, 22, 22, 10, 10, 10, 10, 10, 10, 14, 50, 16, 12, 12, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_asset_inventory_xlsx(db: Session, tenant_id: int, title: str) -> BytesIO:
    """Generate an XLSX workbook containing the tenant's full IT asset inventory."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asset Inventory"

    headers = [
        "ID", "Name", "Asset Type", "Criticality", "Status",
        "Owner", "Location", "Description",
        "Created At", "Updated At",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    assets = (
        db.query(ITAsset)
        .filter(ITAsset.tenant_id == tenant_id)
        .order_by(ITAsset.id.asc())
        .all()
    )
    for a in assets:
        ws.append([
            a.id,
            a.name,
            a.asset_type,
            a.criticality,
            a.status,
            getattr(a, "owner_name", None),
            getattr(a, "location", None),
            getattr(a, "description", None),
            a.created_at.isoformat() if a.created_at else None,
            getattr(a, "updated_at", None).isoformat() if getattr(a, "updated_at", None) else None,
        ])

    widths = [6, 35, 18, 14, 12, 22, 22, 50, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/{artifact_id}/export")
def export_artifact(
    artifact_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Download a tenant artifact as a file.

    For platform-native artifacts (risk register, asset inventory, …) we
    generate an XLSX from the live tenant data so the exported file always
    reflects the current state of the source — not a cached snapshot.

    For non-platform artifacts the caller already has the ``content`` field
    and renders the file client-side; this endpoint returns 400 to make
    that explicit.
    """
    tenant_id = get_user_primary_tenant(current_user, db)
    artifact = (
        db.query(TenantArtifact)
        .filter(TenantArtifact.id == artifact_id, TenantArtifact.tenant_id == tenant_id)
        .first()
    )
    if not artifact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artifact not found")

    if not artifact.is_platform_native:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "This artifact is not a platform-native data export. The "
                "frontend renders its content directly via the existing "
                "download helper — call /artifacts/{id} for the content."
            ),
        )

    pdt = (artifact.platform_data_type or "").strip().lower()
    title = artifact.name or "artifact"

    if pdt == "risk_register":
        buf = _build_risk_register_xlsx(db, tenant_id, title)
    elif pdt == "asset_inventory":
        buf = _build_asset_inventory_xlsx(db, tenant_id, title)
    else:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail=f"Export not implemented for platform_data_type='{pdt}'",
        )

    filename = f"{_safe_filename(title)}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("")
def list_artifacts(
    assessment_id: Optional[int] = Query(None),
    framework_key: Optional[str] = Query(None),
    assessment_type: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    resolved_key = framework_key
    if not resolved_key and assessment_type:
        resolved_key = _resolve_framework_key(assessment_type)

    q = db.query(TenantArtifact).filter(TenantArtifact.tenant_id == tenant_id)
    if assessment_id is not None:
        q = q.filter(TenantArtifact.assessment_id == assessment_id)
    if resolved_key:
        q = q.filter(TenantArtifact.framework_key == resolved_key)
    if status_filter:
        q = q.filter(TenantArtifact.status == status_filter)
    artifacts = q.order_by(TenantArtifact.created_at.desc()).all()
    return [_artifact_out(a) for a in artifacts]


@router.post("")
def create_artifact(
    payload: TenantArtifactCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    user_id = current_user.id

    # Fetch platform record count if native
    platform_count = None
    if payload.is_platform_native and payload.platform_data_type:
        try:
            if payload.platform_data_type == "risk_register":
                platform_count = db.query(func.count(Risk.id)).filter(Risk.tenant_id == tenant_id).scalar()
            elif payload.platform_data_type == "asset_inventory":
                platform_count = db.query(func.count(ITAsset.id)).filter(ITAsset.tenant_id == tenant_id).scalar()
        except Exception:
            pass

    artifact = TenantArtifact(
        tenant_id=tenant_id,
        catalog_item_id=payload.catalog_item_id,
        assessment_id=payload.assessment_id,
        framework_key=payload.framework_key,
        name=payload.name,
        artifact_type=payload.artifact_type,
        stage=payload.stage,
        control_ref=payload.control_ref,
        description=payload.description,
        format=payload.format,
        content=payload.content,
        status="draft",
        assigned_to_id=payload.assigned_to_id,
        created_by_id=user_id,
        is_platform_native=payload.is_platform_native,
        platform_data_type=payload.platform_data_type,
        platform_record_count=platform_count,
    )
    db.add(artifact)
    db.commit()
    db.refresh(artifact)
    return _artifact_out(artifact)


@router.get("/{artifact_id}")
def get_artifact(
    artifact_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    artifact = db.query(TenantArtifact).filter(
        TenantArtifact.id == artifact_id,
        TenantArtifact.tenant_id == tenant_id,
    ).first()
    if not artifact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artifact not found")
    # Refresh platform count if native
    if artifact.is_platform_native and artifact.platform_data_type:
        try:
            if artifact.platform_data_type == "risk_register":
                artifact.platform_record_count = db.query(func.count(Risk.id)).filter(Risk.tenant_id == tenant_id).scalar()
            elif artifact.platform_data_type == "asset_inventory":
                artifact.platform_record_count = db.query(func.count(ITAsset.id)).filter(ITAsset.tenant_id == tenant_id).scalar()
            db.commit()
        except Exception:
            pass
    return _artifact_out(artifact)


@router.put("/{artifact_id}")
def update_artifact(
    artifact_id: int,
    payload: TenantArtifactUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    artifact = db.query(TenantArtifact).filter(
        TenantArtifact.id == artifact_id,
        TenantArtifact.tenant_id == tenant_id,
    ).first()
    if not artifact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artifact not found")

    if payload.name is not None:
        artifact.name = payload.name
    if payload.artifact_type is not None:
        artifact.artifact_type = payload.artifact_type
    if payload.stage is not None:
        artifact.stage = payload.stage
    if payload.description is not None:
        artifact.description = payload.description
    if payload.content is not None:
        artifact.content = payload.content
    if payload.status is not None:
        artifact.status = payload.status
    if payload.assigned_to_id is not None:
        artifact.assigned_to_id = payload.assigned_to_id
    if payload.format is not None:
        artifact.format = payload.format

    db.commit()
    db.refresh(artifact)
    return _artifact_out(artifact)


@router.delete("/{artifact_id}")
def delete_artifact(
    artifact_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    artifact = db.query(TenantArtifact).filter(
        TenantArtifact.id == artifact_id,
        TenantArtifact.tenant_id == tenant_id,
    ).first()
    if not artifact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artifact not found")
    db.delete(artifact)
    db.commit()
    return {"ok": True}


@router.post("/{artifact_id}/assign")
def assign_artifact(
    artifact_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    artifact = db.query(TenantArtifact).filter(
        TenantArtifact.id == artifact_id,
        TenantArtifact.tenant_id == tenant_id,
    ).first()
    if not artifact:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artifact not found")
    user = db.query(GRCUser).filter(GRCUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    artifact.assigned_to_id = user_id
    db.commit()
    db.refresh(artifact)
    return _artifact_out(artifact)
