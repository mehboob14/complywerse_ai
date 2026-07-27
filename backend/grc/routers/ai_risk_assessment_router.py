"""AI Risk Assessment template router.

Owns ingestion + CRUD for entries that come from the "AI Risk Assessment
Template.xlsx" workbook. The data model mirrors the 13 spreadsheet columns
exactly. Two AI assist endpoints generate suggested mitigation, controls,
likelihood, impact, and residual rating from the AI system + risk
description text.

Endpoints (mounted at /erm/ai-risk-assessment):
  GET    /                       list entries
  POST   /                       create entry
  GET    /{entry_id}             read one
  PUT    /{entry_id}             update
  DELETE /{entry_id}             delete
  POST   /upload                 ingest the xlsx workbook
  GET    /template               download empty template
  POST   /{entry_id}/ai-suggest  generate AI suggestions for one entry
  POST   /{entry_id}/accept-ai   accept AI suggestions (copy into primary fields)
  POST   /{entry_id}/bridge-to-risk  create a backing Risk row from the entry
"""

from __future__ import annotations

import io
import os
import json
import logging
from datetime import datetime, date
from typing import Any, Dict, List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from ..models import (
    AIRiskAssessmentEntry,
    AIRiskAssessmentEvidenceLink,
    Evidence,
    GRCUser,
    Risk,
    get_db,
)
from .auth_router import require_auth, get_user_tenants, get_user_primary_tenant

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/erm/ai-risk-assessment", tags=["AI Risk Assessment"])


# ── Lazy table self heal ─────────────────────────────────────────────────────
# The new table is auto created for fresh tenants by Base.metadata.create_all.
# Existing tenant DBs (pre this feature) need a one shot create. Run on each
# bound engine the first time we touch it, memoized so the inspect + create
# only runs once per tenant per process.
_ensured_engines: set[int] = set()


def _ensure_table(db: Session) -> None:
    engine = db.get_bind()
    eid = id(engine)
    if eid in _ensured_engines:
        return
    try:
        AIRiskAssessmentEntry.__table__.create(bind=engine, checkfirst=True)
        _ensured_engines.add(eid)
    except Exception as e:  # pragma: no cover
        logger.warning("ai_risk_assessment table self heal failed: %s", e)


# ── Schemas ───────────────────────────────────────────────────────────────────

class AIRiskEntryCreate(BaseModel):
    risk_id_external: Optional[str] = None
    ai_system_use_case: Optional[str] = None
    risk_description: Optional[str] = None
    risk_category: Optional[str] = None
    likelihood: Optional[int] = Field(default=None, ge=1, le=5)
    impact: Optional[int] = Field(default=None, ge=1, le=5)
    risk_score: Optional[int] = None
    existing_controls: Optional[str] = None
    residual_risk_level: Optional[str] = None
    mitigation_plan: Optional[str] = None
    risk_owner: Optional[str] = None
    risk_owner_user_id: Optional[int] = None
    target_review_date: Optional[date] = None
    status: Optional[str] = "Open"


class AIRiskEntryUpdate(BaseModel):
    risk_id_external: Optional[str] = None
    ai_system_use_case: Optional[str] = None
    risk_description: Optional[str] = None
    risk_category: Optional[str] = None
    likelihood: Optional[int] = Field(default=None, ge=1, le=5)
    impact: Optional[int] = Field(default=None, ge=1, le=5)
    risk_score: Optional[int] = None
    existing_controls: Optional[str] = None
    residual_risk_level: Optional[str] = None
    mitigation_plan: Optional[str] = None
    risk_owner: Optional[str] = None
    risk_owner_user_id: Optional[int] = None
    target_review_date: Optional[date] = None
    status: Optional[str] = None


class AISuggestRequest(BaseModel):
    # Optional caller hint to focus the AI prompt.
    focus: Optional[str] = None


# ── Helpers ──────────────────────────────────────────────────────────────────

def _serialize(e: AIRiskAssessmentEntry) -> Dict[str, Any]:
    return {
        "id": e.id,
        "risk_id_external": e.risk_id_external,
        "ai_system_use_case": e.ai_system_use_case,
        "risk_description": e.risk_description,
        "risk_category": e.risk_category,
        "likelihood": e.likelihood,
        "impact": e.impact,
        "risk_score": e.risk_score,
        "existing_controls": e.existing_controls,
        "residual_risk_level": e.residual_risk_level,
        "mitigation_plan": e.mitigation_plan,
        "risk_owner": e.risk_owner,
        "risk_owner_user_id": e.risk_owner_user_id,
        "target_review_date": e.target_review_date.isoformat() if e.target_review_date else None,
        "status": e.status,
        "bridged_risk_id": e.bridged_risk_id,
        "ai_suggested_mitigation": e.ai_suggested_mitigation,
        "ai_suggested_controls": e.ai_suggested_controls,
        "ai_suggested_likelihood": e.ai_suggested_likelihood,
        "ai_suggested_impact": e.ai_suggested_impact,
        "ai_suggested_residual_level": e.ai_suggested_residual_level,
        "ai_rationale": e.ai_rationale,
        "ai_generated_at": e.ai_generated_at.isoformat() if e.ai_generated_at else None,
        "ai_model": e.ai_model,
        "ai_suggestion_accepted": bool(e.ai_suggestion_accepted),
        "source": e.source,
        "source_file_name": e.source_file_name,
        "created_at": e.created_at.isoformat() if e.created_at else None,
        "updated_at": e.updated_at.isoformat() if e.updated_at else None,
    }


def _residual_from_score(score: Optional[int]) -> str:
    """Map computed risk_score (1..25) to a residual level bucket."""
    if not score:
        return "Low"
    if score >= 15:
        return "High"
    if score >= 8:
        return "Medium"
    return "Low"


def _calc_score(likelihood: Optional[int], impact: Optional[int]) -> Optional[int]:
    if likelihood is None or impact is None:
        return None
    return max(0, min(25, int(likelihood) * int(impact)))


# ── List + read ──────────────────────────────────────────────────────────────

@router.get("")
def list_entries(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    rows = (
        db.query(AIRiskAssessmentEntry)
        .filter(AIRiskAssessmentEntry.tenant_id.in_(user_tenants))
        .order_by(AIRiskAssessmentEntry.created_at.desc().nullslast())
        .all()
    )
    return [_serialize(r) for r in rows]


@router.get("/{entry_id}")
def read_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")
    return _serialize(row)


# ── Create ───────────────────────────────────────────────────────────────────

@router.post("", status_code=status.HTTP_201_CREATED)
def create_entry(
    payload: AIRiskEntryCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    tenant = get_user_primary_tenant(current_user, db)
    if tenant is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No tenant available for current user")

    data = payload.model_dump()
    # Auto compute risk score when both axes set and score is not.
    if data.get("risk_score") is None:
        data["risk_score"] = _calc_score(data.get("likelihood"), data.get("impact"))
    if not data.get("residual_risk_level"):
        data["residual_risk_level"] = _residual_from_score(data.get("risk_score"))
    if not data.get("status"):
        data["status"] = "Open"

    row = AIRiskAssessmentEntry(
        tenant_id=tenant.id,
        source="manual",
        created_by_user_id=current_user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        **data,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _serialize(row)


# ── Update ───────────────────────────────────────────────────────────────────

@router.put("/{entry_id}")
def update_entry(
    entry_id: int,
    payload: AIRiskEntryUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")

    changes = payload.model_dump(exclude_unset=True)
    for k, v in changes.items():
        setattr(row, k, v)
    # Recompute score when either axis changed.
    if "likelihood" in changes or "impact" in changes:
        row.risk_score = _calc_score(row.likelihood, row.impact)
        if not row.residual_risk_level:
            row.residual_risk_level = _residual_from_score(row.risk_score)
    row.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(row)
    return _serialize(row)


# ── Delete ───────────────────────────────────────────────────────────────────

@router.delete("/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")
    db.delete(row)
    db.commit()
    return None


# ── Template download ────────────────────────────────────────────────────────

# Column order MUST stay identical to the supplied workbook so re uploaded files
# round trip cleanly.
TEMPLATE_HEADERS = [
    "Risk ID",
    "AI System / Use Case",
    "Risk Description",
    "Risk Category",
    "Likelihood (1-5)",
    "Impact (1-5)",
    "Risk Score",
    "Existing Controls",
    "Residual Risk Level",
    "Mitigation Plan",
    "Risk Owner",
    "Target Review Date",
    "Status",
]


@router.get("/template")
def download_template(
    current_user: GRCUser = Depends(require_auth),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Title row at row 5 (matches the supplied workbook).
    ws.cell(row=5, column=1, value="AI Risk Assessment Template")
    # Header row at row 6.
    for idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=6, column=idx, value=header)
    # Column widths.
    widths = [10, 28, 50, 22, 14, 12, 12, 38, 18, 50, 28, 18, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="AI Risk Assessment Template.xlsx"'},
    )


# ── Upload + parse xlsx ──────────────────────────────────────────────────────

def _coerce_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _coerce_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.fromisoformat(str(v).strip().split(" ")[0]).date()
    except Exception:
        return None


def _coerce_str(v: Any, max_len: Optional[int] = None) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        s = s[: max_len - 1]
    return s


@router.post("/upload")
async def upload_template(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    tenant = get_user_primary_tenant(current_user, db)
    if tenant is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No tenant available for current user")

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload an .xlsx or .xls file")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot read workbook: {e}")

    ws = wb.active
    # Find the header row. The supplied workbook has the header at row 6 but
    # we tolerate any row in the first 12 rows containing "Risk ID" in col 1.
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and v.strip().lower().startswith("risk id"):
            header_row = r
            break
    if header_row is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Header row not found. Expected 'Risk ID' in column A.")

    created: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for r in range(header_row + 1, ws.max_row + 1):
        # Column ordering from the template.
        risk_id_ext = _coerce_str(ws.cell(r, 1).value, 50)
        ai_system   = _coerce_str(ws.cell(r, 2).value, 255)
        descr       = _coerce_str(ws.cell(r, 3).value)
        category    = _coerce_str(ws.cell(r, 4).value, 100)
        lk          = _coerce_int(ws.cell(r, 5).value)
        imp         = _coerce_int(ws.cell(r, 6).value)
        score       = _coerce_int(ws.cell(r, 7).value)
        controls    = _coerce_str(ws.cell(r, 8).value)
        residual    = _coerce_str(ws.cell(r, 9).value, 20)
        mitigation  = _coerce_str(ws.cell(r, 10).value)
        owner       = _coerce_str(ws.cell(r, 11).value, 255)
        review_d    = _coerce_date(ws.cell(r, 12).value)
        status_v    = _coerce_str(ws.cell(r, 13).value, 50) or "Open"

        # Skip blank rows.
        if not any([risk_id_ext, ai_system, descr, category, lk, imp, score, controls, mitigation, owner, review_d]):
            continue

        if not descr and not ai_system:
            errors.append({"row": r, "error": "Both AI System and Risk Description are blank"})
            continue

        # Sanity clamp on likelihood / impact / score.
        if lk is not None and not (1 <= lk <= 5):
            lk = max(1, min(5, lk))
        if imp is not None and not (1 <= imp <= 5):
            imp = max(1, min(5, imp))
        if score is None:
            score = _calc_score(lk, imp)
        if not residual:
            residual = _residual_from_score(score)

        row = AIRiskAssessmentEntry(
            tenant_id=tenant.id,
            risk_id_external=risk_id_ext,
            ai_system_use_case=ai_system,
            risk_description=descr,
            risk_category=category,
            likelihood=lk,
            impact=imp,
            risk_score=score,
            existing_controls=controls,
            residual_risk_level=residual,
            mitigation_plan=mitigation,
            risk_owner=owner,
            target_review_date=review_d,
            status=status_v,
            source="template_upload",
            source_file_name=file.filename,
            created_by_user_id=current_user.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.add(row)
        db.flush()
        created.append({"id": row.id, "row": r, "ai_system_use_case": ai_system})

    db.commit()

    return {
        "imported": created,
        "errors": errors,
        "summary": {
            "imported_count": len(created),
            "error_count": len(errors),
            "file_name": file.filename,
        },
    }


# ── AI suggestions ───────────────────────────────────────────────────────────

def _ai_available() -> bool:
    # Mirror the heuristic used by other AI features in the codebase.
    return bool(os.environ.get("OPENAI_API_KEY") or os.environ.get("ANTHROPIC_API_KEY"))


def _ai_template_suggestion(entry: AIRiskAssessmentEntry) -> Dict[str, Any]:
    """Deterministic fallback when no LLM key is configured. Produces sensible
    suggestions based on category + score so the feature still works."""
    cat = (entry.risk_category or "").lower()
    sys_name = entry.ai_system_use_case or "the AI system"
    descr = (entry.risk_description or "").lower()

    # Mitigation suggestions per category bucket.
    mitig_lib = {
        "ethical": f"Establish a fairness review board for {sys_name}. Run quarterly bias audits using a diversified evaluation dataset. Document acceptance thresholds per protected attribute.",
        "fairness": f"Run a bias audit on {sys_name} every quarter. Add demographic parity and equal opportunity metrics to the CI pipeline. Block deployment on regression.",
        "data privacy": f"Apply data minimization. Enable purpose limitation tags on the training set used by {sys_name}. Add a DSAR + opt out path. Run a DPIA before each material model change.",
        "regulatory": f"Engage Legal Compliance early. Tag {sys_name} with the applicable regulation (GDPR, EU AI Act, NCA). Schedule a regulatory readiness review aligned to the next supervisory cycle.",
        "operational": f"Add fallback rules around {sys_name}. Define an SLA, monitoring on accuracy and latency, and a tested rollback plan. Run a chaos test once per release.",
        "security": f"Threat model {sys_name} (STRIDE). Add prompt injection and data exfiltration detectors. Run red team exercises twice per year. Patch model dependencies on the SBOM cadence.",
        "intellectual property": f"Plagiarism + copyright detection at output time for {sys_name}. License review on training data. Watermark generated content. Track derivative claims.",
        "explainability": f"Adopt interpretable models or add SHAP / LIME on critical decisions for {sys_name}. Publish a model card. Provide users with a 'why this decision' panel.",
        "trust": f"Publish a model card for {sys_name}. Add user feedback capture. Build an exception path that routes contested decisions to a human reviewer.",
    }
    mitigation = ""
    for k, v in mitig_lib.items():
        if k in cat or k in descr:
            mitigation = v
            break
    if not mitigation:
        mitigation = (
            f"Add monitoring on accuracy + drift for {sys_name}. Define an incident response runbook. "
            f"Schedule quarterly review of model risk and refresh the risk register entry."
        )

    # Existing controls suggestion.
    controls = (
        f"Suggested controls. Model card, monitoring dashboard with drift + bias metrics, "
        f"human in the loop for high impact decisions, periodic external audit, "
        f"role based access on training data, incident response playbook."
    )

    # Likelihood / Impact bucketed from category.
    high_impact_cats = ("regulatory", "security", "data privacy", "ethical", "fairness")
    high_likelihood_cats = ("operational", "explainability", "trust")
    suggested_l = 4 if any(h in cat for h in high_likelihood_cats) else 3
    suggested_i = 5 if any(h in cat for h in high_impact_cats) else 3
    score = suggested_l * suggested_i
    residual = _residual_from_score(score)

    rationale = (
        f"Heuristic: category {entry.risk_category or 'Unknown'} maps to L={suggested_l}, I={suggested_i}. "
        f"Resulting score {score} -> {residual} residual band."
    )
    return {
        "mitigation": mitigation,
        "controls": controls,
        "likelihood": suggested_l,
        "impact": suggested_i,
        "residual_level": residual,
        "rationale": rationale,
        "model": "rules.v1",
    }


def _ai_llm_suggestion(entry: AIRiskAssessmentEntry, focus: Optional[str]) -> Optional[Dict[str, Any]]:
    """Best effort LLM call. Returns None on any failure so the template path
    can take over. Honors OPENAI_API_KEY / ANTHROPIC_API_KEY."""
    try:
        prompt = (
            "You are an AI risk officer. Given the AI risk register entry below, "
            "return JSON with keys: mitigation, controls, likelihood (1-5), impact (1-5), "
            "residual_level (High|Medium|Low), rationale.\n\n"
            f"AI System: {entry.ai_system_use_case or ''}\n"
            f"Risk Description: {entry.risk_description or ''}\n"
            f"Risk Category: {entry.risk_category or ''}\n"
            f"Existing Controls: {entry.existing_controls or ''}\n"
            f"Operator focus: {focus or 'general'}\n"
        )
        # OpenAI path.
        if os.environ.get("OPENAI_API_KEY"):
            try:
                from openai import OpenAI  # type: ignore
                client = OpenAI()
                model = os.environ.get("COMPLYVERSE_AI_MODEL", "gpt-4o-mini")
                resp = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "Respond with valid JSON only."},
                        {"role": "user", "content": prompt},
                    ],
                    response_format={"type": "json_object"},
                    temperature=0.2,
                )
                raw = resp.choices[0].message.content or "{}"
                data = json.loads(raw)
                data["model"] = model
                return data
            except Exception as e:
                logger.warning("OpenAI suggest failed: %s", e)
                return None
        # Anthropic path.
        if os.environ.get("ANTHROPIC_API_KEY"):
            try:
                import time
                import anthropic  # type: ignore
                from ..services.ai_usage import record_provider_attempt, usage_scope
                client = anthropic.Anthropic()
                model = os.environ.get("COMPLYVERSE_AI_MODEL", "claude-haiku-4-5-20251001")
                started_at = time.perf_counter()
                try:
                    with usage_scope(module_key="erm", feature_key="ai_risk_suggest"):
                        msg = client.messages.create(
                            model=model,
                            max_tokens=800,
                            system="Respond with valid JSON only.",
                            messages=[{"role": "user", "content": prompt}],
                        )
                        record_provider_attempt(
                            response=msg, requested_model=model, provider="anthropic",
                            api_family="messages", started_at=started_at,
                        )
                except Exception as exc:
                    record_provider_attempt(
                        error=exc, requested_model=model, provider="anthropic",
                        api_family="messages", started_at=started_at,
                    )
                    raise
                raw = "".join(part.text for part in msg.content if getattr(part, "type", "") == "text")
                # Strip markdown fences if any.
                raw = raw.strip()
                if raw.startswith("```"):
                    raw = raw.strip("`").lstrip("json").strip()
                data = json.loads(raw)
                data["model"] = model
                return data
            except Exception as e:
                logger.warning("Anthropic suggest failed: %s", e)
                return None
    except Exception as e:
        logger.warning("LLM suggest dispatch failed: %s", e)
    return None


@router.post("/{entry_id}/ai-suggest")
def ai_suggest(
    entry_id: int,
    payload: Optional[AISuggestRequest] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")

    focus = payload.focus if payload else None
    suggestion: Optional[Dict[str, Any]] = None
    if _ai_available():
        suggestion = _ai_llm_suggestion(row, focus)
    if not suggestion:
        suggestion = _ai_template_suggestion(row)

    # Persist on the row.
    row.ai_suggested_mitigation = suggestion.get("mitigation")
    row.ai_suggested_controls = suggestion.get("controls")
    try:
        row.ai_suggested_likelihood = int(suggestion.get("likelihood")) if suggestion.get("likelihood") is not None else None
    except (TypeError, ValueError):
        row.ai_suggested_likelihood = None
    try:
        row.ai_suggested_impact = int(suggestion.get("impact")) if suggestion.get("impact") is not None else None
    except (TypeError, ValueError):
        row.ai_suggested_impact = None
    row.ai_suggested_residual_level = suggestion.get("residual_level")
    row.ai_rationale = suggestion.get("rationale")
    row.ai_generated_at = datetime.utcnow()
    row.ai_model = suggestion.get("model")
    row.ai_suggestion_accepted = False
    db.commit()
    db.refresh(row)
    return {
        "source": "llm" if _ai_available() and suggestion.get("model", "").lower() not in ("rules.v1",) else "rules",
        "entry": _serialize(row),
    }


@router.post("/{entry_id}/accept-ai")
def accept_ai(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Copy AI suggested values into the primary fields. Does NOT overwrite
    fields the operator has already filled in (only fills blanks)."""
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")
    if not row.ai_generated_at:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No AI suggestion available. Run /ai-suggest first.")

    if not row.mitigation_plan and row.ai_suggested_mitigation:
        row.mitigation_plan = row.ai_suggested_mitigation
    if not row.existing_controls and row.ai_suggested_controls:
        row.existing_controls = row.ai_suggested_controls
    if row.likelihood is None and row.ai_suggested_likelihood is not None:
        row.likelihood = row.ai_suggested_likelihood
    if row.impact is None and row.ai_suggested_impact is not None:
        row.impact = row.ai_suggested_impact
    if not row.residual_risk_level and row.ai_suggested_residual_level:
        row.residual_risk_level = row.ai_suggested_residual_level
    # Re compute score after either axis updates.
    row.risk_score = _calc_score(row.likelihood, row.impact)
    row.ai_suggestion_accepted = True
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return _serialize(row)


# ── Bridge to general Risk ───────────────────────────────────────────────────

@router.post("/{entry_id}/bridge-to-risk")
def bridge_to_risk(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Create a Risk row from this entry and link it. Idempotent: subsequent
    calls return the existing bridged_risk_id."""
    _ensure_table(db)
    user_tenants = get_user_tenants(current_user, db)
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")
    if row.bridged_risk_id:
        return {"risk_id": row.bridged_risk_id, "created": False}

    # Map template status to Risk status.
    status_map = {
        "open": "open",
        "in progress": "mitigating",
        "closed": "closed",
        "accepted": "accepted",
    }
    risk_status = status_map.get((row.status or "Open").lower(), "open")

    # Map template category to Risk risk_category enum where possible.
    cat = (row.risk_category or "").lower()
    if "regulatory" in cat or "compliance" in cat:
        rcat = "compliance"
    elif "security" in cat:
        rcat = "technology"
    elif "operational" in cat:
        rcat = "operational"
    elif "data privacy" in cat or "intellectual" in cat:
        rcat = "operational"
    elif "ethical" in cat or "fairness" in cat or "explainability" in cat or "trust" in cat:
        rcat = "strategic"
    else:
        rcat = "operational"

    title = f"AI Risk: {row.ai_system_use_case or row.risk_description or 'Untitled'}"
    risk = Risk(
        tenant_id=row.tenant_id,
        title=title[:200],
        description=row.risk_description,
        risk_category=rcat,
        risk_sub_category=row.risk_category,
        inherent_likelihood=row.likelihood,
        inherent_impact=row.impact,
        inherent_score=row.risk_score,
        residual_likelihood=row.likelihood,
        residual_impact=row.impact,
        residual_score=row.risk_score,
        status=risk_status,
        treatment_plan=row.mitigation_plan,
        owner_id=row.risk_owner_user_id,
        source_type="ai_risk_import",
        source_reference=str(row.id),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(risk)
    db.flush()

    row.bridged_risk_id = risk.id
    row.updated_at = datetime.utcnow()
    db.commit()
    return {"risk_id": risk.id, "created": True}


# ── Tenant users (Risk Owner picker) ─────────────────────────────────────────

@router.get("/meta/tenant-users")
def list_tenant_users(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Tenant user list for the Risk Owner picker. Same shape the assets
    page uses (`/assets/tenant-users`) so the frontend can reuse the same
    rendering. Filters to users actually mapped to this tenant via the
    grc_tenant_users join.
    """
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    from ..models import TenantUser
    rows = (
        db.query(GRCUser)
        .join(TenantUser, TenantUser.user_id == GRCUser.id)
        .filter(TenantUser.tenant_id.in_(user_tenants))
        .order_by(GRCUser.display_name.asc().nullslast(), GRCUser.username.asc())
        .all()
    )
    return [
        {
            "id": u.id,
            "display_name": (u.display_name or u.username or u.email or f"User {u.id}"),
            "email": u.email,
            "username": u.username,
        }
        for u in rows
    ]


# ── Evidence link + upload ───────────────────────────────────────────────────

def _ensure_evidence_link_table(db: Session) -> None:
    """Create the AI-risk evidence link table on legacy tenant DBs that
    pre-date the model. Self-heal so we don't need an alembic step."""
    engine = db.get_bind()
    eid = id(engine)
    # Re-use the table-ensure cache for the link table too; key off the
    # table name so it stays per-engine.
    flag_attr = f"ai_risk_link_ensured_{eid}"
    if globals().get(flag_attr):
        return
    try:
        AIRiskAssessmentEvidenceLink.__table__.create(bind=engine, checkfirst=True)
        globals()[flag_attr] = True
    except Exception as e:  # pragma: no cover
        logger.warning("ai_risk_evidence_link table self heal failed: %s", e)


def _serialize_evidence(e: Evidence, link: Optional[AIRiskAssessmentEvidenceLink] = None) -> Dict[str, Any]:
    return {
        "evidence_id": e.id,
        "link_id": link.id if link else None,
        "name": e.name,
        "description": e.description,
        "file_name": e.file_name,
        "file_type": e.file_type,
        "file_path": e.file_path,
        "evidence_type": e.evidence_type,
        "uploaded_at": e.uploaded_at.isoformat() if e.uploaded_at else None,
        "uploaded_by": e.uploaded_by,
        "status": e.status,
        "relationship_type": link.relationship_type if link else "supports",
    }


def _ai_entry_or_404(db: Session, entry_id: int, user_tenants: List[int]) -> AIRiskAssessmentEntry:
    row = (
        db.query(AIRiskAssessmentEntry)
        .filter(
            AIRiskAssessmentEntry.id == entry_id,
            AIRiskAssessmentEntry.tenant_id.in_(user_tenants),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI risk entry not found")
    return row


@router.get("/{entry_id}/evidence")
def list_evidence(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """List every Evidence row linked to this AI risk entry."""
    _ensure_table(db)
    _ensure_evidence_link_table(db)
    user_tenants = get_user_tenants(current_user, db)
    _ai_entry_or_404(db, entry_id, user_tenants)
    links = (
        db.query(AIRiskAssessmentEvidenceLink, Evidence)
        .join(Evidence, Evidence.id == AIRiskAssessmentEvidenceLink.evidence_id)
        .filter(AIRiskAssessmentEvidenceLink.entry_id == entry_id)
        .order_by(Evidence.uploaded_at.desc().nullslast(), Evidence.id.desc())
        .all()
    )
    return [_serialize_evidence(ev, link) for link, ev in links]


class _LinkEvidenceIn(BaseModel):
    evidence_id: int
    relationship_type: Optional[str] = "supports"


@router.post("/{entry_id}/evidence/link", status_code=status.HTTP_201_CREATED)
def link_existing_evidence(
    entry_id: int,
    payload: _LinkEvidenceIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Link an existing Evidence row to this AI risk entry. Idempotent —
    second call with the same evidence_id returns the existing link."""
    _ensure_table(db)
    _ensure_evidence_link_table(db)
    user_tenants = get_user_tenants(current_user, db)
    entry = _ai_entry_or_404(db, entry_id, user_tenants)
    ev = db.query(Evidence).filter(
        Evidence.id == payload.evidence_id,
        Evidence.tenant_id == entry.tenant_id,
    ).first()
    if not ev:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence not found in this tenant")
    existing = db.query(AIRiskAssessmentEvidenceLink).filter(
        AIRiskAssessmentEvidenceLink.entry_id == entry_id,
        AIRiskAssessmentEvidenceLink.evidence_id == payload.evidence_id,
    ).first()
    if existing:
        return _serialize_evidence(ev, existing)
    link = AIRiskAssessmentEvidenceLink(
        entry_id=entry_id,
        evidence_id=payload.evidence_id,
        relationship_type=(payload.relationship_type or "supports"),
        created_at=datetime.utcnow(),
    )
    db.add(link)
    db.commit()
    db.refresh(link)
    return _serialize_evidence(ev, link)


@router.post("/{entry_id}/evidence/upload", status_code=status.HTTP_201_CREATED)
async def upload_and_link_evidence(
    entry_id: int,
    file: UploadFile = File(...),
    evidence_type: Optional[str] = None,
    relationship_type: Optional[str] = "supports",
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Upload a new file as Evidence and link it to the entry in one call.

    File is stored in the platform's evidence directory (same layout the
    generic Evidence uploader uses) so download / OCR / approval workflow
    can re-use it without any special-case handling.
    """
    _ensure_table(db)
    _ensure_evidence_link_table(db)
    user_tenants = get_user_tenants(current_user, db)
    entry = _ai_entry_or_404(db, entry_id, user_tenants)

    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No file provided")
    raw = await file.read()
    # Persist under EVIDENCE_DIR — fall back to ./evidence/ai-risk/<entry>/.
    base_dir = os.environ.get("EVIDENCE_DIR") or os.path.join(os.getcwd(), "evidence")
    target_dir = os.path.join(base_dir, "ai-risk", str(entry_id))
    os.makedirs(target_dir, exist_ok=True)
    # Defensive filename — strip path components, prefix with timestamp so
    # repeated uploads of the same name don't clobber.
    safe_name = os.path.basename(file.filename).replace("/", "_").replace("\\", "_")
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    on_disk = f"{ts}__{safe_name}"
    full_path = os.path.join(target_dir, on_disk)
    with open(full_path, "wb") as f:
        f.write(raw)

    ev = Evidence(
        tenant_id=entry.tenant_id,
        name=safe_name,
        description=f"Uploaded for AI Risk entry #{entry_id} ({entry.ai_system_use_case or entry.risk_description or 'untitled'})",
        file_path=full_path,
        file_name=safe_name,
        file_type=file.content_type or os.path.splitext(safe_name)[1].lstrip("."),
        evidence_type=evidence_type or "ai_risk_supporting",
        status="approved",  # auto-approve for AI risk assessment use
        uploaded_by=current_user.id,
        uploaded_at=datetime.utcnow(),
    )
    db.add(ev)
    db.flush()

    link = AIRiskAssessmentEvidenceLink(
        entry_id=entry_id,
        evidence_id=ev.id,
        relationship_type=relationship_type or "supports",
        created_at=datetime.utcnow(),
    )
    db.add(link)
    db.commit()
    db.refresh(ev)
    db.refresh(link)
    return _serialize_evidence(ev, link)


@router.delete("/{entry_id}/evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_evidence(
    entry_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Remove the link between an entry and an evidence row. Does NOT
    delete the underlying Evidence — it stays in the library so other
    risks / controls that reference it keep working."""
    _ensure_table(db)
    _ensure_evidence_link_table(db)
    user_tenants = get_user_tenants(current_user, db)
    _ai_entry_or_404(db, entry_id, user_tenants)
    link = db.query(AIRiskAssessmentEvidenceLink).filter(
        AIRiskAssessmentEvidenceLink.id == link_id,
        AIRiskAssessmentEvidenceLink.entry_id == entry_id,
    ).first()
    if not link:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence link not found")
    db.delete(link)
    db.commit()
    return None


@router.get("/{entry_id}/evidence/{link_id}/download")
def download_evidence(
    entry_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Stream the underlying file for an entry's evidence link."""
    _ensure_table(db)
    _ensure_evidence_link_table(db)
    user_tenants = get_user_tenants(current_user, db)
    _ai_entry_or_404(db, entry_id, user_tenants)
    row = db.query(AIRiskAssessmentEvidenceLink, Evidence).join(
        Evidence, Evidence.id == AIRiskAssessmentEvidenceLink.evidence_id,
    ).filter(
        AIRiskAssessmentEvidenceLink.id == link_id,
        AIRiskAssessmentEvidenceLink.entry_id == entry_id,
    ).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evidence link not found")
    _, ev = row
    if not ev.file_path or not os.path.exists(ev.file_path):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing on disk")
    with open(ev.file_path, "rb") as f:
        data = f.read()
    headers = {
        "Content-Disposition": f'attachment; filename="{ev.file_name or "evidence.bin"}"'
    }
    return StreamingResponse(
        io.BytesIO(data),
        media_type=ev.file_type or "application/octet-stream",
        headers=headers,
    )
