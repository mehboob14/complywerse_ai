from ....config import get_openai_api_key, get_openai_model

import os
import json
import uuid
import logging
import threading
from typing import List, Optional, Dict, Any
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, UploadFile, File, Form
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func, distinct
from pydantic import BaseModel
from openai import OpenAI

from ....job_status import set_status as set_job_status, get_status as get_job_status, update_status as update_job_status

_jobs_logger = logging.getLogger(__name__)

from ....models import (
    CommonControlGroup, CommonControlGroupMapping, NormalizedControl,
    NormalizedControlLink, Evidence, EvidenceControlMapping, AIEvidenceRecommendation,
    FrameworkControl, FrameworkDomain, ControlObjective, Framework,
    GRCUser, get_db, ParsedFrameworkControl, UploadedFramework, NormalizationRun
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/groups", tags=["Control Library - Groups"])

AI_INTEGRATIONS_OPENAI_API_KEY = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY")
AI_INTEGRATIONS_OPENAI_BASE_URL = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")


class CommonControlGroupCreate(BaseModel):
    code: str
    name: str
    description: Optional[str] = None
    category: Optional[str] = None
    domain: Optional[str] = None
    keywords: Optional[List[str]] = None
    evidence_types: Optional[List[str]] = None


class CommonControlGroupUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    domain: Optional[str] = None
    keywords: Optional[List[str]] = None
    evidence_types: Optional[List[str]] = None
    ai_summary: Optional[str] = None


class GroupMappingCreate(BaseModel):
    normalized_control_ids: Optional[List[int]] = []
    framework_control_ids: Optional[List[int]] = []
    parsed_control_ids: Optional[List[int]] = []


class AutoGroupRequest(BaseModel):
    framework_ids: Optional[List[int]] = None


class GenerateSummaryRequest(BaseModel):
    regenerate_keywords: bool = True


def check_ai_available() -> bool:
    """Check if OpenAI API is configured (Replit AI Integrations or direct API key)."""
    ai_integration_key = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY")
    if ai_integration_key:
        return True
    api_key = os.environ.get("OPENAI_API_KEY")
    if api_key and not api_key.startswith("your-") and len(api_key) >= 20:
        return True
    return False


def raise_ai_unavailable(fallback_available: bool = False):
    """Raise HTTP 503 error when AI features are unavailable."""
    raise HTTPException(
        status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
        detail={
            "error": "AI features unavailable",
            "message": "AI integration is not configured. The platform uses Replit AI Integrations or you can add OPENAI_API_KEY to enable AI features.",
            "fallback_available": fallback_available
        }
    )


def get_openai_client() -> OpenAI:
    if not check_ai_available():
        raise_ai_unavailable(fallback_available=False)
    api_key = get_openai_api_key()
    # Coerce an empty-string base URL to None so the SDK uses its default
    # (https://api.openai.com/v1). A literal "" makes the client connect to an
    # empty URL and fail with APIConnectionError.
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL") or None
    # A per-request timeout is ESSENTIAL: without it a single stalled request
    # blocks the (solo-pool) worker indefinitely, freezing auto-group /
    # normalization mid-run (the classic "stuck at 85%"). max_retries lets the
    # SDK transparently retry transient errors / rate limits with backoff.
    return OpenAI(
        api_key=api_key,
        base_url=base_url,
        timeout=90.0,
        max_retries=2,
    )


def generate_keywords_for_group(name: str, description: str) -> List[str]:
    try:
        client = get_openai_client()
        prompt = f"""Extract 5-10 key compliance/security terms from this control group:

Name: {name}
Description: {description or 'No description provided'}

Return JSON: {{"keywords": ["term1", "term2", ...]}}"""

        response = client.chat.completions.create(
            model=get_openai_model(),
            messages=[
                {"role": "system", "content": "Extract compliance keywords. Respond only with valid JSON."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=500,
            temperature=0.3
        )
        result = json.loads(response.choices[0].message.content or '{"keywords": []}')
        return result.get("keywords", [])
    except Exception:
        return []


def generate_group_summary(controls_text: str) -> dict:
    try:
        client = get_openai_client()
        prompt = f"""Analyze these related compliance controls and generate:
1. A summary describing their common purpose
2. Key terms/keywords that characterize these controls

Controls:
{controls_text[:4000]}

Return JSON:
{{
    "summary": "<2-3 sentence summary of the control group's purpose>",
    "keywords": ["keyword1", "keyword2", ...],
    "evidence_types": ["<suggested evidence types>"]
}}"""

        response = client.chat.completions.create(
            model=get_openai_model(),
            messages=[
                {"role": "system", "content": "You are a compliance expert summarizing control groups. Respond only with valid JSON."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=1000,
            temperature=0.3
        )
        return json.loads(response.choices[0].message.content or '{}')
    except Exception as e:
        error_msg = str(e)
        if "FREE_CLOUD_BUDGET_EXCEEDED" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Cloud budget exceeded. Please upgrade your plan."
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI summary generation failed: {error_msg}"
        )


# Tunables for AI grouping. Single-shot calls work well up to ~80 controls; for
# anything larger we batch + merge so the model never runs out of output tokens
# and silently drops controls (the original bug — `[:100]` slice + max_tokens=4000
# meant ~70% of selected frameworks' controls never made it into a group).
# Batch size tuned for the FULL-text payload (up to ~8KB per control,
# previously 400 chars). Multi-turn pipeline already merges across
# batches so reducing this from 60 just adds more turns — quality
# unchanged, no controls dropped.  Math: 30 controls * 8000 chars / 4
# tokens-per-char ≈ 60k tokens of input, well under gpt-4o's 128k cap.
_AI_GROUP_BATCH_SIZE = 30
_AI_GROUP_MAX_TOKENS = 8000


import re as _re_mod  # local alias to avoid clashing with module-level re uses


def _canonicalize_group_name(name: str) -> str:
    """Strict canonical form for cross-batch group merging.
    'Access Control', 'access-control', 'Access  Control!' → 'accesscontrol'.
    Without this, 'Identity & Access Management' and 'Identity and Access Management'
    stay separate after merging — which is why each framework's controls were
    landing in their own siloed group.
    """
    if not name:
        return ""
    s = name.lower().replace("&", "and")
    return _re_mod.sub(r"[^a-z0-9]+", "", s)


def _interleave_by_framework(controls_list: List[dict]) -> List[dict]:
    """Round-robin interleave controls so each batch sees representation from
    every selected framework. Without this, batches were sequential by
    framework (60 from framework A → 60 from framework B → …) and the AI
    produced framework-siloed groups since each batch only saw one source.
    """
    by_fw: Dict[str, List[dict]] = {}
    fw_order: List[str] = []
    for c in controls_list:
        fw = c.get("framework") or "Unknown"
        if fw not in by_fw:
            by_fw[fw] = []
            fw_order.append(fw)
        by_fw[fw].append(c)
    interleaved: List[dict] = []
    cursors = {fw: 0 for fw in fw_order}
    while True:
        progressed = False
        for fw in fw_order:
            idx = cursors[fw]
            if idx < len(by_fw[fw]):
                interleaved.append(by_fw[fw][idx])
                cursors[fw] = idx + 1
                progressed = True
        if not progressed:
            break
    return interleaved


def _build_grouping_prompt(
    controls_text: str,
    target_groups_hint: str,
    framework_summary: str,
    existing_themes: Optional[List[str]] = None,
) -> str:
    existing_block = ""
    if existing_themes:
        bullets = "\n".join(f"  - {t}" for t in existing_themes[:25])
        existing_block = (
            "\nEXISTING GROUP THEMES from earlier batches (REUSE these names verbatim "
            "when they fit — do not invent a near-duplicate):\n"
            f"{bullets}\n"
        )
    return (
        "Analyze these compliance controls and group them by common cybersecurity "
        "purpose/theme. Create CROSS-FRAMEWORK groups: a single group should pull "
        "in semantically equivalent controls from every framework that has them.\n\n"
        f"Frameworks present in this set: {framework_summary}\n"
        "STRICT RULES:\n"
        "1. EVERY control listed below MUST appear in exactly one group. Do not skip any.\n"
        "2. Use the integer ID and exact type tag (normalized | parsed) when emitting control_ids.\n"
        "3. PREFER cross-framework groups. If two frameworks both cover 'access control', "
        "their controls go in the SAME group — never split into two groups by framework.\n"
        "4. Pick stable, generic group names (e.g. 'Access Control', 'Logging & Monitoring', "
        "'Vulnerability Management') so the same name is used across batches.\n"
        f"5. {target_groups_hint}\n"
        f"{existing_block}\n"
        f"Controls:\n{controls_text}\n\n"
        "Return JSON with groups:\n"
        "{\n"
        '  "groups": [\n'
        "    {\n"
        '      "code": "<short code like CCG-001>",\n'
        '      "name": "<stable, generic group name>",\n'
        '      "description": "<group description>",\n'
        '      "category": "<category like Access Control, Data Protection, etc>",\n'
        '      "domain": "<domain like Security, Privacy, etc>",\n'
        '      "keywords": ["keyword1", "keyword2"],\n'
        '      "control_ids": [ {"id": <control_id>, "type": "normalized|parsed"} ]\n'
        "    }\n"
        "  ]\n"
        "}\n"
    )


def _format_controls_block(controls_list: List[dict]) -> str:
    return "\n\n".join([
        f"Control {i+1} (ID: {c['id']}, Type: {c['type']}, Framework: {c.get('framework', 'N/A')}):\n"
        f"Code: {c['code']}\nName: {c['name']}\nStatement: {(c.get('statement') or '')[:400]}"
        for i, c in enumerate(controls_list)
    ])


def _summarize_frameworks(controls_list: List[dict]) -> str:
    counts: Dict[str, int] = {}
    for c in controls_list:
        fw = c.get("framework") or "Unknown"
        counts[fw] = counts.get(fw, 0) + 1
    return ", ".join(f"{fw} ({n})" for fw, n in sorted(counts.items()))


def _ai_group_single_batch(
    client,
    controls_list: List[dict],
    target_groups_hint: str,
    existing_themes: Optional[List[str]] = None,
) -> List[dict]:
    controls_text = _format_controls_block(controls_list)
    framework_summary = _summarize_frameworks(controls_list)
    prompt = _build_grouping_prompt(controls_text, target_groups_hint, framework_summary, existing_themes)
    response = client.chat.completions.create(
        model=get_openai_model(),
        messages=[
            {"role": "system", "content": (
                "You are a compliance expert grouping related controls across multiple "
                "regulatory frameworks. Every control provided MUST appear in exactly one "
                "group. PREFER cross-framework groups — semantically equivalent controls "
                "from different frameworks belong in the SAME group. Respond only with "
                "valid JSON."
            )},
            {"role": "user", "content": prompt},
        ],
        response_format={"type": "json_object"},
        max_tokens=_AI_GROUP_MAX_TOKENS,
        temperature=0.2,
    )
    result = json.loads(response.choices[0].message.content or '{"groups": []}')
    return result.get("groups", [])


def _merge_ai_groups(group_batches: List[List[dict]]) -> List[dict]:
    """Merge groups produced from separate batches by canonicalized name.
    Uses a strict alphanumeric-only canonical key so 'Access Control',
    'access-control', and 'Access & Control' all collapse to one group.
    """
    merged: Dict[str, dict] = {}
    for batch in group_batches:
        for g in batch:
            name = (g.get("name") or "").strip()
            if not name:
                continue
            key = _canonicalize_group_name(name)
            if not key:
                continue
            if key not in merged:
                merged[key] = {
                    "code": g.get("code"),
                    "name": name,
                    "description": g.get("description"),
                    "category": g.get("category"),
                    "domain": g.get("domain"),
                    "keywords": list(g.get("keywords") or []),
                    "control_ids": [],
                }
            existing = merged[key]
            seen = {(c.get("id"), c.get("type")) for c in existing["control_ids"]}
            for c in g.get("control_ids") or []:
                tup = (c.get("id"), c.get("type"))
                if tup not in seen and c.get("id") is not None:
                    existing["control_ids"].append(c)
                    seen.add(tup)
            for kw in g.get("keywords") or []:
                if kw not in existing["keywords"]:
                    existing["keywords"].append(kw)
            for fld in ("description", "category", "domain"):
                if not existing.get(fld) and g.get(fld):
                    existing[fld] = g[fld]
    return list(merged.values())


class AutoGroupCancelled(Exception):
    """Raised from a progress callback when the operator cancels the job mid-run."""


def ai_auto_group_controls(controls_list: List[dict], progress_cb=None) -> List[dict]:
    """Group an arbitrarily large list of controls via OpenAI.

    Honors every control passed in (no silent truncation). Interleaves
    controls round-robin by framework BEFORE batching so each batch sees
    cross-framework variety; passes existing themes from earlier batches
    into later ones so the model reuses names verbatim; merges by canonical
    name so identical themes consolidate.

    ``progress_cb(done_batches, total_batches)`` — optional. Called after each
    batch so the caller can report granular progress AND abort: raising
    ``AutoGroupCancelled`` from it stops the run cleanly between batches.
    """
    if not controls_list:
        return []
    try:
        client = get_openai_client()

        total = len(controls_list)
        if total <= _AI_GROUP_BATCH_SIZE:
            target_hint = (
                f"Produce as many groups as the {total} controls naturally cluster into "
                "(typically 4-15 groups; aim for ~8-15 controls per group). "
                "Each group should mix controls from every framework that has matching content."
            )
            result = _ai_group_single_batch(client, controls_list, target_hint)
            if progress_cb:
                progress_cb(1, 1)
            return result

        # Round-robin interleave by framework so each batch contains controls
        # from every selected framework (the original bug was sequential
        # batching → each batch only saw 1-2 frameworks → siloed groups).
        controls_list = _interleave_by_framework(controls_list)

        # Batch the (now interleaved) controls and merge results.
        batches: List[List[dict]] = [
            controls_list[i:i + _AI_GROUP_BATCH_SIZE]
            for i in range(0, total, _AI_GROUP_BATCH_SIZE)
        ]
        target_hint = (
            "Produce groups that match common compliance themes (NCA / ISO / NIST domains). "
            "Aim for 4-10 groups per batch with ~6-15 controls each, but cover ALL controls. "
            "Within a single group, mix controls from EVERY framework whose controls match the theme."
        )
        batch_results: List[List[dict]] = []
        existing_theme_names: List[str] = []
        seen_theme_keys: set = set()
        for _bi, batch in enumerate(batches):
            # Pass canonical theme names from earlier batches so the model
            # reuses them verbatim instead of inventing near-duplicates that
            # the merger then has to chase down.
            groups = _ai_group_single_batch(
                client, batch, target_hint, existing_themes=existing_theme_names or None,
            )
            batch_results.append(groups)
            for g in groups:
                nm = (g.get("name") or "").strip()
                key = _canonicalize_group_name(nm)
                if nm and key and key not in seen_theme_keys:
                    existing_theme_names.append(nm)
                    seen_theme_keys.add(key)
            # Report progress + honour a cancel request between batches. Raising
            # AutoGroupCancelled from the callback breaks out cleanly here.
            if progress_cb:
                progress_cb(_bi + 1, len(batches))
        return _merge_ai_groups(batch_results)
    except AutoGroupCancelled:
        raise
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        if "FREE_CLOUD_BUDGET_EXCEEDED" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Cloud budget exceeded. Please upgrade your plan."
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI auto-grouping failed: {error_msg}"
        )


# ─── Reusable helpers for auto-grouping (used by both the sync route and the
#     Celery worker task) ────────────────────────────────────────────────────

def _fetch_controls_for_grouping(
    db: Session, tenant_id: int, framework_ids: Optional[List[int]] = None
) -> List[dict]:
    """Return the canonical control payload list that feeds AI auto-grouping.
    Honours `framework_ids` filter — when set, only that subset is included
    (and normalized controls are excluded so they don't pollute the run).
    """
    controls_list: List[dict] = []
    framework_ids = list(framework_ids or [])
    framework_filter_active = bool(framework_ids)

    if not framework_filter_active:
        for nc in db.query(NormalizedControl).all():
            controls_list.append({
                "id": nc.id,
                "type": "normalized",
                "code": nc.code,
                "name": nc.name,
                "statement": nc.statement or nc.objective or "",
                "framework": "Normalized",
            })

    parsed_query = db.query(ParsedFrameworkControl).join(
        UploadedFramework, ParsedFrameworkControl.uploaded_framework_id == UploadedFramework.id
    ).filter(
        or_(UploadedFramework.tenant_id == tenant_id, UploadedFramework.tenant_id.is_(None))
    )
    if framework_filter_active:
        parsed_query = parsed_query.filter(
            ParsedFrameworkControl.uploaded_framework_id.in_(framework_ids)
        )
    parsed_controls = parsed_query.all()

    fw_id_to_name: Dict[int, str] = {}
    if parsed_controls:
        framework_ids_seen = list({pc.uploaded_framework_id for pc in parsed_controls})
        for fw in db.query(UploadedFramework).filter(UploadedFramework.id.in_(framework_ids_seen)).all():
            fw_id_to_name[fw.id] = fw.name

    for pc in parsed_controls:
        # Use the RICHEST available text for the AI grouping prompt:
        # prefer full_text (the raw control body extracted from the PDF
        # — typically several paragraphs of rationale, requirements,
        # and references). Fall back to description (one paragraph) and
        # finally an empty string. The previous truncation at 400 chars
        # stripped most of the semantic content the model needed to
        # group accurately.
        #
        # We DO clamp to a generous per-control ceiling (8000 chars,
        # roughly 2000 tokens) so a single pathological control can't
        # blow the batch budget. With the batch size below this still
        # leaves comfortable headroom under gpt-4o's 128k context.
        body = (pc.full_text or pc.description or "").strip()
        if len(body) > 8000:
            body = body[:8000] + "\n...[truncated for prompt size]"
        controls_list.append({
            "id": pc.id,
            "type": "parsed",
            "code": pc.original_reference or pc.control_id,
            "name": pc.title,
            "statement": body,
            "framework": fw_id_to_name.get(pc.uploaded_framework_id, "Unknown"),
        })
    return controls_list


def _coerce_int(val: Any) -> Optional[int]:
    """Some LLM responses return numeric ids as strings. Normalise so the
    DB filter doesn't silently miss the row."""
    if val is None:
        return None
    if isinstance(val, int):
        return val
    try:
        return int(str(val).strip())
    except (TypeError, ValueError):
        return None


@router.get("/sessions")
def list_normalization_sessions(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """List normalization SESSIONS (runs) — the owner's baseline plus each custom
    framework-selected run — newest first, so the UI can switch between them."""
    tenant_id = get_user_primary_tenant(current_user, db)
    runs = (db.query(NormalizationRun)
            .filter(NormalizationRun.tenant_id == tenant_id)
            .order_by(NormalizationRun.is_baseline.desc(), NormalizationRun.id.desc())
            .all())
    out = []
    for r in runs:
        nc = db.query(NormalizedControl).filter(
            NormalizedControl.source == "ai_normalized", NormalizedControl.run_id == r.id).count()
        gn = db.query(CommonControlGroup).filter(CommonControlGroup.run_id == r.id).count()
        out.append({
            "id": r.id, "label": r.label, "scope": r.scope,
            "framework_ids": r.framework_ids, "status": r.status,
            "is_baseline": r.is_baseline,
            "started_at": r.started_at.isoformat() if r.started_at else None,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
            "domains": gn, "unified_controls": nc, "summary": r.summary,
        })
    return {"sessions": out}


@router.delete("/sessions/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_normalization_session(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Delete a custom (framework-scoped) normalization SESSION and all of its
    groups/controls. The master baseline can NEVER be deleted here — only the
    disposable scoped sessions a user builds from 'Build Unified View'."""
    tenant_id = get_user_primary_tenant(current_user, db)
    run = db.query(NormalizationRun).filter(
        NormalizationRun.id == run_id, NormalizationRun.tenant_id == tenant_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Session not found")
    if run.is_baseline:
        raise HTTPException(status_code=400, detail="The master baseline cannot be deleted.")
    from sqlalchemy import text as _text
    for stmt in (
        "DELETE FROM grc_common_control_group_mappings WHERE group_id IN (SELECT id FROM grc_common_control_groups WHERE run_id=:r)",
        "DELETE FROM grc_common_control_groups WHERE run_id=:r",
        "DELETE FROM grc_evidence_control_mappings WHERE normalized_control_id IN (SELECT id FROM grc_normalized_controls WHERE run_id=:r)",
        "DELETE FROM grc_ai_evidence_recommendations WHERE normalized_control_id IN (SELECT id FROM grc_normalized_controls WHERE run_id=:r)",
        "DELETE FROM grc_normalized_control_links WHERE normalized_control_id IN (SELECT id FROM grc_normalized_controls WHERE run_id=:r)",
        "DELETE FROM grc_normalized_controls WHERE run_id=:r",
        "DELETE FROM grc_normalization_runs WHERE id=:r",
    ):
        try:
            db.execute(_text(stmt), {"r": run_id})
        except Exception:
            _jobs_logger.exception("delete session %s: stmt failed", run_id)
    db.commit()
    return None


# ── Master-list human review — drive the unified library toward 100% correct ──
class RemoveMemberRequest(BaseModel):
    parsed_control_id: int


@router.get("/review/queue")
def review_queue(
    review_status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Unified controls of the master baseline with their framework members, so an
    admin can APPROVE correct ones or REMOVE a wrong member. Pending first."""
    tenant_id = get_user_primary_tenant(current_user, db)
    from ..services.scoped_session import get_baseline_run
    base = get_baseline_run(db, tenant_id)
    if not base:
        return {"items": [], "total": 0, "counts": {}}
    q = db.query(NormalizedControl).filter(NormalizedControl.run_id == base.id)
    if review_status:
        q = q.filter(NormalizedControl.review_status == review_status)
    total = q.count()
    counts: Dict[str, int] = {}
    for st, cnt in (db.query(NormalizedControl.review_status, func.count())
                    .filter(NormalizedControl.run_id == base.id)
                    .group_by(NormalizedControl.review_status).all()):
        counts[st or "pending"] = cnt
    ncs = (q.order_by(func.coalesce(NormalizedControl.review_status, "pending") == "approved",
                      NormalizedControl.code)
           .offset(skip).limit(limit).all())
    fw_name = {f.id: f.name for f in db.query(UploadedFramework).all()}
    items = []
    for nc in ncs:
        members = []
        for ln in db.query(NormalizedControlLink).filter(
                NormalizedControlLink.normalized_control_id == nc.id).all():
            p = db.query(ParsedFrameworkControl).filter(
                ParsedFrameworkControl.id == ln.parsed_control_id).first()
            if p:
                members.append({
                    "parsed_control_id": p.id,
                    "framework": fw_name.get(p.uploaded_framework_id, ""),
                    "code": p.original_reference or p.control_id or "",
                    "title": p.title or "",
                    "text": (p.description or p.full_text or "")[:240],
                })
        items.append({
            "id": nc.id, "code": nc.code, "name": nc.name,
            "review_status": nc.review_status or "pending",
            "framework_count": len({m["framework"] for m in members}),
            "member_count": len(members), "members": members,
        })
    return {"items": items, "total": total, "counts": counts}


_RICH_CACHE: Dict[str, Any] = {}


def _load_stage4_domains():
    """Load + cache the locked stage4 unified-library dataset (per-domain rich view)."""
    if "domains" not in _RICH_CACHE:
        path = os.path.join(os.path.dirname(__file__), "..", "..", "..",
                            "seed_data", "stage4_all_domains.json")
        with open(path, encoding="utf-8") as f:
            _RICH_CACHE["domains"] = json.load(f)["domains"]
    return _RICH_CACHE["domains"]


def _framework_domain_map():
    """framework -> sorted list of domains it appears in (across the whole library).
    Lets the UI show that a framework 'absent' from one domain is still covered
    elsewhere (it just has no controls of this domain's control-type)."""
    if "fw_doms" not in _RICH_CACHE:
        m: Dict[str, set] = {}
        for d in _load_stage4_domains():
            for f in d.get("frameworks", []):
                m.setdefault(f, set()).add(d["domain"])
        _RICH_CACHE["fw_doms"] = {k: sorted(v) for k, v in m.items()}
    return _RICH_CACHE["fw_doms"]


@router.get("/framework-templates")
def framework_templates(db: Session = Depends(get_db),
                        current_user: GRCUser = Depends(require_auth)):
    """Framework-level artifact catalogs, aggregated ONCE per framework across the
    whole library. (In stage4 these are broadcast identically into every domain;
    here we de-duplicate so each framework's generic, org-wide document templates
    are listed a single time.) These are framework-level — NOT control-level or
    normalized."""
    by_fw: Dict[str, dict] = {}
    for d in _load_stage4_domains():
        for c in d.get("framework_catalog_artifacts", []):
            fw = c.get("framework", "")
            if not fw:
                continue
            entry = by_fw.setdefault(fw, {"framework": fw, "note": c.get("note", ""), "seen": set(), "artifacts": []})
            for a in c.get("artifacts", []):
                key = (a.get("name", ""), a.get("type", ""))
                if key not in entry["seen"]:
                    entry["seen"].add(key)
                    entry["artifacts"].append({"name": a.get("name", ""), "type": a.get("type", "")})
    out = [{"framework": e["framework"], "note": e["note"], "artifacts": e["artifacts"]}
           for e in by_fw.values()]
    out.sort(key=lambda x: -len(x["artifacts"]))
    return {
        "frameworks": out,
        "total_frameworks": len(out),
        "total_artifacts": sum(len(x["artifacts"]) for x in out),
    }


@router.get("/{group_id}/rich")
def group_rich(group_id: int, db: Session = Depends(get_db),
               current_user: GRCUser = Depends(require_auth)):
    """Rich per-domain view for the seeded unified library: cross-framework sets
    with members (one control per framework, original titles preserved),
    normalized evidence (+ what each absorbs), excluded/off-topic evidence with
    reasons, requirement-specific artifacts, framework-level catalog artifacts,
    and absent-framework reasons. Keyed by the group's domain from stage4."""
    group = db.query(CommonControlGroup).filter(CommonControlGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    dom_name = (group.domain or group.name or "").strip()
    domains = _load_stage4_domains()
    dm = next((d for d in domains if d.get("domain") == dom_name), None)
    if dm is None and dom_name:
        dm = next((d for d in domains if d.get("domain", "").startswith(dom_name[:14])), None)
    if dm is None:
        raise HTTPException(status_code=404, detail=f"No rich data for domain '{dom_name}'")
    multi = [s for s in dm.get("sets", []) if s.get("member_count", 0) > 1]
    single = [s for s in dm.get("sets", []) if s.get("member_count", 0) == 1]
    # attach the DB NormalizedControl id to each set/standalone (by heading within
    # this group) so the UI can upload + link evidence to the actual control rows.
    import re as _re
    _nk = lambda s: _re.sub(r"\s+", " ", _re.sub(r"[^a-z0-9 ]", " ", (s or "").lower())).strip()
    name2id: Dict[str, int] = {}
    for nc in db.query(NormalizedControl.id, NormalizedControl.name).filter(
            NormalizedControl.common_group_id == group.id).all():
        name2id.setdefault(_nk(nc.name), nc.id)
    def _attach(lst):
        return [{**s, "nc_id": name2id.get(_nk(s.get("normalized_title")))} for s in lst]
    multi, single = _attach(multi), _attach(single)
    fw_doms = _framework_domain_map()
    # enrich each absent framework with where it IS covered (other domains)
    absent = []
    for a in dm.get("absent_frameworks", []):
        present_in = fw_doms.get(a.get("name"), [])
        absent.append({**a, "present_in": present_in, "present_in_count": len(present_in)})
    return {
        "domain": dm.get("domain"),
        "controls_in": dm.get("controls_in"),
        "frameworks": dm.get("frameworks", []),
        "framework_count": len(dm.get("frameworks", [])),
        "absent_frameworks": absent,
        "framework_catalog_artifacts": dm.get("framework_catalog_artifacts", []),
        "normalized_sets": len(multi),
        "standalone": len(single),
        "sets": multi,
        "standalone_controls": single,
    }


_EVIDENCE_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..",
                             "uploads", "evidence")


@router.get("/normalized/{nc_id}/evidence")
def list_set_evidence(nc_id: int, db: Session = Depends(get_db),
                      current_user: GRCUser = Depends(require_auth)):
    """Evidence uploaded against a normalized set (the linked Evidence records)."""
    ev_ids = [r[0] for r in db.query(EvidenceControlMapping.evidence_id).filter(
        EvidenceControlMapping.normalized_control_id == nc_id).distinct().all()]
    items = []
    for e in (db.query(Evidence).filter(Evidence.id.in_(ev_ids))
              .order_by(Evidence.uploaded_at.desc()).all() if ev_ids else []):
        linked = db.query(EvidenceControlMapping).filter(
            EvidenceControlMapping.evidence_id == e.id,
            EvidenceControlMapping.parsed_control_id.isnot(None)).count()
        items.append({"id": e.id, "name": e.name, "file_name": e.file_name,
                      "evidence_type": e.evidence_type, "status": e.status,
                      "uploaded_at": e.uploaded_at.isoformat() if e.uploaded_at else None,
                      "linked_controls": linked})
    return {"items": items, "count": len(items)}


@router.post("/normalized/{nc_id}/evidence")
async def upload_set_evidence(nc_id: int, name: Optional[str] = Form(None),
                              evidence_type: Optional[str] = Form(None),
                              file: UploadFile = File(...),
                              db: Session = Depends(get_db),
                              current_user: GRCUser = Depends(require_auth)):
    """Upload one evidence file for a normalized set. Stores it in the evidence
    library and auto-links it to the normalized control AND every member framework
    control — so one upload satisfies all frameworks and shows on each requirement."""
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant")
    nc = db.query(NormalizedControl).filter(NormalizedControl.id == nc_id).first()
    if not nc:
        raise HTTPException(status_code=404, detail="Normalized control not found")
    ext = os.path.splitext(file.filename or "")[1].lower()
    tdir = os.path.join(_EVIDENCE_DIR, str(tenant_id))
    os.makedirs(tdir, exist_ok=True)
    path = os.path.join(tdir, f"{uuid.uuid4()}{ext}")
    contents = await file.read()
    with open(path, "wb") as fh:
        fh.write(contents)
    ev = Evidence(tenant_id=tenant_id, name=(name or file.filename or "Evidence"),
                  description=f"Uploaded for unified control: {nc.name}", file_path=path,
                  file_name=file.filename, file_type=file.content_type,
                  evidence_type=evidence_type, uploaded_by=current_user.id,
                  status="approved", ocr_status="not_applicable")
    db.add(ev)
    db.flush()
    # link to the normalized (unified) control itself
    db.add(EvidenceControlMapping(evidence_id=ev.id, normalized_control_id=nc.id,
                                  coverage_type="full", confidence_score=100.0,
                                  created_by_ai=False, control_title=nc.name))
    fw_name = {f.id: f.name for f in db.query(UploadedFramework.id, UploadedFramework.name).all()}
    linked = 0
    for ln in db.query(NormalizedControlLink).filter(
            NormalizedControlLink.normalized_control_id == nc.id).all():
        p = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.id == ln.parsed_control_id).first()
        if not p:
            continue
        db.add(EvidenceControlMapping(
            evidence_id=ev.id, normalized_control_id=nc.id, parsed_control_id=p.id,
            uploaded_framework_id=p.uploaded_framework_id,
            framework_name=fw_name.get(p.uploaded_framework_id),
            control_code=p.original_reference or p.control_id, control_title=p.title,
            coverage_type="full", confidence_score=100.0, created_by_ai=False))
        linked += 1
    db.commit()
    return {"evidence_id": ev.id, "name": ev.name, "linked_controls": linked,
            "message": f"Added to the evidence library and linked to {linked} framework controls."}


def _set_review(db, current_user, nc_id, status_value):
    nc = db.query(NormalizedControl).filter(NormalizedControl.id == nc_id).first()
    if not nc:
        raise HTTPException(status_code=404, detail="Unified control not found")
    nc.review_status = status_value
    nc.reviewed_by = current_user.id
    nc.reviewed_at = datetime.utcnow()
    db.commit()
    return {"id": nc.id, "review_status": nc.review_status}


@router.post("/review/{nc_id}/approve")
def review_approve(nc_id: int, db: Session = Depends(get_db),
                   current_user: GRCUser = Depends(require_auth)):
    return _set_review(db, current_user, nc_id, "approved")


@router.post("/review/{nc_id}/flag")
def review_flag(nc_id: int, db: Session = Depends(get_db),
                current_user: GRCUser = Depends(require_auth)):
    return _set_review(db, current_user, nc_id, "flagged")


@router.post("/review/{nc_id}/remove-member")
def review_remove_member(nc_id: int, body: RemoveMemberRequest,
                         db: Session = Depends(get_db),
                         current_user: GRCUser = Depends(require_auth)):
    """Remove a wrong framework control from a unified control — drops the link
    AND the page's group mapping, keeping the master list correct by hand."""
    nc = db.query(NormalizedControl).filter(NormalizedControl.id == nc_id).first()
    if not nc:
        raise HTTPException(status_code=404, detail="Unified control not found")
    db.query(NormalizedControlLink).filter(
        NormalizedControlLink.normalized_control_id == nc_id,
        NormalizedControlLink.parsed_control_id == body.parsed_control_id,
    ).delete(synchronize_session=False)
    grp = db.query(CommonControlGroup).filter(
        CommonControlGroup.run_id == nc.run_id,
        CommonControlGroup.name == nc.name).first()
    if grp:
        db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == grp.id,
            CommonControlGroupMapping.parsed_control_id == body.parsed_control_id,
        ).delete(synchronize_session=False)
    db.commit()
    remaining = db.query(NormalizedControlLink).filter(
        NormalizedControlLink.normalized_control_id == nc_id).count()
    return {"id": nc_id, "removed": body.parsed_control_id, "remaining_members": remaining}


def persist_ai_groups(
    db: Session, tenant_id: int, user_id: Optional[int], ai_groups: List[dict],
    run_id: Optional[int] = None,
) -> Dict[str, Any]:
    """Take the raw AI-suggested groups and persist them as `CommonControlGroup`
    + `CommonControlGroupMapping` rows. Idempotent on (tenant, group name) and
    (group, control) — re-running an identical grouping won't duplicate rows.
    Returns a summary dict with counts + the affected group ids.
    """
    created_groups: List[CommonControlGroup] = []
    merged_groups: List[CommonControlGroup] = []
    mapping_count = 0

    # Track every code already taken (in the DB AND created earlier in THIS run)
    # so generated codes are guaranteed unique — the AI often reuses the same
    # short code (e.g. "CCG-008") for several groups, which used to collide.
    used_codes: set = {
        c for (c,) in db.query(CommonControlGroup.code).filter(
            CommonControlGroup.tenant_id == tenant_id).all() if c
    }

    for group_data in ai_groups:
        group_name = (group_data.get("name") or "").strip() or "Auto-generated Group"

        existing_by_name = db.query(CommonControlGroup).filter(
            CommonControlGroup.tenant_id == tenant_id,
            func.lower(CommonControlGroup.name) == group_name.lower()
        ).first()

        if existing_by_name:
            group = existing_by_name
            merged_groups.append(group)
        else:
            base = (group_data.get("code") or "CCG").strip() or "CCG"
            code = base
            n = 0
            while code in used_codes:        # unique vs DB + this run
                n += 1
                code = f"{base}-{n}"
            used_codes.add(code)

            group = CommonControlGroup(
                tenant_id=tenant_id,
                run_id=run_id,
                code=code,
                name=group_name,
                description=group_data.get("description"),
                category=group_data.get("category"),
                domain=group_data.get("domain"),
                keywords=group_data.get("keywords") or [],
                created_by=user_id,
            )
            db.add(group)
            db.flush()
            created_groups.append(group)

        for ctrl in group_data.get("control_ids") or []:
            ctrl_id = _coerce_int(ctrl.get("id"))
            ctrl_type = (ctrl.get("type") or "").strip().lower()
            if ctrl_id is None:
                continue

            if ctrl_type == "normalized":
                exists_q = db.query(CommonControlGroupMapping).filter(
                    CommonControlGroupMapping.group_id == group.id,
                    CommonControlGroupMapping.normalized_control_id == ctrl_id,
                ).first()
                if not exists_q:
                    db.add(CommonControlGroupMapping(
                        group_id=group.id,
                        normalized_control_id=ctrl_id,
                        mapping_source="ai",
                        mapping_confidence=0.8,
                    ))
                    mapping_count += 1
            elif ctrl_type in ("parsed", "framework"):
                # The old prompt allowed "framework"; we still accept it for
                # backwards-compat but always store it as a parsed mapping.
                exists_q = db.query(CommonControlGroupMapping).filter(
                    CommonControlGroupMapping.group_id == group.id,
                    CommonControlGroupMapping.parsed_control_id == ctrl_id,
                ).first()
                if not exists_q:
                    db.add(CommonControlGroupMapping(
                        group_id=group.id,
                        parsed_control_id=ctrl_id,
                        mapping_source="ai",
                        mapping_confidence=0.8,
                    ))
                    mapping_count += 1

    db.commit()
    return {
        "created_count": len(created_groups),
        "merged_count": len(merged_groups),
        "mapping_count": mapping_count,
        "created_group_ids": [g.id for g in created_groups],
        "merged_group_ids": [g.id for g in merged_groups],
    }


def serialize_group(group: CommonControlGroup, db: Session, include_controls: bool = False) -> dict:
    normalized_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group.id,
        CommonControlGroupMapping.normalized_control_id.isnot(None)
    ).count()
    
    framework_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group.id,
        CommonControlGroupMapping.framework_control_id.isnot(None)
    ).count()
    
    parsed_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group.id,
        CommonControlGroupMapping.parsed_control_id.isnot(None)
    ).count()

    # Standalone = single-framework-unique controls placed under the domain but
    # not consolidated into a cross-framework set (mapping_source='standalone').
    # These are stored as normalized-control mappings (one NormalizedControl per
    # framework-unique control), so count by normalized_control_id, not parsed.
    standalone_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group.id,
        CommonControlGroupMapping.normalized_control_id.isnot(None),
        CommonControlGroupMapping.mapping_source == "standalone",
    ).count()

    result = {
        "id": group.id,
        "tenant_id": group.tenant_id,
        "code": group.code,
        "name": group.name,
        "description": group.description,
        "category": group.category,
        "domain": group.domain,
        "keywords": group.keywords or [],
        "ai_summary": group.ai_summary,
        "evidence_types": group.evidence_types or [],
        "normalized_control_count": normalized_count,
        "framework_control_count": framework_count,
        "parsed_control_count": parsed_count,
        "standalone_control_count": standalone_count,
        "total_control_count": normalized_count + framework_count + parsed_count,
        "created_at": group.created_at.isoformat() if group.created_at else None,
        "updated_at": group.updated_at.isoformat() if group.updated_at else None,
        "created_by": group.created_by
    }
    
    if include_controls:
        mappings = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group.id
        ).all()
        
        normalized_controls = []
        framework_controls = []
        parsed_controls = []
        
        for mapping in mappings:
            if mapping.normalized_control_id:
                nc = db.query(NormalizedControl).filter(
                    NormalizedControl.id == mapping.normalized_control_id
                ).first()
                if nc:
                    normalized_controls.append({
                        "mapping_id": mapping.id,
                        "control_id": nc.id,
                        "code": nc.code,
                        "name": nc.name,
                        "statement": nc.statement,
                        "mapping_confidence": mapping.mapping_confidence,
                        "mapping_source": mapping.mapping_source
                    })
            
            if mapping.framework_control_id:
                fc = db.query(FrameworkControl).options(
                    joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
                ).filter(
                    FrameworkControl.id == mapping.framework_control_id
                ).first()
                if fc:
                    framework = fc.objective.domain.framework if fc.objective and fc.objective.domain else None
                    framework_controls.append({
                        "mapping_id": mapping.id,
                        "control_id": fc.id,
                        "code": fc.code,
                        "name": fc.name,
                        "statement": fc.statement,
                        "framework_id": framework.id if framework else None,
                        "framework_name": framework.name if framework else None,
                        "framework_code": framework.short_code if framework else None,
                        "mapping_confidence": mapping.mapping_confidence,
                        "mapping_source": mapping.mapping_source
                    })
            
            if mapping.parsed_control_id:
                pc = db.query(ParsedFrameworkControl).filter(
                    ParsedFrameworkControl.id == mapping.parsed_control_id
                ).first()
                if pc:
                    fw = db.query(UploadedFramework).filter(
                        UploadedFramework.id == pc.uploaded_framework_id
                    ).first()
                    parsed_controls.append({
                        "mapping_id": mapping.id,
                        "control_id": pc.id,
                        "code": pc.original_reference,
                        "name": pc.title,
                        "statement": pc.description or (pc.full_text[:500] if pc.full_text else None),
                        "framework_id": fw.id if fw else None,
                        "framework_name": fw.name if fw else None,
                        "mapping_confidence": mapping.mapping_confidence,
                        "mapping_source": mapping.mapping_source
                    })
        
        # Annotate each normalized control with the distinct frameworks it
        # consolidates, so the row shows "common across N frameworks" at a glance.
        nc_ids = [c["control_id"] for c in normalized_controls]
        if nc_ids:
            links = db.query(NormalizedControlLink).filter(
                NormalizedControlLink.normalized_control_id.in_(nc_ids)).all()
            p_map, f_map = _framework_label_maps(
                db,
                [ln.parsed_control_id for ln in links if ln.parsed_control_id],
                [ln.framework_control_id for ln in links if ln.framework_control_id],
            )
            nc_fw: Dict[int, set] = {}
            nc_cnt: Dict[int, int] = {}
            for ln in links:
                nc_cnt[ln.normalized_control_id] = nc_cnt.get(ln.normalized_control_id, 0) + 1
                if ln.parsed_control_id and ln.parsed_control_id in p_map:
                    nc_fw.setdefault(ln.normalized_control_id, set()).add(p_map[ln.parsed_control_id]["framework"])
                elif ln.framework_control_id and ln.framework_control_id in f_map:
                    nc_fw.setdefault(ln.normalized_control_id, set()).add(f_map[ln.framework_control_id]["framework"])
            for c in normalized_controls:
                fws = sorted(nc_fw.get(c["control_id"], set()))
                c["frameworks"] = fws
                c["framework_count"] = len(fws)
                c["linked_control_count"] = nc_cnt.get(c["control_id"], 0)

        result["normalized_controls"] = normalized_controls
        result["framework_controls"] = framework_controls
        result["parsed_controls"] = parsed_controls

    return result


@router.get("/categories")
def get_categories(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    categories = db.query(distinct(CommonControlGroup.category)).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        ),
        CommonControlGroup.category.isnot(None)
    ).all()
    
    return {"categories": [c[0] for c in categories if c[0]]}


@router.get("/domains")
def get_domains(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    domains = db.query(distinct(CommonControlGroup.domain)).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        ),
        CommonControlGroup.domain.isnot(None)
    ).all()
    
    return {"domains": [d[0] for d in domains if d[0]]}


@router.get("")
def list_groups(
    category: Optional[str] = None,
    domain: Optional[str] = None,
    search: Optional[str] = None,
    run_id: Optional[int] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    query = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        )
    )

    # Scope to a normalization session (run). Default: the owner's master
    # baseline, so the page shows the canonical library; the session switcher
    # passes ?run_id=… to view a custom framework-scoped session. Legacy groups
    # (run_id NULL) still show when no baseline exists yet.
    if run_id is None:
        from ..services.scoped_session import get_baseline_run
        _tid = get_user_primary_tenant(current_user, db)
        _base = get_baseline_run(db, _tid) if _tid else None
        run_id = _base.id if _base else None
    if run_id is not None:
        query = query.filter(CommonControlGroup.run_id == run_id)

    if category:
        query = query.filter(CommonControlGroup.category == category)
    if domain:
        query = query.filter(CommonControlGroup.domain == domain)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                CommonControlGroup.name.ilike(search_term),
                CommonControlGroup.code.ilike(search_term),
                CommonControlGroup.description.ilike(search_term)
            )
        )
    
    total = query.count()
    groups = query.order_by(CommonControlGroup.code).offset(skip).limit(limit).all()

    # True count of distinct mapped controls across ALL groups in scope (not just
    # the current page) — the stat card was summing only the visible page.
    group_ids_subq = query.with_entities(CommonControlGroup.id).subquery()
    # Count distinct mapped controls across ALL in-scope groups. Modern baselines
    # map by normalized_control_id (one NormalizedControl per unified/standalone
    # control); older ones mapped by parsed_control_id. Count whichever is present.
    norm_mapped = (db.query(CommonControlGroupMapping.normalized_control_id)
                   .filter(CommonControlGroupMapping.group_id.in_(group_ids_subq),
                           CommonControlGroupMapping.normalized_control_id.isnot(None))
                   .distinct().count())
    parsed_mapped = (db.query(CommonControlGroupMapping.parsed_control_id)
                     .filter(CommonControlGroupMapping.group_id.in_(group_ids_subq),
                             CommonControlGroupMapping.parsed_control_id.isnot(None))
                     .distinct().count())
    total_mapped = norm_mapped + parsed_mapped

    # RAW framework controls behind the library (one link per original framework
    # control) — lets the UI reconcile "3,419 raw → 2,332 unified" instead of
    # presenting shared/unique as if they were additional to the unified count.
    total_raw = 0
    if run_id is not None:
        total_raw = (db.query(NormalizedControlLink.parsed_control_id)
                     .join(NormalizedControl,
                           NormalizedControl.id == NormalizedControlLink.normalized_control_id)
                     .filter(NormalizedControl.run_id == run_id,
                             NormalizedControlLink.parsed_control_id.isnot(None))
                     .distinct().count())

    return {
        "items": [serialize_group(g, db) for g in groups],
        "total": total,
        "total_mapped_controls": total_mapped,
        "total_raw_controls": total_raw,
        "skip": skip,
        "limit": limit
    }


@router.post("", status_code=status.HTTP_201_CREATED)
def create_group(
    group_data: CommonControlGroupCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    
    existing = db.query(CommonControlGroup).filter(
        CommonControlGroup.tenant_id == tenant_id,
        CommonControlGroup.code == group_data.code
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A group with this code already exists"
        )
    
    keywords = group_data.keywords
    if not keywords and group_data.description:
        keywords = generate_keywords_for_group(group_data.name, group_data.description)
    
    group = CommonControlGroup(
        tenant_id=tenant_id,
        code=group_data.code,
        name=group_data.name,
        description=group_data.description,
        category=group_data.category,
        domain=group_data.domain,
        keywords=keywords or [],
        evidence_types=group_data.evidence_types or [],
        created_by=current_user.id
    )
    db.add(group)
    db.commit()
    db.refresh(group)
    
    return serialize_group(group, db)


@router.get("/{group_id}")
def get_group(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    return serialize_group(group, db, include_controls=True)


@router.put("/{group_id}")
def update_group(
    group_id: int,
    group_data: CommonControlGroupUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    if group_data.code and group_data.code != group.code:
        existing = db.query(CommonControlGroup).filter(
            CommonControlGroup.tenant_id == group.tenant_id,
            CommonControlGroup.code == group_data.code,
            CommonControlGroup.id != group_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A group with this code already exists"
            )
    
    update_data = group_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(group, field, value)
    
    group.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(group)
    
    return serialize_group(group, db, include_controls=True)


@router.delete("/{group_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_group(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    db.delete(group)
    db.commit()
    return None


@router.post("/{group_id}/controls")
def add_controls_to_group(
    group_id: int,
    mapping_data: GroupMappingCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    created_mappings = []
    
    for nc_id in (mapping_data.normalized_control_ids or []):
        nc = db.query(NormalizedControl).filter(NormalizedControl.id == nc_id).first()
        if not nc:
            continue
        
        existing = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group_id,
            CommonControlGroupMapping.normalized_control_id == nc_id
        ).first()
        if existing:
            continue
        
        mapping = CommonControlGroupMapping(
            group_id=group_id,
            normalized_control_id=nc_id,
            mapping_source="manual"
        )
        db.add(mapping)
        created_mappings.append({"type": "normalized", "control_id": nc_id})
    
    for fc_id in (mapping_data.framework_control_ids or []):
        fc = db.query(FrameworkControl).filter(FrameworkControl.id == fc_id).first()
        if not fc:
            continue
        
        existing = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group_id,
            CommonControlGroupMapping.framework_control_id == fc_id
        ).first()
        if existing:
            continue
        
        mapping = CommonControlGroupMapping(
            group_id=group_id,
            framework_control_id=fc_id,
            mapping_source="manual"
        )
        db.add(mapping)
        created_mappings.append({"type": "framework", "control_id": fc_id})
    
    for pc_id in (mapping_data.parsed_control_ids or []):
        pc = db.query(ParsedFrameworkControl).filter(ParsedFrameworkControl.id == pc_id).first()
        if not pc:
            continue
        existing = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group_id,
            CommonControlGroupMapping.parsed_control_id == pc_id
        ).first()
        if existing:
            continue
        mapping = CommonControlGroupMapping(
            group_id=group_id,
            parsed_control_id=pc_id,
            mapping_source="manual"
        )
        db.add(mapping)
        created_mappings.append({"type": "parsed", "control_id": pc_id})
    
    db.commit()
    
    return {
        "message": f"Added {len(created_mappings)} controls to group",
        "mappings_created": created_mappings,
        "group": serialize_group(group, db, include_controls=True)
    }


@router.delete("/{group_id}/controls/{mapping_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_control_from_group(
    group_id: int,
    mapping_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    mapping = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.id == mapping_id,
        CommonControlGroupMapping.group_id == group_id
    ).first()
    
    if not mapping:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Mapping not found"
        )
    
    db.delete(mapping)
    db.commit()
    return None


# ─── Async dispatch + status polling for AI auto-grouping ──────────────────
# Long-running grouping was blocking the request thread for minutes. We now
# run the work in a Python background thread (no separate Celery worker
# process required) and write progress to Redis-backed job_status so the UI
# can poll. The Celery task in `tasks/control_library.py` is still available
# for environments that prefer to scale work onto a worker pool.


def _run_auto_grouping_threaded(
    tenant_slug: str,
    job_id: str,
    framework_ids: List[int],
    user_id: Optional[int],
) -> None:
    """Background worker — opens a dedicated tenant session, runs the full
    auto-grouping pipeline, and reports progress via job_status. Designed to
    be launched from a daemon thread so a long AI call doesn't block the
    request thread or require a separate Celery worker.
    """
    namespace = "control_auto_group"
    db_session = None
    try:
        # Acquire a tenant-scoped DB session that lives independently of the
        # request that started the job.
        from ....db import get_tenant_session_factory
        from ....models import Tenant

        SessionLocal = get_tenant_session_factory(tenant_slug)
        db_session = SessionLocal()

        tenant = db_session.query(Tenant).filter(Tenant.slug == tenant_slug).first()
        if not tenant:
            raise RuntimeError(f"Tenant slug '{tenant_slug}' not found")

        _jobs_logger.info("[auto_group] tenant=%s job=%s — loading controls", tenant_slug, job_id)
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "running",
            "phase": "loading_controls",
            "message": "Loading controls from selected frameworks…",
            "progress_percent": 5,
        })

        if (get_job_status(tenant_slug, namespace, job_id) or {}).get("cancel_requested"):
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "cancelled", "phase": "cancelled",
                "message": "Cancelled before it started.", "progress_percent": 100,
            })
            return

        # Preferred path (mirrors the Celery task): if a master baseline exists,
        # build a framework-scoped session by FILTERING it — no full AI re-run.
        from ..services.scoped_session import get_baseline_run, build_scoped_session
        if get_baseline_run(db_session, tenant.id) is not None:
            def _scoped_progress(done, total, msg):
                cur = get_job_status(tenant_slug, namespace, job_id) or {}
                if cur.get("cancel_requested"):
                    raise AutoGroupCancelled()
                cur.update({"status": "running", "phase": "scoping", "message": msg,
                            "progress_percent": int(done * 100 / max(1, total))})
                set_job_status(tenant_slug, namespace, job_id, cur)
            try:
                res = build_scoped_session(
                    db_session, tenant.id, framework_ids or [], user_id=user_id,
                    progress_cb=_scoped_progress,
                    should_cancel=lambda: bool((get_job_status(tenant_slug, namespace, job_id) or {}).get("cancel_requested")),
                )
            except AutoGroupCancelled:
                set_job_status(tenant_slug, namespace, job_id, {
                    "status": "cancelled", "phase": "cancelled",
                    "message": "Cancelled.", "progress_percent": 100})
                return
            except ValueError as ve:
                set_job_status(tenant_slug, namespace, job_id, {
                    "status": "failed", "error": str(ve), "progress_percent": 100})
                return
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "completed", "phase": "done",
                "message": (
                    f"Built a view of {res['unified_controls']} unified + "
                    f"{res.get('standalone', 0)} standalone controls "
                    + ("(new framework classified onto the master list)."
                       if res["ai_used"] else "— reused the master baseline, no AI re-run.")),
                "summary": res, "run_id": res["run_id"], "progress_percent": 100,
            })
            _jobs_logger.info("[auto_group] SCOPED tenant=%s job=%s run=%s ai_used=%s",
                              tenant_slug, job_id, res["run_id"], res["ai_used"])
            return

        controls = _fetch_controls_for_grouping(db_session, tenant.id, framework_ids)
        if len(controls) < 2:
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "failed",
                "phase": "loading_controls",
                "error": "Not enough controls to perform auto-grouping (need at least 2).",
                "progress_percent": 100,
            })
            return

        _jobs_logger.info(
            "[auto_group] tenant=%s job=%s — calling AI on %d controls",
            tenant_slug, job_id, len(controls),
        )
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "running",
            "phase": "ai_grouping",
            "message": f"Asking AI to cluster {len(controls)} controls…",
            "control_count": len(controls),
            "progress_percent": 25,
        })

        def _progress(done, total_batches):
            cur = get_job_status(tenant_slug, namespace, job_id) or {}
            if cur.get("cancel_requested"):
                raise AutoGroupCancelled()
            pct = 25 + int((done / max(1, total_batches)) * 50)
            cur.update({
                "status": "running", "phase": "ai_grouping",
                "message": f"Grouping… batch {done} of {total_batches}",
                "progress_percent": pct,
            })
            set_job_status(tenant_slug, namespace, job_id, cur)

        try:
            ai_groups = ai_auto_group_controls(controls, progress_cb=_progress)
        except AutoGroupCancelled:
            _jobs_logger.info("[auto_group] tenant=%s job=%s — CANCELLED", tenant_slug, job_id)
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "cancelled", "phase": "cancelled",
                "message": "Auto-grouping cancelled by user.", "progress_percent": 100,
            })
            return

        _jobs_logger.info(
            "[auto_group] tenant=%s job=%s — persisting %d groups",
            tenant_slug, job_id, len(ai_groups),
        )
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "running",
            "phase": "persisting",
            "message": f"Saving {len(ai_groups)} group(s) to the database…",
            "control_count": len(controls),
            "group_count": len(ai_groups),
            "progress_percent": 80,
        })

        summary = persist_ai_groups(db_session, tenant.id, user_id, ai_groups)

        _jobs_logger.info(
            "[auto_group] tenant=%s job=%s — DONE created=%d merged=%d mappings=%d",
            tenant_slug, job_id,
            summary["created_count"], summary["merged_count"], summary["mapping_count"],
        )
        # ── Second phase: normalize each new domain (same flow) ──
        norm = {"normalized_controls_created": 0, "links_created": 0}
        new_group_ids = summary.get("created_group_ids") or []
        if new_group_ids:
            from ..services.normalization import run_normalization
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "running", "phase": "normalizing",
                "message": f"Normalizing {len(new_group_ids)} domain(s)…", "progress_percent": 80})

            def _norm_progress(done, total, msg):
                cur = get_job_status(tenant_slug, namespace, job_id) or {}
                if cur.get("cancel_requested"):
                    raise AutoGroupCancelled()
                pct = 80 + int((done / max(1, total)) * 18)
                cur.update({"status": "running", "phase": "normalizing", "message": msg, "progress_percent": pct})
                set_job_status(tenant_slug, namespace, job_id, cur)

            try:
                norm = run_normalization(
                    db_session, tenant.id, new_group_ids, progress_cb=_norm_progress,
                    should_cancel=lambda: bool((get_job_status(tenant_slug, namespace, job_id) or {}).get("cancel_requested")))
            except AutoGroupCancelled:
                set_job_status(tenant_slug, namespace, job_id, {
                    "status": "cancelled", "phase": "cancelled",
                    "message": "Cancelled during normalization.", "progress_percent": 100})
                return

        set_job_status(tenant_slug, namespace, job_id, {
            "status": "completed",
            "phase": "done",
            "message": (
                f"Created {summary['created_count']} domain(s) and "
                f"{norm['normalized_controls_created']} normalized control(s)."
            ),
            "control_count": len(controls),
            "group_count": len(ai_groups),
            "summary": {**summary, **norm},
            "progress_percent": 100,
        })
    except Exception as exc:
        _jobs_logger.exception("[auto_group] tenant=%s job=%s FAILED", tenant_slug, job_id)
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "failed",
            "phase": "error",
            "error": str(exc)[:500],
            "progress_percent": 100,
        })
    finally:
        if db_session is not None:
            try:
                db_session.close()
            except Exception:
                pass


@router.post("/auto-group/dispatch")
def auto_group_dispatch(
    request: AutoGroupRequest,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    if not check_ai_available():
        raise HTTPException(
            status_code=503,
            detail={
                "error": "AI features unavailable",
                "message": "OpenAI API key is not configured.",
                "fallback_available": True,
            },
        )

    tenant_slug = getattr(http_request.state, "tenant_slug", None)
    if not tenant_slug:
        from ....models import Tenant
        t = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        tenant_slug = t.slug if t else None
    if not tenant_slug:
        raise HTTPException(status_code=400, detail="Could not resolve tenant slug")

    job_id = uuid.uuid4().hex
    framework_ids = list(request.framework_ids or [])
    user_id = current_user.id

    set_job_status(tenant_slug, "control_auto_group", job_id, {
        "status": "queued",
        "phase": "queued",
        "message": "Starting…",
        "progress_percent": 1,
        "framework_ids": framework_ids,
    })
    # Pointer to the most-recent job so the UI can resurface an in-flight job
    # after a page reload (persistent progress + stop button).
    set_job_status(tenant_slug, "control_auto_group", "latest", {"job_id": job_id})

    # ── Fast path: a PURE FILTER (every selected framework is already in the
    # master baseline → no AI needed) runs SYNCHRONOUSLY right here — no queue,
    # no Celery worker. A 5-second DB filter shouldn't depend on a background
    # process. Only a brand-new framework (AI classify) or a full rebuild below
    # actually needs the queue.
    try:
        from ..services.scoped_session import get_baseline_run, build_scoped_session
        base = get_baseline_run(db, tenant_id)
        if base and len(framework_ids) >= 2:
            base_gids = [gid for (gid,) in db.query(CommonControlGroup.id).filter(
                CommonControlGroup.run_id == base.id).all()]
            base_fws = set()
            if base_gids:
                base_fws = {fid for (fid,) in db.query(ParsedFrameworkControl.uploaded_framework_id)
                            .join(CommonControlGroupMapping,
                                  CommonControlGroupMapping.parsed_control_id == ParsedFrameworkControl.id)
                            .filter(CommonControlGroupMapping.group_id.in_(base_gids)).distinct().all()}
            new_fws = [f for f in framework_ids if f not in base_fws]
            if not new_fws:   # pure filter — no AI → run inline
                res = build_scoped_session(db, tenant_id, framework_ids, user_id=user_id)
                set_job_status(tenant_slug, "control_auto_group", job_id, {
                    "status": "completed", "phase": "done",
                    "message": (f"Built a unified view of {res['unified_controls']} controls "
                                "— reused the master baseline, no AI."),
                    "summary": res, "run_id": res["run_id"], "progress_percent": 100,
                })
                _jobs_logger.info("[auto_group] SYNC pure-filter tenant=%s job=%s run=%s",
                                  tenant_slug, job_id, res["run_id"])
                return {"job_id": job_id, "status": "completed",
                        "run_id": res["run_id"], "summary": res}
    except ValueError as ve:
        set_job_status(tenant_slug, "control_auto_group", job_id, {"status": "failed", "error": str(ve)})
        raise HTTPException(status_code=400, detail=str(ve))
    except HTTPException:
        raise
    except Exception as exc:   # unexpected → fall through to the queue, don't break the button
        _jobs_logger.warning("[auto_group] sync pure-filter failed (%s) — falling back to queue", exc)

    # Preferred path: dispatch to Celery so the work runs on the dedicated
    # `parsing` queue worker. The worker carries its own DB session per the
    # TenantTask base, status flows through Redis-backed job_status, and a
    # restart of the API process can't lose the job mid-flight.
    #
    # Fallback path (in-process thread) kicks in when Celery's broker is
    # not reachable from this process. Useful for dev boxes without Redis
    # and for catastrophic recovery on prod — surfaces as a log line so
    # ops can see it happened.
    use_thread_fallback = False
    try:
        from ....tasks.control_library import ai_auto_group as _ai_auto_group_task
        # IMPORTANT: TenantTask.__call__ enforces tenant_slug as the FIRST
        # positional arg (backend/grc/tasks/base.py:177-183). Pass it via
        # `args=[...]`, not kwargs, otherwise the worker raises
        # `ValueError: ... called without tenant_slug as first arg`.
        async_result = _ai_auto_group_task.apply_async(
            args=[tenant_slug, job_id],
            kwargs={"framework_ids": framework_ids, "user_id": user_id},
            queue="parsing",
        )
        _jobs_logger.info(
            "[auto_group] DISPATCH (celery) tenant=%s job=%s framework_ids=%s task_id=%s",
            tenant_slug, job_id, framework_ids, async_result.id,
        )
    except Exception as exc:
        _jobs_logger.warning(
            "[auto_group] Celery dispatch failed (%s) — falling back to in-process thread",
            exc,
        )
        use_thread_fallback = True

    if use_thread_fallback:
        thread = threading.Thread(
            target=_run_auto_grouping_threaded,
            args=(tenant_slug, job_id, framework_ids, user_id),
            name=f"auto_group:{job_id}",
            daemon=True,
        )
        thread.start()
        _jobs_logger.info(
            "[auto_group] DISPATCH (thread fallback) tenant=%s job=%s framework_ids=%s",
            tenant_slug, job_id, framework_ids,
        )

    return {"job_id": job_id, "status": "queued"}


@router.get("/auto-group/status/{job_id}")
def auto_group_status(
    job_id: str,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_slug = getattr(http_request.state, "tenant_slug", None)
    if not tenant_slug:
        from ....models import Tenant
        tenant_id = get_user_primary_tenant(current_user, db)
        t = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        tenant_slug = t.slug if t else None
    if not tenant_slug:
        raise HTTPException(status_code=400, detail="Could not resolve tenant slug")
    return get_job_status(tenant_slug, "control_auto_group", job_id)


def _resolve_tenant_slug(http_request: Request, db: Session, current_user: GRCUser) -> str:
    tenant_slug = getattr(http_request.state, "tenant_slug", None)
    if not tenant_slug:
        from ....models import Tenant
        tenant_id = get_user_primary_tenant(current_user, db)
        t = db.query(Tenant).filter(Tenant.id == tenant_id).first()
        tenant_slug = t.slug if t else None
    if not tenant_slug:
        raise HTTPException(status_code=400, detail="Could not resolve tenant slug")
    return tenant_slug


@router.post("/auto-group/cancel/{job_id}")
def auto_group_cancel(
    job_id: str,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Request cancellation of a running/queued auto-group job. Sets a flag the
    worker checks between batches and stops cleanly (status -> cancelled)."""
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    st = get_job_status(tenant_slug, "control_auto_group", job_id) or {}
    cur = st.get("status")
    if cur in ("completed", "failed", "cancelled"):
        return {"status": cur, "message": "Job already finished."}
    st["cancel_requested"] = True
    if cur == "queued":
        # Not started yet — cancel immediately. If a worker later picks it up,
        # the start-of-task cancel check aborts it before any AI work.
        st.update({"status": "cancelled", "phase": "cancelled",
                   "message": "Cancelled before it started.", "progress_percent": 100})
        set_job_status(tenant_slug, "control_auto_group", job_id, st)
        return {"status": "cancelled", "job_id": job_id}
    st.update({"status": "cancelling", "message": "Cancelling — stopping after the current batch…"})
    set_job_status(tenant_slug, "control_auto_group", job_id, st)
    return {"status": "cancelling", "job_id": job_id}


@router.get("/auto-group/active")
def auto_group_active(
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Return the tenant's most-recent auto-group job + whether it's still
    in flight, so the UI can resurface progress (and a Stop button) after a
    page reload or dialog close."""
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    latest = get_job_status(tenant_slug, "control_auto_group", "latest") or {}
    job_id = latest.get("job_id")
    if not job_id:
        return {"active": False}
    st = get_job_status(tenant_slug, "control_auto_group", job_id) or {}
    active = st.get("status") in ("queued", "running", "cancelling")
    return {"active": active, "job_id": job_id, **st}


# ─────────────────────────────────────────────────────────────────────────────
# AI Normalization — for each domain (CommonControlGroup), cluster its controls
# into normalized controls + cross-framework links. Same dispatch/progress/stop
# contract as auto-group, namespace "control_normalization".
# ─────────────────────────────────────────────────────────────────────────────

class NormalizeRequest(BaseModel):
    group_ids: Optional[List[int]] = None


def _run_normalization_threaded(tenant_slug, job_id, group_ids, user_id):
    namespace = "control_normalization"
    db_session = None
    try:
        from ....db import get_tenant_session_factory
        from ....models import Tenant
        SessionLocal = get_tenant_session_factory(tenant_slug)
        db_session = SessionLocal()
        tenant = db_session.query(Tenant).filter(Tenant.slug == tenant_slug).first()
        if not tenant:
            raise RuntimeError(f"Tenant slug '{tenant_slug}' not found")
        from ..services.normalization import run_normalization

        def _progress(done, total, msg):
            cur = get_job_status(tenant_slug, namespace, job_id) or {}
            if cur.get("cancel_requested"):
                raise AutoGroupCancelled()
            pct = 5 + int((done / max(1, total)) * 90)
            cur.update({"status": "running", "phase": "normalizing",
                        "message": msg, "progress_percent": pct})
            set_job_status(tenant_slug, namespace, job_id, cur)

        def _should_cancel():
            return bool((get_job_status(tenant_slug, namespace, job_id) or {}).get("cancel_requested"))

        try:
            summary = run_normalization(db_session, tenant.id, group_ids,
                                        progress_cb=_progress, should_cancel=_should_cancel)
        except AutoGroupCancelled:
            set_job_status(tenant_slug, namespace, job_id, {
                "status": "cancelled", "phase": "cancelled",
                "message": "Normalization cancelled by user.", "progress_percent": 100})
            return
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "completed", "phase": "done",
            "message": (f"Created {summary['normalized_controls_created']} normalized "
                        f"control(s) across {summary['domains_processed']} domain(s)."),
            "summary": summary, "progress_percent": 100})
    except Exception as exc:
        _jobs_logger.exception("[normalize] tenant=%s job=%s FAILED", tenant_slug, job_id)
        set_job_status(tenant_slug, namespace, job_id, {
            "status": "failed", "phase": "error", "error": str(exc)[:500], "progress_percent": 100})
    finally:
        if db_session is not None:
            try:
                db_session.close()
            except Exception:
                pass


@router.post("/normalize/dispatch")
def normalize_dispatch(
    request: NormalizeRequest,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    if not check_ai_available():
        raise HTTPException(status_code=503, detail={
            "error": "AI features unavailable",
            "message": "OpenAI API key is not configured.", "fallback_available": False})
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    job_id = uuid.uuid4().hex
    group_ids = list(request.group_ids or [])
    user_id = current_user.id
    set_job_status(tenant_slug, "control_normalization", job_id, {
        "status": "queued", "phase": "queued", "message": "Starting…",
        "progress_percent": 1, "group_ids": group_ids})
    set_job_status(tenant_slug, "control_normalization", "latest", {"job_id": job_id})

    use_thread_fallback = False
    try:
        from ....tasks.control_library import ai_normalize_controls as _task
        _task.apply_async(args=[tenant_slug, job_id],
                          kwargs={"group_ids": group_ids, "user_id": user_id}, queue="parsing")
        _jobs_logger.info("[normalize] DISPATCH (celery) tenant=%s job=%s groups=%s", tenant_slug, job_id, group_ids)
    except Exception as exc:
        _jobs_logger.warning("[normalize] Celery dispatch failed (%s) — thread fallback", exc)
        use_thread_fallback = True
    if use_thread_fallback:
        threading.Thread(target=_run_normalization_threaded,
                         args=(tenant_slug, job_id, group_ids, user_id),
                         name=f"normalize:{job_id}", daemon=True).start()
    return {"job_id": job_id, "status": "queued"}


@router.get("/normalize/status/{job_id}")
def normalize_status(job_id: str, http_request: Request, db: Session = Depends(get_db),
                     current_user: GRCUser = Depends(require_auth)):
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    return get_job_status(tenant_slug, "control_normalization", job_id)


# ── Master-baseline creation: the first-time / rebuild normalization pipeline ──
# Builds the unified library into a NEW run (is_baseline=False); the admin reviews
# it via the session dropdown and PROMOTES it. The live baseline is never touched
# by the build — only the explicit promote step swaps it.
_BASELINE_NS = "control_baseline_build"


def _run_baseline_build_threaded(tenant_slug, job_id, user_id, label):
    db_session = None
    try:
        from ....db import get_tenant_session_factory
        from ....models import Tenant
        db_session = get_tenant_session_factory(tenant_slug)()
        tenant = db_session.query(Tenant).filter(Tenant.slug == tenant_slug).first()
        if not tenant:
            raise RuntimeError(f"Tenant slug '{tenant_slug}' not found")
        from ..services.baseline_builder import build_baseline_run

        def _progress(done, total, msg):
            cur = get_job_status(tenant_slug, _BASELINE_NS, job_id) or {}
            if cur.get("cancel_requested"):
                raise AutoGroupCancelled()
            cur.update({"status": "running", "phase": "building", "message": msg,
                        "progress_percent": int(done * 100 / max(1, total))})
            set_job_status(tenant_slug, _BASELINE_NS, job_id, cur)

        try:
            res = build_baseline_run(
                db_session, tenant.id, label=label, user_id=user_id, progress_cb=_progress,
                should_cancel=lambda: bool((get_job_status(tenant_slug, _BASELINE_NS, job_id) or {}).get("cancel_requested")))
        except AutoGroupCancelled:
            set_job_status(tenant_slug, _BASELINE_NS, job_id, {
                "status": "cancelled", "phase": "cancelled", "message": "Cancelled.", "progress_percent": 100})
            return
        set_job_status(tenant_slug, _BASELINE_NS, job_id, {
            "status": "completed", "phase": "done",
            "message": (f"Built {res['unified_controls']} unified + {res['standalone']} standalone "
                        f"controls. Review it from the dropdown, then Promote to make it live."),
            "summary": res, "run_id": res["run_id"], "progress_percent": 100})
    except Exception as exc:
        _jobs_logger.exception("[baseline_build] tenant=%s job=%s FAILED", tenant_slug, job_id)
        set_job_status(tenant_slug, _BASELINE_NS, job_id, {
            "status": "failed", "phase": "error", "error": str(exc)[:500], "progress_percent": 100})
    finally:
        if db_session is not None:
            try:
                db_session.close()
            except Exception:
                pass


class BaselineBuildRequest(BaseModel):
    confirm: Optional[str] = None


@router.post("/baseline/build-dispatch")
def baseline_build_dispatch(request: BaselineBuildRequest, http_request: Request,
                            db: Session = Depends(get_db),
                            current_user: GRCUser = Depends(require_auth)):
    """Kick off a one-time / rebuild MASTER baseline (sees ALL frameworks + controls).

    GUARDED so an accidental click can't trigger a ~40-min AI run: the caller MUST
    send the exact confirmation phrase. Runs on the Celery 'parsing' worker ONLY —
    never in the API process — so it is heavy-but-stoppable and never blocks
    requests. Builds a NEW candidate run (is_baseline=False); review + Promote to
    make it live. The live baseline is never touched here.
    """
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    # ── Guard: must type the confirmation phrase exactly (case-insensitive) ──
    if (request.confirm or "").strip().upper() != "REBUILD BASELINE":
        raise HTTPException(status_code=400, detail={
            "error": "confirmation_required",
            "message": ("Rebuilding the master baseline is a heavy one-time AI run over ALL "
                        "frameworks and controls. Type 'REBUILD BASELINE' to confirm.")})
    if not check_ai_available():
        raise HTTPException(status_code=503, detail={
            "error": "AI features unavailable", "message": "OpenAI API key is not configured."})
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    job_id = uuid.uuid4().hex
    set_job_status(tenant_slug, _BASELINE_NS, job_id, {
        "status": "queued", "phase": "queued",
        "message": "Queued on the background worker…", "progress_percent": 1})
    set_job_status(tenant_slug, _BASELINE_NS, "latest", {"job_id": job_id})
    # Celery ONLY — the single heavy job that must run on the worker, never in the
    # API process. (An in-process thread can't be stopped by stopping Celery —
    # exactly what bit us before.) If the broker is unreachable, FAIL clearly
    # instead of silently running in-process.
    try:
        from ....tasks.control_library import build_master_baseline as _baseline_task
        _baseline_task.apply_async(
            args=[tenant_slug, job_id],
            kwargs={"user_id": current_user.id, "label": "Master baseline (candidate)"},
            queue="parsing")
    except Exception as exc:
        _jobs_logger.exception("[baseline_build] celery dispatch failed: %s", exc)
        set_job_status(tenant_slug, _BASELINE_NS, job_id, {
            "status": "failed", "phase": "error", "progress_percent": 100,
            "error": ("Baseline build needs the background worker. Start the Celery "
                      "'parsing' worker and retry.")})
        raise HTTPException(status_code=503, detail={
            "error": "worker_unavailable",
            "message": ("The baseline build runs on the background worker, which isn't "
                        "reachable. Start the Celery 'parsing' worker and retry.")})
    _jobs_logger.info("[baseline_build] DISPATCH (celery) tenant=%s job=%s", tenant_slug, job_id)
    return {"job_id": job_id, "status": "queued"}


@router.get("/baseline/build-status/{job_id}")
def baseline_build_status(job_id: str, http_request: Request, db: Session = Depends(get_db),
                          current_user: GRCUser = Depends(require_auth)):
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    return get_job_status(tenant_slug, _BASELINE_NS, job_id)


@router.post("/baseline/build-cancel/{job_id}")
def baseline_build_cancel(job_id: str, http_request: Request, db: Session = Depends(get_db),
                          current_user: GRCUser = Depends(require_auth)):
    """Request cancellation of a running/queued master-baseline build. Sets a flag
    the worker checks between phases; it stops cleanly. The live baseline is never
    touched (the candidate run is only persisted near the very end)."""
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    st = get_job_status(tenant_slug, _BASELINE_NS, job_id) or {}
    cur = st.get("status")
    if cur in ("completed", "failed", "cancelled"):
        return {"status": cur, "message": "Build already finished."}
    st["cancel_requested"] = True
    if cur == "queued":
        # Not started yet — cancel immediately; the start-of-task check aborts it
        # before any AI work if a worker later picks it up.
        st.update({"status": "cancelled", "phase": "cancelled",
                   "message": "Cancelled before it started.", "progress_percent": 100})
        set_job_status(tenant_slug, _BASELINE_NS, job_id, st)
        return {"status": "cancelled", "job_id": job_id}
    st.update({"status": "cancelling", "message": "Cancelling — stopping after the current phase…"})
    set_job_status(tenant_slug, _BASELINE_NS, job_id, st)
    return {"status": "cancelling", "job_id": job_id}


@router.post("/baseline/promote/{run_id}")
def promote_baseline(run_id: int, db: Session = Depends(get_db),
                     current_user: GRCUser = Depends(require_auth)):
    """Approve a candidate run: demote the current baseline and make this run live."""
    tenant_id = get_user_primary_tenant(current_user, db)
    run = db.query(NormalizationRun).filter(
        NormalizationRun.id == run_id, NormalizationRun.tenant_id == tenant_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.status != "completed":
        raise HTTPException(status_code=400, detail="Only a completed run can be promoted.")
    if run.is_baseline:
        return {"ok": True, "baseline_run_id": run_id, "note": "already the baseline"}
    db.query(NormalizationRun).filter(
        NormalizationRun.tenant_id == tenant_id, NormalizationRun.is_baseline.is_(True)
    ).update({NormalizationRun.is_baseline: False}, synchronize_session=False)
    run.is_baseline = True
    if run.scope == "custom":
        run.scope = "full"
    db.commit()
    _jobs_logger.info("[baseline_build] PROMOTED run=%s tenant=%s", run_id, tenant_id)
    return {"ok": True, "baseline_run_id": run_id}


@router.post("/normalize/cancel/{job_id}")
def normalize_cancel(job_id: str, http_request: Request, db: Session = Depends(get_db),
                     current_user: GRCUser = Depends(require_auth)):
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    st = get_job_status(tenant_slug, "control_normalization", job_id) or {}
    cur = st.get("status")
    if cur in ("completed", "failed", "cancelled"):
        return {"status": cur, "message": "Job already finished."}
    st["cancel_requested"] = True
    if cur == "queued":
        st.update({"status": "cancelled", "phase": "cancelled",
                   "message": "Cancelled before it started.", "progress_percent": 100})
        set_job_status(tenant_slug, "control_normalization", job_id, st)
        return {"status": "cancelled", "job_id": job_id}
    st.update({"status": "cancelling", "message": "Cancelling — stopping after the current domain…"})
    set_job_status(tenant_slug, "control_normalization", job_id, st)
    return {"status": "cancelling", "job_id": job_id}


@router.get("/normalize/active")
def normalize_active(http_request: Request, db: Session = Depends(get_db),
                     current_user: GRCUser = Depends(require_auth)):
    tenant_slug = _resolve_tenant_slug(http_request, db, current_user)
    latest = get_job_status(tenant_slug, "control_normalization", "latest") or {}
    job_id = latest.get("job_id")
    if not job_id:
        return {"active": False}
    st = get_job_status(tenant_slug, "control_normalization", job_id) or {}
    active = st.get("status") in ("queued", "running", "cancelling")
    return {"active": active, "job_id": job_id, **st}


@router.get("/normalize/by-domain")
def normalized_controls_by_domain(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Normalized controls grouped by domain, with how many framework/parsed
    controls each consolidates — drives the per-domain Control Library view."""
    ncs = db.query(NormalizedControl).filter(
        NormalizedControl.source == "ai_normalized"
    ).order_by(NormalizedControl.domain, NormalizedControl.code).all()

    # Load every link once, then resolve each link's framework so we can show,
    # per normalized control, the distinct frameworks it consolidates.
    nc_ids = [nc.id for nc in ncs]
    links = []
    if nc_ids:
        links = db.query(NormalizedControlLink).filter(
            NormalizedControlLink.normalized_control_id.in_(nc_ids)).all()
    p_map, f_map = _framework_label_maps(
        db,
        [ln.parsed_control_id for ln in links if ln.parsed_control_id],
        [ln.framework_control_id for ln in links if ln.framework_control_id],
    )
    link_counts: Dict[int, int] = {}
    nc_frameworks: Dict[int, set] = {}
    for ln in links:
        link_counts[ln.normalized_control_id] = link_counts.get(ln.normalized_control_id, 0) + 1
        if ln.parsed_control_id and ln.parsed_control_id in p_map:
            nc_frameworks.setdefault(ln.normalized_control_id, set()).add(p_map[ln.parsed_control_id]["framework"])
        elif ln.framework_control_id and ln.framework_control_id in f_map:
            nc_frameworks.setdefault(ln.normalized_control_id, set()).add(f_map[ln.framework_control_id]["framework"])

    domains: Dict[str, dict] = {}
    for nc in ncs:
        dom = nc.domain or "Uncategorized"
        domains.setdefault(dom, {"domain": dom, "controls": []})
        domains[dom]["controls"].append({
            "id": nc.id, "code": nc.code, "name": nc.name,
            "statement": nc.statement, "objective": nc.objective,
            "linked_control_count": link_counts.get(nc.id, 0),
            "frameworks": sorted(nc_frameworks.get(nc.id, set())),
            "framework_count": len(nc_frameworks.get(nc.id, set())),
        })
    result = []
    for dom, payload in domains.items():
        payload["normalized_count"] = len(payload["controls"])
        payload["linked_total"] = sum(c["linked_control_count"] for c in payload["controls"])
        result.append(payload)
    result.sort(key=lambda d: d["domain"])
    return {"domains": result, "total_normalized": len(ncs)}


_CL_EVIDENCE_DIR = os.path.join("uploads", "evidence")


@router.post("/{group_id}/upload-evidence")
async def upload_evidence_to_group(
    group_id: int,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    evidence_type: Optional[str] = Form(None),
    collection_date: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Upload evidence ONCE from the Control Library and fan it out to every
    control in the domain group — direct members plus the framework/parsed
    controls behind the group's normalized controls. One upload, linked
    everywhere, so the same evidence satisfies all those framework controls.
    """
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    group = db.query(CommonControlGroup).filter(CommonControlGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Control group not found")

    # Resolve every control this evidence should cover.
    mappings = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id
    ).all()
    framework_ids = {m.framework_control_id for m in mappings if m.framework_control_id}
    parsed_ids = {m.parsed_control_id for m in mappings if m.parsed_control_id}
    normalized_ids = {m.normalized_control_id for m in mappings if m.normalized_control_id}
    # Expand normalized controls -> their underlying framework/parsed controls,
    # so evidence on a normalized control reaches the real framework controls.
    if normalized_ids:
        for ln in db.query(NormalizedControlLink).filter(
            NormalizedControlLink.normalized_control_id.in_(normalized_ids)
        ).all():
            if ln.framework_control_id:
                framework_ids.add(ln.framework_control_id)
            if ln.parsed_control_id:
                parsed_ids.add(ln.parsed_control_id)

    # Persist the file (same layout as the Evidence Library upload).
    os.makedirs(_CL_EVIDENCE_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename)[1] if file.filename else ""
    fpath = os.path.join(_CL_EVIDENCE_DIR, f"{uuid.uuid4().hex}{ext}")
    contents = await file.read()
    with open(fpath, "wb") as fh:
        fh.write(contents)

    parsed_date = None
    if collection_date:
        try:
            parsed_date = datetime.fromisoformat(collection_date.replace("Z", "+00:00"))
        except ValueError:
            parsed_date = None

    ev = Evidence(
        tenant_id=tenant_id,
        name=name,
        description=description or f"Uploaded from control group: {group.name}",
        file_path=fpath,
        file_name=file.filename,
        file_type=file.content_type,
        uploaded_by=current_user.id,
        status="draft",
        evidence_type=evidence_type or None,
        collection_date=parsed_date,
        source_system=f"Control Library · {group.name}",
    )
    db.add(ev)
    db.flush()

    created = 0

    def _add_map(**kw):
        nonlocal created
        exists = db.query(EvidenceControlMapping).filter_by(evidence_id=ev.id, **kw).first()
        if not exists:
            db.add(EvidenceControlMapping(
                evidence_id=ev.id, framework_name=group.name,
                coverage_type="supporting", **kw,
            ))
            created += 1

    for fcid in framework_ids:
        _add_map(framework_control_id=fcid)
    for pcid in parsed_ids:
        _add_map(parsed_control_id=pcid)
    for ncid in normalized_ids:
        _add_map(normalized_control_id=ncid)

    db.commit()
    db.refresh(ev)
    return {
        "evidence_id": ev.id,
        "name": ev.name,
        "group_id": group_id,
        "group_name": group.name,
        "linked_controls": created,
        "breakdown": {
            "framework": len(framework_ids),
            "parsed": len(parsed_ids),
            "normalized": len(normalized_ids),
        },
        "message": (
            f"Evidence uploaded and linked to {created} control(s) across the "
            f"“{group.name}” domain."
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Per-control hub: description, recommended evidence, AI-drafted artifacts, and
# pipeline-backed evidence upload — all fanning out across the frameworks the
# control consolidates. Powers the expandable rows in the Mapped Controls table.
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_control(db: Session, ctrl_type: str, ctrl_id: int):
    """Return {name, code, text, fanout} for a control. `fanout` is the list of
    (kind, id) tuples evidence should link to: the control itself plus — for a
    normalized control — every framework/parsed control it consolidates."""
    ctrl_type = (ctrl_type or "").lower()
    fanout: List[tuple] = []
    if ctrl_type == "normalized":
        nc = db.query(NormalizedControl).filter(NormalizedControl.id == ctrl_id).first()
        if not nc:
            return None
        fanout.append(("normalized", nc.id))
        for ln in db.query(NormalizedControlLink).filter(
            NormalizedControlLink.normalized_control_id == nc.id).all():
            if ln.framework_control_id:
                fanout.append(("framework", ln.framework_control_id))
            if ln.parsed_control_id:
                fanout.append(("parsed", ln.parsed_control_id))
        return {"name": nc.name, "code": nc.code, "text": nc.statement or nc.objective or nc.name, "fanout": fanout}
    if ctrl_type == "framework":
        fc = db.query(FrameworkControl).filter(FrameworkControl.id == ctrl_id).first()
        if not fc:
            return None
        return {"name": fc.name, "code": getattr(fc, "control_id", "") or "",
                "text": fc.statement or getattr(fc, "control_objective", "") or fc.name,
                "fanout": [("framework", fc.id)]}
    if ctrl_type == "parsed":
        pc = db.query(ParsedFrameworkControl).filter(ParsedFrameworkControl.id == ctrl_id).first()
        if not pc:
            return None
        return {"name": pc.title or "", "code": pc.original_reference or pc.control_id or "",
                "text": pc.description or pc.full_text or pc.title or "",
                "fanout": [("parsed", pc.id)]}
    return None


def _fanout_evidence(db: Session, evidence_id: int, fanout: List[tuple], label: str) -> int:
    created = 0
    for kind, cid in fanout:
        kw = ({"framework_control_id": cid} if kind == "framework"
              else {"parsed_control_id": cid} if kind == "parsed"
              else {"normalized_control_id": cid})
        if not db.query(EvidenceControlMapping).filter_by(evidence_id=evidence_id, **kw).first():
            db.add(EvidenceControlMapping(
                evidence_id=evidence_id, framework_name=label, coverage_type="supporting", **kw))
            created += 1
    return created


def _ai_unavailable():
    raise HTTPException(status_code=503, detail={
        "error": "AI features unavailable", "message": "OpenAI API key is not configured."})


def _framework_label_maps(db: Session, parsed_ids: List[int], framework_ids: List[int]):
    """Resolve parsed/framework control ids -> {framework, code, name} so we can
    show, for a normalized control, exactly which framework each member control
    comes from. Bulk-loaded to avoid N+1 queries."""
    p_map: Dict[int, dict] = {}
    f_map: Dict[int, dict] = {}
    if parsed_ids:
        parsed = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.id.in_(set(parsed_ids))).all()
        ufids = list({p.uploaded_framework_id for p in parsed if p.uploaded_framework_id})
        names: Dict[int, str] = {}
        if ufids:
            for fw in db.query(UploadedFramework).filter(UploadedFramework.id.in_(ufids)).all():
                names[fw.id] = fw.name
        for p in parsed:
            p_map[p.id] = {
                "framework": names.get(p.uploaded_framework_id, "Unknown framework"),
                "code": p.original_reference or p.control_id or "",
                "name": p.title or "",
                "text": (p.description or p.full_text or "").strip(),
            }
    if framework_ids:
        fcs = db.query(FrameworkControl).filter(
            FrameworkControl.id.in_(set(framework_ids))).all()
        obj_ids = list({fc.objective_id for fc in fcs if fc.objective_id})
        obj_fw: Dict[int, str] = {}
        if obj_ids:
            objs = db.query(ControlObjective).filter(ControlObjective.id.in_(obj_ids)).all()
            dom_ids = list({o.domain_id for o in objs if o.domain_id})
            dom_fw: Dict[int, str] = {}
            if dom_ids:
                doms = db.query(FrameworkDomain).filter(FrameworkDomain.id.in_(dom_ids)).all()
                fwids = list({d.framework_id for d in doms if d.framework_id})
                fwname: Dict[int, str] = {}
                if fwids:
                    for fw in db.query(Framework).filter(Framework.id.in_(fwids)).all():
                        fwname[fw.id] = fw.name
                for d in doms:
                    dom_fw[d.id] = fwname.get(d.framework_id, "Framework")
            for o in objs:
                obj_fw[o.id] = dom_fw.get(o.domain_id, "Framework")
        for fc in fcs:
            f_map[fc.id] = {
                "framework": obj_fw.get(fc.objective_id, "Framework"),
                "code": fc.code or "",
                "name": fc.name or "",
                "text": (fc.statement or getattr(fc, "control_objective", "") or "").strip(),
            }
    return p_map, f_map


@router.get("/control/{ctrl_type}/{ctrl_id}/coverage")
def control_coverage(
    ctrl_type: str, ctrl_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Cross-framework coverage for a control: WHICH frameworks (and which exact
    framework controls) this normalized control consolidates, plus any evidence
    already linked to any of them. This is how a user verifies that one upload
    here satisfies the control everywhere it exists."""
    info = _resolve_control(db, ctrl_type, ctrl_id)
    if not info:
        raise HTTPException(status_code=404, detail="Control not found")

    parsed_ids = [cid for kind, cid in info["fanout"] if kind == "parsed"]
    framework_ids = [cid for kind, cid in info["fanout"] if kind == "framework"]
    p_map, f_map = _framework_label_maps(db, parsed_ids, framework_ids)

    # Group the member controls under their framework. Each carries its own
    # type+id so the UI can open the real control when its code tag is clicked.
    fw_groups: Dict[str, List[dict]] = {}
    for cid in parsed_ids:
        d = p_map.get(cid)
        if d:
            fw_groups.setdefault(d["framework"], []).append(
                {"code": d["code"], "name": d["name"], "type": "parsed",
                 "control_id": cid, "statement": d.get("text", "")})
    for cid in framework_ids:
        d = f_map.get(cid)
        if d:
            fw_groups.setdefault(d["framework"], []).append(
                {"code": d["code"], "name": d["name"], "type": "framework",
                 "control_id": cid, "statement": d.get("text", "")})
    frameworks = [
        {"framework_name": k, "count": len(v),
         "controls": sorted(v, key=lambda c: c["code"])}
        for k, v in sorted(fw_groups.items())
    ]

    # Evidence already linked to ANY control in the fan-out (proof of linkage).
    ev_filters = []
    if parsed_ids:
        ev_filters.append(EvidenceControlMapping.parsed_control_id.in_(parsed_ids))
    if framework_ids:
        ev_filters.append(EvidenceControlMapping.framework_control_id.in_(framework_ids))
    if (ctrl_type or "").lower() == "normalized":
        ev_filters.append(EvidenceControlMapping.normalized_control_id == ctrl_id)
    evidence = []
    if ev_filters:
        ev_ids = {m.evidence_id for m in db.query(EvidenceControlMapping).filter(or_(*ev_filters)).all()}
        if ev_ids:
            for ev in db.query(Evidence).filter(Evidence.id.in_(ev_ids)).order_by(Evidence.id.desc()).all():
                evidence.append({
                    "id": ev.id, "name": ev.name, "status": ev.status,
                    "file_name": ev.file_name,
                })

    return {
        "control": {"code": info["code"], "name": info["name"], "type": (ctrl_type or "").lower()},
        "framework_count": len(frameworks),
        "linked_control_count": len(parsed_ids) + len(framework_ids),
        "frameworks": frameworks,
        "evidence": evidence,
    }


def _normalize_evidence_reqs(raw) -> List[dict]:
    """evidence_requirements is stored as JSON — usually a list of
    {name, description}, sometimes plain strings. Normalise to dicts."""
    out: List[dict] = []
    if not raw:
        return out
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except (json.JSONDecodeError, TypeError):
            return [{"name": raw, "description": ""}]
    if isinstance(raw, dict):
        raw = [raw]
    if not isinstance(raw, list):
        return out
    for item in raw:
        if isinstance(item, dict):
            nm = (item.get("name") or item.get("title") or "").strip()
            ds = (item.get("description") or item.get("detail") or "").strip()
            if nm or ds:
                out.append({"name": nm or ds[:60], "description": ds})
        elif isinstance(item, str) and item.strip():
            out.append({"name": item.strip(), "description": ""})
    return out


# The artifact catalog keys on CANONICAL framework names + control refs, while
# uploaded frameworks use verbose names + their own ref formats. Map verbose
# uploaded names onto the catalog name by keyword so artifacts still resolve.
_CATALOG_ALIASES = {
    "CIS Controls v8": ["cis"], "COBIT 2019": ["cobit"],
    "DORA": ["dora", "operational resilience"],
    "GDPR": ["gdpr", "general data protection"], "HIPAA": ["hipaa"],
    "ISO 22301:2019": ["22301"], "ISO 41001:2018": ["41001"],
    "ISO/IEC 27001:2022": ["27001"], "NIS 2": ["nis2", "nis 2"],
    "NIST CSF 2.0": ["cybersecurity framework"], "PCI DSS v4.0.1": ["pci"],
    "Saudi PDPL": ["pdpl", "personal data protection law"],
    "SOC 2": ["soc 2", "soc2"], "SOX ITGC": ["sox"], "SWIFT CSCF": ["swift"],
}


def _catalog_framework_for(name: str) -> Optional[str]:
    n = (name or "").lower()
    for cat, kws in _CATALOG_ALIASES.items():
        if any(k in n for k in kws):
            return cat
    return None


def _ref_variants(code: str) -> set:
    """Normalised forms of a control ref so it matches the catalog's ref format
    across frameworks (PCI 'Req ' prefix, CIS parent 'Control N', etc.)."""
    c = (code or "").strip()
    out = {c.lower()}
    if not c:
        return out
    out.add(("req " + c).lower())                 # PCI: 1.2.3 -> Req 1.2.3
    out.add(c.lower().replace("req ", "").strip())
    m = _re_mod.search(r"control\s+(\d+)", c, _re_mod.I)  # CIS safeguard -> parent control
    if m:
        out.add(f"control {m.group(1)}")
    return {v for v in out if v}


_EV_FILLER = {
    "document", "documents", "documented", "approved", "the", "a", "an", "of",
    "for", "and", "to", "current", "version", "record", "records", "report",
    "reports", "evidence", "copy", "latest", "signed", "formal", "written",
    "process", "procedure", "procedures", "policy", "policies",
}


def _ev_key(name: str) -> str:
    """Canonical key for an evidence item so the SAME evidence phrased differently
    across frameworks collapses to one entry (word-order & filler insensitive)."""
    words = [w for w in _re_mod.findall(r"[a-z0-9]+", (name or "").lower())
             if w not in _EV_FILLER and len(w) > 1]
    return " ".join(sorted(set(words)))


def _consolidate_evidence(items: List[dict]) -> List[dict]:
    """Merge evidence items that are the same requirement across frameworks into a
    single recommended item that lists every framework it satisfies — so the user
    uploads ONE thing, not the same thing once per framework."""
    groups: Dict[str, dict] = {}
    for it in items:
        key = _ev_key(it.get("name", "")) or (it.get("name", "").lower().strip())
        if not key:
            continue
        g = groups.get(key)
        if not g:
            g = {"name": it["name"], "description": it.get("description", ""),
                 "frameworks": [], "_seen": set()}
            groups[key] = g
        # Prefer the shortest clean name as the representative label.
        if len(it["name"]) < len(g["name"]):
            g["name"] = it["name"]
        if len(it.get("description", "")) > len(g["description"]):
            g["description"] = it.get("description", "")
        fwkey = (it.get("framework", ""), it.get("code", ""))
        if fwkey not in g["_seen"]:
            g["_seen"].add(fwkey)
            g["frameworks"].append({"framework": it.get("framework", ""), "code": it.get("code", "")})
    out = []
    for g in groups.values():
        g.pop("_seen", None)
        g["framework_count"] = len(g["frameworks"])
        out.append(g)
    out.sort(key=lambda g: -g["framework_count"])
    return out


def _ai_consolidate_evidence(items: List[dict]) -> Optional[List[dict]]:
    """Merge evidence items by MEANING using the AI: items that are the same
    piece of evidence across frameworks (worded differently) become one item
    listing every framework it satisfies. Returns None if AI is unavailable."""
    if not items:
        return []
    if not check_ai_available():
        return None
    client = get_openai_client()
    lines = [f"[{i}] ({it.get('framework','')} {it.get('code','')}) {it.get('name','')}"
             + (f" — {it.get('description','')[:160]}" if it.get('description') else "")
             for i, it in enumerate(items)]
    prompt = (
        "Below are recommended-evidence items drawn from several frameworks for the "
        "SAME control. Merge items that are the SAME piece of evidence (even if "
        "worded differently — e.g. 'Approved HR Policy document' and 'HR security "
        "policy') into ONE. Keep genuinely different evidence separate. For each "
        "merged item give a clear name, a one-line description, and the list of the "
        "input indices [n] it covers.\n\n" + "\n".join(lines) +
        '\n\nRespond ONLY JSON: {"evidence":[{"name":"...","description":"...","members":[0,3]}]}')
    try:
        resp = client.chat.completions.create(
            model=get_openai_model(),
            messages=[{"role": "system", "content": "You consolidate compliance evidence requirements."},
                      {"role": "user", "content": prompt}],
            response_format={"type": "json_object"}, temperature=0.1)
        data = json.loads(resp.choices[0].message.content or "{}")
    except Exception:
        logger.exception("AI evidence consolidation failed")
        return None
    out: List[dict] = []
    for e in data.get("evidence", []) or []:
        idxs = []
        for x in e.get("members", []) or []:
            try:
                idxs.append(int(x))
            except (TypeError, ValueError):
                continue
        seen = set()
        frameworks = []
        for i in idxs:
            if 0 <= i < len(items):
                key = (items[i].get("framework", ""), items[i].get("code", ""))
                if key not in seen:
                    seen.add(key)
                    frameworks.append({"framework": items[i].get("framework", ""), "code": items[i].get("code", "")})
        name = (e.get("name") or "").strip()
        if not name:
            continue
        out.append({"name": name[:255], "description": (e.get("description") or "").strip(),
                    "frameworks": frameworks, "framework_count": len(frameworks)})
    out.sort(key=lambda g: -g["framework_count"])
    return out


@router.get("/control/{ctrl_type}/{ctrl_id}/requirements")
def control_requirements(
    ctrl_type: str, ctrl_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Surface the framework's OWN built-in recommended evidence
    (parsed_control.evidence_requirements) and pre-built artifacts
    (grc_artifact_catalog_items, matched by framework name + control ref) for a
    control — no AI generation. For a normalized control, aggregate across every
    framework control it consolidates, grouped by framework so the user sees
    exactly what each framework already prescribes."""
    from sqlalchemy import text as _text
    info = _resolve_control(db, ctrl_type, ctrl_id)
    if not info:
        raise HTTPException(status_code=404, detail="Control not found")

    parsed_ids = [cid for kind, cid in info["fanout"] if kind == "parsed"]
    rows = (db.query(ParsedFrameworkControl, UploadedFramework.name)
            .join(UploadedFramework, ParsedFrameworkControl.uploaded_framework_id == UploadedFramework.id)
            .filter(ParsedFrameworkControl.id.in_(parsed_ids)).all()) if parsed_ids else []

    evidence_groups: List[dict] = []
    artifact_groups: List[dict] = []
    all_evidence_items: List[dict] = []
    all_artifact_items: List[dict] = []
    total_evidence = 0
    total_artifacts = 0
    for pc, fw_name in rows:
        code = pc.original_reference or pc.control_id or ""
        ev = _normalize_evidence_reqs(pc.evidence_requirements)
        for e in ev:
            all_evidence_items.append({
                "name": e.get("name", ""), "description": e.get("description", ""),
                "framework": fw_name, "code": code})
        # Pre-built artifacts: resolve the catalog framework by alias, then match
        # the control ref against its normalised variants (case-insensitive).
        arts = []
        cat_fw = _catalog_framework_for(fw_name)
        if cat_fw:
            try:
                variants = list(_ref_variants(code))
                res = db.execute(_text(
                    "SELECT name, artifact_type, mandatory, description, format, owner "
                    "FROM grc_artifact_catalog_items "
                    "WHERE framework_name = :fw AND lower(control_ref) = ANY(:refs)"),
                    {"fw": cat_fw, "refs": variants})
                seen_art = set()
                for r in res:
                    if r[0] in seen_art:
                        continue
                    seen_art.add(r[0])
                    arts.append({
                        "name": r[0], "artifact_type": r[1], "mandatory": bool(r[2]),
                        "description": r[3], "format": r[4], "owner": r[5],
                    })
            except Exception:
                logger.exception("artifact lookup failed for %s/%s", cat_fw, code)
        if ev:
            evidence_groups.append({"framework": fw_name, "code": code, "items": ev})
            total_evidence += len(ev)
        if arts:
            artifact_groups.append({"framework": fw_name, "code": code, "items": arts})
            total_artifacts += len(arts)
            for a in arts:
                all_artifact_items.append({
                    "name": a.get("name", ""), "description": a.get("description", ""),
                    "framework": fw_name, "code": code,
                    "artifact_type": a.get("artifact_type"), "mandatory": a.get("mandatory"),
                })

    # Evidence normalization: for a unified control, merge the frameworks'
    # evidence by MEANING using the AI (cached on the control so it's done once).
    # Fall back to fast string-consolidation if AI is unavailable.
    is_norm = (ctrl_type or "").lower() == "normalized"
    consolidated: List[dict] = []
    if all_evidence_items:
        if is_norm:
            nc = db.query(NormalizedControl).filter(NormalizedControl.id == ctrl_id).first()
            cached = getattr(nc, "recommended_evidence", None) if nc else None
            if cached:
                consolidated = cached
            else:
                ai = _ai_consolidate_evidence(all_evidence_items)
                consolidated = ai if ai is not None else _consolidate_evidence(all_evidence_items)
                if ai is not None and nc is not None:
                    try:
                        nc.recommended_evidence = consolidated
                        db.commit()
                    except Exception:
                        db.rollback()
        else:
            consolidated = _consolidate_evidence(all_evidence_items)

    # Keep the recommended-evidence list focused: a control should surface a
    # handful of high-value, broadly-applicable evidence items — not an
    # overwhelming long tail. Rank by how many frameworks each item satisfies
    # (the most "comply once, satisfy many" first) and cap the list.
    MAX_RECOMMENDED_EVIDENCE = 6
    if consolidated and len(consolidated) > MAX_RECOMMENDED_EVIDENCE:
        def _fw_count(e):
            return e.get("framework_count") or len(e.get("frameworks") or [])
        consolidated = sorted(consolidated, key=lambda e: (-_fw_count(e), str(e.get("name") or "")))
        consolidated = consolidated[:MAX_RECOMMENDED_EVIDENCE]

    # Artifact normalization: merge the same pre-built artifact across frameworks
    # into one (lists every framework it satisfies). Deterministic by name.
    consolidated_artifacts = _consolidate_evidence(all_artifact_items) if all_artifact_items else []
    if consolidated_artifacts and len(consolidated_artifacts) > MAX_RECOMMENDED_EVIDENCE:
        consolidated_artifacts = sorted(
            consolidated_artifacts,
            key=lambda e: (-(e.get("framework_count") or len(e.get("frameworks") or [])), str(e.get("name") or "")),
        )[:MAX_RECOMMENDED_EVIDENCE]

    return {
        "control": {"code": info["code"], "name": info["name"], "type": (ctrl_type or "").lower()},
        "is_normalized": is_norm,
        "evidence_total": total_evidence,            # raw count across frameworks
        "unique_evidence_total": len(consolidated),   # after AI normalization
        "artifact_total": total_artifacts,
        "unique_artifact_total": len(consolidated_artifacts),
        "consolidated_evidence": consolidated,        # ← upload once, satisfies many
        "consolidated_artifacts": consolidated_artifacts,
        "evidence_by_control": evidence_groups,       # per-framework breakdown (detail)
        "artifacts_by_control": artifact_groups,
    }


@router.post("/control/{ctrl_type}/{ctrl_id}/upload-evidence")
async def upload_evidence_to_control(
    ctrl_type: str, ctrl_id: int,
    name: str = Form(...),
    description: Optional[str] = Form(None),
    evidence_type: Optional[str] = Form(None),
    collection_date: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Upload evidence for one control THROUGH the Evidence pipeline (file store,
    OCR, AI assessment) and fan it out to every control it covers."""
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User has no tenant assigned")
    info = _resolve_control(db, ctrl_type, ctrl_id)
    if not info:
        raise HTTPException(status_code=404, detail="Control not found")

    from ...evidence.routers.evidence import (
        EVIDENCE_UPLOAD_DIR, OCR_PROCESSABLE_TYPES, process_evidence_background)
    import threading as _th

    tdir = os.path.join(EVIDENCE_UPLOAD_DIR, str(tenant_id))
    os.makedirs(tdir, exist_ok=True)
    ext = os.path.splitext(file.filename)[1].lower() if file.filename else ""
    fpath = os.path.join(tdir, f"{uuid.uuid4().hex}{ext}")
    contents = await file.read()
    with open(fpath, "wb") as fh:
        fh.write(contents)
    ocr = "pending" if ext and ext[1:] in OCR_PROCESSABLE_TYPES else "not_applicable"
    parsed_date = None
    if collection_date:
        try:
            parsed_date = datetime.fromisoformat(collection_date.replace("Z", "+00:00"))
        except ValueError:
            parsed_date = None

    ev = Evidence(
        tenant_id=tenant_id, name=name,
        description=description or f"Uploaded for control {info['code']}",
        file_path=fpath, file_name=file.filename, file_type=file.content_type,
        evidence_type=evidence_type or None, collection_date=parsed_date,
        uploaded_by=current_user.id, status="draft", ocr_status=ocr,
        source_system=f"Control Library · {info['code']}",
    )
    db.add(ev)
    db.flush()
    created = _fanout_evidence(db, ev.id, info["fanout"], info["code"])
    db.commit()
    db.refresh(ev)
    if ocr == "pending":
        _th.Thread(target=process_evidence_background, args=(ev.id,), daemon=True).start()
    return {
        "evidence_id": ev.id, "linked_controls": created,
        "ocr_processing": ocr == "pending",
        "message": (f"Uploaded and linked to {created} control(s)."
                    + (" Running OCR + AI assessment…" if ocr == "pending" else "")),
    }


@router.post("/control/{ctrl_type}/{ctrl_id}/recommend-evidence")
def recommend_evidence_for_control(
    ctrl_type: str, ctrl_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """AI-suggest the evidence an auditor would accept for this control."""
    tenant_id = get_user_primary_tenant(current_user, db)
    info = _resolve_control(db, ctrl_type, ctrl_id)
    if not info:
        raise HTTPException(status_code=404, detail="Control not found")
    if not check_ai_available():
        _ai_unavailable()
    client = get_openai_client()
    prompt = (
        f"Control: {info['name']}\n{(info['text'] or '')[:1500]}\n\n"
        "Suggest 3-5 concrete pieces of EVIDENCE an auditor would accept to prove this "
        "control is implemented. For each give: evidence_type (short label), description "
        "(what the artefact is), priority (high|medium|low).\n"
        'JSON only: {"recommendations":[{"evidence_type":"...","description":"...","priority":"..."}]}'
    )
    resp = client.chat.completions.create(
        model=get_openai_model(),
        messages=[{"role": "system", "content": "You are a meticulous compliance auditor."},
                  {"role": "user", "content": prompt}],
        response_format={"type": "json_object"}, temperature=0.3)
    try:
        recs = (json.loads(resp.choices[0].message.content or "{}").get("recommendations") or [])[:6]
    except json.JSONDecodeError:
        recs = []
    kw = ({"normalized_control_id": ctrl_id} if ctrl_type == "normalized"
          else {"framework_control_id": ctrl_id} if ctrl_type == "framework"
          else {"parsed_control_id": ctrl_id})
    db.query(AIEvidenceRecommendation).filter_by(tenant_id=tenant_id, **kw).delete()
    out = []
    for r in recs:
        rec = AIEvidenceRecommendation(
            tenant_id=tenant_id,
            evidence_type=(r.get("evidence_type") or "document")[:100],
            evidence_description=r.get("description"),
            priority=(r.get("priority") or "medium")[:20], ai_confidence=0.8, **kw)
        db.add(rec)
        out.append({"evidence_type": rec.evidence_type, "description": rec.evidence_description, "priority": rec.priority})
    db.commit()
    return {"recommendations": out}


@router.post("/control/{ctrl_type}/{ctrl_id}/draft-document")
def draft_document_for_control(
    ctrl_type: str, ctrl_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """AI-draft a policy/procedure document (artifact) for this control and
    register it as evidence, fanned out across the frameworks it covers."""
    tenant_id = get_user_primary_tenant(current_user, db)
    info = _resolve_control(db, ctrl_type, ctrl_id)
    if not info:
        raise HTTPException(status_code=404, detail="Control not found")
    if not check_ai_available():
        _ai_unavailable()
    client = get_openai_client()
    prompt = (
        f"Draft a concise, professional policy/procedure document that IMPLEMENTS this control.\n\n"
        f"Control: {info['name']}\n{(info['text'] or '')[:1500]}\n\n"
        "Include sections: Purpose, Scope, Policy Statements, Responsibilities, Review & "
        "Monitoring. Keep it practical and audit-ready. Return Markdown only."
    )
    resp = client.chat.completions.create(
        model=get_openai_model(),
        messages=[{"role": "system", "content": "You are a senior GRC policy author."},
                  {"role": "user", "content": prompt}],
        temperature=0.4)
    import re as _re
    doc_text = (resp.choices[0].message.content or "").strip()
    # The model often wraps the whole document in a ```markdown ... ``` fence —
    # strip it so we store/return clean Markdown (no junk symbols).
    if doc_text.startswith("```"):
        doc_text = _re.sub(r"^```[a-zA-Z]*\s*\n", "", doc_text)
        doc_text = _re.sub(r"\n```\s*$", "", doc_text).strip()

    from ...evidence.routers.evidence import EVIDENCE_UPLOAD_DIR
    tdir = os.path.join(EVIDENCE_UPLOAD_DIR, str(tenant_id))
    os.makedirs(tdir, exist_ok=True)
    fname = f"{(info['code'] or 'control').replace('/', '_')}-policy.md"
    fpath = os.path.join(tdir, f"{uuid.uuid4().hex}.md")
    with open(fpath, "w", encoding="utf-8") as fh:
        fh.write(doc_text)

    ev = Evidence(
        tenant_id=tenant_id,
        name=f"AI Draft: {(info['name'] or info['code'])[:110]} Policy",
        description=f"AI-drafted document artifact for control {info['code']}.",
        file_path=fpath, file_name=fname, file_type="text/markdown",
        evidence_type="policy", uploaded_by=current_user.id, status="draft",
        ocr_status="not_applicable", source_system=f"Control Library · AI Draft · {info['code']}",
    )
    db.add(ev)
    db.flush()
    created = _fanout_evidence(db, ev.id, info["fanout"], info["code"])
    db.commit()
    db.refresh(ev)
    return {
        "evidence_id": ev.id, "name": ev.name, "linked_controls": created,
        "document": doc_text,
        "message": f"Drafted a policy document and linked it as evidence to {created} control(s).",
    }


@router.post("/auto-group")
def auto_group_controls(
    request: AutoGroupRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User has no tenant assigned"
        )
    
    if not check_ai_available():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail={
                "error": "AI features unavailable",
                "message": "OpenAI API key is not configured. Please add OPENAI_API_KEY to enable AI features, or use manual grouping instead.",
                "fallback_available": True,
                "fallback_suggestion": "Use the manual 'Create Group' feature to organize controls"
            }
        )
    
    controls_list = _fetch_controls_for_grouping(db, tenant_id, request.framework_ids)
    if len(controls_list) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Not enough controls to perform auto-grouping"
        )
    
    ai_groups = ai_auto_group_controls(controls_list)
    summary = persist_ai_groups(db, tenant_id, current_user.id, ai_groups)

    all_group_ids = (summary.get("created_group_ids") or []) + (summary.get("merged_group_ids") or [])
    all_groups = db.query(CommonControlGroup).filter(CommonControlGroup.id.in_(all_group_ids)).all() if all_group_ids else []
    return {
        "message": (
            f"Created {summary['created_count']} control groups, "
            f"merged into {summary['merged_count']} existing groups"
        ),
        "groups": [serialize_group(g, db, include_controls=True) for g in all_groups],
    }


@router.get("/{group_id}/frameworks")
def get_group_frameworks(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    fc_mappings = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id,
        CommonControlGroupMapping.framework_control_id.isnot(None)
    ).all()
    
    framework_counts = {}
    for mapping in fc_mappings:
        fc = db.query(FrameworkControl).options(
            joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
        ).filter(FrameworkControl.id == mapping.framework_control_id).first()
        
        if fc and fc.objective and fc.objective.domain and fc.objective.domain.framework:
            fw = fc.objective.domain.framework
            fw_key = f"legacy_{fw.id}"
            if fw_key not in framework_counts:
                framework_counts[fw_key] = {
                    "framework_id": fw.id,
                    "framework_name": fw.name,
                    "framework_code": fw.short_code,
                    "control_count": 0,
                    "source": "legacy"
                }
            framework_counts[fw_key]["control_count"] += 1
    
    pc_mappings = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id,
        CommonControlGroupMapping.parsed_control_id.isnot(None)
    ).all()
    
    for mapping in pc_mappings:
        pc = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.id == mapping.parsed_control_id
        ).first()
        if pc:
            fw = db.query(UploadedFramework).filter(
                UploadedFramework.id == pc.uploaded_framework_id
            ).first()
            if fw:
                fw_key = f"uploaded_{fw.id}"
                if fw_key not in framework_counts:
                    framework_counts[fw_key] = {
                        "framework_id": fw.id,
                        "framework_name": fw.name,
                        "framework_code": fw.name[:10] if fw.name else None,
                        "control_count": 0,
                        "source": "uploaded"
                    }
                framework_counts[fw_key]["control_count"] += 1
    
    normalized_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id,
        CommonControlGroupMapping.normalized_control_id.isnot(None)
    ).count()
    
    parsed_control_count = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id,
        CommonControlGroupMapping.parsed_control_id.isnot(None)
    ).count()
    
    return {
        "group_id": group_id,
        "group_name": group.name,
        "normalized_control_count": normalized_count,
        "parsed_control_count": parsed_control_count,
        "frameworks": list(framework_counts.values())
    }


@router.post("/{group_id}/generate-summary")
def generate_summary(
    group_id: int,
    request: GenerateSummaryRequest = GenerateSummaryRequest(),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if not check_ai_available():
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail={
                "error": "AI features unavailable",
                "message": "OpenAI API key is not configured. Please add OPENAI_API_KEY to enable AI-generated summaries.",
                "fallback_available": False
            }
        )
    
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    mappings = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id
    ).all()
    
    controls_text_parts = []
    for mapping in mappings:
        if mapping.normalized_control_id:
            nc = db.query(NormalizedControl).filter(
                NormalizedControl.id == mapping.normalized_control_id
            ).first()
            if nc:
                controls_text_parts.append(f"- {nc.code}: {nc.name}\n  {nc.statement or ''}")
        
        if mapping.framework_control_id:
            fc = db.query(FrameworkControl).filter(
                FrameworkControl.id == mapping.framework_control_id
            ).first()
            if fc:
                controls_text_parts.append(f"- {fc.code}: {fc.name}\n  {fc.statement or ''}")
        
        if mapping.parsed_control_id:
            pc = db.query(ParsedFrameworkControl).filter(
                ParsedFrameworkControl.id == mapping.parsed_control_id
            ).first()
            if pc:
                controls_text_parts.append(f"- {pc.original_reference}: {pc.title}\n  {pc.description or ''}")
    
    if not controls_text_parts:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Group has no controls to summarize"
        )
    
    controls_text = f"Group: {group.name}\nDescription: {group.description or 'N/A'}\n\nControls:\n" + "\n".join(controls_text_parts)
    
    result = generate_group_summary(controls_text)
    
    group.ai_summary = result.get("summary")
    if request.regenerate_keywords:
        group.keywords = result.get("keywords", group.keywords or [])
    if result.get("evidence_types"):
        group.evidence_types = result.get("evidence_types")
    
    group.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(group)
    
    return serialize_group(group, db, include_controls=True)


@router.post("/{group_id}/populate-from-frameworks")
def populate_group_from_frameworks(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    tenant_id = get_user_primary_tenant(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Group not found")
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    
    group_text = f"{group.name} {group.description or ''} {group.category or ''} {group.domain or ''} {' '.join(group.keywords or [])}"
    group_kw = _extract_keywords(group_text)
    if not group_kw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Group has no keywords for matching")
    
    uploaded_fws = db.query(UploadedFramework).filter(
        or_(
            UploadedFramework.tenant_id == tenant_id,
            UploadedFramework.tenant_id.is_(None)
        )
    ).all()
    
    existing_parsed_ids = set(
        m.parsed_control_id for m in db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group_id,
            CommonControlGroupMapping.parsed_control_id.isnot(None)
        ).all()
    )
    
    added = 0
    for fw in uploaded_fws:
        controls = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.uploaded_framework_id == fw.id
        ).all()
        
        for pc in controls:
            if pc.id in existing_parsed_ids:
                continue
            ctrl_text = f"{pc.title or ''} {pc.description or ''} {pc.category or ''} {pc.domain or ''}"
            ctrl_kw = _extract_keywords(ctrl_text)
            score = _keyword_score(group_kw, ctrl_kw)
            if score >= 0.15:
                mapping = CommonControlGroupMapping(
                    group_id=group_id,
                    parsed_control_id=pc.id,
                    mapping_source="auto",
                    mapping_confidence=round(score, 3)
                )
                db.add(mapping)
                existing_parsed_ids.add(pc.id)
                added += 1
    
    db.commit()
    return {
        "message": f"Added {added} controls from {len(uploaded_fws)} frameworks",
        "added": added,
        "group": serialize_group(group, db, include_controls=False)
    }


@router.post("/populate-all-groups")
def populate_all_groups_from_frameworks(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    user_tenants = get_user_tenants(current_user, db)

    groups = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        )
    ).all()

    # If no groups exist yet, the operator clicked "Populate from
    # Frameworks" without first creating any groups via AI Auto-Grouping.
    # The previous behaviour silently returned "Added 0 controls from N
    # frameworks" and looked broken. Return a clear actionable message
    # instead so the UI can surface "Run AI Auto-Grouping first" instead
    # of an empty success toast.
    if not groups:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "no_groups_to_populate",
                "message": ("No control groups exist yet. Click 'AI Auto-Grouping' "
                            "first to let the AI create groups from the frameworks "
                            "you've uploaded, then come back here to populate."),
                "fix": "ai_auto_group",
            },
        )

    uploaded_fws = db.query(UploadedFramework).filter(
        or_(
            UploadedFramework.tenant_id == tenant_id,
            UploadedFramework.tenant_id.is_(None)
        )
    ).all()
    
    all_parsed = []
    for fw in uploaded_fws:
        controls = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.uploaded_framework_id == fw.id
        ).all()
        for pc in controls:
            ctrl_text = f"{pc.title or ''} {pc.description or ''} {pc.category or ''} {pc.domain or ''}"
            all_parsed.append({
                "id": pc.id,
                "keywords": _extract_keywords(ctrl_text),
                "fw_name": fw.name
            })
    
    total_added = 0
    for group in groups:
        group_text = f"{group.name} {group.description or ''} {group.category or ''} {group.domain or ''} {' '.join(group.keywords or [])}"
        group_kw = _extract_keywords(group_text)
        if not group_kw:
            continue
        
        existing_ids = set(
            m.parsed_control_id for m in db.query(CommonControlGroupMapping).filter(
                CommonControlGroupMapping.group_id == group.id,
                CommonControlGroupMapping.parsed_control_id.isnot(None)
            ).all()
        )
        
        for pc in all_parsed:
            if pc["id"] in existing_ids:
                continue
            score = _keyword_score(group_kw, pc["keywords"])
            if score >= 0.15:
                mapping = CommonControlGroupMapping(
                    group_id=group.id,
                    parsed_control_id=pc["id"],
                    mapping_source="auto",
                    mapping_confidence=round(score, 3)
                )
                db.add(mapping)
                existing_ids.add(pc["id"])
                total_added += 1
    
    db.commit()
    return {
        "message": f"Added {total_added} controls across {len(groups)} groups from {len(uploaded_fws)} frameworks",
        "total_added": total_added,
        "groups_processed": len(groups)
    }


def _extract_keywords(text):
    if not text:
        return set()
    stop_words = {'the', 'and', 'for', 'of', 'to', 'in', 'a', 'an', 'is', 'are', 'be', 'with', 'that', 'this', 'shall', 'must', 'should', 'or', 'its', 'by', 'on', 'as', 'from', 'all', 'has', 'have', 'not', 'at'}
    words = set(w.lower().strip('.,;:()[]') for w in text.split() if len(w) > 2)
    return words - stop_words


def _keyword_score(keywords1, keywords2):
    if not keywords1 or not keywords2:
        return 0.0
    intersection = keywords1 & keywords2
    union = keywords1 | keywords2
    if not union:
        return 0.0
    return len(intersection) / len(union)


@router.get("/{group_id}/similarities")
def get_group_similarities(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    group = db.query(CommonControlGroup).filter(
        CommonControlGroup.id == group_id
    ).first()
    
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Group not found"
        )
    
    if group.tenant_id is not None and group.tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this group"
        )
    
    all_mappings = db.query(CommonControlGroupMapping).filter(
        CommonControlGroupMapping.group_id == group_id
    ).all()
    
    controls_by_framework = {}
    
    for mapping in all_mappings:
        if mapping.parsed_control_id:
            pc = db.query(ParsedFrameworkControl).filter(
                ParsedFrameworkControl.id == mapping.parsed_control_id
            ).first()
            if not pc:
                continue
            fw = db.query(UploadedFramework).filter(
                UploadedFramework.id == pc.uploaded_framework_id
            ).first()
            fw_key = f"uploaded_{fw.id}" if fw else "uploaded_0"
            fw_name = fw.name if fw else "Unknown"
            if fw_key not in controls_by_framework:
                controls_by_framework[fw_key] = {"name": fw_name, "controls": []}
            text = (pc.title or "") + " " + (pc.description or "") + " " + (pc.full_text[:500] if pc.full_text else "")
            controls_by_framework[fw_key]["controls"].append({
                "id": pc.id,
                "code": pc.original_reference,
                "name": pc.title,
                "keywords": _extract_keywords(text),
                "framework_name": fw_name,
                "type": "parsed"
            })
        
        elif mapping.framework_control_id:
            fc = db.query(FrameworkControl).options(
                joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
            ).filter(FrameworkControl.id == mapping.framework_control_id).first()
            if not fc:
                continue
            fw = fc.objective.domain.framework if fc.objective and fc.objective.domain else None
            fw_key = f"legacy_{fw.id}" if fw else "legacy_0"
            fw_name = fw.name if fw else "Unknown"
            if fw_key not in controls_by_framework:
                controls_by_framework[fw_key] = {"name": fw_name, "controls": []}
            text = (fc.name or "") + " " + (fc.statement or "") + " " + (fc.control_objective or "")
            controls_by_framework[fw_key]["controls"].append({
                "id": fc.id,
                "code": fc.code,
                "name": fc.name,
                "keywords": _extract_keywords(text),
                "framework_name": fw_name,
                "type": "framework"
            })
    
    fw_ids = list(controls_by_framework.keys())
    pairs = []
    
    for i in range(len(fw_ids)):
        for j in range(i + 1, len(fw_ids)):
            fw1_controls = controls_by_framework[fw_ids[i]]["controls"]
            fw2_controls = controls_by_framework[fw_ids[j]]["controls"]
            
            for c1 in fw1_controls:
                for c2 in fw2_controls:
                    score = _keyword_score(c1["keywords"], c2["keywords"])
                    if score > 0.1:
                        common_kw = c1["keywords"] & c2["keywords"]
                        common_terms = ", ".join(sorted(list(common_kw))[:5])
                        reasoning = f"Both controls address [{common_terms}] - {c1['framework_name']} via '{c1['name']}' and {c2['framework_name']} via '{c2['name']}'"
                        pairs.append({
                            "control1": c1,
                            "control2": c2,
                            "score": score,
                            "reasoning": reasoning
                        })
    
    pairs.sort(key=lambda x: x["score"], reverse=True)
    top_pairs = pairs[:50]
    
    items = []
    for idx, p in enumerate(top_pairs):
        items.append({
            "id": idx + 1,
            "control1_type": p["control1"].get("type", "parsed"),
            "control1_id": p["control1"]["id"],
            "control1_code": p["control1"]["code"],
            "control1_name": p["control1"]["name"],
            "control1_framework": p["control1"]["framework_name"],
            "control2_type": p["control2"].get("type", "parsed"),
            "control2_id": p["control2"]["id"],
            "control2_code": p["control2"]["code"],
            "control2_name": p["control2"]["name"],
            "control2_framework": p["control2"]["framework_name"],
            "similarity_score": round(p["score"], 4),
            "ai_reasoning": p["reasoning"]
        })
    
    return {"items": items}


# ════════════════════════════════════════════════════════════════════════════
# PIPELINE LAB — extend the baseline with a new framework (test endpoints).
# Safe by design: analyze writes nothing; commit creates a CANDIDATE run and
# never promotes unless asked; demo/candidate deletes are guarded to demo data.
# ════════════════════════════════════════════════════════════════════════════
from .. services import extend_baseline as _EB  # noqa: E402

class _ExtendReq(BaseModel):
    framework_id: int
    promote: bool = False

# (framework list is fetched in the UI from the existing /frameworks/available
#  endpoint — a GET /extend/frameworks here collides with /{group_id}/... routes.)

@router.post("/extend/demo")
def extend_create_demo(db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Create a throwaway DEMO framework (non-canonical domains) to test the pipeline."""
    tenant_id = get_user_primary_tenant(current_user, db)
    demo = [
        ("D-1", "Information Security Policy", "Top-level approved information security policy.", "Governance & Oversight"),
        ("D-2", "Access Control Policy", "Govern user access provisioning and authorization.", "Identity & Access Mgmt"),
        ("D-3", "Incident Response Plan", "Procedures for reporting and managing incidents.", "Incident Handling"),
        ("D-4", "Conduct a security risk assessment", "Identify threats, likelihood and impact.", "Risk & Compliance"),
        ("D-5", "Security Awareness Training", "Periodic security awareness training for staff.", "People & Training"),
        ("D-6", "Anti-Malware Protection", "Deploy and maintain anti-malware across endpoints.", "Threat Defense"),
        ("D-7", "Bespoke Flux Capacitor Attunement Record", "Deliberately unique — should be standalone.", "Misc Ops"),
    ]
    fw = UploadedFramework(tenant_id=tenant_id, name=f"DEMO Framework {uuid.uuid4().hex[:6]} (delete me)",
                           file_name="demo.json", file_path="(demo)", file_type="application/json",
                           upload_status="completed", framework_type="regulatory", source_organization="DEMO",
                           version="1.0", is_active=True, is_shared=False, uploaded_by=current_user.id)
    db.add(fw); db.flush()
    for cid, t, desc, dom in demo:
        db.add(ParsedFrameworkControl(uploaded_framework_id=fw.id, control_id=cid, title=t, description=desc,
                                      full_text=desc, domain=dom, is_mandatory=True, priority="high",
                                      evidence_requirements=[f"{t} document"]))
    db.commit()
    return {"framework_id": fw.id, "name": fw.name, "controls": len(demo)}

@router.post("/extend/analyze")
def extend_analyze(req: _ExtendReq, db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    tenant_id = get_user_primary_tenant(current_user, db)
    return _EB.analyze(db, tenant_id, req.framework_id, get_client=get_openai_client)

@router.post("/extend/commit")
def extend_commit(req: _ExtendReq, db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    tenant_id = get_user_primary_tenant(current_user, db)
    return _EB.commit(db, tenant_id, req.framework_id, get_client=get_openai_client,
                      user_id=current_user.id, label="Pipeline-Lab candidate", promote=req.promote)

@router.delete("/extend/candidate/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
def extend_delete_candidate(run_id: int, db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    tenant_id = get_user_primary_tenant(current_user, db)
    run = db.query(NormalizationRun).filter(NormalizationRun.id == run_id, NormalizationRun.tenant_id == tenant_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.is_baseline:
        raise HTTPException(status_code=400, detail="Refusing to delete the live baseline.")
    from sqlalchemy import text as _text
    for stmt in ["delete from grc_common_control_group_mappings where group_id in (select id from grc_common_control_groups where run_id=:r)",
                 "delete from grc_normalized_control_links where normalized_control_id in (select id from grc_normalized_controls where run_id=:r)",
                 "delete from grc_common_control_groups where run_id=:r",
                 "delete from grc_normalized_controls where run_id=:r",
                 "delete from grc_normalization_runs where id=:r"]:
        db.execute(_text(stmt), {"r": run_id})
    db.commit()

@router.delete("/extend/framework/{framework_id}", status_code=status.HTTP_204_NO_CONTENT)
def extend_delete_framework(framework_id: int, db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    tenant_id = get_user_primary_tenant(current_user, db)
    fw = db.query(UploadedFramework).filter(UploadedFramework.id == framework_id).first()
    if not fw:
        raise HTTPException(status_code=404, detail="Framework not found")
    # Safety: a framework that is part of the LIVE baseline (one of the kept 30)
    # may not be deleted here. Demo frameworks and any newly-uploaded framework
    # that hasn't been kept into the baseline yet are removable.
    base = _scoped_get_baseline_run(db, tenant_id)
    base_ids = set(base.framework_ids or []) if base else set()
    is_demo = (fw.name or "").startswith(("DEMO", "PIPELINE"))
    if not is_demo and framework_id in base_ids:
        raise HTTPException(status_code=400,
                            detail="This framework is part of the live library and cannot be deleted here.")
    from sqlalchemy import text as _text
    # also remove the framework's ingested artifacts from the catalog
    fkey = _EB.framework_key_for(fw.name)
    db.execute(_text("delete from grc_artifact_catalog_items where framework_key=:k"), {"k": fkey})
    for stmt in ["delete from grc_control_evidence_requirements where parsed_control_id in (select id from grc_parsed_framework_controls where uploaded_framework_id=:f)",
                 "delete from grc_normalized_control_links where parsed_control_id in (select id from grc_parsed_framework_controls where uploaded_framework_id=:f)",
                 "delete from grc_common_control_group_mappings where parsed_control_id in (select id from grc_parsed_framework_controls where uploaded_framework_id=:f)",
                 "delete from grc_parsed_framework_controls where uploaded_framework_id=:f",
                 "delete from grc_uploaded_frameworks where id=:f"]:
        db.execute(_text(stmt), {"f": framework_id})
    db.commit()


# ════════════════════════════════════════════════════════════════════════════
# AUTO-ABSORB — the real flow: a new framework enters the system and the pipeline
# triggers on its own. /extend/pending detects new frameworks; /extend/upload
# ingests a seed JSON; /extend/start kicks the phased background job; the UI polls
# /extend/job/{id}; /extend/promote keeps the result; deletes above discard it.
# ════════════════════════════════════════════════════════════════════════════
from .. services.scoped_session import get_baseline_run as _scoped_get_baseline_run  # noqa: E402

_ABSORB_NS = "absorb"


def _phases_meta():
    return [{"key": k, "label": l} for k, l in _EB.PHASES]


def _absorb_worker(slug: str, tenant_id: int, fw_id: int, user_id: int, promote: bool):
    """Background thread: runs the phased absorption with its own tenant session,
    streaming progress into the Redis job-status store the UI polls."""
    from ....db import open_tenant_session
    wdb = open_tenant_session(slug)

    def progress(phase, pct, msg, extra):
        patch = {"status": "running", "phase": phase, "percent": pct, "message": msg}
        if extra:
            patch.update(extra)
        update_job_status(slug, _ABSORB_NS, fw_id, patch)

    try:
        res = _EB.run_absorption(wdb, tenant_id, fw_id, get_client=get_openai_client,
                                 user_id=user_id, promote=promote, progress=progress)
        update_job_status(slug, _ABSORB_NS, fw_id, {
            "status": "done", "phase": "done", "percent": 100,
            "message": "Absorption complete — review and keep or discard.", "result": res})
    except Exception as e:  # noqa: BLE001
        logging.getLogger(__name__).exception("auto-absorb failed for fw=%s", fw_id)
        update_job_status(slug, _ABSORB_NS, fw_id, {"status": "error", "message": str(e)})
    finally:
        wdb.close()


@router.get("/extend/pending")
def extend_pending(db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Active frameworks not yet part of the live baseline — what auto-absorb picks up."""
    tenant_id = get_user_primary_tenant(current_user, db)
    return _EB.pending_frameworks(db, tenant_id)


@router.post("/extend/upload")
async def extend_upload(request: Request, file: UploadFile = File(...),
                        db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Upload a raw framework seed (.json, controls-only is fine). Ingests it as a
    new framework (like a developer seeding it), then it shows up as pending."""
    tenant_id = get_user_primary_tenant(current_user, db)
    raw = await file.read()
    try:
        data = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="File is not valid JSON.")
    if not isinstance(data, dict) or "metadata" not in data or "controls" not in data:
        raise HTTPException(status_code=400,
                            detail="Seed must be an object with 'metadata' and 'controls'.")
    from ....seed_frameworks import seed_framework_from_json, framework_exists
    name = (data.get("metadata") or {}).get("name")
    if name and framework_exists(db, name, tenant_id):
        existing = db.query(UploadedFramework).filter(
            UploadedFramework.name == name, UploadedFramework.tenant_id == tenant_id).first()
        n = db.query(ParsedFrameworkControl).filter(
            ParsedFrameworkControl.uploaded_framework_id == existing.id).count()
        return {"framework_id": existing.id, "name": existing.name, "controls": n, "already_existed": True}
    fw = seed_framework_from_json(db, data, tenant_id=tenant_id, uploaded_by=current_user.id, force=False)
    if not fw:
        raise HTTPException(status_code=400, detail="Could not ingest the framework.")
    # SANDBOX ISOLATION: the uploaded framework is hidden from the rest of the app
    # (Frameworks / Coverage / Gap / main library all filter is_active) until the
    # user presses Keep. This guarantees testing never affects the live library.
    fw.is_active = False
    fw.is_shared = False
    fw.upload_status = "sandbox"
    db.commit()
    n = db.query(ParsedFrameworkControl).filter(
        ParsedFrameworkControl.uploaded_framework_id == fw.id).count()
    # The framework brings its own artifacts too — ingest them into the catalog.
    artifacts_in = 0
    if isinstance(data.get("artifacts"), list) and data["artifacts"]:
        artifacts_in = _EB.ingest_artifacts(db, fw.id, data["artifacts"])
    return {"framework_id": fw.id, "name": fw.name, "controls": n,
            "artifacts": artifacts_in, "already_existed": False}


@router.post("/extend/start")
def extend_start(req: _ExtendReq, request: Request,
                 db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Kick the phased auto-absorb job for a framework. Returns immediately; the UI
    polls /extend/job/{framework_id}. Builds a CANDIDATE (never promotes unless asked)."""
    tenant_id = get_user_primary_tenant(current_user, db)
    slug = getattr(request.state, "tenant_slug", None)
    if not slug:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    fw = db.query(UploadedFramework).filter(UploadedFramework.id == req.framework_id).first()
    if not fw:
        raise HTTPException(status_code=404, detail="Framework not found")
    # Concurrency guard: never run two absorptions for the same framework at once
    # (a browser-tab watcher + a manual start, or several open tabs, would otherwise
    # each spawn a candidate run). If one is already in flight, return it as-is.
    existing = get_job_status(slug, _ABSORB_NS, fw.id, default={"status": "idle"})
    if existing.get("status") == "running":
        return {"job_id": fw.id, "framework_id": fw.id, "status": "running",
                "already_running": True, "phases": _phases_meta()}
    # If a finished candidate already exists for this framework (and is still
    # present, i.e. not yet discarded), don't build a duplicate — return it.
    if existing.get("status") == "done":
        cand_id = (existing.get("result") or {}).get("candidate_run_id")
        if cand_id and db.query(NormalizationRun).filter(NormalizationRun.id == cand_id).first():
            return {"job_id": fw.id, "framework_id": fw.id, "status": "done",
                    "already_done": True, "phases": _phases_meta()}
    set_job_status(slug, _ABSORB_NS, fw.id, {
        "status": "running", "phase": "read", "percent": 1,
        "message": "Starting…", "framework": fw.name, "phases": _phases_meta()})
    threading.Thread(target=_absorb_worker,
                     args=(slug, tenant_id, fw.id, current_user.id, bool(req.promote)),
                     daemon=True).start()
    return {"job_id": fw.id, "framework_id": fw.id, "status": "running", "phases": _phases_meta()}


@router.get("/extend/job/{framework_id}")
def extend_job(framework_id: int, request: Request,
               db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    slug = getattr(request.state, "tenant_slug", None)
    if not slug:
        raise HTTPException(status_code=400, detail="Tenant context required.")
    st = get_job_status(slug, _ABSORB_NS, framework_id, default={"status": "idle"})
    if "phases" not in st:
        st["phases"] = _phases_meta()
    return st


@router.get("/extend/candidate/{run_id}/placements")
def extend_placements(run_id: int, framework_id: int = Query(...),
                      db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """The mock/duplicate library: faithful copy of the live library + where every
    new-framework control landed in it. The live library is never touched."""
    tenant_id = get_user_primary_tenant(current_user, db)
    try:
        return _EB.candidate_placements(db, tenant_id, run_id, framework_id)
    except RuntimeError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/extend/candidate/{run_id}/export")
def extend_export(run_id: int, framework_id: int = Query(...),
                  db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Download the pipeline result as an Excel workbook (Summary / Placements /
    Evidence sheets) — the 'goal answers in Excel'."""
    from fastapi import Response
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    tenant_id = get_user_primary_tenant(current_user, db)
    try:
        data = _EB.candidate_placements(db, tenant_id, run_id, framework_id)
    except RuntimeError as e:
        raise HTTPException(status_code=404, detail=str(e))

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="6D28D9")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def _autosize(ws, widths):
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 1 — Summary
    ws = wb.active; ws.title = "Summary"
    ws.append(["New-Framework Absorption — Result"]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    summary = [
        ("Framework", data["framework"]),
        ("Candidate (mock) library run", f"#{data['candidate_run_id']}"),
        ("Live library (untouched)", f"run #{data['live']['run_id']} · {data['live']['total']} entries · {data['live']['domains']} domains"),
        ("Mock library", f"{data['mock']['total']} entries · {data['mock']['domains']} domains"),
        ("Controls absorbed", len(data["placements"])),
        ("Joined existing sets", data["mock"]["added_join"]),
        ("New standalone entries", data["mock"]["added_standalone"]),
        ("Recommended evidence items generated", len(data["evidence"])),
    ]
    for k, v in summary:
        ws.append([k, v]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    _autosize(ws, [40, 70])

    # Sheet 2 — Placements (the framework → library pipeline)
    ws2 = wb.create_sheet("Placements")
    cols = ["Control ID", "Title", "Framework domain", "→ Library domain", "Disposition", "Joined set"]
    ws2.append(cols); _style_header(ws2, len(cols))
    for p in data["placements"]:
        ws2.append([p["control_id"], p["title"], p["framework_domain"],
                    p["canonical_domain"], p["disposition"], p["joined_set"] or ""])
    _autosize(ws2, [16, 50, 28, 30, 14, 40])

    # Sheet 3 — Per-domain composition
    ws3 = wb.create_sheet("By domain")
    cols3 = ["Library domain", "Existing (live)", "Added from framework", "Total in mock"]
    ws3.append(cols3); _style_header(ws3, len(cols3))
    for d in data["per_domain"]:
        ws3.append([d["domain"], d["baseline"], d["added"], d["total"]])
    _autosize(ws3, [38, 16, 22, 16])

    # Sheet 4 — Evidence (carried + generated)
    ws4 = wb.create_sheet("Evidence")
    cols4 = ["Control ID", "Control title", "Recommended evidence", "Type"]
    ws4.append(cols4); _style_header(ws4, len(cols4))
    for e in data["evidence"]:
        ws4.append([e["control_id"], e["control_title"], e["evidence_title"], e["evidence_type"]])
    _autosize(ws4, [16, 45, 50, 16])

    # Sheet 5 — Artifact normalization (new vs deduped against the unified catalog)
    art = data.get("artifacts") or {}
    ws5 = wb.create_sheet("Artifacts")
    ws5.append([f"Artifacts: {art.get('artifacts_total', 0)} total · "
                f"{art.get('artifacts_new', 0)} new · {art.get('artifacts_duplicate', 0)} deduped"])
    ws5["A1"].font = Font(bold=True)
    ws5.append([])
    ws5.append(["Disposition", "Artifact", "Type", "Matches existing / control ref"])
    _style_header(ws5, 4)
    # openpyxl freezes row 1; header is row 3 here — re-apply header style on row 3
    for c in range(1, 5):
        cell = ws5.cell(row=3, column=c); cell.fill = hdr_fill; cell.font = hdr_font
    for a in art.get("artifacts_new_sample", []):
        ws5.append(["NEW", a.get("artifact"), a.get("type"), a.get("control_ref") or ""])
    for a in art.get("artifacts_dup_sample", []):
        ws5.append(["DEDUPED", a.get("artifact"), a.get("type"),
                    f"{a.get('matches')} ({a.get('in_framework')})"])
    _autosize(ws5, [14, 44, 16, 50])

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = "".join(ch for ch in (data["framework"] or "framework") if ch.isalnum() or ch in " -_")[:50].strip().replace(" ", "_")
    fname = f"absorption_{safe}_run{data['candidate_run_id']}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/extend/promote/{run_id}")
def extend_promote(run_id: int, db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Keep a candidate run: make it the live baseline (the only step that changes
    what the rest of the UI shows). Fully reversible — the old baseline run remains."""
    tenant_id = get_user_primary_tenant(current_user, db)
    run = db.query(NormalizationRun).filter(
        NormalizationRun.id == run_id, NormalizationRun.tenant_id == tenant_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    db.query(NormalizationRun).filter(
        NormalizationRun.tenant_id == tenant_id, NormalizationRun.is_baseline.is_(True)).update(
        {NormalizationRun.is_baseline: False})
    run.is_baseline = True
    # Activate any sandbox framework(s) this run added — now that it's live, the
    # framework and its artifacts become visible in the rest of the app.
    ids = list(run.framework_ids or [])
    if ids:
        db.query(UploadedFramework).filter(
            UploadedFramework.id.in_(ids),
            UploadedFramework.upload_status == "sandbox").update(
            {UploadedFramework.is_active: True, UploadedFramework.is_shared: True,
             UploadedFramework.upload_status: "completed"}, synchronize_session=False)
    db.commit()
    return {"promoted_run_id": run_id, "is_baseline": True}
