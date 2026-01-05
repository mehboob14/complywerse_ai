import os
import io
import json
import uuid
from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func, distinct
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI

from ....models import (
    CommonControlGroup, CommonControlGroupMapping, NormalizedControl,
    FrameworkControl, FrameworkDomain, ControlObjective, Framework,
    ControlSimilarityMapping, AIEvidenceRecommendation,
    Evidence, EvidenceControlMapping, GRCUser, get_db
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/reports", tags=["Control Library - Harmonization Reports"])

AI_INTEGRATIONS_OPENAI_API_KEY = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY")
AI_INTEGRATIONS_OPENAI_BASE_URL = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "uploads", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

generated_reports = {}


class ExportRequest(BaseModel):
    format: str = "xlsx"
    framework_ids: Optional[List[int]] = []
    include_sections: Optional[List[str]] = ["summary", "groups", "mappings", "evidence"]


class ExecutiveSummaryRequest(BaseModel):
    include_recommendations: bool = True
    focus_areas: Optional[List[str]] = None


def get_openai_client() -> OpenAI:
    if not AI_INTEGRATIONS_OPENAI_API_KEY:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="OpenAI integration not configured"
        )
    return OpenAI(
        api_key=AI_INTEGRATIONS_OPENAI_API_KEY,
        base_url=AI_INTEGRATIONS_OPENAI_BASE_URL
    )


def get_user_tenant_id(user: GRCUser, db: Session) -> int:
    tenant_id = get_user_primary_tenant(user, db)
    if not tenant_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="User is not assigned to any tenant"
        )
    return tenant_id


def get_framework_controls_map(db: Session) -> dict:
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    fw_map = {}
    for fw in frameworks:
        fw_map[fw.id] = {
            "framework": fw,
            "controls": []
        }
        for domain in fw.domains:
            for objective in domain.objectives:
                controls = db.query(FrameworkControl).filter(
                    FrameworkControl.objective_id == objective.id
                ).all()
                for ctrl in controls:
                    fw_map[fw.id]["controls"].append({
                        "control": ctrl,
                        "domain": domain,
                        "objective": objective
                    })
    return fw_map


def get_group_data(db: Session, tenant_id: int, include_controls: bool = True) -> List[dict]:
    groups = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id == tenant_id,
            CommonControlGroup.tenant_id.is_(None)
        )
    ).all()
    
    result = []
    for group in groups:
        mappings = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group.id
        ).all()
        
        normalized_controls = []
        framework_controls = []
        
        if include_controls:
            for mapping in mappings:
                if mapping.normalized_control_id:
                    nc = db.query(NormalizedControl).filter(
                        NormalizedControl.id == mapping.normalized_control_id
                    ).first()
                    if nc:
                        normalized_controls.append({
                            "id": nc.id,
                            "code": nc.code,
                            "name": nc.name,
                            "statement": nc.statement,
                            "mapping_confidence": mapping.mapping_confidence
                        })
                
                if mapping.framework_control_id:
                    fc = db.query(FrameworkControl).options(
                        joinedload(FrameworkControl.objective)
                        .joinedload(ControlObjective.domain)
                        .joinedload(FrameworkDomain.framework)
                    ).filter(
                        FrameworkControl.id == mapping.framework_control_id
                    ).first()
                    if fc:
                        framework = fc.objective.domain.framework if fc.objective and fc.objective.domain else None
                        framework_controls.append({
                            "id": fc.id,
                            "code": fc.code,
                            "name": fc.name,
                            "statement": fc.statement,
                            "framework_id": framework.id if framework else None,
                            "framework_name": framework.name if framework else None,
                            "framework_code": framework.short_code if framework else None,
                            "mapping_confidence": mapping.mapping_confidence
                        })
        
        result.append({
            "id": group.id,
            "code": group.code,
            "name": group.name,
            "description": group.description,
            "category": group.category,
            "domain": group.domain,
            "keywords": group.keywords or [],
            "evidence_types": group.evidence_types or [],
            "ai_summary": group.ai_summary,
            "normalized_control_count": len(normalized_controls),
            "framework_control_count": len(framework_controls),
            "normalized_controls": normalized_controls,
            "framework_controls": framework_controls
        })
    
    return result


def get_similarity_mappings(db: Session, tenant_id: int, min_score: float = 0.5) -> List[dict]:
    mappings = db.query(ControlSimilarityMapping).filter(
        ControlSimilarityMapping.tenant_id == tenant_id,
        ControlSimilarityMapping.similarity_score >= min_score
    ).order_by(ControlSimilarityMapping.similarity_score.desc()).all()
    
    result = []
    for m in mappings:
        source_control = None
        target_control = None
        
        if m.source_type == "normalized":
            nc = db.query(NormalizedControl).filter(NormalizedControl.id == m.source_control_id).first()
            if nc:
                source_control = {"id": nc.id, "type": "normalized", "code": nc.code, "name": nc.name}
        elif m.source_type == "framework":
            fc = db.query(FrameworkControl).options(
                joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
            ).filter(FrameworkControl.id == m.source_control_id).first()
            if fc:
                fw = fc.objective.domain.framework if fc.objective and fc.objective.domain else None
                source_control = {
                    "id": fc.id, "type": "framework", "code": fc.code, "name": fc.name,
                    "framework": fw.short_code if fw else None
                }
        
        if m.target_type == "normalized":
            nc = db.query(NormalizedControl).filter(NormalizedControl.id == m.target_control_id).first()
            if nc:
                target_control = {"id": nc.id, "type": "normalized", "code": nc.code, "name": nc.name}
        elif m.target_type == "framework":
            fc = db.query(FrameworkControl).options(
                joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
            ).filter(FrameworkControl.id == m.target_control_id).first()
            if fc:
                fw = fc.objective.domain.framework if fc.objective and fc.objective.domain else None
                target_control = {
                    "id": fc.id, "type": "framework", "code": fc.code, "name": fc.name,
                    "framework": fw.short_code if fw else None
                }
        
        result.append({
            "id": m.id,
            "source": source_control,
            "target": target_control,
            "similarity_score": m.similarity_score,
            "similarity_type": m.similarity_type,
            "ai_reasoning": m.ai_reasoning,
            "verified": m.verified
        })
    
    return result


def get_evidence_recommendations(db: Session, tenant_id: int) -> List[dict]:
    recs = db.query(AIEvidenceRecommendation).filter(
        AIEvidenceRecommendation.tenant_id == tenant_id
    ).order_by(AIEvidenceRecommendation.priority).all()
    
    result = []
    for r in recs:
        control_info = None
        if r.normalized_control_id:
            nc = db.query(NormalizedControl).filter(NormalizedControl.id == r.normalized_control_id).first()
            if nc:
                control_info = {"type": "normalized", "code": nc.code, "name": nc.name}
        elif r.framework_control_id:
            fc = db.query(FrameworkControl).filter(FrameworkControl.id == r.framework_control_id).first()
            if fc:
                control_info = {"type": "framework", "code": fc.code, "name": fc.name}
        
        result.append({
            "id": r.id,
            "group_id": r.group_id,
            "control": control_info,
            "evidence_type": r.evidence_type,
            "evidence_description": r.evidence_description,
            "priority": r.priority,
            "ai_confidence": r.ai_confidence,
            "sample_evidence_names": r.sample_evidence_names or []
        })
    
    return result


def calculate_summary_stats(db: Session, tenant_id: int) -> dict:
    user_tenants = [tenant_id]
    
    groups_count = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        )
    ).count()
    
    normalized_count = db.query(NormalizedControl).count()
    
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    framework_stats = []
    total_framework_controls = 0
    
    for fw in frameworks:
        control_count = 0
        for domain in fw.domains:
            for objective in domain.objectives:
                control_count += db.query(FrameworkControl).filter(
                    FrameworkControl.objective_id == objective.id
                ).count()
        total_framework_controls += control_count
        framework_stats.append({
            "id": fw.id,
            "name": fw.name,
            "code": fw.short_code,
            "control_count": control_count
        })
    
    similarity_count = db.query(ControlSimilarityMapping).filter(
        ControlSimilarityMapping.tenant_id == tenant_id
    ).count()
    
    verified_mappings = db.query(ControlSimilarityMapping).filter(
        ControlSimilarityMapping.tenant_id == tenant_id,
        ControlSimilarityMapping.verified == True
    ).count()
    
    evidence_count = db.query(Evidence).filter(Evidence.tenant_id == tenant_id).count()
    approved_evidence = db.query(Evidence).filter(
        Evidence.tenant_id == tenant_id,
        Evidence.status == "approved"
    ).count()
    
    tenant_evidence_ids = db.query(Evidence.id).filter(Evidence.tenant_id == tenant_id).subquery()
    covered_controls = db.query(func.count(distinct(EvidenceControlMapping.framework_control_id))).filter(
        EvidenceControlMapping.evidence_id.in_(tenant_evidence_ids),
        EvidenceControlMapping.framework_control_id.isnot(None)
    ).scalar() or 0
    
    recommendations_count = db.query(AIEvidenceRecommendation).filter(
        AIEvidenceRecommendation.tenant_id == tenant_id
    ).count()
    
    return {
        "generated_at": datetime.utcnow().isoformat(),
        "common_control_groups": groups_count,
        "normalized_controls": normalized_count,
        "total_framework_controls": total_framework_controls,
        "frameworks": framework_stats,
        "similarity_mappings": similarity_count,
        "verified_mappings": verified_mappings,
        "total_evidence": evidence_count,
        "approved_evidence": approved_evidence,
        "controls_with_evidence": covered_controls,
        "evidence_coverage_percent": round((covered_controls / total_framework_controls * 100) if total_framework_controls > 0 else 0, 2),
        "evidence_recommendations": recommendations_count
    }


def create_xlsx_report(
    summary: dict,
    groups: List[dict],
    mappings: List[dict],
    recommendations: List[dict],
    include_sections: List[str]
) -> io.BytesIO:
    wb = Workbook()
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    if "summary" in include_sections:
        ws = wb.active
        ws.title = "Summary"
        
        ws["A1"] = "Harmonization Report Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:B1")
        
        ws["A3"] = "Generated At"
        ws["B3"] = summary.get("generated_at", "")
        
        row = 5
        summary_data = [
            ("Common Control Groups", summary.get("common_control_groups", 0)),
            ("Normalized Controls", summary.get("normalized_controls", 0)),
            ("Total Framework Controls", summary.get("total_framework_controls", 0)),
            ("Similarity Mappings", summary.get("similarity_mappings", 0)),
            ("Verified Mappings", summary.get("verified_mappings", 0)),
            ("Total Evidence", summary.get("total_evidence", 0)),
            ("Approved Evidence", summary.get("approved_evidence", 0)),
            ("Controls with Evidence", summary.get("controls_with_evidence", 0)),
            ("Evidence Coverage %", f"{summary.get('evidence_coverage_percent', 0)}%"),
            ("Evidence Recommendations", summary.get("evidence_recommendations", 0))
        ]
        
        for label, value in summary_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1
        
        row += 2
        ws[f"A{row}"] = "Frameworks"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        ws[f"A{row}"] = "Framework"
        ws[f"B{row}"] = "Code"
        ws[f"C{row}"] = "Control Count"
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = header_font
            ws[f"{col}{row}"].fill = header_fill
        row += 1
        
        for fw in summary.get("frameworks", []):
            ws[f"A{row}"] = fw.get("name", "")
            ws[f"B{row}"] = fw.get("code", "")
            ws[f"C{row}"] = fw.get("control_count", 0)
            row += 1
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
    
    if "groups" in include_sections:
        ws = wb.create_sheet("Control Groups")
        
        headers = ["Group Code", "Group Name", "Category", "Domain", "Normalized Controls", "Framework Controls", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row = 2
        for group in groups:
            ws.cell(row=row, column=1, value=group.get("code", ""))
            ws.cell(row=row, column=2, value=group.get("name", ""))
            ws.cell(row=row, column=3, value=group.get("category", ""))
            ws.cell(row=row, column=4, value=group.get("domain", ""))
            ws.cell(row=row, column=5, value=group.get("normalized_control_count", 0))
            ws.cell(row=row, column=6, value=group.get("framework_control_count", 0))
            ws.cell(row=row, column=7, value=group.get("description", "")[:500] if group.get("description") else "")
            row += 1
        
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    if "mappings" in include_sections:
        ws = wb.create_sheet("Similarity Mappings")
        
        headers = ["Source Type", "Source Code", "Source Name", "Target Type", "Target Code", "Target Name", 
                   "Similarity Score", "Similarity Type", "Verified", "AI Reasoning"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row = 2
        for m in mappings:
            source = m.get("source") or {}
            target = m.get("target") or {}
            ws.cell(row=row, column=1, value=source.get("type", ""))
            ws.cell(row=row, column=2, value=source.get("code", ""))
            ws.cell(row=row, column=3, value=source.get("name", ""))
            ws.cell(row=row, column=4, value=target.get("type", ""))
            ws.cell(row=row, column=5, value=target.get("code", ""))
            ws.cell(row=row, column=6, value=target.get("name", ""))
            ws.cell(row=row, column=7, value=m.get("similarity_score", 0))
            ws.cell(row=row, column=8, value=m.get("similarity_type", ""))
            ws.cell(row=row, column=9, value="Yes" if m.get("verified") else "No")
            ws.cell(row=row, column=10, value=(m.get("ai_reasoning", "") or "")[:500])
            row += 1
        
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    if "evidence" in include_sections:
        ws = wb.create_sheet("Evidence Recommendations")
        
        headers = ["Control Type", "Control Code", "Control Name", "Evidence Type", "Priority", 
                   "Confidence", "Description", "Sample Names"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row = 2
        for rec in recommendations:
            control = rec.get("control") or {}
            ws.cell(row=row, column=1, value=control.get("type", ""))
            ws.cell(row=row, column=2, value=control.get("code", ""))
            ws.cell(row=row, column=3, value=control.get("name", ""))
            ws.cell(row=row, column=4, value=rec.get("evidence_type", ""))
            ws.cell(row=row, column=5, value=rec.get("priority", ""))
            ws.cell(row=row, column=6, value=rec.get("ai_confidence"))
            ws.cell(row=row, column=7, value=(rec.get("evidence_description", "") or "")[:500])
            ws.cell(row=row, column=8, value=", ".join(rec.get("sample_evidence_names", [])))
            row += 1
        
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_csv_report(data: List[dict], headers: List[str]) -> io.BytesIO:
    import csv
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    
    bytes_output = io.BytesIO()
    bytes_output.write(output.getvalue().encode("utf-8"))
    bytes_output.seek(0)
    return bytes_output


@router.get("/harmonization")
def get_harmonization_report(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    summary = calculate_summary_stats(db, tenant_id)
    groups = get_group_data(db, tenant_id, include_controls=True)
    mappings = get_similarity_mappings(db, tenant_id)
    recommendations = get_evidence_recommendations(db, tenant_id)
    
    return {
        "report_type": "harmonization",
        "generated_at": datetime.utcnow().isoformat(),
        "summary": summary,
        "common_control_groups": groups,
        "similarity_mappings": mappings,
        "evidence_recommendations": recommendations
    }


@router.get("/framework/{framework_id}")
def get_framework_report(
    framework_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    framework = db.query(Framework).filter(Framework.id == framework_id).first()
    if not framework:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework not found"
        )
    
    controls = []
    for domain in framework.domains:
        for objective in domain.objectives:
            fc_list = db.query(FrameworkControl).filter(
                FrameworkControl.objective_id == objective.id
            ).all()
            for fc in fc_list:
                controls.append({
                    "control": fc,
                    "domain": domain,
                    "objective": objective
                })
    
    tenant_evidence_ids = db.query(Evidence.id).filter(Evidence.tenant_id == tenant_id).subquery()
    covered_fc_ids = set(
        row[0] for row in db.query(EvidenceControlMapping.framework_control_id).filter(
            EvidenceControlMapping.evidence_id.in_(tenant_evidence_ids),
            EvidenceControlMapping.framework_control_id.isnot(None)
        ).distinct().all()
    )
    
    cross_mappings = []
    control_data = []
    
    for item in controls:
        fc = item["control"]
        has_evidence = fc.id in covered_fc_ids
        
        similar_controls = db.query(ControlSimilarityMapping).filter(
            ControlSimilarityMapping.tenant_id == tenant_id,
            or_(
                ControlSimilarityMapping.source_control_id == fc.id,
                ControlSimilarityMapping.target_control_id == fc.id
            ),
            ControlSimilarityMapping.source_type == "framework"
        ).all()
        
        mapped_to = []
        for sim in similar_controls:
            other_id = sim.target_control_id if sim.source_control_id == fc.id else sim.source_control_id
            other_type = sim.target_type if sim.source_control_id == fc.id else sim.source_type
            
            if other_type == "framework":
                other_fc = db.query(FrameworkControl).options(
                    joinedload(FrameworkControl.objective).joinedload(ControlObjective.domain).joinedload(FrameworkDomain.framework)
                ).filter(FrameworkControl.id == other_id).first()
                if other_fc and other_fc.objective and other_fc.objective.domain and other_fc.objective.domain.framework:
                    if other_fc.objective.domain.framework.id != framework_id:
                        mapped_to.append({
                            "framework_id": other_fc.objective.domain.framework.id,
                            "framework_code": other_fc.objective.domain.framework.short_code,
                            "control_code": other_fc.code,
                            "control_name": other_fc.name,
                            "similarity_score": sim.similarity_score
                        })
        
        recommendations = db.query(AIEvidenceRecommendation).filter(
            AIEvidenceRecommendation.tenant_id == tenant_id,
            AIEvidenceRecommendation.framework_control_id == fc.id
        ).all()
        
        control_data.append({
            "id": fc.id,
            "code": fc.code,
            "name": fc.name,
            "domain": item["domain"].name,
            "objective": item["objective"].name,
            "has_evidence": has_evidence,
            "mapped_to_other_frameworks": mapped_to,
            "evidence_recommendations": [
                {"type": r.evidence_type, "priority": r.priority}
                for r in recommendations
            ]
        })
    
    covered_count = sum(1 for c in control_data if c["has_evidence"])
    mapped_count = sum(1 for c in control_data if c["mapped_to_other_frameworks"])
    
    return {
        "framework_id": framework.id,
        "framework_name": framework.name,
        "framework_code": framework.short_code,
        "total_controls": len(control_data),
        "controls_with_evidence": covered_count,
        "controls_mapped_to_others": mapped_count,
        "evidence_coverage_percent": round((covered_count / len(control_data) * 100) if control_data else 0, 2),
        "controls": control_data
    }


@router.post("/export")
def export_harmonization_report(
    request: ExportRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    summary = calculate_summary_stats(db, tenant_id)
    groups = get_group_data(db, tenant_id, include_controls=False)
    mappings = get_similarity_mappings(db, tenant_id)
    recommendations = get_evidence_recommendations(db, tenant_id)
    
    report_id = str(uuid.uuid4())
    filename = f"harmonization_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    
    if request.format == "xlsx":
        output = create_xlsx_report(
            summary, groups, mappings, recommendations, request.include_sections or []
        )
        file_path = os.path.join(REPORTS_DIR, f"{filename}.xlsx")
        with open(file_path, "wb") as f:
            f.write(output.getvalue())
        
        generated_reports[report_id] = {
            "id": report_id,
            "filename": f"{filename}.xlsx",
            "file_path": file_path,
            "format": "xlsx",
            "generated_at": datetime.utcnow().isoformat(),
            "generated_by": current_user.id,
            "tenant_id": tenant_id
        }
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    
    elif request.format == "csv":
        flat_data = []
        for group in groups:
            flat_data.append({
                "group_code": group.get("code"),
                "group_name": group.get("name"),
                "category": group.get("category"),
                "domain": group.get("domain"),
                "normalized_controls": group.get("normalized_control_count"),
                "framework_controls": group.get("framework_control_count")
            })
        
        headers = ["group_code", "group_name", "category", "domain", "normalized_controls", "framework_controls"]
        output = create_csv_report(flat_data, headers)
        
        file_path = os.path.join(REPORTS_DIR, f"{filename}.csv")
        with open(file_path, "wb") as f:
            f.write(output.getvalue())
        
        generated_reports[report_id] = {
            "id": report_id,
            "filename": f"{filename}.csv",
            "file_path": file_path,
            "format": "csv",
            "generated_at": datetime.utcnow().isoformat(),
            "generated_by": current_user.id,
            "tenant_id": tenant_id
        }
        
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported format: {request.format}. Supported formats: xlsx, csv"
        )


@router.get("/audit-ready")
def get_audit_ready_summary(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    
    tenant_evidence_ids = db.query(Evidence.id).filter(Evidence.tenant_id == tenant_id).subquery()
    covered_fc_ids = set(
        row[0] for row in db.query(EvidenceControlMapping.framework_control_id).filter(
            EvidenceControlMapping.evidence_id.in_(tenant_evidence_ids),
            EvidenceControlMapping.framework_control_id.isnot(None)
        ).distinct().all()
    )
    
    framework_summaries = []
    total_controls = 0
    total_covered = 0
    total_gaps = 0
    
    for fw in frameworks:
        control_count = 0
        covered = 0
        uncovered_controls = []
        
        for domain in fw.domains:
            for objective in domain.objectives:
                controls = db.query(FrameworkControl).filter(
                    FrameworkControl.objective_id == objective.id
                ).all()
                for ctrl in controls:
                    control_count += 1
                    if ctrl.id in covered_fc_ids:
                        covered += 1
                    else:
                        uncovered_controls.append({
                            "code": ctrl.code,
                            "name": ctrl.name,
                            "domain": domain.name
                        })
        
        gaps = control_count - covered
        total_controls += control_count
        total_covered += covered
        total_gaps += gaps
        
        framework_summaries.append({
            "framework_id": fw.id,
            "framework_name": fw.name,
            "framework_code": fw.short_code,
            "total_controls": control_count,
            "controls_with_evidence": covered,
            "gaps": gaps,
            "coverage_percent": round((covered / control_count * 100) if control_count > 0 else 0, 2),
            "top_gaps": uncovered_controls[:10]
        })
    
    total_evidence = db.query(Evidence).filter(Evidence.tenant_id == tenant_id).count()
    approved_evidence = db.query(Evidence).filter(
        Evidence.tenant_id == tenant_id,
        Evidence.status == "approved"
    ).count()
    
    groups_count = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id == tenant_id,
            CommonControlGroup.tenant_id.is_(None)
        )
    ).count()
    
    return {
        "report_type": "audit_ready_summary",
        "generated_at": datetime.utcnow().isoformat(),
        "overall_summary": {
            "total_frameworks": len(frameworks),
            "total_controls": total_controls,
            "controls_with_evidence": total_covered,
            "total_gaps": total_gaps,
            "overall_coverage_percent": round((total_covered / total_controls * 100) if total_controls > 0 else 0, 2),
            "common_control_groups": groups_count,
            "total_evidence_items": total_evidence,
            "approved_evidence": approved_evidence
        },
        "frameworks": framework_summaries,
        "audit_readiness_score": round((total_covered / total_controls * 100) if total_controls > 0 else 0, 1)
    }


@router.get("/cross-framework-mapping")
def get_cross_framework_mapping(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    user_tenants = get_user_tenants(current_user, db)
    
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    framework_map = {fw.id: fw.short_code for fw in frameworks}
    
    groups = db.query(CommonControlGroup).filter(
        or_(
            CommonControlGroup.tenant_id.in_(user_tenants),
            CommonControlGroup.tenant_id.is_(None)
        )
    ).all()
    
    mapping_table = []
    
    for group in groups:
        row = {
            "group_id": group.id,
            "group_code": group.code,
            "group_name": group.name,
            "category": group.category,
            "normalized_controls": []
        }
        
        for fw_id, fw_code in framework_map.items():
            row[fw_code] = []
        
        mappings = db.query(CommonControlGroupMapping).filter(
            CommonControlGroupMapping.group_id == group.id
        ).all()
        
        for mapping in mappings:
            if mapping.normalized_control_id:
                nc = db.query(NormalizedControl).filter(
                    NormalizedControl.id == mapping.normalized_control_id
                ).first()
                if nc:
                    row["normalized_controls"].append({
                        "code": nc.code,
                        "name": nc.name
                    })
            
            if mapping.framework_control_id:
                fc = db.query(FrameworkControl).options(
                    joinedload(FrameworkControl.objective)
                    .joinedload(ControlObjective.domain)
                    .joinedload(FrameworkDomain.framework)
                ).filter(FrameworkControl.id == mapping.framework_control_id).first()
                
                if fc and fc.objective and fc.objective.domain and fc.objective.domain.framework:
                    fw_code = fc.objective.domain.framework.short_code
                    if fw_code in row:
                        row[fw_code].append({
                            "code": fc.code,
                            "name": fc.name
                        })
        
        mapping_table.append(row)
    
    columns = ["group_code", "group_name", "category", "normalized_controls"] + list(framework_map.values())
    
    return {
        "columns": columns,
        "framework_codes": list(framework_map.values()),
        "mappings": mapping_table,
        "total_groups": len(mapping_table)
    }


@router.get("/evidence-requirements")
def get_evidence_requirements_report(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    tenant_evidence_ids = db.query(Evidence.id).filter(Evidence.tenant_id == tenant_id).subquery()
    
    fc_evidence_map = {}
    evidence_mappings = db.query(EvidenceControlMapping).filter(
        EvidenceControlMapping.evidence_id.in_(tenant_evidence_ids),
        EvidenceControlMapping.framework_control_id.isnot(None)
    ).all()
    
    for em in evidence_mappings:
        if em.framework_control_id not in fc_evidence_map:
            fc_evidence_map[em.framework_control_id] = []
        evidence = db.query(Evidence).filter(Evidence.id == em.evidence_id).first()
        if evidence:
            fc_evidence_map[em.framework_control_id].append({
                "id": evidence.id,
                "name": evidence.name,
                "type": evidence.evidence_type,
                "status": evidence.status
            })
    
    recommendations = db.query(AIEvidenceRecommendation).filter(
        AIEvidenceRecommendation.tenant_id == tenant_id,
        AIEvidenceRecommendation.framework_control_id.isnot(None)
    ).all()
    
    fc_rec_map = {}
    for rec in recommendations:
        if rec.framework_control_id not in fc_rec_map:
            fc_rec_map[rec.framework_control_id] = []
        fc_rec_map[rec.framework_control_id].append({
            "type": rec.evidence_type,
            "priority": rec.priority,
            "description": rec.evidence_description
        })
    
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    
    report_data = []
    
    for fw in frameworks:
        for domain in fw.domains:
            for objective in domain.objectives:
                controls = db.query(FrameworkControl).filter(
                    FrameworkControl.objective_id == objective.id
                ).all()
                
                for ctrl in controls:
                    actual_evidence = fc_evidence_map.get(ctrl.id, [])
                    required_evidence = fc_rec_map.get(ctrl.id, [])
                    
                    actual_types = set(e["type"] for e in actual_evidence if e["type"])
                    required_types = set(r["type"] for r in required_evidence)
                    
                    missing_types = required_types - actual_types
                    
                    report_data.append({
                        "framework_id": fw.id,
                        "framework_code": fw.short_code,
                        "control_id": ctrl.id,
                        "control_code": ctrl.code,
                        "control_name": ctrl.name,
                        "domain": domain.name,
                        "required_evidence_types": list(required_types),
                        "actual_evidence": actual_evidence,
                        "actual_evidence_count": len(actual_evidence),
                        "missing_evidence_types": list(missing_types),
                        "is_compliant": len(missing_types) == 0 and len(actual_evidence) > 0
                    })
    
    compliant_count = sum(1 for r in report_data if r["is_compliant"])
    
    return {
        "generated_at": datetime.utcnow().isoformat(),
        "total_controls": len(report_data),
        "compliant_controls": compliant_count,
        "non_compliant_controls": len(report_data) - compliant_count,
        "compliance_percent": round((compliant_count / len(report_data) * 100) if report_data else 0, 2),
        "controls": report_data
    }


@router.post("/generate-executive-summary")
def generate_executive_summary(
    request: ExecutiveSummaryRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    summary = calculate_summary_stats(db, tenant_id)
    
    frameworks = db.query(Framework).filter(Framework.is_active == True).all()
    
    tenant_evidence_ids = db.query(Evidence.id).filter(Evidence.tenant_id == tenant_id).subquery()
    covered_fc_ids = set(
        row[0] for row in db.query(EvidenceControlMapping.framework_control_id).filter(
            EvidenceControlMapping.evidence_id.in_(tenant_evidence_ids),
            EvidenceControlMapping.framework_control_id.isnot(None)
        ).distinct().all()
    )
    
    framework_gaps = []
    for fw in frameworks:
        total = 0
        covered = 0
        for domain in fw.domains:
            for objective in domain.objectives:
                controls = db.query(FrameworkControl).filter(
                    FrameworkControl.objective_id == objective.id
                ).all()
                for ctrl in controls:
                    total += 1
                    if ctrl.id in covered_fc_ids:
                        covered += 1
        
        gap = total - covered
        framework_gaps.append({
            "name": fw.name,
            "code": fw.short_code,
            "total": total,
            "covered": covered,
            "gap": gap,
            "coverage_percent": round((covered / total * 100) if total > 0 else 0, 1)
        })
    
    context = f"""
Organization Compliance Summary:
- Total Frameworks: {len(frameworks)}
- Common Control Groups: {summary.get('common_control_groups', 0)}
- Total Framework Controls: {summary.get('total_framework_controls', 0)}
- Controls with Evidence: {summary.get('controls_with_evidence', 0)}
- Evidence Coverage: {summary.get('evidence_coverage_percent', 0)}%
- Total Evidence Items: {summary.get('total_evidence', 0)}
- Approved Evidence: {summary.get('approved_evidence', 0)}
- Similarity Mappings: {summary.get('similarity_mappings', 0)}
- Verified Mappings: {summary.get('verified_mappings', 0)}

Framework-specific Status:
"""
    for fw in framework_gaps:
        context += f"\n- {fw['name']} ({fw['code']}): {fw['coverage_percent']}% coverage, {fw['gap']} gaps"
    
    if request.focus_areas:
        context += f"\n\nFocus areas: {', '.join(request.focus_areas)}"
    
    try:
        client = get_openai_client()
        
        prompt = f"""Based on the following compliance data, generate an executive summary for management:

{context}

Provide a JSON response with:
{{
    "overall_posture": "<1-2 sentence summary of overall compliance status>",
    "posture_rating": "<strong|adequate|needs_improvement|critical>",
    "key_strengths": ["<strength 1>", "<strength 2>", ...],
    "key_gaps": ["<gap 1>", "<gap 2>", ...],
    "recommendations": ["<recommendation 1>", "<recommendation 2>", ...],
    "priority_actions": [
        {{"action": "<action>", "priority": "<high|medium|low>", "rationale": "<why>"}}
    ],
    "risk_summary": "<overall risk assessment>",
    "next_steps": ["<next step 1>", "<next step 2>", ...]
}}

Be specific, actionable, and focus on business impact."""

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "You are a compliance expert generating executive summaries for management. Respond only with valid JSON."
                },
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=2000,
            temperature=0.4
        )
        
        ai_summary = json.loads(response.choices[0].message.content or '{}')
        
        return {
            "generated_at": datetime.utcnow().isoformat(),
            "data_summary": summary,
            "framework_coverage": framework_gaps,
            "executive_summary": ai_summary
        }
        
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        if "FREE_CLOUD_BUDGET_EXCEEDED" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Cloud budget exceeded. Please upgrade your plan."
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI summary generation failed: {error_msg}"
        )


@router.get("/download/{report_id}")
def download_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    if report_id not in generated_reports:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report not found. It may have expired or been deleted."
        )
    
    report = generated_reports[report_id]
    
    if report.get("tenant_id") != tenant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this report"
        )
    
    file_path = report.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file not found"
        )
    
    with open(file_path, "rb") as f:
        content = f.read()
    
    if report.get("format") == "xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif report.get("format") == "csv":
        media_type = "text/csv"
    else:
        media_type = "application/octet-stream"
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={report.get('filename', 'report')}"}
    )


@router.get("/history")
def get_report_history(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_tenant_id(current_user, db)
    
    tenant_reports = [
        {
            "id": r["id"],
            "filename": r["filename"],
            "format": r["format"],
            "generated_at": r["generated_at"],
            "generated_by": r["generated_by"]
        }
        for r in generated_reports.values()
        if r.get("tenant_id") == tenant_id
    ]
    
    tenant_reports.sort(key=lambda x: x["generated_at"], reverse=True)
    
    paginated = tenant_reports[skip:skip + limit]
    
    return {
        "total": len(tenant_reports),
        "skip": skip,
        "limit": limit,
        "reports": paginated
    }
