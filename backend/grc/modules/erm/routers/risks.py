from typing import Any, Dict, List, Optional
from datetime import datetime, date
from io import BytesIO
import os
import json
import logging
import re
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from pydantic import BaseModel
import openpyxl
from openai import OpenAI

logger = logging.getLogger(__name__)

from ....models import (
    Risk, RiskControlLink, RiskAssetLink, RiskEvidenceLink,
    RiskFrameworkControlLink, RiskGovernanceLink,
    NormalizedControl, FrameworkControl, ITAsset, Evidence,
    GovernanceObjective, Issue, GRCUser, Tenant, get_db,
    ParsedFrameworkControl, UploadedFramework, RiskMitigationAction,
    Vulnerability, VulnerabilityAssetLink, RiskAssessmentRisk,
    Team, BusinessUnit,
)
from ....schemas import (
    RiskCreate, RiskUpdate, RiskResponse,
    RiskAssessment, RiskTreatment,
    RiskControlLinkCreate, RiskAssetLinkCreate, RiskEvidenceLinkCreate,
    RiskFrameworkControlLinkCreate, RiskGovernanceLinkCreate,
    RiskDetailResponse, RiskHeatmapData, MessageResponse,
    RiskMitigationActionCreate, RiskMitigationActionResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/risks", tags=["ERM - Risk Register"])

UBL_TEMPLATE_REGISTER_TYPE = "UBL Template"
UBL_REGISTER_TYPE_ALIASES = {
    "ubl",
    "ubltemplate",
    "ublriskregister",
    "ublriskregistertemplate",
}
UBL_ALLOWED_CATEGORIES = {"technology", "third_party", "isms", "process", "other"}
UBL_LOCATION_OPTIONS = ("Bahrain", "International", "Pakistan", "Qatar", "UAE")
UBL_NON_EDITABLE_KEYS = {"risk_id", "source_sheet"}
UBL_PLATFORM_MAPPED_KEYS = {
    "likelihood_raw",
    "impact_raw",
    "risk_value_raw",
    "risk_level_raw",
    "residual_risk_raw",
    "status_raw",
    "mapped_category",
    "mapped_status",
    "inherent_score",
    "residual_score",
}
UBL_SUPPRESSED_EXTRA_KEYS = {
    "likelihood",
    "likelihood_of_vulnerability_exploitation",
    "impact",
    "impact_level",
    "risk_value",
    "risk_level",
    "risk_score",
    "inherent_risk_rating",
    "residual_risk",
    "residual_risk_if_controls_are_implemented",
    "residual_risks_after_implementation_of_controls",
    "status",
    "inherent_score",
    "residual_score",
}
_UBL_RISK_ID_PATTERN = re.compile(r"^\s*r\s*[-_ ]?\s*0*(\d+)\s*$", re.IGNORECASE)
_VULN_ID_PATTERN = re.compile(r"^\s*VULN-(\d+)\s*$", re.IGNORECASE)


def normalize_key(value: Optional[str]) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _resolve_team_to_business_unit(
    *,
    team_id: Optional[int],
    fallback_business_unit_id: Optional[int],
    tenant_id: int,
    db: Session,
) -> Optional[int]:
    """Map a Team selection to a BusinessUnit id, creating the BU on demand.

    Frontends now pick a Team (admin/teams) when assigning a risk; this
    helper translates that choice to a BusinessUnit row so the rest of
    the platform that joins on `business_unit_id` keeps working without
    any other code changes. Auto-mirroring matches by case-insensitive
    team name. When no team is selected, falls back to the caller's
    `fallback_business_unit_id` if supplied — useful for API callers
    that already know the BU id directly.
    """
    if team_id is None:
        return fallback_business_unit_id
    team = db.query(Team).filter(
        Team.id == team_id,
        Team.tenant_id == tenant_id,
    ).first()
    if not team:
        return fallback_business_unit_id
    name_key = (team.name or "").strip().lower()
    if not name_key:
        return fallback_business_unit_id
    # Reuse an existing BU with the same name when present; otherwise
    # create a fresh BU row mirroring the team.
    existing = db.query(BusinessUnit).filter(
        BusinessUnit.tenant_id == tenant_id,
    ).all()
    for bu in existing:
        if (bu.name or "").strip().lower() == name_key:
            return bu.id
    new_bu = BusinessUnit(tenant_id=tenant_id, name=team.name)
    db.add(new_bu)
    db.flush()
    return new_bu.id


def is_ubl_template_register_type(value: Optional[str]) -> bool:
    return normalize_key(value) in UBL_REGISTER_TYPE_ALIASES


def normalize_header(value: object) -> str:
    return " ".join("".join(ch if ch.isalnum() else " " for ch in str(value).lower()).split())


def parse_scale_1_to_5(val, default: Optional[int] = None) -> Optional[int]:
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return max(1, min(5, int(val)))

    text = str(val).strip().lower()
    if not text:
        return default

    num_match = re.search(r"([1-5])", text)
    if num_match:
        return int(num_match.group(1))

    keyword_map = [
        ("almost certain", 5),
        ("critical", 5),
        ("major", 4),
        ("likely", 4),
        ("high", 4),
        ("moderate", 3),
        ("medium", 3),
        ("possible", 3),
        ("minor", 2),
        ("unlikely", 2),
        ("low", 2),
        ("insignificant", 1),
        ("rare", 1),
    ]
    for keyword, score in keyword_map:
        if keyword in text:
            return score
    return default


def parse_numeric_score(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)

    text = str(val).strip()
    if not text or text.startswith("="):
        return None

    cleaned = text.replace(",", "")
    number_match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if number_match:
        try:
            return float(number_match.group(0))
        except Exception:
            return None
    return None


def parse_risk_level_to_score(val) -> Optional[float]:
    if val is None:
        return None
    text = str(val).strip().lower()
    if not text:
        return None
    if "critical" in text:
        return 20.0
    if "high" in text:
        return 15.0
    if "moderate" in text or "medium" in text:
        return 10.0
    if "low" in text:
        return 5.0
    return None


def parse_excel_date(val) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())

    text = str(val).strip()
    if not text:
        return None

    known_formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%b-%Y",
        "%d-%B-%Y",
    ]
    for fmt in known_formats:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def parse_ubl_risk_sequence(value: Optional[object]) -> Optional[int]:
    match = _UBL_RISK_ID_PATTERN.match(str(value or ""))
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def format_ubl_risk_id(seq: int) -> str:
    width = max(2, len(str(seq)))
    return f"R-{seq:0{width}d}"


def get_next_ubl_risk_sequence(tenant_id: int, db: Session) -> int:
    max_seq = 0
    rows = db.query(Risk.register_type, Risk.ubl_fields).filter(Risk.tenant_id == tenant_id).all()
    for register_type, ubl_fields in rows:
        if not is_ubl_template_register_type(register_type):
            continue
        payload: Dict[str, Any] = {}
        if isinstance(ubl_fields, dict):
            payload = ubl_fields
        elif isinstance(ubl_fields, str):
            try:
                parsed = json.loads(ubl_fields)
                if isinstance(parsed, dict):
                    payload = parsed
            except Exception:
                payload = {}
        seq = parse_ubl_risk_sequence(payload.get("risk_id"))
        if seq and seq > max_seq:
            max_seq = seq
    return max_seq + 1


def get_next_vulnerability_sequence(tenant_id: int, db: Session) -> int:
    max_seq = 0
    rows = db.query(Vulnerability.vuln_id).filter(Vulnerability.tenant_id == tenant_id).all()
    for (vuln_id,) in rows:
        match = _VULN_ID_PATTERN.match(str(vuln_id or ""))
        if not match:
            continue
        try:
            seq = int(match.group(1))
        except Exception:
            continue
        if seq > max_seq:
            max_seq = seq
    return max_seq + 1


def normalize_ubl_location(value: Optional[object]) -> Optional[str]:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = normalize_key(text)
    for allowed in UBL_LOCATION_OPTIONS:
        if normalize_key(allowed) == normalized:
            return allowed
    return text


def normalize_asset_criticality(value: Optional[object]) -> str:
    text = str(value or "").strip().lower()
    if "critical" in text:
        return "critical"
    if "high" in text:
        return "high"
    if "low" in text:
        return "low"
    return "medium"


def infer_vulnerability_severity(risk_level: Optional[object], score: Optional[float]) -> str:
    text = str(risk_level or "").strip().lower()
    if "critical" in text:
        return "critical"
    if "high" in text:
        return "high"
    if "moderate" in text or "medium" in text:
        return "medium"
    if "low" in text:
        return "low"
    if score is not None:
        if score >= 20:
            return "critical"
        if score >= 12:
            return "high"
        if score >= 6:
            return "medium"
        return "low"
    return "medium"


def sanitize_ubl_fields(payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    cleaned: Dict[str, Any] = {}
    for key, value in payload.items():
        if key in UBL_NON_EDITABLE_KEYS or key in UBL_PLATFORM_MAPPED_KEYS:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
        if value is None:
            continue
        cleaned[key] = value
    return cleaned


def resolve_ubl_category(
    requested_category: Optional[str],
    ubl_fields: Optional[Dict[str, Any]],
    default_category: str = "technology",
) -> str:
    base = (requested_category or "").strip().lower()
    if base not in UBL_ALLOWED_CATEGORIES:
        base = default_category
    raw_category = (ubl_fields or {}).get("risk_category_raw")
    mapped = map_ubl_category(raw_category, base)
    return mapped if mapped in UBL_ALLOWED_CATEGORIES else default_category


def map_ubl_category(raw_category: Optional[object], sheet_default_category: str) -> str:
    text = str(raw_category or "").strip().lower()
    if not text:
        return sheet_default_category

    if "third party" in text or "3rd party" in text or "vendor" in text or "supplier" in text or "outsourc" in text:
        return "third_party"
    if "process" in text:
        return "process"
    if "isms" in text:
        return "isms"
    if "technology" in text:
        return "technology"
    if "other" in text:
        return "other"
    return sheet_default_category


def map_uploaded_status(raw_status, implementation_status, mitigation_option, residual_score: Optional[float]) -> str:
    def _status_from_text(value: Optional[object]) -> Optional[str]:
        text = str(value or "").strip().lower()
        if not text:
            return None
        if "accept" in text:
            return "accepted"
        if "close" in text or "resolved" in text:
            return "closed"
        if any(k in text for k in ["implemented", "complete", "fixed", "mitigated"]):
            return "mitigated"
        if any(k in text for k in ["in progress", "ongoing", "treat", "mitigat", "transfer", "reduce"]):
            return "in_treatment"
        if any(k in text for k in ["open", "not implemented", "not started", "pending"]):
            return "open"
        return None

    for candidate in (raw_status, implementation_status, mitigation_option):
        mapped = _status_from_text(candidate)
        if mapped:
            return mapped

    if residual_score is not None and residual_score < 10:
        return "mitigated"
    return "open"


def find_header_row(ws, header_keywords: List[str], max_scan_rows: int = 30) -> int:
    scan_until = min(max_scan_rows, ws.max_row)
    for row_num in range(1, scan_until + 1):
        row_values = [cell.value for cell in ws[row_num]]
        row_text = " ".join(str(v).lower() for v in row_values if v is not None)
        matches = sum(1 for kw in header_keywords if kw in row_text)
        if matches >= 3:
            return row_num
    return 1


def import_ubl_risk_register(
    wb,
    tenant_id: int,
    register_type: Optional[str],
    db: Session,
    current_user: GRCUser,
):
    sheet_configs = [
        {"match": ["technology risk register"], "category": "technology"},
        {"match": ["3rdparty risk assessment", "3rd party risk assessment", "third party risk assessment"], "category": "third_party"},
        {"match": ["process risk register"], "category": "process"},
        {"match": ["isms"], "category": "isms"},
    ]

    available_sheets = [(sheet_name, normalize_header(sheet_name)) for sheet_name in wb.sheetnames]
    target_sheets = []
    for cfg in sheet_configs:
        for sheet_name, normalized_sheet in available_sheets:
            if any(pattern in normalized_sheet for pattern in cfg["match"]):
                target_sheets.append((sheet_name, cfg["category"]))
                break

    if not target_sheets:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="UBL template sheets not found. Expected sheets like Technology Risk Register, 3rdParty Risk Assessment, Process Risk Register, or ISMS.",
        )

    final_register_type = (register_type or "").strip() or UBL_TEMPLATE_REGISTER_TYPE
    next_ubl_risk_sequence = get_next_ubl_risk_sequence(tenant_id, db)
    next_vuln_sequence = get_next_vulnerability_sequence(tenant_id, db)
    created_count = 0
    auto_assets_count = 0
    auto_vulnerabilities_count = 0
    skipped_count = 0
    errors: List[str] = []
    asset_cache: Dict[str, ITAsset] = {}
    vulnerability_cache: Dict[str, Vulnerability] = {}

    def _json_value(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.isoformat()
        if isinstance(val, date):
            return datetime.combine(val, datetime.min.time()).isoformat()
        if isinstance(val, (int, float, bool)):
            return val
        text = str(val).strip()
        return text or None

    def _put(target: dict, key: str, val):
        parsed = _json_value(val)
        if parsed is not None:
            target[key] = parsed

    def get_or_create_asset(
        *,
        category: str,
        asset_name: Optional[object],
        ip_or_url: Optional[object],
        location: Optional[object],
        criticality: Optional[object],
        externally_exposed: Optional[object],
    ) -> Optional[ITAsset]:
        nonlocal auto_assets_count

        asset_name_text = str(asset_name or "").strip()
        ip_or_url_text = str(ip_or_url or "").strip()
        if not asset_name_text and not ip_or_url_text:
            return None

        cache_key = f"{normalize_key(asset_name_text)}|{normalize_key(ip_or_url_text)}|{normalize_key(location)}"
        if cache_key in asset_cache:
            return asset_cache[cache_key]

        existing: Optional[ITAsset] = None
        if asset_name_text:
            existing = db.query(ITAsset).filter(
                ITAsset.tenant_id == tenant_id,
                func.lower(func.coalesce(ITAsset.name, "")) == asset_name_text.lower(),
            ).first()
        if not existing and ip_or_url_text:
            existing = db.query(ITAsset).filter(
                ITAsset.tenant_id == tenant_id,
                func.lower(func.coalesce(ITAsset.ip_address, "")) == ip_or_url_text.lower(),
            ).first()
        if existing:
            asset_cache[cache_key] = existing
            return existing

        inferred_type = "third_party" if category == "third_party" else "application"
        created_asset = ITAsset(
            tenant_id=tenant_id,
            name=(asset_name_text or ip_or_url_text)[:255],
            description="Auto-created from UBL Template risk register upload",
            asset_type=inferred_type,
            owner_id=current_user.id,
            owner_name=(current_user.display_name or current_user.email),
            host_name=asset_name_text[:255] if asset_name_text else None,
            ip_address=ip_or_url_text[:50] if ip_or_url_text else None,
            criticality=normalize_asset_criticality(criticality),
            location=normalize_ubl_location(location),
            status="active",
            cde_environment=("yes" in str(externally_exposed or "").strip().lower()),
        )
        db.add(created_asset)
        asset_cache[cache_key] = created_asset
        auto_assets_count += 1
        return created_asset

    def get_or_create_vulnerability(
        *,
        vulnerability_text: Optional[object],
        threat_text: Optional[object],
        associated_risks_text: Optional[object],
        scenario_text: Optional[object],
        asset_name_text: Optional[object],
        ip_or_url_text: Optional[object],
        recommended_controls_text: Optional[object],
        risk_level_text: Optional[object],
        inherent_score_value: Optional[float],
        discovered_at: Optional[datetime],
        due_date: Optional[datetime],
    ) -> Optional[Vulnerability]:
        nonlocal next_vuln_sequence, auto_vulnerabilities_count

        title_source = vulnerability_text or threat_text or associated_risks_text or scenario_text
        title = str(title_source or "").strip()
        if not title:
            return None

        component_text = str(asset_name_text or "").strip()
        cache_key = f"{normalize_key(title)}|{normalize_key(component_text)}"
        if cache_key in vulnerability_cache:
            return vulnerability_cache[cache_key]

        existing = db.query(Vulnerability).filter(
            Vulnerability.tenant_id == tenant_id,
            func.lower(func.coalesce(Vulnerability.title, "")) == title.lower(),
            func.lower(func.coalesce(Vulnerability.affected_component, "")) == component_text.lower(),
        ).first()
        if existing:
            vulnerability_cache[cache_key] = existing
            return existing

        vuln_id = f"VULN-{next_vuln_sequence:05d}"
        next_vuln_sequence += 1

        description_parts = []
        if vulnerability_text:
            description_parts.append(f"Vulnerability: {vulnerability_text}")
        if threat_text:
            description_parts.append(f"Threat: {threat_text}")
        if scenario_text:
            description_parts.append(f"Scenario: {scenario_text}")
        if associated_risks_text:
            description_parts.append(f"Associated Risk: {associated_risks_text}")
        vuln_description = "\n\n".join(description_parts)[:8000] if description_parts else None

        ip_or_url = str(ip_or_url_text or "").strip()
        affected_url = ip_or_url if ip_or_url.lower().startswith(("http://", "https://")) else None

        created_vulnerability = Vulnerability(
            tenant_id=tenant_id,
            vuln_id=vuln_id,
            title=title[:500],
            description=vuln_description,
            severity=infer_vulnerability_severity(risk_level_text, inherent_score_value),
            cvss_score=None,
            affected_component=component_text[:255] if component_text else None,
            affected_host=component_text[:255] if component_text else None,
            affected_url=affected_url[:500] if affected_url else None,
            recommendation=(str(recommended_controls_text).strip()[:8000] if recommended_controls_text else None),
            status="open",
            discovered_at=discovered_at or datetime.utcnow(),
            due_date=due_date,
        )
        db.add(created_vulnerability)
        vulnerability_cache[cache_key] = created_vulnerability
        auto_vulnerabilities_count += 1
        return created_vulnerability

    header_keywords = [
        "risk id",
        "risk category",
        "likelihood",
        "impact",
        "risk value",
        "risk level",
        "status",
    ]

    for sheet_name, sheet_default_category in target_sheets:
        ws = wb[sheet_name]
        header_row = find_header_row(ws, header_keywords)
        headers = [cell.value for cell in ws[header_row]]
        header_map = {
            normalize_header(header): idx
            for idx, header in enumerate(headers)
            if header is not None and str(header).strip()
        }

        def get_value(row, *aliases):
            for alias in aliases:
                idx = header_map.get(normalize_header(alias))
                if idx is None or idx >= len(row):
                    continue
                value = row[idx]
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                return value
            return None

        for row_num, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(row):
                continue

            source_risk_id = get_value(row, "Risk ID#", "Risk ID", "Risk Id", "Risk ID ")
            risk_category_raw = get_value(row, "Risk Category", "Category")
            source = get_value(row, "Source", "Threat Source")
            sub_source = get_value(row, "Sub-Source  Activity", "Sub-Source Activity", "Activity", "Sub Source")
            location = normalize_ubl_location(get_value(row, "Location"))
            asset_name = get_value(row, "Application Name / Hostname / URL", "Assets", "Asset Name", "Application Name", "Asset")
            ip_or_url = get_value(row, "IP Address / URL", "IP Address", "URL")
            asset_criticality = get_value(row, "Asset Criticality")
            external_exposure = get_value(row, "Externally Exposed")
            vulnerability_count = get_value(row, "Count of Vulnerabilities")
            vulnerability = get_value(
                row,
                "Vulnerabilities Identified via Vulnerability Assessment",
                "Vulnerability",
                "Vulnerabilities",
            )
            threat = get_value(
                row,
                "Threat Due to Identified Vulnerability",
                "Threats Due to Identified Vulnerabilities",
                "Threats Due to Identified Vulnerabilities2",
                "Threat Source",
            )
            associated_risks = get_value(row, "Associated Risks")
            scenario = get_value(
                row,
                "Risk Description (Scenario-Based)",
                "Associated Risks",
                "Risk Description",
                "Key Risk Description",
            )
            business_impact = get_value(row, "Impact (Business / Regulatory / Financial)")
            recommended_controls = get_value(row, "Recommended Controls")
            mitigation_option = get_value(row, "Mitigation Option", "Risk Treatment Option")
            fixed_vulnerability_count = get_value(row, "Count of Fixed Vulnerabilities")
            frequency = get_value(row, "Frequancy", "Frequency")
            business_justification = get_value(row, "Business Justification")
            timelines = get_value(row, "Timelines")
            compensating_controls = get_value(row, "Compensating Controls")
            implementation_status = get_value(row, "Implementation Status")
            residual_risk = get_value(
                row,
                "Residual Risks (after implementation of controls)",
                "Residual Risk (If Controls are Implemented)",
                "Residual Risk",
            )
            row_status = get_value(row, "Status")
            mitigation_date = get_value(row, "Mitigation Date", "Date")
            reported_date = get_value(row, "Reported Date", "Date")
            likelihood_raw = get_value(row, "Likelihood of Vulnerability Exploitation", "Likelihood")
            impact_raw = get_value(row, "Impact Level", "Impact")
            risk_value_raw = get_value(row, "Risk Value", "Inherent Risk Rating", "Risk Score", "Inherent Risk Rating")
            risk_level_raw = get_value(row, "Risk Level", "Inherent Risk Rating")
            risk_owner = get_value(row, "Risk Owner", "Responsibility", "Owner")
            type_value = get_value(row, "Type", "Security Triad")
            cia_impacted = get_value(row, "CIA Impacted")
            annex_a = get_value(row, "Annex A")

            has_minimum_data = any(
                [
                    source_risk_id,
                    scenario,
                    threat,
                    vulnerability,
                    asset_name,
                    recommended_controls,
                    risk_category_raw,
                ]
            )
            if not has_minimum_data:
                skipped_count += 1
                continue

            raw_header_values: dict = {}
            for raw_header, idx in header_map.items():
                if idx >= len(row):
                    continue
                value = _json_value(row[idx])
                if value is None:
                    continue
                raw_header_values[raw_header] = value

            generated_risk_id = format_ubl_risk_id(next_ubl_risk_sequence)

            title_source = associated_risks or scenario or threat or vulnerability or asset_name or generated_risk_id or "Imported UBL Risk"
            title = str(title_source).replace("\n", " ").strip()
            if asset_name and title and str(asset_name).strip().lower() not in title.lower():
                title = f"{str(asset_name).strip()} - {title}"
            title = title[:200] if title else "Imported UBL Risk"

            description_lines = []
            description_lines.append(f"Risk ID: {generated_risk_id}")
            if source_risk_id:
                description_lines.append(f"Source Risk ID: {source_risk_id}")
            if source:
                description_lines.append(f"Source: {source}")
            if sub_source:
                description_lines.append(f"Activity/Sub-Source: {sub_source}")
            if location:
                description_lines.append(f"Location: {location}")
            if asset_name:
                description_lines.append(f"Asset/Application: {asset_name}")
            if ip_or_url:
                description_lines.append(f"IP/URL: {ip_or_url}")
            if asset_criticality:
                description_lines.append(f"Asset Criticality: {asset_criticality}")
            if external_exposure:
                description_lines.append(f"Externally Exposed: {external_exposure}")
            if type_value:
                description_lines.append(f"Type/Security Triad: {type_value}")
            if vulnerability:
                description_lines.append(f"Vulnerability: {vulnerability}")
            if threat:
                description_lines.append(f"Threat: {threat}")
            if scenario:
                description_lines.append(f"Scenario: {scenario}")
            if business_impact:
                description_lines.append(f"Business Impact: {business_impact}")
            if cia_impacted:
                description_lines.append(f"CIA Impacted: {cia_impacted}")
            if annex_a:
                description_lines.append(f"Annex A: {annex_a}")
            if risk_owner:
                description_lines.append(f"Risk Owner: {risk_owner}")
            description = "\n".join(description_lines)[:8000]

            inherent_likelihood = parse_scale_1_to_5(likelihood_raw, default=3) or 3
            inherent_impact = parse_scale_1_to_5(impact_raw, default=3) or 3
            inherent_score = parse_numeric_score(risk_value_raw)
            if inherent_score is None:
                inherent_score = parse_risk_level_to_score(risk_level_raw)
            if inherent_score is None:
                inherent_score = float(inherent_likelihood * inherent_impact)

            residual_score = parse_numeric_score(residual_risk)
            if residual_score is None:
                residual_score = parse_risk_level_to_score(residual_risk)

            category = map_ubl_category(risk_category_raw, sheet_default_category)
            if category not in UBL_ALLOWED_CATEGORIES:
                category = "other"

            status_value = map_uploaded_status(
                raw_status=row_status,
                implementation_status=implementation_status,
                mitigation_option=mitigation_option,
                residual_score=residual_score,
            )

            treatment_plan_parts = []
            if recommended_controls:
                treatment_plan_parts.append(f"Recommended Controls: {recommended_controls}")
            if mitigation_option:
                treatment_plan_parts.append(f"Mitigation Option: {mitigation_option}")
            if compensating_controls:
                treatment_plan_parts.append(f"Compensating Controls: {compensating_controls}")
            if business_justification:
                treatment_plan_parts.append(f"Business Justification: {business_justification}")
            if timelines:
                treatment_plan_parts.append(f"Timeline: {timelines}")
            treatment_plan = "\n\n".join(treatment_plan_parts) if treatment_plan_parts else None

            due_date = parse_excel_date(mitigation_date) or parse_excel_date(timelines)
            review_date = parse_excel_date(reported_date)

            risk_sub_category = str(sub_source).strip() if sub_source else None
            if risk_sub_category and len(risk_sub_category) > 100:
                risk_sub_category = risk_sub_category[:100]

            ubl_fields: dict = {"source_sheet": sheet_name, "risk_id": generated_risk_id}
            _put(ubl_fields, "source_risk_id", source_risk_id)
            _put(ubl_fields, "risk_category_raw", risk_category_raw)
            _put(ubl_fields, "source", source)
            _put(ubl_fields, "sub_source_activity", sub_source)
            _put(ubl_fields, "location", location)
            _put(ubl_fields, "type_or_security_triad", type_value)
            _put(ubl_fields, "application_name_or_asset", asset_name)
            _put(ubl_fields, "ip_or_url", ip_or_url)
            _put(ubl_fields, "asset_criticality", asset_criticality)
            _put(ubl_fields, "externally_exposed", external_exposure)
            _put(ubl_fields, "vulnerability_count", vulnerability_count)
            _put(ubl_fields, "vulnerabilities_identified", vulnerability)
            _put(ubl_fields, "threat_due_to_vulnerability", threat)
            _put(ubl_fields, "associated_risks", associated_risks)
            _put(ubl_fields, "risk_description_scenario", scenario)
            _put(ubl_fields, "impact_business_regulatory_financial", business_impact)
            _put(ubl_fields, "recommended_controls", recommended_controls)
            _put(ubl_fields, "reported_date", reported_date)
            _put(ubl_fields, "mitigation_option", mitigation_option)
            _put(ubl_fields, "fixed_vulnerability_count", fixed_vulnerability_count)
            _put(ubl_fields, "frequency", frequency)
            _put(ubl_fields, "business_justification", business_justification)
            _put(ubl_fields, "timeline", timelines)
            _put(ubl_fields, "compensating_controls", compensating_controls)
            _put(ubl_fields, "implementation_status", implementation_status)
            _put(ubl_fields, "mitigation_date", mitigation_date)
            _put(ubl_fields, "risk_owner", risk_owner)
            _put(ubl_fields, "cia_impacted", cia_impacted)
            _put(ubl_fields, "annex_a", annex_a)
            for raw_key, raw_value in raw_header_values.items():
                safe_key = re.sub(r"[^a-z0-9]+", "_", raw_key.lower()).strip("_")
                if (
                    not safe_key
                    or safe_key in ubl_fields
                    or safe_key in UBL_PLATFORM_MAPPED_KEYS
                    or safe_key in UBL_SUPPRESSED_EXTRA_KEYS
                ):
                    continue
                _put(ubl_fields, safe_key, raw_value)

            try:
                db_risk = Risk(
                    tenant_id=tenant_id,
                    title=title or "Imported UBL Risk",
                    description=description,
                    category=category,
                    risk_category=category,
                    risk_sub_category=risk_sub_category,
                    register_type=final_register_type,
                    ubl_fields=ubl_fields,
                    inherent_likelihood=inherent_likelihood,
                    inherent_impact=inherent_impact,
                    inherent_score=inherent_score,
                    residual_score=residual_score,
                    treatment_plan=treatment_plan,
                    status=status_value,
                    due_date=due_date,
                    review_date=review_date,
                    owner_id=current_user.id,
                )
                db.add(db_risk)

                asset = get_or_create_asset(
                    category=category,
                    asset_name=asset_name,
                    ip_or_url=ip_or_url,
                    location=location,
                    criticality=asset_criticality,
                    externally_exposed=external_exposure,
                )
                if asset:
                    db_risk.asset_links.append(RiskAssetLink(asset=asset))
                    _put(ubl_fields, "linked_asset_name", getattr(asset, "name", None))

                vulnerability_entry = get_or_create_vulnerability(
                    vulnerability_text=vulnerability,
                    threat_text=threat,
                    associated_risks_text=associated_risks,
                    scenario_text=scenario,
                    asset_name_text=asset_name,
                    ip_or_url_text=ip_or_url,
                    recommended_controls_text=recommended_controls,
                    risk_level_text=risk_level_raw,
                    inherent_score_value=inherent_score,
                    discovered_at=review_date,
                    due_date=due_date,
                )
                if vulnerability_entry and asset:
                    has_link = False
                    if getattr(vulnerability_entry, "id", None) and getattr(asset, "id", None):
                        has_link = db.query(VulnerabilityAssetLink).filter(
                            VulnerabilityAssetLink.vulnerability_id == vulnerability_entry.id,
                            VulnerabilityAssetLink.asset_id == asset.id,
                        ).first() is not None
                    if not has_link:
                        vulnerability_entry.asset_links.append(
                            VulnerabilityAssetLink(
                                asset=asset,
                                notes="Linked from UBL Template risk register import",
                                created_by=current_user.id,
                                link_source="nca_bridge",
                                auto_linked=True,
                            )
                        )
                if vulnerability_entry:
                    _put(ubl_fields, "linked_vulnerability_id", getattr(vulnerability_entry, "vuln_id", None))

                next_ubl_risk_sequence += 1
                created_count += 1
            except Exception as e:
                errors.append(f"{sheet_name} row {row_num}: {str(e)}")

    db.commit()
    return {
        "message": f"Successfully imported {created_count} risks",
        "created": created_count,
        "skipped": skipped_count,
        "assets_created": auto_assets_count,
        "vulnerabilities_created": auto_vulnerabilities_count,
        "errors": errors[:10] if errors else [],
    }


def resolve_risk_category(category: Optional[str], risk_category: Optional[str]) -> str:
    primary = (category or "").strip().lower()
    secondary = (risk_category or "").strip().lower()
    return primary or secondary or "operational"


def calculate_risk_score(likelihood: int, impact: int) -> float:
    return likelihood * impact


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


@router.get("", response_model=List[RiskResponse])
def list_risks(
    tenant_id: Optional[int] = None,
    category: Optional[str] = None,
    register_type: Optional[str] = None,
    status_filter: Optional[str] = None,
    min_score: Optional[float] = None,
    max_score: Optional[float] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)

    normalized_risk_category = func.replace(
        func.replace(
            func.replace(
                func.replace(
                    func.lower(func.coalesce(Risk.risk_category, Risk.category, "")),
                    " ",
                    ""
                ),
                "_",
                ""
            ),
            "-",
            ""
        ),
        "/",
        ""
    )
    normalized_risk_register_type = func.replace(
        func.replace(
            func.replace(
                func.replace(
                    func.lower(func.coalesce(Risk.register_type, "")),
                    " ",
                    ""
                ),
                "_",
                ""
            ),
            "-",
            ""
        ),
        "/",
        ""
    )

    if category:
        category_value = re.sub(r"[^a-z0-9]", "", category.strip().lower())
        query = query.filter(normalized_risk_category == category_value)
    if register_type:
        register_type_value = re.sub(r"[^a-z0-9]", "", register_type.strip().lower())
        query = query.filter(normalized_risk_register_type == register_type_value)
    if status_filter:
        query = query.filter(Risk.status == status_filter)
    if min_score is not None:
        query = query.filter(Risk.inherent_score >= min_score)
    if max_score is not None:
        query = query.filter(Risk.inherent_score <= max_score)
    
    risks = query.order_by(Risk.created_at.desc()).offset(skip).limit(limit).all()
    return risks


@router.post("", response_model=RiskResponse, status_code=status.HTTP_201_CREATED)
def create_risk(
    risk: RiskCreate,
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    tenant = db.query(Tenant).filter(Tenant.id == tenant_id).first()
    if not tenant:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tenant not found"
        )
    
    is_ubl_register = is_ubl_template_register_type(risk.register_type)
    sanitized_ubl_fields = sanitize_ubl_fields(risk.ubl_fields if isinstance(risk.ubl_fields, dict) else None) if is_ubl_register else (
        risk.ubl_fields if isinstance(risk.ubl_fields, dict) else None
    )

    resolved_category = resolve_risk_category(risk.category, getattr(risk, 'risk_category', None))
    if is_ubl_register:
        resolved_category = resolve_ubl_category(
            requested_category=resolved_category,
            ubl_fields=sanitized_ubl_fields if isinstance(sanitized_ubl_fields, dict) else None,
            default_category="technology",
        )
    inherent_score = risk.inherent_score
    if inherent_score is None and risk.inherent_likelihood and risk.inherent_impact:
        inherent_score = calculate_risk_score(risk.inherent_likelihood, risk.inherent_impact)

    residual_score = risk.residual_score
    if residual_score is None and risk.residual_likelihood and risk.residual_impact:
        residual_score = calculate_risk_score(risk.residual_likelihood, risk.residual_impact)

    risk_sub_category = risk.risk_sub_category
    if is_ubl_register and isinstance(sanitized_ubl_fields, dict):
        ubl_sub_source = str(sanitized_ubl_fields.get("sub_source_activity") or "").strip()
        if ubl_sub_source:
            risk_sub_category = ubl_sub_source[:100]
        sanitized_ubl_fields["risk_id"] = format_ubl_risk_id(get_next_ubl_risk_sequence(tenant_id, db))

    # Resolve team assignment to a BusinessUnit row. Teams (admin/teams)
    # are the canonical "department" source; we auto-mirror the team to
    # a BusinessUnit (matched by case-insensitive name) so the rest of
    # the platform that joins on business_unit_id keeps working.
    resolved_bu_id = _resolve_team_to_business_unit(
        team_id=getattr(risk, "team_id", None),
        fallback_business_unit_id=getattr(risk, "business_unit_id", None),
        tenant_id=tenant_id,
        db=db,
    )

    db_risk = Risk(
        tenant_id=tenant_id,
        title=risk.title,
        description=risk.description,
        category=resolved_category,
        risk_category=resolved_category,
        risk_sub_category=risk_sub_category,
        register_type=risk.register_type,
        ubl_fields=sanitized_ubl_fields if isinstance(sanitized_ubl_fields, dict) else None,
        owner_id=risk.owner_id,
        business_owner_id=risk.business_owner_id,
        business_unit_id=resolved_bu_id,
        affected_department_ids=risk.affected_department_ids or [],
        inherent_likelihood=risk.inherent_likelihood,
        inherent_impact=risk.inherent_impact,
        inherent_score=inherent_score,
        residual_likelihood=risk.residual_likelihood,
        residual_impact=risk.residual_impact,
        residual_score=residual_score,
        risk_appetite=risk.risk_appetite,
        status=risk.status or "open",
        treatment_plan=risk.treatment_plan,
        closure_status=risk.closure_status,
    )
    db.add(db_risk)
    db.commit()
    db.refresh(db_risk)
    return db_risk


@router.get("/dashboard")
def get_risk_dashboard(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {
            "total_risks": 0,
            "by_category": {},
            "by_status": {},
            "by_score_range": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "avg_inherent_score": 0,
            "avg_residual_score": 0,
            "open_risks": 0,
            "risks_needing_review": 0
        }
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    by_category = {}
    by_status = {}
    by_score_range = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    total_inherent_score = 0
    total_residual_score = 0
    risks_with_score = 0
    
    for risk in risks:
        cat = risk.risk_category or risk.category or "operational"
        by_category[cat] = by_category.get(cat, 0) + 1
        
        status_val = risk.status or "open"
        by_status[status_val] = by_status.get(status_val, 0) + 1
        
        score = risk.residual_score or risk.inherent_score or 0
        if score >= 20:
            by_score_range["critical"] += 1
        elif score >= 12:
            by_score_range["high"] += 1
        elif score >= 6:
            by_score_range["medium"] += 1
        else:
            by_score_range["low"] += 1
        
        if risk.inherent_score:
            total_inherent_score += risk.inherent_score
            risks_with_score += 1
        if risk.residual_score:
            total_residual_score += risk.residual_score
    
    return {
        "total_risks": len(risks),
        "by_category": by_category,
        "by_status": by_status,
        "by_score_range": by_score_range,
        "avg_inherent_score": round(total_inherent_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "avg_residual_score": round(total_residual_score / risks_with_score, 1) if risks_with_score > 0 else 0,
        "open_risks": by_status.get("open", 0),
        "risks_needing_review": sum(1 for r in risks if r.review_date and r.review_date < datetime.utcnow())
    }


RISK_STATUS_BUCKETS = ("open", "in_treatment", "mitigated", "accepted", "closed")


def _bucket_status(raw: Optional[str]) -> str:
    """Normalize Risk.status into one of the five canonical UI buckets.

    Canonical: open, in_treatment, mitigated, accepted, closed.
    Anything unrecognized falls back to "open" so it's still surfaced rather
    than silently dropped from the dashboard counts.
    """
    s = (raw or "open").strip().lower().replace(" ", "_")
    if s in ("closed", "resolved", "cancelled", "canceled"):
        return "closed"
    if s in ("in_treatment", "treating", "under_treatment", "in_progress", "under_review", "mitigating"):
        # "mitigating" reads as in-progress treatment; reserve "mitigated" for completed
        return "in_treatment" if s != "mitigating" else "in_treatment"
    if s == "mitigated":
        return "mitigated"
    if s == "accepted":
        return "accepted"
    return "open"


@router.get("/dashboard/by-register")
def get_risk_dashboard_by_register(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Per-register-type breakdown for the Risk Register dashboard.

    Returns one entry per distinct register_type (NULL bucketed as "Standard"),
    each with totals, status mix (open/in_progress/closed/pending), category
    mix, and assignee workload (top 10 owners by count + total contributors).
    """
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"registers": [], "total_risks": 0}

    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)

    risks = query.all()

    # Pre-fetch owner names in one round-trip
    owner_ids = {r.owner_id for r in risks if r.owner_id}
    owners_by_id: Dict[int, str] = {}
    if owner_ids:
        for u in db.query(GRCUser).filter(GRCUser.id.in_(owner_ids)).all():
            owners_by_id[u.id] = (getattr(u, "display_name", None) or getattr(u, "username", None) or getattr(u, "email", None) or f"User #{u.id}")

    buckets: Dict[str, Dict[str, Any]] = {}
    for risk in risks:
        key = (risk.register_type or "Standard").strip() or "Standard"
        b = buckets.setdefault(key, {
            "register_type": key,
            "total": 0,
            "by_status": {s: 0 for s in RISK_STATUS_BUCKETS},
            "by_category": {},
            "by_owner": {},
            "by_score_range": {"critical": 0, "high": 0, "medium": 0, "low": 0},
            "owner_ids": set(),
            "avg_residual_score": 0.0,
            "_residual_total": 0.0,
            "_residual_count": 0,
        })
        b["total"] += 1
        b["by_status"][_bucket_status(risk.status)] += 1
        cat = risk.risk_category or risk.category or "operational"
        b["by_category"][cat] = b["by_category"].get(cat, 0) + 1
        if risk.owner_id:
            label = owners_by_id.get(risk.owner_id, f"User #{risk.owner_id}")
            b["by_owner"][label] = b["by_owner"].get(label, 0) + 1
            b["owner_ids"].add(risk.owner_id)
        score = risk.residual_score or risk.inherent_score or 0
        if score >= 20:
            b["by_score_range"]["critical"] += 1
        elif score >= 12:
            b["by_score_range"]["high"] += 1
        elif score >= 6:
            b["by_score_range"]["medium"] += 1
        else:
            b["by_score_range"]["low"] += 1
        if risk.residual_score:
            b["_residual_total"] += float(risk.residual_score)
            b["_residual_count"] += 1

    out = []
    for key, b in buckets.items():
        owners_sorted = sorted(b["by_owner"].items(), key=lambda kv: kv[1], reverse=True)[:10]
        avg = (b["_residual_total"] / b["_residual_count"]) if b["_residual_count"] else 0.0
        out.append({
            "register_type": b["register_type"],
            "total": b["total"],
            "by_status": b["by_status"],
            "by_category": b["by_category"],
            "by_score_range": b["by_score_range"],
            "top_owners": [{"owner": name, "count": cnt} for name, cnt in owners_sorted],
            "contributors": len(b["owner_ids"]),
            "avg_residual_score": round(avg, 2),
        })
    out.sort(key=lambda x: x["total"], reverse=True)
    return {"registers": out, "total_risks": len(risks)}


def _infer_source_type(risk: Risk, assessed_risk_ids: set) -> str:
    """Best-effort provenance for risks created before source_type was tracked.

    Reads only signals already on the row (register_type, presence in an
    assessment join) — no DB writes. Once the writers start populating
    source_type on new risks, those values win and this fallback is skipped.
    """
    explicit = (getattr(risk, "source_type", None) or "").strip()
    if explicit:
        return explicit
    rt = (risk.register_type or "").strip().lower()
    if rt in ("ubl template", "ubltemplate", "ubl"):
        return "ubl_import"
    if rt in ("nca template", "ncatemplate", "nca"):
        return "nca_import"
    if risk.id in assessed_risk_ids:
        return "assessment"
    if rt:
        # Any other named register (PCI-DSS, ISO 27001, SOX, NIST, GDPR, …)
        return "register_import"
    return "manual"


@router.get("/dashboard/by-source")
def get_risk_dashboard_by_source(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Provenance dashboard — counts grouped by source_type with status mix.

    For risks created before provenance tracking, source_type is inferred
    from register_type and assessment membership rather than shown as
    "Unspecified". Explicit source_type on the row always wins.
    """
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"sources": [], "total_risks": 0}

    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)

    risks = query.all()

    # Single batched lookup of every risk_id that appears in any assessment,
    # so the per-row inference is O(1).
    assessed_risk_ids: set = set()
    if risks:
        risk_ids = [r.id for r in risks]
        rows = (
            db.query(RiskAssessmentRisk.risk_id)
            .filter(RiskAssessmentRisk.risk_id.in_(risk_ids))
            .distinct()
            .all()
        )
        assessed_risk_ids = {row[0] for row in rows}

    buckets: Dict[str, Dict[str, Any]] = {}
    for risk in risks:
        key = _infer_source_type(risk, assessed_risk_ids)
        b = buckets.setdefault(key, {
            "source_type": key,
            "total": 0,
            "by_status": {s: 0 for s in RISK_STATUS_BUCKETS},
        })
        b["total"] += 1
        b["by_status"][_bucket_status(risk.status)] += 1

    sources = sorted(buckets.values(), key=lambda x: x["total"], reverse=True)
    return {"sources": sources, "total_risks": len(risks)}


@router.get("/heatmap")
def get_risk_heatmap(
    risk_type: Optional[str] = None,
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.all()
    
    heatmap = {}
    risk_type_prefix = risk_type if risk_type in ["inherent", "residual"] else "inherent"
    
    for risk in risks:
        likelihood = getattr(risk, f"{risk_type_prefix}_likelihood") or 0
        impact = getattr(risk, f"{risk_type_prefix}_impact") or 0
        
        if likelihood > 0 and impact > 0:
            key = f"{likelihood}-{impact}"
            if key not in heatmap:
                heatmap[key] = {"likelihood": likelihood, "impact": impact, "count": 0, "risks": []}
            heatmap[key]["count"] += 1
            heatmap[key]["risks"].append({
                "id": risk.id,
                "title": risk.title,
                "score": getattr(risk, f"{risk_type_prefix}_score")
            })
    
    return list(heatmap.values())


@router.get("/{risk_id}/detail")
def get_risk_detail(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links).joinedload(RiskControlLink.normalized_control),
        joinedload(Risk.asset_links).joinedload(RiskAssetLink.asset),
        joinedload(Risk.evidence_links).joinedload(RiskEvidenceLink.evidence),
        joinedload(Risk.framework_control_links).joinedload(RiskFrameworkControlLink.framework_control),
        joinedload(Risk.governance_links).joinedload(RiskGovernanceLink.governance_objective),
        joinedload(Risk.owner)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    linked_controls = []
    for link in risk.control_links:
        if link.normalized_control:
            linked_controls.append({
                "id": link.id,
                "control_id": link.normalized_control.id,
                "code": link.normalized_control.code,
                "name": link.normalized_control.name
            })
    
    linked_framework_controls = []
    for link in risk.framework_control_links:
        if link.framework_control:
            linked_framework_controls.append({
                "id": link.id,
                "framework_control_id": link.framework_control.id,
                "code": link.framework_control.code,
                "name": link.framework_control.name,
                "mitigation_effectiveness": link.mitigation_effectiveness,
                "notes": link.notes
            })
    
    linked_assets = []
    for link in risk.asset_links:
        if link.asset:
            linked_assets.append({
                "id": link.id,
                "asset_id": link.asset.id,
                "name": link.asset.name,
                "asset_type": link.asset.asset_type
            })
    
    linked_evidence = []
    for link in risk.evidence_links:
        if link.evidence:
            linked_evidence.append({
                "id": link.id,
                "evidence_id": link.evidence.id,
                "name": link.evidence.name,
                "status": link.evidence.status
            })
    
    linked_governance = []
    for link in risk.governance_links:
        if link.governance_objective:
            linked_governance.append({
                "id": link.id,
                "governance_objective_id": link.governance_objective.id,
                "name": link.governance_objective.name,
                "impact_level": link.impact_level
            })
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "risk_sub_category": risk.risk_sub_category,
        "register_type": risk.register_type,
        "ubl_fields": risk.ubl_fields,
        "owner_id": risk.owner_id,
        "owner_name": risk.owner.display_name if risk.owner else None,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "due_date": risk.due_date.isoformat() if risk.due_date else None,
        "review_date": risk.review_date.isoformat() if risk.review_date else None,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": linked_controls,
        "linked_framework_controls": linked_framework_controls,
        "linked_assets": linked_assets,
        "linked_evidence": linked_evidence,
        "linked_governance": linked_governance
    }


@router.get("/{risk_id}", response_model=dict)
def get_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).options(
        joinedload(Risk.control_links),
        joinedload(Risk.asset_links),
        joinedload(Risk.evidence_links)
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    return {
        "id": risk.id,
        "tenant_id": risk.tenant_id,
        "title": risk.title,
        "description": risk.description,
        "category": risk.category,
        "risk_category": risk.risk_category,
        "risk_sub_category": risk.risk_sub_category,
        "register_type": risk.register_type,
        "ubl_fields": risk.ubl_fields,
        "owner_id": risk.owner_id,
        "inherent_likelihood": risk.inherent_likelihood,
        "inherent_impact": risk.inherent_impact,
        "inherent_score": risk.inherent_score,
        "residual_likelihood": risk.residual_likelihood,
        "residual_impact": risk.residual_impact,
        "residual_score": risk.residual_score,
        "risk_appetite": risk.risk_appetite,
        "status": risk.status,
        "treatment_plan": risk.treatment_plan,
        "created_at": risk.created_at.isoformat(),
        "updated_at": risk.updated_at.isoformat(),
        "linked_controls": [link.normalized_control_id for link in risk.control_links],
        "linked_assets": [link.asset_id for link in risk.asset_links],
        "linked_evidence": [link.evidence_id for link in risk.evidence_links]
    }


@router.put("/{risk_id}", response_model=RiskResponse)
def update_risk(
    risk_id: int,
    risk_update: RiskUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    update_data = risk_update.model_dump(exclude_unset=True)
    target_register_type = update_data.get("register_type", risk.register_type)
    is_target_ubl = is_ubl_template_register_type(target_register_type)

    if is_target_ubl:
        existing_ubl_fields = risk.ubl_fields if isinstance(risk.ubl_fields, dict) else {}
        incoming_ubl_fields = update_data.get("ubl_fields") if "ubl_fields" in update_data else existing_ubl_fields
        sanitized_ubl_fields = sanitize_ubl_fields(incoming_ubl_fields if isinstance(incoming_ubl_fields, dict) else None)

        existing_risk_id = (existing_ubl_fields or {}).get("risk_id")
        if parse_ubl_risk_sequence(existing_risk_id) is None:
            existing_risk_id = format_ubl_risk_id(get_next_ubl_risk_sequence(risk.tenant_id, db))
        sanitized_ubl_fields["risk_id"] = str(existing_risk_id)
        if existing_ubl_fields.get("source_sheet"):
            sanitized_ubl_fields["source_sheet"] = existing_ubl_fields.get("source_sheet")
        update_data["ubl_fields"] = sanitized_ubl_fields

        mapped_category = resolve_ubl_category(
            requested_category=resolve_risk_category(
                update_data.get("category", risk.category),
                update_data.get("risk_category", risk.risk_category),
            ),
            ubl_fields=sanitized_ubl_fields,
            default_category="technology",
        )
        update_data["category"] = mapped_category
        update_data["risk_category"] = mapped_category

        if "risk_sub_category" not in update_data:
            ubl_sub_source = str(sanitized_ubl_fields.get("sub_source_activity") or "").strip()
            if ubl_sub_source:
                update_data["risk_sub_category"] = ubl_sub_source[:100]
    elif "category" in update_data or "risk_category" in update_data:
        update_data["category"] = resolve_risk_category(
            update_data.get("category", risk.category),
            update_data.get("risk_category", risk.risk_category)
        )
        update_data["risk_category"] = update_data["category"]

    # team_id is a transient mapping field — resolve it to a real
    # business_unit_id and don't try to setattr it onto the ORM model
    # (Risk has no `team_id` column today).
    if "team_id" in update_data:
        team_id_value = update_data.pop("team_id")
        resolved = _resolve_team_to_business_unit(
            team_id=team_id_value,
            fallback_business_unit_id=update_data.get("business_unit_id"),
            tenant_id=risk.tenant_id,
            db=db,
        )
        if resolved is not None:
            update_data["business_unit_id"] = resolved
        else:
            update_data.pop("business_unit_id", None)

    for field, value in update_data.items():
        setattr(risk, field, value)

    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.delete("/{risk_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    db.delete(risk)
    db.commit()
    return None


@router.post("/{risk_id}/assess", response_model=RiskResponse)
def assess_risk(
    risk_id: int,
    assessment: RiskAssessment,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.inherent_likelihood = assessment.inherent_likelihood
    risk.inherent_impact = assessment.inherent_impact
    risk.inherent_score = calculate_risk_score(
        assessment.inherent_likelihood,
        assessment.inherent_impact
    )
    
    if assessment.residual_likelihood and assessment.residual_impact:
        risk.residual_likelihood = assessment.residual_likelihood
        risk.residual_impact = assessment.residual_impact
        risk.residual_score = calculate_risk_score(
            assessment.residual_likelihood,
            assessment.residual_impact
        )
    
    if assessment.risk_appetite:
        risk.risk_appetite = assessment.risk_appetite
    
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/treatment", response_model=RiskResponse)
def add_treatment_plan(
    risk_id: int,
    treatment: RiskTreatment,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    risk.treatment_plan = treatment.treatment_plan
    risk.status = "mitigating"
    risk.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/controls", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_control(
    risk_id: int,
    link: RiskControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    control = db.query(NormalizedControl).filter(
        NormalizedControl.id == link.normalized_control_id
    ).first()
    if not control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control not found"
        )
    
    existing = db.query(RiskControlLink).filter(
        RiskControlLink.risk_id == risk_id,
        RiskControlLink.normalized_control_id == link.normalized_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskControlLink(
        risk_id=risk_id,
        normalized_control_id=link.normalized_control_id
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Control linked successfully")


@router.post("/{risk_id}/link-framework-control", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_framework_control(
    risk_id: int,
    link: RiskFrameworkControlLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    framework_control = db.query(FrameworkControl).filter(
        FrameworkControl.id == link.framework_control_id
    ).first()
    if not framework_control:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control not found"
        )
    
    existing = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.risk_id == risk_id,
        RiskFrameworkControlLink.framework_control_id == link.framework_control_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskFrameworkControlLink(
        risk_id=risk_id,
        framework_control_id=link.framework_control_id,
        mitigation_effectiveness=link.mitigation_effectiveness,
        notes=link.notes
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Framework control linked successfully")


@router.delete("/{risk_id}/link-framework-control/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_framework_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskFrameworkControlLink).filter(
        RiskFrameworkControlLink.id == link_id,
        RiskFrameworkControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Framework control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/link-governance", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_governance(
    risk_id: int,
    link: RiskGovernanceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    governance_objective = db.query(GovernanceObjective).filter(
        GovernanceObjective.id == link.governance_objective_id,
        GovernanceObjective.tenant_id.in_(user_tenants)
    ).first()
    if not governance_objective:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance objective not found"
        )
    
    existing = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.risk_id == risk_id,
        RiskGovernanceLink.governance_objective_id == link.governance_objective_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskGovernanceLink(
        risk_id=risk_id,
        governance_objective_id=link.governance_objective_id,
        impact_level=link.impact_level
    )
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Governance objective linked successfully")


@router.delete("/{risk_id}/link-governance/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_governance(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskGovernanceLink).filter(
        RiskGovernanceLink.id == link_id,
        RiskGovernanceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Governance link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.delete("/{risk_id}/controls/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_control(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskControlLink).filter(
        RiskControlLink.id == link_id,
        RiskControlLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Control link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/assets", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_asset(
    risk_id: int,
    link: RiskAssetLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    asset = db.query(ITAsset).filter(
        ITAsset.id == link.asset_id,
        ITAsset.tenant_id.in_(user_tenants)
    ).first()
    if not asset:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset not found"
        )
    
    existing = db.query(RiskAssetLink).filter(
        RiskAssetLink.risk_id == risk_id,
        RiskAssetLink.asset_id == link.asset_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskAssetLink(risk_id=risk_id, asset_id=link.asset_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Asset linked successfully")


@router.delete("/{risk_id}/assets/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_asset(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskAssetLink).filter(
        RiskAssetLink.id == link_id,
        RiskAssetLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Asset link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.post("/{risk_id}/evidence", response_model=MessageResponse, status_code=status.HTTP_201_CREATED)
def link_risk_to_evidence(
    risk_id: int,
    link: RiskEvidenceLinkCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    evidence = db.query(Evidence).filter(
        Evidence.id == link.evidence_id,
        Evidence.tenant_id.in_(user_tenants)
    ).first()
    if not evidence:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence not found"
        )
    
    existing = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.risk_id == risk_id,
        RiskEvidenceLink.evidence_id == link.evidence_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Link already exists"
        )
    
    db_link = RiskEvidenceLink(risk_id=risk_id, evidence_id=link.evidence_id)
    db.add(db_link)
    db.commit()
    
    return MessageResponse(message="Evidence linked successfully")


@router.delete("/{risk_id}/evidence/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
def unlink_risk_from_evidence(
    risk_id: int,
    link_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    link = db.query(RiskEvidenceLink).filter(
        RiskEvidenceLink.id == link_id,
        RiskEvidenceLink.risk_id == risk_id
    ).first()
    if not link:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evidence link not found"
        )
    
    db.delete(link)
    db.commit()
    return None


@router.get("/template/download")
def download_risk_register_template(
    _current_user: GRCUser = Depends(require_auth),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    headers = [
        "Ref",
        "Risk Title",
        "Risk Description",
        "Risk Category",
        "Risk Sub Category",
        "Register Type",
        "Asset Name",
        "Threat",
        "Vulnerability",
        "Likelihood",
        "Impact",
        "Risk Score",
        "Post-Treatment Likelihood",
        "Post-Treatment Impact",
        "Residual Risk",
        "Risk Treatment Option",
        "Mitigating Action Controls",
        "Action Plan",
        "Responsibility",
        "Gaps",
        "Recommendations",
    ]
    ws.append(headers)

    sample_row = [
        "R-001",
        "Privileged Access Misconfiguration",
        "Admin access is not consistently restricted and reviewed.",
        "Technology",
        "Cybersecurity",
        "ISO 27001",
        "Identity Platform",
        "Unauthorized privileged access",
        "No periodic privileged account recertification",
        4,
        5,
        20,
        3,
        4,
        12,
        "Mitigate",
        "MFA, PAM, and quarterly recertification",
        "Implement PAM and enforce quarterly access reviews",
        "Security Team",
        "Legacy accounts not monitored",
        "Automate recertification and alerting",
    ]
    ws.append(sample_row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=risk_register_template.xlsx"},
    )


@router.post("/upload")
async def upload_risk_register(
    file: UploadFile = File(...),
    register_type: Optional[str] = Query(None),
    tenant_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User is not assigned to any tenant"
            )
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))

        if is_ubl_template_register_type(register_type):
            return import_ubl_risk_register(
                wb=wb,
                tenant_id=tenant_id,
                register_type=register_type,
                db=db,
                current_user=current_user,
            )
        
        ws = None
        for sheet_name in ['Risk Assessment', 'Risks', 'Risk Register', 'Sheet1']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        
        headers = []
        header_row = 1
        header_keywords = [
            'asset name', 'threat', 'likelihood', 'impact', 'risk score',
            'risk title', 'risk category', 'residual score', 'status'
        ]
        for row_num in range(1, 10):
            row_values = [cell.value for cell in ws[row_num]]
            row_str = ' '.join([str(v).lower() for v in row_values if v])
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches >= 3:
                headers = row_values
                header_row = row_num
                break
        
        def normalize_header(value: str) -> str:
            return " ".join("".join(ch if ch.isalnum() else " " for ch in str(value).lower()).split())

        header_map = {}
        for idx, h in enumerate(headers):
            if h:
                header_map[normalize_header(h)] = idx
        
        def get_value(row, *possible_names):
            for name in possible_names:
                normalized_name = normalize_header(name)
                if normalized_name in header_map:
                    idx = header_map[normalized_name]
                    if idx < len(row):
                        return row[idx]
            return None
        
        def parse_int(val, default=1):
            if val is None:
                return default
            if isinstance(val, (int, float)):
                return max(1, min(5, int(val)))
            try:
                return max(1, min(5, int(float(str(val).strip()))))
            except:
                return default
        
        def parse_score(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and val.startswith('='):
                return None
            try:
                return float(val)
            except:
                return None
        
        def map_category(threat_or_category):
            if not threat_or_category:
                return 'operational'
            text = str(threat_or_category).lower()
            if any(w in text for w in ['strategic', 'business', 'market']):
                return 'strategic'
            if any(w in text for w in ['financial', 'money', 'cost', 'budget']):
                return 'financial'
            if any(w in text for w in ['compliance', 'regulatory', 'legal', 'pci', 'gdpr', 'privacy', 'lawfulness', 'data subject']):
                return 'compliance'
            if any(w in text for w in ['technology', 'system', 'network', 'cyber', 'malware', 'phishing', 'security']):
                return 'technology'
            if any(w in text for w in ['vendor', 'supplier', 'third', 'partner', 'outsourcing']):
                return 'third_party'
            return 'operational'
        
        def map_status(treatment_option, residual_score):
            if not treatment_option:
                return 'open'
            text = str(treatment_option).lower()
            if 'accept' in text:
                return 'accepted'
            if 'avoid' in text or 'close' in text:
                return 'closed'
            if 'mitigat' in text or 'reduc' in text or 'treat' in text:
                if residual_score and residual_score < 10:
                    return 'mitigated'
                return 'in_treatment'
            if 'transfer' in text:
                return 'in_treatment'
            return 'open'
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(row):
                continue
            
            ref = get_value(row, 'ref', 'ref.', 'id', 'risk id', 'risk_id')
            risk_title = get_value(row, 'risk title', 'title', 'risk name', 'risk scenario', 'scenario')
            risk_description = get_value(row, 'risk description', 'description', 'details')
            risk_category_value = get_value(row, 'risk category', 'category', 'domain', 'threat / category', 'threat category')
            asset_name = get_value(row, 'asset name', 'asset', 'asset_name')
            threat = get_value(row, 'threat', 'threat description')
            vulnerability = get_value(row, 'vulnerabilities', 'vulnerability', 'vuln')
            
            if not risk_title and not asset_name and not threat and not vulnerability and not risk_description:
                skipped_count += 1
                continue
            
            title_parts = []
            if risk_title:
                title_parts.append(str(risk_title).strip())
            if asset_name:
                title_parts.append(str(asset_name).strip())
            if threat:
                threat_clean = str(threat).strip().replace('\n', ' ')[:80]
                if threat_clean:
                    title_parts.append(threat_clean)
            
            if not title_parts:
                if vulnerability:
                    title_parts.append(str(vulnerability).strip()[:80])
                elif ref:
                    title_parts.append(str(ref))
            
            if not title_parts:
                skipped_count += 1
                continue
            
            title = " - ".join(title_parts)[:200]
            
            description_parts = []
            if risk_description:
                description_parts.append(str(risk_description).strip())
            if threat:
                description_parts.append(f"Threat: {threat}")
            if vulnerability:
                description_parts.append(f"Vulnerability: {vulnerability}")
            gaps = get_value(row, 'gaps', 'gap')
            if gaps:
                description_parts.append(f"Gaps: {gaps}")
            recommendations = get_value(row, 'recommendations', 'recommendation')
            if recommendations:
                description_parts.append(f"Recommendations: {recommendations}")
            
            description = "\n\n".join(description_parts) if description_parts else None
            
            inherent_likelihood = parse_int(get_value(row, 'likelihood', 'inherent likelihood', 'probability'))
            inherent_impact = parse_int(get_value(row, 'impact', 'inherent impact', 'consequence'))
            inherent_score = parse_score(get_value(row, 'risk score', 'inherent score', 'inherent risk'))
            if inherent_score is None:
                inherent_score = inherent_likelihood * inherent_impact
            
            residual_likelihood = parse_int(get_value(row, 'post-treatment likelihood', 'residual likelihood'), default=None)
            residual_impact = parse_int(get_value(row, 'post-treatment impact', 'residual impact'), default=None)
            residual_score = parse_score(get_value(row, 'residual risk', 'residual score', 'post-treatment risk'))
            if residual_score is None and residual_likelihood and residual_impact:
                residual_score = residual_likelihood * residual_impact
            
            mitigating_controls = get_value(row, 'mitigating action controls', 'controls', 'existing controls', 'mitigating controls')
            action_plan = get_value(row, 'action plan', 'treatment plan', 'plan')
            treatment_parts = []
            if mitigating_controls:
                treatment_parts.append(f"Existing Controls: {mitigating_controls}")
            if action_plan:
                treatment_parts.append(f"Action Plan: {action_plan}")
            treatment_plan = "\n\n".join(treatment_parts) if treatment_parts else None
            
            treatment_option = get_value(row, 'risk treatment option', 'treatment option', 'treatment')
            row_register_type = get_value(row, 'register type', 'framework', 'risk type')
            final_register_type = (
                register_type.strip() if register_type and register_type.strip()
                else (str(row_register_type).strip() if row_register_type else None)
            )

            category = map_category(risk_category_value or threat or risk_title or final_register_type)

            row_status = get_value(row, 'status', 'risk status')
            if row_status:
                status_text = str(row_status).strip().lower().replace(' ', '_').replace('-', '_')
                allowed_statuses = {'open', 'in_treatment', 'mitigated', 'accepted', 'closed'}
                risk_status = status_text if status_text in allowed_statuses else map_status(treatment_option, residual_score)
            else:
                risk_status = map_status(treatment_option, residual_score)
            
            owner_name = get_value(row, 'responsibility', 'owner', 'risk owner')
            
            try:
                db_risk = Risk(
                    tenant_id=tenant_id,
                    title=title,
                    description=description,
                    category=category,
                    risk_category=category,
                    inherent_likelihood=inherent_likelihood,
                    inherent_impact=inherent_impact,
                    inherent_score=inherent_score,
                    residual_likelihood=residual_likelihood if residual_likelihood else None,
                    residual_impact=residual_impact if residual_impact else None,
                    residual_score=residual_score,
                    treatment_plan=treatment_plan,
                    register_type=final_register_type,
                    status=risk_status,
                    owner_id=current_user.id
                )
                db.add(db_risk)
                created_count += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Successfully imported {created_count} risks",
            "created": created_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse Excel file: {str(e)}"
        )


@router.get("/aging")
def get_risks_with_aging(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(Risk).filter(Risk.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Risk.tenant_id == tenant_id)
    
    risks = query.order_by(Risk.created_at.asc()).all()
    
    now = datetime.utcnow()
    result = []
    for risk in risks:
        days_since_created = (now - risk.created_at).days if risk.created_at else 0
        days_since_updated = (now - risk.updated_at).days if risk.updated_at else days_since_created
        
        result.append({
            "id": risk.id,
            "title": risk.title,
            "category": risk.risk_category or risk.category,
            "status": risk.status,
            "inherent_score": risk.inherent_score,
            "residual_score": risk.residual_score,
            "created_at": risk.created_at.isoformat() if risk.created_at else None,
            "updated_at": risk.updated_at.isoformat() if risk.updated_at else None,
            "days_since_created": days_since_created,
            "days_since_updated": days_since_updated,
            "owner_id": risk.owner_id
        })
    
    return result


@router.post("/{risk_id}/close", response_model=RiskResponse)
def close_risk(
    risk_id: int,
    closure_notes: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    if risk.status == "closed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk is already closed"
        )
    
    risk.status = "closed"
    risk.closure_status = "closed"
    risk.closed_at = datetime.utcnow()
    risk.closed_by = current_user.id
    risk.closure_notes = closure_notes
    risk.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(risk)
    return risk


@router.post("/{risk_id}/reopen", response_model=RiskResponse)
def reopen_risk(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )
    
    if risk.status != "closed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk is not closed"
        )
    
    risk.status = "open"
    risk.closure_status = None
    risk.closed_at = None
    risk.closed_by = None
    risk.closure_notes = None
    risk.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(risk)
    return risk


class RiskAISuggestionRequest(BaseModel):
    name: str
    category: Optional[str] = None
    sub_category: Optional[str] = None
    description: Optional[str] = None


class RecommendedControl(BaseModel):
    control_id: int
    control_name: str
    control_code: Optional[str] = None
    relevance: str
    rationale: str


class RiskAISuggestionResponse(BaseModel):
    suggested_description: str
    suggested_causes: List[str]
    suggested_consequences: List[str]
    recommended_controls: List[RecommendedControl]
    suggested_likelihood: int
    suggested_impact: int
    risk_treatment_options: List[str]


def get_openai_client() -> OpenAI:
    api_key = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    is_modelfarm = base_url and "modelfarm" in base_url
    if not api_key and not is_modelfarm:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI features unavailable. OpenAI API key not configured."
        )
    if not is_modelfarm and (api_key.startswith("_DUMMY") or api_key == "your-api-key-here" or len(api_key) < 20):
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI features unavailable. OpenAI API key not configured."
        )
    return OpenAI(
        api_key=api_key,
        base_url=base_url
    )


def parse_ai_response(response_text: str) -> dict:
    try:
        cleaned = response_text.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        return json.loads(cleaned.strip())
    except json.JSONDecodeError:
        return {
            "suggested_description": "Unable to generate description",
            "suggested_causes": [],
            "suggested_consequences": [],
            "suggested_likelihood": 3,
            "suggested_impact": 3,
            "risk_treatment_options": ["Mitigate", "Accept", "Transfer"]
        }


def get_available_controls_for_matching(db: Session, user_tenants: List[int]) -> str:
    controls_info = []
    
    normalized_controls = db.query(NormalizedControl).limit(100).all()
    for ctrl in normalized_controls:
        controls_info.append(f"- ID:{ctrl.id} | Code:{ctrl.code} | Name:{ctrl.name}")
    
    for tenant_id in user_tenants:
        parsed_controls = db.query(ParsedFrameworkControl).join(
            UploadedFramework
        ).filter(
            (UploadedFramework.tenant_id == tenant_id) | 
            (UploadedFramework.is_shared == True),
            UploadedFramework.is_active == True
        ).limit(50).all()
        
        for ctrl in parsed_controls:
            ctrl_code = ctrl.original_reference or ctrl.control_id
            controls_info.append(f"- ID:{ctrl.id} | Code:{ctrl_code} | Title:{ctrl.title[:80] if ctrl.title else 'N/A'}")
    
    return "\n".join(controls_info[:100])


RISK_AI_SUGGESTION_PROMPT = """You are an expert Enterprise Risk Management (ERM) consultant with 20+ years of experience. Analyze the risk information provided and generate comprehensive suggestions.

RISK INFORMATION:
Name: {name}
Category: {category}
Sub-category: {sub_category}
Existing Description: {description}

AVAILABLE CONTROLS FOR RECOMMENDATION (select the most relevant ones):
{available_controls}

Based on this risk, provide suggestions in the following JSON format:
{{
    "suggested_description": "<A comprehensive 2-4 sentence professional risk description that explains what the risk is, its context, and potential business impact>",
    
    "suggested_causes": [
        "<Root cause 1 - specific and actionable>",
        "<Root cause 2 - specific and actionable>",
        "<Root cause 3 - specific and actionable>"
    ],
    
    "suggested_consequences": [
        "<Business consequence 1 - specific impact on operations, finances, reputation, or compliance>",
        "<Business consequence 2 - specific impact>",
        "<Business consequence 3 - specific impact>"
    ],
    
    "recommended_control_ids": [
        {{
            "control_id": <ID number from the available controls list>,
            "relevance": "<high|medium|low>",
            "rationale": "<Explain why this control helps mitigate this specific risk>"
        }}
    ],
    
    "suggested_likelihood": <1-5 scale where 1=Rare, 2=Unlikely, 3=Possible, 4=Likely, 5=Almost Certain>,
    
    "suggested_impact": <1-5 scale where 1=Negligible, 2=Minor, 3=Moderate, 4=Major, 5=Catastrophic>,
    
    "risk_treatment_options": [
        "<Primary treatment recommendation: Mitigate/Accept/Transfer/Avoid>",
        "<Alternative treatment option>",
        "<Supporting action>"
    ]
}}

GUIDELINES:
1. Base likelihood and impact on industry standards for the risk category
2. Select 2-4 most relevant controls from the provided list
3. Be specific and actionable in causes and consequences
4. Match the professional tone expected in enterprise GRC systems
5. ONLY recommend controls that exist in the provided list - use exact IDs

Return ONLY valid JSON, no additional text."""


@router.get("/{risk_id}/mitigation-actions", response_model=List[RiskMitigationActionResponse])
def get_risk_mitigation_actions(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    risk = db.query(Risk).filter(Risk.id == risk_id, Risk.tenant_id.in_(user_tenants)).first()
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    actions = db.query(RiskMitigationAction).filter(
        RiskMitigationAction.risk_id == risk_id
    ).order_by(RiskMitigationAction.created_at.desc()).all()
    return actions


@router.post("/{risk_id}/mitigation-actions", response_model=RiskMitigationActionResponse, status_code=status.HTTP_201_CREATED)
def create_risk_mitigation_action(
    risk_id: int,
    data: RiskMitigationActionCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    risk = db.query(Risk).filter(Risk.id == risk_id, Risk.tenant_id.in_(user_tenants)).first()
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    action = RiskMitigationAction(
        risk_id=risk_id,
        title=data.title,
        description=data.description,
        action_type=data.action_type,
        priority=data.priority,
        owner_id=data.owner_id,
        due_date=data.due_date,
        expected_residual_reduction=data.expected_residual_reduction,
        notes=data.notes,
        status="pending",
    )
    db.add(action)
    db.commit()
    db.refresh(action)
    return action


@router.post("/ai-suggest", response_model=RiskAISuggestionResponse)
def get_risk_ai_suggestions(
    request: RiskAISuggestionRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if not request.name or len(request.name.strip()) < 3:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Risk name must be at least 3 characters"
        )
    
    user_tenants = get_user_tenants(current_user, db)
    
    try:
        client = get_openai_client()
    except HTTPException:
        return RiskAISuggestionResponse(
            suggested_description=f"Risk related to {request.name} in the {request.category or 'operational'} category.",
            suggested_causes=["Process failure", "Human error", "Inadequate controls"],
            suggested_consequences=["Operational disruption", "Financial loss", "Reputational damage"],
            recommended_controls=[],
            suggested_likelihood=3,
            suggested_impact=3,
            risk_treatment_options=["Mitigate", "Accept", "Transfer"]
        )
    
    available_controls = get_available_controls_for_matching(db, user_tenants)
    
    prompt = RISK_AI_SUGGESTION_PROMPT.format(
        name=request.name,
        category=request.category or "Not specified",
        sub_category=request.sub_category or "Not specified",
        description=request.description or "Not provided",
        available_controls=available_controls if available_controls else "No controls available"
    )
    
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an expert ERM consultant. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1500
        )
        
        ai_response = parse_ai_response(response.choices[0].message.content)
        
        recommended_controls = []
        ai_control_recs = ai_response.get("recommended_control_ids", [])
        
        for rec in ai_control_recs[:5]:
            control_id = rec.get("control_id")
            if not control_id:
                continue
            
            normalized_ctrl = db.query(NormalizedControl).filter(
                NormalizedControl.id == control_id
            ).first()
            
            if normalized_ctrl:
                recommended_controls.append(RecommendedControl(
                    control_id=normalized_ctrl.id,
                    control_name=normalized_ctrl.name,
                    control_code=normalized_ctrl.code,
                    relevance=rec.get("relevance", "medium"),
                    rationale=rec.get("rationale", "Relevant to this risk category")
                ))
            else:
                parsed_ctrl = db.query(ParsedFrameworkControl).filter(
                    ParsedFrameworkControl.id == control_id
                ).first()
                
                if parsed_ctrl:
                    recommended_controls.append(RecommendedControl(
                        control_id=parsed_ctrl.id,
                        control_name=parsed_ctrl.title or "Control",
                        control_code=parsed_ctrl.original_reference or parsed_ctrl.control_id,
                        relevance=rec.get("relevance", "medium"),
                        rationale=rec.get("rationale", "Relevant to this risk category")
                    ))
        
        return RiskAISuggestionResponse(
            suggested_description=ai_response.get("suggested_description", f"Risk related to {request.name}"),
            suggested_causes=ai_response.get("suggested_causes", [])[:5],
            suggested_consequences=ai_response.get("suggested_consequences", [])[:5],
            recommended_controls=recommended_controls,
            suggested_likelihood=min(5, max(1, ai_response.get("suggested_likelihood", 3))),
            suggested_impact=min(5, max(1, ai_response.get("suggested_impact", 3))),
            risk_treatment_options=ai_response.get("risk_treatment_options", ["Mitigate", "Accept", "Transfer"])[:4]
        )
        
    except Exception as e:
        logger.error(f"AI suggestion error: {str(e)}")
        return RiskAISuggestionResponse(
            suggested_description=f"Risk related to {request.name} requiring assessment and mitigation.",
            suggested_causes=["Process failure", "External factors", "Resource constraints"],
            suggested_consequences=["Operational impact", "Financial impact", "Compliance impact"],
            recommended_controls=[],
            suggested_likelihood=3,
            suggested_impact=3,
            risk_treatment_options=["Mitigate", "Accept", "Transfer"]
        )


class AITreatmentPlanResponse(BaseModel):
    treatment_plan: str


@router.post("/{risk_id}/ai-treatment-plan", response_model=AITreatmentPlanResponse)
def generate_ai_treatment_plan(
    risk_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    risk = db.query(Risk).options(
        joinedload(Risk.control_links).joinedload(RiskControlLink.normalized_control),
        joinedload(Risk.framework_control_links).joinedload(RiskFrameworkControlLink.framework_control),
    ).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()

    if not risk:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Risk not found"
        )

    linked_controls = []
    for link in risk.control_links:
        if link.normalized_control:
            linked_controls.append(f"{link.normalized_control.code}: {link.normalized_control.name}")
    for link in risk.framework_control_links:
        if link.framework_control:
            linked_controls.append(f"{link.framework_control.code}: {link.framework_control.name}")

    controls_text = "\n".join(f"- {c}" for c in linked_controls) if linked_controls else "No controls currently linked."

    prompt = f"""You are an enterprise risk management expert. Generate a detailed, actionable treatment plan for the following risk.

Risk Title: {risk.title}
Description: {risk.description or 'N/A'}
Category: {risk.risk_category or risk.category or 'N/A'}
Inherent Likelihood: {risk.inherent_likelihood or 'N/A'}/5
Inherent Impact: {risk.inherent_impact or 'N/A'}/5
Inherent Score: {risk.inherent_score or 'N/A'}
Residual Likelihood: {risk.residual_likelihood or 'N/A'}/5
Residual Impact: {risk.residual_impact or 'N/A'}/5
Residual Score: {risk.residual_score or 'N/A'}
Current Status: {risk.status or 'N/A'}

Linked Controls:
{controls_text}

Generate a comprehensive treatment plan that includes:
1. Treatment Strategy (mitigate, transfer, accept, or avoid - with justification)
2. Specific Action Items (3-5 concrete steps with responsible parties and timelines)
3. Control Improvements (enhancements to existing controls or new controls needed)
4. Monitoring & Review (KPIs, review frequency, escalation triggers)
5. Expected Residual Risk (target likelihood and impact after treatment)

Write the plan in clear, professional language suitable for a risk committee. Return ONLY the treatment plan text, no JSON formatting."""

    try:
        client = get_openai_client()

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an enterprise risk management expert who generates detailed, actionable risk treatment plans."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1500
        )

        treatment_plan = response.choices[0].message.content.strip()
        return AITreatmentPlanResponse(treatment_plan=treatment_plan)

    except Exception as e:
        logger.error(f"AI treatment plan generation error: {str(e)}")
        category = risk.risk_category or risk.category or "operational"
        return AITreatmentPlanResponse(
            treatment_plan=f"""Treatment Plan for: {risk.title}

1. Treatment Strategy: Mitigate
   Reduce risk through enhanced controls and monitoring.

2. Action Items:
   - Conduct detailed risk assessment and root cause analysis (Week 1-2)
   - Implement additional preventive controls specific to {category} risks (Week 2-4)
   - Establish monitoring procedures and key risk indicators (Week 3-4)
   - Train relevant staff on updated procedures (Week 4-6)
   - Conduct effectiveness review (Week 8)

3. Control Improvements:
   - Review and strengthen existing control framework
   - Add detective controls for early warning
   - Implement automated monitoring where feasible

4. Monitoring & Review:
   - Monthly KRI reporting
   - Quarterly treatment plan review
   - Immediate escalation if risk materializes

5. Expected Residual Risk:
   - Target Likelihood: {max(1, (risk.inherent_likelihood or 3) - 1)}/5
   - Target Impact: {max(1, (risk.inherent_impact or 3) - 1)}/5"""
        )
