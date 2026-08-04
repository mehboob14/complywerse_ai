from ....config import get_openai_api_key, get_openai_model

from typing import List, Optional
from datetime import datetime, timedelta
from io import BytesIO
import re
import os
import json
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_  # Used in other routers filtering
from pydantic import BaseModel

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ....models import (
    Risk, RiskKRI, RiskKRIMeasurement, GRCUser, get_db
)
from ....schemas import (
    RiskKRICreate, RiskKRIUpdate, RiskKRIResponse,
    RiskKRIMeasurementCreate, RiskKRIMeasurementResponse,
    MessageResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant
from ....services import kri_feeds, metric_catalog

router = APIRouter(prefix="/kris", tags=["ERM - Key Risk Indicators"])


class KRIAISuggestRequest(BaseModel):
    name: str
    description: Optional[str] = None
    risk_id: Optional[int] = None
    metric_hint: Optional[str] = None


def _get_openai_client():
    from openai import OpenAI

    api_key = get_openai_api_key()
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    if not api_key:
        return None
    is_modelfarm = "modelfarm" in (base_url or "")
    if not is_modelfarm and (api_key.startswith("_DUMMY") or api_key == "your-api-key-here" or len(api_key) < 20):
        return None
    return OpenAI(api_key=api_key, base_url=base_url)


def _fallback_kri_suggestion(name: str, description: Optional[str] = None) -> dict:
    text = f"{name} {description or ''}".lower()
    metric_type = "numeric"
    unit = None
    threshold_direction = "lower_is_better"
    frequency = "monthly"
    green_threshold = 5.0
    amber_threshold = 10.0

    if any(k in text for k in ["%", "percent", "rate", "ratio"]):
        metric_type = "percentage"
        unit = "%"
        green_threshold = 2.0
        amber_threshold = 5.0
    elif any(k in text for k in ["count", "incidents", "tickets", "cases"]):
        metric_type = "count"
        unit = "count"
        green_threshold = 10.0
        amber_threshold = 20.0

    if any(k in text for k in ["uptime", "availability", "compliance"]):
        threshold_direction = "higher_is_better"
        if metric_type == "percentage":
            green_threshold = 99.0
            amber_threshold = 97.0

    if any(k in text for k in ["daily", "real-time", "realtime"]):
        frequency = "daily"
    elif "weekly" in text:
        frequency = "weekly"
    elif "quarterly" in text:
        frequency = "quarterly"

    return {
        "description": description or f"KRI tracking for {name}",
        "metric_type": metric_type,
        "unit": unit,
        "threshold_direction": threshold_direction,
        "frequency": frequency,
        "green_threshold": green_threshold,
        "amber_threshold": amber_threshold,
        "suggested_name": name,
        "data_source": "AI-suggested",
        "rationale": "Heuristic suggestion generated from indicator name/description"
    }


def _ai_suggest_kri_payload(name: str, description: Optional[str] = None, risk_context: Optional[str] = None, metric_hint: Optional[str] = None) -> dict:
    client = _get_openai_client()
    if not client:
        return _fallback_kri_suggestion(name, description)

    prompt = f"""You are a GRC and ERM specialist. Suggest KRI setup values for manual KRI entry.
Return ONLY valid JSON with keys:
suggested_name, description, metric_type, unit, threshold_direction, frequency, green_threshold, amber_threshold, data_source, rationale

KRI Name: {name or 'Not provided'}
Description: {description or 'Not provided'}
Metric Hint: {metric_hint or 'Not provided'}
Risk Context: {risk_context or 'Not provided'}

Constraints:
- suggested_name: concise KRI indicator name that fits the risk context (e.g. "Number of Critical Incidents", "IT Uptime Rate")
- metric_type in [numeric, percentage, count, boolean]
- threshold_direction in [lower_is_better, higher_is_better]
- frequency in [daily, weekly, monthly, quarterly, annually]
- green_threshold and amber_threshold as numbers where applicable
"""

    try:
        response = client.chat.completions.create(
            model=get_openai_model(),
            messages=[
                {"role": "system", "content": "You are a precise assistant. Return JSON only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2,
            max_tokens=900
        )
        content = (response.choices[0].message.content or "").strip()
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        parsed = json.loads(content.strip())
        fallback = _fallback_kri_suggestion(name, description)
        fallback.update({k: v for k, v in parsed.items() if v is not None})
        return fallback
    except Exception:
        return _fallback_kri_suggestion(name, description)


def calculate_kri_status(value: float, kri: RiskKRI) -> str:
    if kri.green_threshold is None or kri.amber_threshold is None:
        return "unknown"
    
    if kri.threshold_direction == "lower_is_better":
        if value <= kri.green_threshold:
            return "green"
        elif value <= kri.amber_threshold:
            return "amber"
        else:
            return "red"
    else:
        if value >= kri.green_threshold:
            return "green"
        elif value >= kri.amber_threshold:
            return "amber"
        else:
            return "red"


def _hydrate(db: Session, kri: RiskKRI, tenant_ids: List[int]) -> RiskKRIResponse:
    """Build the response for a KRI, resolving its value LIVE from the platform
    metric layer when it's bound to a metric_key, then computing RAG status."""
    data = RiskKRIResponse.model_validate(kri)
    value = kri.current_value
    metric_key = getattr(kri, "metric_key", None)
    if metric_key:
        data.is_live = True
        meta = kri_feeds.catalog_meta(metric_key)
        if meta:
            data.module = meta.module
            data.module_label = metric_catalog.MODULE_LABELS.get(meta.module, meta.module)
            if not data.unit:
                data.unit = meta.unit
        live = kri_feeds.current_value(db, tenant_ids, metric_key)
        if live is not None:
            value = live
            data.current_value = live
    if value is not None:
        data.current_status = calculate_kri_status(value, kri)
    return data


@router.get("", response_model=List[RiskKRIResponse])
def list_kris(
    risk_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    is_active: Optional[bool] = True,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    kri_feeds.ensure_kri_columns(db)

    # Scope directly on RiskKRI.tenant_id (backfilled from the parent risk) so
    # standalone / live / uploaded KRIs are included, not just risk-linked ones.
    query = db.query(RiskKRI).filter(RiskKRI.tenant_id.in_(user_tenants))

    if risk_id:
        query = query.filter(RiskKRI.risk_id == risk_id)
    if is_active is not None:
        query = query.filter(RiskKRI.is_active == is_active)

    kris = query.offset(skip).limit(limit).all()
    result = [_hydrate(db, kri, user_tenants) for kri in kris]

    if status_filter:
        result = [k for k in result if k.current_status == status_filter]

    return result


@router.post("", response_model=RiskKRIResponse, status_code=status.HTTP_201_CREATED)
def create_kri(
    kri: RiskKRICreate,
    ai_assist: bool = False,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    tenant_id = get_user_primary_tenant(current_user, db)
    kri_feeds.ensure_kri_columns(db)

    # Frontend historically sent risk_id=0 when no risk was selected.
    risk_id = kri.risk_id if kri.risk_id and kri.risk_id > 0 else None

    if risk_id:
        risk = db.query(Risk).filter(
            Risk.id == risk_id,
            Risk.tenant_id.in_(user_tenants)
        ).first()
        if not risk:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Risk not found"
            )

    enriched = None
    if ai_assist:
        risk_context = None
        if risk_id:
            linked_risk = db.query(Risk).filter(
                Risk.id == risk_id,
                Risk.tenant_id.in_(user_tenants)
            ).first()
            if linked_risk:
                risk_context = f"{linked_risk.title} | category={linked_risk.category} | status={linked_risk.status}"
        enriched = _ai_suggest_kri_payload(kri.name, kri.description, risk_context=risk_context)

    description = kri.description or (enriched.get("description") if enriched else None)
    metric_type = kri.metric_type or (enriched.get("metric_type") if enriched else "numeric")
    unit = kri.unit or (enriched.get("unit") if enriched else None)
    green_threshold = kri.green_threshold if kri.green_threshold is not None else (enriched.get("green_threshold") if enriched else None)
    amber_threshold = kri.amber_threshold if kri.amber_threshold is not None else (enriched.get("amber_threshold") if enriched else None)
    threshold_direction = kri.threshold_direction or (enriched.get("threshold_direction") if enriched else "lower_is_better")
    frequency = kri.frequency or (enriched.get("frequency") if enriched else "monthly")
    data_source = kri.data_source or (enriched.get("data_source") if enriched else None)
    
    db_kri = RiskKRI(
        risk_id=risk_id,
        tenant_id=tenant_id,
        metric_key=kri.metric_key,
        name=kri.name,
        description=description,
        metric_type=metric_type,
        unit=unit,
        green_threshold=green_threshold,
        amber_threshold=amber_threshold,
        threshold_direction=threshold_direction,
        frequency=frequency,
        data_source=data_source,
        owner_id=kri.owner_id,
        kind=kri.kind or "kri",
        category=kri.category,
        formula=kri.formula,
        target=kri.target,
        reporting_period=kri.reporting_period,
        next_due_date=kri.next_due_date,
        data_provider_id=kri.data_provider_id,
        reviewer_id=kri.reviewer_id,
        linked_control_ids=kri.linked_control_ids or [],
        linked_objective_ids=kri.linked_objective_ids or [],
        linked_framework_id=kri.linked_framework_id,
    )
    db.add(db_kri)
    db.commit()
    db.refresh(db_kri)
    return _hydrate(db, db_kri, user_tenants)


@router.post("/ai-suggest")
def suggest_kri_with_ai(
    request: KRIAISuggestRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    risk_context = None

    if request.risk_id:
        risk = db.query(Risk).filter(
            Risk.id == request.risk_id,
            Risk.tenant_id.in_(user_tenants)
        ).first()
        if not risk:
            raise HTTPException(status_code=404, detail="Risk not found")
        risk_context = f"{risk.title} | category={risk.category} | status={risk.status}"

    suggestion = _ai_suggest_kri_payload(
        name=request.name,
        description=request.description,
        risk_context=risk_context,
        metric_hint=request.metric_hint
    )

    return {
        "name": request.name,
        "suggestion": suggestion
    }


@router.get("/alerts", response_model=List[RiskKRIResponse])
def get_kri_alerts(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    kri_feeds.ensure_kri_columns(db)
    kris = db.query(RiskKRI).filter(
        RiskKRI.tenant_id.in_(user_tenants),
        RiskKRI.is_active == True,
    ).all()

    alerts = [_hydrate(db, kri, user_tenants) for kri in kris]
    alerts = [k for k in alerts if k.current_status in ("red", "amber")]
    return sorted(alerts, key=lambda x: 0 if x.current_status == "red" else 1)


@router.get("/metric-options")
def kri_metric_options(current_user: GRCUser = Depends(require_auth)):
    """The catalog of platform metrics a KRI can bind to for a live value."""
    return {"metrics": kri_feeds.metric_options()}


@router.get("/templates")
def kri_templates(current_user: GRCUser = Depends(require_auth)):
    """Ready-made KPI/KRI definitions for the create-from-dropdown picker."""
    from ....services import metric_templates
    return {"templates": metric_templates.templates()}


@router.get("/due")
def kri_due(db: Session = Depends(get_db), current_user: GRCUser = Depends(require_auth)):
    """Manual KRIs/KPIs whose next measurement is due or overdue."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"due": []}
    kri_feeds.ensure_kri_columns(db)
    now = datetime.utcnow()
    kris = (db.query(RiskKRI)
            .filter(RiskKRI.tenant_id.in_(user_tenants), RiskKRI.is_active == True,
                    RiskKRI.metric_key.is_(None), RiskKRI.next_due_date.isnot(None),
                    RiskKRI.next_due_date <= now)
            .order_by(RiskKRI.next_due_date.asc()).all())
    return {"due": [
        {"id": k.id, "name": k.name, "kind": k.kind, "frequency": k.frequency,
         "next_due_date": k.next_due_date.isoformat() if k.next_due_date else None,
         "data_provider_id": k.data_provider_id,
         "days_overdue": (now - k.next_due_date).days if k.next_due_date else None}
        for k in kris]}


@router.get("/report")
def kri_report(days: int = 90, kind: Optional[str] = None, db: Session = Depends(get_db),
               current_user: GRCUser = Depends(require_auth)):
    """Governance KRI rollup: every active KRI with its live (or manual) value, RAG
    status, source module, and a trend series — plus a summary for the board view."""
    user_tenants = get_user_tenants(current_user, db)
    empty = {"kris": [], "days": days,
             "summary": {"total": 0, "green": 0, "amber": 0, "red": 0, "unknown": 0, "live": 0, "breached": 0}}
    if not user_tenants:
        return empty
    kri_feeds.ensure_kri_columns(db)
    q = db.query(RiskKRI).filter(RiskKRI.tenant_id.in_(user_tenants), RiskKRI.is_active == True)
    if kind in ("kpi", "kri"):
        q = q.filter(RiskKRI.kind == kind)
    kris = q.all()

    items = []
    counts = {"green": 0, "amber": 0, "red": 0, "unknown": 0}
    by_category: dict = {}
    for kri in kris:
        data = _hydrate(db, kri, user_tenants)
        st = data.current_status or "unknown"
        counts[st] = counts.get(st, 0) + 1
        cat = data.category or "Uncategorized"
        bc = by_category.setdefault(cat, {"green": 0, "amber": 0, "red": 0, "unknown": 0, "total": 0})
        bc[st] = bc.get(st, 0) + 1
        bc["total"] += 1
        d = data.model_dump()
        if getattr(kri, "metric_key", None):
            d["history"] = kri_feeds.history(db, user_tenants, kri.metric_key, days=days)
        else:
            d["history"] = [
                {"date": m.measured_at.date().isoformat() if m.measured_at else None, "value": m.value}
                for m in sorted(kri.measurements, key=lambda x: x.measured_at or datetime.min)
                if m.value is not None
            ]
        items.append(d)

    return {
        "kris": items,
        "days": days,
        "summary": {
            "total": len(kris), **counts,
            "kpi": sum(1 for k in kris if (getattr(k, "kind", None) or "kri") == "kpi"),
            "kri": sum(1 for k in kris if (getattr(k, "kind", None) or "kri") != "kpi"),
            "live": sum(1 for k in kris if getattr(k, "metric_key", None)),
            "manual": sum(1 for k in kris if not getattr(k, "metric_key", None)),
            "breached": counts["red"] + counts["amber"],
        },
        "by_category": [{"category": c, **v} for c, v in sorted(by_category.items())],
    }


@router.get("/{kri_id}", response_model=RiskKRIResponse)
def get_kri(
    kri_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri_feeds.ensure_kri_columns(db)
    kri = db.query(RiskKRI).options(
        joinedload(RiskKRI.measurements)
    ).filter(
        RiskKRI.id == kri_id,
        RiskKRI.tenant_id.in_(user_tenants)
    ).first()

    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    return _hydrate(db, kri, user_tenants)


@router.put("/{kri_id}", response_model=RiskKRIResponse)
def update_kri(
    kri_id: int,
    kri_update: RiskKRIUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri_feeds.ensure_kri_columns(db)
    kri = db.query(RiskKRI).filter(
        RiskKRI.id == kri_id,
        RiskKRI.tenant_id.in_(user_tenants)
    ).first()

    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )

    update_data = kri_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(kri, key, value)

    db.commit()
    db.refresh(kri)
    return _hydrate(db, kri, user_tenants)


@router.delete("/{kri_id}", response_model=MessageResponse)
def delete_kri(
    kri_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri_feeds.ensure_kri_columns(db)
    kri = db.query(RiskKRI).filter(
        RiskKRI.id == kri_id,
        RiskKRI.tenant_id.in_(user_tenants)
    ).first()

    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )

    db.delete(kri)
    db.commit()
    return {"message": "KRI deleted successfully", "id": kri_id}


@router.post("/{kri_id}/measure", response_model=RiskKRIMeasurementResponse)
def record_kri_measurement(
    kri_id: int,
    measurement: RiskKRIMeasurementCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    kri_feeds.ensure_kri_columns(db)

    kri = db.query(RiskKRI).filter(
        RiskKRI.id == kri_id,
        RiskKRI.tenant_id.in_(user_tenants)
    ).first()

    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )

    status_val = calculate_kri_status(measurement.value, kri)
    review_status = measurement.review_status or "approved"

    db_measurement = RiskKRIMeasurement(
        kri_id=kri_id,
        value=measurement.value,
        status=status_val,
        measured_by=current_user.id,
        notes=measurement.notes,
        period_label=measurement.period_label,
        target=measurement.target if measurement.target is not None else kri.target,
        review_status=review_status,
        reviewed_by=current_user.id if review_status == "approved" else None,
        reviewed_at=datetime.utcnow() if review_status == "approved" else None,
    )
    db.add(db_measurement)

    kri.current_value = measurement.value
    kri.last_measured_at = datetime.utcnow()

    db.commit()
    # ── v2 Issue Management hook: when the KRI flips to red AND the tenant
    # has opted in (kri_red_breach flag), auto-spawn an Issue. Gated +
    # de-duplicated inside the helper so it can never double-fire.
    # Wrapped in try/except so issue-management drift can never break the
    # core KRI measurement write.
    if status_val == "red":
        try:
            from ....modules.issue_management.services.auto_create import from_event
            risk_for_tenant = db.query(Risk).filter(Risk.id == kri.risk_id).first() if kri.risk_id else None
            tenant_id = kri.tenant_id or (risk_for_tenant.tenant_id if risk_for_tenant else None)
            if tenant_id:
                from_event(
                    db=db,
                    tenant_id=tenant_id,
                    source_type="kri_breach",
                    source_id=kri.id,
                    title=f"KRI breach (RED): {kri.name}",
                    description=(
                        f"KRI '{kri.name}' measured at {measurement.value} "
                        f"{kri.unit or ''} — red-threshold breach.\n"
                        f"Notes: {measurement.notes or '—'}"
                    ),
                    impact="high",
                    urgency="high",
                    issue_type="incident",
                    category="operations",
                    reporter_id=current_user.id,
                    feature_flag="kri_red_breach",
                )
                db.commit()
        except Exception:
            db.rollback()

    db.refresh(db_measurement)
    return db_measurement


@router.get("/{kri_id}/trend", response_model=List[RiskKRIMeasurementResponse])
def get_kri_trend(
    kri_id: int,
    days: int = 365,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    kri = db.query(RiskKRI).join(Risk).filter(
        RiskKRI.id == kri_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not kri:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="KRI not found"
        )
    
    start_date = datetime.utcnow() - timedelta(days=days)
    measurements = db.query(RiskKRIMeasurement).filter(
        RiskKRIMeasurement.kri_id == kri_id,
        RiskKRIMeasurement.measured_at >= start_date
    ).order_by(RiskKRIMeasurement.measured_at.asc()).all()
    
    return measurements


HEADER_KEYWORDS = {
    'name': ['kri name', 'kri_name', 'indicator name', 'indicator', 'name', 'key risk indicator'],
    'kri_id': ['kri id', 'kri_id', 'id', 'indicator id', 'ref', 'reference'],
    'category': ['risk category', 'category', 'risk_category', 'risk area', 'domain'],
    'description': ['description', 'desc', 'details', 'definition'],
    'measurement': ['measurement', 'metric', 'measure', 'kpi', 'metric_type', 'measurement type'],
    'unit': ['unit', 'unit of measure', 'uom', 'units'],
    'frequency': ['frequency', 'monitoring frequency', 'reporting frequency', 'review frequency', 'cadence'],
    'owner': ['owner', 'kri owner', 'responsible', 'assigned to', 'assignee'],
    'current_value': ['current value', 'current_value', 'value', 'current', 'latest value', 'actual'],
    'previous_value': ['previous value', 'previous_value', 'prior value', 'last value'],
    'trend': ['trend', 'direction', 'movement'],
    'green_threshold': ['green threshold', 'green_threshold', 'green', 'target', 'green limit'],
    'amber_threshold': ['amber threshold', 'amber_threshold', 'amber', 'warning', 'amber limit', 'yellow threshold', 'yellow'],
    'red_threshold': ['red threshold', 'red_threshold', 'red', 'critical', 'red limit', 'breach'],
    'status': ['status', 'current status', 'rag status', 'rag', 'indicator status'],
    'data_source': ['data source', 'data_source', 'source', 'system'],
    'threshold_direction': ['threshold direction', 'direction', 'threshold_direction', 'polarity'],
}


def _match_header(header_text: str) -> Optional[str]:
    if not header_text:
        return None
    text = str(header_text).strip().lower()
    for field, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            if text == kw or text.replace('_', ' ') == kw:
                return field
    for field, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            if kw in text or text in kw:
                return field
    return None


def _parse_threshold(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    nums = re.findall(r'[-+]?\d*\.?\d+', text)
    if nums:
        return float(nums[0])
    return None


def _parse_frequency(value) -> str:
    if not value:
        return "monthly"
    text = str(value).strip().lower()
    freq_map = {
        'daily': 'daily', 'day': 'daily',
        'weekly': 'weekly', 'week': 'weekly',
        'monthly': 'monthly', 'month': 'monthly',
        'quarterly': 'quarterly', 'quarter': 'quarterly',
        'annually': 'annually', 'annual': 'annually', 'yearly': 'annually', 'year': 'annually',
    }
    for key, val in freq_map.items():
        if key in text:
            return val
    return "monthly"


def _detect_threshold_direction(green_val, red_val) -> str:
    if green_val is not None and red_val is not None:
        if green_val < red_val:
            return "lower_is_better"
        else:
            return "higher_is_better"
    return "lower_is_better"


@router.post("/upload", response_model=None)
def upload_kris(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx)")

    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="No tenant found for user")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    ws = None
    preferred_names = ['kris', 'kri', 'key risk indicators', 'indicators', 'sheet1']
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in preferred_names:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header_row_idx = None
    column_map = {}
    for idx, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c else '' for c in row]
        matched = {}
        for col_idx, cell in enumerate(cells):
            field = _match_header(cell)
            if field and field not in matched:
                matched[field] = col_idx
        if 'name' in matched and len(matched) >= 3:
            header_row_idx = idx
            column_map = matched
            break

    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not detect header row. Ensure columns include at least 'KRI Name' and 2 other recognized columns.")

    data_rows = rows[header_row_idx + 1:]
    created = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(data_rows, start=header_row_idx + 2):
        try:
            def get_val(field):
                col = column_map.get(field)
                if col is not None and col < len(row):
                    v = row[col]
                    return v if v is not None else None
                return None

            name = get_val('name')
            if not name or str(name).strip() == '':
                skipped += 1
                continue
            name = str(name).strip()

            description_parts = []
            if get_val('description'):
                description_parts.append(str(get_val('description')).strip())
            if get_val('category'):
                description_parts.append(f"Category: {str(get_val('category')).strip()}")
            if get_val('kri_id'):
                description_parts.append(f"KRI ID: {str(get_val('kri_id')).strip()}")

            description = ' | '.join(description_parts) if description_parts else None

            measurement = get_val('measurement')
            unit = get_val('unit')
            metric_type = "numeric"
            if measurement:
                m_text = str(measurement).strip().lower()
                if any(k in m_text for k in ['percent', '%', 'ratio', 'rate']):
                    metric_type = "percentage"
                    if not unit:
                        unit = "%"
                elif any(k in m_text for k in ['count', 'number', 'total', '#']):
                    metric_type = "count"
                elif any(k in m_text for k in ['bool', 'yes/no', 'true/false']):
                    metric_type = "boolean"
                if not unit:
                    unit = str(measurement).strip()

            green = _parse_threshold(get_val('green_threshold'))
            amber = _parse_threshold(get_val('amber_threshold'))
            red = _parse_threshold(get_val('red_threshold'))

            if green is None and red is not None and amber is not None:
                green = amber
                amber = red
            elif green is not None and amber is None and red is not None:
                amber = (green + red) / 2

            threshold_dir = get_val('threshold_direction')
            if threshold_dir:
                td_text = str(threshold_dir).strip().lower()
                if any(k in td_text for k in ['higher', 'more', 'above', 'increase']):
                    threshold_direction = "higher_is_better"
                else:
                    threshold_direction = "lower_is_better"
            else:
                threshold_direction = _detect_threshold_direction(green, red)

            current_value = _parse_threshold(get_val('current_value'))
            frequency = _parse_frequency(get_val('frequency'))

            data_source_parts = []
            if get_val('data_source'):
                data_source_parts.append(str(get_val('data_source')).strip())
            if get_val('kri_id'):
                data_source_parts.append(f"Ref: {str(get_val('kri_id')).strip()}")
            data_source = ' | '.join(data_source_parts) if data_source_parts else None

            db_kri = RiskKRI(
                risk_id=None,
                name=name,
                description=description,
                metric_type=metric_type,
                unit=unit,
                current_value=current_value,
                green_threshold=green,
                amber_threshold=amber,
                threshold_direction=threshold_direction,
                frequency=frequency,
                data_source=data_source,
                is_active=True,
            )
            db.add(db_kri)
            created += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
            skipped += 1

    if created > 0:
        db.commit()

    return {
        "message": f"Successfully imported {created} KRIs",
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
