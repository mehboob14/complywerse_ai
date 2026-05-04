from typing import List, Optional
from datetime import datetime
from io import BytesIO
import os
import json

from pydantic import BaseModel
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, subqueryload
from sqlalchemy import func, and_, or_

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    openpyxl = None

try:
    from openai import OpenAI
    client = OpenAI()
except Exception:
    client = None

EVIDENCE_KEYWORD_MAP = {
    "mfa_authentication": {
        "keywords": ["mfa", "multi-factor", "multifactor", "authentication", "two-factor", "2fa", "login", "privileged account", "remote access", "access control", "identity"],
        "suggestion": "Demonstrate that strong authentication controls are enforced by providing configuration evidence, access reviews, and policy documentation.",
        "evidence": [
            {"evidence_type": "MFA Configuration Evidence", "description": "Screenshots or exports showing MFA is enabled and enforced across systems", "example_files": ["mfa-config-screenshot.png", "azure-ad-mfa-policy.pdf"]},
            {"evidence_type": "Access Control Policy", "description": "Formal policy defining authentication requirements and access control standards", "example_files": ["access-control-policy.pdf", "authentication-standard.docx"]},
            {"evidence_type": "Privileged Account Inventory", "description": "List of privileged accounts with MFA status and last review date", "example_files": ["privileged-accounts.xlsx", "admin-account-review.pdf"]},
            {"evidence_type": "Access Review Records", "description": "Periodic user access review results and recertification evidence", "example_files": ["access-review-q4.xlsx", "user-recertification.pdf"]},
        ],
    },
    "patching_vulnerability": {
        "keywords": ["patch", "patching", "vulnerability", "vulnerabilities", "update", "updates", "cve", "remediation", "security update", "critical vulnerability"],
        "suggestion": "Provide evidence of timely patch deployment, vulnerability scanning, and remediation tracking for critical systems.",
        "evidence": [
            {"evidence_type": "Patch Management Report", "description": "Report showing patch deployment status, timelines, and compliance rates", "example_files": ["patch-status-report.xlsx", "wsus-deployment-summary.pdf"]},
            {"evidence_type": "Vulnerability Scan Results", "description": "Latest vulnerability scan output showing identified and remediated vulnerabilities", "example_files": ["nessus-scan-results.pdf", "qualys-report.xlsx"]},
            {"evidence_type": "Remediation Tracker", "description": "Tracking document for open vulnerabilities with target remediation dates", "example_files": ["remediation-tracker.xlsx", "vuln-remediation-plan.pdf"]},
            {"evidence_type": "Patch Policy Document", "description": "Formal patch management policy with SLAs for critical/high/medium patches", "example_files": ["patch-management-policy.pdf", "vulnerability-management-sop.docx"]},
        ],
    },
    "training_awareness": {
        "keywords": ["training", "awareness", "security awareness", "phishing", "education", "e-learning", "employee training", "staff training", "cyber awareness"],
        "suggestion": "Demonstrate that security awareness training is delivered regularly with completion tracking and effectiveness measurement.",
        "evidence": [
            {"evidence_type": "Training Completion Records", "description": "Report showing employee training completion rates and dates", "example_files": ["training-completion-report.xlsx", "lms-completion-extract.pdf"]},
            {"evidence_type": "Training Material / Curriculum", "description": "Course content, slides, or syllabus for the security awareness program", "example_files": ["security-awareness-slides.pdf", "training-curriculum.docx"]},
            {"evidence_type": "Phishing Simulation Results", "description": "Results from phishing simulation campaigns showing click rates and improvements", "example_files": ["phishing-sim-results.pdf", "knowbe4-report.xlsx"]},
            {"evidence_type": "Training Attendance Records", "description": "Sign-in sheets or attendance logs from training sessions", "example_files": ["training-attendance.xlsx", "session-sign-in-sheet.pdf"]},
        ],
    },
    "backup_dr": {
        "keywords": ["backup", "disaster recovery", "bcp", "business continuity", "recovery", "restore", "rto", "rpo", "failover", "data recovery"],
        "suggestion": "Provide evidence of regular backup verification, disaster recovery testing, and documented recovery objectives.",
        "evidence": [
            {"evidence_type": "Backup Verification Report", "description": "Reports confirming successful backup completion and restore testing", "example_files": ["backup-verification-report.pdf", "restore-test-results.xlsx"]},
            {"evidence_type": "DR Test Results", "description": "Documentation from disaster recovery exercises including RTO/RPO metrics", "example_files": ["dr-test-report.pdf", "failover-test-results.docx"]},
            {"evidence_type": "BCP / DR Plan", "description": "Current business continuity or disaster recovery plan document", "example_files": ["bcp-plan-v3.pdf", "disaster-recovery-plan.docx"]},
            {"evidence_type": "Backup Configuration Evidence", "description": "Screenshots or exports showing backup schedules, retention, and scope", "example_files": ["backup-schedule-config.png", "backup-policy-settings.pdf"]},
        ],
    },
    "encryption_data_protection": {
        "keywords": ["encryption", "encrypt", "data protection", "dlp", "data loss", "data classification", "sensitive data", "pii", "cryptograph", "key management", "data at rest", "data in transit"],
        "suggestion": "Demonstrate that encryption and data protection controls are in place with proper key management and classification.",
        "evidence": [
            {"evidence_type": "Encryption Configuration Evidence", "description": "Screenshots or settings showing encryption for data at rest and in transit", "example_files": ["tls-config-screenshot.png", "disk-encryption-status.pdf"]},
            {"evidence_type": "Data Classification Inventory", "description": "Inventory of data assets with classification levels and protection requirements", "example_files": ["data-classification-inventory.xlsx", "data-asset-register.pdf"]},
            {"evidence_type": "Key Management Records", "description": "Key rotation schedules, access controls, and lifecycle management evidence", "example_files": ["key-rotation-log.xlsx", "kms-access-policy.pdf"]},
            {"evidence_type": "DLP Policy & Reports", "description": "Data loss prevention policy and recent DLP incident/alert reports", "example_files": ["dlp-policy.pdf", "dlp-incident-report.xlsx"]},
        ],
    },
    "logging_monitoring": {
        "keywords": ["logging", "monitoring", "siem", "log", "logs", "alert", "alerting", "detection", "security monitoring", "audit trail", "audit log", "event management"],
        "suggestion": "Provide evidence that security events are logged, monitored, and that alerts are reviewed and actioned in a timely manner.",
        "evidence": [
            {"evidence_type": "SIEM Dashboard / Configuration", "description": "Screenshots of SIEM dashboards showing log sources, alert rules, and coverage", "example_files": ["siem-dashboard.png", "splunk-config-summary.pdf"]},
            {"evidence_type": "Log Retention Configuration", "description": "Evidence of log retention policies and storage configuration", "example_files": ["log-retention-policy.pdf", "log-storage-config.png"]},
            {"evidence_type": "Alert Review Records", "description": "Records showing security alerts are reviewed and investigated", "example_files": ["alert-review-log.xlsx", "soc-triage-report.pdf"]},
            {"evidence_type": "Monitoring Coverage Report", "description": "Report showing which systems and events are covered by monitoring", "example_files": ["monitoring-coverage-matrix.xlsx", "log-source-inventory.pdf"]},
        ],
    },
    "firewall_network": {
        "keywords": ["firewall", "network security", "network segmentation", "ids", "ips", "intrusion", "perimeter", "dmz", "network access", "port", "traffic"],
        "suggestion": "Demonstrate that network security controls are properly configured, reviewed, and tested through rule reviews and penetration testing.",
        "evidence": [
            {"evidence_type": "Firewall Rule Review", "description": "Recent firewall rule review showing approved rules and cleanup actions", "example_files": ["firewall-rule-review.xlsx", "fw-rule-audit-report.pdf"]},
            {"evidence_type": "Network Architecture Diagram", "description": "Current network diagrams showing segmentation, DMZ, and security zones", "example_files": ["network-diagram.pdf", "security-zone-architecture.vsdx"]},
            {"evidence_type": "Penetration Test Report", "description": "Recent penetration test findings and remediation status", "example_files": ["pentest-report.pdf", "external-pentest-findings.docx"]},
            {"evidence_type": "IDS/IPS Configuration", "description": "Intrusion detection/prevention system configuration and alert tuning evidence", "example_files": ["ids-config-export.pdf", "ips-rule-policy.png"]},
        ],
    },
    "change_management": {
        "keywords": ["change management", "change control", "change request", "cab", "change advisory", "release management", "deployment", "rollback"],
        "suggestion": "Provide evidence that changes follow a formal approval process with proper documentation, testing, and rollback procedures.",
        "evidence": [
            {"evidence_type": "Change Request Records", "description": "Sample change requests showing approval workflow and implementation details", "example_files": ["change-requests-log.xlsx", "sample-cr-approval.pdf"]},
            {"evidence_type": "CAB Meeting Minutes", "description": "Change Advisory Board meeting minutes documenting change review decisions", "example_files": ["cab-meeting-minutes.pdf", "cab-decisions-log.xlsx"]},
            {"evidence_type": "Change Management Policy", "description": "Formal change management policy and standard operating procedure", "example_files": ["change-management-policy.pdf", "change-control-sop.docx"]},
        ],
    },
    "vendor_third_party": {
        "keywords": ["vendor", "third party", "third-party", "outsourcing", "supplier", "contractor", "service provider", "sla", "soc2", "soc 2", "supply chain"],
        "suggestion": "Demonstrate that third-party risks are assessed, monitored, and that vendor compliance is verified through audits and SLA tracking.",
        "evidence": [
            {"evidence_type": "Vendor Risk Assessment", "description": "Completed risk assessments for critical third-party vendors", "example_files": ["vendor-risk-assessment.xlsx", "critical-vendor-review.pdf"]},
            {"evidence_type": "Third-Party Audit Reports", "description": "SOC 2, ISO 27001, or other audit reports from key vendors", "example_files": ["vendor-soc2-report.pdf", "supplier-iso27001-cert.pdf"]},
            {"evidence_type": "SLA Performance Reports", "description": "Service level agreement performance tracking and compliance reports", "example_files": ["sla-performance-report.xlsx", "vendor-scorecard.pdf"]},
            {"evidence_type": "Vendor Contracts / Agreements", "description": "Contracts with security clauses, data processing agreements, or NDAs", "example_files": ["vendor-dpa.pdf", "service-agreement.docx"]},
        ],
    },
    "incident_response": {
        "keywords": ["incident", "breach", "incident response", "forensic", "tabletop", "escalation", "security incident", "data breach", "cyber incident"],
        "suggestion": "Provide evidence of incident response preparedness including plans, testing, and post-incident review processes.",
        "evidence": [
            {"evidence_type": "Incident Response Plan", "description": "Current incident response plan with roles, escalation paths, and procedures", "example_files": ["incident-response-plan.pdf", "ir-playbook.docx"]},
            {"evidence_type": "Incident Post-Mortem Reports", "description": "Post-incident review reports with root cause analysis and lessons learned", "example_files": ["incident-postmortem.pdf", "root-cause-analysis.docx"]},
            {"evidence_type": "Tabletop Exercise Records", "description": "Documentation from incident response tabletop exercises and drills", "example_files": ["tabletop-exercise-report.pdf", "ir-drill-results.docx"]},
            {"evidence_type": "Incident Log", "description": "Log of security incidents with classification, response times, and resolution", "example_files": ["security-incident-log.xlsx", "incident-tracker.pdf"]},
        ],
    },
    "policy_governance": {
        "keywords": ["policy", "policies", "governance", "compliance framework", "regulatory", "regulation", "standard", "iso", "nist", "governance framework"],
        "suggestion": "Provide evidence that governance policies are documented, approved, communicated, and regularly reviewed.",
        "evidence": [
            {"evidence_type": "Policy Documents", "description": "Approved policy documents with version control and review dates", "example_files": ["information-security-policy.pdf", "acceptable-use-policy.docx"]},
            {"evidence_type": "Governance Committee Minutes", "description": "Minutes from governance/risk committee meetings showing oversight activities", "example_files": ["governance-committee-minutes.pdf", "risk-committee-report.docx"]},
            {"evidence_type": "Policy Acknowledgment Records", "description": "Evidence that staff have read and acknowledged relevant policies", "example_files": ["policy-acknowledgment-log.xlsx", "staff-sign-off-records.pdf"]},
        ],
    },
    "physical_security": {
        "keywords": ["physical security", "physical access", "data center", "cctv", "badge", "visitor", "facility", "building security", "server room", "environmental"],
        "suggestion": "Demonstrate that physical access controls are in place with monitoring, access logs, and environmental safeguards.",
        "evidence": [
            {"evidence_type": "Physical Access Logs", "description": "Badge/swipe access logs for secure areas and data centers", "example_files": ["access-log-export.xlsx", "datacenter-access-log.pdf"]},
            {"evidence_type": "CCTV / Surveillance Records", "description": "Evidence that surveillance systems are operational and footage is retained", "example_files": ["cctv-coverage-map.pdf", "surveillance-retention-policy.docx"]},
            {"evidence_type": "Visitor Management Records", "description": "Visitor sign-in logs and escort procedures documentation", "example_files": ["visitor-log.xlsx", "visitor-management-procedure.pdf"]},
            {"evidence_type": "Environmental Monitoring Reports", "description": "Temperature, humidity, and fire suppression monitoring evidence", "example_files": ["environmental-monitoring-report.pdf", "ups-maintenance-log.xlsx"]},
        ],
    },
    "asset_inventory": {
        "keywords": ["asset", "inventory", "cmdb", "configuration management", "hardware", "software inventory", "it asset", "asset management", "end of life", "eol"],
        "suggestion": "Provide evidence of a maintained IT asset inventory with classification, ownership, and lifecycle tracking.",
        "evidence": [
            {"evidence_type": "IT Asset Inventory", "description": "Current inventory of hardware and software assets with classification and ownership", "example_files": ["it-asset-inventory.xlsx", "cmdb-export.csv"]},
            {"evidence_type": "Asset Lifecycle Records", "description": "Records showing asset procurement, deployment, and decommissioning", "example_files": ["asset-lifecycle-tracker.xlsx", "eol-hardware-list.pdf"]},
            {"evidence_type": "Software License Records", "description": "Software license inventory with compliance status and renewal dates", "example_files": ["software-license-inventory.xlsx", "license-compliance-report.pdf"]},
        ],
    },
}


def get_question_specific_evidence(question_text: str, question_type: str = "", risk_category: str = "", control_objective: str = ""):
    text_lower = (question_text or "").lower()
    context_lower = f"{text_lower} {(risk_category or '').lower()} {(control_objective or '').lower()}"

    best_match = None
    best_score = 0

    for category_key, category_data in EVIDENCE_KEYWORD_MAP.items():
        score = 0
        for keyword in category_data["keywords"]:
            if keyword in context_lower:
                score += len(keyword)
        if score > best_score:
            best_score = score
            best_match = category_key

    if best_match:
        matched = EVIDENCE_KEYWORD_MAP[best_match]
        return matched["suggestion"], matched["evidence"][:4]

    if question_type == "risk_rating":
        return (
            f"Assess the likelihood and impact based on historical data, current controls, and industry benchmarks.",
            [
                {"evidence_type": "Risk Register Extract", "description": "Current risk register showing identified risks and ratings", "example_files": ["risk-register.xlsx", "risk-register-extract.pdf"]},
                {"evidence_type": "Risk Assessment Report", "description": "Formal risk assessment documenting methodology and findings", "example_files": ["risk-assessment-report.pdf", "annual-risk-review.docx"]},
                {"evidence_type": "Historical Incident Records", "description": "Records of past incidents and losses related to this risk area", "example_files": ["incident-log.xlsx", "loss-event-database.csv"]},
            ],
        )
    elif question_type == "control_rating":
        return (
            f"Evaluate the design and operating effectiveness of controls through testing evidence and audit results.",
            [
                {"evidence_type": "Control Testing Results", "description": "Results from control testing and walkthroughs", "example_files": ["control-test-results.xlsx", "walkthrough-evidence.pdf"]},
                {"evidence_type": "Internal Audit Report", "description": "Internal audit findings related to this control area", "example_files": ["internal-audit-report.pdf", "audit-findings-summary.docx"]},
                {"evidence_type": "Process Documentation", "description": "Documented procedures and process flows", "example_files": ["process-flowchart.pdf", "standard-operating-procedure.docx"]},
            ],
        )
    else:
        return (
            f"Provide relevant documentation demonstrating compliance with this requirement.",
            [
                {"evidence_type": "Supporting Documentation", "description": "Relevant documents supporting your response", "example_files": ["supporting-doc.pdf", "evidence-package.zip"]},
                {"evidence_type": "Process Evidence", "description": "Screenshots or exports demonstrating process execution", "example_files": ["process-screenshot.png", "system-export.xlsx"]},
                {"evidence_type": "Compliance Records", "description": "Records demonstrating compliance with requirements", "example_files": ["compliance-checklist.xlsx", "attestation-record.pdf"]},
            ],
        )

from ....models import (
    RCSATemplate, RCSAQuestion, RCSACampaign, RCSAAssessment,
    RCSAResponse, RCSAFinding, RCSAApprovalWorkflow, RCSAApprovalTier,
    RCSAApprovalHistory, Risk, InternalControl, RiskMitigationAction,
    BusinessUnit, GRCUser, Tenant, get_db
)
from ....schemas import (
    RCSATemplateCreate, RCSATemplateUpdate, RCSATemplateResponse, RCSATemplateDetailResponse,
    RCSAQuestionCreate, RCSAQuestionUpdate, RCSAQuestionResponse, RCSAQuestionUpsert,
    RCSACampaignCreate, RCSACampaignUpdate, RCSACampaignResponse,
    RCSAAssessmentResponse, RCSAAssessmentDetailResponse, RCSAQuestionWithResponse, RCSAResponseDetail, RCSAEvidenceFile,
    RCSAResponseCreate, RCSAResponseUpdate, RCSAResponseResponse,
    RCSABulkResponseSave, RCSAFindingCreate, RCSAFindingUpdate, RCSAFindingResponse,
    RCSAApprovalWorkflowCreate, RCSAApprovalWorkflowUpdate, RCSAApprovalWorkflowResponse,
    RCSAApprovalTierCreate, RCSAApprovalTierResponse, RCSAApprovalHistoryResponse,
    RCSAApprovalAction, RCSADelegateAction, RCSABUAssignRequest,
    RCSADashboardSummary, RCSAFindingsBySeverity, RCSABUProgress, RCSAAISuggestionResponse,
    EvidenceRecommendation, MessageResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/rcsa", tags=["RCSA - Risk and Control Self-Assessment"])


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


# =============================================================================
# Template Management Endpoints
# =============================================================================

@router.get("/templates", response_model=List[RCSATemplateResponse])
def list_templates(
    tenant_id: Optional[int] = None,
    category: Optional[str] = None,
    source: Optional[str] = None,
    include_system: bool = True,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSATemplate).filter(
        or_(
            RCSATemplate.tenant_id.in_(user_tenants),
            and_(RCSATemplate.is_system_template == True, include_system)
        ),
        RCSATemplate.is_active == True
    )
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(
            or_(
                RCSATemplate.tenant_id == tenant_id,
                RCSATemplate.is_system_template == True
            )
        )
    if category:
        query = query.filter(RCSATemplate.category == category)
    if source:
        query = query.filter(RCSATemplate.source == source)
    
    templates = query.order_by(RCSATemplate.name).offset(skip).limit(limit).all()
    
    return [
        RCSATemplateResponse(
            id=t.id,
            tenant_id=t.tenant_id,
            name=t.name,
            description=t.description,
            category=t.category,
            source=t.source,
            version=t.version,
            is_system_template=t.is_system_template,
            is_active=t.is_active,
            risk_categories=t.risk_categories or [],
            regulatory_mapping=t.regulatory_mapping or {},
            created_by=t.created_by,
            created_at=t.created_at,
            updated_at=t.updated_at,
            question_count=len(t.questions)
        )
        for t in templates
    ]


@router.get("/templates/{template_id}", response_model=RCSATemplateDetailResponse)
def get_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    template = db.query(RCSATemplate).options(
        joinedload(RCSATemplate.questions)
    ).filter(
        RCSATemplate.id == template_id,
        or_(
            RCSATemplate.tenant_id.in_(user_tenants),
            RCSATemplate.is_system_template == True
        )
    ).first()
    
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    return RCSATemplateDetailResponse(
        id=template.id,
        tenant_id=template.tenant_id,
        name=template.name,
        description=template.description,
        category=template.category,
        source=template.source,
        version=template.version,
        is_system_template=template.is_system_template,
        is_active=template.is_active,
        risk_categories=template.risk_categories or [],
        regulatory_mapping=template.regulatory_mapping or {},
        created_by=template.created_by,
        created_at=template.created_at,
        updated_at=template.updated_at,
        question_count=len(template.questions),
        questions=[
            RCSAQuestionResponse(
                id=q.id,
                template_id=q.template_id,
                section=q.section,
                question_order=q.question_order,
                question_text=q.question_text,
                question_type=q.question_type,
                is_required=q.is_required,
                options=q.options or [],
                risk_category=q.risk_category,
                control_objective=q.control_objective,
                guidance_text=q.guidance_text,
                ai_suggestion_enabled=q.ai_suggestion_enabled,
                created_at=q.created_at
            )
            for q in sorted(template.questions, key=lambda x: (x.section or "", x.question_order))
        ]
    )


@router.post("/templates", response_model=RCSATemplateResponse, status_code=status.HTTP_201_CREATED)
def create_template(
    template: RCSATemplateCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User not assigned to any tenant")
    
    db_template = RCSATemplate(
        tenant_id=tenant_id,
        name=template.name,
        description=template.description,
        category=template.category,
        source=template.source,
        version=template.version,
        is_system_template=False,
        risk_categories=template.risk_categories,
        regulatory_mapping=template.regulatory_mapping,
        created_by=current_user.id
    )
    db.add(db_template)
    db.flush()
    
    for q in template.questions:
        db_question = RCSAQuestion(
            template_id=db_template.id,
            section=q.section,
            question_order=q.question_order,
            question_text=q.question_text,
            question_type=q.question_type,
            is_required=q.is_required,
            options=q.options,
            risk_category=q.risk_category,
            control_objective=q.control_objective,
            guidance_text=q.guidance_text,
            ai_suggestion_enabled=q.ai_suggestion_enabled
        )
        db.add(db_question)
    
    db.commit()
    db.refresh(db_template)
    
    return RCSATemplateResponse(
        id=db_template.id,
        tenant_id=db_template.tenant_id,
        name=db_template.name,
        description=db_template.description,
        category=db_template.category,
        source=db_template.source,
        version=db_template.version,
        is_system_template=db_template.is_system_template,
        is_active=db_template.is_active,
        risk_categories=db_template.risk_categories or [],
        regulatory_mapping=db_template.regulatory_mapping or {},
        created_by=db_template.created_by,
        created_at=db_template.created_at,
        updated_at=db_template.updated_at,
        question_count=len(template.questions)
    )


@router.put("/templates/{template_id}", response_model=RCSATemplateResponse)
def update_template(
    template_id: int,
    template: RCSATemplateUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)

    db_template = db.query(RCSATemplate).filter(
        RCSATemplate.id == template_id,
        RCSATemplate.tenant_id.in_(user_tenants),
        RCSATemplate.is_system_template == False
    ).first()

    if not db_template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found or not editable")

    update_data = template.model_dump(exclude_unset=True)
    questions_data = update_data.pop('questions', None)

    # Update scalar template fields
    for key, value in update_data.items():
        setattr(db_template, key, value)

    # Upsert / delete questions when provided
    if questions_data is not None:
        # Direct query — avoids any refresh/lazy-load ordering issues
        existing_questions = db.query(RCSAQuestion).filter(
            RCSAQuestion.template_id == template_id
        ).all()
        existing_by_id = {q.id: q for q in existing_questions}

        payload_existing_ids = {
            q['id'] for q in questions_data
            if q.get('id') and q['id'] in existing_by_id
        }

        # Delete questions that were removed from the payload
        for q in existing_questions:
            if q.id not in payload_existing_ids:
                db.delete(q)
        db.flush()

        # Update existing / insert new
        for idx, q_data in enumerate(questions_data):
            q_id = q_data.get('id')
            order = q_data.get('sequence') or q_data.get('question_order') or (idx + 1)
            section = q_data.get('section') or q_data.get('category') or None
            guidance = q_data.get('guidance_text') or q_data.get('guidance') or None

            if q_id and q_id in existing_by_id:
                db_q = existing_by_id[q_id]
                db_q.question_text = q_data.get('question_text', db_q.question_text)
                db_q.question_type = q_data.get('question_type', db_q.question_type)
                db_q.question_order = order
                db_q.section = section
                db_q.is_required = q_data.get('is_required', db_q.is_required)
                db_q.options = q_data.get('options') or []
                db_q.risk_category = q_data.get('risk_category') or None
                db_q.control_objective = q_data.get('control_objective') or None
                db_q.guidance_text = guidance
                db_q.ai_suggestion_enabled = q_data.get('ai_suggestion_enabled', True)
            else:
                db.add(RCSAQuestion(
                    template_id=template_id,
                    question_text=q_data.get('question_text') or '',
                    question_type=q_data.get('question_type') or 'text',
                    question_order=order,
                    section=section,
                    is_required=q_data.get('is_required', True),
                    options=q_data.get('options') or [],
                    risk_category=q_data.get('risk_category') or None,
                    control_objective=q_data.get('control_objective') or None,
                    guidance_text=guidance,
                    ai_suggestion_enabled=q_data.get('ai_suggestion_enabled', True),
                ))

    db.commit()
    db.refresh(db_template)

    return RCSATemplateResponse(
        id=db_template.id,
        tenant_id=db_template.tenant_id,
        name=db_template.name,
        description=db_template.description,
        category=db_template.category,
        source=db_template.source,
        version=db_template.version,
        is_system_template=db_template.is_system_template,
        is_active=db_template.is_active,
        risk_categories=db_template.risk_categories or [],
        regulatory_mapping=db_template.regulatory_mapping or {},
        created_by=db_template.created_by,
        created_at=db_template.created_at,
        updated_at=db_template.updated_at,
        question_count=len(db_template.questions)
    )


@router.delete("/templates/{template_id}", response_model=MessageResponse)
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_template = db.query(RCSATemplate).filter(
        RCSATemplate.id == template_id,
        RCSATemplate.tenant_id.in_(user_tenants),
        RCSATemplate.is_system_template == False
    ).first()
    
    if not db_template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found or cannot be deleted")
    
    campaign_count = db.query(RCSACampaign).filter(RCSACampaign.template_id == template_id).count()
    if campaign_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete template. It is used by {campaign_count} campaign(s)"
        )
    
    db.delete(db_template)
    db.commit()
    
    return MessageResponse(message="Template deleted successfully", id=template_id)


class CloneTemplateRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None

@router.post("/templates/{template_id}/clone", response_model=RCSATemplateResponse)
def clone_template(
    template_id: int,
    clone_data: Optional[CloneTemplateRequest] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    tenant_id = get_user_primary_tenant(current_user, db)
    
    source_template = db.query(RCSATemplate).options(
        joinedload(RCSATemplate.questions)
    ).filter(
        RCSATemplate.id == template_id,
        or_(
            RCSATemplate.tenant_id.in_(user_tenants),
            RCSATemplate.is_system_template == True
        )
    ).first()
    
    if not source_template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    new_name = clone_data.name if clone_data and clone_data.name else None
    new_description = clone_data.description if clone_data and clone_data.description else None
    
    cloned_template = RCSATemplate(
        tenant_id=tenant_id,
        name=new_name or f"{source_template.name} (Copy)",
        description=new_description or source_template.description,
        category=source_template.category,
        source="custom",
        version="1.0",
        is_system_template=False,
        risk_categories=source_template.risk_categories,
        regulatory_mapping=source_template.regulatory_mapping,
        created_by=current_user.id
    )
    db.add(cloned_template)
    db.flush()
    
    for q in source_template.questions:
        db_question = RCSAQuestion(
            template_id=cloned_template.id,
            section=q.section,
            question_order=q.question_order,
            question_text=q.question_text,
            question_type=q.question_type,
            is_required=q.is_required,
            options=q.options,
            risk_category=q.risk_category,
            control_objective=q.control_objective,
            guidance_text=q.guidance_text,
            ai_suggestion_enabled=q.ai_suggestion_enabled
        )
        db.add(db_question)
    
    db.commit()
    db.refresh(cloned_template)
    
    return RCSATemplateResponse(
        id=cloned_template.id,
        tenant_id=cloned_template.tenant_id,
        name=cloned_template.name,
        description=cloned_template.description,
        category=cloned_template.category,
        source=cloned_template.source,
        version=cloned_template.version,
        is_system_template=cloned_template.is_system_template,
        is_active=cloned_template.is_active,
        risk_categories=cloned_template.risk_categories or [],
        regulatory_mapping=cloned_template.regulatory_mapping or {},
        created_by=cloned_template.created_by,
        created_at=cloned_template.created_at,
        updated_at=cloned_template.updated_at,
        question_count=len(source_template.questions)
    )


@router.post("/templates/upload", response_model=RCSATemplateResponse)
async def upload_template(
    file: UploadFile = File(...),
    name: str = Query(...),
    category: str = Query(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if openpyxl is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel support not available")
    
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User not assigned to any tenant")
    
    content = await file.read()
    
    ALL_HEADER_KEYWORDS = [
        'question_text', 'question text', 'question', 'assessment question',
        'control question', 'checklist item', 'requirement', 'statement',
        'assessment item', 'risk question', 'control description',
        'assessment criteria', 'criteria', 'assessment',
        'section', 'domain', 'control area', 'risk area',
        'control domain', 'pillar', 'control category', 'risk domain', 'process area',
        'question_type', 'question type', 'response type', 'answer type',
        'risk_category', 'risk category', 'risk type', 'risk classification',
        'control_objective', 'control objective', 'objective',
        'control name', 'control id', 'control ref',
        'guidance_text', 'guidance text', 'guidance', 'help text',
        'is_required', 'is required', 'mandatory',
        'risk id', 'risk_id', 'id', 'ref', 'reference', 'sr', 'sr.', 's.no', 'no.',
        'likelihood', 'impact', 'rating', 'score', 'risk rating', 'control rating',
        'control effectiveness', 'effectiveness', 'residual risk', 'inherent risk',
        'owner', 'risk owner', 'control owner', 'responsible', 'responsibility',
        'status', 'action', 'mitigation', 'treatment', 'response',
        'finding', 'observation', 'gap', 'weakness', 'issue',
        'evidence', 'testing', 'test result', 'result',
        'frequency', 'last review', 'next review', 'due date',
        'notes', 'remarks', 'comment', 'comments', 'description', 'details',
    ]
    
    HEADER_MAP = {
        'question_text': [
            'question_text', 'question text', 'question', 'assessment question',
            'control question', 'checklist item', 'requirement', 'statement',
            'assessment item', 'risk question', 'control description',
            'assessment criteria', 'criteria', 'assessment',
            'finding', 'observation', 'risk description', 'risk statement',
        ],
        'section': [
            'section', 'domain', 'category', 'area', 'control area', 'risk area',
            'group', 'topic', 'control domain', 'function', 'pillar', 'theme',
            'control category', 'risk domain', 'process area', 'department',
        ],
        'risk_id': [
            'risk id', 'risk_id', 'id', 'ref', 'reference', 'sr', 'sr.', 's.no',
            'no.', 'control id', 'control_id', 'item no', 'item_no', 'serial',
        ],
        'question_type': [
            'question_type', 'question type', 'type', 'response type', 'answer type',
        ],
        'risk_category': [
            'risk_category', 'risk category', 'risk type', 'risk classification',
        ],
        'control_objective': [
            'control_objective', 'control objective', 'objective', 'control',
            'control name', 'control ref',
        ],
        'guidance_text': [
            'guidance_text', 'guidance text', 'guidance', 'help text', 'notes',
            'explanation', 'details', 'additional info',
            'additional information', 'hint', 'comment', 'remarks',
        ],
        'is_required': [
            'is_required', 'is required', 'required', 'mandatory',
        ],
    }
    
    METADATA_LABELS = {
        'business unit', 'business unit:', 'period', 'period:', 'date', 'date:',
        'prepared by', 'prepared by:', 'reviewed by', 'reviewed by:',
        'approved by', 'approved by:', 'department', 'department:',
        'version', 'version:', 'status', 'status:', 'template', 'template:',
        'organization', 'organization:', 'company', 'company:',
    }
    
    def score_row_as_header(row_values):
        if not row_values:
            return 0
        non_empty = [str(v).strip().lower() for v in row_values if v is not None and str(v).strip()]
        if len(non_empty) < 2:
            return 0
        matches = 0
        for val in non_empty:
            for kw in ALL_HEADER_KEYWORDS:
                if val == kw or val.replace('_', ' ') == kw:
                    matches += 1
                    break
        return matches
    
    def find_header_row(all_rows):
        best_row_idx = 0
        best_score = 0
        for idx, row in enumerate(all_rows):
            s = score_row_as_header(row)
            if s > best_score:
                best_score = s
                best_row_idx = idx
        return best_row_idx if best_score >= 2 else 0
    
    def normalize_headers(raw_headers):
        header_mapping = {}
        normalized = [str(h).strip().lower() if h else '' for h in raw_headers]
        for field, aliases in HEADER_MAP.items():
            for alias in aliases:
                for idx, nh in enumerate(normalized):
                    if nh == alias and idx not in header_mapping.values():
                        header_mapping[field] = idx
                        break
                if field in header_mapping:
                    break
        if 'question_text' not in header_mapping:
            longest_text_col = -1
            longest_avg = 0
            for idx, nh in enumerate(normalized):
                if nh and idx not in header_mapping.values():
                    if longest_text_col == -1:
                        longest_text_col = idx
            if longest_text_col >= 0:
                header_mapping['question_text'] = longest_text_col
        return header_mapping
    
    def is_metadata_row(row_values):
        if not row_values:
            return True
        first_val = str(row_values[0]).strip().lower() if row_values[0] else ''
        if first_val in METADATA_LABELS or first_val.rstrip(':') in METADATA_LABELS:
            return True
        non_empty = [v for v in row_values if v is not None and str(v).strip()]
        if len(non_empty) <= 1 and first_val and not any(c.isdigit() for c in first_val[:5]):
            if len(first_val) < 30 and ':' in first_val:
                return True
        return False
    
    def extract_row(row_values, header_mapping):
        def get_val(field):
            idx = header_mapping.get(field)
            if idx is not None and idx < len(row_values):
                val = row_values[idx]
                return str(val).strip() if val is not None else ''
            return ''
        return get_val
    
    def build_questions(data_rows, header_mapping):
        questions = []
        q_order = 0
        for row_values in data_rows:
            if is_metadata_row(row_values):
                continue
            get_val = extract_row(row_values, header_mapping)
            q_text = get_val('question_text')
            if not q_text:
                non_empty_vals = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
                for val in non_empty_vals:
                    if len(val) > 10:
                        q_text = val
                        break
            if not q_text:
                continue
            section_val = get_val('section')
            risk_id_val = get_val('risk_id')
            if risk_id_val and not section_val:
                section_val = risk_id_val
            questions.append(RCSAQuestionCreate(
                section=section_val,
                question_order=q_order,
                question_text=q_text,
                question_type=get_val('question_type') or 'risk_rating',
                is_required=get_val('is_required').lower() in ('true', 'yes', '1', 'mandatory', ''),
                risk_category=get_val('risk_category') or None,
                control_objective=get_val('control_objective') or None,
                guidance_text=get_val('guidance_text') or None
            ))
            q_order += 1
        return questions
    
    try:
        if file.filename.endswith('.csv'):
            import csv
            from io import StringIO
            csv_content = content.decode('utf-8')
            reader = csv.reader(StringIO(csv_content))
            all_rows = list(reader)
            if not all_rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty CSV file")
            header_idx = find_header_row(all_rows)
            raw_headers = all_rows[header_idx]
            header_mapping = normalize_headers(raw_headers)
            data_rows = all_rows[header_idx + 1:]
            questions = build_questions(data_rows, header_mapping)
        else:
            wb = openpyxl.load_workbook(BytesIO(content))
            question_sheet_names = [
                'questions', 'questionnaire', 'assessment questionnaire',
                'assessment questions', 'survey', 'checklist', 'rcsa',
                'rcsa questions', 'self assessment', 'self-assessment',
            ]
            ws = None
            for sheet_name in wb.sheetnames:
                if sheet_name.strip().lower() in question_sheet_names:
                    ws = wb[sheet_name]
                    break
            if ws is None:
                for sheet_name in wb.sheetnames:
                    sn_lower = sheet_name.strip().lower()
                    for qsn in question_sheet_names:
                        if qsn in sn_lower or sn_lower in qsn:
                            ws = wb[sheet_name]
                            break
                    if ws is not None:
                        break
            if ws is None:
                ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row))
            if not all_rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty Excel file")
            header_idx = find_header_row(all_rows)
            raw_headers = all_rows[header_idx]
            header_mapping = normalize_headers(raw_headers)
            data_rows = all_rows[header_idx + 1:]
            questions = build_questions(data_rows, header_mapping)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse file: {str(e)}")
    
    db_template = RCSATemplate(
        tenant_id=tenant_id,
        name=name,
        description=f"Uploaded from {file.filename}",
        category=category,
        source="custom",
        version="1.0",
        is_system_template=False,
        created_by=current_user.id
    )
    db.add(db_template)
    db.flush()
    
    for q in questions:
        db_question = RCSAQuestion(
            template_id=db_template.id,
            section=q.section,
            question_order=q.question_order,
            question_text=q.question_text,
            question_type=q.question_type,
            is_required=q.is_required,
            options=q.options,
            risk_category=q.risk_category,
            control_objective=q.control_objective,
            guidance_text=q.guidance_text,
            ai_suggestion_enabled=q.ai_suggestion_enabled
        )
        db.add(db_question)
    
    db.commit()
    db.refresh(db_template)
    
    return RCSATemplateResponse(
        id=db_template.id,
        tenant_id=db_template.tenant_id,
        name=db_template.name,
        description=db_template.description,
        category=db_template.category,
        source=db_template.source,
        version=db_template.version,
        is_system_template=db_template.is_system_template,
        is_active=db_template.is_active,
        risk_categories=db_template.risk_categories or [],
        regulatory_mapping=db_template.regulatory_mapping or {},
        created_by=db_template.created_by,
        created_at=db_template.created_at,
        updated_at=db_template.updated_at,
        question_count=len(questions)
    )


@router.get("/templates/download/{template_id}")
def download_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if openpyxl is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel support not available")
    
    user_tenants = get_user_tenants(current_user, db)
    
    template = db.query(RCSATemplate).options(
        joinedload(RCSATemplate.questions)
    ).filter(
        RCSATemplate.id == template_id,
        or_(
            RCSATemplate.tenant_id.in_(user_tenants),
            RCSATemplate.is_system_template == True
        )
    ).first()
    
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "RCSA Template"
    
    headers = ["section", "question_text", "question_type", "is_required", "risk_category", "control_objective", "guidance_text"]
    ws.append(headers)
    
    for q in sorted(template.questions, key=lambda x: (x.section or "", x.question_order)):
        ws.append([
            q.section,
            q.question_text,
            q.question_type,
            str(q.is_required),
            q.risk_category,
            q.control_objective,
            q.guidance_text
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rcsa_template_{template.name.replace(' ', '_')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# =============================================================================
# Campaign Management Endpoints
# =============================================================================

@router.get("/campaigns", response_model=List[RCSACampaignResponse])
def list_campaigns(
    tenant_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSACampaign).options(
        joinedload(RCSACampaign.template),
        subqueryload(RCSACampaign.assessments)
    ).filter(RCSACampaign.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSACampaign.tenant_id == tenant_id)
    if status_filter:
        query = query.filter(RCSACampaign.status == status_filter)
    
    campaigns = query.order_by(RCSACampaign.created_at.desc()).offset(skip).limit(limit).all()
    
    return [
        RCSACampaignResponse(
            id=c.id,
            tenant_id=c.tenant_id,
            template_id=c.template_id,
            template_name=c.template.name if c.template else None,
            name=c.name,
            description=c.description,
            period_type=c.period_type,
            period_label=c.period_label,
            start_date=c.start_date,
            due_date=c.due_date,
            status=c.status,
            approval_workflow_id=c.approval_workflow_id,
            reminder_days_before=c.reminder_days_before,
            escalation_days_after=c.escalation_days_after,
            created_by=c.created_by,
            created_at=c.created_at,
            updated_at=c.updated_at,
            assessment_count=len(c.assessments),
            completed_count=sum(1 for a in c.assessments if a.status == "approved")
        )
        for c in campaigns
    ]


@router.get("/campaigns/{campaign_id}", response_model=RCSACampaignResponse)
def get_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    campaign = db.query(RCSACampaign).options(
        joinedload(RCSACampaign.template),
        joinedload(RCSACampaign.assessments).joinedload(RCSAAssessment.business_unit),
        joinedload(RCSACampaign.assessments).joinedload(RCSAAssessment.assessor)
    ).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    return RCSACampaignResponse(
        id=campaign.id,
        tenant_id=campaign.tenant_id,
        template_id=campaign.template_id,
        template_name=campaign.template.name if campaign.template else None,
        name=campaign.name,
        description=campaign.description,
        period_type=campaign.period_type,
        period_label=campaign.period_label,
        start_date=campaign.start_date,
        due_date=campaign.due_date,
        status=campaign.status,
        approval_workflow_id=campaign.approval_workflow_id,
        reminder_days_before=campaign.reminder_days_before,
        escalation_days_after=campaign.escalation_days_after,
        created_by=campaign.created_by,
        created_at=campaign.created_at,
        updated_at=campaign.updated_at,
        assessment_count=len(campaign.assessments),
        completed_count=sum(1 for a in campaign.assessments if a.status == "approved")
    )


@router.get("/campaigns/{campaign_id}/detail")
def get_campaign_detail(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    campaign = db.query(RCSACampaign).options(
        joinedload(RCSACampaign.template),
        joinedload(RCSACampaign.assessments).joinedload(RCSAAssessment.business_unit),
        joinedload(RCSACampaign.assessments).joinedload(RCSAAssessment.assessor),
        joinedload(RCSACampaign.assessments).joinedload(RCSAAssessment.findings)
    ).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    assessments_data = []
    total_findings = 0
    risk_scores = []
    control_scores = []
    completed_count = 0
    
    for a in campaign.assessments:
        findings_count = len(a.findings) if a.findings else 0
        total_findings += findings_count
        
        progress = 0
        if a.status == "approved":
            progress = 100
            completed_count += 1
        elif a.status == "submitted" or a.status == "under_review":
            progress = 100
        elif a.status == "in_progress":
            total_responses = db.query(func.count(RCSAResponse.id)).filter(
                RCSAResponse.assessment_id == a.id
            ).scalar() or 0
            total_questions = 0
            if campaign.template:
                total_questions = db.query(func.count(RCSAQuestion.id)).filter(
                    RCSAQuestion.template_id == campaign.template_id
                ).scalar() or 0
            progress = int((total_responses / total_questions * 100) if total_questions > 0 else 0)
        
        if a.overall_risk_score is not None:
            risk_scores.append(a.overall_risk_score)
        if a.overall_control_score is not None:
            control_scores.append(a.overall_control_score)
        
        assessment_data = {
            "id": a.id,
            "business_unit_id": a.business_unit_id,
            "business_unit_name": a.business_unit.name if a.business_unit else "Unknown",
            "assessor_id": a.assessor_id,
            "assessor_name": a.assessor.display_name or a.assessor.username if a.assessor else None,
            "assessor_email": a.assessor.email if a.assessor else None,
            "status": a.status,
            "progress": progress,
            "risk_score": a.overall_risk_score,
            "control_score": a.overall_control_score,
            "findings_count": findings_count,
            "submitted_at": a.submitted_at.isoformat() if a.submitted_at else None,
            "reviewed_at": a.completed_at.isoformat() if a.completed_at else None,
        }
        assessments_data.append(assessment_data)
    
    assigned_units = len(campaign.assessments)
    overall_progress = int((completed_count / assigned_units * 100) if assigned_units > 0 else 0)
    pending_assessments = sum(1 for a in campaign.assessments if a.status in ["not_started", "in_progress"])
    
    avg_risk = round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else None
    avg_control = round(sum(control_scores) / len(control_scores), 2) if control_scores else None
    
    return {
        "id": campaign.id,
        "tenant_id": campaign.tenant_id,
        "name": campaign.name,
        "description": campaign.description,
        "template_id": campaign.template_id,
        "template_name": campaign.template.name if campaign.template else None,
        "status": campaign.status,
        "period": campaign.period_label or campaign.period_type,
        "start_date": campaign.start_date.isoformat() if campaign.start_date else None,
        "end_date": campaign.due_date.isoformat() if campaign.due_date else None,
        "progress": overall_progress,
        "assigned_units": assigned_units,
        "completed_units": completed_count,
        "pending_assessments": pending_assessments,
        "assessments": assessments_data,
        "total_findings": total_findings,
        "avg_risk_score": avg_risk,
        "avg_control_score": avg_control,
        "created_at": campaign.created_at.isoformat() if campaign.created_at else None,
        "updated_at": campaign.updated_at.isoformat() if campaign.updated_at else None,
    }


@router.post("/campaigns", response_model=RCSACampaignResponse, status_code=status.HTTP_201_CREATED)
def create_campaign(
    campaign: RCSACampaignCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User not assigned to any tenant")
    
    template = db.query(RCSATemplate).filter(RCSATemplate.id == campaign.template_id).first()
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    db_campaign = RCSACampaign(
        tenant_id=tenant_id,
        template_id=campaign.template_id,
        name=campaign.name,
        description=campaign.description,
        period_type=campaign.period_type,
        period_label=campaign.period_label,
        start_date=campaign.start_date,
        due_date=campaign.due_date,
        approval_workflow_id=campaign.approval_workflow_id,
        reminder_days_before=campaign.reminder_days_before,
        escalation_days_after=campaign.escalation_days_after,
        created_by=current_user.id
    )
    db.add(db_campaign)
    db.flush()
    
    for bu_id in campaign.business_unit_ids:
        bu = db.query(BusinessUnit).filter(
            BusinessUnit.id == bu_id,
            BusinessUnit.tenant_id == tenant_id
        ).first()
        if bu:
            assessment = RCSAAssessment(
                tenant_id=tenant_id,
                campaign_id=db_campaign.id,
                business_unit_id=bu_id
            )
            db.add(assessment)
    
    db.commit()
    db.refresh(db_campaign)
    
    return RCSACampaignResponse(
        id=db_campaign.id,
        tenant_id=db_campaign.tenant_id,
        template_id=db_campaign.template_id,
        template_name=template.name,
        name=db_campaign.name,
        description=db_campaign.description,
        period_type=db_campaign.period_type,
        period_label=db_campaign.period_label,
        start_date=db_campaign.start_date,
        due_date=db_campaign.due_date,
        status=db_campaign.status,
        approval_workflow_id=db_campaign.approval_workflow_id,
        reminder_days_before=db_campaign.reminder_days_before,
        escalation_days_after=db_campaign.escalation_days_after,
        created_by=db_campaign.created_by,
        created_at=db_campaign.created_at,
        updated_at=db_campaign.updated_at,
        assessment_count=len(campaign.business_unit_ids),
        completed_count=0
    )


@router.put("/campaigns/{campaign_id}", response_model=RCSACampaignResponse)
def update_campaign(
    campaign_id: int,
    campaign: RCSACampaignUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_campaign = db.query(RCSACampaign).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    update_data = campaign.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_campaign, key, value)
    
    db.commit()
    db.refresh(db_campaign)
    
    template = db.query(RCSATemplate).filter(RCSATemplate.id == db_campaign.template_id).first()
    assessment_count = db.query(RCSAAssessment).filter(RCSAAssessment.campaign_id == campaign_id).count()
    completed_count = db.query(RCSAAssessment).filter(
        RCSAAssessment.campaign_id == campaign_id,
        RCSAAssessment.status == "approved"
    ).count()
    
    return RCSACampaignResponse(
        id=db_campaign.id,
        tenant_id=db_campaign.tenant_id,
        template_id=db_campaign.template_id,
        template_name=template.name if template else None,
        name=db_campaign.name,
        description=db_campaign.description,
        period_type=db_campaign.period_type,
        period_label=db_campaign.period_label,
        start_date=db_campaign.start_date,
        due_date=db_campaign.due_date,
        status=db_campaign.status,
        approval_workflow_id=db_campaign.approval_workflow_id,
        reminder_days_before=db_campaign.reminder_days_before,
        escalation_days_after=db_campaign.escalation_days_after,
        created_by=db_campaign.created_by,
        created_at=db_campaign.created_at,
        updated_at=db_campaign.updated_at,
        assessment_count=assessment_count,
        completed_count=completed_count
    )


@router.delete("/campaigns/{campaign_id}", response_model=MessageResponse)
def delete_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_campaign = db.query(RCSACampaign).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    if db_campaign.status == "active":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot delete an active campaign")
    
    db.delete(db_campaign)
    db.commit()
    
    return MessageResponse(message="Campaign deleted successfully", id=campaign_id)


@router.post("/campaigns/{campaign_id}/activate", response_model=RCSACampaignResponse)
def activate_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_campaign = db.query(RCSACampaign).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    if db_campaign.status != "draft":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Campaign can only be activated from draft status")
    
    db_campaign.status = "active"
    
    existing_assessments = db.query(RCSAAssessment).filter(
        RCSAAssessment.campaign_id == campaign_id
    ).count()
    
    if existing_assessments == 0:
        business_units = db.query(BusinessUnit).filter(
            BusinessUnit.tenant_id == db_campaign.tenant_id
        ).all()
        for bu in business_units:
            assessment = RCSAAssessment(
                tenant_id=db_campaign.tenant_id,
                campaign_id=campaign_id,
                business_unit_id=bu.id
            )
            db.add(assessment)
    
    db.commit()
    db.refresh(db_campaign)
    
    template = db.query(RCSATemplate).filter(RCSATemplate.id == db_campaign.template_id).first()
    assessment_count = db.query(RCSAAssessment).filter(RCSAAssessment.campaign_id == campaign_id).count()
    
    return RCSACampaignResponse(
        id=db_campaign.id,
        tenant_id=db_campaign.tenant_id,
        template_id=db_campaign.template_id,
        template_name=template.name if template else None,
        name=db_campaign.name,
        description=db_campaign.description,
        period_type=db_campaign.period_type,
        period_label=db_campaign.period_label,
        start_date=db_campaign.start_date,
        due_date=db_campaign.due_date,
        status=db_campaign.status,
        approval_workflow_id=db_campaign.approval_workflow_id,
        reminder_days_before=db_campaign.reminder_days_before,
        escalation_days_after=db_campaign.escalation_days_after,
        created_by=db_campaign.created_by,
        created_at=db_campaign.created_at,
        updated_at=db_campaign.updated_at,
        assessment_count=assessment_count,
        completed_count=0
    )


@router.post("/campaigns/{campaign_id}/close", response_model=RCSACampaignResponse)
def close_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_campaign = db.query(RCSACampaign).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    db_campaign.status = "closed"
    db.commit()
    db.refresh(db_campaign)
    
    template = db.query(RCSATemplate).filter(RCSATemplate.id == db_campaign.template_id).first()
    assessment_count = db.query(RCSAAssessment).filter(RCSAAssessment.campaign_id == campaign_id).count()
    completed_count = db.query(RCSAAssessment).filter(
        RCSAAssessment.campaign_id == campaign_id,
        RCSAAssessment.status == "approved"
    ).count()
    
    return RCSACampaignResponse(
        id=db_campaign.id,
        tenant_id=db_campaign.tenant_id,
        template_id=db_campaign.template_id,
        template_name=template.name if template else None,
        name=db_campaign.name,
        description=db_campaign.description,
        period_type=db_campaign.period_type,
        period_label=db_campaign.period_label,
        start_date=db_campaign.start_date,
        due_date=db_campaign.due_date,
        status=db_campaign.status,
        approval_workflow_id=db_campaign.approval_workflow_id,
        reminder_days_before=db_campaign.reminder_days_before,
        escalation_days_after=db_campaign.escalation_days_after,
        created_by=db_campaign.created_by,
        created_at=db_campaign.created_at,
        updated_at=db_campaign.updated_at,
        assessment_count=assessment_count,
        completed_count=completed_count
    )


@router.post("/campaigns/{campaign_id}/assign", response_model=MessageResponse)
def assign_business_units(
    campaign_id: int,
    request: RCSABUAssignRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_campaign = db.query(RCSACampaign).filter(
        RCSACampaign.id == campaign_id,
        RCSACampaign.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    created_count = 0
    for bu_id in request.business_unit_ids:
        existing = db.query(RCSAAssessment).filter(
            RCSAAssessment.campaign_id == campaign_id,
            RCSAAssessment.business_unit_id == bu_id
        ).first()
        
        if not existing:
            assessor_id = request.assessor_ids.get(bu_id) if request.assessor_ids else None
            assessment = RCSAAssessment(
                tenant_id=db_campaign.tenant_id,
                campaign_id=campaign_id,
                business_unit_id=bu_id,
                assessor_id=assessor_id,
                assigned_at=datetime.utcnow() if assessor_id else None
            )
            db.add(assessment)
            created_count += 1
    
    db.commit()
    
    return MessageResponse(message=f"Assigned {created_count} business units to campaign", id=campaign_id)


# =============================================================================
# Assessment Management Endpoints
# =============================================================================

@router.get("/assessments", response_model=List[RCSAAssessmentResponse])
def list_assessments(
    campaign_id: Optional[int] = None,
    business_unit_id: Optional[int] = None,
    status_filter: Optional[str] = None,
    assessor_id: Optional[int] = None,
    mine: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor),
        joinedload(RCSAAssessment.campaign),
        joinedload(RCSAAssessment.responses),
        joinedload(RCSAAssessment.findings)
    ).filter(RCSAAssessment.tenant_id.in_(user_tenants))
    
    if mine:
        query = query.filter(RCSAAssessment.assessor_id == current_user.id)
    if campaign_id:
        query = query.filter(RCSAAssessment.campaign_id == campaign_id)
    if business_unit_id:
        query = query.filter(RCSAAssessment.business_unit_id == business_unit_id)
    if status_filter:
        query = query.filter(RCSAAssessment.status == status_filter)
    if assessor_id:
        query = query.filter(RCSAAssessment.assessor_id == assessor_id)
    
    assessments = query.order_by(RCSAAssessment.created_at.desc()).offset(skip).limit(limit).all()
    
    return [
        RCSAAssessmentResponse(
            id=a.id,
            tenant_id=a.tenant_id,
            campaign_id=a.campaign_id,
            campaign_name=a.campaign.name if a.campaign else None,
            business_unit_id=a.business_unit_id,
            business_unit_name=a.business_unit.name if a.business_unit else None,
            status=a.status,
            current_approval_tier=a.current_approval_tier,
            assessor_id=a.assessor_id,
            assessor_name=a.assessor.display_name if a.assessor else None,
            due_date=a.campaign.due_date if a.campaign else None,
            assigned_at=a.assigned_at,
            started_at=a.started_at,
            submitted_at=a.submitted_at,
            completed_at=a.completed_at,
            overall_risk_score=a.overall_risk_score,
            overall_control_score=a.overall_control_score,
            ai_quality_score=a.ai_quality_score,
            ai_suggestions_used=a.ai_suggestions_used,
            ai_gaps_identified=a.ai_gaps_identified,
            notes=a.notes,
            created_at=a.created_at,
            updated_at=a.updated_at,
            response_count=len(a.responses),
            finding_count=len(a.findings)
        )
        for a in assessments
    ]


@router.get("/assessments/{assessment_id}", response_model=RCSAAssessmentResponse)
def get_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor),
        joinedload(RCSAAssessment.responses).joinedload(RCSAResponse.question),
        joinedload(RCSAAssessment.findings),
        joinedload(RCSAAssessment.approval_history)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=assessment.business_unit.name if assessment.business_unit else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=len(assessment.responses),
        finding_count=len(assessment.findings)
    )


@router.get("/assessments/{assessment_id}/detail", response_model=RCSAAssessmentDetailResponse)
def get_assessment_detail(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Get detailed assessment with questions and responses for the assessment form"""
    from ....models import RCSAResponseEvidence
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor),
        joinedload(RCSAAssessment.responses).joinedload(RCSAResponse.evidence_links).joinedload(RCSAResponseEvidence.evidence),
        joinedload(RCSAAssessment.campaign)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    campaign = assessment.campaign
    if not campaign:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Campaign not found")
    
    template = db.query(RCSATemplate).options(
        joinedload(RCSATemplate.questions)
    ).filter(RCSATemplate.id == campaign.template_id).first()
    
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    questions = []
    for q in sorted(template.questions, key=lambda x: (x.section or "", x.question_order or 0)):
        questions.append(RCSAQuestionWithResponse(
            id=q.id,
            section=q.section,
            question_text=q.question_text,
            guidance=q.guidance_text,
            question_type=q.question_type,
            is_required=q.is_required,
            sequence=q.question_order or 0,
            question_order=q.question_order or 0,
            ai_suggestion_enabled=q.ai_suggestion_enabled or False,
            risk_category=q.risk_category,
            control_objective=q.control_objective
        ))
    
    responses = []
    for r in assessment.responses:
        evidence_files = []
        if hasattr(r, 'evidence_links') and r.evidence_links:
            for link in r.evidence_links:
                if link.evidence:
                    evidence_files.append(RCSAEvidenceFile(
                        id=link.evidence.id,
                        filename=link.evidence.file_name or link.evidence.name,
                        file_size=0,
                        uploaded_at=link.evidence.uploaded_at
                    ))
        responses.append(RCSAResponseDetail(
            question_id=r.question_id,
            likelihood=r.likelihood_rating,
            impact=r.impact_rating,
            effectiveness=r.control_effectiveness,
            yes_no_value=r.response_value == "yes" if r.response_value in ["yes", "no"] else None,
            text_value=r.response_value if r.response_value not in ["yes", "no"] else None,
            evidence=evidence_files
        ))
    
    total_required = len([q for q in questions if q.is_required])
    answered = len([r for r in responses if r.likelihood or r.impact or r.effectiveness or r.yes_no_value is not None or r.text_value])
    progress = (answered / total_required * 100) if total_required > 0 else 0
    
    return RCSAAssessmentDetailResponse(
        id=assessment.id,
        campaign_id=assessment.campaign_id,
        campaign_name=campaign.name,
        business_unit=assessment.business_unit.name if assessment.business_unit else None,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        status=assessment.status,
        due_date=campaign.due_date,
        progress=progress,
        questions=questions,
        responses=responses
    )


@router.post("/assessments/{assessment_id}/start", response_model=RCSAAssessmentResponse)
def start_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    if assessment.status not in ["not_started", "requires_changes"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment cannot be started")
    
    assessment.status = "in_progress"
    assessment.started_at = datetime.utcnow()
    if not assessment.assessor_id:
        assessment.assessor_id = current_user.id
        assessment.assigned_at = datetime.utcnow()
    
    db.commit()
    db.refresh(assessment)
    
    bu = db.query(BusinessUnit).filter(BusinessUnit.id == assessment.business_unit_id).first()
    assessor = db.query(GRCUser).filter(GRCUser.id == assessment.assessor_id).first()
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=bu.name if bu else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessor.display_name if assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=0,
        finding_count=0
    )


@router.post("/assessments/{assessment_id}/save", response_model=MessageResponse)
def save_assessment_responses(
    assessment_id: int,
    request: RCSABulkResponseSave,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    if assessment.status not in ["in_progress", "requires_changes"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not in editable status")
    
    saved_count = 0
    for resp in request.responses:
        existing = db.query(RCSAResponse).filter(
            RCSAResponse.assessment_id == assessment_id,
            RCSAResponse.question_id == resp.question_id
        ).first()
        
        risk_score = None
        if resp.likelihood_rating and resp.impact_rating:
            risk_score = resp.likelihood_rating * resp.impact_rating
        
        if existing:
            existing.response_value = resp.response_value
            existing.likelihood_rating = resp.likelihood_rating
            existing.impact_rating = resp.impact_rating
            existing.risk_score = risk_score
            existing.control_effectiveness = resp.control_effectiveness
            existing.control_description = resp.control_description
            existing.last_tested_date = resp.last_tested_date
            existing.responded_by = current_user.id
            existing.responded_at = datetime.utcnow()
        else:
            db_response = RCSAResponse(
                assessment_id=assessment_id,
                question_id=resp.question_id,
                response_value=resp.response_value,
                likelihood_rating=resp.likelihood_rating,
                impact_rating=resp.impact_rating,
                risk_score=risk_score,
                control_effectiveness=resp.control_effectiveness,
                control_description=resp.control_description,
                last_tested_date=resp.last_tested_date,
                responded_by=current_user.id,
                responded_at=datetime.utcnow()
            )
            db.add(db_response)
        saved_count += 1
    
    db.commit()
    
    return MessageResponse(message=f"Saved {saved_count} responses", id=assessment_id)


@router.post("/assessments/{assessment_id}/submit", response_model=RCSAAssessmentResponse)
def submit_assessment(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.responses),
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    if assessment.status not in ["in_progress", "requires_changes"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment cannot be submitted")
    
    risk_scores = [r.risk_score for r in assessment.responses if r.risk_score]
    if risk_scores:
        assessment.overall_risk_score = sum(risk_scores) / len(risk_scores)
    
    effectiveness_map = {"effective": 3, "partially_effective": 2, "ineffective": 1, "not_applicable": None}
    control_scores = [
        effectiveness_map.get(r.control_effectiveness)
        for r in assessment.responses
        if r.control_effectiveness and effectiveness_map.get(r.control_effectiveness)
    ]
    if control_scores:
        assessment.overall_control_score = sum(control_scores) / len(control_scores)
    
    assessment.status = "submitted"
    assessment.submitted_at = datetime.utcnow()
    assessment.current_approval_tier = 1
    
    history = RCSAApprovalHistory(
        assessment_id=assessment_id,
        action="submitted",
        tier_number=0,
        performed_by=current_user.id
    )
    db.add(history)
    
    db.commit()
    db.refresh(assessment)
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=assessment.business_unit.name if assessment.business_unit else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=len(assessment.responses),
        finding_count=len(assessment.findings) if hasattr(assessment, 'findings') else 0
    )


@router.get("/assessments/{assessment_id}/ai-suggestions", response_model=List[RCSAAISuggestionResponse])
def get_ai_suggestions(
    assessment_id: int,
    question_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    campaign = db.query(RCSACampaign).filter(RCSACampaign.id == assessment.campaign_id).first()
    template = db.query(RCSATemplate).options(
        joinedload(RCSATemplate.questions)
    ).filter(RCSATemplate.id == campaign.template_id).first()
    
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template not found")
    
    bu = db.query(BusinessUnit).filter(BusinessUnit.id == assessment.business_unit_id).first()
    bu_name = bu.name if bu else "Unknown Business Unit"
    
    suggestions = []

    if question_id is not None:
        target_questions = [q for q in template.questions if q.id == question_id]
        if not target_questions:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found in template")
    else:
        target_questions = list(template.questions)

    for question in target_questions:
        if not question.ai_suggestion_enabled:
            continue
        
        existing_response = db.query(RCSAResponse).filter(
            RCSAResponse.assessment_id == assessment_id,
            RCSAResponse.question_id == question.id
        ).first()
        
        suggestion_text = ""
        confidence = 0.0
        gaps = []
        
        evidence_recs = []
        
        if client:
            try:
                prompt = f"""You are an enterprise GRC expert. For the following RCSA assessment question, suggest what specific types of evidence the user should upload to demonstrate compliance.

Business Unit: {bu_name}
Question: {question.question_text}
Question Type: {question.question_type}
Risk Category: {question.risk_category or 'General'}
Control Objective: {question.control_objective or 'Not specified'}

Provide:
1. A concise suggestion for how to respond
2. 2-4 specific evidence types the user should upload, with descriptions and example file names

Format your response as JSON:
{{"suggestion": "your suggestion text", "confidence": 0.85, "gaps": ["gap1"], "evidence_recommendations": [{{"evidence_type": "Policy Document", "description": "Formal policy covering this control area", "example_files": ["access-control-policy.pdf", "security-policy-v2.docx"]}}, ...]}}"""

                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[{"role": "user", "content": prompt}],
                    response_format={"type": "json_object"}
                )
                
                result = json.loads(response.choices[0].message.content)
                suggestion_text = result.get("suggestion", "")
                confidence = result.get("confidence", 0.7)
                gaps = result.get("gaps", [])
                evidence_recs = result.get("evidence_recommendations", [])
            except Exception:
                suggestion_text, evidence_recs = get_question_specific_evidence(
                    question.question_text, question.question_type,
                    question.risk_category, question.control_objective
                )
                confidence = 0.5
        else:
            suggestion_text, evidence_recs = get_question_specific_evidence(
                question.question_text, question.question_type,
                question.risk_category, question.control_objective
            )
            confidence = 0.5
        
        suggestions.append(RCSAAISuggestionResponse(
            question_id=question.id,
            suggestion=suggestion_text,
            confidence=confidence,
            reasoning=f"Based on {bu_name}'s operational context",
            gaps_detected=gaps,
            evidence_recommendations=evidence_recs
        ))
    
    return suggestions


# =============================================================================
# Approval Workflow Endpoints
# =============================================================================

@router.get("/approval-workflows", response_model=List[RCSAApprovalWorkflowResponse])
def list_approval_workflows(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSAApprovalWorkflow).options(
        joinedload(RCSAApprovalWorkflow.tiers)
    ).filter(
        RCSAApprovalWorkflow.tenant_id.in_(user_tenants),
        RCSAApprovalWorkflow.is_active == True
    )
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSAApprovalWorkflow.tenant_id == tenant_id)
    
    workflows = query.all()
    
    return [
        RCSAApprovalWorkflowResponse(
            id=w.id,
            tenant_id=w.tenant_id,
            name=w.name,
            description=w.description,
            is_default=w.is_default,
            is_active=w.is_active,
            created_by=w.created_by,
            created_at=w.created_at,
            updated_at=w.updated_at,
            tier_count=len(w.tiers),
            tiers=[
                RCSAApprovalTierResponse(
                    id=t.id,
                    workflow_id=t.workflow_id,
                    tier_order=t.tier_order,
                    tier_name=t.tier_name,
                    approver_type=t.approver_type,
                    approver_role_id=t.approver_role_id,
                    approver_user_id=t.approver_user_id,
                    can_delegate=t.can_delegate,
                    auto_approve_days=t.auto_approve_days
                )
                for t in sorted(w.tiers, key=lambda x: x.tier_order)
            ]
        )
        for w in workflows
    ]


@router.post("/approval-workflows", response_model=RCSAApprovalWorkflowResponse, status_code=status.HTTP_201_CREATED)
def create_approval_workflow(
    workflow: RCSAApprovalWorkflowCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="User not assigned to any tenant")
    
    db_workflow = RCSAApprovalWorkflow(
        tenant_id=tenant_id,
        name=workflow.name,
        description=workflow.description,
        is_default=workflow.is_default,
        created_by=current_user.id
    )
    db.add(db_workflow)
    db.flush()
    
    for tier in workflow.tiers:
        db_tier = RCSAApprovalTier(
            workflow_id=db_workflow.id,
            tier_order=tier.tier_order,
            tier_name=tier.tier_name,
            approver_type=tier.approver_type,
            approver_role_id=tier.approver_role_id,
            approver_user_id=tier.approver_user_id,
            can_delegate=tier.can_delegate,
            auto_approve_days=tier.auto_approve_days
        )
        db.add(db_tier)
    
    db.commit()
    db.refresh(db_workflow)
    
    return RCSAApprovalWorkflowResponse(
        id=db_workflow.id,
        tenant_id=db_workflow.tenant_id,
        name=db_workflow.name,
        description=db_workflow.description,
        is_default=db_workflow.is_default,
        is_active=db_workflow.is_active,
        created_by=db_workflow.created_by,
        created_at=db_workflow.created_at,
        updated_at=db_workflow.updated_at,
        tier_count=len(workflow.tiers),
        tiers=[]
    )


@router.put("/approval-workflows/{workflow_id}", response_model=RCSAApprovalWorkflowResponse)
def update_approval_workflow(
    workflow_id: int,
    workflow: RCSAApprovalWorkflowUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_workflow = db.query(RCSAApprovalWorkflow).filter(
        RCSAApprovalWorkflow.id == workflow_id,
        RCSAApprovalWorkflow.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_workflow:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Workflow not found")
    
    update_data = workflow.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_workflow, key, value)
    
    db.commit()
    db.refresh(db_workflow)
    
    return RCSAApprovalWorkflowResponse(
        id=db_workflow.id,
        tenant_id=db_workflow.tenant_id,
        name=db_workflow.name,
        description=db_workflow.description,
        is_default=db_workflow.is_default,
        is_active=db_workflow.is_active,
        created_by=db_workflow.created_by,
        created_at=db_workflow.created_at,
        updated_at=db_workflow.updated_at,
        tier_count=len(db_workflow.tiers),
        tiers=[]
    )


@router.post("/assessments/{assessment_id}/approve", response_model=RCSAAssessmentResponse)
def approve_assessment(
    assessment_id: int,
    action: RCSAApprovalAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    if assessment.status not in ["submitted", "under_review"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not pending approval")
    
    campaign = db.query(RCSACampaign).filter(RCSACampaign.id == assessment.campaign_id).first()
    workflow = None
    if campaign and campaign.approval_workflow_id:
        workflow = db.query(RCSAApprovalWorkflow).options(
            joinedload(RCSAApprovalWorkflow.tiers)
        ).filter(RCSAApprovalWorkflow.id == campaign.approval_workflow_id).first()
    
    current_tier = assessment.current_approval_tier
    max_tiers = len(workflow.tiers) if workflow else 1
    
    history = RCSAApprovalHistory(
        assessment_id=assessment_id,
        action="approved",
        tier_number=current_tier,
        performed_by=current_user.id,
        comments=action.comments
    )
    db.add(history)
    
    if current_tier >= max_tiers:
        assessment.status = "approved"
        assessment.completed_at = datetime.utcnow()
    else:
        assessment.status = "under_review"
        assessment.current_approval_tier = current_tier + 1
    
    db.commit()
    db.refresh(assessment)
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=assessment.business_unit.name if assessment.business_unit else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=0,
        finding_count=0
    )


@router.post("/assessments/{assessment_id}/reject", response_model=RCSAAssessmentResponse)
def reject_assessment(
    assessment_id: int,
    action: RCSAApprovalAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    if assessment.status not in ["submitted", "under_review"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Assessment is not pending approval")
    
    history = RCSAApprovalHistory(
        assessment_id=assessment_id,
        action="rejected",
        tier_number=assessment.current_approval_tier,
        performed_by=current_user.id,
        comments=action.comments
    )
    db.add(history)
    
    assessment.status = "rejected"
    
    db.commit()
    db.refresh(assessment)
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=assessment.business_unit.name if assessment.business_unit else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=0,
        finding_count=0
    )


@router.post("/assessments/{assessment_id}/return", response_model=RCSAAssessmentResponse)
def return_assessment(
    assessment_id: int,
    action: RCSAApprovalAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit),
        joinedload(RCSAAssessment.assessor)
    ).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    history = RCSAApprovalHistory(
        assessment_id=assessment_id,
        action="returned",
        tier_number=assessment.current_approval_tier,
        performed_by=current_user.id,
        comments=action.comments
    )
    db.add(history)
    
    assessment.status = "requires_changes"
    assessment.current_approval_tier = 0
    
    db.commit()
    db.refresh(assessment)
    
    return RCSAAssessmentResponse(
        id=assessment.id,
        tenant_id=assessment.tenant_id,
        campaign_id=assessment.campaign_id,
        business_unit_id=assessment.business_unit_id,
        business_unit_name=assessment.business_unit.name if assessment.business_unit else None,
        status=assessment.status,
        current_approval_tier=assessment.current_approval_tier,
        assessor_id=assessment.assessor_id,
        assessor_name=assessment.assessor.display_name if assessment.assessor else None,
        assigned_at=assessment.assigned_at,
        started_at=assessment.started_at,
        submitted_at=assessment.submitted_at,
        completed_at=assessment.completed_at,
        overall_risk_score=assessment.overall_risk_score,
        overall_control_score=assessment.overall_control_score,
        ai_quality_score=assessment.ai_quality_score,
        ai_suggestions_used=assessment.ai_suggestions_used,
        ai_gaps_identified=assessment.ai_gaps_identified,
        notes=assessment.notes,
        created_at=assessment.created_at,
        updated_at=assessment.updated_at,
        response_count=0,
        finding_count=0
    )


@router.post("/assessments/{assessment_id}/delegate", response_model=MessageResponse)
def delegate_approval(
    assessment_id: int,
    action: RCSADelegateAction,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    delegate_user = db.query(GRCUser).filter(GRCUser.id == action.delegate_to_user_id).first()
    if not delegate_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Delegate user not found")
    
    history = RCSAApprovalHistory(
        assessment_id=assessment_id,
        action="delegated",
        tier_number=assessment.current_approval_tier,
        performed_by=current_user.id,
        delegated_to=action.delegate_to_user_id,
        comments=action.comments
    )
    db.add(history)
    db.commit()
    
    return MessageResponse(
        message=f"Approval delegated to {delegate_user.display_name or delegate_user.username}",
        id=assessment_id
    )


# =============================================================================
# Findings Endpoints
# =============================================================================

@router.get("/findings", response_model=List[RCSAFindingResponse])
def list_findings(
    tenant_id: Optional[int] = None,
    assessment_id: Optional[int] = None,
    severity: Optional[str] = None,
    status_filter: Optional[str] = None,
    finding_type: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSAFinding).options(
        joinedload(RCSAFinding.remediation_owner)
    ).filter(RCSAFinding.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSAFinding.tenant_id == tenant_id)
    if assessment_id:
        query = query.filter(RCSAFinding.assessment_id == assessment_id)
    if severity:
        query = query.filter(RCSAFinding.severity == severity)
    if status_filter:
        query = query.filter(RCSAFinding.status == status_filter)
    if finding_type:
        query = query.filter(RCSAFinding.finding_type == finding_type)
    
    findings = query.order_by(RCSAFinding.created_at.desc()).offset(skip).limit(limit).all()
    
    return [
        RCSAFindingResponse(
            id=f.id,
            tenant_id=f.tenant_id,
            assessment_id=f.assessment_id,
            finding_type=f.finding_type,
            severity=f.severity,
            title=f.title,
            description=f.description,
            risk_category=f.risk_category,
            affected_controls=f.affected_controls or [],
            ai_generated=f.ai_generated,
            ai_recommendation=f.ai_recommendation,
            linked_risk_id=f.linked_risk_id,
            linked_internal_control_id=f.linked_internal_control_id,
            linked_mitigation_action_id=f.linked_mitigation_action_id,
            status=f.status,
            remediation_due_date=f.remediation_due_date,
            remediation_owner_id=f.remediation_owner_id,
            remediation_owner_name=f.remediation_owner.display_name if f.remediation_owner else None,
            created_at=f.created_at,
            updated_at=f.updated_at,
            closed_at=f.closed_at
        )
        for f in findings
    ]


@router.post("/findings", response_model=RCSAFindingResponse, status_code=status.HTTP_201_CREATED)
def create_finding(
    finding: RCSAFindingCreate,
    assessment_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    assessment = db.query(RCSAAssessment).filter(
        RCSAAssessment.id == assessment_id,
        RCSAAssessment.tenant_id.in_(user_tenants)
    ).first()
    
    if not assessment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assessment not found")
    
    db_finding = RCSAFinding(
        tenant_id=assessment.tenant_id,
        assessment_id=assessment_id,
        finding_type=finding.finding_type,
        severity=finding.severity,
        title=finding.title,
        description=finding.description,
        risk_category=finding.risk_category,
        affected_controls=finding.affected_controls,
        remediation_due_date=finding.remediation_due_date,
        remediation_owner_id=finding.remediation_owner_id
    )
    db.add(db_finding)
    db.commit()
    db.refresh(db_finding)
    
    return RCSAFindingResponse(
        id=db_finding.id,
        tenant_id=db_finding.tenant_id,
        assessment_id=db_finding.assessment_id,
        finding_type=db_finding.finding_type,
        severity=db_finding.severity,
        title=db_finding.title,
        description=db_finding.description,
        risk_category=db_finding.risk_category,
        affected_controls=db_finding.affected_controls or [],
        ai_generated=db_finding.ai_generated,
        ai_recommendation=db_finding.ai_recommendation,
        linked_risk_id=db_finding.linked_risk_id,
        linked_internal_control_id=db_finding.linked_internal_control_id,
        linked_mitigation_action_id=db_finding.linked_mitigation_action_id,
        status=db_finding.status,
        remediation_due_date=db_finding.remediation_due_date,
        remediation_owner_id=db_finding.remediation_owner_id,
        remediation_owner_name=None,
        created_at=db_finding.created_at,
        updated_at=db_finding.updated_at,
        closed_at=db_finding.closed_at
    )


@router.put("/findings/{finding_id}", response_model=RCSAFindingResponse)
def update_finding(
    finding_id: int,
    finding: RCSAFindingUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_finding = db.query(RCSAFinding).filter(
        RCSAFinding.id == finding_id,
        RCSAFinding.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_finding:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Finding not found")
    
    update_data = finding.model_dump(exclude_unset=True)
    
    if "status" in update_data and update_data["status"] == "closed":
        db_finding.closed_at = datetime.utcnow()
    
    for key, value in update_data.items():
        setattr(db_finding, key, value)
    
    db.commit()
    db.refresh(db_finding)
    
    owner = db.query(GRCUser).filter(GRCUser.id == db_finding.remediation_owner_id).first() if db_finding.remediation_owner_id else None
    
    return RCSAFindingResponse(
        id=db_finding.id,
        tenant_id=db_finding.tenant_id,
        assessment_id=db_finding.assessment_id,
        finding_type=db_finding.finding_type,
        severity=db_finding.severity,
        title=db_finding.title,
        description=db_finding.description,
        risk_category=db_finding.risk_category,
        affected_controls=db_finding.affected_controls or [],
        ai_generated=db_finding.ai_generated,
        ai_recommendation=db_finding.ai_recommendation,
        linked_risk_id=db_finding.linked_risk_id,
        linked_internal_control_id=db_finding.linked_internal_control_id,
        linked_mitigation_action_id=db_finding.linked_mitigation_action_id,
        status=db_finding.status,
        remediation_due_date=db_finding.remediation_due_date,
        remediation_owner_id=db_finding.remediation_owner_id,
        remediation_owner_name=owner.display_name if owner else None,
        created_at=db_finding.created_at,
        updated_at=db_finding.updated_at,
        closed_at=db_finding.closed_at
    )


@router.post("/findings/{finding_id}/link-risk", response_model=RCSAFindingResponse)
def link_finding_to_risk(
    finding_id: int,
    risk_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_finding = db.query(RCSAFinding).filter(
        RCSAFinding.id == finding_id,
        RCSAFinding.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_finding:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Finding not found")
    
    risk = db.query(Risk).filter(
        Risk.id == risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    
    db_finding.linked_risk_id = risk_id
    db.commit()
    db.refresh(db_finding)
    
    return RCSAFindingResponse(
        id=db_finding.id,
        tenant_id=db_finding.tenant_id,
        assessment_id=db_finding.assessment_id,
        finding_type=db_finding.finding_type,
        severity=db_finding.severity,
        title=db_finding.title,
        description=db_finding.description,
        risk_category=db_finding.risk_category,
        affected_controls=db_finding.affected_controls or [],
        ai_generated=db_finding.ai_generated,
        ai_recommendation=db_finding.ai_recommendation,
        linked_risk_id=db_finding.linked_risk_id,
        linked_internal_control_id=db_finding.linked_internal_control_id,
        linked_mitigation_action_id=db_finding.linked_mitigation_action_id,
        status=db_finding.status,
        remediation_due_date=db_finding.remediation_due_date,
        remediation_owner_id=db_finding.remediation_owner_id,
        remediation_owner_name=None,
        created_at=db_finding.created_at,
        updated_at=db_finding.updated_at,
        closed_at=db_finding.closed_at
    )


@router.post("/findings/{finding_id}/link-control", response_model=RCSAFindingResponse)
def link_finding_to_control(
    finding_id: int,
    control_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_finding = db.query(RCSAFinding).filter(
        RCSAFinding.id == finding_id,
        RCSAFinding.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_finding:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Finding not found")
    
    control = db.query(InternalControl).filter(
        InternalControl.id == control_id,
        InternalControl.tenant_id.in_(user_tenants)
    ).first()
    
    if not control:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Internal control not found")
    
    db_finding.linked_internal_control_id = control_id
    db.commit()
    db.refresh(db_finding)
    
    return RCSAFindingResponse(
        id=db_finding.id,
        tenant_id=db_finding.tenant_id,
        assessment_id=db_finding.assessment_id,
        finding_type=db_finding.finding_type,
        severity=db_finding.severity,
        title=db_finding.title,
        description=db_finding.description,
        risk_category=db_finding.risk_category,
        affected_controls=db_finding.affected_controls or [],
        ai_generated=db_finding.ai_generated,
        ai_recommendation=db_finding.ai_recommendation,
        linked_risk_id=db_finding.linked_risk_id,
        linked_internal_control_id=db_finding.linked_internal_control_id,
        linked_mitigation_action_id=db_finding.linked_mitigation_action_id,
        status=db_finding.status,
        remediation_due_date=db_finding.remediation_due_date,
        remediation_owner_id=db_finding.remediation_owner_id,
        remediation_owner_name=None,
        created_at=db_finding.created_at,
        updated_at=db_finding.updated_at,
        closed_at=db_finding.closed_at
    )


@router.post("/findings/{finding_id}/create-action", response_model=MessageResponse)
def create_mitigation_action_from_finding(
    finding_id: int,
    risk_id: Optional[int] = Query(None, description="Risk ID to link the action to. Required if finding is not linked to a risk."),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    
    db_finding = db.query(RCSAFinding).filter(
        RCSAFinding.id == finding_id,
        RCSAFinding.tenant_id.in_(user_tenants)
    ).first()
    
    if not db_finding:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Finding not found")
    
    target_risk_id = risk_id or db_finding.linked_risk_id
    if not target_risk_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Finding must be linked to a risk or risk_id must be provided to create a mitigation action"
        )
    
    risk = db.query(Risk).filter(
        Risk.id == target_risk_id,
        Risk.tenant_id.in_(user_tenants)
    ).first()
    
    if not risk:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Risk not found")
    
    action = RiskMitigationAction(
        risk_id=target_risk_id,
        title=f"Remediate: {db_finding.title}",
        description=db_finding.description or f"Mitigation action for RCSA finding: {db_finding.title}",
        action_type="mitigate",
        priority="high" if db_finding.severity in ["critical", "high"] else "medium",
        status="open",
        due_date=db_finding.remediation_due_date,
        owner_id=db_finding.remediation_owner_id
    )
    db.add(action)
    db.flush()
    
    db_finding.linked_mitigation_action_id = action.id
    if not db_finding.linked_risk_id:
        db_finding.linked_risk_id = target_risk_id
    db.commit()
    
    return MessageResponse(message="Mitigation action created and linked to risk", id=action.id)


# =============================================================================
# Dashboard Endpoints
# =============================================================================

@router.get("/dashboard/summary", response_model=RCSADashboardSummary)
def get_dashboard_summary(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return RCSADashboardSummary(
            total_campaigns=0, active_campaigns=0, total_assessments=0,
            completed_assessments=0, pending_approval=0, overdue_assessments=0,
            completion_rate=0.0, avg_risk_score=None, avg_control_score=None,
            pending_assessments=0, open_findings=0
        )
    
    campaign_query = db.query(RCSACampaign).filter(RCSACampaign.tenant_id.in_(user_tenants))
    assessment_query = db.query(RCSAAssessment).filter(RCSAAssessment.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        campaign_query = campaign_query.filter(RCSACampaign.tenant_id == tenant_id)
        assessment_query = assessment_query.filter(RCSAAssessment.tenant_id == tenant_id)
    
    total_campaigns = campaign_query.count()
    active_campaigns = campaign_query.filter(RCSACampaign.status == "active").count()
    
    assessments = assessment_query.all()
    total_assessments = len(assessments)
    completed_assessments = sum(1 for a in assessments if a.status == "approved")
    pending_approval = sum(1 for a in assessments if a.status in ["submitted", "under_review"])
    
    overdue_count = 0
    for a in assessments:
        if a.status not in ["approved", "rejected"]:
            campaign = db.query(RCSACampaign).filter(RCSACampaign.id == a.campaign_id).first()
            if campaign and campaign.due_date and campaign.due_date < datetime.utcnow():
                overdue_count += 1
    
    completion_rate = (completed_assessments / total_assessments * 100) if total_assessments > 0 else 0.0
    
    risk_scores = [a.overall_risk_score for a in assessments if a.overall_risk_score]
    control_scores = [a.overall_control_score for a in assessments if a.overall_control_score]
    
    pending_assessments = sum(1 for a in assessments if a.status in ["not_started", "in_progress"])
    
    finding_query = db.query(RCSAFinding).filter(RCSAFinding.tenant_id.in_(user_tenants))
    if tenant_id:
        finding_query = finding_query.filter(RCSAFinding.tenant_id == tenant_id)
    open_findings = finding_query.filter(RCSAFinding.status.in_(["open", "in_progress"])).count()
    
    return RCSADashboardSummary(
        total_campaigns=total_campaigns,
        active_campaigns=active_campaigns,
        total_assessments=total_assessments,
        completed_assessments=completed_assessments,
        pending_approval=pending_approval,
        overdue_assessments=overdue_count,
        completion_rate=round(completion_rate, 1),
        avg_risk_score=round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else None,
        avg_control_score=round(sum(control_scores) / len(control_scores), 2) if control_scores else None,
        pending_assessments=pending_assessments,
        open_findings=open_findings
    )


@router.get("/dashboard/findings-by-severity", response_model=RCSAFindingsBySeverity)
def get_findings_by_severity(
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return RCSAFindingsBySeverity(critical=0, high=0, medium=0, low=0, total=0, by_type={})
    
    query = db.query(RCSAFinding).filter(RCSAFinding.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSAFinding.tenant_id == tenant_id)
    
    findings = query.all()
    
    by_severity = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    by_type = {}
    
    for f in findings:
        sev = f.severity.lower() if f.severity else "medium"
        if sev in by_severity:
            by_severity[sev] += 1
        
        ft = f.finding_type or "other"
        by_type[ft] = by_type.get(ft, 0) + 1
    
    return RCSAFindingsBySeverity(
        critical=by_severity["critical"],
        high=by_severity["high"],
        medium=by_severity["medium"],
        low=by_severity["low"],
        total=len(findings),
        by_type=by_type
    )


@router.get("/dashboard/business-unit-progress", response_model=List[RCSABUProgress])
def get_business_unit_progress(
    campaign_id: Optional[int] = None,
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSAAssessment).options(
        joinedload(RCSAAssessment.business_unit)
    ).filter(RCSAAssessment.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSAAssessment.tenant_id == tenant_id)
    if campaign_id:
        query = query.filter(RCSAAssessment.campaign_id == campaign_id)
    
    assessments = query.all()
    
    bu_stats = {}
    for a in assessments:
        bu_id = a.business_unit_id
        if bu_id not in bu_stats:
            bu_stats[bu_id] = {
                "name": a.business_unit.name if a.business_unit else f"BU {bu_id}",
                "total": 0,
                "completed": 0,
                "in_progress": 0,
                "not_started": 0,
                "risk_scores": []
            }
        
        bu_stats[bu_id]["total"] += 1
        
        if a.status == "approved":
            bu_stats[bu_id]["completed"] += 1
        elif a.status in ["in_progress", "submitted", "under_review"]:
            bu_stats[bu_id]["in_progress"] += 1
        else:
            bu_stats[bu_id]["not_started"] += 1
        
        if a.overall_risk_score:
            bu_stats[bu_id]["risk_scores"].append(a.overall_risk_score)
    
    return [
        RCSABUProgress(
            business_unit_id=bu_id,
            business_unit_name=stats["name"],
            total_assessments=stats["total"],
            completed=stats["completed"],
            in_progress=stats["in_progress"],
            not_started=stats["not_started"],
            completion_rate=round(stats["completed"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0.0,
            avg_risk_score=round(sum(stats["risk_scores"]) / len(stats["risk_scores"]), 2) if stats["risk_scores"] else None
        )
        for bu_id, stats in bu_stats.items()
    ]


@router.get("/dashboard/recent-campaigns")
def get_recent_campaigns(
    tenant_id: Optional[int] = None,
    limit: int = 5,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []
    
    query = db.query(RCSACampaign).options(
        joinedload(RCSACampaign.template),
        subqueryload(RCSACampaign.assessments)
    ).filter(RCSACampaign.tenant_id.in_(user_tenants))
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(RCSACampaign.tenant_id == tenant_id)
    
    campaigns = query.order_by(RCSACampaign.updated_at.desc()).limit(limit).all()
    
    result = []
    for c in campaigns:
        assigned_units = len(c.assessments)
        completed_units = sum(1 for a in c.assessments if a.status == "approved")
        progress = int((completed_units / assigned_units * 100) if assigned_units > 0 else 0)
        
        result.append({
            "id": c.id,
            "name": c.name,
            "template_name": c.template.name if c.template else None,
            "status": c.status,
            "period": c.period_label or c.period_type,
            "start_date": c.start_date.isoformat() if c.start_date else None,
            "end_date": c.due_date.isoformat() if c.due_date else None,
            "progress": progress,
            "assigned_units": assigned_units,
            "completed_units": completed_units,
        })
    
    return result


# RCSA Evidence Upload Endpoints
from fastapi import UploadFile, File, Form
import os
import uuid

@router.post("/evidence/upload")
async def upload_rcsa_evidence(
    file: UploadFile = File(...),
    assessment_id: int = Form(...),
    question_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Upload evidence file for an RCSA response"""
    from ....models import Evidence, RCSAResponseEvidence, RCSAResponse, RCSAAssessment
    
    if not assessment_id or not question_id:
        raise HTTPException(status_code=400, detail="assessment_id and question_id are required")
    
    # Get assessment to verify access
    assessment = db.query(RCSAAssessment).filter(RCSAAssessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    
    validate_tenant_access(current_user, assessment.tenant_id, db)
    
    # Create uploads directory if it doesn't exist
    upload_dir = os.path.join(os.getcwd(), "uploads", "rcsa_evidence")
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate unique filename
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(upload_dir, unique_filename)
    
    # Save file
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Create evidence record
    evidence = Evidence(
        tenant_id=assessment.tenant_id,
        name=file.filename,
        description=f"RCSA evidence for assessment {assessment_id}, question {question_id}",
        file_path=file_path,
        file_name=file.filename,
        file_type=file.content_type,
        uploaded_by=current_user.id,
        uploaded_at=datetime.utcnow(),
        status="approved",
        evidence_type="rcsa_evidence"
    )
    db.add(evidence)
    db.flush()
    
    # Get or create response record
    response = db.query(RCSAResponse).filter(
        RCSAResponse.assessment_id == assessment_id,
        RCSAResponse.question_id == question_id
    ).first()
    
    if not response:
        response = RCSAResponse(
            assessment_id=assessment_id,
            question_id=question_id,
            responded_by=current_user.id,
            responded_at=datetime.utcnow()
        )
        db.add(response)
        db.flush()
    
    # Link evidence to response
    evidence_link = RCSAResponseEvidence(
        response_id=response.id,
        evidence_id=evidence.id,
        uploaded_by=current_user.id
    )
    db.add(evidence_link)
    db.commit()
    
    return {
        "id": evidence.id,
        "filename": file.filename,
        "file_size": len(content),
        "uploaded_at": evidence.uploaded_at.isoformat()
    }


@router.delete("/evidence/{evidence_id}")
def delete_rcsa_evidence(
    evidence_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Delete an RCSA evidence file"""
    from ....models import Evidence, RCSAResponseEvidence
    
    evidence = db.query(Evidence).filter(Evidence.id == evidence_id).first()
    if not evidence:
        raise HTTPException(status_code=404, detail="Evidence not found")
    
    validate_tenant_access(current_user, evidence.tenant_id, db)
    
    # Delete file from disk
    if evidence.file_path and os.path.exists(evidence.file_path):
        os.remove(evidence.file_path)
    
    # Delete link records
    db.query(RCSAResponseEvidence).filter(RCSAResponseEvidence.evidence_id == evidence_id).delete()
    
    # Delete evidence record
    db.delete(evidence)
    db.commit()
    
    return {"message": "Evidence deleted successfully"}


@router.get("/evidence/response/{response_id}")
def get_response_evidence(
    response_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    """Get all evidence files for a specific response"""
    from ....models import Evidence, RCSAResponseEvidence, RCSAResponse, RCSAAssessment
    
    response = db.query(RCSAResponse).filter(RCSAResponse.id == response_id).first()
    if not response:
        raise HTTPException(status_code=404, detail="Response not found")
    
    assessment = db.query(RCSAAssessment).filter(RCSAAssessment.id == response.assessment_id).first()
    validate_tenant_access(current_user, assessment.tenant_id, db)
    
    evidence_links = db.query(RCSAResponseEvidence).filter(
        RCSAResponseEvidence.response_id == response_id
    ).all()
    
    result = []
    for link in evidence_links:
        evidence = db.query(Evidence).filter(Evidence.id == link.evidence_id).first()
        if evidence:
            file_size = 0
            if evidence.file_path and os.path.exists(evidence.file_path):
                file_size = os.path.getsize(evidence.file_path)
            result.append({
                "id": evidence.id,
                "filename": evidence.file_name,
                "file_size": file_size,
                "uploaded_at": evidence.uploaded_at.isoformat() if evidence.uploaded_at else None
            })
    
    return result
