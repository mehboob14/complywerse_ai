"""Bulk import Risk Incidents from CSV / Excel."""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ....models import (
    RiskIncident, IncidentAssetLink, IncidentVulnerabilityLink, IncidentRiskLink,
    Risk, ITAsset, Vulnerability, GRCUser, get_db,
)
from ....routers.auth_router import require_auth, get_user_primary_tenant
from ..schema_migrations import ensure_incident_schema

router = APIRouter(prefix="/incidents", tags=["ERM - Incidents Import"])

TEMPLATE_COLUMNS = [
    ("title", "Required — incident title", "Payment gateway outage"),
    ("description", "Optional detail", "API gateway returned 502 for 45 minutes"),
    ("incident_date", "Required YYYY-MM-DD", "2026-07-01"),
    ("severity", "low|medium|high|critical", "high"),
    ("status", "open|investigating|mitigating|resolved|closed", "open"),
    ("financial_impact", "Optional number", "25000"),
    ("operational_impact", "Optional text", "Checkout unavailable"),
    ("root_cause", "Optional", "Misconfigured load balancer"),
    ("corrective_actions", "Optional", "Rollback + add health checks"),
    ("tags", "Comma-separated labels", "availability,payments,p1"),
    ("assignee_email", "Match existing user email", "owner@example.com"),
    ("risk_title", "Optional — match existing risk by title", "Payment processing downtime"),
    ("asset_names", "Comma-separated asset names to link", "gw-prod-01,checkout-api"),
    ("vulnerability_titles", "Comma-separated vuln titles/CVE to link", "CVE-2024-1234"),
]


VALID_SEVERITY = {"low", "medium", "high", "critical"}
VALID_STATUS = {"open", "investigating", "mitigating", "contained", "resolved", "closed"}


def _norm(h: Any) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _cell(row: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        for rk, rv in row.items():
            if _norm(rk) == _norm(k):
                if rv is None:
                    return ""
                return str(rv).strip()
    return ""


def _parse_date(raw: str) -> Optional[datetime]:
    if not raw:
        return None
    text = raw.strip().replace("Z", "+00:00")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _split_list(raw: str) -> List[str]:
    if not raw:
        return []
    return [p.strip() for p in raw.replace(";", ",").split(",") if p.strip()]


@router.get("/template/download")
def download_incident_template(current_user: GRCUser = Depends(require_auth)):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c[0] for c in TEMPLATE_COLUMNS])
    writer.writerow([c[1] for c in TEMPLATE_COLUMNS])
    writer.writerow([c[2] for c in TEMPLATE_COLUMNS])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=incidents_import_template.csv"},
    )


@router.post("/import/upload")
async def upload_incidents_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    ensure_incident_schema(db)
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User is not assigned to any tenant")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="File must be CSV or Excel (.csv, .xlsx, .xls)")

    content = await file.read()
    rows: List[Dict[str, Any]] = []
    try:
        if filename.endswith(".csv"):
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            rows = [dict(r) for r in reader]
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers: List[str] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                first = str(row[0]).lower() if row and row[0] is not None else ""
                if idx == 1 and ("required" in first or "optional" in first):
                    continue
                if idx == 2 and first.startswith("payment"):
                    continue
                if not any(row):
                    continue
                rows.append({
                    headers[i]: row[i]
                    for i in range(min(len(headers), len(row)))
                    if headers[i]
                })
            wb.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {exc}") from exc

    cleaned = []
    for r in rows:
        title = _cell(r, "title")
        if not title or title.lower().startswith("required"):
            continue
        cleaned.append(r)
    rows = cleaned
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    email_to_id = {
        (getattr(u, "email", None) or "").strip().lower(): u.id
        for u in db.query(GRCUser).all()
        if getattr(u, "email", None)
    }
    risks = db.query(Risk).filter(Risk.tenant_id == tenant_id).all()
    risk_by_title = {(r.title or "").strip().lower(): r.id for r in risks}
    assets = db.query(ITAsset).filter(ITAsset.tenant_id == tenant_id).all()
    asset_by_name = {(a.name or "").strip().lower(): a.id for a in assets}
    vulns = db.query(Vulnerability).filter(Vulnerability.tenant_id == tenant_id).all()
    vuln_by_key: Dict[str, int] = {}
    for v in vulns:
        for key in (getattr(v, "title", None), getattr(v, "cve_id", None), getattr(v, "vuln_id", None)):
            if key:
                vuln_by_key[str(key).strip().lower()] = v.id

    imported = 0
    errors: List[str] = []

    for i, row in enumerate(rows, start=1):
        try:
            with db.begin_nested():
                title = _cell(row, "title")
                if not title:
                    raise ValueError("title is required")
                incident_date = _parse_date(_cell(row, "incident_date")) or datetime.utcnow()
                severity = (_cell(row, "severity") or "medium").lower()
                if severity not in VALID_SEVERITY:
                    raise ValueError(f"invalid severity '{severity}'")
                status = (_cell(row, "status") or "open").lower()
                if status not in VALID_STATUS:
                    status = "open"

                fin_raw = _cell(row, "financial_impact")
                financial_impact = float(fin_raw) if fin_raw else None

                assignee_email = _cell(row, "assignee_email").lower()
                assigned_to = email_to_id.get(assignee_email) if assignee_email else None

                risk_title = _cell(row, "risk_title").lower()
                risk_id = risk_by_title.get(risk_title) if risk_title else None

                tags = _split_list(_cell(row, "tags"))

                incident = RiskIncident(
                    tenant_id=tenant_id,
                    risk_id=risk_id,
                    title=title[:255],
                    description=_cell(row, "description") or None,
                    incident_date=incident_date,
                    severity=severity,
                    status=status,
                    financial_impact=financial_impact,
                    operational_impact=_cell(row, "operational_impact") or None,
                    root_cause=_cell(row, "root_cause") or None,
                    corrective_actions=_cell(row, "corrective_actions") or None,
                    reported_by=current_user.id,
                    assigned_to=assigned_to,
                    tags=tags or None,
                )
                db.add(incident)
                db.flush()

                for name in _split_list(_cell(row, "asset_names")):
                    aid = asset_by_name.get(name.lower())
                    if aid:
                        db.add(IncidentAssetLink(
                            incident_id=incident.id, asset_id=aid, created_by=current_user.id,
                        ))

                for key in _split_list(_cell(row, "vulnerability_titles")):
                    vid = vuln_by_key.get(key.lower())
                    if vid:
                        db.add(IncidentVulnerabilityLink(
                            incident_id=incident.id, vulnerability_id=vid, created_by=current_user.id,
                        ))

                if risk_id:
                    # Also mirror primary risk into the multi-risk link table for consistency.
                    exists = (
                        db.query(IncidentRiskLink.id)
                        .filter(
                            IncidentRiskLink.incident_id == incident.id,
                            IncidentRiskLink.risk_id == risk_id,
                        )
                        .first()
                    )
                    if not exists:
                        db.add(IncidentRiskLink(
                            incident_id=incident.id, risk_id=risk_id, created_by=current_user.id,
                        ))

            imported += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save imported incidents: {exc}") from exc

    return {
        "success": imported > 0,
        "imported": imported,
        "total_rows": len(rows),
        "errors": errors[:50],
        "total_errors": len(errors),
        "message": (
            f"Imported {imported} of {len(rows)} incidents"
            + (f" with {len(errors)} row error(s)" if errors else "")
        ),
    }
