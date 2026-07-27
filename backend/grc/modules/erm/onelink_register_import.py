"""1LINK RCSA register import — parse their 71-column risk-register workbook into
Risk rows, preserving every column verbatim in Risk.template_fields.

Isolated + additive: this module owns the parsing and row creation; the existing
`/risks/upload` path is deliberately left untouched. It is exposed through a
dedicated onboarding endpoint. Column locating is header-based (matches by header
text across the two-row grouped header), so it tolerates minor column shifts in
the client's real file rather than assuming fixed positions.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session

REGISTER_LABEL = "1LINK"

# ERM classification (their label) → the platform's category slug. Free-text on the
# model, so unmatched values pass through lower-cased; this only tidies the common ones.
CLASSIFICATION_MAP = {
    "operational": "operational", "operational risk": "operational",
    "financial": "financial", "credit": "financial", "liquidity": "financial",
    "market": "financial", "settlement": "financial", "financial reporting": "compliance",
    "strategic": "strategic", "compliance": "compliance", "legal": "compliance",
    "data privacy": "compliance", "information technology": "technology", "it": "technology",
    "technology": "technology", "cyber": "technology", "cybersecurity": "technology",
    "third party": "third_party", "third-party": "third_party", "outsourcing": "third_party",
    "reputational": "reputational", "fraud": "operational",
}


def _s(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _num(v: Any) -> Optional[float]:
    try:
        if v is None or _s(v) == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _label(ws, ci: int) -> str:
    """Combined header label for a column: 'Group – Sub' from the two header rows."""
    r1 = _s(ws.cell(row=1, column=ci).value)
    r2 = _s(ws.cell(row=2, column=ci).value)
    if r1 and r2 and r1 != r2:
        return f"{r1} - {r2}"
    return r2 or r1


def _labels(ws) -> Dict[int, str]:
    return {ci: _label(ws, ci) for ci in range(1, ws.max_column + 1)}


def _find(labels: Dict[int, str], *keywords: str) -> Optional[int]:
    """First column whose label contains ALL keywords (case-insensitive)."""
    kws = [k.lower() for k in keywords]
    for ci, lab in labels.items():
        low = lab.lower()
        if all(k in low for k in kws):
            return ci
    return None


def parse_onelink_workbook(wb) -> List[Dict[str, Any]]:
    ws = wb[wb.sheetnames[0]]
    labels = _labels(ws)

    c_id = _find(labels, "risk id")
    c_desc = _find(labels, "risk description")
    c_class = _find(labels, "erm risk classification") or _find(labels, "classification")
    c_sub = _find(labels, "sub-categ") or _find(labels, "sub categ")
    c_owner = _find(labels, "risk owner")
    c_dept = _find(labels, "department") and _find(labels, "department")  # 'Department' (not function)
    c_impact_rating = _find(labels, "overall impact rating")
    c_likelihood = _find(labels, "likelihood")
    c_inherent_rating = _find(labels, "inherent risk rating")
    c_residual_score = _find(labels, "residual risk score")
    c_residual_rating = _find(labels, "residual risk rating")
    c_response = _find(labels, "risk response")
    c_mit_plan = _find(labels, "mitigation plan")
    c_mit_time = _find(labels, "mitigation timeline")
    c_impl_status = _find(labels, "implementation status")

    def cell(row: int, ci: Optional[int]):
        return ws.cell(row=row, column=ci).value if ci else None

    out: List[Dict[str, Any]] = []
    for r in range(3, ws.max_row + 1):
        risk_id = _s(cell(r, c_id))
        title = _s(cell(r, c_desc))
        if not risk_id and not title:
            continue  # blank template row

        # Preserve every populated column, verbatim, keyed by its header label.
        raw: Dict[str, Any] = {}
        for ci, lab in labels.items():
            val = ws.cell(row=r, column=ci).value
            if lab and _s(val) != "":
                raw[lab] = _s(val)

        classification = _s(cell(r, c_class))
        category = CLASSIFICATION_MAP.get(classification.lower(), classification.lower() or "operational")
        il = _num(cell(r, c_likelihood))
        ii = _num(cell(r, c_impact_rating))
        ih = _num(cell(r, c_inherent_rating))
        rs = _num(cell(r, c_residual_score))

        out.append({
            "risk_id": risk_id or f"ROW-{r}",
            "title": (title or risk_id)[:255],
            "description": title,
            "category": category,
            "sub_category": _s(cell(r, c_sub))[:100] or None,
            "owner": _s(cell(r, c_owner)),
            "department": _s(cell(r, c_dept)),
            "inherent_likelihood": int(il) if il else None,
            "inherent_impact": int(ii) if ii else None,
            "inherent_score": float(ih) if ih is not None else (float(il * ii) if (il and ii) else None),
            "residual_score": float(rs) if rs is not None else None,
            "residual_rating": _s(cell(r, c_residual_rating)),
            "risk_response": _s(cell(r, c_response)),
            "mitigation_plan": _s(cell(r, c_mit_plan)),
            "mitigation_timeline": _s(cell(r, c_mit_time)),
            "implementation_status": _s(cell(r, c_impl_status)),
            "template_fields": {
                "register": REGISTER_LABEL,
                "risk_id": risk_id,
                "erm_classification": classification,
                "columns": raw,
            },
        })
    return out


# Risk-response text → mitigation action_type.
_RESPONSE_TO_ACTION = {
    "mitigate": "mitigate", "reduce": "mitigate", "reduction": "mitigate",
    "transfer": "transfer", "avoid": "avoid", "avoidance": "avoid", "accept": "accept",
}


def import_onelink_register(db: Session, tenant_id: int, file_bytes: bytes) -> Dict[str, Any]:
    """Parse an uploaded 1LINK register and create Risk rows (+ a mitigation action
    where a plan is given). Idempotent per (tenant, source_reference)."""
    import openpyxl
    from ...models import Risk, RiskMitigationAction

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"created": 0, "skipped": 0, "errors": [f"Could not read workbook: {exc}"]}

    rows = parse_onelink_workbook(wb)
    created = skipped = 0
    errors: List[str] = []

    # first user as owner fallback
    from ...models import GRCUser
    owner_row = db.query(GRCUser.id).order_by(GRCUser.id.asc()).first()
    owner_id = owner_row[0] if owner_row else None

    for row in rows:
        source_ref = f"1link-register:{row['risk_id']}"
        try:
            exists = (db.query(Risk.id)
                      .filter(Risk.tenant_id == tenant_id, Risk.source_reference == source_ref).first())
            if exists:
                skipped += 1
                continue
            risk = Risk(
                tenant_id=tenant_id, title=row["title"], description=row["description"] or row["title"],
                category=row["category"], risk_category=row["category"],
                risk_sub_category=row["sub_category"], register_type=REGISTER_LABEL,
                owner_id=owner_id, business_owner_id=owner_id,
                inherent_likelihood=row["inherent_likelihood"], inherent_impact=row["inherent_impact"],
                inherent_score=row["inherent_score"], residual_score=row["residual_score"],
                status="open", source_type="register_import", source_reference=source_ref,
                template_fields=row["template_fields"],
            )
            db.add(risk)
            db.flush()
            created += 1

            plan = row.get("mitigation_plan")
            if plan and plan.upper() != "N/A":
                action_type = _RESPONSE_TO_ACTION.get((row.get("risk_response") or "").lower(), "mitigate")
                db.add(RiskMitigationAction(
                    risk_id=risk.id, title=plan[:255], description=plan, action_type=action_type,
                    status="open", priority="medium", owner_id=owner_id))
        except Exception as exc:  # noqa: BLE001 — one bad row shouldn't abort the import
            errors.append(f"{row.get('risk_id')}: {exc}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors,
            "register": REGISTER_LABEL, "parsed_rows": len(rows)}
