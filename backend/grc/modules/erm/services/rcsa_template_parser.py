"""Parse an uploaded RCSA Excel template into a schema the platform can drive.

Banks like UBL maintain a long-standing in-house RCSA matrix in Excel with
two-row merged headers, group cells, and a domain-specific column order. We
do not try to canonicalise their schema into a fixed model — instead we
*read* the file, extract its exact column structure (groups + subcolumns +
1-based column indices), and persist that as JSON on
``RCSACustomTemplate.column_schema``. Every downstream feature
(CRUD form rendering, AI suggestion, re-export) reads that schema rather
than guessing, so the operator's existing format is preserved verbatim.

Public entry point:
    parse_rcsa_template(file_bytes, original_filename) -> ParsedTemplate

The parser is intentionally tolerant — it accepts:
- one-row OR two-row header layouts;
- merged header cells (UBL's template has 19 of them);
- arbitrary number of data rows (zero is fine — we may upload the template
  before any data has been entered);
- numeric, text, score, date, and Y/N columns (data-type detection is best-
  effort, used only as a UI hint).

No DB or LLM imports here — the result is a pure dataclass + JSON spec that
the router layer persists into ``RCSACustomTemplate.column_schema`` and
optionally a list of seed-rows into ``RCSACustomRow.data``.
"""

from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ─── Public dataclasses ──────────────────────────────────────────────────────


@dataclass
class ColumnSpec:
    """One leaf column in the template — either standalone (no group) or one
    of several sub-columns of a parent group (e.g. "Inherent Risk Assessment"
    → Impact / Likelihood / Overall / Concept).
    """
    key: str                # snake_case identifier used as the JSON payload key
    label: str              # the human header label, taken from the source cell
    col_index: int          # 1-based source column for re-export fidelity
    group: Optional[str] = None  # parent group label (e.g. "Inherent Risk Assessment")
    group_key: Optional[str] = None  # snake_case version of the group
    data_type: str = "text"  # text|number|score|yes_no|date|enum — UI hint only
    sample_values: List[str] = field(default_factory=list)  # up to 3, for UI


@dataclass
class ColumnGroup:
    """A merged top-row header that owns 1+ sub-columns underneath."""
    label: str
    key: str
    columns: List[ColumnSpec]


@dataclass
class ParsedTemplate:
    sheet_name: str
    title: Optional[str]              # banner row, e.g. "UBL - Risk & Control …"
    header_row_count: int             # 1 or 2
    data_start_row: int               # first row containing actual data
    groups: List[ColumnGroup]
    flat_columns: List[ColumnSpec]    # all columns in source order
    data_rows: List[Dict[str, Any]]   # parsed seed rows keyed by ColumnSpec.key
    file_sha256: str
    warnings: List[str] = field(default_factory=list)

    def to_schema_json(self) -> Dict[str, Any]:
        """Schema persisted to RCSACustomTemplate.column_schema."""
        return {
            "title": self.title,
            "sheet_name": self.sheet_name,
            "header_row_count": self.header_row_count,
            "data_start_row": self.data_start_row,
            "groups": [
                {
                    "label": g.label,
                    "key": g.key,
                    "columns": [_column_to_dict(c) for c in g.columns],
                }
                for g in self.groups
            ],
            "flat_columns": [_column_to_dict(c) for c in self.flat_columns],
            "warnings": list(self.warnings),
        }


def _column_to_dict(c: ColumnSpec) -> Dict[str, Any]:
    return {
        "key": c.key,
        "label": c.label,
        "col_index": c.col_index,
        "group": c.group,
        "group_key": c.group_key,
        "data_type": c.data_type,
        "sample_values": list(c.sample_values),
    }


# ─── Header detection ────────────────────────────────────────────────────────


_BANNER_ROW_MAX_NON_EMPTY = 1  # a banner row only has one cell populated (col A)
_KEY_NORMALISE_RE = re.compile(r"[^a-z0-9]+")


def _slugify(label: str, *, taken: Optional[Dict[str, int]] = None) -> str:
    """Stable snake_case identifier for the column. Disambiguates collisions
    by appending _2/_3/… so two columns named "Impact/Severity" inside
    different groups produce distinct top-level keys."""
    s = (label or "").strip().lower()
    s = _KEY_NORMALISE_RE.sub("_", s).strip("_")
    if not s:
        s = "col"
    if taken is None:
        return s
    n = taken.get(s, 0) + 1
    taken[s] = n
    return s if n == 1 else f"{s}_{n}"


def _row_non_empty_count(ws: Worksheet, row: int, max_col: int) -> int:
    return sum(
        1 for c in range(1, max_col + 1) if ws.cell(row=row, column=c).value not in (None, "")
    )


def _merged_value(ws: Worksheet, row: int, col: int) -> Any:
    """Return the value of the cell, traversing merged ranges so a cell
    inside a merged region returns the anchor cell's value (openpyxl returns
    None for non-anchor cells of a merged range otherwise)."""
    v = ws.cell(row=row, column=col).value
    if v not in (None, ""):
        return v
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def _detect_header_layout(ws: Worksheet) -> Tuple[Optional[str], int, int, int]:
    """Return (banner_title, header_row_start, header_row_count, data_start).

    Layouts supported:
      * 1-row header: data starts row 2 (or row 3 if row 1 is a banner)
      * 2-row header: row N has merged group headers, row N+1 has sub-labels.
    Detected by checking how many cells of row N are populated vs N+1.
    A row that's a banner has exactly one populated cell — typically col A
    spanning the whole sheet width.
    """
    max_col = ws.max_column or 1
    max_row = min(ws.max_row or 1, 10)

    banner_title: Optional[str] = None
    cursor = 1

    # Detect optional banner row(s).
    while cursor <= max_row:
        nonempty = _row_non_empty_count(ws, cursor, max_col)
        if nonempty <= _BANNER_ROW_MAX_NON_EMPTY:
            v = ws.cell(row=cursor, column=1).value
            if isinstance(v, str) and v.strip():
                banner_title = v.strip()
            cursor += 1
        else:
            break

    header_row_start = cursor
    if header_row_start > max_row:
        # Nothing looked like a header — assume row 1 is it.
        return banner_title, 1, 1, 2

    # Is the next row also a header? It is if a meaningful fraction of cells
    # in that row are non-empty AND it is NOT the data row (heuristic: a data
    # row's first column tends to be a string used as a category not a header).
    next_row = header_row_start + 1
    if next_row <= max_row and _row_non_empty_count(ws, next_row, max_col) >= 3:
        # Treat 2-row header if row+1 has different distinct headers than row
        # (e.g. row has merged group "Inherent Risk Assessment" once, row+1
        # has Impact / Likelihood / Overall / Concept across 4 cells).
        row_distinct = {
            _merged_value(ws, header_row_start, c) for c in range(1, max_col + 1)
        }
        next_distinct = {
            ws.cell(row=next_row, column=c).value for c in range(1, max_col + 1)
        }
        # If next-row cell labels are mostly NOT in row distinct (i.e. they
        # introduce new vocabulary), treat as a 2-row header.
        overlap = sum(1 for v in next_distinct if v in row_distinct and v not in (None, ""))
        if overlap < max(1, len(next_distinct) // 3):
            return banner_title, header_row_start, 2, next_row + 1

    return banner_title, header_row_start, 1, header_row_start + 1


# ─── Column extraction ──────────────────────────────────────────────────────


def _detect_data_type(samples: List[Any]) -> str:
    """Best-effort: text | number | score | yes_no | date | enum."""
    if not samples:
        return "text"
    cleaned = [s for s in samples if s not in (None, "")]
    if not cleaned:
        return "text"
    # Y/N detector — every value is one of yes/no/y/n
    yn = {"yes", "no", "y", "n", "true", "false"}
    if all(isinstance(s, str) and s.strip().lower() in yn for s in cleaned):
        return "yes_no"
    # Date detector
    if all(isinstance(s, datetime) for s in cleaned):
        return "date"
    # Number detector — every value parses to a number
    numeric_count = 0
    for s in cleaned:
        if isinstance(s, (int, float)):
            numeric_count += 1
            continue
        try:
            float(str(s))
            numeric_count += 1
        except (TypeError, ValueError):
            continue
    if numeric_count == len(cleaned):
        # Score columns are short small integers (0..100 typically).
        try:
            vals = [float(s) if not isinstance(s, (int, float)) else s for s in cleaned]
            if all(0 <= v <= 100 and float(int(v)) == float(v) for v in vals):
                return "score"
        except Exception:  # noqa: BLE001
            pass
        return "number"
    # Enum detector — small, repeated label set
    distinct = {str(s).strip() for s in cleaned}
    if len(distinct) <= max(3, len(cleaned) // 2):
        return "enum"
    return "text"


def _extract_columns(
    ws: Worksheet,
    *,
    header_row: int,
    header_row_count: int,
    data_start: int,
) -> Tuple[List[ColumnGroup], List[ColumnSpec], List[Dict[str, Any]]]:
    """Walk every populated column index. For each leaf column produce a
    ColumnSpec. Group columns by the row-above label (None for ungrouped).
    Then build flat data rows keyed by each column's snake_case key.
    """
    max_col = ws.max_column or 1
    last_row = ws.max_row or data_start

    taken: Dict[str, int] = {}
    flat: List[ColumnSpec] = []

    # Walk leaf columns. If 2-row header, leaf labels live on header_row+1;
    # group labels live on header_row.
    for col in range(1, max_col + 1):
        if header_row_count == 2:
            leaf_label_cell = ws.cell(row=header_row + 1, column=col).value
            leaf_label = (
                str(leaf_label_cell).strip()
                if leaf_label_cell not in (None, "")
                else None
            )
            group_label_raw = _merged_value(ws, header_row, col)
            group_label = (
                str(group_label_raw).strip()
                if group_label_raw not in (None, "")
                else None
            )
            # If a column has no sub-header, the group label IS the leaf
            # label (e.g. a column that spans both header rows merged).
            if not leaf_label and group_label:
                leaf_label = group_label
                group_label = None
        else:
            leaf_label = ws.cell(row=header_row, column=col).value
            leaf_label = str(leaf_label).strip() if leaf_label not in (None, "") else None
            group_label = None

        if not leaf_label:
            continue  # truly empty column — skip
        # Defensive: don't repeat the group label as its own column's group.
        if group_label and group_label == leaf_label:
            group_label = None

        # Collect samples for data-type detection (capped).
        samples: List[Any] = []
        for r in range(data_start, min(last_row, data_start + 20) + 1):
            v = ws.cell(row=r, column=col).value
            if v not in (None, ""):
                samples.append(v)

        # Build a slug scoped under the group so two identical labels in
        # different groups don't collide.
        slug_seed = (
            f"{_slugify(group_label)}_{_slugify(leaf_label)}"
            if group_label
            else _slugify(leaf_label)
        )
        col_key = _slugify(slug_seed, taken=taken)
        group_key = _slugify(group_label) if group_label else None

        col_spec = ColumnSpec(
            key=col_key,
            label=leaf_label,
            col_index=col,
            group=group_label,
            group_key=group_key,
            data_type=_detect_data_type(samples),
            sample_values=[str(s)[:80] for s in samples[:3]],
        )
        flat.append(col_spec)

    # Group columns by group_label preserving source order.
    groups: List[ColumnGroup] = []
    seen_group_keys: Dict[str, ColumnGroup] = {}
    for c in flat:
        gk = c.group_key or f"__ungrouped__{c.key}"
        gl = c.group or c.label
        if gk in seen_group_keys:
            seen_group_keys[gk].columns.append(c)
        else:
            group = ColumnGroup(label=gl, key=gk, columns=[c])
            seen_group_keys[gk] = group
            groups.append(group)

    # Now extract data rows.
    data_rows: List[Dict[str, Any]] = []
    for r in range(data_start, last_row + 1):
        # Skip blank rows.
        if not any(
            ws.cell(row=r, column=c.col_index).value not in (None, "") for c in flat
        ):
            continue
        row_payload: Dict[str, Any] = {}
        for c in flat:
            v = ws.cell(row=r, column=c.col_index).value
            if isinstance(v, datetime):
                row_payload[c.key] = v.isoformat()
            else:
                row_payload[c.key] = v
        data_rows.append(row_payload)

    return groups, flat, data_rows


# ─── Public entry point ─────────────────────────────────────────────────────


def parse_rcsa_template(
    file_bytes: bytes,
    *,
    original_filename: str,
    sheet_name: Optional[str] = None,
) -> ParsedTemplate:
    """Parse the uploaded .xlsx into a ParsedTemplate.

    Raises ValueError on a fundamentally unusable file (no workbook, no
    sheets, header detection produced zero columns). Callers in the router
    layer turn that into a 422 with the error text so the operator knows
    what's wrong before the row import path runs.
    """
    if not file_bytes:
        raise ValueError("Uploaded file is empty")
    sha = hashlib.sha256(file_bytes).hexdigest()

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Could not open as Excel workbook: {exc}") from exc

    if not wb.sheetnames:
        raise ValueError("Workbook has no sheets")

    target_sheet = sheet_name or wb.sheetnames[0]
    if target_sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{target_sheet}' not found; sheets present: {wb.sheetnames}")
    ws = wb[target_sheet]

    warnings: List[str] = []
    banner, header_start, header_count, data_start = _detect_header_layout(ws)
    if header_start <= 0:
        raise ValueError("Could not detect a header row in the worksheet")

    groups, flat, rows = _extract_columns(
        ws,
        header_row=header_start,
        header_row_count=header_count,
        data_start=data_start,
    )

    if not flat:
        raise ValueError(
            "Header detection found no columns. The first non-banner row must "
            "contain column labels."
        )

    if len(flat) < 5:
        warnings.append(
            f"Only {len(flat)} columns detected. A useful RCSA template "
            "typically has 15+. Confirm the upload is the correct file."
        )

    return ParsedTemplate(
        sheet_name=target_sheet,
        title=banner,
        header_row_count=header_count,
        data_start_row=data_start,
        groups=groups,
        flat_columns=flat,
        data_rows=rows,
        file_sha256=sha,
        warnings=warnings,
    )


# ─── Round-trip: export rows back to Excel preserving original layout ───────


def export_rows_to_excel(
    *,
    template_bytes: bytes,
    schema: Dict[str, Any],
    rows_data: List[Dict[str, Any]],
) -> bytes:
    """Re-emit the rows as an Excel file using the original layout.

    Strategy: open the operator's original workbook (which already carries
    their formatting, merged headers, banner, fonts, column widths), wipe
    everything from ``data_start_row`` downward, then write each row by
    placing each cell at the schema's recorded 1-based column index.
    This guarantees the exported file is visually identical to the upload
    they recognise, just populated with our system's current state.
    """
    if not template_bytes:
        raise ValueError("Original template bytes are required for export")
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
    sheet_name = schema.get("sheet_name") or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in original template")
    ws = wb[sheet_name]

    data_start = int(schema.get("data_start_row") or 2)
    flat_columns = schema.get("flat_columns") or []
    last_col = max((c.get("col_index") or 0) for c in flat_columns) if flat_columns else 0

    # Clear any pre-existing data rows.
    if ws.max_row >= data_start:
        for r in range(data_start, ws.max_row + 1):
            for c in range(1, (last_col or ws.max_column) + 1):
                ws.cell(row=r, column=c).value = None

    # Write rows.
    for offset, row in enumerate(rows_data):
        target_row = data_start + offset
        for col in flat_columns:
            key = col.get("key")
            col_index = col.get("col_index")
            if not key or not col_index:
                continue
            v = row.get(key)
            if v is None:
                continue
            ws.cell(row=target_row, column=col_index).value = v

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


__all__ = [
    "ColumnSpec",
    "ColumnGroup",
    "ParsedTemplate",
    "parse_rcsa_template",
    "export_rows_to_excel",
]
