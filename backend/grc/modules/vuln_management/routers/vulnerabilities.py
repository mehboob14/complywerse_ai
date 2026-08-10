from typing import List, Optional
import io
import csv
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from ....models import (
    Vulnerability, VulnerabilityReport, VulnerabilitySLAConfig,
    VulnerabilityAssetLink, VulnerabilityDependency, GRCUser, get_db
)
from ....schemas import (
    VulnerabilityCreate, VulnerabilityUpdate, VulnerabilityResponse,
    VulnerabilityAssign, VulnerabilityStatusChange, MessageResponse
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant, require_tenant_permission

router = APIRouter(tags=["Vulnerabilities"])


def validate_tenant_access(user: GRCUser, tenant_id: int, db: Session) -> None:
    user_tenants = get_user_tenants(user, db)
    if tenant_id not in user_tenants:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied to this tenant's data"
        )


def get_vuln_or_404(vuln_id: int, user_tenants: List[int], db: Session) -> Vulnerability:
    vuln = db.query(Vulnerability).filter(
        Vulnerability.id == vuln_id,
        Vulnerability.tenant_id.in_(user_tenants)
    ).first()
    if not vuln:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Vulnerability not found"
        )
    return vuln


def get_sla_days(tenant_id: int, severity: str, db: Session) -> int:
    sla = db.query(VulnerabilitySLAConfig).filter(
        VulnerabilitySLAConfig.tenant_id == tenant_id,
        VulnerabilitySLAConfig.severity == severity,
        VulnerabilitySLAConfig.is_active == True
    ).first()
    if sla:
        return sla.remediation_days
    defaults = {"critical": 7, "high": 30, "medium": 90, "low": 180, "info": 365}
    return defaults.get(severity, 90)


def generate_vuln_id(tenant_id: int, db: Session) -> str:
    count = db.query(Vulnerability).filter(Vulnerability.tenant_id == tenant_id).count()
    return f"VULN-{count + 1:05d}"


def _build_vulnerability_response(v: Vulnerability, solution_count=None) -> VulnerabilityResponse:
    linked_assets = []
    if getattr(v, "asset_links", None):
        linked_assets = [
            link.asset.name
            for link in v.asset_links
            if getattr(link, "asset", None) and getattr(link.asset, "name", None)
        ]

    return VulnerabilityResponse(
        id=v.id,
        tenant_id=v.tenant_id,
        report_id=v.report_id,
        vuln_id=v.vuln_id,
        title=v.title,
        description=v.description,
        severity=v.severity,
        cvss_score=v.cvss_score,
        cvss_vector=v.cvss_vector,
        cve_id=v.cve_id,
        cwe_id=v.cwe_id,
        cwe_ids=getattr(v, "cwe_ids", None),
        nvd_cvss_score=getattr(v, "nvd_cvss_score", None),
        nvd_cvss_vector=getattr(v, "nvd_cvss_vector", None),
        vpr_score=getattr(v, "vpr_score", None),
        cpe=getattr(v, "cpe", None),
        affected_component=v.affected_component,
        affected_host=v.affected_host,
        affected_port=v.affected_port,
        affected_url=v.affected_url,
        plugin_family=getattr(v, "plugin_family", None),
        evidence=v.evidence,
        reproduction_steps=v.reproduction_steps,
        recommendation=v.recommendation,
        ai_recommendation=v.ai_recommendation,
        ai_impact_assessment=v.ai_impact_assessment,
        status=v.status,
        resolution_notes=v.resolution_notes,
        discovered_at=v.discovered_at,
        due_date=v.due_date,
        resolved_at=v.resolved_at,
        assigned_to=v.assigned_to,
        verified_by=v.verified_by,
        verified_at=v.verified_at,
        is_exception=v.is_exception,
        exception_reason=v.exception_reason,
        exception_approved_by=v.exception_approved_by,
        exception_expiry=v.exception_expiry,
        created_at=v.created_at,
        updated_at=v.updated_at,
        assignee_name=v.assignee.display_name if v.assignee else None,
        verifier_name=v.verifier.display_name if v.verifier else None,
        linked_assets=linked_assets,
        template_type=getattr(v, "template_type", None),
        template_fields=getattr(v, "template_fields", None) or None,
        # Threat-intelligence enrichment — all None on un-enriched rows. The
        # frontend hides each chip/badge when the value is None so rows without
        # a CVE-ID render exactly as before.
        epss_score=getattr(v, "epss_score", None),
        epss_percentile=getattr(v, "epss_percentile", None),
        kev_flag=getattr(v, "kev_flag", None),
        kev_date_added=getattr(v, "kev_date_added", None),
        nvd_published_at=getattr(v, "nvd_published_at", None),
        nvd_last_modified_at=getattr(v, "nvd_last_modified_at", None),
        nvd_last_synced_at=getattr(v, "nvd_last_synced_at", None),
        exploit_references=list(getattr(v, "exploit_references", None) or []) or None,
        composite_priority=getattr(v, "composite_priority", None),
        # These three are declared on the schema but were never populated
        # here, so they came back null on every response no matter what was
        # stored. That made the UI blind to exploit maturity and attack
        # vector — two of the seven scoring signals — and its score
        # permanently disagreed with the backend's.
        public_exploit_count=getattr(v, "public_exploit_count", None),
        public_exploit_refs=list(getattr(v, "public_exploit_refs", None) or []) or None,
        # Exploit-DB corroboration (the second public-exploit source) — both columns,
        # so cheap to expose on every row; the frontend ORs them for "exploit exists".
        exploitdb_count=getattr(v, "exploitdb_count", None),
        exploitdb_verified_count=getattr(v, "exploitdb_verified_count", None),
        # Populated only by the detail endpoint (a query); None on list rows.
        solution_count=solution_count,
        # Phase 6 — vendor patch intelligence.
        patch_references=list(getattr(v, "patch_references", None) or []) or None,
        vendor_advisory_ids=list(getattr(v, "vendor_advisory_ids", None) or []) or None,
        remediation_guidance=getattr(v, "remediation_guidance", None),
        psirt_synced_at=getattr(v, "psirt_synced_at", None),
        psirt_source=getattr(v, "psirt_source", None),
        # Phase 8 — exception workflow.
        exception_status=getattr(v, "exception_status", None) or None,
        exception_requested_by_id=getattr(v, "exception_requested_by_id", None),
        exception_requested_at=getattr(v, "exception_requested_at", None),
        exception_justification=getattr(v, "exception_justification", None),
        exception_compensating_controls=list(
            getattr(v, "exception_compensating_controls", None) or []
        ) or None,
        exception_approved_at=getattr(v, "exception_approved_at", None),
        exception_expires_at=getattr(v, "exception_expires_at", None),
        exception_denial_reason=getattr(v, "exception_denial_reason", None),
        exception_revoked_by_id=getattr(v, "exception_revoked_by_id", None),
        exception_revoked_at=getattr(v, "exception_revoked_at", None),
        exception_revocation_reason=getattr(v, "exception_revocation_reason", None),
        exception_metadata=dict(getattr(v, "exception_metadata", None) or {}) or None,
        # Scanner closure loop — provenance + verified-close evidence.
        connection_id=getattr(v, "connection_id", None),
        source=getattr(v, "source", None),
        external_vuln_id=getattr(v, "external_vuln_id", None),
        scanner_status=getattr(v, "scanner_status", None),
        first_detected=getattr(v, "first_detected", None),
        last_seen=getattr(v, "last_seen", None),
        last_seen_scan_id=getattr(v, "last_seen_scan_id", None),
        closed_at=getattr(v, "closed_at", None),
        closed_by=getattr(v, "closed_by", None),
        closure_evidence=dict(getattr(v, "closure_evidence", None) or {}) or None,
        reopened_at=getattr(v, "reopened_at", None),
        reopen_count=getattr(v, "reopen_count", None),
    )


# Statuses that should be hidden from the default vulnerability register list.
# Mirrors `RESOLVED_STATUSES` in the dashboard router so the two views stay
# consistent on what counts as "closed/mitigated".
# Terminal/closed statuses — hidden from the default list view. The new
# `auto_closed_decommissioned` value is set by the asset-lifecycle auto-close
# hook (asset_lifecycle._auto_close_linked_vulns) when a linked asset enters
# decommissioned/retired; treated identically to other terminal states for
# filtering, but reports can distinguish it via the literal value.
_LIST_CLOSED_STATUSES = [
    "resolved", "remediated", "verified", "closed",
    "accepted", "false_positive", "auto_closed_decommissioned",
    # Set by the scanner closure engine when a completed re-scan covering the
    # host no longer reports the finding (evidence in closure_evidence).
    "auto_closed_fixed",
]


@router.get("/vulnerabilities", response_model=List[VulnerabilityResponse])
def list_vulnerabilities(
    tenant_id: Optional[int] = None,
    report_id: Optional[int] = None,
    severity: Optional[str] = None,
    status_filter: Optional[str] = None,
    assigned_to: Optional[int] = None,
    cve_id: Optional[str] = None,
    # Domain filter — the scanner family a finding belongs to. Used by the
    # grouped register view to load one domain's findings on expand.
    plugin_family: Optional[str] = None,
    is_exception: Optional[bool] = None,
    is_overdue: Optional[bool] = None,
    # Filter by template source. Used by the NCA register to surface only the
    # vulnerabilities ingested from the NCA template (template_type="NCA Template").
    # Special sentinel "_general" returns vulns with template_type IS NULL.
    template_type: Optional[str] = None,
    # New: closed/mitigated vulnerabilities are hidden by default to keep the
    # operational register focused on what still needs work. Two toggles:
    #   include_closed=true → include closed/mitigated alongside open ones
    #   closed_only=true    → show ONLY closed/mitigated (the "show closed" view)
    # An explicit `status_filter=<value>` always wins over these toggles so
    # existing callers that pin a specific status (and any saved bookmarks)
    # behave exactly as before.
    include_closed: bool = False,
    closed_only: bool = False,
    search: Optional[str] = None,
    # Public-exploit filter (GitHub PoC OR Exploit-DB). Distinct from KEV.
    # True → only findings with a public exploit; False → only those without.
    has_exploit: Optional[bool] = None,
    # ATT&CK tactic richness filter — True keeps findings whose mapped chain
    # spans ≥ HIGH_TACTICS_MIN distinct tactics (see attack.selection).
    high_tactics: Optional[bool] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:view"))
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return []

    # Lapse any risk acceptances that are past their review date before we read.
    #
    # The proper home for this is the nightly Celery task in grc/tasks/exceptions.py,
    # but no worker runs in this deployment, which meant acceptances never actually
    # expired — a finding parked as "accepted" stayed hidden forever. Doing it here
    # is a bounded, indexed query (exception_status + exception_expires_at are both
    # indexed) scoped to the caller's tenants, and it runs once per register load
    # rather than once per record. When a worker does exist this stays correct;
    # it just finds nothing to do.
    try:
        from ....services.vuln_exception import expire_due_exceptions
        expire_due_exceptions(db, tenant_ids=user_tenants)
    except Exception:  # noqa: BLE001 — never let the sweep break the register
        import logging
        logging.getLogger(__name__).exception("exception expiry sweep failed")
        db.rollback()

    # Same rationale as the sweep above (no Celery worker in this deployment):
    # composite_priority (the register's Contextual score) is otherwise written only
    # when a finding is individually opened / edited / enriched, so a register nobody
    # has clicked through shows "not scored" for every row despite every input
    # (CVSS/EPSS/KEV/…) already being present. Score the NULL rows once, in place, from
    # stored fields only — no external calls. Bounded + best-effort; converges to a
    # no-op after the first load.
    try:
        from ..enrichment.enrichment_service import backfill_composite_priorities
        backfill_composite_priorities(db, user_tenants)
    except Exception:  # noqa: BLE001 — never let the backfill break the register
        import logging
        logging.getLogger(__name__).exception("composite priority backfill failed")
        db.rollback()

    query = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.tenant_id.in_(user_tenants))

    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        query = query.filter(Vulnerability.tenant_id == tenant_id)
    if report_id:
        query = query.filter(Vulnerability.report_id == report_id)
    if severity:
        query = query.filter(Vulnerability.severity == severity)
    if status_filter:
        query = query.filter(Vulnerability.status == status_filter)
    else:
        # Apply the closed/open toggle only when no explicit status was given.
        if closed_only:
            query = query.filter(Vulnerability.status.in_(_LIST_CLOSED_STATUSES))
        elif not include_closed:
            query = query.filter(Vulnerability.status.notin_(_LIST_CLOSED_STATUSES))
    if assigned_to:
        query = query.filter(Vulnerability.assigned_to == assigned_to)
    if cve_id:
        query = query.filter(Vulnerability.cve_id.ilike(f"%{cve_id}%"))
    if plugin_family:
        # "Uncategorized" is the UI label for findings with no scanner family.
        if plugin_family == "Uncategorized":
            query = query.filter(Vulnerability.plugin_family.is_(None))
        else:
            query = query.filter(Vulnerability.plugin_family == plugin_family)
    if template_type:
        if template_type == "_general":
            query = query.filter(Vulnerability.template_type.is_(None))
        else:
            query = query.filter(Vulnerability.template_type == template_type)
    if is_exception is not None:
        query = query.filter(Vulnerability.is_exception == is_exception)
    if is_overdue:
        query = query.filter(
            Vulnerability.due_date < datetime.utcnow(),
            Vulnerability.status.notin_(_LIST_CLOSED_STATUSES)
        )
    if search:
        query = query.filter(
            (Vulnerability.title.ilike(f"%{search}%")) |
            (Vulnerability.vuln_id.ilike(f"%{search}%")) |
            (Vulnerability.cve_id.ilike(f"%{search}%"))
        )
    if has_exploit is not None:
        from sqlalchemy import or_, func
        exploit_expr = or_(
            func.coalesce(Vulnerability.public_exploit_count, 0) > 0,
            func.coalesce(Vulnerability.exploitdb_count, 0) > 0,
        )
        query = query.filter(exploit_expr if has_exploit else ~exploit_expr)

    # High-tactics is mapping-derived (CWE→CAPEC→ATT&CK), not a SQL column.
    # Evaluate after the cheap filters, before pagination — register sizes are
    # hundreds, not tens of thousands, and select_techniques is pure/in-memory.
    if high_tactics is not None:
        from ..attack.selection import is_high_tactics
        candidates = query.order_by(Vulnerability.created_at.desc()).all()
        matched = [
            v for v in candidates
            if is_high_tactics(v.cwe_id, v.cvss_vector) is high_tactics
        ]
        page = matched[skip: skip + limit]
        return [_build_vulnerability_response(v) for v in page]

    vulns = query.order_by(Vulnerability.created_at.desc()).offset(skip).limit(limit).all()

    return [_build_vulnerability_response(v) for v in vulns]


@router.get("/vulnerabilities/domains")
def list_vulnerability_domains(
    tenant_id: Optional[int] = None,
    template_type: Optional[str] = None,
    include_closed: bool = False,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:view")),
):
    """Group the register into runtime-derived domains (the scanner's own plugin
    family). Computed server-side over the WHOLE register so the counts match
    the KPI cards — a client tally would only see the loaded page (the list caps
    at `limit`). Domains are ordered worst-severity-first so buckets holding
    criticals surface at the top even when the family is a catch-all like "Misc.".
    """
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"domains": [], "total": 0}

    q = db.query(
        Vulnerability.plugin_family,
        Vulnerability.severity,
        func.count(Vulnerability.id),
    ).filter(Vulnerability.tenant_id.in_(user_tenants))

    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
        q = q.filter(Vulnerability.tenant_id == tenant_id)
    if template_type:
        if template_type == "_general":
            q = q.filter(Vulnerability.template_type.is_(None))
        else:
            q = q.filter(Vulnerability.template_type == template_type)
    if not include_closed:
        q = q.filter(Vulnerability.status.notin_(_LIST_CLOSED_STATUSES))

    q = q.group_by(Vulnerability.plugin_family, Vulnerability.severity)

    SEV_KEYS = ["critical", "high", "medium", "low", "info"]
    SEV_RANK = {"critical": 4, "high": 3, "medium": 2, "low": 1, "info": 0}
    fam_map: dict = {}
    for family, severity, cnt in q.all():
        key = family or "Uncategorized"
        d = fam_map.setdefault(key, {s: 0 for s in SEV_KEYS})
        sev = (severity or "info").lower()
        d[sev] = d.get(sev, 0) + int(cnt)

    domains = []
    for family, d in fam_map.items():
        total = sum(d[s] for s in SEV_KEYS)
        worst = next((s for s in SEV_KEYS if d[s] > 0), "info")
        domains.append({
            "family": family,
            "total": total,
            "by_severity": d,
            "worst_severity": worst,
        })
    # worst severity first, then biggest bucket — so "Misc." (criticals) and
    # "Databases" float up and the 120-row info "Windows" pile sinks.
    domains.sort(key=lambda x: (SEV_RANK[x["worst_severity"]], x["total"]), reverse=True)
    return {"domains": domains, "total": sum(x["total"] for x in domains)}


@router.post("/vulnerabilities", response_model=VulnerabilityResponse, status_code=status.HTTP_201_CREATED)
def create_vulnerability(
    request: VulnerabilityCreate,
    tenant_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:create"))
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="User not associated with any tenant")
    
    if tenant_id:
        validate_tenant_access(current_user, tenant_id, db)
    else:
        tenant_id = get_user_primary_tenant(current_user, db)
        if not tenant_id:
            tenant_id = user_tenants[0]
    
    vuln_id = request.vuln_id or generate_vuln_id(tenant_id, db)
    
    sla_days = get_sla_days(tenant_id, request.severity, db)
    due_date = request.due_date or (datetime.utcnow() + timedelta(days=sla_days))
    
    vuln = Vulnerability(
        tenant_id=tenant_id,
        report_id=request.report_id,
        vuln_id=vuln_id,
        title=request.title,
        description=request.description,
        severity=request.severity,
        cvss_score=request.cvss_score,
        cvss_vector=request.cvss_vector,
        cve_id=request.cve_id,
        cwe_id=request.cwe_id,
        affected_component=request.affected_component,
        affected_host=request.affected_host,
        affected_port=request.affected_port,
        affected_url=request.affected_url,
        evidence=request.evidence,
        reproduction_steps=request.reproduction_steps,
        recommendation=request.recommendation,
        status="open",
        discovered_at=request.discovered_at or datetime.utcnow(),
        due_date=due_date
    )
    db.add(vuln)
    db.commit()
    db.refresh(vuln)
    
    db.refresh(vuln)
    return _build_vulnerability_response(vuln)


@router.post("/vulnerabilities/bulk-upload")
async def bulk_upload_vulnerabilities(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:create"))
):
    """Bulk import vulnerabilities from CSV or Excel file."""
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        user_tenants = get_user_tenants(current_user, db)
        if not user_tenants:
            raise HTTPException(status_code=403, detail="User not associated with any tenant")
        tenant_id = user_tenants[0]

    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail=f"Unsupported file type '.{ext}'. Please upload a CSV or Excel file.")

    contents = await file.read()

    rows: list[dict] = []
    try:
        if ext == "csv":
            decoded = contents.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = [dict(r) for r in reader]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

            # NCA-style workbooks ship with a Cover Page first; pick the sheet
            # whose name suggests it is the data register, fall back to active.
            preferred = None
            for s in wb.sheetnames:
                sl = s.lower()
                if ("vulnerability" in sl or "register" in sl) and "legend" not in sl and "cover" not in sl:
                    preferred = s
                    break
            ws = wb[preferred] if preferred else wb.active

            # Scan first 20 rows for the actual header row — NCA template puts
            # headers at row 11, generic CSV-style sheets at row 1.
            all_rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
            header_row_idx = 0
            header_keywords = ("title", "vulnerability id", "cve")
            for r_idx, raw_row in enumerate(all_rows):
                joined = " ".join(str(c).lower() for c in raw_row if c is not None)
                if any(k in joined for k in header_keywords):
                    header_row_idx = r_idx
                    break
            headers = [str(c).strip() if c is not None else "" for c in all_rows[header_row_idx]]

            # Iterate the remaining rows after the header
            for raw_row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                row_dict: dict = {}
                for i, v in enumerate(raw_row):
                    if i >= len(headers):
                        continue
                    key = headers[i]
                    if not key:
                        continue
                    row_dict[key] = (str(v).strip() if v is not None else "")
                if row_dict:
                    rows.append(row_dict)
            wb.close()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {str(e)}. Ensure it matches the template format.")

    if not rows:
        raise HTTPException(status_code=422, detail="The file is empty or contains no data rows.")

    VALID_SEVERITIES = {"critical", "high", "medium", "low", "info"}
    VALID_STATUSES = {
        "open", "in_progress", "remediated", "verified", "closed",
        "accepted", "false_positive", "auto_closed_decommissioned",
    }

    created = 0
    skipped = 0  # rows with no usable title (usually trailing blank spreadsheet rows)
    errors: list[str] = []

    # Case-insensitive, whitespace-tolerant column lookup so templates with
    # `Title` / `TITLE` / ` title ` headers all work. Tolerates extra template
    # columns by matching the first key whose normalized form starts with the
    # requested name.
    def _lookup(row: dict, *aliases: str):
        if not row:
            return None
        norm_keys = {k: str(k).lower().strip() for k in row.keys()}
        for alias in aliases:
            target = alias.lower().strip()
            for orig_key, norm in norm_keys.items():
                if norm == target or norm.startswith(target):
                    val = row.get(orig_key)
                    if val is not None and str(val).strip():
                        return str(val).strip()
        return None

    # Calculate base count once so each row in this batch gets a unique ID
    existing_count = db.query(Vulnerability).filter(Vulnerability.tenant_id == tenant_id).count()
    id_counter = existing_count

    for idx, row in enumerate(rows, start=2):
        title = _lookup(row, "title", "vulnerability title", "name", "summary") or ""
        if not title:
            skipped += 1
            continue

        severity = (_lookup(row, "severity", "risk severity", "risk level") or "medium").lower()
        if severity not in VALID_SEVERITIES:
            errors.append(f"Row {idx}: invalid severity '{severity}' — using 'medium'")
            severity = "medium"

        vul_status = (_lookup(row, "status", "remediation status") or "open").lower().replace(" ", "_")
        if vul_status not in VALID_STATUSES:
            vul_status = "open"

        cvss_raw = _lookup(row, "cvss_score", "cve score", "cvss") or ""
        # Templates may store the score as 'CVSS:3.0 7.5' — extract the trailing
        # decimal so the value still comes through.
        import re as _re
        score_matches = _re.findall(r"\d+(?:\.\d+)?", cvss_raw)
        try:
            cvss_score = float(score_matches[-1]) if score_matches else None
        except ValueError:
            cvss_score = None

        due_date_raw = _lookup(row, "due_date", "due date") or ""
        due_date = None
        if due_date_raw:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    due_date = datetime.strptime(due_date_raw, fmt)
                    break
                except ValueError:
                    continue
        if not due_date:
            sla_days = get_sla_days(tenant_id, severity, db)
            due_date = datetime.utcnow() + timedelta(days=sla_days)

        id_counter += 1
        vuln_id_str = f"VULN-{id_counter:05d}"
        # If this ID already exists for the tenant (e.g. from a prior partial import), keep incrementing
        while db.query(Vulnerability).filter(
            Vulnerability.tenant_id == tenant_id,
            Vulnerability.vuln_id == vuln_id_str
        ).count() > 0:
            id_counter += 1
            vuln_id_str = f"VULN-{id_counter:05d}"

        vuln = Vulnerability(
            tenant_id=tenant_id,
            vuln_id=vuln_id_str,
            title=title,
            description=_lookup(row, "description", "vulnerability description"),
            severity=severity,
            status=vul_status,
            cvss_score=cvss_score,
            cve_id=_lookup(row, "cve_id", "cve number", "cve"),
            affected_component=_lookup(row, "affected_asset", "affected_component", "affected technology", "affected_assets", "affected assets"),
            recommendation=_lookup(row, "remediation", "recommendation", "threat analysis"),
            discovered_at=datetime.utcnow(),
            due_date=due_date,
        )
        db.add(vuln)
        created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error while saving vulnerabilities: {str(e)}")

    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/vulnerabilities/{vuln_id}", response_model=VulnerabilityResponse)
def get_vulnerability(
    vuln_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:view"))
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="User not associated with any tenant")
    
    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(
        Vulnerability.id == vuln_id,
        Vulnerability.tenant_id.in_(user_tenants)
    ).first()
    
    if not vuln:
        raise HTTPException(status_code=404, detail="Vulnerability not found")

    # Keep the stored contextual score honest on every view: recompute it from the
    # finding's current signals + linked asset (no external feeds) so the saved
    # value can never drift stale from the evidence the page shows — the "stored
    # score is stale (42 vs 37)" case. Best-effort: a scoring hiccup must never
    # block viewing the finding.
    try:
        from grc.modules.vuln_management.enrichment.enrichment_service import recompute_composite_priority
        _before = vuln.composite_priority
        recompute_composite_priority(vuln, db)
        if vuln.composite_priority != _before:
            db.commit()
    except Exception:
        db.rollback()

    # Solution count for the Analysis "Patch" tile — so a finding with a synced
    # remediation (e.g. a Nessus solution) reads "Guidance available" instead of the
    # misleading "Not found" when no vendor patch reference is stored.
    from ....models import VulnerabilitySolution
    solution_count = db.query(VulnerabilitySolution).filter(
        VulnerabilitySolution.vulnerability_id == vuln.id
    ).count()

    return _build_vulnerability_response(vuln, solution_count=solution_count)


@router.put("/vulnerabilities/{vuln_id}", response_model=VulnerabilityResponse)
def update_vulnerability(
    vuln_id: int,
    request: VulnerabilityUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit"))
):
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    
    update_data = request.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(vuln, field, value)
    
    vuln.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(vuln)

    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()

    return _build_vulnerability_response(vuln)


@router.delete("/vulnerabilities/{vuln_id}", response_model=MessageResponse)
def delete_vulnerability(
    vuln_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:delete"))
):
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    
    db.delete(vuln)
    db.commit()
    
    return MessageResponse(message="Vulnerability deleted successfully")


@router.post("/vulnerabilities/{vuln_id}/assign", response_model=VulnerabilityResponse)
def assign_vulnerability(
    vuln_id: int,
    request: VulnerabilityAssign,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit"))
):
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    
    user = db.query(GRCUser).filter(GRCUser.id == request.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    vuln.assigned_to = request.user_id
    vuln.updated_at = datetime.utcnow()
    
    if vuln.status == "open":
        vuln.status = "in_progress"
    
    db.commit()
    db.refresh(vuln)

    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()

    return _build_vulnerability_response(vuln)


class LinkAssetIn(BaseModel):
    asset_id: int


def _reload_vuln_full(db: Session, vuln_id: int) -> Vulnerability:
    return db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln_id).first()


@router.post("/vulnerabilities/{vuln_id}/link-asset", response_model=VulnerabilityResponse)
def link_asset_to_vulnerability(
    vuln_id: int,
    request: LinkAssetIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit"))
):
    """Manually attach a finding to an IT asset. Without a linked asset the
    contextual score falls back to defaults for internet-exposure and asset
    criticality — this is the operator's control to attach the right host so
    those signals become real. Recomputes the score immediately."""
    from ....models import ITAsset
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    asset = db.query(ITAsset).filter(
        ITAsset.id == request.asset_id,
        ITAsset.tenant_id.in_(user_tenants),
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    existing = db.query(VulnerabilityAssetLink).filter(
        VulnerabilityAssetLink.vulnerability_id == vuln.id,
        VulnerabilityAssetLink.asset_id == asset.id,
    ).first()
    if not existing:
        db.add(VulnerabilityAssetLink(
            vulnerability_id=vuln.id, asset_id=asset.id,
            impact_on_asset="Manually linked", link_source="manual", auto_linked=False))
        db.flush()
    try:
        from grc.modules.vuln_management.enrichment.enrichment_service import recompute_composite_priority
        recompute_composite_priority(vuln, db)
    except Exception:
        pass
    db.commit()
    return _build_vulnerability_response(_reload_vuln_full(db, vuln.id))


@router.delete("/vulnerabilities/{vuln_id}/link-asset/{asset_id}", response_model=VulnerabilityResponse)
def unlink_asset_from_vulnerability(
    vuln_id: int,
    asset_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit"))
):
    """Detach an asset from a finding (recomputes the score without it)."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    link = db.query(VulnerabilityAssetLink).filter(
        VulnerabilityAssetLink.vulnerability_id == vuln.id,
        VulnerabilityAssetLink.asset_id == asset_id,
    ).first()
    if link:
        db.delete(link)
        db.flush()
        try:
            from grc.modules.vuln_management.enrichment.enrichment_service import recompute_composite_priority
            recompute_composite_priority(vuln, db)
        except Exception:
            pass
        db.commit()
    return _build_vulnerability_response(_reload_vuln_full(db, vuln.id))


class AcceptRiskIn(BaseModel):
    """Accepting a risk is a dated decision, not a status flip."""
    justification: str = Field(min_length=1, max_length=4000)
    review_by: Optional[datetime] = Field(
        default=None,
        description="When the acceptance lapses and the finding returns to the queue.",
    )


@router.post("/vulnerabilities/{vuln_id}/accept-risk", response_model=VulnerabilityResponse)
def accept_risk(
    vuln_id: int,
    body: AcceptRiskIn,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Accept the risk instead of fixing it — WITH an expiry.

    This exists because setting `status = 'accepted'` on its own is a trap: the
    finding leaves the queue and never comes back. Here the review date is
    written to `exception_expires_at`, which is what the expiry sweep keys off,
    so the acceptance genuinely lapses.

    An acceptance with no review date is allowed but flagged as permanent —
    the UI warns, because "accepted forever" is almost never what anyone means.
    """
    user_tenants = get_user_tenants(current_user, db)
    vuln = db.query(Vulnerability).filter(
        Vulnerability.id == vuln_id,
        Vulnerability.tenant_id.in_(user_tenants),
    ).first()
    if not vuln:
        raise HTTPException(status_code=404, detail="Vulnerability not found")

    now = datetime.utcnow()
    previous = vuln.status
    who = getattr(current_user, "display_name", None) or current_user.username

    vuln.status = "accepted"
    vuln.resolved_at = now
    vuln.resolution_notes = body.justification.strip()
    # The bit that actually makes expiry work.
    vuln.exception_status = "approved"
    vuln.exception_justification = body.justification.strip()
    vuln.exception_expires_at = body.review_by
    vuln.exception_expiry = body.review_by  # legacy column kept in sync
    vuln.exception_approved_by = current_user.id
    vuln.exception_approved_at = now
    vuln.updated_at = now

    try:
        from ....models import AuditLog
        until = body.review_by.strftime("%d %b %Y") if body.review_by else "no review date (permanent)"
        db.add(AuditLog(
            tenant_id=vuln.tenant_id,
            user_id=current_user.id,
            action="vulnerability.risk_accepted",
            resource_type="vulnerability",
            resource_id=vuln.id,
            changes={
                "detail": f"Risk accepted by {who} until {until}",
                "from": previous,
                "to": "accepted",
                "review_by": body.review_by.isoformat() if body.review_by else None,
            },
        ))
    except Exception:  # noqa: BLE001
        import logging
        logging.getLogger(__name__).exception("accept_risk audit write failed vuln_id=%s", vuln_id)

    db.commit()
    db.refresh(vuln)

    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()
    return _build_vulnerability_response(vuln)


@router.post("/vulnerabilities/{vuln_id}/status", response_model=VulnerabilityResponse)
def change_vulnerability_status(
    vuln_id: int,
    request: VulnerabilityStatusChange,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit"))
):
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)
    
    valid_statuses = ["open", "in_progress", "remediated", "verified", "closed", "resolved", "accepted", "false_positive"]
    if request.status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")
    
    previous_status = vuln.status
    vuln.status = request.status
    vuln.updated_at = datetime.utcnow()

    # Journal the transition so the finding's History tab shows it. Without
    # this, only configurable-workflow transitions were recorded and a plain
    # status change left no trace at all.
    try:
        from ....models import AuditLog
        db.add(AuditLog(
            tenant_id=vuln.tenant_id,
            user_id=current_user.id,
            action="vulnerability.status_changed",
            resource_type="vulnerability",
            resource_id=vuln.id,
            changes={
                "detail": f"Status changed from {previous_status or 'unset'} to {request.status}",
                "from": previous_status,
                "to": request.status,
            },
        ))
    except Exception:  # noqa: BLE001 — auditing must never block the change
        import logging
        logging.getLogger(__name__).exception(
            "Failed to write status-change audit row for vuln_id=%s", vuln_id
        )

    if request.resolution_notes:
        vuln.resolution_notes = request.resolution_notes
    
    if request.status in [
        "resolved", "remediated", "verified", "closed",
        "accepted", "false_positive", "auto_closed_decommissioned",
    ]:
        vuln.resolved_at = datetime.utcnow()
    if request.status in ["verified", "resolved", "closed"]:
        vuln.verified_by = current_user.id
        vuln.verified_at = datetime.utcnow()

    # Scanner write-back: translate the decision into outbox actions for the
    # finding's scanner connection (no-op for manual/report findings). The
    # push itself is best-effort after commit; unsupported/disabled actions
    # are recorded as skipped with the reason — never silently dropped.
    try:
        from ....modules.integrations.services.writeback_service import WritebackService
        WritebackService.on_status_change(
            db, vuln, previous_status, request.status, user_id=current_user.id,
        )
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            "Writeback enqueue failed for vuln %s (non-fatal)", vuln.id
        )

    db.commit()
    db.refresh(vuln)

    try:
        from ....modules.integrations.services.writeback_service import WritebackService
        WritebackService.try_process_now(db, vuln)
    except Exception:
        pass

    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()

    return _build_vulnerability_response(vuln)


# ---------------------------------------------------------------------------
# CVE auto-fill from title
# ---------------------------------------------------------------------------
# When the operator types a title like "Log4Shell RCE" or "CVE-2021-44228
# in payment-gateway", we offer to pre-fill the CVE-ID / CVSS score /
# severity / CWE / description fields by looking up NVD directly.
#
# Two match strategies:
#   1. Explicit CVE-ID in the title (regex). Highest confidence.
#   2. Known nickname → CVE map (Log4Shell, Spring4Shell, EternalBlue, …).
# Both paths converge on the same NVD lookup so the response shape is
# uniform regardless of how the match was made.

# Famous vulnerability nicknames operators frequently type as titles. We
# keep this short and curated rather than fuzzy-matching every title —
# fuzzy matches against the full CVE catalogue produce too many false
# positives to be useful as an auto-fill default.
_VULN_NICKNAMES: dict[str, str] = {
    "log4shell": "CVE-2021-44228",
    "log4j": "CVE-2021-44228",
    "spring4shell": "CVE-2022-22965",
    "springshell": "CVE-2022-22965",
    "eternalblue": "CVE-2017-0144",
    "heartbleed": "CVE-2014-0160",
    "shellshock": "CVE-2014-6271",
    "dirtycow": "CVE-2016-5195",
    "dirty cow": "CVE-2016-5195",
    "bluekeep": "CVE-2019-0708",
    "printnightmare": "CVE-2021-34527",
    "follina": "CVE-2022-30190",
    "proxyshell": "CVE-2021-34473",
    "proxylogon": "CVE-2021-26855",
    "zerologon": "CVE-2020-1472",
    "ghostcat": "CVE-2020-1938",
    "shitrix": "CVE-2019-19781",
    "kr00k": "CVE-2019-15126",
    "ripple20": "CVE-2020-11896",
    "smbghost": "CVE-2020-0796",
    "curveball": "CVE-2020-0601",
    "moveit": "CVE-2023-34362",
    "citrix bleed": "CVE-2023-4966",
    "regresshion": "CVE-2024-6387",
    "xz backdoor": "CVE-2024-3094",
    "xz utils": "CVE-2024-3094",
}

_CVE_PATTERN = __import__("re").compile(r"CVE-\d{4}-\d{4,7}", __import__("re").IGNORECASE)


def _extract_cve_from_title(title: str) -> Optional[str]:
    """Return the first CVE-ID found in the title, uppercased. None if absent."""
    if not title:
        return None
    m = _CVE_PATTERN.search(title)
    return m.group(0).upper() if m else None


def _match_nickname(title: str) -> Optional[str]:
    """Map well-known vulnerability nicknames to their CVE-ID."""
    if not title:
        return None
    t = title.lower()
    # Longest-prefix wins so "dirty cow" beats a stray "dirty".
    for key in sorted(_VULN_NICKNAMES.keys(), key=len, reverse=True):
        if key in t:
            return _VULN_NICKNAMES[key]
    return None


def _severity_from_cvss(score: Optional[float]) -> Optional[str]:
    """Map a CVSS base score to the severity bucket the rest of the platform
    uses. Mirrors the bucketing in the natural-language summary helpers."""
    if score is None:
        return None
    try:
        s = float(score)
    except (TypeError, ValueError):
        return None
    if s >= 9.0:
        return "critical"
    if s >= 7.0:
        return "high"
    if s >= 4.0:
        return "medium"
    if s > 0.0:
        return "low"
    return "info"


def _fetch_nvd_full(cve_id: str) -> Optional[dict]:
    """Raw NVD GET for a single CVE returning the parsed JSON or None.

    Separate from the enrichment client because we need CVSS + CWE fields
    the cached `NvdResult` shape doesn't carry. Best-effort: never raises.
    """
    import os
    import requests as _requests
    headers = {"User-Agent": "complywerse-vuln-lookup/1.0"}
    api_key = (os.environ.get("NVD_API_KEY") or "").strip()
    if api_key:
        headers["apiKey"] = api_key
    try:
        response = _requests.get(
            "https://services.nvd.nist.gov/rest/json/cves/2.0",
            params={"cveId": cve_id.upper()},
            headers=headers,
            timeout=8,
        )
    except Exception:
        return None
    if response.status_code != 200:
        return None
    try:
        data = response.json()
    except Exception:
        return None
    return data if isinstance(data, dict) else None


class CveLookupRequest(BaseModel):
    title: str
    cve_id: Optional[str] = None  # optional explicit override


class CveLookupResponse(BaseModel):
    matched: bool
    match_source: Optional[str] = None  # 'cve_in_title' | 'nickname' | 'explicit'
    cve_id: Optional[str] = None
    cvss_score: Optional[float] = None
    cvss_vector: Optional[str] = None
    severity: Optional[str] = None
    cwe_id: Optional[str] = None
    description: Optional[str] = None
    nvd_url: Optional[str] = None


@router.post("/vulnerabilities/lookup-by-title", response_model=CveLookupResponse)
def lookup_cve_by_title(
    payload: CveLookupRequest,
    current_user: GRCUser = Depends(require_auth),
):
    """Find a likely CVE for a free-text vulnerability title.

    Returns the highest-confidence match it finds (explicit CVE-ID in title
    > known nickname). When nothing matches, returns `matched=False` and
    the frontend keeps the form blank — never auto-fills speculative data.
    """
    title = (payload.title or "").strip()
    explicit = (payload.cve_id or "").strip().upper()

    if explicit and _CVE_PATTERN.match(explicit):
        cve_id = explicit
        match_source = "explicit"
    else:
        extracted = _extract_cve_from_title(title)
        if extracted:
            cve_id = extracted
            match_source = "cve_in_title"
        else:
            nickname = _match_nickname(title)
            if nickname:
                cve_id = nickname
                match_source = "nickname"
            else:
                return CveLookupResponse(matched=False)

    raw = _fetch_nvd_full(cve_id)
    if not raw:
        # We know the CVE-ID — return that much so the user can at least
        # accept the ID into the form even if NVD is unreachable.
        return CveLookupResponse(
            matched=True,
            match_source=match_source,
            cve_id=cve_id,
            nvd_url=f"https://nvd.nist.gov/vuln/detail/{cve_id}",
        )

    vulns = raw.get("vulnerabilities") or []
    if not vulns:
        return CveLookupResponse(
            matched=True,
            match_source=match_source,
            cve_id=cve_id,
            nvd_url=f"https://nvd.nist.gov/vuln/detail/{cve_id}",
        )

    cve = (vulns[0] or {}).get("cve") or {}

    # English description.
    description: Optional[str] = None
    for desc in cve.get("descriptions") or []:
        if (desc or {}).get("lang") == "en":
            description = (desc.get("value") or "").strip() or None
            if description and len(description) > 1000:
                description = description[:1000].rsplit(" ", 1)[0] + "…"
            break

    # CVSS — prefer v3.1, fall back to v3.0, then v2.0.
    cvss_score: Optional[float] = None
    cvss_vector: Optional[str] = None
    metrics = cve.get("metrics") or {}
    for key in ("cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        entries = metrics.get(key) or []
        if not entries:
            continue
        cvss_data = (entries[0] or {}).get("cvssData") or {}
        if cvss_data.get("baseScore") is not None:
            try:
                cvss_score = float(cvss_data["baseScore"])
            except (TypeError, ValueError):
                cvss_score = None
        cvss_vector = cvss_data.get("vectorString") or cvss_vector
        if cvss_score is not None:
            break

    # CWE — first weakness entry, English.
    cwe_id: Optional[str] = None
    for w in cve.get("weaknesses") or []:
        for d in (w or {}).get("description") or []:
            if (d or {}).get("lang") == "en":
                val = (d.get("value") or "").strip()
                if val and val.upper().startswith("CWE-"):
                    cwe_id = val.upper()
                    break
        if cwe_id:
            break

    return CveLookupResponse(
        matched=True,
        match_source=match_source,
        cve_id=cve_id,
        cvss_score=cvss_score,
        cvss_vector=cvss_vector,
        severity=_severity_from_cvss(cvss_score),
        cwe_id=cwe_id,
        description=description,
        nvd_url=f"https://nvd.nist.gov/vuln/detail/{cve_id}",
    )


# ---------------------------------------------------------------------------
# Vulnerability dependency chains
# ---------------------------------------------------------------------------
# A "dependent" vuln requires its "prerequisite" to be exploitable in
# practice (e.g. a local privilege-escalation flaw needs an RCE foothold
# first). Storing the chain lets a triager push the right one to the top
# of the queue and gives auditors a structured "kill-chain" view.
#
# We deliberately do NOT auto-modify composite_priority based on chain
# state — that would surprise users whose SLAs/dashboards are calibrated
# to the current numbers. The chain is informational + visualised; humans
# choose how to act on it.


class VulnDependencyCreate(BaseModel):
    prerequisite_vuln_id: int
    notes: Optional[str] = None
    chain_stage: Optional[str] = None  # initial_access / execution / priv_esc / lateral / exfil


class VulnDependencyResponse(BaseModel):
    id: int
    dependent_vuln_id: int
    prerequisite_vuln_id: int
    notes: Optional[str] = None
    chain_stage: Optional[str] = None
    # Denormalised prereq summary so the UI doesn't need a second fetch.
    prerequisite_title: Optional[str] = None
    prerequisite_vuln_code: Optional[str] = None
    prerequisite_severity: Optional[str] = None
    prerequisite_status: Optional[str] = None
    prerequisite_cve_id: Optional[str] = None
    prerequisite_kev_flag: Optional[bool] = None
    prerequisite_composite_priority: Optional[float] = None
    created_at: Optional[datetime] = None

    class Config:
        from_attributes = True


def _serialize_dependency(dep: VulnerabilityDependency, prereq: Optional[Vulnerability]) -> dict:
    """Build the response shape with prereq fields flattened in."""
    return {
        "id": dep.id,
        "dependent_vuln_id": dep.dependent_vuln_id,
        "prerequisite_vuln_id": dep.prerequisite_vuln_id,
        "notes": dep.notes,
        "chain_stage": dep.chain_stage,
        "prerequisite_title": prereq.title if prereq else None,
        "prerequisite_vuln_code": prereq.vuln_id if prereq else None,
        "prerequisite_severity": prereq.severity if prereq else None,
        "prerequisite_status": prereq.status if prereq else None,
        "prerequisite_cve_id": prereq.cve_id if prereq else None,
        "prerequisite_kev_flag": bool(prereq.kev_flag) if prereq and prereq.kev_flag is not None else None,
        "prerequisite_composite_priority": prereq.composite_priority if prereq else None,
        "created_at": dep.created_at,
    }


@router.get("/vulnerabilities/{vuln_id}/dependencies")
def list_vulnerability_dependencies(
    vuln_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """List the chain: vulns this one *depends on* (prerequisites) and
    vulns that *depend on* this one (dependents). Bidirectional view in
    one round-trip so the UI renders the whole graph for this row.
    """
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    # Prerequisites this vuln depends on.
    prereq_rows = (
        db.query(VulnerabilityDependency)
        .options(joinedload(VulnerabilityDependency.prerequisite_vuln))
        .filter(VulnerabilityDependency.dependent_vuln_id == vuln.id)
        .all()
    )
    prerequisites = [
        _serialize_dependency(d, d.prerequisite_vuln) for d in prereq_rows
    ]

    # Other vulns that depend on this one.
    dep_rows = (
        db.query(VulnerabilityDependency)
        .options(joinedload(VulnerabilityDependency.dependent_vuln))
        .filter(VulnerabilityDependency.prerequisite_vuln_id == vuln.id)
        .all()
    )
    dependents = []
    for d in dep_rows:
        dv = d.dependent_vuln
        dependents.append({
            "id": d.id,
            "dependent_vuln_id": d.dependent_vuln_id,
            "prerequisite_vuln_id": d.prerequisite_vuln_id,
            "notes": d.notes,
            "chain_stage": d.chain_stage,
            "dependent_title": dv.title if dv else None,
            "dependent_vuln_code": dv.vuln_id if dv else None,
            "dependent_severity": dv.severity if dv else None,
            "dependent_status": dv.status if dv else None,
            "dependent_cve_id": dv.cve_id if dv else None,
            "dependent_kev_flag": bool(dv.kev_flag) if dv and dv.kev_flag is not None else None,
            "dependent_composite_priority": dv.composite_priority if dv else None,
            "created_at": d.created_at,
        })

    # Chain warning summary — surfaces in the UI as a coloured banner.
    # If any unresolved prereq is KEV-listed or has composite > 9, the
    # operator should treat THIS vuln with extra urgency.
    open_high_prereqs = [
        p for p in prerequisites
        if (p.get("prerequisite_status") or "").lower() not in (
            "resolved", "remediated", "verified", "closed", "false_positive",
        )
        and (
            bool(p.get("prerequisite_kev_flag"))
            or (isinstance(p.get("prerequisite_composite_priority"), (int, float))
                and (p["prerequisite_composite_priority"] or 0) >= 9.0)
        )
    ]
    chain_warning = None
    if open_high_prereqs:
        chain_warning = (
            f"{len(open_high_prereqs)} unresolved prerequisite"
            f"{'s' if len(open_high_prereqs) != 1 else ''} with high real-world "
            f"urgency. Closing those raises the effective exploitability of this vuln."
        )

    return {
        "prerequisites": prerequisites,
        "dependents": dependents,
        "chain_warning": chain_warning,
    }


@router.post("/vulnerabilities/{vuln_id}/dependencies", response_model=VulnDependencyResponse)
def add_vulnerability_dependency(
    vuln_id: int,
    payload: VulnDependencyCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Declare that `vuln_id` depends on `prerequisite_vuln_id`."""
    user_tenants = get_user_tenants(current_user, db)
    dependent = get_vuln_or_404(vuln_id, user_tenants, db)

    if payload.prerequisite_vuln_id == vuln_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="A vulnerability cannot depend on itself.",
        )

    prerequisite = get_vuln_or_404(payload.prerequisite_vuln_id, user_tenants, db)
    if prerequisite.tenant_id != dependent.tenant_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot link vulnerabilities across tenants.",
        )

    # Reject simple cycles (A→B→A). Deeper cycles still possible — surfaced
    # in the UI but not blocked, since real chains can legitimately loop
    # (e.g. mutually reinforcing exploits).
    inverse = (
        db.query(VulnerabilityDependency)
        .filter(
            VulnerabilityDependency.dependent_vuln_id == prerequisite.id,
            VulnerabilityDependency.prerequisite_vuln_id == dependent.id,
        )
        .first()
    )
    if inverse:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Would create a direct cycle — '{prerequisite.title}' already "
                f"depends on '{dependent.title}'."
            ),
        )

    existing = (
        db.query(VulnerabilityDependency)
        .filter(
            VulnerabilityDependency.dependent_vuln_id == dependent.id,
            VulnerabilityDependency.prerequisite_vuln_id == prerequisite.id,
        )
        .first()
    )
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Dependency already exists.",
        )

    dep = VulnerabilityDependency(
        tenant_id=dependent.tenant_id,
        dependent_vuln_id=dependent.id,
        prerequisite_vuln_id=prerequisite.id,
        notes=payload.notes,
        chain_stage=payload.chain_stage,
        created_by=current_user.id,
    )
    db.add(dep)
    db.commit()
    db.refresh(dep)
    return VulnDependencyResponse(**_serialize_dependency(dep, prerequisite))


@router.delete("/vulnerabilities/{vuln_id}/dependencies/{dependency_id}")
def remove_vulnerability_dependency(
    vuln_id: int,
    dependency_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Remove a single dependency edge."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    dep = (
        db.query(VulnerabilityDependency)
        .filter(
            VulnerabilityDependency.id == dependency_id,
            VulnerabilityDependency.dependent_vuln_id == vuln.id,
        )
        .first()
    )
    if not dep:
        raise HTTPException(status_code=404, detail="Dependency not found.")
    db.delete(dep)
    db.commit()
    return {"deleted": True}


# ---------------------------------------------------------------------------
# Threat-intelligence enrichment
# ---------------------------------------------------------------------------
# Pulls EPSS exploit-probability, CISA KEV flag, and NVD canonical metadata
# from free public APIs, writes them to the vuln row, and recomputes
# composite_priority. Synchronous and best-effort — every external call has
# its own try/except inside the service, so this endpoint never 5xx's because
# a third-party service is slow or down.


@router.post("/vulnerabilities/{vuln_id}/enrich", response_model=VulnerabilityResponse)
def enrich_vulnerability_endpoint(
    vuln_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Run NVD + EPSS + CISA KEV enrichment for one vuln. Returns the
    updated record. Idempotent — calling repeatedly just refreshes the
    enrichment fields (EPSS especially changes daily)."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    # Lazy import keeps the optional `requests` + redis deps out of this
    # module's import path for callers that never hit the enrich endpoint.
    from ..enrichment import enrich_vulnerability

    try:
        enrich_vulnerability(vuln, db)
    except Exception:
        # The service is designed never to raise; this catch is paranoia.
        # Don't fail the request — return whatever state the row is in.
        import logging
        logging.getLogger(__name__).exception(
            "Unexpected exception during enrichment for vuln %s", vuln.id
        )

    db.refresh(vuln)
    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()
    return _build_vulnerability_response(vuln)


@router.post("/vulnerabilities/enrich-all")
def bulk_enrich_endpoint(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Queue a bulk-enrichment Celery job for the caller's tenant. Returns
    immediately — actual work runs in the worker. Use this to backfill
    EPSS/KEV/NVD on rows that pre-date the enrichment feature."""
    tenant_id = get_user_primary_tenant(current_user, db)
    # Resolve slug for TenantTask. Lazy import keeps Celery deps out of the
    # request path until someone actually clicks the button.
    from ....db import MasterSession
    from ....models import Tenant as _Tenant

    master = MasterSession()
    try:
        row = master.query(_Tenant.slug).filter(_Tenant.id == tenant_id).first()
    finally:
        master.close()

    if not row or not row[0]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not resolve tenant slug for bulk enrichment.",
        )

    try:
        from ...tasks.vulnerabilities import bulk_enrich_open_vulns
        result = bulk_enrich_open_vulns.delay(tenant_slug=row[0], only_with_cve=True)
        return {"queued": True, "task_id": result.id}
    except Exception as exc:
        # Broker down — give the user a clear message instead of a 500.
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Could not queue bulk enrichment: {exc}",
        )


# ---------------------------------------------------------------------------
# Phase 6 — Vendor patch intelligence (MSRC first)
# ---------------------------------------------------------------------------
# Asks the responsible PSIRT (currently only MSRC; Red Hat / Cisco land in
# future PRs) for the KB articles / vendor advisory IDs / remediation text
# that fix this vuln. Synchronous and best-effort — same failure semantics as
# /enrich.


@router.post("/vulnerabilities/{vuln_id}/sync-patch-info", response_model=VulnerabilityResponse)
def sync_patch_info_endpoint(
    vuln_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Ask the vendor PSIRT (MSRC for now) for patch metadata. Returns the
    updated record so the UI doesn't need a follow-up GET. Idempotent — a
    re-sync replaces this vuln's MSRC entries with the fresh data while
    preserving entries from other PSIRTs."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    # Lazy import — keeps `requests` + redis off the import path for callers
    # that never click the Sync button.
    from ..patch_intel import sync_patch_intel

    try:
        sync_patch_intel(vuln, db)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            "Unexpected exception during patch-intel sync for vuln %s", vuln.id
        )

    db.refresh(vuln)
    vuln = db.query(Vulnerability).options(
        joinedload(Vulnerability.assignee),
        joinedload(Vulnerability.verifier),
        joinedload(Vulnerability.asset_links).joinedload(VulnerabilityAssetLink.asset),
    ).filter(Vulnerability.id == vuln.id).first()
    return _build_vulnerability_response(vuln)


@router.post("/vulnerabilities/sync-patch-info-all")
def bulk_sync_patch_info_endpoint(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Queue a tenant-wide patch-intel sync via Celery. Returns immediately;
    the worker walks every open CVE-bearing vuln and asks MSRC for each.
    Non-Microsoft CVEs are cached as negative hits so subsequent runs skip
    them quickly."""
    tenant_id = get_user_primary_tenant(current_user, db)
    from ....db import MasterSession
    from ....models import Tenant as _Tenant

    master = MasterSession()
    try:
        row = master.query(_Tenant.slug).filter(_Tenant.id == tenant_id).first()
    finally:
        master.close()

    if not row or not row[0]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not resolve tenant slug for bulk patch-intel sync.",
        )

    try:
        from ...tasks.patch_intel import bulk_sync_msrc
        result = bulk_sync_msrc.delay(tenant_slug=row[0])
        return {"queued": True, "task_id": result.id}
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Could not queue bulk patch-intel sync: {exc}",
        )


# ---------------------------------------------------------------------------
# Phase 8 — Exception Workflow
# ---------------------------------------------------------------------------
# Formal request → approve|deny → revoke|expire state machine. Separation of
# duties enforced server-side: the user who requested an exception cannot
# also approve or deny it. The legacy `is_exception`/`exception_reason`/
# `exception_approved_by`/`exception_expiry` columns are kept in sync so
# existing dashboards and reports keep working.

from ....schemas import (  # noqa: E402  (intentional inline import for grouping)
    ExceptionRequestBody,
    ExceptionApproveBody,
    ExceptionDenyBody,
    ExceptionRevokeBody,
)


def _exception_fsm_response(vuln, summary: dict) -> dict:
    """Combine the FSM summary with the canonical VulnerabilityResponse so
    the UI can refresh either the panel or the whole row from one call."""
    return {
        "vulnerability": _build_vulnerability_response(vuln).model_dump(),
        "exception": summary,
    }


@router.post("/vulnerabilities/{vuln_id}/exception/request")
def request_vuln_exception(
    vuln_id: int,
    body: ExceptionRequestBody,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Open an exception request for a vulnerability. Justification mandatory."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    from ....services.vuln_exception import (
        ExceptionWorkflowError,
        request_exception,
    )

    try:
        summary = request_exception(
            vuln=vuln,
            actor_id=current_user.id,
            justification=body.justification,
            compensating_controls=body.compensating_controls,
            expires_at=body.expires_at,
        )
    except ExceptionWorkflowError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc),
        ) from exc

    db.commit()
    db.refresh(vuln)
    return _exception_fsm_response(vuln, summary)


@router.post("/vulnerabilities/{vuln_id}/exception/approve")
def approve_vuln_exception(
    vuln_id: int,
    body: ExceptionApproveBody,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    # Approval is a stronger action than edit, but until a dedicated
    # `approve_exception` permission is rolled out per-tenant we gate it
    # behind edit. The FSM still enforces separation of duties (the
    # requester cannot also approve), which is the auditor-relevant control.
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Approve a `requested` exception. Separation-of-duties enforced."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    from ....services.vuln_exception import (
        ExceptionWorkflowError,
        approve_exception,
    )

    try:
        summary = approve_exception(
            vuln=vuln,
            actor_id=current_user.id,
            comment=body.comment,
            expires_at=body.expires_at,
        )
    except ExceptionWorkflowError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc),
        ) from exc

    db.commit()
    db.refresh(vuln)
    return _exception_fsm_response(vuln, summary)


@router.post("/vulnerabilities/{vuln_id}/exception/deny")
def deny_vuln_exception(
    vuln_id: int,
    body: ExceptionDenyBody,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Deny a `requested` exception. Reason mandatory; separation-of-duties enforced."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    from ....services.vuln_exception import (
        ExceptionWorkflowError,
        deny_exception,
    )

    try:
        summary = deny_exception(
            vuln=vuln,
            actor_id=current_user.id,
            denial_reason=body.denial_reason,
        )
    except ExceptionWorkflowError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc),
        ) from exc

    db.commit()
    db.refresh(vuln)
    return _exception_fsm_response(vuln, summary)


class BulkExceptionRequestBody(BaseModel):
    vulnerability_ids: List[int]
    justification: str
    compensating_controls: Optional[List[str]] = None
    expires_at: Optional[datetime] = None


@router.post("/vulnerabilities/exception/bulk-request")
def bulk_request_vuln_exception(
    body: BulkExceptionRequestBody,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Open exception requests on a batch of vulns at once. Each row gets
    its own state-machine entry (per-vuln history preserved); a single
    justification + expiry applies to all. Rows that aren't in a
    transitionable state are skipped with a per-row error message.

    Returns a summary `{requested: int, skipped: int, errors: [{id, msg}]}`."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No tenant access.")
    if not body.vulnerability_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="vulnerability_ids must contain at least one id.",
        )
    if not (body.justification or "").strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Justification is required.",
        )

    from ....services.vuln_exception import (
        ExceptionWorkflowError,
        request_exception,
    )

    rows = (
        db.query(Vulnerability)
        .filter(Vulnerability.id.in_(body.vulnerability_ids))
        .filter(Vulnerability.tenant_id.in_(user_tenants))
        .all()
    )
    found = {v.id for v in rows}
    requested = 0
    skipped = 0
    errors: List[dict] = []

    for vid in body.vulnerability_ids:
        if vid not in found:
            errors.append({"id": vid, "msg": "not_found_or_no_access"})
            skipped += 1
            continue
    for vuln in rows:
        try:
            request_exception(
                vuln=vuln,
                actor_id=current_user.id,
                justification=body.justification,
                compensating_controls=body.compensating_controls,
                expires_at=body.expires_at,
            )
            requested += 1
        except ExceptionWorkflowError as exc:
            errors.append({"id": vuln.id, "msg": str(exc)})
            skipped += 1

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Commit failed.")

    return {"requested": requested, "skipped": skipped, "errors": errors}


@router.get("/vulnerabilities/exception-queue")
def list_exception_queue(
    state: Optional[str] = Query(None, description="Filter by exception_status; defaults to all non-none states."),
    skip: int = 0,
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Queue view across the tenant — vulns currently in any exception
    state. Used by the Exceptions queue page so reviewers don't have to
    open each vuln to see what's pending."""
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"total": 0, "rows": []}

    query = (
        db.query(Vulnerability)
        .filter(Vulnerability.tenant_id.in_(user_tenants))
        .filter(Vulnerability.exception_status.isnot(None))
        .filter(Vulnerability.exception_status != "none")
    )
    if state:
        query = query.filter(Vulnerability.exception_status == state)

    total = query.count()
    rows = (
        query.order_by(Vulnerability.exception_requested_at.desc().nulls_last())
        .offset(skip).limit(limit).all()
    )

    return {
        "total": total,
        "rows": [
            {
                "id": v.id,
                "vuln_id": v.vuln_id,
                "title": v.title,
                "severity": v.severity,
                "cve_id": v.cve_id,
                "exception_status": v.exception_status,
                "exception_requested_by_id": v.exception_requested_by_id,
                "exception_requested_at": v.exception_requested_at.isoformat() if v.exception_requested_at else None,
                "exception_approved_at": v.exception_approved_at.isoformat() if v.exception_approved_at else None,
                "exception_expires_at": v.exception_expires_at.isoformat() if v.exception_expires_at else None,
                "exception_justification": (v.exception_justification or "")[:300] if v.exception_justification else None,
                "composite_priority": v.composite_priority,
            }
            for v in rows
        ],
    }


@router.post("/vulnerabilities/{vuln_id}/exception/revoke")
def revoke_vuln_exception(
    vuln_id: int,
    body: ExceptionRevokeBody,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
    _perm: bool = Depends(require_tenant_permission("vulnerabilities:vulnerability_register:edit")),
):
    """Revoke an `approved` exception. Used when a compensating control fails
    or new threat intel makes the exception untenable."""
    user_tenants = get_user_tenants(current_user, db)
    vuln = get_vuln_or_404(vuln_id, user_tenants, db)

    from ....services.vuln_exception import (
        ExceptionWorkflowError,
        revoke_exception,
    )

    try:
        summary = revoke_exception(
            vuln=vuln,
            actor_id=current_user.id,
            reason=body.reason,
        )
    except ExceptionWorkflowError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc),
        ) from exc

    db.commit()
    db.refresh(vuln)
    return _exception_fsm_response(vuln, summary)
