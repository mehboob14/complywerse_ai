"""Bulk import Issues from CSV / Excel.

Endpoints (mounted under /issue-management):
  GET  /issues/template/download
  POST /issues/import/upload
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ....models import Issue, IssueAction, IssueActivity, GRCUser, get_db
from ....routers.auth_router import (
    require_auth,
    get_user_primary_tenant,
    require_tenant_permission,
)
from ..services.severity_resolver import resolve_severity
from ..services.code_generator import next_issue_code

router = APIRouter(prefix="/issues", tags=["Issue Management - Import"])

_require_create = require_tenant_permission("issue_management:issues:create")

# Column layout for the downloadable template (header, description, example).
ISSUE_TEMPLATE_COLUMNS = [
    ("title", "Required — short issue title", "Unpatched SSL on payment gateway"),
    ("description", "Optional — detail / context", "TLS 1.0 still enabled on gw-prod-01"),
    ("impact", "high | medium | low (feeds severity matrix)", "high"),
    ("urgency", "high | medium | low (feeds severity matrix)", "medium"),
    ("severity", "Optional override: critical|high|medium|low|informational", ""),
    ("issue_type", "incident|audit_finding|non_conformance|vendor_breach|process_gap|capa|other", "incident"),
    ("category", "security|privacy|operations|contract|data|regulatory|safety", "security"),
    ("root_cause", "Optional short root-cause label", "Missing patch window"),
    ("assignee_email", "Optional — match an existing tenant user email", "owner@example.com"),
    ("due_date", "Optional ISO date YYYY-MM-DD", "2026-08-15"),
    ("capa_title", "Optional — creates an initial CAPA action", "Apply TLS 1.2+ and verify"),
    ("capa_type", "corrective|preventive|containment|verification (default corrective)", "corrective"),
]


VALID_IMPACT = {"high", "medium", "low"}
VALID_URGENCY = {"high", "medium", "low"}
VALID_SEVERITY = {"critical", "high", "medium", "low", "informational"}
VALID_ISSUE_TYPE = {
    "incident", "audit_finding", "non_conformance", "vendor_breach",
    "process_gap", "capa", "other",
}
VALID_CATEGORY = {
    "security", "privacy", "operations", "contract", "data", "regulatory", "safety",
}
VALID_CAPA_TYPE = {"corrective", "preventive", "containment", "verification"}


def _norm_header(h: Any) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _cell(row: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        for rk, rv in row.items():
            if _norm_header(rk) == _norm_header(k):
                if rv is None:
                    return ""
                return str(rv).strip()
    return ""


def _parse_date(raw: str) -> Optional[datetime]:
    if not raw:
        return None
    text = raw.strip().replace("Z", "+00:00")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


@router.get("/template/download", dependencies=[Depends(_require_create)])
def download_issue_template(
    current_user: GRCUser = Depends(require_auth),
):
    """CSV template for bulk issue import (Excel-compatible)."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([c[0] for c in ISSUE_TEMPLATE_COLUMNS])
    writer.writerow([c[1] for c in ISSUE_TEMPLATE_COLUMNS])
    writer.writerow([c[2] for c in ISSUE_TEMPLATE_COLUMNS])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=issues_import_template.csv"},
    )


@router.post("/import/upload", dependencies=[Depends(_require_create)])
async def upload_issues_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    """Bulk-import Issues (+ optional initial CAPA) from CSV or Excel."""
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=400, detail="User is not assigned to any tenant")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(
            status_code=400,
            detail="File must be CSV or Excel (.csv, .xlsx, .xls)",
        )

    content = await file.read()
    rows: List[Dict[str, Any]] = []

    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        else:
            from openpyxl import load_workbook

            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers: List[str] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [str(c).strip() if c is not None else "" for c in row]
                    continue
                # Skip description / example rows from the template.
                first = str(row[0]).lower() if row and row[0] is not None else ""
                if idx == 1 and ("required" in first or "optional" in first or "short issue" in first):
                    continue
                if idx == 2 and first.startswith("unpatched"):
                    continue
                if not any(row):
                    continue
                row_dict: Dict[str, Any] = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_dict[headers[col_idx]] = cell
                rows.append(row_dict)
            wb.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {exc}") from exc

    # Drop empty / instruction rows that slipped through.
    cleaned: List[Dict[str, Any]] = []
    for r in rows:
        title = _cell(r, "title")
        if not title:
            continue
        if title.lower().startswith("required") or title.lower().startswith("optional"):
            continue
        cleaned.append(r)
    rows = cleaned

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    # Preload user emails for assignee matching (per-tenant DB — all users here).
    users = db.query(GRCUser).all()
    email_to_id: Dict[str, int] = {}
    for u in users:
        email = (getattr(u, "email", None) or "").strip().lower()
        if email:
            email_to_id[email] = u.id

    imported = 0
    capa_created = 0
    errors: List[str] = []

    for i, row in enumerate(rows, start=1):
        try:
            with db.begin_nested():
                title = _cell(row, "title")
                if not title:
                    raise ValueError("title is required")

                impact = (_cell(row, "impact") or "medium").lower()
                urgency = (_cell(row, "urgency") or "medium").lower()
                if impact not in VALID_IMPACT:
                    raise ValueError(f"invalid impact '{impact}'")
                if urgency not in VALID_URGENCY:
                    raise ValueError(f"invalid urgency '{urgency}'")

                severity_override = (_cell(row, "severity") or "").lower() or None
                if severity_override and severity_override not in VALID_SEVERITY:
                    raise ValueError(f"invalid severity '{severity_override}'")

                issue_type = (_cell(row, "issue_type") or "incident").lower()
                category = (_cell(row, "category") or "operations").lower()
                if issue_type not in VALID_ISSUE_TYPE:
                    issue_type = "other"
                if category not in VALID_CATEGORY:
                    category = "operations"

                computed_severity, _ack, resolve_hours = resolve_severity(
                    impact=impact, urgency=urgency, tenant_id=tenant_id, db=db,
                )
                final_severity = severity_override or computed_severity

                assignee_email = _cell(row, "assignee_email").lower()
                assignee_id = email_to_id.get(assignee_email) if assignee_email else None

                due_date = _parse_date(_cell(row, "due_date"))
                detected_at = datetime.utcnow()
                target_closure = detected_at + timedelta(hours=resolve_hours)

                issue = Issue(
                    tenant_id=tenant_id,
                    title=title[:255],
                    description=_cell(row, "description") or None,
                    severity=final_severity,
                    severity_override=severity_override,
                    severity_override_reason=(
                        "Set via bulk import" if severity_override else None
                    ),
                    impact=impact,
                    urgency=urgency,
                    issue_type=issue_type,
                    category=category,
                    root_cause=_cell(row, "root_cause") or None,
                    detected_at=detected_at,
                    target_closure_date=target_closure,
                    due_date=due_date,
                    reporter_id=current_user.id,
                    assignee_id=assignee_id,
                    source_type="import",
                    workflow_state="new",
                    status="open",
                )
                db.add(issue)
                db.flush()
                issue.code = next_issue_code(tenant_id, db)

                db.add(IssueActivity(
                    issue_id=issue.id,
                    user_id=current_user.id,
                    type="created",
                    payload={"source": "bulk_import", "row": i},
                ))

                capa_title = _cell(row, "capa_title")
                if capa_title:
                    capa_type = (_cell(row, "capa_type") or "corrective").lower()
                    if capa_type not in VALID_CAPA_TYPE:
                        capa_type = "corrective"
                    action = IssueAction(
                        issue_id=issue.id,
                        action_type=capa_type,
                        title=capa_title[:255],
                        description=None,
                        assignee_id=assignee_id,
                        due_date=due_date,
                        status="planned",
                        created_by=current_user.id,
                    )
                    db.add(action)
                    db.flush()
                    db.add(IssueActivity(
                        issue_id=issue.id,
                        user_id=current_user.id,
                        type="action_added",
                        payload={
                            "action_id": action.id,
                            "action_type": capa_type,
                            "title": capa_title,
                            "source": "bulk_import",
                        },
                    ))
                    capa_created += 1

            imported += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save imported issues: {exc}") from exc

    return {
        "success": imported > 0,
        "imported": imported,
        "capa_created": capa_created,
        "total_rows": len(rows),
        "errors": errors[:50],
        "total_errors": len(errors),
        "message": (
            f"Imported {imported} of {len(rows)} issues"
            + (f" ({capa_created} CAPA actions)" if capa_created else "")
            + (f" with {len(errors)} row error(s)" if errors else "")
        ),
    }
