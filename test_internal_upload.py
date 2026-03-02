#!/usr/bin/env python3
"""
Quick test to verify Internal Risk Template upload works
"""
import requests
import json
from pathlib import Path

# Configuration
API_BASE_URL = "http://localhost:4000"
TEST_TEMPLATE_PATH = Path("backend/risks_templates/Internal_Risk_Template.xlsx")
TENANT_ID = 1

# Test authentication - would need actual token in production
headers = {
    "Authorization": "Bearer test-token"  # This will fail in real scenario
}

def test_upload():
    if not TEST_TEMPLATE_PATH.exists():
        print(f"❌ Template not found: {TEST_TEMPLATE_PATH}")
        return False
    
    print(f"📄 Testing upload with: {TEST_TEMPLATE_PATH}")
    
    with open(TEST_TEMPLATE_PATH, 'rb') as f:
        files = {
            'file': (TEST_TEMPLATE_PATH.name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        data = {
            'register_type': 'internal'
        }
        
        try:
            # Mock the upload to see what happens
            print(f"  Would POST to: {API_BASE_URL}/api/erm/risks/upload")
            print(f"  Register Type: internal")
            print(f"  File: {TEST_TEMPLATE_PATH.name}")
            print(f"  File size: {TEST_TEMPLATE_PATH.stat().st_size} bytes")
            
            # Read and analyze the template structure
            import openpyxl
            wb = openpyxl.load_workbook(TEST_TEMPLATE_PATH)
            print(f"\n📊 Template Analysis:")
            print(f"  Sheets: {wb.sheetnames}")
            
            # Check Risk Register sheet
            if 'Risk Register' in wb.sheetnames:
                ws = wb['Risk Register']
                print(f"\n  📋 Risk Register Sheet:")
                for row_num in range(1, 10):
                    row_values = [cell.value for cell in ws[row_num]]
                    non_empty = [v for v in row_values if v]
                    if non_empty:
                        print(f"    Row {row_num}: {non_empty[:5]}{'...' if len(non_empty) > 5 else ''}")
                
                # Count data rows
                data_row_count = 0
                for row_num in range(8, ws.max_row + 1):
                    row_values = [cell.value for cell in ws[row_num]]
                    if any(row_values):
                        data_row_count += 1
                
                print(f"\n  ✅ Data rows (from row 8): {data_row_count}")
            
            return True
            
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

if __name__ == "__main__":
    print("=" * 60)
    print("INTERNAL RISK TEMPLATE UPLOAD TEST")
    print("=" * 60)
    success = test_upload()
    print("\n" + "=" * 60)
    if success:
        print("✅ Analysis complete - ready to test actual upload")
    else:
        print("❌ Test failed")
    print("=" * 60)
