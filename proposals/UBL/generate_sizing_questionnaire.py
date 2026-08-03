r"""Generate UBL Pakistan on-prem infrastructure sizing questionnaire.

Output: UBL_Pakistan_Infrastructure_Sizing_Questionnaire.xlsx
        in the same directory as this script.

Structure (per plan twinkling-meandering-panda.md):
    1. Cover                  visible
    2. Questionnaire          visible (10 questions)
    3. Local-LLM Reference    visible (3 GPU tiers)
    4. Internal Sizing Tiers  hidden (Tier S/M/L compute + data volumes)
    5. Internal Calc          veryHidden (tier auto-pick formula)

Run:
    .\.venv\Scripts\python.exe proposals\UBL\generate_sizing_questionnaire.py
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Style constants ──────────────────────────────────────────────────────
NAVY = "1F3864"
SLATE = "475569"
LIGHT_GREY = "F1F5F9"
HEADER_GREY = "E2E8F0"
ACCENT_BLUE = "DBEAFE"
GREEN = "DCFCE7"
AMBER = "FEF3C7"
RED = "FEE2E2"
WHITE = "FFFFFF"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=SLATE)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=NAVY)
FONT_LABEL = Font(name="Calibri", size=10, bold=True, color=SLATE)
FONT_BODY = Font(name="Calibri", size=10, color="1E293B")
FONT_QUESTION = Font(name="Calibri", size=11, bold=True, color="0F172A")
FONT_WHY = Font(name="Calibri", size=9, italic=True, color=SLATE)
FONT_TABLE_HDR = Font(name="Calibri", size=10, bold=True, color=WHITE)
FONT_INPUT = Font(name="Calibri", size=10, color="0F172A")

FILL_HEADER_GREY = PatternFill("solid", fgColor=HEADER_GREY)
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT_GREY)
FILL_INPUT = PatternFill("solid", fgColor=WHITE)
FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_AMBER = PatternFill("solid", fgColor=AMBER)
FILL_RED = PatternFill("solid", fgColor=RED)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT_BLUE)

THIN = Side(style="thin", color="CBD5E1")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def style_cell(cell, *, font=None, fill=None, align=None, border=BORDER_ALL):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if border is not None:
        cell.border = border


# ── Sheet 1: Cover ───────────────────────────────────────────────────────
def build_cover(ws):
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 18

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "Compliverse — Infrastructure Sizing Questionnaire"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT

    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = f"Prepared for: United Bank Limited (UBL) Pakistan   ·   {date.today().strftime('%d %B %Y')}"
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[5].height = 6

    ws.merge_cells("B6:F6")
    c = ws["B6"]
    c.value = "Purpose"
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT

    purpose_text = (
        "To deliver an accurate on-premises deployment of Compliverse (our multi-tenant GRC platform) "
        "for 100 named users at UBL Pakistan, we need to align on ten infrastructure parameters. "
        "Your answers will let us right-size compute, storage, database, backup, and — if you opt for "
        "on-prem LLM — the GPU footprint required to preserve the document-parsing and ComplyChat "
        "accuracy you saw in the demo."
    )
    ws.merge_cells("B7:F9")
    c = ws["B7"]
    c.value = purpose_text
    c.font = FONT_BODY
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 22

    # Working assumptions block
    ws.merge_cells("B11:F11")
    c = ws["B11"]
    c.value = "Our working assumptions about the UBL deployment"
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT

    assumptions = [
        ("Named GRC users", "100 (compliance team + auditors + risk owners + admins)"),
        ("Peak concurrency", "~30 (30% of named users, aligned with PKT 09:00–18:00 working hours)"),
        ("IT estate in scope for vuln scanning", "~5,000 assets (typical for a Tier-1 Pakistani bank — branches + DCs + cloud workloads)"),
        ("Expected vulnerability volume", "50K–150K findings per quarterly scan cycle; 200K–600K records/year"),
        ("Governance document corpus", "200–400 policies + procedures at steady state; 12 audit packages/year"),
        ("Regulatory burden", "SBP cybersecurity guidelines, PCI-DSS, ISO 27001, internal bank policies"),
        ("Data retention", "7 years for audit logs and evidence (banking regulatory minimum)"),
        ("Deployment mode", "On-premises in UBL data centre(s); no public cloud component"),
    ]
    r = 12
    for label, val in assumptions:
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        c = ws.cell(row=r, column=3, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value="How to complete this workbook")
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT

    steps = [
        ("1.", "Open the 'Questionnaire' tab. Each question explains what it drives in our sizing so you can answer with the right level of detail."),
        ("2.", "Type your answers directly into the highlighted input cells. Where a dropdown is provided, please use it — that lets our sizing template auto-derive the BOM."),
        ("3.", "Question 4 references the 'Storage Tier Guide' tab; Question 5 references the 'Local-LLM Reference' tab. Review both before answering those questions."),
        ("4.", "Where exact values are not known, please give a range or 'TBD' rather than leaving blank — we'd rather over-provision a known gap than guess silently."),
        ("5.", "Return the completed file to your Compliverse account contact. We will respond within 5 business days with a sized BOM and indicative pricing."),
    ]
    r += 1
    for n, text in steps:
        ws.cell(row=r, column=2, value=n).font = FONT_LABEL
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        c = ws.cell(row=r, column=3, value=text)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value="Contact")
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(
        row=r,
        column=2,
        value="Compliverse Solutions  ·  proposals@compliverse.io  ·  Reference: UBL-PK-2026-001",
    )
    c.font = FONT_BODY
    c.alignment = ALIGN_LEFT

    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(
        row=r,
        column=2,
        value=(
            "Confidentiality — This document contains pre-sales architectural information shared "
            "under the executed NDA. Please do not redistribute outside UBL infrastructure and security teams."
        ),
    )
    c.font = FONT_WHY
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[r].height = 28


# ── Sheet 2: Questionnaire ───────────────────────────────────────────────
QUESTIONS = [
    {
        "n": 1,
        "title": "User concurrency profile",
        "ask": (
            "Confirm the 100 named users assumption. What is your expected peak concurrent user "
            "count (users actively using the system at the same time)? What are your working "
            "hours / time zone? How many branch offices will access the platform?"
        ),
        "why": "Sizes the FastAPI app tier, load balancer, and session store. 30% concurrency is our default for banks.",
        "format": "Number + free text",
        "validation": None,
    },
    {
        "n": 2,
        "title": "Deployment topology",
        "ask": (
            "Single data centre, active-passive with DR (cold / warm / hot standby), or "
            "active-active dual DC? Will we deploy on bare metal, VMware / Hyper-V, or "
            "Kubernetes (OpenShift / Rancher / vanilla)?"
        ),
        "why": "Drives HA multiplier, DR replication design, and whether we ship VM images or container manifests.",
        "format": "Dropdown + free text",
        "validation": [
            "Single DC, no DR",
            "Single DC + cold DR",
            "Active-Passive (warm standby)",
            "Active-Passive (hot standby)",
            "Active-Active dual DC",
        ],
    },
    {
        "n": 3,
        "title": "Compute platform available",
        "ask": (
            "CPU vendor and generation (e.g. Intel Xeon Scalable Gen 4, AMD EPYC Gen 4). "
            "Maximum vCPU and RAM you can allocate to us per host. Can we have dedicated "
            "hosts, or only shared VMs in a virtualised pool?"
        ),
        "why": "Tells us whether our recommended 8 vCPU / 32 GB profile fits, or we need to re-shape across more smaller hosts.",
        "format": "Free text + numbers",
        "validation": None,
    },
    {
        "n": 4,
        "title": "Storage profile & target tier",
        "ask": (
            "Available storage tiers (NVMe / SAS SSD / spinning). Backing array (Dell "
            "PowerStore, NetApp, Pure, Ceph, vSAN, other). IOPS guarantee per LUN if any. "
            "Reviewing the 'Storage Tier Guide' tab, which target tier do you plan to provision: "
            "100 GB (pilot only) / 500 GB / 1 TB (recommended) / 3 TB / 5 TB?"
        ),
        "why": "Postgres needs NVMe with ≥10K IOPS sustained for nightly enrichment. Target tier sets the Y1–Y3 storage budget — 1 TB is our default recommendation for 100 users + 5K-asset bank scanning + 3-year retention.",
        "format": "Dropdown + free text",
        "validation": [
            "100 GB — pilot only",
            "500 GB — Y1 production",
            "1 TB — recommended (3-year retention)",
            "3 TB — HA + DR + 7-year retention",
            "5 TB — future-proof (DR + growth headroom)",
            "TBD — please advise",
        ],
    },
    {
        "n": 5,
        "title": "GPU availability for on-prem LLM",
        "ask": (
            "Do you already own GPUs (A100 80GB / H100 / L40S / A10 / L4)? If procurement "
            "is required, what is the power and cooling budget per rack-U? Reviewing the "
            "'Local-LLM Reference' tab, which tier do you target — A (1× A100), B (2× A100, "
            "recommended), or C (2× H100 or 4× A100)?"
        ),
        "why": "Picks the local-LLM tier. Tier B is our recommendation for 100 users; Tier C if accuracy gap must be ≤3%.",
        "format": "Dropdown + free text",
        "validation": [
            "Cloud LLM (no on-prem GPU)",
            "Tier A — 1× A100 80GB (budget)",
            "Tier B — 2× A100 80GB (recommended)",
            "Tier C — 2× H100 or 4× A100 (no compromise)",
            "GPUs already owned (specify in notes)",
            "TBD — please advise",
        ],
    },
    {
        "n": 6,
        "title": "Network & egress policy",
        "ask": (
            "Internal LAN speed (1 / 10 / 25 GbE). Is internet egress allowed? We need it "
            "for daily threat-intelligence sync (NVD, EPSS, CISA KEV, MSRC) — roughly 50 MB / day. "
            "If fully air-gapped, can we deploy an internal mirror / relay?"
        ),
        "why": "Determines worker design for threat-intel and whether we ship the offline DB sync tool.",
        "format": "Dropdown + free text",
        "validation": [
            "10 GbE LAN, full internet egress",
            "10 GbE LAN, proxied egress (whitelist)",
            "10 GbE LAN, fully air-gapped",
            "1 GbE LAN, full internet egress",
            "1 GbE LAN, fully air-gapped",
            "Other (specify)",
        ],
    },
    {
        "n": 7,
        "title": "Backup infrastructure & RPO/RTO",
        "ask": (
            "Existing backup target (Veeam, NetBackup, Commvault, Rubrik, MinIO, tape, other). "
            "Acceptable RPO and RTO from the dropdowns below. Retention years for audit logs "
            "and evidence (regulatory minimum)."
        ),
        "why": "Picks Postgres replication mode (sync vs async), WAL archive cadence, and evidence replication frequency.",
        "format": "Dropdown + free text",
        "validation_rpo": ["5 minutes", "15 minutes", "1 hour", "4 hours", "24 hours"],
        "validation_rto": ["15 minutes", "30 minutes", "2 hours", "4 hours", "8 hours"],
        "validation": None,
    },
    {
        "n": 8,
        "title": "High-availability requirement",
        "ask": (
            "Is HA mandatory for the application and database tier from day one, or is a "
            "single-node production acceptable for year 1? If HA is required, target SLA "
            "(99.5% / 99.9% / 99.95%)?"
        ),
        "why": "HA mandatory doubles compute + DB footprint. Single-node production is viable for year-1 with planned HA upgrade.",
        "format": "Dropdown",
        "validation": [
            "Single-node production (year-1)",
            "HA mandatory — 99.5% SLA",
            "HA mandatory — 99.9% SLA",
            "HA mandatory — 99.95% SLA",
        ],
    },
    {
        "n": 9,
        "title": "Security infrastructure constraints",
        "ask": (
            "Required encryption-at-rest standard (AES-256, FIPS 140-2, self-encrypting drives). "
            "HSM / KMS available (Thales, Gemalto, CyberArk, HashiCorp Vault). Hardened OS baseline "
            "(CIS RHEL, Ubuntu STIG, custom). Mandatory antivirus / EDR on application hosts "
            "(CrowdStrike, SentinelOne, Defender for Endpoint, other)?"
        ),
        "why": "Tells us whether to ship FIPS-mode builds, how to wire secret injection, and the EDR exclusion list we need to publish.",
        "format": "Multi-select + free text",
        "validation": None,
    },
    {
        "n": 10,
        "title": "Existing UBL services to integrate with",
        "ask": (
            "AD / LDAP domain (single forest or multi-domain?). SSO IdP (ADFS, Azure AD, Okta, "
            "Ping, other). SMTP relay for outbound notifications. Syslog target for our audit "
            "stream (QRadar, Splunk, ArcSight, other). Monitoring platform we can plug observability "
            "into (Prometheus / Grafana, Nagios, Zabbix, Dynatrace, other)."
        ),
        "why": "Wires SSO, mail delivery, audit-log shipping, and the monitoring exporters we enable at install time.",
        "format": "Free text",
        "validation": None,
    },
]


def build_questionnaire(ws):
    ws.title = "Questionnaire"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    widths = [4, 16, 50, 32, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Infrastructure Sizing Questionnaire — 10 Questions"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = (
        "Please complete the highlighted cells. Dropdowns are provided where applicable. "
        "Each question has a 'Why we ask' note so you can answer at the right level of detail."
    )
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 22

    # Column headers
    headers = ["#", "Topic", "Question", "Your Answer", "Notes / Context"]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = FONT_TABLE_HDR
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 22

    r = 5
    for q in QUESTIONS:
        # Two rows per question: question + why
        q_row = r
        why_row = r + 1
        ws.row_dimensions[q_row].height = 70
        ws.row_dimensions[why_row].height = 22

        # # column
        c = ws.cell(row=q_row, column=1, value=q["n"])
        c.font = FONT_QUESTION
        c.alignment = ALIGN_CENTER
        c.fill = FILL_HEADER_GREY
        c.border = BORDER_ALL
        ws.merge_cells(start_row=q_row, start_column=1, end_row=why_row, end_column=1)
        ws.cell(row=why_row, column=1).fill = FILL_HEADER_GREY
        ws.cell(row=why_row, column=1).border = BORDER_ALL

        # Topic
        c = ws.cell(row=q_row, column=2, value=q["title"])
        c.font = FONT_QUESTION
        c.alignment = ALIGN_LEFT
        c.fill = FILL_LIGHT
        c.border = BORDER_ALL
        c = ws.cell(row=why_row, column=2, value=f"Format: {q['format']}")
        c.font = FONT_WHY
        c.alignment = ALIGN_LEFT
        c.fill = FILL_LIGHT
        c.border = BORDER_ALL

        # Question
        c = ws.cell(row=q_row, column=3, value=q["ask"])
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        c.border = BORDER_ALL
        c = ws.cell(row=why_row, column=3, value=f"Why we ask — {q['why']}")
        c.font = FONT_WHY
        c.alignment = ALIGN_LEFT
        c.fill = FILL_LIGHT
        c.border = BORDER_ALL

        # Your Answer (input cell, white fill with blue accent border emphasis)
        ans = ws.cell(row=q_row, column=4, value="")
        ans.font = FONT_INPUT
        ans.alignment = ALIGN_LEFT
        ans.fill = FILL_ACCENT
        ans.border = BORDER_ALL
        ws.merge_cells(start_row=q_row, start_column=4, end_row=why_row, end_column=4)

        # Notes
        notes = ws.cell(row=q_row, column=5, value="")
        notes.font = FONT_INPUT
        notes.alignment = ALIGN_LEFT
        notes.fill = FILL_INPUT
        notes.border = BORDER_ALL
        ws.merge_cells(start_row=q_row, start_column=5, end_row=why_row, end_column=5)

        # Data validation
        if q["n"] == 7:
            # Q7 has two dropdowns side-by-side using prompt text in the answer cell
            ans.value = "RPO: <pick>\nRTO: <pick>\nBackup target: <free text>\nRetention years: <number>"
            dv_rpo = DataValidation(
                type="list",
                formula1='"' + ",".join(q["validation_rpo"]) + '"',
                allow_blank=True,
                showDropDown=False,
            )
            dv_rpo.add(f"D{q_row}")
            ws.add_data_validation(dv_rpo)
        elif q["validation"]:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(q["validation"]) + '"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.add(f"D{q_row}")
            ws.add_data_validation(dv)

        r += 2

    # Sign-off block
    sign_row = r + 1
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=5)
    c = ws.cell(row=sign_row, column=1, value="Sign-off")
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT

    for i, label in enumerate(("Name", "Role / Department", "Date", "Email"), start=0):
        rr = sign_row + 1 + i
        ws.cell(row=rr, column=2, value=label).font = FONT_LABEL
        ws.cell(row=rr, column=2).alignment = ALIGN_RIGHT
        ws.cell(row=rr, column=2).border = BORDER_ALL
        ws.cell(row=rr, column=2).fill = FILL_LIGHT
        ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=5)
        c = ws.cell(row=rr, column=3, value="")
        c.font = FONT_INPUT
        c.fill = FILL_ACCENT
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[rr].height = 20


# ── Sheet 3: Local-LLM Reference ─────────────────────────────────────────
LLM_TIERS = [
    {
        "tier": "A — Budget / Pilot",
        "gpu": "1 × A100 80 GB",
        "model": "Llama 3.3 70B Instruct 4-bit (AWQ/GPTQ) via vLLM. Must quantize — 70B at FP16 needs 140 GB VRAM, won't fit on a single A100.",
        "concurrency": "5–8 concurrent RAG queries; ComplyChat p95 ≈ 10–14 s; OCR throughput ≈ 1 page / 4 s",
        "ocr": "Qwen2-VL 7B time-sharing the same GPU (adapter swap)",
        "gap": "5–10% reasoning gap vs GPT-4o; 10–15% on multi-column OCR. ComplyChat hallucination rate ~2× higher on long-context (>8K tokens).",
        "fit": "Pilot or proof-of-concept (<50 active users). Not recommended for production audit-evidence workflows.",
        "fill": FILL_AMBER,
    },
    {
        "tier": "B — UBL Recommended",
        "gpu": "2 × A100 80 GB (NVLink) OR 1 × H100 80 GB",
        "model": "Llama 3.3 70B Instruct FP16 across both A100s via tensor parallelism (vLLM TP=2). Dedicated Qwen2-VL 72B for vision if budget allows third GPU; otherwise Qwen2-VL 7B FP16 on GPU#2 alongside the LLM.",
        "concurrency": "15–25 concurrent ComplyChat sessions; p95 ≈ 5–7 s; OCR throughput ≈ 1 page / 1.5 s",
        "ocr": "Qwen2-VL 7B FP16 on GPU#2 (recommended) or Qwen2-VL 72B if 3rd GPU added",
        "gap": "3–5% reasoning gap vs GPT-4o; ~5% OCR gap. Within UBL's acceptable accuracy band for ComplyChat + governance document parsing.",
        "fit": "UBL — 100 users + nightly batch enrichment. Sweet spot for accuracy × cost × concurrency.",
        "fill": FILL_GREEN,
    },
    {
        "tier": "C — No-compromise / Future-proof",
        "gpu": "4 × A100 80 GB OR 2 × H100 80 GB",
        "model": "Llama 3.1 405B 4-bit (TP=4 across A100s) for ComplyChat + cross-framework reasoning. Dedicated Qwen2-VL 72B FP8 on its own GPU.",
        "concurrency": "25–40 concurrent; p95 ≈ 4–6 s",
        "ocr": "Qwen2-VL 72B FP8 on dedicated GPU — bank-form-grade accuracy",
        "gap": "1–3% reasoning gap vs GPT-4o (essentially parity on banking workflows). Best-in-class OCR.",
        "fit": "Hard 'no accuracy loss' mandate; or anticipated growth to 200+ users in 18 months.",
        "fill": FILL_ACCENT,
    },
]


def build_llm_reference(ws):
    ws.title = "Local-LLM Reference"
    ws.sheet_view.showGridLines = False

    widths = [4, 22, 24, 38, 22, 30, 24, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "Local-LLM Deployment Tiers"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:H3")
    c = ws["B3"]
    c.value = (
        "Three GPU-footprint options for on-prem LLM. The 'Accuracy gap' column is our internal benchmark "
        "vs the cloud GPT-4o baseline the platform is built on. Tier B is the sweet spot for 100 users."
    )
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[3].height = 26

    headers = ["", "Tier", "GPU footprint", "LLM strategy", "Concurrency", "Vision (OCR) model", "Accuracy gap vs GPT-4o", "Recommended for"]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=val)
        cell.font = FONT_TABLE_HDR
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[5].height = 24

    for i, t in enumerate(LLM_TIERS, start=6):
        ws.cell(row=i, column=1).border = BORDER_ALL
        values = [None, t["tier"], t["gpu"], t["model"], t["concurrency"], t["ocr"], t["gap"], t["fit"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
            if col >= 2:
                cell.fill = t["fill"]
        ws.row_dimensions[i].height = 56

    # Footnotes
    r = 6 + len(LLM_TIERS) + 1
    notes = [
        "Embedding model — bge-large-en-v1.5 (or bge-m3 for multilingual Urdu/English) runs CPU-only, ~5K req/min.",
        "Vector store — Qdrant on NVMe; ~10–15 GB for UBL's expected year-1 document corpus.",
        "Hybrid option — sensitive parsing on local Tier B, cross-framework drafting routed to cloud with PII redaction. Available on request.",
        "All tiers include vLLM (or TGI) serving, Prometheus exporters, and zero-downtime model swap via blue-green hot-reload.",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        c = ws.cell(row=r, column=2, value="• " + note)
        c.font = FONT_WHY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 18
        r += 1


# ── Sheet 4: Storage Tier Guide (visible) ────────────────────────────────
STORAGE_TIERS = [
    {
        "name": "100 GB",
        "fits": "Pilot (≤ 30 users), 1-year retention, single framework active, light vuln scanning (~5K assets, 1 scan/quarter)",
        "live": "Postgres ~5 GB · Files ~30 GB · Vectors ~1 GB · WAL ~10 GB · Working ~50 GB",
        "backup": "Not included — bring own backup target",
        "fit_ubl": "No — too small for 100 GRC users at a Tier-1 bank",
        "fill": FILL_RED,
    },
    {
        "name": "500 GB",
        "fits": "Production for 100 users, 1–2 year retention, all frameworks active, quarterly bank-wide scans (~5K assets), 12 audit packages/yr",
        "live": "Postgres ~15 GB · Files ~80 GB · Vectors ~2 GB · WAL ~30 GB · Working ~150 GB",
        "backup": "+ 200 GB local backup (3× live, 30-day daily retention)",
        "fit_ubl": "Acceptable for Year 1 + first half of Year 2. Will outgrow by month 18.",
        "fill": FILL_AMBER,
    },
    {
        "name": "1 TB (UBL recommended)",
        "fits": "Production for 100 users, 3-year retention, all frameworks active, quarterly bank-wide scans, full audit-package archive, governance versioning",
        "live": "Postgres ~40 GB · Files ~200 GB · Vectors ~5 GB · WAL ~60 GB · Working ~350 GB",
        "backup": "+ 500 GB local backup (3× live + 30 daily + 12 monthly + room for 7-year cold tier)",
        "fit_ubl": "Recommended baseline. Covers 3 years before next storage refresh.",
        "fill": FILL_GREEN,
    },
    {
        "name": "3 TB",
        "fits": "Production + active DR site + 7-year regulatory retention + uncompressed scanned evidence + multi-framework audit cycles in parallel",
        "live": "~600 GB live · 1.4 TB backup · 1 TB DR replica",
        "backup": "Active-passive DR with synchronous Postgres replication; offsite copy via MinIO mirror",
        "fit_ubl": "Pick if UBL mandates HA + DR from go-live, or if regulator requires immutable 7-year archive on-system.",
        "fill": FILL_ACCENT,
    },
    {
        "name": "5 TB",
        "fits": "Production + active DR + 7-year retention + capacity for 2× growth (e.g. UBL adds subsidiaries, vendor risk module, expanded vuln coverage)",
        "live": "~1 TB live · 2 TB backup · 2 TB DR replica with headroom",
        "backup": "Same as 3 TB plus offline tape/object archive on-system",
        "fit_ubl": "Future-proof option if UBL anticipates merging in subsidiary banks or expanding regulatory scope (e.g. SBP DFI rules).",
        "fill": FILL_ACCENT,
    },
]


def build_storage_tier_guide(ws):
    ws.title = "Storage Tier Guide"
    ws.sheet_view.showGridLines = False

    widths = [4, 22, 44, 36, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "Storage Tier Guide"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = (
        "Five tiers from pilot to fully future-proofed. Pick the tier closest to UBL's planned retention "
        "and DR posture in Question 4 / Question 7. Numbers are usable storage (after RAID/EC overhead)."
    )
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[3].height = 26

    headers = ["", "Tier", "Fits this profile", "Live data breakdown", "Backup/DR included", "Fit for UBL (100 users)"]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=val)
        cell.font = FONT_TABLE_HDR
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[5].height = 24

    for i, t in enumerate(STORAGE_TIERS, start=6):
        ws.cell(row=i, column=1).border = BORDER_ALL
        values = [None, t["name"], t["fits"], t["live"], t["backup"], t["fit_ubl"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
            if col >= 2:
                cell.fill = t["fill"]
        ws.row_dimensions[i].height = 72

    # Footnotes
    r = 6 + len(STORAGE_TIERS) + 1
    notes = [
        "Postgres NVMe is mandatory for any tier ≥ 500 GB — bank-scale vuln scans push the DB hard at 02:00 PKT nightly enrichment.",
        "Object store (evidence + audit packages + framework PDFs) accounts for ~60-70% of total — bias provisioning here.",
        "Backup target sizing assumes 3× live + 30 daily + 12 monthly + (for ≥ 3 TB) 7-year cold tier on tape or S3-compatible object archive.",
        "WAL archive sized for 30-day point-in-time recovery; trim to 7 days for tighter storage budget if RPO permits.",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value="• " + note)
        c.font = FONT_WHY
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = 22
        r += 1


# ── Sheet 5: Internal Sizing Tiers (hidden) ──────────────────────────────
SIZING_ROWS = [
    ("FastAPI app nodes",        "1 × (8 vCPU / 16 GB)",          "2 × (8 vCPU / 32 GB)",                       "2 × (8 vCPU / 32 GB) per site × 2 sites"),
    ("Celery worker nodes",      "1 × (8 vCPU / 16 GB) — parsing+default queues co-located", "2 × (8 vCPU / 32 GB) — parsing queue dedicated, default queue shared", "3 × (8 vCPU / 32 GB) per site — parsing, default, enrichment isolated"),
    ("PostgreSQL primary",       "1 × (8 vCPU / 32 GB / 500 GB NVMe)", "1 × (12 vCPU / 48 GB / 1 TB NVMe, 15K+ IOPS)", "1 × (16 vCPU / 64 GB / 2 TB NVMe, 25K+ IOPS)"),
    ("PostgreSQL replica",       "—",                              "1 × sync hot standby (same shape as primary)", "1 sync local + 1 async DR (same shape)"),
    ("Redis",                    "1 × (2 vCPU / 4 GB)",            "2 × (2 vCPU / 8 GB) Sentinel HA",            "3-node cluster per site"),
    ("Qdrant vector store",      "1 × (4 vCPU / 8 GB / 50 GB NVMe)","1 × (4 vCPU / 16 GB / 100 GB NVMe)",         "2-node replicated per site"),
    ("MinIO / object store",     "1 × node, 500 GB usable",        "4-node erasure-coded, 4 TB usable (8 TB raw)", "4-node EC per site, 8 TB usable (16 TB raw) per site"),
    ("Next.js frontend",         "1 × (2 vCPU / 4 GB)",            "2 × (2 vCPU / 4 GB)",                        "2 per site"),
    ("LB / reverse proxy",       "nginx single",                   "HAProxy + keepalived VIP",                   "Site-LB + global LB (F5/Citrix)"),
    ("PgBouncer (connection pool)", "Optional",                    "1 × (2 vCPU / 4 GB) — recommended",          "1 × per app site"),
    ("Monitoring (Prom+Grafana)", "Shared with app node",          "1 × (2 vCPU / 8 GB / 200 GB) — 30d metrics",  "1 × per site"),
    ("GPU node (if local LLM)",  "Tier A (1× A100 80GB)",          "Tier B (2× A100 80GB or 1× H100)",            "Tier C (4× A100 or 2× H100) per site"),
    ("Subtotal CPU (excl. GPU)", "~30 vCPU",                       "~68 vCPU",                                   "~160 vCPU"),
    ("Subtotal RAM (excl. GPU)", "~80 GB",                         "~220 GB",                                    "~520 GB"),
    ("Subtotal usable storage",  "1–2 TB",                         "5–8 TB",                                     "20–30 TB across sites"),
]

DATA_VOLUME_ROWS = [
    # Bank-scale: ~15K employees, ~5K IT assets in scope, 100 GRC users, Tier-1 regulatory burden.
    ("PostgreSQL: vulnerabilities + asset links + enrichment (5K assets quarterly-scanned)", "8–15 GB",      "30–60 GB"),
    ("PostgreSQL: audit logs (compressed, 2K events/day at bank scale)",                     "1–2 GB",        "6–10 GB"),
    ("PostgreSQL: governance, risk, controls, evidence metadata",                            "0.5–1 GB",      "2–4 GB"),
    ("PostgreSQL WAL archive (30-day rolling)",                                              "20–40 GB",      "80–120 GB"),
    ("Vector store (Qdrant) — ~5K policy chunks + ~1K evidence summaries",                   "1–2 GB",        "4–8 GB"),
    ("Object store: governance documents + versions (300 policies × 2 MB avg × 3 versions)", "2–4 GB",        "8–15 GB"),
    ("Object store: evidence files (audit cycles, screenshots, scan reports)",               "20–60 GB",      "80–200 GB"),
    ("Object store: audit packages (12 packages/yr × 300–800 MB)",                           "4–10 GB",        "20–50 GB"),
    ("Object store: uploaded framework PDFs + parsed artifacts",                             "1–2 GB",        "2–4 GB"),
    ("Object store: scanner XML/JSON imports (quarterly bulk feeds)",                        "5–15 GB",        "25–60 GB"),
    ("──── Total live data (Y1 → Y3) ────",                                                  "~60–150 GB",    "~250–500 GB"),
    ("Backup target (3× live + 30d daily + 12 monthly + 7yr cold tier)",                     "~500 GB",        "~1.5–3 TB"),
]


def build_sizing_tiers(ws):
    ws.title = "Internal Sizing Tiers"
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False

    widths = [4, 28, 28, 36, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Internal Sizing Tiers — DO NOT SEND TO CUSTOMER"
    c.font = FONT_TITLE
    c.alignment = ALIGN_LEFT
    c.fill = FILL_RED
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "Tier S = single-node / no HA · Tier M = UBL default (HA, no DR) · Tier L = HA + active DR"
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_LEFT

    # Compute table
    headers = ["", "Component", "Tier S", "Tier M (UBL default)", "Tier L"]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=val)
        cell.font = FONT_TABLE_HDR
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[5].height = 22

    for i, row in enumerate(SIZING_ROWS, start=6):
        ws.cell(row=i, column=1).border = BORDER_ALL
        for col, val in enumerate(row, start=2):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
            if col == 4:
                cell.fill = FILL_GREEN
        ws.row_dimensions[i].height = 20

    # Data volume table
    r = 6 + len(SIZING_ROWS) + 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c = ws.cell(row=r, column=2, value="Data-volume projection (UBL, 100 users)")
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT
    r += 1
    headers = ["", "Data category", "Year 1", "Year 3", ""]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = FONT_TABLE_HDR
        cell.fill = FILL_NAVY
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[r].height = 22
    for row in DATA_VOLUME_ROWS:
        r += 1
        ws.cell(row=r, column=1).border = BORDER_ALL
        for col, val in enumerate(row, start=2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
        ws.cell(row=r, column=5).border = BORDER_ALL

    # Concurrency targets
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    c = ws.cell(row=r, column=2, value="Throughput & latency targets (Tier M)")
    c.font = FONT_SECTION
    c.alignment = ALIGN_LEFT
    targets = [
        ("Peak concurrent web users", "30 (confirm via Q1)"),
        ("API p95 latency", "< 400 ms list / detail endpoints"),
        ("ComplyChat p95", "< 6 s cloud LLM · < 8 s local Tier B"),
        ("Document parse p95", "< 45 s for 20-page PDF (cloud) · < 60 s (local Tier B)"),
        ("Vuln enrichment burst", "500 vulns/h cloud · 200 vulns/h local Tier B"),
        ("Nightly jobs window", "02:00 PKT — NVD 30m, EPSS 10m, MSRC 20m, KEV 5m"),
    ]
    for label, val in targets:
        r += 1
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = BORDER_ALL
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        c = ws.cell(row=r, column=3, value=val)
        c.font = FONT_BODY
        c.alignment = ALIGN_LEFT
        c.border = BORDER_ALL


# ── Sheet 5: Internal Calc (veryHidden) ──────────────────────────────────
def build_internal_calc(ws):
    ws.title = "Internal Calc"
    ws.sheet_state = "veryHidden"

    for i, w in enumerate([26, 38, 28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = "UBL sizing auto-derivation (read after questionnaire is returned)"
    ws["A1"].font = FONT_SECTION

    # Pull selected answers from Questionnaire.
    # Question rows (D column = answer): Q1=5, Q2=7, Q3=9, Q4=11, Q5=13, Q6=15, Q7=17, Q8=19, Q9=21, Q10=23
    refs = [
        ("Q1 — concurrency answer",     "=Questionnaire!D5"),
        ("Q2 — topology",               "=Questionnaire!D7"),
        ("Q3 — compute platform",       "=Questionnaire!D9"),
        ("Q4 — storage",                "=Questionnaire!D11"),
        ("Q5 — GPU tier",               "=Questionnaire!D13"),
        ("Q6 — network/egress",         "=Questionnaire!D15"),
        ("Q7 — backup RPO/RTO",         "=Questionnaire!D17"),
        ("Q8 — HA requirement",         "=Questionnaire!D19"),
        ("Q9 — security infra",         "=Questionnaire!D21"),
        ("Q10 — integrations",          "=Questionnaire!D23"),
    ]
    for i, (label, formula) in enumerate(refs, start=3):
        ws.cell(row=i, column=1, value=label).font = FONT_LABEL
        ws.cell(row=i, column=1).fill = FILL_LIGHT
        ws.cell(row=i, column=1).border = BORDER_ALL
        c = ws.cell(row=i, column=2, value=formula)
        c.font = FONT_BODY
        c.border = BORDER_ALL

    # Tier-pick formula based on Q8 (HA) + Q2 (topology containing 'Active-Active')
    out_row = 3 + len(refs) + 2
    ws.cell(row=out_row, column=1, value="Recommended Tier").font = FONT_LABEL
    ws.cell(row=out_row, column=1).fill = FILL_LIGHT
    ws.cell(row=out_row, column=1).border = BORDER_ALL
    # If Q2 contains "Active-Active", Tier L. Elif Q8 starts with "HA mandatory", Tier M. Else Tier S.
    tier_formula = (
        '=IF(ISNUMBER(SEARCH("Active-Active",B4)),"L (HA + active DR)",'
        'IF(ISNUMBER(SEARCH("HA mandatory",B10)),"M (HA, no DR)","S (single-node)"))'
    )
    c = ws.cell(row=out_row, column=2, value=tier_formula)
    c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    c.fill = FILL_GREEN
    c.border = BORDER_ALL

    # GPU tier pick from Q5
    out_row += 1
    ws.cell(row=out_row, column=1, value="LLM / GPU footprint").font = FONT_LABEL
    ws.cell(row=out_row, column=1).fill = FILL_LIGHT
    ws.cell(row=out_row, column=1).border = BORDER_ALL
    gpu_formula = (
        '=IF(ISNUMBER(SEARCH("Cloud",B7)),"Cloud LLM — no GPU",'
        'IF(ISNUMBER(SEARCH("Tier A",B7)),"1× A100 80GB",'
        'IF(ISNUMBER(SEARCH("Tier B",B7)),"2× A100 80GB",'
        'IF(ISNUMBER(SEARCH("Tier C",B7)),"2× H100 or 4× A100","TBD — clarify with customer"))))'
    )
    c = ws.cell(row=out_row, column=2, value=gpu_formula)
    c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    c.fill = FILL_GREEN
    c.border = BORDER_ALL

    # Rollups
    out_row += 2
    ws.cell(row=out_row, column=1, value="Compute rollup (vCPU / RAM)").font = FONT_LABEL
    ws.cell(row=out_row, column=1).fill = FILL_LIGHT
    ws.cell(row=out_row, column=1).border = BORDER_ALL
    rollup_formula = (
        '=IF(B15="L (HA + active DR)","~140 vCPU / ~500 GB RAM / ~12 TB storage",'
        'IF(B15="M (HA, no DR)","~58 vCPU / ~210 GB RAM / ~6 TB storage",'
        '"~30 vCPU / ~120 GB RAM / ~3 TB storage"))'
    )
    c = ws.cell(row=out_row, column=2, value=rollup_formula)
    c.font = FONT_BODY
    c.border = BORDER_ALL

    # Action note
    out_row += 2
    ws.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=3)
    c = ws.cell(
        row=out_row,
        column=1,
        value=(
            "When the workbook returns from UBL, unhide this sheet via VBA "
            "(ActiveWorkbook.Sheets(\"Internal Calc\").Visible = xlSheetVisible) "
            "or python: ws.sheet_state = 'visible'. The Recommended Tier + GPU formulas above "
            "produce the BOM lookup that drives the proposal deck."
        ),
    )
    c.font = FONT_WHY
    c.alignment = ALIGN_LEFT


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # First sheet exists by default — rename and reuse.
    build_cover(wb.active)
    build_questionnaire(wb.create_sheet())
    build_storage_tier_guide(wb.create_sheet())
    build_llm_reference(wb.create_sheet())
    build_sizing_tiers(wb.create_sheet())
    build_internal_calc(wb.create_sheet())

    out_path = Path(__file__).parent / "UBL_Pakistan_Infrastructure_Sizing_Questionnaire.xlsx"
    wb.save(out_path)
    print(f"Wrote: {out_path}")
    print(f"Size : {out_path.stat().st_size / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
